#!/usr/bin/env python3
"""Avia Solutions - 21 August 2026. John: the departure-timing chart currently plots
"route total carried" already clipped to the 87.5% load-factor capacity ceiling. Because
the clipped total is flat across BOTH the 00:00-06:00 window and the evening window at the
identical ceiling value, the chart cannot show that the unrestricted optimum (typically
around 00:30) would carry materially more than the chosen, curfew-compliant departure -
which is exactly the point Mark agreed to make on the 21 August call. This rebuilds the
chart with the TRUE, uncapped total demand as the plotted line and the capacity ceiling
drawn as a flat reference line instead of a clip.

NO DATABASE ACCESS NEEDED. This reads the "Departure curve" sheet that cortex_workbook.py
already writes into every Meridian Excel export where an airline is named (added 19
August). That sheet already carries, per candidate departure time: the raw connecting
score, the SCALED-BUT-UNCAPPED two-way connecting figure, and the CAPPED two-way route
total. Point-to-point is constant across the day (the sheet's own footnote states the
figure), so:

    true uncapped total(t) = connecting, two-way(t) + point-to-point constant

This is arithmetic on the sheet's own numbers, not a new model call - consistent with the
house rule against re-deriving figures that are already sourced in front of you. A
self-check (below) confirms the reconstruction: capping true_total(t) at the sheet's own
stated ceiling must reproduce the sheet's own "Route total carried" column exactly, for
every row. If it doesn't, STOP - something about this workbook's sheet has changed shape
and the chart must not be trusted.

TESTED against Meridian_SJC_TPE (27).xlsx (CI, 20 August export): self-check passed on
all 27 rows; true demand peaks circa 265,270 two-way at the 00:30 unrestricted optimum
against a 151,515 two-way capacity ceiling - the gap the current, capped chart cannot
show.

21 AUGUST, SECOND VERSION: John flagged that the single total-demand line lost the
point-to-point / connecting split the original two-line chart had, and with it the
fact that P2P is constant across the day (route_feed's own finding) while connecting
varies. Since P2P sits well under the capacity ceiling on its own, every passenger the
cap costs is connecting feed, never point-to-point - a script assertion confirms this
holds before the chart is drawn, rather than asserting it in a caption. Now plotted as
a stacked band (P2P base, connecting on top) with the region above the ceiling shaded
separately as "connecting demand the cap costs".

USAGE. Run against ANY Meridian Excel export for an airline that has a "Departure curve"
sheet - CI, BR or JX, the current run or a fresh one:

    py -3.12 departure_curve_uncapped.py "Meridian_SJC_TPE (27).xlsx"

Writes <input file stem>_uncapped_curve.png beside the input file.

Avia Solutions Limited. All rights reserved.
"""
import re
import sys
from pathlib import Path

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

NAVY = "#1F3864"
BLUE = "#4A90A4"
GREY = "#8C8C8C"


def _mins(hhmm):
    h, m = hhmm.split(":")
    return int(h) * 60 + int(m)


def load_curve(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Departure curve" not in wb.sheetnames:
        raise SystemExit(
            f"No 'Departure curve' sheet in {xlsx_path!r}. This sheet only exists when "
            "the run named an airline (cortex_workbook.py's own gate). Re-export from "
            "Meridian with an airline selected, then re-run this script - do not guess "
            "the figures from another file.")
    ws = wb["Departure curve"]

    hdr_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        if row and row[0] == "Departure (origin local)":
            hdr_row = i
            break
    if hdr_row is None:
        raise SystemExit("Could not find the header row ('Departure (origin local)') - "
                          "sheet layout has changed, do not trust a guessed column order.")

    rows = []
    note_text = ""
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if row[0] is None:
            continue
        if isinstance(row[0], str) and row[0].startswith("Chosen departure"):
            note_text = row[0]
            continue
        hhmm, permitted, raw, conn2, beyond2, behind2, total2 = row[:7]
        rows.append({"hhmm": hhmm, "mins": _mins(hhmm), "permitted": (permitted == "yes"),
                     "conn2": float(conn2), "total2_capped": float(total2)})
    if not note_text:
        # footnote sometimes sits a row or two further down (blank spacer rows) - scan on
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if row[0] and isinstance(row[0], str) and row[0].startswith("Chosen departure"):
                note_text = row[0]
                break
    if not note_text:
        raise SystemExit("Could not find the footnote row with the point-to-point constant "
                          "and capacity ceiling - do not fabricate them.")

    m_p2p = re.search(r"constant across the day \(([\d,]+) two-way\)", note_text)
    m_cap = re.search(r"capacity ceiling|aircraft ceiling.*?\(([\d,]+) two-way", note_text)
    m_chosen = re.search(r"Chosen departure ([\d:]+)", note_text)
    m_unres = re.search(r"unrestricted optimum ([\d:]+)", note_text)
    if not (m_p2p and m_cap and m_chosen):
        raise SystemExit(f"Footnote did not parse as expected:\n{note_text}\n"
                          "Fix the parser before trusting any figure below.")
    p2p2 = float(m_p2p.group(1).replace(",", ""))
    cap2 = float(m_cap.group(1).replace(",", ""))
    chosen = m_chosen.group(1)
    unrestricted = m_unres.group(1) if m_unres else None

    rows.sort(key=lambda r: r["mins"])
    return rows, p2p2, cap2, chosen, unrestricted, note_text


def build_chart(xlsx_path, out_path=None):
    rows, p2p2, cap2, chosen, unrestricted, note_text = load_curve(xlsx_path)

    # SELF-CHECK before anything is plotted. true uncapped total, capped at the sheet's own
    # ceiling, must reproduce the sheet's own capped column for every row.
    bad = []
    for r in rows:
        r["true_total2"] = r["conn2"] + p2p2
        reconstructed = min(r["true_total2"], cap2)
        if abs(reconstructed - r["total2_capped"]) > 1.0:
            bad.append((r["hhmm"], reconstructed, r["total2_capped"]))
    if bad:
        print("SELF-CHECK FAILED on", len(bad), "of", len(rows), "rows - do not trust this "
              "chart. First mismatches:")
        for hhmm, got, sheet in bad[:5]:
            print(f"  {hhmm}: reconstructed {got:,.0f} vs sheet {sheet:,.0f}")
        raise SystemExit(1)
    print(f"Self-check passed on all {len(rows)} rows: connecting(t) + point-to-point "
          f"({p2p2:,.0f}), capped at the ceiling ({cap2:,.0f}), reproduces the sheet's own "
          "'Route total carried' column exactly. The uncapped reconstruction is trustworthy.")

    # JOHN, 21 AUGUST: the single total-demand line lost the P2P/connecting split - and
    # with it, the fact that P2P is constant across the day (route_feed's own finding) so
    # the capacity ceiling never binds on it. P2P (64,290 two-way here) sits well under the
    # ceiling (151,515) on its own, so every passenger given up above the ceiling is
    # CONNECTING feed, never point-to-point. Confirmed, not assumed:
    if p2p2 >= cap2:
        raise SystemExit(f"P2P constant ({p2p2:,.0f}) exceeds the capacity ceiling "
                          f"({cap2:,.0f}) - the 'everything lost is connecting' framing "
                          "below does not hold for this run. Do not use it blind.")
    print(f"Confirmed: P2P ({p2p2:,.0f}) is always below the ceiling ({cap2:,.0f}), so "
          "every passenger the cap costs is connecting feed, never point-to-point.")

    xs = [r["mins"] / 60.0 for r in rows]
    p2p_line = [p2p2 for _ in rows]
    true_total = [r["true_total2"] for r in rows]
    chosen_mins = _mins(chosen) / 60.0

    fig, ax = plt.subplots(figsize=(10, 5.6))

    # shade restricted hours
    in_block = False
    start = None
    for r in rows + [rows[0]]:
        blocked = not r["permitted"]
        t = r["mins"] / 60.0
        if blocked and not in_block:
            start = t
            in_block = True
        elif not blocked and in_block:
            ax.axvspan(start, t, color="#E7EAF0", zorder=0)
            in_block = False
    if in_block:
        ax.axvspan(start, 24, color="#E7EAF0", zorder=0)

    # STACKED: P2P as the constant base band, connecting stacked on top - so the shape of
    # the top edge is still the total-demand curve, but the split is visible throughout.
    ax.fill_between(xs, 0, p2p_line, color=NAVY, alpha=0.85, zorder=2,
                     label=f"Point-to-point, constant all day ({p2p2:,.0f} two-way)")
    ax.fill_between(xs, p2p_line, true_total, color=BLUE, alpha=0.85, zorder=2,
                     label="Connecting, varies by departure time")

    # The region above the ceiling is demand the cap costs - shade it distinctly and
    # label it, since it is entirely connecting (proven above), never point-to-point.
    lost = [max(t - cap2, 0) for t in true_total]
    if any(lost):
        ax.fill_between(xs, [min(t, cap2) for t in true_total], true_total,
                         where=[t > cap2 for t in true_total], color="#C0504D", alpha=0.55,
                         hatch="///", edgecolor="#C0504D", linewidth=0, zorder=3,
                         label="Connecting demand the capacity cap costs")

    ax.axhline(cap2, color="#404040", linewidth=1.6, linestyle="--",
               zorder=4, label=f"Capacity ceiling, {cap2:,.0f} two-way (87.5% load factor)")

    chosen_true = next(r["true_total2"] for r in rows if r["hhmm"] == chosen)
    chosen_carried = min(chosen_true, cap2)
    ax.annotate(f"chosen {chosen}: {chosen_carried:,.0f} carried",
                xy=(chosen_mins, cap2), xytext=(chosen_mins + 0.6, cap2 * 0.60),
                arrowprops=dict(arrowstyle="->", color="#1F1F1F"), color="#1F1F1F", fontsize=10)
    if unrestricted:
        un_row = next((r for r in rows if r["hhmm"] == unrestricted), None)
        if un_row:
            un_lost = un_row["true_total2"] - min(un_row["true_total2"], cap2)
            ax.annotate(
                f"unrestricted optimum {unrestricted}: {un_row['true_total2']:,.0f} true "
                f"demand\n{un_lost:,.0f} of it connecting feed the cap would still cost",
                xy=(_mins(unrestricted) / 60.0, un_row["true_total2"]),
                xytext=(_mins(unrestricted) / 60.0 + 0.6, un_row["true_total2"] * 0.90),
                arrowprops=dict(arrowstyle="->", color="#1F1F1F"), color="#1F1F1F", fontsize=10)

    ax.set_xlim(0, 24)
    ax.set_ylim(0, None)
    ax.xaxis.set_major_locator(mticker.MultipleLocator(3))
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):02d}:00"))
    ax.set_xlabel("Departure time, origin local")
    ax.set_ylabel("Passengers a year, two-way")
    ax.set_title("Demand by outbound departure time, point-to-point vs connecting - "
                 "shaded hours are restricted departures",
                 fontsize=11, loc="left", color="#404040")
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=2, frameon=False, fontsize=9)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}"))
    fig.tight_layout()

    out_path = out_path or (Path(xlsx_path).with_name(Path(xlsx_path).stem + "_uncapped_curve.png"))
    fig.savefig(out_path, dpi=150)
    print(f"Written: {out_path}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: py -3.12 departure_curve_uncapped.py <Meridian export>.xlsx")
    build_chart(sys.argv[1])
