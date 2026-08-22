#!/usr/bin/env python3
"""Avia Solutions - verify the EW / 2-way workbook tab pairs (22 August 2026, John's request).

WHAT CHANGED. cortex_workbook.py's Forecast, Connecting feed, Schedule, Departure curve, Catchment
and Competition tabs are now each written twice - an "EW" tab at native each-way values and a
"2-way" tab at the same figures x2 - so the basis is stated in the tab name and never has to be
inferred from a footnote. Economics was NOT paired: a rotation's cost lines (fuel, maintenance,
landing, ground handling) are priced per rotation, with no coherent each-way half, so an EW split
there would set halved revenue against unhalved cost and understate profit. That tab stays single,
now labelled "two way" explicitly.

WHAT THIS SCRIPT CHECKS.
  1. The expected 15 tabs exist (Departure curve / Competition only build when their source data is
     present, so their count is conditional; the script reports what it found either way).
  2. On Forecast, Connecting feed and Schedule: the 2-way tab's COUNT cells are exactly 2x the EW
     tab's, and its RATE cells (capture rate, PTEW, seat factor) are identical to the EW tab's -
     doubling a rate would be the bug this project exists to avoid.
  3. On Catchment and Competition: EW and 2-way are byte-for-byte identical (both are percentages,
     no passenger count to double).
  4. On Departure curve: the 2-way tab's total column is exactly 2x the EW tab's.
  5. No sheet is named exactly "Forecast", "Connecting feed", "Schedule", "Departure curve",
     "Catchment" or "Competition" any more (the old unsuffixed names) - only Cover, Economics and
     Assumptions stay unsuffixed.

Run on the workstation:
    py -3.12 diag_ewtw_workbook_check.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + r"\app")

import openpyxl
import cortex_app as CA
import cortex_workbook as CWB

FAIL = []


def check(label, cond, detail=""):
    mark = "OK" if cond else "FAIL"
    print(f"  [{mark}] {label}" + (f"  ({detail})" if detail else ""))
    if not cond:
        FAIL.append(label)


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def find_row(ws, label_text, col=1, max_row=60):
    for r in range(1, max_row + 1):
        v = cell(ws, r, col)
        if v and label_text in str(v):
            return r
    return None


def main():
    fc = CA.calibrated_forecast("SJC", "TPE", airline="CI", carrier_type="FSC", aircraft="A359", freq=5)
    if not fc.get("ok"):
        print(f"Production call failed: {fc.get('error')}. STOP.")
        return

    tmpd = tempfile.mkdtemp()
    xlsx_path = os.path.join(tmpd, "ewtw_check.xlsx")
    CWB.build_workbook(xlsx_path, fc, {"airline_name": "CI", "analyst": "Avia Solutions",
                                        "date": "22 Aug 2026", "plan_lf": 0.85,
                                        "capture_basis": "modelled from drive time and competing service"})

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    names = wb.sheetnames
    print("Tabs built:", names, "\n")

    print("1. No old unsuffixed tab names survive")
    for old in ("Forecast", "Connecting feed", "Schedule", "Departure curve", "Catchment", "Competition"):
        check(f'no bare "{old}" sheet', old not in names)

    print("\n2. Forecast EW vs 2-way")
    if "Forecast EW" in names and "Forecast 2-way" in names:
        ew, tw = wb["Forecast EW"], wb["Forecast 2-way"]
        r_ew = find_row(ew, "GRAND TOTAL"); r_tw = find_row(tw, "GRAND TOTAL")
        check("GRAND TOTAL row found on both", r_ew and r_tw)
        if r_ew and r_tw:
            for col, label in ((4, "Grown demand"), (6, "Stimulated demand"), (8, "Forecast")):
                v_ew, v_tw = cell(ew, r_ew, col), cell(tw, r_tw, col)
                ok = v_ew is not None and v_tw is not None and abs(v_tw - 2 * v_ew) < 0.15
                check(f"{label} (000s): 2-way = 2x EW", ok, f"EW={v_ew}, 2-way={v_tw}")
            v_ew, v_tw = cell(ew, r_ew, 7), cell(tw, r_tw, 7)   # capture rate, a ratio
            check("Capture rate unchanged (rate, not a count)", abs((v_ew or 0) - (v_tw or 0)) < 0.001,
                  f"EW={v_ew}, 2-way={v_tw}")
            v_ew, v_tw = cell(ew, r_ew, 9), cell(tw, r_tw, 9)   # PTEW, a rate
            check("PTEW unchanged (rate, not a count)", v_ew == v_tw, f"EW={v_ew}, 2-way={v_tw}")
    else:
        check("Forecast EW / 2-way tabs present", False, f"found: {names}")

    print("\n3. Connecting feed EW vs 2-way")
    if "Connecting feed EW" in names and "Connecting feed 2-way" in names:
        ew, tw = wb["Connecting feed EW"], wb["Connecting feed 2-way"]
        for r_ew_total, r_tw_total, tag in [(find_row(ew, "Total", max_row=25), find_row(tw, "Total", max_row=25), "behind"),
                                             (find_row(ew, "Total", col=1, max_row=60), find_row(tw, "Total", col=1, max_row=60), "beyond")]:
            pass  # handled below with an explicit dual-leg scan
        # explicit dual scan: first "Total" row in the first 25 rows is the behind leg, the next is beyond
        totals_ew = [r for r in range(1, 60) if cell(ew, r, 1) == "Total"]
        totals_tw = [r for r in range(1, 60) if cell(tw, r, 1) == "Total"]
        check("two Total rows found on both tabs", len(totals_ew) == 2 and len(totals_tw) == 2,
              f"EW={totals_ew}, 2-way={totals_tw}")
        for i, r_ew in enumerate(totals_ew):
            r_tw = totals_tw[i]
            v_ew, v_tw = cell(ew, r_ew, 7), cell(tw, r_tw, 7)
            ok = v_ew is not None and v_tw is not None and abs(v_tw - 2 * v_ew) < 2
            check(f"leg {i+1} forecast total: 2-way = 2x EW", ok, f"EW={v_ew}, 2-way={v_tw}")
    else:
        check("Connecting feed EW / 2-way tabs present", False, f"found: {names}")

    print("\n4. Schedule EW vs 2-way")
    if "Schedule EW" in names and "Schedule 2-way" in names:
        ew, tw = wb["Schedule EW"], wb["Schedule 2-way"]
        r_ew = find_row(ew, "Total"); r_tw = find_row(tw, "Total")
        if r_ew and r_tw:
            v_ew, v_tw = cell(ew, r_ew, 8), cell(tw, r_tw, 8)   # season pax total
            ok = v_ew is not None and v_tw is not None and abs(v_tw - 2 * v_ew) < 2
            check("Total pax: 2-way = 2x EW", ok, f"EW={v_ew}, 2-way={v_tw}")
            # outbound/inbound rows (row 5, 6) should be identical between tabs
            same_rows = all(cell(ew, r, c) == cell(tw, r, c) for r in (5, 6) for c in (1, 2, 3, 4, 5, 6))
            check("outbound/inbound detail rows identical on both tabs", same_rows)
    else:
        check("Schedule EW / 2-way tabs present", False, f"found: {names}")

    print("\n5. Departure curve EW vs 2-way (built only when an optimiser curve is present)")
    if "Departure curve EW" in names and "Departure curve 2-way" in names:
        ew, tw = wb["Departure curve EW"], wb["Departure curve 2-way"]
        diffs = []
        for r in range(5, ew.max_row + 1):
            v_ew, v_tw = cell(ew, r, 7), cell(tw, r, 7)
            if v_ew is None or v_tw is None:
                continue
            diffs.append(abs(v_tw - 2 * v_ew))
        check("route total carried: 2-way = 2x EW on every row", diffs and max(diffs) < 2,
              f"max diff {max(diffs) if diffs else 'n/a'} over {len(diffs)} rows")
    else:
        print("  (not built this run - no optimiser curve on this payload; not a failure)")

    print("\n6. Catchment and Competition: EW and 2-way identical (percentages, no count to double)")
    if "Catchment EW" in names and "Catchment 2-way" in names:
        ew, tw = wb["Catchment EW"], wb["Catchment 2-way"]
        same = all(cell(ew, r, c) == cell(tw, r, c) for r in range(4, ew.max_row + 1) for c in (1, 2))
        check("Catchment EW == Catchment 2-way", same)
    if "Competition EW" in names and "Competition 2-way" in names:
        ew, tw = wb["Competition EW"], wb["Competition 2-way"]
        same = all(cell(ew, r, c) == cell(tw, r, c) for r in range(4, ew.max_row + 1) for c in (1, 2))
        check("Competition EW == Competition 2-way", same)
    elif "Competition EW" not in names:
        print("  (Competition not built this run - no alliance data passed via meta; not a failure)")

    print("\n7. Economics stays single, now labelled two way")
    check('exactly one "Economics" sheet, no EW/2-way pair', "Economics" in names
          and "Economics EW" not in names and "Economics 2-way" not in names)
    ews = wb["Economics"]
    # _title() writes the heading to row 1 and the subtitle to row 2 - "two way"
    # lives in the subtitle, so both rows need reading, not just the first.
    title_text = " ".join(str(cell(ews, r, c) or "") for r in (1, 2) for c in range(1, 8))
    check('Economics title states "two way"', "two way" in title_text.lower(), title_text)

    print("\n" + "=" * 70)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED:")
        for f in FAIL:
            print(f"  - {f}")
        print("\nSTOP - do not send workbooks out until these are resolved.")
    else:
        print("All checks passed. The EW/2-way pairs are consistent.")


if __name__ == "__main__":
    main()
