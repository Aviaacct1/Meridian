#!/usr/bin/env python3
"""
Avia Solutions - Deck Data Contract emitter.
==============================================================================
A SINGLE structured object per route that holds every field the house pitch deck
needs, so the deck builds with no manual data entry. This is an OUTPUT LAYER ONLY:
it exposes the forecast model's outputs (and the deck's published reference numbers
for the BA acceptance test) in one consistent, named structure. It does not change
the calibrated forecast method or any number it produces.

Two deliverables:
  1. emit_json()      - the contract as one JSON object.
  2. emit_workbook()  - an .xlsx mirroring it, one sheet per block (a client deliverable).

The contract is route-agnostic and driven by a RouteCase + the model's computed
outputs, so Genoa-New York or any pair fills the same shape. build_contract() maps
the model's output dict into the contract; ba_lhr_sjc_reference() returns the fully
populated BA London Heathrow - San Jose example from the validated 2015 deck numbers,
used as the acceptance test.

Derived metrics are implemented exactly per the spec:
  pdew         = annual two-way passengers / 728      (52 x 7 x 2)
  airline_share= airline annual forecast / city annual demand
  yield (RPK)  = passenger_revenue / (passengers x distance_km)
  ASK          = seats x distance_km x frequency x 52 x 2
  PRASK        = passenger_revenue / ASK
  TRASK        = total_revenue / ASK
Author of any generated file: Avia Solutions.
"""
from __future__ import annotations
import json, math, os, argparse

DAYS_2WAY = 728  # fallback only: 52 weeks x 7 daily x 2 directions, when no route freq is supplied


# ----------------------------------------------------------------- derived metrics
def haversine_km(lat1, lon1, lat2, lon2):
    r = 6371.0088
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1); dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def pdew(annual_two_way, freq=None, weeks=52.0):
    """PDEW = annual two-way passengers / the route's actual scheduled departures that year, two-way
    (freq/week x weeks x 2). Fixed 22 August 2026, same defect as route_feed.py's PTEW fix earlier
    the same day: this always divided by the fallback DAYS_2WAY (728, a DAILY each-way service
    assumed year-round), regardless of the route's real frequency. On a non-daily route the two never
    reconciled - CI/SJC-TPE at 5x/week has 260 scheduled departures each way, not the 364 the flat
    constant assumed, understating PDEW by roughly 29%. freq is optional and falls back to the old
    728 basis only when the caller has no route defined yet (ba_lhr_sjc_reference's fixture is daily,
    freq=7, so it already sits exactly on this basis and is left untouched, not converted)."""
    dep_two_way = (freq * weeks * 2) if freq else DAYS_2WAY
    return round(annual_two_way / dep_two_way, 1) if annual_two_way and dep_two_way else 0.0


def ask(seats, distance_km, freq_per_week):
    return seats * distance_km * freq_per_week * 52 * 2


def yield_per_rpk(passenger_revenue, passengers, distance_km):
    rpk = passengers * distance_km
    return round(passenger_revenue / rpk, 4) if rpk else None


def prask(passenger_revenue, ask_):
    return round(passenger_revenue / ask_, 4) if ask_ else None


def trask(total_revenue, ask_):
    return round(total_revenue / ask_, 4) if ask_ else None


def airline_share(forecast, demand):
    return round(forecast / demand, 4) if demand else 0.0


# ----------------------------------------------------------------- BA reference data
# From "British Airways Master forecast slides (OS) 06Mar2015.pptx" (Egnyte, the validated
# 2015 deck that won the route). All figures are the deck's published values. 000s expanded
# to absolute where the contract holds absolutes; the deck table itself is in 000s.
_BA_SEGMENTS = [
    # name, base(000s), growth, demand_yr(000s), stim, after_stim(000s), capture, forecast(000s)
    ("Origin (UK) Business",               71.4, 0.090, 77.9, 1.15, 89.6, 0.400, 35.8),
    ("Origin (UK) Leisure/VFR Primary",    36.4, 0.090, 39.7, 1.00, 39.7, 0.200,  7.9),
    ("Origin (UK) Leisure/VFR Secondary",  17.4, 0.090, 19.0, 1.00, 19.0, 0.200,  3.8),
    ("Origin (UK) Leisure/VFR Contested",   4.6, 0.090,  5.0, 1.00,  5.0, 0.050,  0.3),
    ("Destination (US) Business",          65.9, 0.090, 71.9, 1.15, 82.7, 0.220, 18.2),
    ("Destination (US) Leisure/VFR Primary",33.6, 0.075, 36.1, 1.00, 36.1, 0.300, 10.8),
    ("Destination (US) Leisure/VFR Secondary",16.1,0.075,17.3, 1.00, 17.3, 0.300,  5.2),
    ("Destination (US) Leisure/VFR Contested",4.3,0.075, 4.6, 1.00,  4.6, 0.150,  0.7),
]
# Connecting-at-hub (London) cities: code, name, country, demand, share, forecast
_BA_CNX_HUB = [
    ("PAR","Paris","France",140521,0.030,4216),("DUB","Dublin","Ireland",28174,0.090,2536),
    ("MUC","Munich","Germany",43810,0.054,2366),("CPT","Cape Town","South Africa",5893,0.375,2212),
    ("FRA","Frankfurt","Germany",100810,0.020,2016),("DUS","Dusseldorf","Germany",13088,0.137,1793),
    ("AMS","Amsterdam","Netherlands",60675,0.029,1767),("STO","Stockholm","Sweden",25155,0.062,1555),
    ("GVA","Geneva","Switzerland",12128,0.126,1531),("HEL","Helsinki","Finland",6606,0.196,1294),
    ("CAI","Cairo","Egypt",5501,0.235,1294),("MIL","Milan","Italy",19567,0.065,1278),
    ("BER","Berlin","Germany",18020,0.070,1269),("GLA","Glasgow","United Kingdom",4028,0.303,1219),
    ("CPH","Copenhagen","Denmark",32449,0.036,1173),("BCN","Barcelona","Spain",22072,0.052,1155),
    ("TLV","Tel Aviv-Yafo","Israel",45892,0.023,1075),("ZRH","Zurich","Switzerland",36924,0.029,1057),
    ("ORK","Cork","Ireland",3037,0.306,930),("JED","Jeddah","Saudi Arabia",5336,0.153,815),
    ("EDI","Edinburgh","United Kingdom",5944,0.129,768),("OSL","Oslo","Norway",9265,0.080,744),
    ("ROM","Rome","Italy",27549,0.026,715),("VIE","Vienna","Austria",14304,0.049,697),
    ("PRG","Prague","Czech Republic",8455,0.068,576),("MAD","Madrid","Spain",19730,0.029,576),
    ("LYS","Grenoble/Lyon","France",10060,0.057,569),("DXB","Dubai","UAE",13038,0.043,557),
    ("MAN","Manchester","United Kingdom",13139,0.042,553),("BRU","Brussels","Belgium",16663,0.031,516),
    ("JNB","Johannesburg","South Africa",9752,0.044,431),("EAP","Basel/Mulhouse","France",6575,0.065,427),
    ("NCE","Nice","France",7306,0.057,418),("HAM","Hamburg","Germany",12265,0.031,384),
    ("IST","Istanbul","Turkey",15646,0.023,356),("BUD","Budapest","Hungary",8959,0.036,325),
    ("ATH","Athens","Greece",7547,0.043,324),("BFS","Belfast","United Kingdom",1587,0.194,308),
    ("VCE","Venice","Italy",10673,0.028,296),("HAJ","Hannover","Germany",4183,0.059,247),
    ("ABV","Abuja","Nigeria",546,0.435,237),("SNN","Shannon","Ireland",4740,0.049,232),
    ("GOT","Goteborg","Sweden",8092,0.028,225),("ABZ","Aberdeen","United Kingdom",975,0.207,202),
    ("BLQ","Bologna","Italy",3872,0.042,162),("NCL","Newcastle","United Kingdom",1897,0.075,142),
    ("MOW","Moscow","Russia",10826,0.013,138),("BUH","Bucharest","Romania",3326,0.038,128),
    ("LIS","Lisbon","Portugal",3808,0.030,115),("LAD","Luanda","Angola",304,0.346,105),
    ("TLS","Toulouse","France",4835,0.021,99),("MRS","Marseille","France",5654,0.016,91),
    ("PSA","Pisa","Italy",1311,0.066,87),("RUH","Riyadh","Saudi Arabia",5400,0.015,81),
    ("WAW","Warsaw","Poland",5511,0.014,78),("LBA","Leeds Bradford","United Kingdom",473,0.159,75),
    ("STR","Stuttgart","Germany",7098,0.009,63),("BEY","Beirut","Lebanon",2502,0.021,53),
    ("KWI","Kuwait","Kuwait",1306,0.038,50),("DOH","Doha","Qatar",1476,0.032,47),
    ("ZAG","Zagreb","Croatia",1668,0.025,43),("BIO","Bilbao","Spain",1586,0.023,36),
    ("LUX","Luxembourg","Luxembourg",1643,0.020,33),("NBO","Nairobi","Kenya",3757,0.007,28),
    ("BGO","Bergen","Norway",1251,0.016,21),("LOS","Lagos","Nigeria",2750,0.007,18),
    ("EBB","Entebbe","Uganda",1294,0.013,17),("ACC","Accra","Ghana",1194,0.012,15),
    ("SVG","Stavanger","Norway",1052,0.013,14),("CMB","Colombo","Sri Lanka",979,0.014,14),
    ("AUH","Abu Dhabi","UAE",783,0.010,8),("BAH","Bahrain","Bahrain",799,0.007,6),
    ("SPU","Split","Croatia",518,0.009,4),("IBZ","Ibiza","Spain",115,0.027,3),
    ("Other","Other","",45675,0.000,6),
]


def ba_lhr_sjc_reference():
    """The fully populated BA London Heathrow - San Jose contract, from the validated 2015 deck.
    Used as the acceptance test. Hub = London (origin, BA's hub); Destination = San Jose."""
    # geometry
    LHR = (51.4775, -0.4614); SJC = (37.3626, -121.9291)
    dist_km = round(haversine_km(*LHR, *SJC), 1)
    dist_nm = round(dist_km / 1.852, 1)
    seats, freq = 214, 7
    annual_seats = seats * freq * 52 * 2          # 155,792
    ask_yr = ask(seats, dist_km, freq)

    # block 3 - segment table (deck is in 000s; expand forecast/demand to absolute pax)
    seg_rows = []
    for name, base, g, dy, stim, after, cap, fc in _BA_SEGMENTS:
        fc_abs = round(fc * 1000)
        seg_rows.append({
            "segment": name, "base_annual_demand": round(base * 1000),
            "annual_growth_rate": g, "demand_at_service_year": round(dy * 1000),
            "stimulation_factor": stim, "demand_after_stimulation": round(after * 1000),
            "capture_rate": cap, "forecast": fc_abs, "pdew": pdew(fc_abs * 2 if False else fc_abs),
        })
    # NOTE on pdew: the deck's PDEW is forecast / 728 (forecast already two-way annual). Use that.
    for r in seg_rows:
        r["pdew"] = pdew(r["forecast"])

    p2p_forecast = 82708
    cnx_hub_forecast = 45011
    cnx_dest_forecast = 2628
    grand_total = 130346

    # block 4 - connecting at hub (London)
    hub_cities = [{"nr": i + 1, "city_code": c, "city_name": n, "country": ctry,
                   "annual_demand": d, "airline_share": s, "annual_forecast": f, "pdew": pdew(f)}
                  for i, (c, n, ctry, d, s, f) in enumerate(_BA_CNX_HUB)]

    # block 6 - revenue 3yr (deck published)
    rev = {
        "years": [2016, 2017, 2018],
        "passengers": {
            "point_to_point": [82708, 85106, 87574],
            "connecting_at_hub": [45011, 46316, 47659],
            "connecting_at_destination": [2628, 2704, 2782],
            "total": [130346, 134126, 138016],
        },
        "annual_capacity": [annual_seats, annual_seats, annual_seats],
        "revenue": {
            "point_to_point": [84645831, 87100560, 89626476],
            "connecting_at_hub": [26636396, 27408851, 28203708],
            "connecting_at_destination": [2120647, 2182145, 2245428],
            "cargo": [7280000, 7644000, 8026200],
            "ancillary": [1036250, 1066302, 1097224],
            "total": [121719124, 125401858, 129199036],
        },
    }
    rev["implied_load_factor"] = [round(p / annual_seats, 3) for p in rev["passengers"]["total"]]

    # block 7 - detailed economics Yr1 (derived from the revenue table + geometry)
    pax_rev_y1 = rev["revenue"]["point_to_point"][0] + rev["revenue"]["connecting_at_hub"][0] + rev["revenue"]["connecting_at_destination"][0]
    total_rev_y1 = rev["revenue"]["total"][0]
    cnx_pax_y1 = rev["passengers"]["connecting_at_hub"][0] + rev["passengers"]["connecting_at_destination"][0]
    cnx_rev_y1 = rev["revenue"]["connecting_at_hub"][0] + rev["revenue"]["connecting_at_destination"][0]
    econ = {
        "equipment": "B788", "weekly_departures": freq, "total_departures_annual_two_way": freq * 52 * 2,
        "block_hours_per_departure": 10.5,   # ASSUMPTION: westbound ~11h / eastbound ~10h; refine from schedule
        "cabin_seats": {"business": 35, "premium_coach": 25, "coach": 154},   # BA 787-8 config (35J/25W/154Y)
        "total_seats": seats, "seats_per_departure": seats,
        "cabin_load_factor": {"business": None, "premium_coach": None, "coach": None},  # needs cabin demand split
        "total_load_factor": rev["implied_load_factor"][0],
        "avg_ow_fare_point_to_point": round(rev["revenue"]["point_to_point"][0] / rev["passengers"]["point_to_point"][0], 2),
        "avg_ow_fare_connecting": round(cnx_rev_y1 / cnx_pax_y1, 2),
        "avg_ow_fare_blended": round(pax_rev_y1 / grand_total, 2),
        "yield_rev_per_rpk": yield_per_rpk(pax_rev_y1, grand_total, dist_km),
        "prask": prask(pax_rev_y1, ask_yr),
        "passenger_revenue": pax_rev_y1, "cargo_revenue": rev["revenue"]["cargo"][0],
        "ancillary_revenue": rev["revenue"]["ancillary"][0], "total_revenue": total_rev_y1,
        "trask": trask(total_rev_y1, ask_yr),
        "cask": None, "breakeven_load_factor": None,   # model economics module produces these; not in the deck revenue table
    }

    contract = {
        "_contract": "Avia deck data contract v1", "_author": "Avia Solutions",
        "_source": "BA LHR-SJC validated 2015 deck (British Airways Master forecast slides (OS) 06Mar2015); acceptance test",
        "route_metadata": {
            "airline_name": "British Airways", "airline_iata": "BA",
            "origin_airport": "LHR", "origin_city_code": "LON",
            "destination_airport": "SJC", "destination_city_code": "SJC",
            "hub_airport": "LHR", "aircraft_type": "B788", "seats": seats,
            "frequency_per_week": freq, "service_year": 2016,
            "distance_km": dist_km, "distance_nm": dist_nm,
            "catchment_headline": {"point_to_point_market": 249800,
                                   "connecting_market_over_hub": 904500,
                                   "connecting_market_over_destination": 1107200},
        },
        "summary_and_schedule": {
            "point_to_point_market": 249800,
            "connecting_market_over_hub": 904500,
            "connecting_market_over_destination": 1107200,
            "catchment_note": "Based on AviaSolutions' San Jose Service Area catchment analysis",
            "schedule": [
                {"sector": "LHR-SJC", "dep_time": "11:30", "arr_time": "14:30", "operating_days": "Daily",
                 "aircraft": "B788", "seats": seats, "annual_seats": annual_seats // 2,
                 "annual_pax": grand_total // 2, "seat_factor": rev["implied_load_factor"][0]},
                {"sector": "SJC-LHR", "dep_time": "16:30", "arr_time": "10:55+1", "operating_days": "Daily",
                 "aircraft": "B788", "seats": seats, "annual_seats": annual_seats // 2,
                 "annual_pax": grand_total // 2, "seat_factor": rev["implied_load_factor"][0]},
                {"sector": "TOTAL", "dep_time": "", "arr_time": "", "operating_days": "7/wk each way",
                 "aircraft": "B788", "seats": seats, "annual_seats": annual_seats,
                 "annual_pax": grand_total, "seat_factor": rev["implied_load_factor"][0]},
            ],
            "_schedule_times_note": "ASSUMPTION - representative schedule; real times are a deck input",
        },
        "segment_forecast": {
            "rows": seg_rows,
            "summary": {
                "point_to_point_total": {"base_annual_demand": 249800, "demand_at_service_year": 271500,
                    "demand_after_stimulation": 293900, "capture_rate": 0.281, "forecast": p2p_forecast, "pdew": pdew(p2p_forecast)},
                "connecting_at_hub_total": {"base_annual_demand": 904500, "demand_at_service_year": 991300,
                    "demand_after_stimulation": 991300, "capture_rate": 0.045, "forecast": cnx_hub_forecast, "pdew": pdew(cnx_hub_forecast)},
                "connecting_at_destination_total": {"base_annual_demand": 1107200, "demand_at_service_year": 1206800,
                    "demand_after_stimulation": 1206800, "capture_rate": 0.002, "forecast": cnx_dest_forecast, "pdew": pdew(cnx_dest_forecast)},
                "grand_total": {"base_annual_demand": 2261500, "demand_at_service_year": 2469700,
                    "demand_after_stimulation": 2492100, "capture_rate": 0.052, "forecast": grand_total, "pdew": pdew(grand_total)},
            },
            "_competition_buckets": {
                "connecting_at_hub": [{"bucket": "No direct competition", "base": 904500, "capture": 0.045, "forecast": 45000}],
                "connecting_at_destination": [
                    {"bucket": "Direct competition", "base": 977800, "capture": 0.001, "forecast": 800},
                    {"bucket": "No direct competition", "base": 129400, "capture": 0.013, "forecast": 1800}],
            },
        },
        "connecting_at_hub": {"hub": "LHR", "cities": hub_cities,
            "total": {"annual_demand": 991300, "annual_forecast": cnx_hub_forecast, "pdew": pdew(cnx_hub_forecast)}},
        "connecting_at_destination": {"destination": "SJC", "cities": [],
            "total": {"annual_demand": 1206800, "annual_forecast": cnx_dest_forecast, "pdew": pdew(cnx_dest_forecast)},
            "_note": "Per-city table extractable identically to connecting_at_hub; the SJC behind-feed is a 2,628-pax flow (US domestic over SJC). Buckets: direct-comp 800, no-direct-comp 1,800."},
        "revenue_forecast": rev,
        "economics_year1": econ,
        "revenue_build": {
            "by_flow": {"years": [2016, 2017, 2018],
                "point_to_point": rev["revenue"]["point_to_point"],
                "connecting_at_hub": rev["revenue"]["connecting_at_hub"],
                "connecting_at_destination": rev["revenue"]["connecting_at_destination"],
                "cargo": rev["revenue"]["cargo"], "ancillary": rev["revenue"]["ancillary"]},
            "by_cabin": {"years": [2016, 2017, 2018],
                "business": [None, None, None], "premium_coach": [None, None, None], "coach": [None, None, None],
                "_note": "Cabin revenue split needs the cabin fare x cabin pax split (Market Forecast Scenario slide / model cabin allocation); not in the revenue table extract."},
        },
        "catchment": {
            "zones": {
                "primary": {"definition": "Core San Jose service area - the inner catchment where SJC is the clear nearest gateway."},
                "secondary": {"definition": "Wider Bay Area where SJC competes with SFO/OAK on access time."},
                "contested": {"definition": "Outer overlap shared with SFO/OAK; demand split by generalised cost."},
                "_note": "Zone geometry and per-band population/demand come from the catchment module (cell-level apportionment); definitions here, figures wired from the model."},
            "top_markets_beyond_hub": [{"city": c["city_name"], "city_code": c["city_code"], "annual_demand": c["annual_demand"]}
                                       for c in sorted(hub_cities, key=lambda x: -x["annual_demand"]) if c["city_code"] != "Other"][:15],
        },
    }
    return contract


# ----------------------------------------------------------------- generic builder (model-driven)
def _seats_for(aircraft, cabin_config, aircraft_db):
    """(business, premium_coach, coach, total) for the aircraft. cabin_config overrides;
    else the AIRCRAFT 2-class default (business=bus_seats, premium_coach=0, coach=econ_seats)."""
    if cabin_config:
        b = cabin_config.get("business", 0); p = cabin_config.get("premium_coach", 0); c = cabin_config.get("coach", 0)
        return b, p, c, (b + p + c)
    ac = (aircraft_db or {}).get(aircraft, {})
    b = ac.get("bus_seats", 0); c = ac.get("econ_seats", 0)
    return b, 0, c, (b + c)


def _cabin_economics(cabin):
    """From a chosen LOPA + the route's cabin demand + fares, the per-cabin seats, load factor and
    two-way revenue for blocks 7 and 8. cabin = {"lopa","demand_ew","fares","freq"}."""
    lopa, dem, fares, freq = cabin["lopa"], cabin["demand_ew"], cabin["fares"], cabin["freq"]
    cabs = ["first", "business", "premium_coach", "coach"]
    seats = {c: lopa.get(c, 0) for c in cabs}
    cap = {c: seats[c] * freq * 52 for c in cabs}
    filled = {c: min(dem.get(c, 0), cap[c]) for c in cabs}
    lf = {c: (filled[c] / cap[c] if cap[c] else None) for c in cabs}
    rev = {c: round(filled[c] * fares.get(c, 0) * 2) for c in cabs}
    return {"seats": seats, "load_factor": lf, "revenue": rev}


def build_contract(case: dict, outputs: dict, connecting: dict = None, growth_rate: float = None,
                   ancillary_per_pax: float = None, aircraft_db: dict = None, segment_rows: list = None,
                   cabin: dict = None) -> dict:
    """Generic, RouteCase-driven contract assembly from the model's computed outputs - the live
    counterpart to ba_lhr_sjc_reference(). `case` = RouteCase.to_dict(); `outputs` = the assess
    result dict (population, natural, current, capture, directional_demand, frequency, econ_lf,
    bus_lf, route_pnl, annual_pnl, observed_split, dest_od_total). Optional `connecting` =
    connecting_feed result {"hub_cities":[...], "dest_cities":[...], "hub_market", "dest_market"};
    `growth_rate` projects revenue to 3 years; `ancillary_per_pax` adds the ancillary line.

    Fields the model does not yet produce are emitted as None with a sibling _need note, so the
    deck degrades gracefully and the gaps are explicit. This does NOT recompute any forecast number.
    """
    if aircraft_db is None:
        try:
            from aircraft_economics import AIRCRAFT as aircraft_db
        except Exception:
            aircraft_db = {}
    rp = outputs.get("route_pnl", {}) or {}
    ap = outputs.get("annual_pnl", {}) or {}
    freq = outputs.get("frequency") or case.get("frequency")
    aircraft = case.get("aircraft")
    b_seats, p_seats, c_seats, seats = _seats_for(aircraft, case.get("cabin_config"), aircraft_db)
    ce = _cabin_economics(cabin) if cabin else None
    if ce:   # a chosen LOPA overrides the 2-class default
        b_seats = ce["seats"]["business"]; p_seats = ce["seats"]["premium_coach"]; c_seats = ce["seats"]["coach"]
        f_seats = ce["seats"]["first"]; seats = f_seats + b_seats + p_seats + c_seats
    dist_nm = case.get("sector_nm")
    dist_km = round(dist_nm * 1.852, 1) if dist_nm else None
    home = case.get("home"); dest = case.get("primary_dest")
    hub = case.get("hub_airport") or home
    svc_year = case.get("service_year")

    natural = outputs.get("natural") or 0.0
    each_way = outputs.get("directional_demand") or 0.0
    capture = outputs.get("capture")
    carried = ap.get("annual_pax") or round((each_way * 2) * (rp.get("load_factor") or 0))
    annual_seats = seats * freq * 52 * 2 if seats and freq else None
    total_lf = rp.get("load_factor")

    # THE TOTAL IS NOT THE POINT TO POINT LEG. Corrected 14 August 2026, and the cause is a shape
    # change nobody carried through. This module was written against assess(), where
    # directional_demand was the LOCAL leg. cortex_app's payload sets demand.total to carried_ew,
    # and route_forecast line 823 computes carried = min(captured + feed, capacity x plan cap), so
    # the connecting passengers are ALREADY INSIDE IT. Writing that figure into
    # point_to_point_total presented an entire route as local traffic, and grand_total then added
    # the two connecting legs a second time. Measured on SJC-TPE CI 4x 2027: 109,764 carried was
    # labelled point to point, and grand_total read 123,266, which is 96.8% of the aircraft against
    # a plan cap of 87.5%. The impossible load factor was the symptom; the double count was the
    # cause.
    #
    # None rather than a fallback where the payload cannot answer: a contract that quietly
    # substitutes the nearest number to hand is what produced the fault above.
    _p2p_ew = outputs.get("p2p_carried_ew")
    _cnx_ew = outputs.get("connecting_carried_ew")
    p2p_carried = round(_p2p_ew * 2) if _p2p_ew else None
    cnx_carried = round(_cnx_ew * 2) if _cnx_ew else None
    p2p_demand = round((outputs.get("p2p_demand_ew") or 0) * 2) or None
    NEED_LEG = ("the payload's demand.total is the TOTAL carried; run through forecast_to_contract "
                "so p2p_carried_ew and connecting_carried_ew reach this module")

    # THE CONNECTING LEG SPLIT. The city tables cannot supply it: cortex_app._feed_list takes
    # top=15, so both lists are the fifteen largest cities and their sum is a subtotal, not a leg.
    # Measured on the same case, the two tables sum to 13,502 against an implied leg of 26,356.
    # forecast_to_contract's own docstring says to verify this on the first deck; this is that
    # verification, and it failed. The carried leg is therefore split on the pre-cap feed sides as
    # a RATIO, which survives the cap because the cap scales both sides together.
    _fb, _fh = outputs.get("feed_beyond_ew") or 0.0, outputs.get("feed_behind_ew") or 0.0
    _fsum = _fb + _fh
    cnx_hub_carried = round(cnx_carried * _fb / _fsum) if (cnx_carried and _fsum) else None
    cnx_dest_carried = (cnx_carried - cnx_hub_carried) if (cnx_carried and cnx_hub_carried is not None) else None
    NEED_TOPN = ("the fifteen largest cities only, from cortex_app._feed_list(top=15); their sum is "
                 "a subtotal and the leg total is stated separately")

    turns = ap.get("annual_turnarounds") or (freq * 52 * 2 if freq else 0)
    pax_rev_y1 = round((rp.get("econ_rev", 0) + rp.get("bus_rev", 0)) * turns) if turns else None
    cargo_y1 = round(rp.get("cargo_rev", 0) * turns) if turns else None
    anc_y1 = round(ancillary_per_pax * carried) if (ancillary_per_pax and carried) else None
    total_rev_y1 = (pax_rev_y1 or 0) + (cargo_y1 or 0) + (anc_y1 or 0) or ap.get("annual_gross_rev")
    ask_yr = ask(seats, dist_km, freq) if (seats and dist_km and freq) else None

    def proj(v, yrs=3):
        if v is None:
            return [None] * yrs
        if not growth_rate:
            return [v] + [None] * (yrs - 1)
        return [round(v * (1 + growth_rate) ** i) for i in range(yrs)]

    cnx = connecting or {}
    hub_cities = cnx.get("hub_cities") or []
    dest_cities = cnx.get("dest_cities") or []
    # SECOND EACH-WAY FIGURE FOUND IN AN OTHERWISE TWO-WAY CONTRACT (22 August 2026, found while
    # fixing the PDEW departures basis above). connecting_from_forecast hands over hub_cities/
    # dest_cities' annual_demand/annual_forecast RAW, each way, by design (forecast_to_contract's
    # own job is to pass the payload through unmodified; deck_contract is the one place that
    # doubles, per the _hub_mkt2/_dest_mkt2 fix from 20 August, "doubled once here, at the single
    # place both legs and every consumer draw from"). The per-city cities[] list and its own total
    # were never brought into that rule, so connecting_at_hub.total.annual_forecast sat at roughly
    # half of segment_forecast.summary.connecting_at_hub_total.forecast in the SAME contract, two
    # "connecting at hub" figures a page apart reading a factor of two off each other, which is
    # exactly the shape of confusion this whole EW/two-way project exists to close.
    cnx_hub_fc = round(sum(c.get("annual_forecast", 0) for c in hub_cities) * 2) if hub_cities else None
    cnx_dest_fc = round(sum(c.get("annual_forecast", 0) for c in dest_cities) * 2) if dest_cities else None

    # CONNECTING MARKET SIZE, TWO-WAY (20 August 2026, Jol's review of the SJC-TPE packs, the
    # basis mix he and John caught live: "connecting market over Taipei 719,500 both directions...
    # but this says each way"). cnx["hub_market"]/["dest_market"] arrive EACH WAY from
    # connecting_from_forecast (dem["feed_beyond_base"]/["feed_behind_base"], the same figure
    # verify_connecting_build.py itself names "each way"). Every OTHER figure this function builds
    # - natural*2 four lines above, p2p_carried, cnx_carried, p2p_demand a few lines above - is
    # explicitly doubled to two-way, so the connecting legs' market size was the one each-way
    # number left in an otherwise two-way contract. Two measured consequences: the process-visual
    # chart printed this each-way figure captioned "both directions", and the connecting legs'
    # capture rate below divided a two-way carried figure by this each-way market, reading roughly
    # double the true rate (circa 8.1% shown where circa 4.0% is correct on the TPE-beyond leg).
    # Doubled once here, at the single place both legs and every consumer draw from.
    _hub_mkt2 = round((cnx.get("hub_market") or 0) * 2) or None
    _dest_mkt2 = round((cnx.get("dest_market") or 0) * 2) or None

    NEED_SEG = "8-segment split needs business/leisure ratio + per-zone demand + per-segment growth & capture (see field note)"
    NEED_CNX_DEST = "behind-destination home feed needs the home side of the connecting layer"

    contract = {
        "_contract": "Avia deck data contract v1", "_author": "Avia Solutions",
        "_source": f"model run: assess() for {outputs.get('case_id', case.get('case_id'))}",
        "route_metadata": {
            "airline_name": case.get("airline_name") or "Generic (airline-agnostic)",
            "airline_iata": case.get("airline_iata") or "",
            "origin_airport": home, "origin_city_code": None, "_origin_city_need": "airport->city lookup",
            "destination_airport": dest, "destination_city_code": None, "_dest_city_need": "airport->city lookup",
            "hub_airport": hub, "aircraft_type": aircraft, "seats": seats,
            "frequency_per_week": freq, "service_year": svc_year,
            "distance_km": dist_km, "distance_nm": dist_nm,
            "catchment_headline": {"point_to_point_market": round(natural * 2),
                                   "connecting_market_over_hub": _hub_mkt2,
                                   "connecting_market_over_destination": _dest_mkt2,
                                   "_connecting_need": None if cnx else NEED_CNX_DEST},
        },
        "summary_and_schedule": {
            "point_to_point_market": round(natural * 2),
            "connecting_market_over_hub": _hub_mkt2,
            "connecting_market_over_destination": _dest_mkt2,
            "catchment_note": f"Based on AviaSolutions' {home} Service Area catchment analysis",
            "schedule": [
                {"sector": f"{home}-{dest}", "dep_time": None, "arr_time": None, "operating_days": f"{freq}/wk",
                 "aircraft": aircraft, "seats": seats,
                 "annual_seats": (annual_seats // 2 if annual_seats else None),
                 "annual_pax": (carried // 2 if carried else None), "seat_factor": total_lf},
                {"sector": f"{dest}-{home}", "dep_time": None, "arr_time": None, "operating_days": f"{freq}/wk",
                 "aircraft": aircraft, "seats": seats,
                 "annual_seats": (annual_seats // 2 if annual_seats else None),
                 "annual_pax": (carried // 2 if carried else None), "seat_factor": total_lf},
                {"sector": "TOTAL", "dep_time": "", "arr_time": "", "operating_days": f"{freq}/wk each way",
                 "aircraft": aircraft, "seats": seats, "annual_seats": annual_seats,
                 "annual_pax": carried, "seat_factor": total_lf},
            ],
            "_schedule_times_need": "dep/arr times are a schedule input assumption",
        },
        "segment_forecast": {
            "rows": (segment_rows or []),
            "_rows_need": (None if segment_rows else NEED_SEG),
            "_rows_source": ("segment_model.build_segment_table (route-current Sabre cabin mix + zone bands + analyst capture)" if segment_rows else None),
            "summary": {
                "point_to_point_total": {"base_annual_demand": round(natural * 2),
                    # demand_at_service_year is the leg BEFORE stimulation and the payload does not
                    # carry it, so it is named as a gap rather than filled with the figure beside it.
                    "demand_at_service_year": None,
                    "_demand_at_service_year_need": "no pre-stimulation local leg in the payload",
                    "demand_after_stimulation": p2p_demand,
                    "capture_rate": capture,
                    "forecast": p2p_carried,
                    "_forecast_need": (None if p2p_carried else NEED_LEG),
                    "pdew": pdew(p2p_carried or 0, freq=freq)},
                "connecting_at_hub_total": ({"base_annual_demand": _hub_mkt2,
                    "demand_at_service_year": _hub_mkt2, "demand_after_stimulation": _hub_mkt2,
                    "capture_rate": (airline_share(cnx_hub_carried, _hub_mkt2) if _hub_mkt2 else None),
                    "forecast": cnx_hub_carried,
                    "_forecast_need": (None if cnx_hub_carried else NEED_LEG),
                    "top_cities_forecast": cnx_hub_fc, "_top_cities_need": NEED_TOPN,
                    "pdew": pdew(cnx_hub_carried or 0, freq=freq)} if hub_cities else
                    {"forecast": None, "_need": NEED_CNX_DEST}),
                # base_annual_demand/demand_at_service_year/demand_after_stimulation/capture_rate
                # added here 20 August 2026, same fix as the hub leg above: this block previously
                # carried forecast only, so the printed table's demand column read "-" for this leg
                # while the hub leg above showed a number, an asymmetry with no reason behind it now
                # that _dest_mkt2 is available on the same two-way basis as everything around it.
                "connecting_at_destination_total": ({"base_annual_demand": _dest_mkt2,
                    "demand_at_service_year": _dest_mkt2, "demand_after_stimulation": _dest_mkt2,
                    "capture_rate": (airline_share(cnx_dest_carried, _dest_mkt2) if _dest_mkt2 else None),
                    "forecast": cnx_dest_carried,
                    "_forecast_need": (None if cnx_dest_carried else NEED_LEG),
                    "top_cities_forecast": cnx_dest_fc, "_top_cities_need": NEED_TOPN,
                    "pdew": pdew(cnx_dest_carried or 0, freq=freq)}
                    if dest_cities else {"forecast": None, "_need": NEED_CNX_DEST}),
                # carried ALREADY contains both connecting legs. Adding them here is what produced a
                # load factor above the plan cap, so the total is taken and never summed.
                "grand_total": {"forecast": carried, "pdew": pdew(carried or 0, freq=freq),
                    "_basis": "carried, after the plan load factor cap; the legs below sum to it"},
            },
        },
        "connecting_at_hub": {"hub": hub,
            "cities": [{"nr": i + 1, "city_code": c.get("city_code") or c.get("market"),
                        "city_name": c.get("city_name"), "country": c.get("country"),
                        "annual_demand": round((c.get("annual_demand") or 0) * 2) or None,
                        "airline_share": c.get("airline_share"),
                        "annual_forecast": round((c.get("annual_forecast") or 0) * 2) or None,
                        "pdew": pdew((c.get("annual_forecast") or 0) * 2, freq=freq)}
                       for i, c in enumerate(hub_cities)],
            "total": {"annual_forecast": cnx_hub_fc, "pdew": pdew(cnx_hub_fc or 0, freq=freq)},
            "_need": (None if hub_cities else "run connecting_feed and pass connecting=")},
        # dest_cities was passed straight through un-doubled until today - the same gap as the hub
        # side above, fixed the same way and now rebuilt here rather than passed through, so both
        # legs of the connecting table double at the one point, not two.
        "connecting_at_destination": {"destination": dest,
            "cities": [{"nr": i + 1, "city_code": c.get("city_code") or c.get("market"),
                        "city_name": c.get("city_name"), "country": c.get("country"),
                        "annual_demand": round((c.get("annual_demand") or 0) * 2) or None,
                        "airline_share": c.get("airline_share"),
                        "annual_forecast": round((c.get("annual_forecast") or 0) * 2) or None,
                        "pdew": pdew((c.get("annual_forecast") or 0) * 2, freq=freq)}
                       for i, c in enumerate(dest_cities)],
            "total": {"annual_forecast": cnx_dest_fc, "pdew": pdew(cnx_dest_fc or 0, freq=freq)},
            "_need": (None if dest_cities else NEED_CNX_DEST)},
        "revenue_forecast": {
            "years": ([svc_year, (svc_year + 1 if svc_year else None), (svc_year + 2 if svc_year else None)]),
            "passengers": {"point_to_point": proj(carried), "connecting_at_hub": proj(cnx_hub_fc),
                           "connecting_at_destination": proj(cnx_dest_fc),
                           "total": proj((carried or 0) + (cnx_hub_fc or 0) + (cnx_dest_fc or 0))},
            "annual_capacity": [annual_seats, annual_seats, annual_seats],
            "revenue": {"point_to_point": proj(pax_rev_y1), "cargo": proj(cargo_y1),
                        "ancillary": proj(anc_y1), "total": proj(total_rev_y1),
                        "_connecting_revenue_need": (None if hub_cities else "needs connecting pax x connecting fare")},
            "_projection_need": (None if growth_rate else "years 2-3 need a growth rate"),
            "_ancillary_need": (None if ancillary_per_pax else "ancillary needs a per-pax benchmark"),
        },
        "economics_year1": {
            "equipment": aircraft, "weekly_departures": freq,
            "total_departures_annual_two_way": (freq * 52 * 2 if freq else None),
            "block_hours_per_departure": (round(case.get("block_min") / 60.0, 2) if case.get("block_min") else None),
            "cabin_seats": ({"first": ce["seats"]["first"], "business": b_seats, "premium_coach": p_seats, "coach": c_seats}
                            if ce else {"business": b_seats, "premium_coach": p_seats, "coach": c_seats,
                            "_need": (None if case.get("cabin_config") else "3-class split needs a cabin config / LOPA; using the 2-class default")}),
            "total_seats": seats, "seats_per_departure": seats,
            "cabin_load_factor": (ce["load_factor"] if ce else
                {"business": outputs.get("bus_lf"), "premium_coach": None, "coach": outputs.get("econ_lf"),
                 "_need": "per-cabin LF needs a chosen LOPA + cabin demand (pass cabin=)"}),
            "total_load_factor": total_lf,
            "avg_ow_fare_point_to_point": (round(pax_rev_y1 / carried, 2) if (pax_rev_y1 and carried) else None),
            "avg_ow_fare_connecting": None, "avg_ow_fare_blended": (round(pax_rev_y1 / carried, 2) if (pax_rev_y1 and carried) else None),
            "yield_rev_per_rpk": (yield_per_rpk(pax_rev_y1, carried, dist_km) if (pax_rev_y1 and carried and dist_km) else None),
            "prask": (prask(pax_rev_y1, ask_yr) if (pax_rev_y1 and ask_yr) else None),
            "passenger_revenue": pax_rev_y1, "cargo_revenue": cargo_y1, "ancillary_revenue": anc_y1,
            "total_revenue": total_rev_y1, "trask": (trask(total_rev_y1, ask_yr) if (total_rev_y1 and ask_yr) else None),
            "cask": rp.get("cask"), "breakeven_load_factor": rp.get("breakeven_lf"),
        },
        "revenue_build": {
            "by_flow": {"years": [svc_year], "point_to_point": [pax_rev_y1], "connecting_at_hub": [None],
                        "connecting_at_destination": [None], "cargo": [cargo_y1], "ancillary": [anc_y1]},
            "by_cabin": ({"years": [svc_year], "first": [ce["revenue"]["first"]], "business": [ce["revenue"]["business"]],
                          "premium_coach": [ce["revenue"]["premium_coach"]], "coach": [ce["revenue"]["coach"]]}
                         if ce else {"years": [svc_year], "business": [None], "premium_coach": [None], "coach": [None],
                         "_need": "cabin revenue split needs a chosen LOPA + cabin demand (pass cabin=)"}),
        },
        "catchment": {
            "zones": {"primary": {"definition": f"Core {home} service area."},
                      "secondary": {"definition": f"Wider area where {home} competes on access time."},
                      "contested": {"definition": "Outer overlap split by generalised cost."},
                      "_observed_split": outputs.get("observed_split"),
                      "_note": "Zone geometry/population from the catchment module; the observed split is the apportionment."},
            # same each-way source as the connecting_at_hub cities above; doubled here too, or this
            # block would print a third, again-different "connecting market" figure per city.
            "top_markets_beyond_hub": ([{"city": c.get("city_name"), "city_code": c.get("city_code") or c.get("market"),
                                         "annual_demand": round((c.get("annual_demand") or 0) * 2) or None}
                                        for c in sorted(hub_cities, key=lambda x: -(x.get("annual_demand") or 0))[:15]]
                                       if hub_cities else []),
        },
    }
    return contract


# ----------------------------------------------------------------- writers
def emit_json(contract: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(contract, f, indent=2, ensure_ascii=False)
    return path


def _author(wb):
    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"


def emit_workbook(contract: dict, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    NAVY = "FF1F3864"; CYAN = "FFDDEBF7"; HEADF = Font(bold=True, color="FFFFFFFF", name="Arial")
    HEAD = PatternFill("solid", fgColor=NAVY); SUB = PatternFill("solid", fgColor=CYAN)
    BOLD = Font(bold=True, name="Arial"); BODY = Font(name="Arial")
    wb = Workbook(); _author(wb)
    ws0 = wb.active; ws0.title = "Contract"
    rm = contract["route_metadata"]
    # THE SAME PDEW FIX AS build_contract() (22 August 2026), applied to the Excel formulas below,
    # which hardcoded DAYS_2WAY as literal text and so bypassed pdew() entirely: the route's real
    # scheduled departures, two-way, or the old daily-service fallback when frequency is unknown.
    _dep2 = round((rm.get("frequency_per_week") or 0) * 52 * 2) or DAYS_2WAY
    info = [["Avia deck data contract", ""], ["Author", "Avia Solutions"],
            ["Route", f"{rm['airline_name']} {rm['origin_airport']}-{rm['destination_airport']}"],
            ["Service year", rm["service_year"]], ["Aircraft", f"{rm['aircraft_type']} ({rm['seats']} seats)"],
            ["Frequency/week", rm["frequency_per_week"]],
            ["Distance", f"{rm['distance_km']:.0f} km / {rm['distance_nm']:.0f} nm"], ["", ""],
            ["Source", contract.get("_source", "")]]
    for r in info:
        ws0.append(r)
    ws0["A1"].font = Font(bold=True, size=14, name="Arial", color=NAVY)
    for c in "AB":
        ws0.column_dimensions[c].width = 34

    def header(ws, cols):
        ws.append(cols)
        for j in range(1, len(cols) + 1):
            cell = ws.cell(row=ws.max_row, column=j); cell.font = HEADF; cell.fill = HEAD
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def widths(ws, w):
        for i, x in enumerate(w):
            ws.column_dimensions[chr(65 + i)].width = x

    # 1 Route metadata
    ws = wb.create_sheet("1_Route_Metadata")
    header(ws, ["Field", "Value"])
    ch = rm["catchment_headline"]
    for k, v in [("Airline", rm["airline_name"]), ("Airline IATA", rm["airline_iata"]),
                 ("Origin airport", rm["origin_airport"]), ("Origin city code", rm["origin_city_code"]),
                 ("Destination airport", rm["destination_airport"]), ("Destination city code", rm["destination_city_code"]),
                 ("Hub airport", rm["hub_airport"]), ("Aircraft type", rm["aircraft_type"]), ("Seats", rm["seats"]),
                 ("Frequency per week", rm["frequency_per_week"]), ("Service year", rm["service_year"]),
                 ("Distance (km)", rm["distance_km"]), ("Distance (nm)", rm["distance_nm"]),
                 ("P2P market", ch["point_to_point_market"]), ("Connecting market over hub", ch["connecting_market_over_hub"]),
                 ("Connecting market over destination", ch["connecting_market_over_destination"])]:
        ws.append([k, v]); ws.cell(row=ws.max_row, column=1).font = BODY
    widths(ws, [34, 22])

    # 2 Summary & schedule
    ws = wb.create_sheet("2_Summary_Schedule")
    ss = contract["summary_and_schedule"]
    ws.append(["Point to point market", ss["point_to_point_market"]])
    ws.append(["Connecting market over hub", ss["connecting_market_over_hub"]])
    ws.append(["Connecting market over destination", ss["connecting_market_over_destination"]])
    ws.append([ss["catchment_note"]]); ws.append([])
    header(ws, ["Sector", "Dep", "Arr", "Op Days", "Aircraft", "Seats", "Annual Seats", "Annual Pax", "Seat Factor"])
    sched_start = ws.max_row + 1
    for s in ss["schedule"]:
        ws.append([s["sector"], s["dep_time"], s["arr_time"], s["operating_days"], s["aircraft"],
                   s["seats"], s["annual_seats"], s["annual_pax"], s["seat_factor"]])
        ws.cell(row=ws.max_row, column=9).number_format = "0.0%"
    widths(ws, [12, 9, 10, 12, 10, 8, 13, 12, 11])

    # 3 Segment forecast
    ws = wb.create_sheet("3_Segment_Forecast")
    header(ws, ["Market segment", "Base Annual Demand", "Annual Growth Rate", "Demand at Service Year",
                "Stimulation", "Demand After Stimulation", "Capture Rate", "Forecast", "PTEW"])
    first = ws.max_row + 1
    for r in contract["segment_forecast"]["rows"]:
        ws.append([r["segment"], r["base_annual_demand"], r["annual_growth_rate"], r["demand_at_service_year"],
                   r["stimulation_factor"], r["demand_after_stimulation"], r["capture_rate"], r["forecast"], None])
        row = ws.max_row
        ws.cell(row=row, column=9).value = f"=H{row}/{_dep2}"          # PTEW formula
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=7).number_format = "0.0%"
        ws.cell(row=row, column=9).number_format = "0.0"
    last = ws.max_row
    for key, lbl in [("point_to_point_total", "Point-to-point total"),
                     ("connecting_at_hub_total", "Connecting at hub total"),
                     ("connecting_at_destination_total", "Connecting at destination total"),
                     ("grand_total", "GRAND TOTAL")]:
        # A MISSING FIGURE IS A BLANK CELL, NOT A CRASH. build_contract writes the full set of
        # demand columns for point_to_point_total and connecting_at_hub_total, and only forecast and
        # pdew for connecting_at_destination_total and grand_total, so this line raised
        # KeyError 'base_annual_demand' on every live contract. It had only ever been run against
        # ba_lhr_sjc_reference(), which fills all four rows by hand, so no live output had ever
        # reached the workbook: the JSON wrote and the .xlsx did not.
        #
        # .get() rather than a new required key, because those two rows genuinely have no base-year
        # demand to state: the connecting build starts at the market, not at a base that grew into
        # it. A blank says that and a zero would not.
        s = contract["segment_forecast"]["summary"][key]
        ws.append([lbl, s.get("base_annual_demand"), None, s.get("demand_at_service_year"), None,
                   s.get("demand_after_stimulation"), s.get("capture_rate"), s.get("forecast"), None])
        row = ws.max_row
        ws.cell(row=row, column=9).value = f"=H{row}/{_dep2}"
        ws.cell(row=row, column=7).number_format = "0.0%"; ws.cell(row=row, column=9).number_format = "0.0"
        for j in range(1, 10):
            ws.cell(row=row, column=j).font = BOLD
    widths(ws, [34, 16, 14, 16, 11, 18, 12, 12, 8])

    # 4 Connecting at hub
    ws = wb.create_sheet("4_Connecting_at_Hub")
    header(ws, ["Nr", "City Code", "City Name", "Country", "Annual Demand", "Airline Share", "Annual Forecast", "PTEW"])
    for c in contract["connecting_at_hub"]["cities"]:
        ws.append([c["nr"], c["city_code"], c["city_name"], c["country"], c["annual_demand"], None, c["annual_forecast"], None])
        row = ws.max_row
        ws.cell(row=row, column=6).value = f"=IF(E{row}=0,0,G{row}/E{row})"   # airline_share formula
        ws.cell(row=row, column=8).value = f"=G{row}/{_dep2}"             # ptew
        ws.cell(row=row, column=6).number_format = "0.0%"; ws.cell(row=row, column=8).number_format = "0.0"
    n = ws.max_row
    ws.append(["", "", "TOTAL", "", f"=SUM(E2:E{n})", None, f"=SUM(G2:G{n})", None])
    row = ws.max_row
    ws.cell(row=row, column=6).value = f"=IF(E{row}=0,0,G{row}/E{row})"; ws.cell(row=row, column=6).number_format = "0.0%"
    ws.cell(row=row, column=8).value = f"=G{row}/{_dep2}"; ws.cell(row=row, column=8).number_format = "0.0"
    for j in range(1, 9):
        ws.cell(row=row, column=j).font = BOLD
    widths(ws, [5, 10, 18, 18, 14, 13, 14, 8])

    # 5 Connecting at destination
    ws = wb.create_sheet("5_Connecting_at_Dest")
    header(ws, ["Nr", "City Code", "City Name", "Country", "Annual Demand", "Airline Share", "Annual Forecast", "PTEW"])
    cd = contract["connecting_at_destination"]
    for c in cd["cities"]:
        ws.append([c["nr"], c["city_code"], c["city_name"], c["country"], c["annual_demand"], None, c["annual_forecast"], None])
    # THE SAME MISMATCH AS THE SUMMARY ROWS, on the destination total. build_contract writes
    # connecting_at_destination["total"] with annual_forecast and pdew only, and never an
    # annual_demand, so this raised KeyError 'annual_demand' on every live contract. The hub side
    # is not affected because build_contract builds that total itself with the full set.
    #
    # .get() again rather than a new required key: the destination total genuinely has no base-year
    # demand figure to state, and a blank cell says that where a zero would not.
    _cdt = cd.get("total") or {}
    ws.append(["", "", "TOTAL", "", _cdt.get("annual_demand"), None,
               _cdt.get("annual_forecast"), _cdt.get("pdew")])
    for j in range(1, 9):
        ws.cell(row=ws.max_row, column=j).font = BOLD
    ws.append([]); ws.append([cd.get("_note", "")])
    widths(ws, [5, 10, 18, 18, 14, 13, 14, 8])

    # 6 Revenue 3yr
    ws = wb.create_sheet("6_Revenue_3yr")
    rv = contract["revenue_forecast"]; yrs = rv["years"]
    header(ws, ["Passenger demand"] + [str(y) for y in yrs])
    for k, lbl in [("point_to_point", "Point to point"), ("connecting_at_hub", "Connecting at hub"),
                   ("connecting_at_destination", "Connecting at destination"), ("total", "Total")]:
        ws.append([lbl] + (rv["passengers"].get(k) or [None] * len(yrs)))
        if k == "total":
            for j in range(2, 5):
                ws.cell(row=ws.max_row, column=j).font = BOLD
    ws.append(["Annual capacity"] + rv["annual_capacity"])
    cap_row = ws.max_row; tot_row = cap_row - 1
    ws.append(["Implied load factor"] + [f"={chr(65+j)}{tot_row}/{chr(65+j)}{cap_row}" for j in range(1, 4)])
    for j in range(2, 5):
        ws.cell(row=ws.max_row, column=j).number_format = "0.0%"
    ws.append([])
    header(ws, ["Revenue ($)"] + [str(y) for y in yrs])
    for k, lbl in [("point_to_point", "Point to point"), ("connecting_at_hub", "Connecting at hub"),
                   ("connecting_at_destination", "Connecting at destination"), ("cargo", "Cargo"),
                   ("ancillary", "Ancillary"), ("total", "Total")]:
        # THE BUILDER PRODUCES FOUR OF THESE SIX AND SAYS SO. revenue_forecast["revenue"] holds
        # point_to_point, cargo, ancillary and total, and carries _connecting_revenue_need reading
        # "needs connecting pax x connecting fare", so the two connecting lines are a stated gap
        # rather than an oversight. This loop asked for all six and raised KeyError
        # 'connecting_at_hub' on every live contract.
        #
        # A missing flow becomes an EMPTY ROW of the right width rather than a crash, so the sheet
        # still shows the line and leaves it blank, which is what the _need note describes.
        ws.append([lbl] + (rv["revenue"].get(k) or [None] * len(yrs)))
        for j in range(2, 5):
            ws.cell(row=ws.max_row, column=j).number_format = "$#,##0"
            if k == "total":
                ws.cell(row=ws.max_row, column=j).font = BOLD
    widths(ws, [26, 16, 16, 16])

    # 7 Economics Yr1
    ws = wb.create_sheet("7_Economics_Yr1")
    e = contract["economics_year1"]; cs = e["cabin_seats"]
    header(ws, ["Metric", "Value"])
    rows7 = [("Equipment", e["equipment"]), ("Weekly departures", e["weekly_departures"]),
             ("Total departures (annual, 2-way)", e["total_departures_annual_two_way"]),
             ("Block hours per departure", e["block_hours_per_departure"]),
             ("Business seats", cs["business"]), ("Premium coach seats", cs["premium_coach"]),
             ("Coach seats", cs["coach"]), ("Total seats", e["total_seats"]),
             ("Seats per departure", e["seats_per_departure"]), ("Total load factor", e["total_load_factor"]),
             ("Avg OW fare - point to point", e["avg_ow_fare_point_to_point"]),
             ("Avg OW fare - connecting", e["avg_ow_fare_connecting"]),
             ("Avg OW fare - blended", e["avg_ow_fare_blended"]),
             ("Yield (rev/RPK)", e["yield_rev_per_rpk"]), ("PRASK", e["prask"]),
             ("Passenger revenue", e["passenger_revenue"]), ("Cargo revenue", e["cargo_revenue"]),
             ("Ancillary revenue", e["ancillary_revenue"]), ("Total revenue", e["total_revenue"]),
             ("TRASK", e["trask"]), ("CASK", e["cask"]), ("Breakeven load factor", e["breakeven_load_factor"])]
    for k, v in rows7:
        ws.append([k, v]); ws.cell(row=ws.max_row, column=1).font = BODY
        if "load factor" in k.lower() and isinstance(v, float):
            ws.cell(row=ws.max_row, column=2).number_format = "0.0%"
        if "revenue" in k.lower() and isinstance(v, (int, float)):
            ws.cell(row=ws.max_row, column=2).number_format = "$#,##0"
        if "fare" in k.lower() and isinstance(v, (int, float)):
            ws.cell(row=ws.max_row, column=2).number_format = "$#,##0"
    widths(ws, [32, 18])

    # 8 Revenue build
    ws = wb.create_sheet("8_Revenue_Build")
    rb = contract["revenue_build"]
    ws.append(["Revenue build by flow ($)"]); ws.cell(row=ws.max_row, column=1).font = BOLD
    header(ws, ["Flow"] + [str(y) for y in rb["by_flow"]["years"]])
    for k, lbl in [("point_to_point", "Point to point"), ("connecting_at_hub", "Connecting at hub"),
                   ("connecting_at_destination", "Connecting at destination"), ("cargo", "Cargo"), ("ancillary", "Ancillary")]:
        # Same shape as the revenue_forecast loop above: by_flow carries the flows the builder can
        # fill and this asked for all five.
        ws.append([lbl] + (rb["by_flow"].get(k) or [None] * len(rb["by_flow"].get("years") or [])))
        for j in range(2, 5):
            ws.cell(row=ws.max_row, column=j).number_format = "$#,##0"
    ws.append([])
    ws.append(["Revenue build by cabin ($)"]); ws.cell(row=ws.max_row, column=1).font = BOLD
    header(ws, ["Cabin"] + [str(y) for y in rb["by_cabin"]["years"]])
    for k, lbl in [("business", "Business"), ("premium_coach", "Premium coach"), ("coach", "Coach")]:
        ws.append([lbl] + rb["by_cabin"][k])
    ws.append([]); ws.append([rb["by_cabin"].get("_note", "")])
    widths(ws, [26, 16, 16, 16])

    # 9 Catchment
    ws = wb.create_sheet("9_Catchment")
    cat = contract["catchment"]
    header(ws, ["Zone", "Definition"])
    for z in ("primary", "secondary", "contested"):
        ws.append([z.title(), cat["zones"][z]["definition"]]); ws.cell(row=ws.max_row, column=1).font = BODY
    ws.append([]); ws.append(["Top markets beyond hub", ""]); ws.cell(row=ws.max_row, column=1).font = BOLD
    header(ws, ["City", "City Code", "Annual Demand"])
    for m in cat["top_markets_beyond_hub"]:
        ws.append([m["city"], m["city_code"], m["annual_demand"]])
    widths(ws, [22, 12, 16])

    wb.save(path)
    return path


def emit_from_assess(case_path, assess_out_path, out_dir=".", connecting_path=None,
                     growth_rate=None, ancillary_per_pax=None, stem=None):
    """Emit the contract from a live model run: a RouteCase JSON + the assess() output JSON
    (and optionally a connecting_feed JSON). This is the wiring the deck generator calls."""
    case = json.load(open(case_path))
    outputs = json.load(open(assess_out_path))
    connecting = json.load(open(connecting_path)) if connecting_path else None
    c = build_contract(case, outputs, connecting=connecting, growth_rate=growth_rate,
                       ancillary_per_pax=ancillary_per_pax)
    stem = stem or f"{outputs.get('case_id', case.get('case_id', 'route'))}_deck_contract"
    j = emit_json(c, os.path.join(out_dir, stem + ".json"))
    x = emit_workbook(c, os.path.join(out_dir, stem + ".xlsx"))
    return c, j, x


def main():
    ap = argparse.ArgumentParser(description="Emit a deck data contract (JSON + workbook).")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--case", default=None, help="RouteCase JSON (model-run path; with --assess)")
    ap.add_argument("--assess", default=None, help="assess() output JSON (the <case_id>_case.json)")
    ap.add_argument("--connecting", default=None, help="optional connecting_feed JSON")
    ap.add_argument("--growth", type=float, default=None, help="annual growth rate for the 3-year revenue projection")
    ap.add_argument("--ancillary-per-pax", type=float, default=None)
    a = ap.parse_args()
    if a.case and a.assess:
        c, j, x = emit_from_assess(a.case, a.assess, a.out_dir, a.connecting, a.growth, a.ancillary_per_pax)
        print("wrote", j); print("wrote", x)
        pt = c["segment_forecast"]["summary"]["point_to_point_total"]
        print("P2P forecast", pt["forecast"], "PTEW", pt["pdew"], "| aircraft",
              c["route_metadata"]["aircraft_type"], c["route_metadata"]["seats"], "seats")
        return
    # default: the BA reference worked example
    c = ba_lhr_sjc_reference()
    j = emit_json(c, os.path.join(a.out_dir, "ba_lhr_sjc_deck_contract.json"))
    x = emit_workbook(c, os.path.join(a.out_dir, "ba_lhr_sjc_deck_contract.xlsx"))
    print("wrote", j); print("wrote", x)
    print("P2P forecast", c["segment_forecast"]["summary"]["point_to_point_total"]["forecast"],
          "PTEW", c["segment_forecast"]["summary"]["point_to_point_total"]["pdew"])
    print("Grand total", c["segment_forecast"]["summary"]["grand_total"]["forecast"],
          "PTEW", c["segment_forecast"]["summary"]["grand_total"]["pdew"])


if __name__ == "__main__":
    main()
