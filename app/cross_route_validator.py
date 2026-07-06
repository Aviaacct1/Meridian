#!/usr/bin/env python3
"""
Avia Solutions  Cross-Route Validation Framework (Chat 20)
============================================================
Tests the forecast pipeline computation logic against 5 completed
case studies from Avia Solutions presentations.

Validation approach:
    Since we only have raw QSI Excel data for BA LHR-SJC, we can't run the
    full pipeline end-to-end for other routes. Instead, we use the presentation
    outputs as BOTH input data and validation targets to verify:

    1. FORECAST TABLE MATHEMATICS  given base demand, growth, stimulation,
       and capture rates, does our assembly logic produce the correct forecasts?
    2. PTEW CALCULATIONS  do computed PTEW values match presentations?
    3. LOAD FACTOR  does capacity  frequency  weeks produce correct LF?
    4. CONNECTING CITY AGGREGATION  do individual city forecasts sum correctly?
    5. PARAMETER REASONABLENESS  are calibration parameters within expected
       ranges for each route type?
    6. CROSS-ROUTE PATTERN ANALYSIS  do calibration patterns vary as expected
       across different carrier types, hub sizes, and market characteristics?

Case studies:
    1. BA LHR-SJC (2015)   Full-service, major hub, Europe-US, 7x weekly
    2. KE ICN-SJC (2024)   Full-service, Asian hub, Korea-US, 7x & 5x weekly
    3. SQ SIN-SJC (2024)   Full-service, Asian hub, Singapore-US, 4x weekly
    4. CX HKG-SJC (2025)   Full-service, Asian hub, HK-US, 4x weekly
    5. FI KEF-SJC (2023)   Full-service, small hub, Iceland-US, 4x weekly
"""

from config import OUTPUT_DIR, ensure_output_dir
import sys
import os
import json
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Tuple
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class ConnectingMarketTarget:
    """Connecting market from presentation forecast table."""
    hub_code: str
    hub_name: str
    base_demand: float          # '000s
    growth_compound: float      # e.g. 0.127 for 12.7%
    demand_after_growth: float  # '000s
    stimulation: float          # e.g. 1.00
    demand_after_stim: float    # '000s
    capture_rate: float         # blended across direct/no-direct
    forecast_pax: float         # '000s
    ptew: float
    # Breakdown if available
    direct_comp_demand: float = 0
    direct_comp_capture: float = 0
    direct_comp_forecast: float = 0
    no_direct_demand: float = 0
    no_direct_capture: float = 0
    no_direct_forecast: float = 0


@dataclass
class CityForecast:
    """Individual connecting city from presentation tables."""
    city_code: str
    city_name: str
    country: str
    base_demand: float
    capture_rate: float
    forecast_pax: float
    ptew: float


@dataclass
class RouteValidationCase:
    """Complete validation case extracted from a presentation."""
    # Route identity
    airline_name: str
    airline_code: str
    origin: str
    destination: str
    frequency: int           # flights per week
    aircraft: str
    seats: int
    base_year: int
    forecast_year: int

    # P2P market
    p2p_base_demand: float   # '000s
    p2p_growth: float
    p2p_demand_grown: float  # '000s
    p2p_stimulation: float
    p2p_demand_stim: float   # '000s
    p2p_capture_rate: float
    p2p_forecast: float      # '000s
    p2p_ptew: float

    # Connecting markets
    connecting_markets: List[ConnectingMarketTarget] = field(default_factory=list)

    # Grand total from presentation
    grand_total_pax: float = 0    # '000s
    grand_total_ptew: float = 0
    annual_seats: int = 0
    load_factor: float = 0

    # Connecting city detail (top cities)
    hub_cities: Dict[str, List[CityForecast]] = field(default_factory=dict)

    @property
    def computed_annual_seats(self) -> int:
        return self.seats * self.frequency * 52 * 2

    @property
    def total_connecting(self) -> float:
        return sum(m.forecast_pax for m in self.connecting_markets)

    @property
    def computed_grand_total(self) -> float:
        return self.p2p_forecast + self.total_connecting


# ============================================================================
# CASE STUDY DEFINITIONS  Extracted from presentations
# ============================================================================

def case_ba_lhr_sjc() -> RouteValidationCase:
    """BA LHR-SJC from BA Fcst LHRSJC (Feb 2015), 7x weekly, without India."""
    case = RouteValidationCase(
        airline_name="British Airways", airline_code="BA",
        origin="LHR", destination="SJC",
        frequency=7, aircraft="787-800", seats=214,
        base_year=2013, forecast_year=2015,
        # P2P  from forecast file (4 segments combined)
        p2p_base_demand=249.8,  # sum of UK Biz + UK Leis + US Biz + US Leis base
        p2p_growth=0.10,
        p2p_demand_grown=274.8,  # approx
        p2p_stimulation=1.0,  # blended (some segments have 1.15)
        p2p_demand_stim=274.8,
        p2p_capture_rate=0.284,  # blended across segments
        p2p_forecast=78.110,
        p2p_ptew=107.3,  # 78110 / (7*52)
        grand_total_pax=129.162,
        grand_total_ptew=177.4,  # 129162 / (7*52)
        annual_seats=155792,  # 214 * 7 * 52 * 2
        load_factor=0.829,
    )
    # Connecting at LHR (home hub)
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="LHR", hub_name="London Heathrow",
        base_demand=0, growth_compound=0.09,
        demand_after_growth=0, stimulation=1.0, demand_after_stim=0,
        capture_rate=0.0,  # varies by city via QSI
        forecast_pax=48.115, ptew=66.1,
    ))
    # Connecting at SJC (dest hub)
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SJC", hub_name="San Jose",
        base_demand=0, growth_compound=0.10,
        demand_after_growth=0, stimulation=1.0, demand_after_stim=0,
        capture_rate=0.0,
        forecast_pax=2.937, ptew=4.0,
    ))
    return case


def case_ke_icn_sjc_7x() -> RouteValidationCase:
    """Korean Air ICN-SJC, 7x weekly, B77W, from Feb 2024 presentation."""
    case = RouteValidationCase(
        airline_name="Korean Air", airline_code="KE",
        origin="ICN", destination="SJC",
        frequency=7, aircraft="B77W", seats=338,
        base_year=2019, forecast_year=2025,
        p2p_base_demand=184.4,
        p2p_growth=0.127,  # 12.7% compound 2019-2025
        p2p_demand_grown=207.8,
        p2p_stimulation=1.27,
        p2p_demand_stim=262.8,
        p2p_capture_rate=0.195,
        p2p_forecast=51.2,
        p2p_ptew=70.4,
        grand_total_pax=184.4,  # 184,353
        grand_total_ptew=253.2,
        annual_seats=246064,  # 338 * 7 * 52 * 2
        load_factor=0.749,
    )
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="ICN", hub_name="Seoul Incheon",
        base_demand=3169.4, growth_compound=0.127,
        demand_after_growth=3571.9, stimulation=1.0,
        demand_after_stim=3571.9,
        capture_rate=0.034,
        forecast_pax=121.5, ptew=166.9,
        direct_comp_demand=2799.7, direct_comp_capture=0.021, direct_comp_forecast=59.5,
        no_direct_demand=772.2, no_direct_capture=0.08, no_direct_forecast=62.1,
    ))
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SJC", hub_name="San Jose",
        base_demand=1888.6, growth_compound=0.127,
        demand_after_growth=2128.5, stimulation=1.0,
        demand_after_stim=2128.5,
        capture_rate=0.005,
        forecast_pax=11.6, ptew=15.9,
        direct_comp_demand=1910.3, direct_comp_capture=0.003, direct_comp_forecast=6.0,
        no_direct_demand=218.1, no_direct_capture=0.026, no_direct_forecast=5.6,
    ))
    # Top connecting cities at ICN
    case.hub_cities['ICN'] = [
        CityForecast('TYO', 'Tokyo', 'Japan', 415461, 0.165, 25316, 34.7),
        CityForecast('SGN', 'Ho Chi Minh City', 'Viet Nam', 211197, 0.087, 18415, 25.2),
        CityForecast('MNL', 'Manila', 'Philippines', 454876, 0.050, 10369, 14.2),
        CityForecast('BKK', 'Bangkok', 'Thailand', 172463, 0.060, 10317, 14.1),
        CityForecast('SIN', 'Singapore', 'Singapore', 245532, 0.083, 7017, 9.6),
        CityForecast('SHA', 'Shanghai', 'China', 354202, 0.042, 5796, 7.9),
    ]
    case.hub_cities['SJC'] = [
        CityForecast('LAX', 'Los Angeles', 'USA', 749483, 0.036, 3907, 5.4),
        CityForecast('SAN', 'San Diego', 'USA', 27415, 0.093, 2558, 3.5),
        CityForecast('LAS', 'Las Vegas', 'USA', 105584, 0.030, 1293, 1.8),
        CityForecast('PDX', 'Portland', 'USA', 28893, 0.027, 775, 1.1),
    ]
    return case


def case_ke_icn_sjc_5x() -> RouteValidationCase:
    """Korean Air ICN-SJC, 5x weekly, A350, from Feb 2024 appendix."""
    case = RouteValidationCase(
        airline_name="Korean Air", airline_code="KE",
        origin="ICN", destination="SJC",
        frequency=5, aircraft="A350-900", seats=311,
        base_year=2019, forecast_year=2025,
        p2p_base_demand=184.4,
        p2p_growth=0.127,
        p2p_demand_grown=207.8,
        p2p_stimulation=1.27,
        p2p_demand_stim=262.8,
        p2p_capture_rate=0.195,
        p2p_forecast=51.2,
        p2p_ptew=98.6,  # higher PTEW due to fewer flights
        grand_total_pax=140.9,  # 140,858
        grand_total_ptew=270.9,
        annual_seats=161720,  # 311 * 5 * 52 * 2
        load_factor=0.871,
    )
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="ICN", hub_name="Seoul Incheon",
        base_demand=3169.4, growth_compound=0.127,
        demand_after_growth=3571.9, stimulation=1.0,
        demand_after_stim=3571.9,
        capture_rate=0.023,
        forecast_pax=82.0, ptew=157.7,
        direct_comp_demand=2799.7, direct_comp_capture=0.018, direct_comp_forecast=50.8,
        no_direct_demand=772.2, no_direct_capture=0.04, no_direct_forecast=31.3,
    ))
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SJC", hub_name="San Jose",
        base_demand=1888.6, growth_compound=0.127,
        demand_after_growth=2128.5, stimulation=1.0,
        demand_after_stim=2128.5,
        capture_rate=0.004,
        forecast_pax=7.6, ptew=14.6,
        direct_comp_demand=1910.3, direct_comp_capture=0.002, direct_comp_forecast=3.6,
        no_direct_demand=218.1, no_direct_capture=0.018, no_direct_forecast=4.0,
    ))
    return case


def case_sq_sin_sjc() -> RouteValidationCase:
    """Singapore Airlines SIN-SJC, 4x weekly, A350-900ULR, Feb 2024."""
    case = RouteValidationCase(
        airline_name="Singapore Airlines", airline_code="SQ",
        origin="SIN", destination="SJC",
        frequency=4, aircraft="A350-900ULR", seats=161,
        base_year=2019, forecast_year=2025,
        p2p_base_demand=59.3,
        p2p_growth=0.127,
        p2p_demand_grown=66.9,
        p2p_stimulation=1.30,
        p2p_demand_stim=86.9,
        p2p_capture_rate=0.350,
        p2p_forecast=30.4,
        p2p_ptew=73.1,
        grand_total_pax=54.8,  # 54,778
        grand_total_ptew=131.7,
        annual_seats=66976,  # 161 * 4 * 52 * 2 = 66,976
        load_factor=0.818,
    )
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SIN", hub_name="Singapore",
        base_demand=1210.4, growth_compound=0.193,
        demand_after_growth=1444.1, stimulation=1.0,
        demand_after_stim=1444.1,
        capture_rate=0.016,
        forecast_pax=23.1, ptew=55.6,
        direct_comp_demand=0, direct_comp_capture=0, direct_comp_forecast=0,
        no_direct_demand=1444.1, no_direct_capture=0.016, no_direct_forecast=23.1,
    ))
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SJC", hub_name="San Jose",
        base_demand=630.2, growth_compound=0.147,
        demand_after_growth=722.7, stimulation=1.0,
        demand_after_stim=722.7,
        capture_rate=0.002,
        forecast_pax=1.2, ptew=2.9,
        direct_comp_demand=408.5, direct_comp_capture=0.001, direct_comp_forecast=0.5,
        no_direct_demand=314.2, no_direct_capture=0.002, no_direct_forecast=0.7,
    ))
    # Top connecting cities at SIN
    case.hub_cities['SIN'] = []  # Not broken out in presentation with individual cities
    case.hub_cities['SJC'] = [
        CityForecast('PDX', 'Portland', 'USA', 14118, 0.016, 229, 0.5),
        CityForecast('SEA', 'Seattle', 'USA', 47827, 0.006, 222, 0.5),
        CityForecast('LAX', 'Los Angeles', 'USA', 166433, 0.002, 200, 0.5),
        CityForecast('LAS', 'Las Vegas', 'USA', 20520, 0.006, 129, 0.3),
    ]
    return case


def case_cx_hkg_sjc() -> RouteValidationCase:
    """Cathay Pacific HKG-SJC, 4x weekly, A350-900, Sep 2025."""
    case = RouteValidationCase(
        airline_name="Cathay Pacific", airline_code="CX",
        origin="HKG", destination="SJC",
        frequency=4, aircraft="A350-900", seats=280,
        base_year=2025, forecast_year=2028,
        p2p_base_demand=66.5,
        p2p_growth=0.321,  # compound: (1+g)^3 - 1 = 0.321 to get from 66.5 to 87.8
        p2p_demand_grown=87.8,
        p2p_stimulation=1.40,
        p2p_demand_stim=123.0,
        p2p_capture_rate=0.350,
        p2p_forecast=43.0,
        p2p_ptew=103,
        grand_total_pax=95.1,  # 95,084
        grand_total_ptew=229,
        annual_seats=116480,  # 280 * 4 * 52 * 2
        load_factor=0.816,
    )
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="HKG", hub_name="Hong Kong",
        base_demand=1415.6, growth_compound=0.321,  # compound over 3 years
        demand_after_growth=1870.0, stimulation=1.0,
        demand_after_stim=1870.0,
        capture_rate=0.025,
        forecast_pax=47.5, ptew=114,
        direct_comp_demand=0, direct_comp_capture=0, direct_comp_forecast=0,
        no_direct_demand=1870.0, no_direct_capture=0.025, no_direct_forecast=47.5,
    ))
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SJC", hub_name="San Jose",
        base_demand=682.7, growth_compound=0.321,  # compound over 3 years
        demand_after_growth=901.9, stimulation=1.0,
        demand_after_stim=901.9,
        capture_rate=0.005,
        forecast_pax=4.5, ptew=11,
        direct_comp_demand=518.0, direct_comp_capture=0.001, direct_comp_forecast=0.7,
        no_direct_demand=383.9, no_direct_capture=0.010, no_direct_forecast=3.8,
    ))
    # Top connecting cities at HKG
    case.hub_cities['HKG'] = [
        CityForecast('SZX', 'Shenzhen', 'China', 68179, 0.135, 9177, 22),
        CityForecast('DEL', 'Delhi', 'India', 244675, 0.033, 8093, 19),
        CityForecast('MNL', 'Manila', 'Philippines', 232580, 0.022, 5188, 12),
        CityForecast('BLR', 'Bengaluru', 'India', 163557, 0.029, 4741, 11),
        CityForecast('CAN', 'Guangzhou', 'China', 52880, 0.072, 3800, 9),
    ]
    case.hub_cities['SJC'] = [
        CityForecast('SEA', 'Seattle', 'USA', 74630, 0.012, 899, 2),
        CityForecast('LAX', 'Los Angeles', 'USA', 383379, 0.007, 700, 2),
        CityForecast('GDL', 'Guadalajara', 'Mexico', 1914, 0.328, 627, 2),
    ]
    return case


def case_fi_kef_sjc() -> RouteValidationCase:
    """Icelandair KEF-SJC, 4x weekly, B757-200, Oct 2023."""
    case = RouteValidationCase(
        airline_name="Icelandair", airline_code="FI",
        origin="KEF", destination="SJC",
        frequency=4, aircraft="B757-200", seats=183,
        base_year=2019, forecast_year=2025,
        p2p_base_demand=38.1,
        p2p_growth=0.024,  # 2.4% compound
        p2p_demand_grown=39.0,
        p2p_stimulation=1.00,  # no stim  direct service existed from SFO
        p2p_demand_stim=39.0,
        p2p_capture_rate=0.500,
        p2p_forecast=19.5,
        p2p_ptew=46.8,
        grand_total_pax=62.8,  # 62,794
        grand_total_ptew=150.9,
        annual_seats=76128,  # 183 * 4 * 52 * 2
        load_factor=0.825,
    )
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="KEF", hub_name="Reykjavik Keflavik",
        base_demand=1785.7, growth_compound=0.025,
        demand_after_growth=1831.0, stimulation=1.0,
        demand_after_stim=1831.0,
        capture_rate=0.023,
        forecast_pax=42.2, ptew=101.4,
        direct_comp_demand=499.1, direct_comp_capture=0.011, direct_comp_forecast=5.4,
        no_direct_demand=1331.9, no_direct_capture=0.028, no_direct_forecast=36.8,
    ))
    case.connecting_markets.append(ConnectingMarketTarget(
        hub_code="SJC", hub_name="San Jose",
        base_demand=64.0, growth_compound=0.134,
        demand_after_growth=72.6, stimulation=1.0,
        demand_after_stim=72.6,
        capture_rate=0.015,
        forecast_pax=1.1, ptew=2.7,
        direct_comp_demand=18.3, direct_comp_capture=0.001, direct_comp_forecast=0.0,
        no_direct_demand=54.3, no_direct_capture=0.020, no_direct_forecast=1.1,
    ))
    # Top connecting cities at KEF
    case.hub_cities['KEF'] = [
        CityForecast('STO', 'Stockholm', 'Sweden', 43401, 0.150, 6490, 15.6),
        CityForecast('LON', 'London', 'United Kingdom', 499133, 0.012, 5428, 13.0),
        CityForecast('DUB', 'Dublin', 'Ireland', 87924, 0.054, 4762, 11.4),
        CityForecast('OSL', 'Oslo', 'Norway', 25003, 0.169, 4219, 10.1),
        CityForecast('CPH', 'Copenhagen', 'Denmark', 45614, 0.080, 3661, 8.8),
        CityForecast('PAR', 'Paris', 'France', 224684, 0.015, 3471, 8.3),
    ]
    case.hub_cities['SJC'] = [
        CityForecast('LAX', 'Los Angeles', 'USA', 29182, 0.032, 935, 2.2),
        CityForecast('SAN', 'San Diego', 'USA', 5648, 0.005, 28, 0.1),
    ]
    return case


# ============================================================================
# VALIDATION TESTS
# ============================================================================

class ValidationResult:
    def __init__(self, test_name: str, route: str):
        self.test_name = test_name
        self.route = route
        self.checks = []
        self.passed = 0
        self.failed = 0
        self.warnings = 0

    def check(self, label: str, actual, expected, tolerance=0.02, unit=''):
        """Check a value against expected with tolerance."""
        if expected == 0:
            ok = abs(actual) < 0.1
            pct_diff = 0
        else:
            pct_diff = abs(actual - expected) / abs(expected)
            ok = pct_diff <= tolerance

        status = 'PASS' if ok else 'FAIL'
        if ok:
            self.passed += 1
        else:
            self.failed += 1

        self.checks.append({
            'label': label, 'actual': actual, 'expected': expected,
            'pct_diff': pct_diff, 'tolerance': tolerance, 'status': status, 'unit': unit
        })

    def warn(self, label: str, message: str):
        self.warnings += 1
        self.checks.append({
            'label': label, 'status': 'WARN', 'message': message
        })

    def info(self, label: str, message: str):
        self.checks.append({
            'label': label, 'status': 'INFO', 'message': message
        })


def test_forecast_table_math(case: RouteValidationCase) -> ValidationResult:
    """Test 1: Verify forecast table arithmetic."""
    r = ValidationResult('Forecast Table Mathematics',
                         f"{case.airline_code} {case.origin}-{case.destination} ({case.frequency}x)")

    # P2P calculations
    computed_p2p_grown = case.p2p_base_demand * (1 + case.p2p_growth)
    r.check('P2P demand after growth', computed_p2p_grown, case.p2p_demand_grown,
            tolerance=0.01, unit="'000s pax")

    computed_p2p_stim = case.p2p_demand_grown * case.p2p_stimulation
    r.check('P2P demand after stimulation', computed_p2p_stim, case.p2p_demand_stim,
            tolerance=0.01, unit="'000s pax")

    computed_p2p_fcst = case.p2p_demand_stim * case.p2p_capture_rate
    r.check('P2P forecast', computed_p2p_fcst, case.p2p_forecast,
            tolerance=0.02, unit="'000s pax")

    # Connecting market calculations
    for cm in case.connecting_markets:
        prefix = f"Cnx@{cm.hub_code}"

        if cm.base_demand > 0:
            computed_grown = cm.base_demand * (1 + cm.growth_compound)
            r.check(f'{prefix} demand after growth', computed_grown, cm.demand_after_growth,
                    tolerance=0.01, unit="'000s")

        # Direct competition sub-segment
        if cm.direct_comp_demand > 0:
            computed_dc = cm.direct_comp_demand * cm.direct_comp_capture
            r.check(f'{prefix} direct comp forecast', computed_dc, cm.direct_comp_forecast,
                    tolerance=0.05, unit="'000s")

        # No direct competition sub-segment
        if cm.no_direct_demand > 0:
            computed_ndc = cm.no_direct_demand * cm.no_direct_capture
            r.check(f'{prefix} no-direct forecast', computed_ndc, cm.no_direct_forecast,
                    tolerance=0.05, unit="'000s")

        # Total connecting
        if cm.direct_comp_forecast > 0 or cm.no_direct_forecast > 0:
            computed_total = cm.direct_comp_forecast + cm.no_direct_forecast
            r.check(f'{prefix} total forecast', computed_total, cm.forecast_pax,
                    tolerance=0.05, unit="'000s")

    # Grand total
    computed_grand = case.p2p_forecast + sum(m.forecast_pax for m in case.connecting_markets)
    r.check('Grand total', computed_grand, case.grand_total_pax,
            tolerance=0.02, unit="'000s pax")

    return r


def test_ptew_calculations(case: RouteValidationCase) -> ValidationResult:
    """Test 2: Verify PTEW (Passengers per Trip Each Way) calculations."""
    r = ValidationResult('PTEW Calculations',
                         f"{case.airline_code} {case.origin}-{case.destination} ({case.frequency}x)")

    # PTEW = Passengers per Trip Each Way
    # Total pax / (freq  52 weeks  2 directions)
    annual_flights_both = case.frequency * 52 * 2

    # P2P PTEW
    computed_ptew = (case.p2p_forecast * 1000) / annual_flights_both
    r.check('P2P PTEW', computed_ptew, case.p2p_ptew, tolerance=0.02, unit='pax/trip')

    # Connecting PTEWs
    for cm in case.connecting_markets:
        if cm.ptew > 0:
            computed = (cm.forecast_pax * 1000) / annual_flights_both
            r.check(f'Cnx@{cm.hub_code} PTEW', computed, cm.ptew,
                    tolerance=0.03, unit='pax/trip')

    # Grand total PTEW
    if case.grand_total_ptew > 0:
        computed_total = (case.grand_total_pax * 1000) / annual_flights_both
        r.check('Grand total PTEW', computed_total, case.grand_total_ptew,
                tolerance=0.02, unit='pax/trip')

    # Individual city PTEW checks
    for hub, cities in case.hub_cities.items():
        for city in cities[:3]:  # Top 3 per hub
            if city.ptew > 0:
                computed = city.forecast_pax / annual_flights_both
                r.check(f'{hub}{city.city_code} PTEW', computed, city.ptew,
                        tolerance=0.10, unit='pax/trip')

    return r


def test_capacity_load_factor(case: RouteValidationCase) -> ValidationResult:
    """Test 3: Verify capacity and load factor calculations."""
    r = ValidationResult('Capacity & Load Factor',
                         f"{case.airline_code} {case.origin}-{case.destination} ({case.frequency}x)")

    # Annual seats = seats  freq  52  2 (both directions)
    computed_seats = case.seats * case.frequency * 52 * 2
    r.check('Annual seats', computed_seats, case.annual_seats, tolerance=0.001, unit='seats')

    # Load factor = total pax / annual seats
    computed_lf = (case.grand_total_pax * 1000) / computed_seats
    r.check('Load factor', computed_lf, case.load_factor, tolerance=0.01, unit='%')

    # Reasonableness
    if case.load_factor < 0.60:
        r.warn('Low LF', f"Load factor {case.load_factor:.1%} below 60%  route may not be viable")
    if case.load_factor > 0.92:
        r.warn('High LF', f"Load factor {case.load_factor:.1%} above 92%  may be capacity constrained")

    return r


def test_connecting_city_aggregation(case: RouteValidationCase) -> ValidationResult:
    """Test 4: Verify connecting city forecasts sum to market totals."""
    r = ValidationResult('Connecting City Aggregation',
                         f"{case.airline_code} {case.origin}-{case.destination} ({case.frequency}x)")

    for cm in case.connecting_markets:
        if cm.hub_code in case.hub_cities and case.hub_cities[cm.hub_code]:
            cities = case.hub_cities[cm.hub_code]
            city_sum = sum(c.forecast_pax for c in cities) / 1000  # convert to '000s
            total = cm.forecast_pax

            if total > 0:
                coverage = city_sum / total
                r.info(f'{cm.hub_code} city coverage',
                       f"{len(cities)} cities cover {coverage:.1%} of total {total:.1f}k")

                # Verify individual city math: forecast = demand  capture_rate
                # Note: city base_demand in presentations is ALREADY the grown/forecast-year figure
                for city in cities[:5]:
                    if city.base_demand > 0 and city.capture_rate > 0:
                        computed = city.base_demand * city.capture_rate
                        # Allow wider tolerance  presentations round capture rates
                        r.check(f'{cm.hub_code}{city.city_code} forecast',
                                computed, city.forecast_pax, tolerance=0.10, unit='pax')

    return r


def test_parameter_reasonableness(case: RouteValidationCase) -> ValidationResult:
    """Test 5: Check calibration parameters fall within expected ranges."""
    r = ValidationResult('Parameter Reasonableness',
                         f"{case.airline_code} {case.origin}-{case.destination} ({case.frequency}x)")

    # P2P capture rate bounds
    if case.p2p_capture_rate > 0.60:
        r.warn('High P2P capture', f"{case.p2p_capture_rate:.0%} exceeds 60%")
    elif case.p2p_capture_rate < 0.05:
        r.warn('Low P2P capture', f"{case.p2p_capture_rate:.0%} below 5%")
    else:
        r.check('P2P capture in range', case.p2p_capture_rate, 0.25,
                tolerance=1.0)  # wide  just checking it's nonzero

    # Stimulation bounds
    if case.p2p_stimulation > 1.50:
        r.warn('High stimulation', f"{case.p2p_stimulation:.2f}x exceeds 1.50x")
    if case.p2p_stimulation < 1.00:
        r.warn('Negative stimulation', f"{case.p2p_stimulation:.2f}x below 1.00x")

    # Connecting capture rates should be much lower than P2P
    for cm in case.connecting_markets:
        if cm.capture_rate > 0 and cm.capture_rate > case.p2p_capture_rate:
            r.warn(f'Cnx@{cm.hub_code} capture > P2P',
                   f"Connecting {cm.capture_rate:.1%} > P2P {case.p2p_capture_rate:.1%}")

    # Hub connecting capture: home hub should generally > dest hub
    home_cm = case.connecting_markets[0] if case.connecting_markets else None
    dest_cm = case.connecting_markets[1] if len(case.connecting_markets) > 1 else None
    if home_cm and dest_cm and home_cm.capture_rate > 0 and dest_cm.capture_rate > 0:
        if home_cm.capture_rate < dest_cm.capture_rate:
            r.warn('Unusual hub balance',
                   f"Home hub {home_cm.hub_code} capture {home_cm.capture_rate:.1%} "
                   f"< dest {dest_cm.hub_code} {dest_cm.capture_rate:.1%}")
        else:
            r.info('Hub balance OK',
                   f"Home {home_cm.hub_code}: {home_cm.capture_rate:.1%}, "
                   f"Dest {dest_cm.hub_code}: {dest_cm.capture_rate:.1%}")

    # No-direct should have higher capture than direct-competition
    for cm in case.connecting_markets:
        if cm.direct_comp_capture > 0 and cm.no_direct_capture > 0:
            if cm.no_direct_capture <= cm.direct_comp_capture:
                r.warn(f'{cm.hub_code} no-direct  direct capture',
                       f"No-direct {cm.no_direct_capture:.1%} should exceed "
                       f"direct {cm.direct_comp_capture:.1%}")
            else:
                ratio = cm.no_direct_capture / cm.direct_comp_capture
                r.info(f'{cm.hub_code} no-direct/direct ratio', f"{ratio:.1f}x")

    return r


def test_cross_route_patterns(cases: List[RouteValidationCase]) -> ValidationResult:
    """Test 6: Cross-route pattern analysis."""
    r = ValidationResult('Cross-Route Pattern Analysis', 'All Routes')

    # Compare P2P capture rates
    r.info('P2P Capture Rate Comparison', '')
    for c in cases:
        r.info(f'  {c.airline_code} {c.origin}-{c.destination} ({c.frequency}x)',
               f"P2P: {c.p2p_capture_rate:.1%}, Stim: {c.p2p_stimulation:.2f}x, "
               f"LF: {c.load_factor:.1%}")

    # Frequency sensitivity: KE 7x vs 5x
    ke7 = next((c for c in cases if c.airline_code == 'KE' and c.frequency == 7), None)
    ke5 = next((c for c in cases if c.airline_code == 'KE' and c.frequency == 5), None)
    if ke7 and ke5:
        freq_ratio = ke5.frequency / ke7.frequency
        pax_ratio = ke5.grand_total_pax / ke7.grand_total_pax
        r.info('KE frequency sensitivity',
               f"5x/7x freq ratio = {freq_ratio:.2f}, pax ratio = {pax_ratio:.2f}")
        # Pax should drop less than linearly with frequency
        if pax_ratio > freq_ratio:
            r.info('  Sublinear pax decline', 'Expected  fewer flights capture proportionally more')
        else:
            r.warn('  Superlinear pax decline',
                   'Unusual  pax dropped faster than frequency')

        # Home hub capture should decrease with lower frequency
        ke7_home = ke7.connecting_markets[0].capture_rate
        ke5_home = ke5.connecting_markets[0].capture_rate
        if ke5_home < ke7_home:
            r.info('  ICN capture decreases with freq',
                   f"7x: {ke7_home:.1%}, 5x: {ke5_home:.1%}")
        else:
            r.warn('  ICN capture unchanged/increased',
                   f"7x: {ke7_home:.1%}, 5x: {ke5_home:.1%}")

    # Hub size effect: larger hubs should show lower capture rates
    # (more competition from other routings)
    hubs_by_demand = []
    for c in cases:
        for cm in c.connecting_markets:
            if cm.base_demand > 100 and cm.hub_code != 'SJC':
                hubs_by_demand.append((c.airline_code, cm.hub_code,
                                       cm.demand_after_stim, cm.capture_rate))
    hubs_by_demand.sort(key=lambda x: x[2], reverse=True)
    r.info('Hub connecting pools (descending demand)', '')
    for al, hub, demand, cap in hubs_by_demand:
        r.info(f'  {al} {hub}', f"Pool: {demand:,.0f}k, Capture: {cap:.1%}")

    # Stimulation patterns
    r.info('Stimulation analysis', '')
    for c in cases:
        stim_note = ""
        if c.p2p_stimulation == 1.0:
            stim_note = "No stimulation (service already exists nearby)"
        elif c.p2p_stimulation > 1.2:
            stim_note = "High stimulation (new market/strong demand creation)"
        else:
            stim_note = "Moderate stimulation"
        r.info(f'  {c.airline_code} {c.origin}-{c.destination}',
               f"{c.p2p_stimulation:.2f}x  {stim_note}")

    return r


# ============================================================================
# MAIN VALIDATION RUNNER
# ============================================================================

def run_all_validations():
    """Run complete cross-route validation suite."""
    cases = [
        case_ba_lhr_sjc(),
        case_ke_icn_sjc_7x(),
        case_ke_icn_sjc_5x(),
        case_sq_sin_sjc(),
        case_cx_hkg_sjc(),
        case_fi_kef_sjc(),
    ]

    all_results = []

    print("=" * 70)
    print("AVIA SOLUTIONS  CROSS-ROUTE VALIDATION FRAMEWORK")
    print(f"Date: {datetime.now():%Y-%m-%d %H:%M}")
    print(f"Cases: {len(cases)} route configurations")
    print("=" * 70)

    # Per-route tests
    for case in cases:
        print(f"\n{'' * 70}")
        print(f"  {case.airline_name}: {case.origin}-{case.destination} "
              f"({case.frequency}x/wk, {case.aircraft}, {case.seats}s)")
        print(f"  Target: {case.grand_total_pax * 1000:,.0f} pax, "
              f"{case.load_factor:.1%} LF, {case.annual_seats:,} seats")
        print(f"{'' * 70}")

        for test_fn in [test_forecast_table_math, test_ptew_calculations,
                        test_capacity_load_factor, test_connecting_city_aggregation,
                        test_parameter_reasonableness]:
            result = test_fn(case)
            all_results.append(result)
            _print_result(result)

    # Cross-route tests
    print(f"\n{'' * 70}")
    print("  CROSS-ROUTE PATTERN ANALYSIS")
    print(f"{'' * 70}")
    xr = test_cross_route_patterns(cases)
    all_results.append(xr)
    _print_result(xr)

    # Summary
    total_pass = sum(r.passed for r in all_results)
    total_fail = sum(r.failed for r in all_results)
    total_warn = sum(r.warnings for r in all_results)

    print(f"\n{'=' * 70}")
    print(f"  VALIDATION SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Routes tested:  {len(cases)}")
    print(f"  Tests run:      {len(all_results)}")
    print(f"  Checks passed:  {total_pass}")
    print(f"  Checks failed:  {total_fail}")
    print(f"  Warnings:       {total_warn}")
    print(f"  Result:         {' ALL PASSED' if total_fail == 0 else ' FAILURES DETECTED'}")
    print(f"{'=' * 70}")

    return all_results, cases


def _print_result(result: ValidationResult):
    """Print a single validation result."""
    status = '' if result.failed == 0 else ''
    print(f"\n  {status} {result.test_name} ({result.passed}P/{result.failed}F/{result.warnings}W)")

    for check in result.checks:
        if check['status'] == 'INFO':
            print(f"     {check['label']}: {check.get('message', '')}")
        elif check['status'] == 'WARN':
            print(f"     {check['label']}: {check.get('message', '')}")
        elif check['status'] == 'PASS':
            exp = check['expected']
            act = check['actual']
            if isinstance(exp, float) and exp < 1 and check.get('unit') == '%':
                print(f"     {check['label']}: {act:.1%} vs {exp:.1%} "
                      f"({check['pct_diff']:.1%} diff)")
            else:
                print(f"     {check['label']}: {act:,.1f} vs {exp:,.1f} "
                      f"({check['pct_diff']:.1%} diff)")
        else:  # FAIL
            exp = check['expected']
            act = check['actual']
            print(f"     {check['label']}: {act:,.1f} vs {exp:,.1f} "
                  f"({check['pct_diff']:.1%} diff, tol {check['tolerance']:.0%})")


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

def write_validation_workbook(results: list, cases: list, outpath: str):
    """Generate validation results workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hfill = PatternFill('solid', fgColor='002060')
    df = Font(name='Arial', size=10)
    bf = Font(name='Arial', size=10, bold=True)
    tf = Font(name='Arial', size=14, bold=True, color='002060')
    pass_fill = PatternFill('solid', fgColor='C6EFCE')
    fail_fill = PatternFill('solid', fgColor='FFC7CE')
    warn_fill = PatternFill('solid', fgColor='FFEB9C')
    info_fill = PatternFill('solid', fgColor='DDEBF7')
    thin = Side(style='thin', color='999999')
    border = Border(bottom=thin)

    def hdr(ws, cols, row=1):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row, c, v)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    #  Sheet 1: Summary Dashboard 
    ws = wb.active
    ws.title = 'Validation Summary'
    ws.cell(1, 1, 'Avia Solutions  Cross-Route Validation').font = tf
    ws.cell(2, 1, f'Generated: {datetime.now():%Y-%m-%d %H:%M}').font = df

    total_pass = sum(r.passed for r in results)
    total_fail = sum(r.failed for r in results)
    total_warn = sum(r.warnings for r in results)

    ws.cell(4, 1, 'Overall Result').font = bf
    cell = ws.cell(4, 2, 'ALL PASSED' if total_fail == 0 else 'FAILURES DETECTED')
    cell.font = Font(name='Arial', size=10, bold=True,
                     color='006100' if total_fail == 0 else '9C0006')
    cell.fill = pass_fill if total_fail == 0 else fail_fill

    row = 6
    for label, val in [('Routes Tested', len(cases)), ('Tests Run', len(results)),
                        ('Checks Passed', total_pass), ('Checks Failed', total_fail),
                        ('Warnings', total_warn)]:
        ws.cell(row, 1, label).font = df
        ws.cell(row, 2, val).font = bf
        row += 1

    # Route summary table
    row += 1
    ws.cell(row, 1, 'Route Summary').font = tf
    row += 1
    hdr(ws, ['Route', 'Airline', 'Freq', 'Aircraft', 'Seats', 'Total Pax',
             'Load Factor', 'P2P Capture', 'Stimulation', 'Cnx Share'], row)
    row += 1
    for case in cases:
        cnx_total = sum(m.forecast_pax for m in case.connecting_markets)
        cnx_share = cnx_total / case.grand_total_pax if case.grand_total_pax > 0 else 0
        ws.cell(row, 1, f'{case.origin}-{case.destination}').font = bf
        ws.cell(row, 2, case.airline_code).font = df
        ws.cell(row, 3, f'{case.frequency}x').font = df
        ws.cell(row, 4, case.aircraft).font = df
        ws.cell(row, 5, case.seats).font = df
        c = ws.cell(row, 6, case.grand_total_pax * 1000)
        c.font = df; c.number_format = '#,##0'
        c = ws.cell(row, 7, case.load_factor)
        c.font = df; c.number_format = '0.0%'
        c = ws.cell(row, 8, case.p2p_capture_rate)
        c.font = df; c.number_format = '0.0%'
        c = ws.cell(row, 9, case.p2p_stimulation)
        c.font = df; c.number_format = '0.00'
        c = ws.cell(row, 10, cnx_share)
        c.font = df; c.number_format = '0.0%'
        row += 1

    for c in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14
    ws.column_dimensions['A'].width = 18

    #  Sheet 2: Detailed Check Results 
    ws2 = wb.create_sheet('Detailed Checks')
    hdr(ws2, ['Route', 'Test', 'Check', 'Status', 'Actual', 'Expected',
              'Difference', 'Tolerance', 'Detail'])
    row = 2
    for result in results:
        for check in result.checks:
            ws2.cell(row, 1, result.route).font = df
            ws2.cell(row, 2, result.test_name).font = df
            ws2.cell(row, 3, check['label']).font = df

            status_cell = ws2.cell(row, 4, check['status'])
            status_cell.font = bf
            if check['status'] == 'PASS':
                status_cell.fill = pass_fill
            elif check['status'] == 'FAIL':
                status_cell.fill = fail_fill
            elif check['status'] == 'WARN':
                status_cell.fill = warn_fill
            else:
                status_cell.fill = info_fill

            if 'actual' in check:
                c = ws2.cell(row, 5, check['actual'])
                c.font = df
                if isinstance(check['actual'], float) and check['actual'] < 2:
                    c.number_format = '0.000'
                else:
                    c.number_format = '#,##0.0'
                c = ws2.cell(row, 6, check['expected'])
                c.font = df
                if isinstance(check['expected'], float) and check['expected'] < 2:
                    c.number_format = '0.000'
                else:
                    c.number_format = '#,##0.0'
                c = ws2.cell(row, 7, check['pct_diff'])
                c.font = df; c.number_format = '0.0%'
                c = ws2.cell(row, 8, check['tolerance'])
                c.font = df; c.number_format = '0.0%'

            if 'message' in check:
                ws2.cell(row, 9, check['message']).font = df

            row += 1

    for c in range(1, 10):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['I'].width = 50

    #  Sheet 3: Cross-Route Comparison 
    ws3 = wb.create_sheet('Route Comparison')
    ws3.cell(1, 1, 'Cross-Route Parameter Comparison').font = tf

    # Parameter comparison table
    params = [
        'P2P Base Demand (000s)', 'P2P Growth', 'P2P Stimulation',
        'P2P Capture Rate', 'P2P Forecast (000s)',
        'Home Hub Pool (000s)', 'Home Hub Capture',
        'Home Hub Forecast (000s)', 'Dest Hub Pool (000s)',
        'Dest Hub Capture', 'Dest Hub Forecast (000s)',
        'Grand Total (000s)', 'Annual Seats', 'Load Factor',
        'Connecting Share', 'PTEW Grand Total',
    ]

    row = 3
    ws3.cell(row, 1, 'Parameter').font = bf
    for i, case in enumerate(cases):
        col = i + 2
        ws3.cell(row, col, f'{case.airline_code} {case.origin}-{case.destination}\n{case.frequency}x').font = bf
        ws3.cell(row, col).alignment = Alignment(horizontal='center', wrap_text=True)
    row += 1

    for p_idx, param in enumerate(params):
        ws3.cell(row, 1, param).font = df
        for i, case in enumerate(cases):
            col = i + 2
            val = _get_param(case, p_idx)
            c = ws3.cell(row, col, val)
            c.font = df
            if param.endswith('(000s)'):
                c.number_format = '#,##0.0'
            elif 'Capture' in param or 'Growth' in param or 'Load Factor' in param or 'Share' in param:
                c.number_format = '0.0%'
            elif 'Stimulation' in param:
                c.number_format = '0.00'
            elif 'Seats' in param:
                c.number_format = '#,##0'
            elif 'PTEW' in param:
                c.number_format = '0.0'
        row += 1

    ws3.column_dimensions['A'].width = 28
    for i in range(len(cases)):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i + 2)].width = 16

    wb.save(outpath)
    return outpath


def _get_param(case: RouteValidationCase, idx: int):
    """Extract parameter value by index for comparison table."""
    home = case.connecting_markets[0] if case.connecting_markets else None
    dest = case.connecting_markets[1] if len(case.connecting_markets) > 1 else None
    cnx_total = sum(m.forecast_pax for m in case.connecting_markets)
    cnx_share = cnx_total / case.grand_total_pax if case.grand_total_pax > 0 else 0

    values = [
        case.p2p_base_demand,
        case.p2p_growth,
        case.p2p_stimulation,
        case.p2p_capture_rate,
        case.p2p_forecast,
        home.demand_after_stim if home else 0,
        home.capture_rate if home else 0,
        home.forecast_pax if home else 0,
        dest.demand_after_stim if dest else 0,
        dest.capture_rate if dest else 0,
        dest.forecast_pax if dest else 0,
        case.grand_total_pax,
        case.annual_seats,
        case.load_factor,
        cnx_share,
        case.grand_total_ptew,
    ]
    return values[idx] if idx < len(values) else 0


# ============================================================================
# CLI
# ============================================================================

if __name__ == '__main__':
    results, cases = run_all_validations()

    ensure_output_dir()
    outpath = str(OUTPUT_DIR / 'Cross_Route_Validation.xlsx')
    write_validation_workbook(results, cases, outpath)
    print(f"\nOutput: {outpath}")
