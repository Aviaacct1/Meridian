#!/usr/bin/env python3
"""
Avia Solutions  Departure Time Grid Search (Chat 27)
=====================================================
Makes the Time Grid tab functional by enabling departure time shifting
within the SingleExtractOAGProvider and orchestrating multi-scenario
pipeline runs.

Architecture:
    TimeShiftProvider  wraps SingleExtractOAGProvider, modifying the
        proposed carrier's direct flights to shift departure/arrival
        times while keeping flying time constant. All other flights
        (competitor services, connecting flights) remain unchanged.
        Connections are rebuilt at each hub with the new timing.

    TimeGridRunner  runs the full pipeline at each candidate departure
        time, collects results, compares them, and identifies the
        optimal schedule.

Validated against BA LHR-SJC with three known time scenarios:
    Original (21:30 SJC dep) = 129,162 pax
    New time (22:00 SJC dep) = ~136,000 pax
    5pm SJC  (17:00 SJC dep) = ~139,302 pax

The key mechanism: shifting the proposed service's departure time
changes its arrival time at the hub, which changes the connection
window  making some connections viable and others invalid. Through
the QSI's steep elapsed-time decay curve, even small changes in
connection quality create significant capture rate differences.

Dependencies:
    - single_extract_oag_provider.py (SingleExtractOAGProvider)
    - closed_loop_pipeline_v2.py (run_pipeline)
    - route_config.py (RouteConfig)
    - connection_builder.py (build_connections)
    - providers.py (Itinerary, ScheduleProvider)
"""

from config import REFERENCE_CASE_DIR
import os
import sys
import copy
import time as time_module
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Any
from collections import defaultdict


from providers import Itinerary, ScheduleProvider
from single_extract_oag_provider import (
    SingleExtractOAGProvider,
    load_legs_from_qsi_template,
    connection_to_itinerary,
    _safe_str, _safe_int, _parse_elapsed,
)
from connection_builder import (
    build_connections, parse_time_hhmm, minutes_to_hhmm,
    load_alliance_data, ONEWORLD, STAR_ALLIANCE, SKYTEAM,
)
from oag_parser import read_mct_xls, build_mct_lookup


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class TimeScenarioResult:
    """Results from a single departure time scenario."""
    dep_time_origin: str           # e.g., "17:00"  departure from origin (local)
    arr_time_hub: str              # arrival at hub (local)
    dep_time_hub: str              # departure from hub (local, return service)
    arr_time_origin: str           # arrival at origin (local, return service)
    grand_total: int = 0           # total annual passengers
    p2p_total: float = 0.0
    cnx_home_total: float = 0.0
    cnx_dest_total: float = 0.0
    load_factor: float = 0.0
    n_cities_home: int = 0
    n_cities_dest: int = 0
    total_capture_home: float = 0.0
    total_capture_dest: float = 0.0
    itineraries_qsi1: int = 0
    itineraries_qsi2: int = 0
    elapsed_seconds: float = 0.0
    error: str = ''
    full_results: Dict = field(default_factory=dict)

    @property
    def label(self) -> str:
        return f"{self.dep_time_origin} dep origin"


@dataclass
class GridSearchResult:
    """Complete grid search output."""
    origin: str = ''
    destination: str = ''
    carrier: str = ''
    flying_time_outbound: int = 0   # minutes
    flying_time_return: int = 0     # minutes
    scenarios: List[TimeScenarioResult] = field(default_factory=list)
    best_scenario: Optional[TimeScenarioResult] = None
    ranked: List[TimeScenarioResult] = field(default_factory=list)
    city_sensitivity: List[Dict] = field(default_factory=list)
    total_elapsed: float = 0.0

    def summary(self) -> str:
        lines = [
            f"Grid Search: {self.carrier} {self.origin}{self.destination}",
            f"Scenarios: {len(self.scenarios)}",
        ]
        if self.best_scenario:
            lines.append(f"Best: {self.best_scenario.dep_time_origin} dep  "
                         f"{self.best_scenario.grand_total:,} pax, "
                         f"{self.best_scenario.load_factor:.1%} LF")
        if len(self.ranked) >= 2:
            worst = self.ranked[-1]
            delta = self.best_scenario.grand_total - worst.grand_total
            lines.append(f"Range: {delta:+,} pax between best and worst")
        return '\n'.join(lines)


# ============================================================================
# TIME-SHIFT PROVIDER
# ============================================================================

class TimeShiftProvider(ScheduleProvider):
    """
    Wraps SingleExtractOAGProvider with modified departure times for the
    proposed carrier's direct service.

    The mechanism:
    1. Load all leg data from the QSI template (same as SingleExtractOAGProvider)
    2. Identify the proposed carrier's direct flights between origin and destination
    3. Modify their dep_time/arr_time to the target time (keeping flying time constant)
    4. Rebuild connections at each hub with the modified leg data

    This mirrors what the analyst does manually: paste different times into the
    OAG schedule, re-run the connection builder, re-run QSI.

    For bidirectional QSI (home + dest perspectives), the runner should create
    TWO providers  one wrapping QSILHR, one wrapping QSISJC. The primary
    provider can reference the dest provider via the qsi2_file attribute so the
    pipeline's run_pipeline() finds it.

    Args:
        base_provider: The original SingleExtractOAGProvider with all config
        target_dep_time_outbound: New departure time from origin (HHMM string, e.g., '1700')
        target_dep_time_return: New departure time from destination/hub (HHMM string)
            If None, calculated from outbound using block time symmetry
        flying_time_outbound: Flying time in minutes (origin  destination)
        flying_time_return: Flying time in minutes (destination  origin)
        dest_provider: Optional companion TimeShiftProvider for dest perspective
    """

    def __init__(self,
                 base_provider: SingleExtractOAGProvider,
                 target_dep_time_outbound: str,
                 target_dep_time_return: Optional[str] = None,
                 flying_time_outbound: int = 625,
                 flying_time_return: int = 660,
                 dest_provider: Optional['TimeShiftProvider'] = None):

        self._base = base_provider
        self._target_dep_out = target_dep_time_outbound
        self._flying_out = flying_time_outbound
        self._flying_ret = flying_time_return
        self._dest_provider = dest_provider

        # Calculate arrival times from flying time
        dep_out_mins = parse_time_hhmm(target_dep_time_outbound)
        arr_hub_mins = (dep_out_mins + flying_time_outbound) % 1440
        self._arr_hub = minutes_to_hhmm(arr_hub_mins)

        # Return service: if not specified, calculate to maintain turnaround
        if target_dep_time_return:
            self._target_dep_ret = target_dep_time_return
            dep_ret_mins = parse_time_hhmm(target_dep_time_return)
        else:
            # Default: departure from hub = arrival at hub + 90 min turnaround
            dep_ret_mins = (arr_hub_mins + 90) % 1440
            self._target_dep_ret = minutes_to_hhmm(dep_ret_mins)

        arr_origin_mins = (dep_ret_mins + flying_time_return) % 1440
        self._arr_origin = minutes_to_hhmm(arr_origin_mins)

        self._cache: Dict[str, List[Itinerary]] = {}
        self._build_log: List[str] = []
        self._stats: Dict[str, Any] = {}

        self._log(f"TimeShiftProvider: {base_provider.origin_airport}{base_provider.dest_airport}")
        self._log(f"  Outbound: dep {target_dep_time_outbound}  arr {self._arr_hub} "
                  f"(flying {flying_time_outbound} min)")
        self._log(f"  Return:   dep {self._target_dep_ret}  arr {self._arr_origin} "
                  f"(flying {flying_time_return} min)")

    @property
    def qsi2_file(self):
        """Expose dest provider reference for pipeline compatibility.
        
        The pipeline's run_pipeline() checks hasattr(provider, 'qsi2_file') to 
        decide whether to create a separate ExcelScheduleProvider for the dest QSI.
        By returning None here but providing a dest_schedule_provider property,
        we signal that dest QSI should use the dest provider instead.
        """
        if self._dest_provider:
            return '__dest_provider__'  # Signal that we have a dest provider
        return None

    @property
    def dest_schedule_provider(self):
        """Return the destination-perspective TimeShiftProvider."""
        return self._dest_provider

        self._base = base_provider
        self._target_dep_out = target_dep_time_outbound
        self._flying_out = flying_time_outbound
        self._flying_ret = flying_time_return

        # Calculate arrival times from flying time
        dep_out_mins = parse_time_hhmm(target_dep_time_outbound)
        arr_hub_mins = (dep_out_mins + flying_time_outbound) % 1440
        self._arr_hub = minutes_to_hhmm(arr_hub_mins)

        # Return service: if not specified, calculate to maintain turnaround
        if target_dep_time_return:
            self._target_dep_ret = target_dep_time_return
            dep_ret_mins = parse_time_hhmm(target_dep_time_return)
        else:
            # Default: departure from hub = arrival at hub + reasonable turnaround (90 mins)
            dep_ret_mins = (arr_hub_mins + 90) % 1440
            self._target_dep_ret = minutes_to_hhmm(dep_ret_mins)

        arr_origin_mins = (dep_ret_mins + flying_time_return) % 1440
        self._arr_origin = minutes_to_hhmm(arr_origin_mins)

        self._cache: Dict[str, List[Itinerary]] = {}
        self._build_log: List[str] = []
        self._stats: Dict[str, Any] = {}

        self._log(f"TimeShiftProvider: {base_provider.origin_airport}{base_provider.dest_airport}")
        self._log(f"  Outbound: dep {target_dep_time_outbound}  arr {self._arr_hub} "
                  f"(flying {flying_time_outbound} min)")
        self._log(f"  Return:   dep {self._target_dep_ret}  arr {self._arr_origin} "
                  f"(flying {flying_time_return} min)")

    def _log(self, msg: str):
        self._build_log.append(msg)

    #  ScheduleProvider interface 

    def get_itineraries(self, direction: str) -> List[Itinerary]:
        if direction in self._cache:
            return self._cache[direction]

        if direction == 'qsi1':
            result = self._build_shifted_qsi1()
        elif direction == 'qsi2':
            result = self._build_shifted_qsi2()
        else:
            result = []

        self._cache[direction] = result
        return result

    def get_metadata(self) -> Dict[str, Any]:
        base_meta = self._base.get_metadata()
        return {
            'provider_type': 'TimeShiftProvider',
            'base_provider': base_meta.get('provider_type', 'unknown'),
            'qsi_file': base_meta.get('qsi_file', ''),
            'origin': self._base.origin_airport,
            'destination': self._base.dest_airport,
            'target_dep_outbound': self._target_dep_out,
            'target_arr_hub': self._arr_hub,
            'target_dep_return': self._target_dep_ret,
            'target_arr_origin': self._arr_origin,
            'stats': self._stats,
            'build_log': self._build_log,
        }

    #  Time-shifted connection building 

    def _modify_proposed_flights(self, legs: List[Dict],
                                 new_dep: str, new_arr: str,
                                 new_flying_mins: int) -> List[Dict]:
        """
        Find and modify the proposed carrier's direct flights in the leg data.

        Identifies flights by: carrier matches proposed_carrier AND
        route connects the origin and destination airports (in EITHER direction,
        since the leg sheets contain flights in various directions).

        Modifies dep_time, arr_time, and flying_mins. All other flights untouched.

        Returns a new list (does not mutate the originals).
        """
        carrier = self._base.proposed_carrier
        # Get catchment airports for both origin and destination
        origin_catchment = self._get_catchment_airports(self._base.origin_airport)
        dest_catchment = self._get_catchment_airports(self._base.dest_airport)
        # The proposed service flies between origin and destination
        # It could be dep=origin,arr=dest OR dep=dest,arr=origin
        modified = []
        n_changed = 0

        for leg in legs:
            is_proposed = False
            if leg['carrier'] == carrier:
                dep = leg['dep_airport']
                arr = leg['arr_airport']
                # Check both directions
                if ((dep in origin_catchment and arr in dest_catchment) or
                    (dep in dest_catchment and arr in origin_catchment)):
                    is_proposed = True

            if is_proposed:
                # Clone and modify
                new_leg = dict(leg)
                new_dep_mins = parse_time_hhmm(new_dep)
                new_arr_mins = parse_time_hhmm(new_arr)
                new_leg['dep_time'] = new_dep
                new_leg['arr_time'] = new_arr
                new_leg['dep_time_mins'] = new_dep_mins
                new_leg['arr_time_mins'] = new_arr_mins
                new_leg['flying_mins'] = new_flying_mins
                ft_hhmm = (new_flying_mins // 60) * 100 + (new_flying_mins % 60)
                new_leg['flying_time'] = str(ft_hhmm).zfill(4)
                modified.append(new_leg)
                n_changed += 1
                self._log(f"     Shifted {carrier} {dep}{arr}: dep={new_dep} arr={new_arr}")
            else:
                modified.append(leg)

        self._log(f"    Modified {n_changed} proposed {carrier} flights to dep={new_dep} arr={new_arr}")
        return modified

    def _get_catchment_airports(self, airport: str) -> set:
        """Get catchment airports for a given airport code.

        For major multi-airport cities, include all airports:
          LHR: also LGW, STN, LTN, SEN, LCY
          SJC: also SFO, OAK
        For most airports: just the airport itself.
        """
        # Bay Area
        if airport in ('SJC', 'SFO', 'OAK'):
            return {'SJC', 'SFO', 'OAK'}
        # London
        if airport in ('LHR', 'LGW', 'STN', 'LTN', 'SEN', 'LCY'):
            return {'LHR', 'LGW', 'STN', 'LTN', 'SEN', 'LCY'}
        # New York
        if airport in ('JFK', 'EWR', 'LGA'):
            return {'JFK', 'EWR', 'LGA'}
        # Washington DC
        if airport in ('IAD', 'DCA', 'BWI'):
            return {'IAD', 'DCA', 'BWI'}
        return {airport}

    def _build_shifted_qsi1(self) -> List[Itinerary]:
        """
        Build QSI 1 with shifted proposed service times.

        QSI 1 uses Leg 1.1 (originhub) + Leg 2.1 (hubbeyond).
        The proposed service appears in Leg 1.1 as originhub.
        Shifting its arrival time at the hub changes which Leg 2.1
        connections are valid.
        """
        self._log(f"\n{'='*60}")
        self._log(f"TimeShift QSI 1: dep={self._target_dep_out}")
        self._log(f"{'='*60}")

        # Load original leg data
        leg1_all = load_legs_from_qsi_template(self._base.qsi_file, 'Leg 1.1')
        leg2_all = load_legs_from_qsi_template(self._base.qsi_file, 'Leg 2.1')
        self._log(f"  Leg 1.1: {len(leg1_all)} flights, Leg 2.1: {len(leg2_all)} flights")

        if not leg1_all or not leg2_all:
            return []

        # Modify proposed service in Leg 1.1 (outbound: origin  hub)
        # The proposed carrier's direct flight connects origindestination
        # (e.g., BA SJCLHR). Shifting its arrival time at the hub changes
        # which Leg 2.1 connections are valid.
        leg1_modified = self._modify_proposed_flights(
            leg1_all,
            new_dep=self._target_dep_out,
            new_arr=self._arr_hub,
            new_flying_mins=self._flying_out,
        )

        # Also modify any proposed carrier flights in Leg 2.1 that are
        # hub  catchment (the "second leg" of the return journey).
        # These appear as hub  dest in Leg 2.1 for bidirectional connections.
        # Actually, Leg 2.1 = hub  beyond  the proposed service from hub
        # to origin would appear here only in the return direction context.
        # For the outbound QSI (origin  hub  beyond), the proposed service
        # is only in Leg 1.1. Leg 2.1 contains competitor connections beyond hub.
        # So we DON'T modify Leg 2.1  only the proposed service's arrival at
        # the hub matters for connection timing.

        # Auto-detect hubs
        leg1_arr = set(l['arr_airport'] for l in leg1_modified)
        leg2_dep = set(l['dep_airport'] for l in leg2_all)
        hubs = sorted(leg1_arr & leg2_dep)
        self._log(f"  Hubs: {len(hubs)}")

        # Build connections at each hub
        all_itineraries = []
        hub_stats = {}

        for hub in hubs:
            leg1 = [l for l in leg1_modified if l['arr_airport'] == hub]
            leg2 = [l for l in leg2_all if l['dep_airport'] == hub]
            if not leg1 or not leg2:
                continue

            mct = self._base._load_mct(hub)
            valid, failed = build_connections(
                leg1, leg2, self._base._alliances, mct, self._base._lcc_set,
                self._base.min_connect, self._base.max_connect, self._base.default_mct,
                hub_airport=hub)

            hub_itineraries = []
            for cnx in valid:
                beyond_apt = cnx.get('airport_label', '') or cnx.get('arr_airport', '')
                beyond_city = self._base._apt_to_city(beyond_apt)
                it = connection_to_itinerary(cnx, city_code=beyond_city,
                                             airport_code=beyond_apt)
                hub_itineraries.append(it)

            cities = set(it.city for it in hub_itineraries)
            hub_stats[hub] = {
                'itineraries': len(hub_itineraries),
                'cities': len(cities),
            }

            all_itineraries.extend(hub_itineraries)

        total_cities = set(it.city for it in all_itineraries)
        self._stats['qsi1'] = {
            'total_itineraries': len(all_itineraries),
            'total_cities': len(total_cities),
            'hubs': len(hub_stats),
        }
        self._log(f"  QSI 1 total: {len(all_itineraries)} itineraries, "
                  f"{len(total_cities)} cities")
        return all_itineraries

    def _build_shifted_qsi2(self) -> List[Itinerary]:
        """
        Build QSI 2 with shifted proposed service times.

        QSI 2 uses Leg 1.2 (beyondhub) + Leg 2.2 (huborigin).
        The proposed service appears in Leg 2.2 as huborigin (return).
        Shifting its departure time from hub changes which Leg 1.2
        arrivals can connect to it.
        """
        self._log(f"\n{'='*60}")
        self._log(f"TimeShift QSI 2: return dep={self._target_dep_ret}")
        self._log(f"{'='*60}")

        leg1_all = load_legs_from_qsi_template(self._base.qsi_file, 'Leg 1.2')
        leg2_all = load_legs_from_qsi_template(self._base.qsi_file, 'Leg 2.2')
        self._log(f"  Leg 1.2: {len(leg1_all)} flights, Leg 2.2: {len(leg2_all)} flights")

        if not leg1_all or not leg2_all:
            return []

        # Modify proposed service in Leg 2.2 (return: hub  origin)
        # e.g., BA LHR  SJC
        leg2_modified = self._modify_proposed_flights(
            leg2_all,
            new_dep=self._target_dep_ret,
            new_arr=self._arr_origin,
            new_flying_mins=self._flying_ret,
        )

        # Leg 1.2 (beyond  hub) is not modified  competitor schedules stay fixed

        # Auto-detect hubs
        leg1_arr = set(l['arr_airport'] for l in leg1_all)
        leg2_dep = set(l['dep_airport'] for l in leg2_modified)
        hubs = sorted(leg1_arr & leg2_dep)
        self._log(f"  Hubs: {len(hubs)}")

        all_itineraries = []
        hub_stats = {}

        for hub in hubs:
            leg1 = [l for l in leg1_all if l['arr_airport'] == hub]
            leg2 = [l for l in leg2_modified if l['dep_airport'] == hub]
            if not leg1 or not leg2:
                continue

            mct = self._base._load_mct(hub)
            valid, failed = build_connections(
                leg1, leg2, self._base._alliances, mct, self._base._lcc_set,
                self._base.min_connect, self._base.max_connect, self._base.default_mct,
                hub_airport=hub)

            hub_itineraries = []
            for cnx in valid:
                beyond_apt = cnx.get('dep_airport', '')
                beyond_city = self._base._apt_to_city(beyond_apt)
                it = connection_to_itinerary(cnx, city_code=beyond_city,
                                             airport_code=beyond_apt)
                hub_itineraries.append(it)

            cities = set(it.city for it in hub_itineraries)
            hub_stats[hub] = {
                'itineraries': len(hub_itineraries),
                'cities': len(cities),
            }

            all_itineraries.extend(hub_itineraries)

        total_cities = set(it.city for it in all_itineraries)
        self._stats['qsi2'] = {
            'total_itineraries': len(all_itineraries),
            'total_cities': len(total_cities),
            'hubs': len(hub_stats),
        }
        self._log(f"  QSI 2 total: {len(all_itineraries)} itineraries, "
                  f"{len(total_cities)} cities")
        return all_itineraries


def _run_pipeline_with_dest_override(config, dest_provider, output_path=None):
    """
    Run the pipeline with explicit dest schedule provider override.
    
    This works around the pipeline's assumption that dest QSI comes from a 
    qsi2_file path on the ExcelScheduleProvider. Instead, we monkey-patch the
    run_pipeline function's dest schedule resolution to use our TimeShiftProvider.
    
    The approach: we temporarily add a dest_schedule_provider attribute to the
    schedule_provider, then import and call the original run_pipeline. If the
    pipeline doesn't recognize it (original code), we fall back to a manual
    orchestration.
    """
    from closed_loop_pipeline_v2 import QSIEngine, ForecastAssembler, calibration_analysis, validate, write_output
    from route_config import RouteConfig as RC
    
    # Step 1: QSI scoring for home hub
    qsi_home = QSIEngine(config)
    lhr_captures = qsi_home.run(config.schedule_provider)
    
    # Step 2: QSI scoring for destination using our dest TimeShiftProvider
    dest_config = RC()
    dest_config.airline_code = config.airline_code
    dest_config.home_airport_code = config.dest_airport_code
    dest_config.dest_airport_code = config.home_airport_code
    dest_config.aircraft_type = config.aircraft_type
    dest_config.seats = config.seats
    dest_config.frequency = config.frequency
    dest_config.online_coeff = config.online_coeff
    dest_config.alliance_coeff = config.alliance_coeff
    dest_config.interline_coeff = config.interline_coeff
    dest_config.et_decay_factor = config.et_decay_factor
    dest_config.et_decay_interval = config.et_decay_interval
    
    # Use the time-shifted dest provider (or fall back to home provider)
    dest_schedule = dest_provider if dest_provider else config.schedule_provider
    
    qsi_dest = QSIEngine(dest_config)
    sjc_captures = qsi_dest.run(dest_schedule)
    
    # Step 3: Forecast assembly
    assembler = ForecastAssembler(config)
    results = assembler.run(config.demand_provider, lhr_captures, sjc_captures)
    results['assembler_audit'] = assembler.audit
    
    # Step 4: Calibration analysis
    cal = calibration_analysis(results['home_results'])
    results['calibration'] = cal
    
    # Step 5: Validation
    passed = validate(config, results)
    results['validation_passed'] = passed
    
    # Step 6: Output
    if output_path:
        write_output(config, results, qsi_home, output_path)
    
    return results


# ============================================================================
# TIME GRID RUNNER
# ============================================================================

class TimeGridRunner:
    """
    Runs the full pipeline at multiple departure times and compares results.

    Usage:
        runner = TimeGridRunner(
            base_config=config,                    # RouteConfig with providers
            base_provider=single_extract_provider,  # The SingleExtractOAGProvider
            origin='SJC',
            destination='LHR',
            carrier='BA',
            flying_time_outbound=625,   # mins SJCLHR
            flying_time_return=660,     # mins LHRSJC
        )

        # Define grid
        scenarios = runner.generate_grid(
            start='14:00', end='23:00', step_minutes=60
        )

        # Run
        result = runner.run(scenarios)
        print(result.summary())
    """

    def __init__(self,
                 base_config,               # RouteConfig
                 base_provider: SingleExtractOAGProvider,
                 origin: str = 'SJC',
                 destination: str = 'LHR',
                 carrier: str = 'BA',
                 flying_time_outbound: int = 625,
                 flying_time_return: int = 660,
                 dest_qsi_file: Optional[str] = None,
                 dest_base_provider: Optional[SingleExtractOAGProvider] = None,
                 callback=None):            # Optional progress callback
        """
        Args:
            base_config: RouteConfig (will be cloned per scenario)
            base_provider: The SingleExtractOAGProvider for home QSI (e.g., QSILHR)
            origin: Origin airport code
            destination: Destination airport code
            carrier: Proposed carrier code
            flying_time_outbound: Block time origindestination (minutes)
            flying_time_return: Block time destinationorigin (minutes)
            dest_qsi_file: Optional path to dest QSI file (e.g., QSISJC.xlsx)
                If provided, a dest-side TimeShiftProvider is created for each scenario
            dest_base_provider: Optional pre-built SingleExtractOAGProvider for dest
            callback: Optional function(scenario_idx, n_total, dep_time, status)
        """
        self._config = base_config
        self._base_provider = base_provider
        self._origin = origin
        self._dest = destination
        self._carrier = carrier
        self._ft_out = flying_time_outbound
        self._ft_ret = flying_time_return
        self._callback = callback

        # Dest provider for bidirectional QSI
        self._dest_base = dest_base_provider
        if not dest_base_provider and dest_qsi_file and os.path.exists(dest_qsi_file):
            self._dest_base = SingleExtractOAGProvider(
                qsi_file=dest_qsi_file,
                origin_airport=destination,  # QSISJC has SJC as "home"
                dest_airport=origin,
                proposed_carrier=carrier,
                use_city_codes=base_provider.use_city_codes,
                mct_files=base_provider.mct_files,
            )

    def generate_grid(self, start: str = '14:00', end: str = '23:00',
                      step_minutes: int = 60, max_scenarios: int = 12,
                      return_dep_times: Optional[List[str]] = None) -> List[Dict]:
        """
        Generate a list of departure time scenarios.

        Args:
            start: Grid start time (HH:MM format, origin departure)
            end: Grid end time (HH:MM format)
            step_minutes: Step interval in minutes
            max_scenarios: Maximum number of scenarios
            return_dep_times: Optional explicit return departure times
                If None, calculated from outbound arrival + 90min turnaround

        Returns:
            List of dicts with 'dep_outbound' and 'dep_return' keys (HHMM strings)
        """
        start_mins = _parse_hhmm(start)
        end_mins = _parse_hhmm(end)

        scenarios = []
        current = start_mins

        while current <= end_mins and len(scenarios) < max_scenarios:
            dep_out = minutes_to_hhmm(current)
            arr_hub = minutes_to_hhmm((current + self._ft_out) % 1440)

            # Return departure
            if return_dep_times and len(return_dep_times) > len(scenarios):
                dep_ret = return_dep_times[len(scenarios)]
            else:
                # Default: arrival at hub + 90 min turnaround
                dep_ret_mins = ((current + self._ft_out) % 1440 + 90) % 1440
                dep_ret = minutes_to_hhmm(dep_ret_mins)

            arr_origin = minutes_to_hhmm(
                (parse_time_hhmm(dep_ret) + self._ft_ret) % 1440)

            scenarios.append({
                'dep_outbound': dep_out,
                'arr_hub': arr_hub,
                'dep_return': dep_ret,
                'arr_origin': arr_origin,
            })

            current += step_minutes

        return scenarios

    def run(self, scenarios: List[Dict],
            output_dir: Optional[str] = None) -> GridSearchResult:
        """
        Run the full pipeline for each scenario.

        Args:
            scenarios: List from generate_grid()
            output_dir: Optional directory to save per-scenario output files

        Returns:
            GridSearchResult with ranked scenarios
        """
        from closed_loop_pipeline_v2 import run_pipeline

        result = GridSearchResult(
            origin=self._origin,
            destination=self._dest,
            carrier=self._carrier,
            flying_time_outbound=self._ft_out,
            flying_time_return=self._ft_ret,
        )

        t0 = time_module.time()

        for idx, scenario in enumerate(scenarios):
            dep_out = scenario['dep_outbound']
            dep_ret = scenario['dep_return']

            if self._callback:
                self._callback(idx, len(scenarios), dep_out, 'running')

            t_start = time_module.time()
            try:
                # Create time-shifted provider for home perspective
                shifted_home = TimeShiftProvider(
                    base_provider=self._base_provider,
                    target_dep_time_outbound=dep_out,
                    target_dep_time_return=dep_ret,
                    flying_time_outbound=self._ft_out,
                    flying_time_return=self._ft_ret,
                )

                # Create time-shifted provider for dest perspective (if available)
                shifted_dest = None
                if self._dest_base:
                    # For the dest perspective, the "outbound" direction is reversed:
                    # dep from hub = dep_ret, arr at origin = arr_origin
                    shifted_dest = TimeShiftProvider(
                        base_provider=self._dest_base,
                        target_dep_time_outbound=dep_ret,     # Dest "outbound" = hub  origin
                        target_dep_time_return=dep_out,        # Dest "return" = origin  hub
                        flying_time_outbound=self._ft_ret,    # Flying time huborigin
                        flying_time_return=self._ft_out,       # Flying time originhub
                    )
                    shifted_home._dest_provider = shifted_dest

                # Clone config and swap provider
                cfg = _clone_config(self._config)
                cfg.schedule_provider = shifted_home

                # Output path for this scenario
                output_path = None
                if output_dir:
                    output_path = os.path.join(
                        output_dir, f"grid_{dep_out.replace(':','')}_{dep_ret.replace(':','')}.xlsx")

                # Run pipeline with dest schedule override
                pipeline_results = _run_pipeline_with_dest_override(
                    cfg, shifted_dest, output_path
                )

                # Extract metrics
                elapsed = time_module.time() - t_start
                sr = TimeScenarioResult(
                    dep_time_origin=dep_out,
                    arr_time_hub=scenario.get('arr_hub', ''),
                    dep_time_hub=dep_ret,
                    arr_time_origin=scenario.get('arr_origin', ''),
                    grand_total=int(round(pipeline_results.get('grand_total', 0))),
                    p2p_total=pipeline_results.get('p2p_total', 0),
                    cnx_home_total=pipeline_results.get('home_total', 0),
                    cnx_dest_total=pipeline_results.get('dest_total', 0),
                    load_factor=pipeline_results.get('load_factor', 0),
                    n_cities_home=len(pipeline_results.get('home_results', [])),
                    n_cities_dest=len(pipeline_results.get('dest_results', [])),
                    total_capture_home=sum(
                        c.get('adj_avg_fair_share', 0)
                        for c in pipeline_results.get('home_results', [])
                    ),
                    total_capture_dest=sum(
                        c.get('adj_avg_fair_share', 0)
                        for c in pipeline_results.get('dest_results', [])
                    ),
                    itineraries_qsi1=len(shifted_home.get_itineraries('qsi1')),
                    itineraries_qsi2=len(shifted_home.get_itineraries('qsi2')),
                    elapsed_seconds=elapsed,
                    full_results=pipeline_results,
                )
                result.scenarios.append(sr)

            except Exception as e:
                elapsed = time_module.time() - t_start
                sr = TimeScenarioResult(
                    dep_time_origin=dep_out,
                    dep_time_hub=dep_ret,
                    error=str(e),
                    elapsed_seconds=elapsed,
                )
                result.scenarios.append(sr)

            if self._callback:
                status = 'error' if sr.error else 'done'
                self._callback(idx, len(scenarios), dep_out, status)

        # Rank scenarios by grand_total (descending)
        successful = [s for s in result.scenarios if not s.error]
        result.ranked = sorted(successful, key=lambda s: -s.grand_total)
        if result.ranked:
            result.best_scenario = result.ranked[0]

        # City-level sensitivity analysis
        if len(successful) >= 2:
            result.city_sensitivity = self._compute_city_sensitivity(successful)

        result.total_elapsed = time_module.time() - t0
        return result

    def _compute_city_sensitivity(self, scenarios: List[TimeScenarioResult]) -> List[Dict]:
        """
        Compare per-city connecting traffic across scenarios to identify
        which cities are most sensitive to departure time.
        """
        # Collect home connecting traffic per city across scenarios
        city_data = defaultdict(dict)
        for s in scenarios:
            home_results = s.full_results.get('home_results', [])
            for city_rec in home_results:
                city = city_rec.get('city', '')
                if city and city != 'Total':
                    city_data[city][s.dep_time_origin] = city_rec.get('connecting_pax', 0)

        # Compute sensitivity metrics
        sensitivity = []
        for city, time_vals in city_data.items():
            vals = list(time_vals.values())
            if not vals:
                continue
            max_v = max(vals)
            min_v = min(vals)
            rng = max_v - min_v
            best_time = max(time_vals, key=time_vals.get)
            sensitivity.append({
                'city': city,
                'max_pax': max_v,
                'min_pax': min_v,
                'range': rng,
                'range_pct': (rng / max_v * 100) if max_v > 0 else 0,
                'best_time': best_time,
                'values_by_time': dict(time_vals),
            })

        sensitivity.sort(key=lambda x: -x['range'])
        return sensitivity


# ============================================================================
# HELPERS
# ============================================================================

def _parse_hhmm(time_str: str) -> int:
    """Parse HH:MM or HHMM to minutes since midnight."""
    s = time_str.strip().replace(':', '')
    if not s:
        return 0
    if len(s) <= 2:
        return int(s) * 60
    h = int(s[:-2])
    m = int(s[-2:])
    return h * 60 + m


def _clone_config(config):
    """Deep-clone a RouteConfig, preserving provider references."""
    from route_config import RouteConfig
    new = RouteConfig()
    # Copy all simple attributes
    for attr in dir(config):
        if attr.startswith('_'):
            continue
        try:
            val = getattr(config, attr)
            if not callable(val):
                setattr(new, attr, val)
        except (AttributeError, TypeError):
            pass
    return new


# ============================================================================
# EXCEL OUTPUT FOR GRID RESULTS
# ============================================================================

def write_grid_output(result: GridSearchResult, output_path: str):
    """Write grid search results to a branded Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Styles
    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='0a1628')
    gold_fill = PatternFill('solid', fgColor='e8a83e')
    gold_font = Font(name='Arial', bold=True, size=10, color='0a1628')
    data_font = Font(name='Arial', size=10)
    good_fill = PatternFill('solid', fgColor='C6EFCE')
    bad_fill = PatternFill('solid', fgColor='FFC7CE')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    #  Sheet 1: Summary 
    ws = wb.active
    ws.title = 'Grid Summary'
    ws.sheet_properties.tabColor = '0a1628'

    ws.merge_cells('A1:K1')
    ws['A1'] = f'Departure Time Grid  {result.carrier} {result.origin}{result.destination}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='0a1628')

    ws['A3'] = 'Route:'
    ws['B3'] = f'{result.origin}  {result.destination}'
    ws['A4'] = 'Carrier:'
    ws['B4'] = result.carrier
    ws['A5'] = 'Scenarios:'
    ws['B5'] = len(result.scenarios)
    ws['A6'] = 'Block time out:'
    ws['B6'] = f'{result.flying_time_outbound // 60}h {result.flying_time_outbound % 60:02d}m'
    ws['A7'] = 'Block time ret:'
    ws['B7'] = f'{result.flying_time_return // 60}h {result.flying_time_return % 60:02d}m'
    for r in range(3, 8):
        ws[f'A{r}'].font = Font(name='Arial', bold=True, size=10)
        ws[f'B{r}'].font = data_font

    # Scenario table
    row = 9
    headers = ['Rank', f'Dep {result.origin}', f'Arr {result.destination}',
               f'Dep {result.destination}', f'Arr {result.origin}',
               'Grand Total', 'P2P', 'Cnx Home', 'Cnx Dest',
               'Load Factor', ' vs Best']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for rank, s in enumerate(result.ranked, 1):
        row += 1
        best_total = result.best_scenario.grand_total if result.best_scenario else 0
        delta = s.grand_total - best_total

        vals = [rank, s.dep_time_origin, s.arr_time_hub,
                s.dep_time_hub, s.arr_time_origin,
                s.grand_total, round(s.p2p_total),
                round(s.cnx_home_total), round(s.cnx_dest_total),
                s.load_factor, delta]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = data_font
            cell.border = thin_border
            if c == 6:
                cell.number_format = '#,##0'
            elif c in (7, 8, 9):
                cell.number_format = '#,##0'
            elif c == 10:
                cell.number_format = '0.0%'
            elif c == 11:
                cell.number_format = '+#,##0;-#,##0;0'

        if rank == 1:
            for c in range(1, len(vals) + 1):
                ws.cell(row=row, column=c).fill = good_fill

    # Best recommendation
    row += 3
    if result.best_scenario:
        ws.cell(row=row, column=1,
                value='RECOMMENDATION').font = Font(name='Arial', bold=True,
                                                    size=12, color='0a1628')
        row += 1
        best = result.best_scenario
        ws.cell(row=row, column=1,
                value=f'Optimal: {best.dep_time_origin} dep {result.origin}  '
                      f'{best.dep_time_hub} dep {result.destination}').font = data_font
        row += 1
        ws.cell(row=row, column=1,
                value=f'Forecast: {best.grand_total:,} pax, '
                      f'{best.load_factor:.1%} LF').font = data_font

        if len(result.ranked) >= 2:
            worst = result.ranked[-1]
            diff = best.grand_total - worst.grand_total
            pct = (diff / worst.grand_total * 100) if worst.grand_total else 0
            row += 1
            ws.cell(row=row, column=1,
                    value=f'Range: {diff:+,} pax ({pct:+.1f}%) between '
                          f'best and worst').font = data_font

    for c, w in enumerate([8, 12, 12, 12, 12, 12, 10, 10, 10, 12, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    #  Sheet 2: City Sensitivity 
    if result.city_sensitivity:
        ws2 = wb.create_sheet('City Sensitivity')
        ws2.sheet_properties.tabColor = '336699'

        ws2.merge_cells('A1:F1')
        ws2['A1'] = 'Per-City Sensitivity to Departure Time'
        ws2['A1'].font = Font(name='Arial', bold=True, size=12, color='0a1628')

        row = 3
        time_labels = sorted(set(
            t for s in result.city_sensitivity for t in s.get('values_by_time', {})
        ))
        headers2 = ['City'] + time_labels + ['Best Time', 'Range (pax)', 'Range %']
        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for cs in result.city_sensitivity[:50]:  # Top 50 most sensitive
            row += 1
            ws2.cell(row=row, column=1, value=cs['city']).font = data_font
            ws2.cell(row=row, column=1).border = thin_border

            col = 2
            for t in time_labels:
                val = cs.get('values_by_time', {}).get(t, 0)
                cell = ws2.cell(row=row, column=col, value=round(val))
                cell.font = data_font
                cell.number_format = '#,##0'
                cell.border = thin_border
                col += 1

            ws2.cell(row=row, column=col, value=cs.get('best_time', '')).font = data_font
            ws2.cell(row=row, column=col).border = thin_border
            col += 1
            cell = ws2.cell(row=row, column=col, value=round(cs.get('range', 0)))
            cell.font = data_font
            cell.number_format = '#,##0'
            cell.border = thin_border
            col += 1
            cell = ws2.cell(row=row, column=col, value=cs.get('range_pct', 0) / 100)
            cell.font = data_font
            cell.number_format = '0.0%'
            cell.border = thin_border

        for c in range(1, len(headers2) + 1):
            ws2.column_dimensions[get_column_letter(c)].width = 14
        ws2.column_dimensions['A'].width = 8

    wb.save(output_path)
    return output_path


# ============================================================================
# STANDALONE VALIDATION
# ============================================================================

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Departure Time Grid Search')
    parser.add_argument('--qsi-file', default=str(REFERENCE_CASE_DIR / 'QSILHR_v1_OS_JZ_17Feb15.xlsx'))
    parser.add_argument('--origin', default='LHR')
    parser.add_argument('--dest', default='SJC')
    parser.add_argument('--carrier', default='BA')
    parser.add_argument('--start', default='15:00', help='Grid start (HH:MM)')
    parser.add_argument('--end', default='22:00', help='Grid end (HH:MM)')
    parser.add_argument('--step', type=int, default=60, help='Step minutes')
    parser.add_argument('--quick', action='store_true',
                        help='Quick test  just build providers, no pipeline run')
    args = parser.parse_args()

    # Create base provider
    mct_files = {
        'LHR': str(REFERENCE_CASE_DIR / 'LHR_MCTs.xls'),
        'SJC': str(REFERENCE_CASE_DIR / 'Minimum_Cnx_Times_SJC.xls'),
    }

    print(f"Creating SingleExtractOAGProvider from {os.path.basename(args.qsi_file)}")
    base = SingleExtractOAGProvider(
        qsi_file=args.qsi_file,
        origin_airport=args.origin,
        dest_airport=args.dest,
        proposed_carrier=args.carrier,
        use_city_codes=True,
        mct_files=mct_files,
    )

    if args.quick:
        # Quick test: just verify time-shifting produces different itinerary counts
        print(f"\nQuick test: comparing itinerary counts at different times")
        print(f"{'Time':>8s}  {'QSI1 its':>10s}  {'QSI1 cities':>12s}  {'QSI2 its':>10s}  {'QSI2 cities':>12s}")
        print('-' * 60)

        times = ['1700', '2000', '2130', '2200']
        for t in times:
            shifted = TimeShiftProvider(
                base_provider=base,
                target_dep_time_outbound=t,
                flying_time_outbound=625,
                flying_time_return=660,
            )
            q1 = shifted.get_itineraries('qsi1')
            q2 = shifted.get_itineraries('qsi2')
            n1 = len(q1)
            c1 = len(set(it.city for it in q1))
            n2 = len(q2)
            c2 = len(set(it.city for it in q2))
            print(f"{t:>8s}  {n1:>10,}  {c1:>12}  {n2:>10,}  {c2:>12}")
        print("\nDone. Different counts confirm time-shifting affects connections.")
    else:
        print("Full grid search requires pipeline run  use --quick for provider-only test")
