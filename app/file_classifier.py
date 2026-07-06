#!/usr/bin/env python3
"""
Avia Solutions - Data File Classifier
=======================================
Classifies uploaded Excel files by inspecting actual cell content,
not filenames. Built from 20 years of Avia analyst data file patterns.

DATA SIGNATURES (cell-level):

  QSI Template / Scored:
    - Cells contain 'Carrier1', 'Carrier1Name', 'FlightNo1' in header row
    - Cells contain 'From Airport X to Transit Points - Leg'
    - Cells contain 'Route Label', 'Dep. Aprt.', 'Connection'

  MCT (Minimum Connection Times):
    - Cells contain 'MCT Time (min)' or column header 'MCT'
    - Cells contain 'Domestic to Domestic', 'International to'
    - OAG Analyser header text present
    - Typically small files with airport pairs and time values

  MIDT Connecting Demand (Sabre):
    - Sheet named 'Connecting Demand' or similar
    - Cells contain 'Mod Org City Code', 'Direct & Indirect Demand'
    - Cells contain 'Competing Hubs for Indirect Demand'
    - Has 'City Lookup' sheet with airport/city/country columns

  P2P Demand (Sabre):
    - Sheet named 'Point to Point' or 'P2P Demand'
    - Cells contain 'Point-to-Point Demand', 'Competing Hubs for P2P'
    - Has city lookup and pivot sheets like MIDT but P2P-labelled

  City Lookup:
    - Columns: 'Airport Code', 'City Code', 'City Name', 'Country Name'
    - Or simpler: 'City', 'Airport' two-column format
    - Small reference file, no demand data

  OAG Schedule:
    - OAG ANALYSER header text
    - 'Schedules Power Table' reference
    - Destination/arrival airport columns with IATA codes
    - FROM/VIA routing columns

  Forecast Workbook:
    - Sheet 'Proposed Schedule' with Flight/From/Dep.Time columns
    - Sheet 'Demand settings' with city codes and traffic numbers
    - Sheet 'QSI Weightings' or 'Conneting Info'

  Fare Data:
    - Columns with fare/yield/revenue headers
    - Booking class breakdowns (F, J, W, Y etc.)

USAGE:
    from file_classifier import classify_file, FileType

    file_type = classify_file(uploaded_file_or_path)
    # Returns FileType enum: QSI_TEMPLATE, MCT, MIDT_CONNECTING,
    # P2P_DEMAND, CITY_LOOKUP, OAG_SCHEDULE, FORECAST, FARE_DATA, UNKNOWN
"""

from config import REFERENCE_CASE_DIR
from enum import Enum
from typing import Optional, Dict, List, Tuple
import os


class FileType(Enum):
    QSI_TEMPLATE = "qsi_template"        # QSI template with Leg sheets and flight data
    QSI_SCORED = "qsi_scored"            # QSI after scoring (has Calc sheets)
    MCT = "mct"                          # Minimum Connection Times
    MIDT_CONNECTING = "midt_connecting"  # Sabre MIDT connecting demand
    P2P_DEMAND = "p2p_demand"            # Sabre P2P demand
    CITY_LOOKUP = "city_lookup"          # Airport/city code mapping
    OAG_SCHEDULE = "oag_schedule"        # OAG schedule/capacity data
    FORECAST = "forecast"                # Completed forecast workbook
    CNX_BUILDER = "cnx_builder"          # Connection builder workbook
    FARE_DATA = "fare_data"              # Fare/yield data
    CALIBRATION = "calibration"          # QSI calibration file (variant of QSI_SCORED)
    UNKNOWN = "unknown"


# ============================================================================
# SHEET-LEVEL SIGNATURES
# ============================================================================

def _get_sheet_names(file_or_path) -> List[str]:
    """Get sheet names from a file (path string or file-like object)."""
    try:
        if isinstance(file_or_path, str):
            if file_or_path.lower().endswith('.xls') and not file_or_path.lower().endswith(('.xlsx', '.xlsm')):
                import xlrd
                wb = xlrd.open_workbook(file_or_path)
                return wb.sheet_names()
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_or_path, read_only=True, data_only=True)
                names = wb.sheetnames
                wb.close()
                return names
        else:
            # File-like object (Streamlit upload)
            name = getattr(file_or_path, 'name', '').upper()
            file_or_path.seek(0)
            if name.endswith('.XLS') and not name.endswith(('.XLSX', '.XLSM')):
                import xlrd
                data = file_or_path.read()
                file_or_path.seek(0)
                wb = xlrd.open_workbook(file_contents=data)
                return wb.sheet_names()
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_or_path, read_only=True, data_only=True)
                names = wb.sheetnames
                wb.close()
                file_or_path.seek(0)
                return names
    except Exception:
        if hasattr(file_or_path, 'seek'):
            file_or_path.seek(0)
        return []


def _read_cells(file_or_path, sheet_index=0, max_rows=10, max_cols=15) -> List[List]:
    """Read first N rows of a specific sheet. Returns list of row-lists."""
    rows = []
    try:
        if isinstance(file_or_path, str):
            fname = file_or_path
        else:
            fname = getattr(file_or_path, 'name', '').upper()

        is_xls = fname.upper().endswith('.XLS') and not fname.upper().endswith(('.XLSX', '.XLSM'))

        if is_xls:
            import xlrd
            if isinstance(file_or_path, str):
                wb = xlrd.open_workbook(file_or_path)
            else:
                file_or_path.seek(0)
                data = file_or_path.read()
                file_or_path.seek(0)
                wb = xlrd.open_workbook(file_contents=data)
            if sheet_index < wb.nsheets:
                ws = wb.sheet_by_index(sheet_index)
                for r in range(min(max_rows, ws.nrows)):
                    row = []
                    for c in range(min(max_cols, ws.ncols)):
                        row.append(ws.cell_value(r, c))
                    rows.append(row)
        else:
            import openpyxl
            if not isinstance(file_or_path, str):
                file_or_path.seek(0)
            wb = openpyxl.load_workbook(file_or_path, read_only=True, data_only=True)
            if sheet_index < len(wb.worksheets):
                ws = wb.worksheets[sheet_index]
                for r_idx, row in enumerate(ws.iter_rows(
                        max_row=max_rows, max_col=max_cols, values_only=True)):
                    rows.append(list(row))
                    if r_idx >= max_rows - 1:
                        break
            wb.close()
            if not isinstance(file_or_path, str):
                file_or_path.seek(0)
    except Exception:
        if hasattr(file_or_path, 'seek'):
            file_or_path.seek(0)
    return rows


def _flatten_cells(rows: List[List]) -> str:
    """Flatten all cell values into a single uppercase string for searching."""
    parts = []
    for row in rows:
        for cell in row:
            if cell is not None:
                parts.append(str(cell).upper())
    return ' '.join(parts)


# ============================================================================
# CLASSIFICATION ENGINE
# ============================================================================

def classify_file(file_or_path, origin_code: str = '', dest_code: str = '') -> FileType:
    """
    Classify a data file by inspecting its content.

    Args:
        file_or_path: File path (str) or file-like object (Streamlit upload)
        origin_code: Origin airport/city code (optional, helps direction detection)
        dest_code: Destination airport/city code (optional)

    Returns:
        FileType enum value
    """
    sheets = _get_sheet_names(file_or_path)
    sheets_upper = [s.upper() for s in sheets]
    sheet_text = ' '.join(sheets_upper)

    # ------------------------------------------------------------------
    # CHECK 1: Sheet name signatures (most reliable for structured Avia files)
    # ------------------------------------------------------------------

    # QSI Template: has "Leg 1.1", "Leg 2.1" etc in sheet names
    leg_sheets = [s for s in sheets_upper if 'LEG' in s and any(
        x in s for x in ['1.1', '2.1', '1.2', '2.2'])]
    if len(leg_sheets) >= 2:
        # Template vs Scored distinction:
        # Templates have "QSI 1 > Leg 1.1" format (with > separator)
        # Scored files have just "Leg 1.1" without the > prefix
        has_arrow_format = any('>' in s and 'LEG' in s.upper() for s in sheets)
        has_calc = any('CALC' in s for s in sheets_upper)

        if has_arrow_format:
            # This is the working template (even if it has Calc sheets)
            return FileType.QSI_TEMPLATE
        elif has_calc:
            if any('CALIB' in s for s in sheets_upper):
                return FileType.CALIBRATION
            return FileType.QSI_SCORED
        return FileType.QSI_TEMPLATE

    # Forecast: has "Proposed Schedule" AND ("Demand Settings" or "Forecast TABLE")
    # Check BEFORE P2P/MIDT because forecast workbooks contain P2P and demand sheets
    has_proposed = any('PROPOSED SCHEDULE' in s for s in sheets_upper)
    has_demand_settings = any('DEMAND SETTING' in s for s in sheets_upper)
    has_forecast_table = any('FORECAST TABLE' in s or 'FORECAST FINALISED' in s
                             for s in sheets_upper)
    has_forecast_sheet = any(s.strip() == 'FORECAST' for s in sheets_upper)
    if has_proposed and (has_demand_settings or has_forecast_table or has_forecast_sheet):
        return FileType.FORECAST
    if has_forecast_table and has_demand_settings:
        return FileType.FORECAST

    # MIDT Connecting Demand: has "Connecting Demand" sheet
    if any('CONNECTING DEMAND' in s for s in sheets_upper):
        if any('P2P' in s or 'POINT TO POINT' in s for s in sheets_upper):
            return FileType.P2P_DEMAND
        return FileType.MIDT_CONNECTING

    # VS-era Sabre files: have 'Pivot-P2P', 'Pivot-CNX', 'Data' sheets
    has_pivot_p2p = any('PIVOT' in s and 'P2P' in s for s in sheets_upper)
    has_data_sheet = any(s.strip() == 'DATA' for s in sheets_upper)

    if has_pivot_p2p and has_data_sheet:
        # Content-only classification: read the Pivot-P2P sheet
        pivot_idx = next((i for i, s in enumerate(sheets_upper)
                          if 'PIVOT' in s and 'P2P' in s), 0)
        pivot_rows = _read_cells(file_or_path, sheet_index=pivot_idx,
                                 max_rows=12, max_cols=5)

        # Check 1: Look for 'Mod Dest City' row -- if it names a specific city, it's P2P
        found_mod_dest = False
        for pr in pivot_rows:
            if pr and pr[0] and 'mod dest' in str(pr[0]).lower():
                found_mod_dest = True
                dest_val = str(pr[1]).strip() if len(pr) > 1 and pr[1] else ''
                if dest_val == '(All)' or not dest_val:
                    return FileType.MIDT_CONNECTING
                elif len(dest_val) == 3 and dest_val.isalpha():
                    # Specific city pair -- this is P2P demand
                    return FileType.P2P_DEMAND

        # Check 2: No 'Mod Dest City' found -- count data rows with 3-letter city codes
        # If many city rows, it's connecting demand; if few (1-2), it's P2P
        if not found_mod_dest:
            city_count = 0
            for pr in pivot_rows:
                if pr and pr[0]:
                    val = str(pr[0]).strip()
                    if len(val) == 3 and val.isalpha() and val.isupper():
                        city_count += 1
            if city_count >= 3:
                return FileType.MIDT_CONNECTING
            else:
                return FileType.P2P_DEMAND

        return FileType.P2P_DEMAND

    # P2P Demand: has "P2P Demand" or "Point to Point" sheet
    if any('P2P' in s for s in sheets_upper) or \
       any('POINT TO POINT' in s for s in sheets_upper):
        return FileType.P2P_DEMAND

    # OAG Schedule: has sheets like 'Raw Data', 'Schedule Analysis', 'Dest.beyond'
    if any('RAW DATA' in s for s in sheets_upper) or \
       any('SCHEDULE ANALYSIS' in s for s in sheets_upper) or \
       any('DEST.BEYOND' in s or 'DEST BEYOND' in s for s in sheets_upper):
        return FileType.OAG_SCHEDULE

    # Connection builder: has "leg1"/"leg2" and "Run"
    if any('LEG1' in s or 'LEG2' in s for s in sheets_upper) and \
       any('RUN' in s for s in sheets_upper):
        return FileType.CNX_BUILDER

    # OAG Template (empty template with Input/Export/Output structure)
    if any('INPUT' in s for s in sheets_upper) and \
       any('EXPORT' in s for s in sheets_upper):
        return FileType.OAG_SCHEDULE

    # Output/report files (Forecast TABLE, Cnx TABLE etc.)
    if any('FORECAST TABLE' in s for s in sheets_upper) and \
       any('CNX' in s and 'TABLE' in s for s in sheets_upper):
        return FileType.FORECAST

    # Pipeline test output (has Summary + Connecting Home/Dest)
    if any('CONNECTING HOME' in s for s in sheets_upper) and \
       any('CONNECTING DEST' in s for s in sheets_upper):
        return FileType.FORECAST

    # ------------------------------------------------------------------
    # CHECK 2: Cell content signatures (for files without distinctive sheet names)
    # ------------------------------------------------------------------

    # Read first sheet content
    rows = _read_cells(file_or_path, sheet_index=0, max_rows=10, max_cols=15)
    content = _flatten_cells(rows)

    # MCT: contains 'MCT Time' or 'MCT' column with domestic/international patterns
    if 'MCT TIME' in content or \
       ('MCT' in content and ('DOMESTIC' in content or 'INTERNATIONAL' in content)):
        return FileType.MCT

    # MCT variant: OAG format with From-To and time values
    if 'DOMESTIC TO DOMESTIC' in content or 'DOMESTIC TO INTERNATIONAL' in content:
        return FileType.MCT

    # OAG Schedule: contains 'OAG ANALYSER' and schedule-related text
    # Check first three sheets (OAG exports vary in sheet ordering)
    oag_content = content
    if 'OAG ANALYSER' not in oag_content:
        for si in range(1, min(3, len(sheets))):
            rows_si = _read_cells(file_or_path, sheet_index=si, max_rows=8, max_cols=10)
            oag_content = _flatten_cells(rows_si)
            if 'OAG ANALYSER' in oag_content:
                break
    if 'OAG ANALYSER' in oag_content:
        # Could be MCT or schedule - check further
        if 'MCT' in oag_content or 'CONNECTING TIME' in oag_content:
            return FileType.MCT
        return FileType.OAG_SCHEDULE

    # OAG Schedule: first sheet has carrier/flight/airport columns (ADF export)
    if ('CARRIER CODE' in content or 'CARRIER1' in content) and \
       ('DEP AIRPORT' in content or 'ARR AIRPORT' in content) and \
       ('LOCAL DEP TIME' in content or 'FREQUENCY' in content or 'SEATS' in content):
        return FileType.OAG_SCHEDULE

    # QSI by cell content: 'Carrier1', 'FlightNo1', 'From Airport X'
    if 'CARRIER1' in content and ('FLIGHTNO1' in content or 'CARRIER1NAME' in content):
        if any('CALC' in s for s in sheets_upper):
            return FileType.QSI_SCORED
        return FileType.QSI_TEMPLATE

    # QSI by content: 'Route Label', 'Dep. Aprt.'
    if 'ROUTE LABEL' in content and 'DEP' in content:
        return FileType.QSI_SCORED

    # City Lookup: 'Airport Code', 'City Code', 'City Name' columns
    if ('AIRPORT CODE' in content or 'AIRPORT' in content) and \
       ('CITY CODE' in content or 'CITY NAME' in content):
        # Make sure it's not a demand file that happens to have a lookup tab
        if 'CONNECTING DEMAND' not in sheet_text and 'P2P' not in sheet_text:
            return FileType.CITY_LOOKUP

    # City Lookup: simple 'City' + 'Airport' two-column format
    if content.count('CITY') >= 1 and content.count('AIRPORT') >= 1:
        # Check if this is primarily a lookup (small, reference-style)
        if len(sheets) <= 3 and 'DEMAND' not in sheet_text:
            return FileType.CITY_LOOKUP

    # Fare data: contains fare/yield/revenue columns
    if 'FARE' in content and ('YIELD' in content or 'REVENUE' in content or 'BOOKING' in content):
        return FileType.FARE_DATA
    if any('FARE' in s for s in sheets_upper):
        return FileType.FARE_DATA

    # Forecast by content: 'Flight', 'From', 'Dep. Time', 'Arr. Time'
    if 'PROPOSED SCHEDULE' in content or \
       ('FLIGHT' in content and 'DEP' in content and 'ARR' in content):
        return FileType.FORECAST

    # ------------------------------------------------------------------
    # CHECK 3: Try reading other sheets if first sheet was empty/separator
    # ------------------------------------------------------------------
    if not content.strip() or len(content) < 20:
        # First sheet was empty (separator sheet like "Connecting Demand >>")
        # Try second sheet
        if len(sheets) > 1:
            rows2 = _read_cells(file_or_path, sheet_index=1, max_rows=8, max_cols=12)
            content2 = _flatten_cells(rows2)

            if 'DIRECT & INDIRECT DEMAND' in content2 or \
               'EXPAND DATA SOURCE' in content2:
                # MIDT file - check for P2P
                if any('P2P' in s for s in sheets_upper):
                    return FileType.P2P_DEMAND
                return FileType.MIDT_CONNECTING

    # ------------------------------------------------------------------
    # CHECK 4: Structural heuristics
    # ------------------------------------------------------------------

    # .xls files with 1-2 sheets and small data are often MCT exports
    fname = file_or_path if isinstance(file_or_path, str) else getattr(file_or_path, 'name', '')
    if fname.upper().endswith('.XLS') and not fname.upper().endswith(('.XLSX', '.XLSM')):
        if len(sheets) <= 3:
            # Check if data looks like MCT (airport codes + numbers)
            if rows:
                has_times = any(
                    any(isinstance(c, (int, float)) and 10 <= c <= 300 for c in row if c)
                    for row in rows
                )
                has_airports = any(
                    any(isinstance(c, str) and len(c) == 3 and c.isalpha() for c in row if c)
                    for row in rows
                )
                if has_times and has_airports:
                    return FileType.MCT

    # ------------------------------------------------------------------
    # CHECK 5: Sheet name contains 'Lookup' but file has demand sheets too
    # (Some MIDT files include a City Lookup tab)
    # ------------------------------------------------------------------
    if any('LOOKUP' in s for s in sheets_upper):
        if len(sheets) <= 2:
            return FileType.CITY_LOOKUP
        # Multi-sheet file with lookup - probably demand file
        # Already handled above

    return FileType.UNKNOWN


def classify_direction(file_or_path, file_type: FileType,
                       origin_code: str = '', dest_code: str = '') -> str:
    """
    For MIDT_CONNECTING files, determine if this is home or dest side.

    Returns: 'home' or 'dest'
    """
    if file_type != FileType.MIDT_CONNECTING:
        return ''

    orig = origin_code.upper()[:3] if origin_code else ''
    dest = dest_code.upper()[:3] if dest_code else ''

    if not orig or not dest:
        return 'home'  # default

    # Check sheet content for city codes
    sheets = _get_sheet_names(file_or_path)
    for i, sname in enumerate(sheets):
        rows = _read_cells(file_or_path, sheet_index=i, max_rows=6, max_cols=5)
        content = _flatten_cells(rows)

        # Look for 'Mod Org City Code' followed by a city
        if 'MOD ORG CITY' in content:
            # The city code near this label tells us the origin of the demand
            if orig in content:
                return 'home'  # demand originating from home side
            if dest in content:
                return 'dest'  # demand originating from dest side

    # Fallback: check filename for city codes
    fname = (file_or_path if isinstance(file_or_path, str)
             else getattr(file_or_path, 'name', '')).upper()
    if orig and dest:
        orig_pos = fname.find(orig)
        dest_pos = fname.find(dest)
        if orig_pos >= 0 and dest_pos >= 0:
            return 'home' if orig_pos < dest_pos else 'dest'
        if orig_pos >= 0:
            return 'home'
        if dest_pos >= 0:
            return 'dest'

    return 'home'


# ============================================================================
# BATCH CLASSIFICATION
# ============================================================================

def classify_batch(files, origin_code: str = '', dest_code: str = '') -> Dict:
    """
    Classify a batch of files and sort them into role buckets.

    Args:
        files: List of file paths or file-like objects
        origin_code: Origin airport code
        dest_code: Destination airport code

    Returns:
        Dict with keys: qsi_template, mct, home_cnx, dest_cnx, p2p,
        city_lookup, forecast, oag_schedule, fare, unknown
        Each value is a list of (file, FileType) tuples.
    """
    result = {
        'qsi_template': None,       # Single file
        'mct': [],                   # List
        'home_cnx': [],              # List
        'dest_cnx': [],              # List
        'p2p': [],                   # List
        'city_lookup': None,         # Single file
        'forecast': [],              # List
        'oag_schedule': [],          # List
        'fare': [],                  # List
        'unknown': [],               # List
        'classifications': {},       # fname -> FileType for display
    }

    for f in files:
        ft = classify_file(f, origin_code, dest_code)
        fname = f if isinstance(f, str) else getattr(f, 'name', str(f))
        result['classifications'][fname] = ft

        if ft in (FileType.QSI_TEMPLATE, FileType.QSI_SCORED, FileType.CALIBRATION):
            if result['qsi_template'] is None:
                result['qsi_template'] = f
            else:
                # Additional QSI files - treat as supplementary
                result['unknown'].append(f)

        elif ft == FileType.MCT:
            result['mct'].append(f)

        elif ft == FileType.MIDT_CONNECTING:
            direction = classify_direction(f, ft, origin_code, dest_code)
            if direction == 'dest':
                result['dest_cnx'].append(f)
            else:
                result['home_cnx'].append(f)

        elif ft == FileType.P2P_DEMAND:
            result['p2p'].append(f)

        elif ft == FileType.CITY_LOOKUP:
            result['city_lookup'] = f

        elif ft == FileType.OAG_SCHEDULE:
            result['oag_schedule'].append(f)

        elif ft == FileType.FORECAST:
            result['forecast'].append(f)

        elif ft == FileType.FARE_DATA:
            result['fare'].append(f)

        elif ft == FileType.CNX_BUILDER:
            if result['qsi_template'] is None:
                result['qsi_template'] = f
            else:
                result['unknown'].append(f)

        else:
            result['unknown'].append(f)

    return result


# ============================================================================
# SELF-TEST
# ============================================================================

if __name__ == '__main__':
    import sys

    print("Avia File Classifier - Self Test")
    print("=" * 60)

    test_dir = str(REFERENCE_CASE_DIR)
    if not os.path.exists(test_dir):
        print("Project directory not available. Skipping file tests.")
        sys.exit(0)

    test_cases = [
        ("QSI__HomeAirport.xlsm", FileType.QSI_TEMPLATE),
        ("QSILHR.xlsx", FileType.QSI_SCORED),
        ("QSILHR_v1_OS_JZ_17Feb15.xlsx", FileType.QSI_SCORED),
        ("QSI_Caibration_with_new_service_UK_Leisure_1.xlsx", FileType.CALIBRATION),
        ("LHR_MCTs.xls", FileType.MCT),
        ("Minimum_Cnx_Times_SJC.xls", FileType.MCT),
        ("OAG_Minimum_Connecting_Time_DS_20Jan11.xls", FileType.MCT),
        ("LONSJCXXX.xlsx", FileType.MIDT_CONNECTING),
        ("SJCLONXXX__2013_CUT_4_data.xlsx", FileType.MIDT_CONNECTING),
        ("P2P_LONBAY_AREA_2013.xlsx", FileType.P2P_DEMAND),
        ("Bay_Area_Demandxlsx.xlsx", FileType.P2P_DEMAND),
        ("OAG_Airport__City_Lookup_DS_25Feb11.xlsx", FileType.CITY_LOOKUP),
        ("OAG__LHR__WORLD__LHR_AUG2014.xlsx", FileType.OAG_SCHEDULE),
        ("BA_Fcst_LHRSJC_JZ_23Feb2015_FINAL_without_INDIA.xlsm", FileType.FORECAST),
        ("Cnx_Leg1_1_Leg2_1.xlsm", FileType.CNX_BUILDER),
    ]

    passed = 0
    failed = 0
    for fname, expected in test_cases:
        path = os.path.join(test_dir, fname)
        if not os.path.exists(path):
            print(f"  SKIP: {fname} (not found)")
            continue
        actual = classify_file(path)
        ok = actual == expected
        status = "PASS" if ok else "FAIL"
        if ok:
            passed += 1
        else:
            failed += 1
        print(f"  {status}: {fname:55s} -> {actual.value:20s} {'(expected ' + expected.value + ')' if not ok else ''}")

    print(f"\nResults: {passed} passed, {failed} failed")
