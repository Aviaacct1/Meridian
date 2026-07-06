#!/usr/bin/env python3
"""
Avia Solutions  Closed-Loop Pipeline V2 (Chat 12)
====================================================
Refactored pipeline that consumes providers instead of hardcoded file paths.

Architecture:
    RouteConfig  provides ScheduleProvider + DemandProvider
    QSIEngine    scores itineraries from ScheduleProvider
    ForecastAssembler  combines P2P + connecting using DemandProvider + QSI captures
    OutputWriter  generates validated Excel output

Same inputs, same outputs, better architecture.
Regression target: BA LHR-SJC = 129,162 passengers.
"""

from config import OUTPUT_DIR, ensure_output_dir
import os
import sys
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required"); sys.exit(1)

from providers import (
    Itinerary, ScheduleProvider, DemandProvider,
    P2PSegmentData, P2PSubsegmentData, ConnectingCityData,
)
from route_config import RouteConfig


# ============================================================================
# QSI ENGINE  Pure computation, no data knowledge
# ============================================================================

class QSIEngine:
    """
    Scores itineraries and computes competitive market shares.
    
    This is the core QSI computation extracted from closed_loop_pipeline.py.
    It has NO knowledge of where data comes from  it receives Itinerary
    objects from a ScheduleProvider and returns capture rates.
    """

    def __init__(self, config: RouteConfig):
        self.config = config
        self.cnx_coeffs = config.cnx_coeffs
        self.et_factor = config.et_decay_factor
        self.et_interval = config.et_decay_interval
        # Service-level (non-stop vs one-stop) coefficients. These reproduce the
        # established Excel QSI method, whose 'QSI Coefficients' sheet weights a
        # non-stop at 1.0 and a one-stop at 0.20 (two-stop 0.40). Omitting this
        # term understates a non-stop service's fair share by ~4-5x. Config-driven
        # and defaulted so existing configs keep working.
        self.nonstop_coeff = getattr(config, 'nonstop_coeff', 1.0)
        self.onestop_coeff = getattr(config, 'onestop_coeff', 0.20)
        self.audit = []

    def log(self, msg):
        self.audit.append(msg)
        print(msg)

    def et_coeff(self, elapsed_mins: int, min_elapsed_mins: int) -> float:
        """Elapsed time decay coefficient."""
        excess_hrs = (elapsed_mins - min_elapsed_mins) / 60.0
        if excess_hrs <= 0:
            return 1.0
        step = int(excess_hrs / self.et_interval)
        return 1.0 / ((step + 1) ** self.et_factor)

    def score_itineraries(self, itineraries: List[Itinerary]) -> List[Itinerary]:
        """Score all itineraries: compute et_coeff, cnx_coeff, qsi per itinerary."""
        # Find minimum elapsed time per city
        min_el = {}
        for it in itineraries:
            if it.city not in min_el or it.elapsed < min_el[it.city]:
                min_el[it.city] = it.elapsed

        for it in itineraries:
            me = min_el[it.city]
            it.et_coeff = self.et_coeff(it.elapsed, me)
            it.cnx_coeff = self.cnx_coeffs.get(it.cnx_type, 0)
            # Non-stop (no connection point) vs one-stop service-level weight.
            service_coeff = self.nonstop_coeff if not str(it.cnx_airport).strip() else self.onestop_coeff
            it.service_coeff = service_coeff
            it.qsi = it.freq * it.et_coeff * it.cnx_coeff * service_coeff

        return itineraries

    def aggregate_shares(self, scored: List[Itinerary]) -> Tuple[Dict, Dict]:
        """Aggregate QSI by route label and compute fair shares per city."""
        route_qsi = defaultdict(lambda: {'qsi': 0.0, 'city': '', 'airport': ''})
        city_market = defaultdict(float)

        for it in scored:
            rl = it.route_label
            route_qsi[rl]['qsi'] += it.qsi
            route_qsi[rl]['city'] = it.city
            route_qsi[rl]['airport'] = it.airport
            city_market[it.city] += it.qsi

        result = {}
        for rl, data in route_qsi.items():
            cy = data['city']
            mkt = city_market.get(cy, 0)
            result[rl] = {
                'city': cy,
                'airport': data['airport'],
                'carrier_qsi': data['qsi'],
                'market_qsi': mkt,
                'fair_share': data['qsi'] / mkt if mkt > 0 else 0.0,
            }
        return result, city_market

    def bidirectional_calc(self, shares1, shares2, cm1, cm2) -> Dict:
        """Compute bidirectional average fair shares."""
        all_labels = set(shares1.keys()) | set(shares2.keys())
        result = {}
        for rl in sorted(all_labels):
            s1 = shares1.get(rl, {})
            s2 = shares2.get(rl, {})
            city = s1.get('city', '') or s2.get('city', '')
            airport = s1.get('airport', '') or s2.get('airport', '')
            q1c = s1.get('carrier_qsi', 0)
            q1m = s1.get('market_qsi', 0) or cm1.get(city, 0)
            fs1 = s1.get('fair_share', 0) if s1 else 0
            q2c = s2.get('carrier_qsi', 0)
            q2m = s2.get('market_qsi', 0) or cm2.get(city, 0)
            fs2 = s2.get('fair_share', 0) if s2 else 0
            avg = (fs1 + fs2) / 2.0
            rt = 'Exclude' if (q1m == 0 or q2m == 0) else 'Include'
            adj = 0.0 if rt == 'Exclude' else avg
            result[rl] = {
                'city': city, 'airport': airport,
                'q1_carrier': q1c, 'q1_market': q1m, 'fs1': fs1,
                'q2_carrier': q2c, 'q2_market': q2m, 'fs2': fs2,
                'avg_share': avg, 'rt_check': rt, 'adj_share': adj,
            }
        return result

    def extract_carrier_captures(self, bidir: Dict,
                                  carrier_code: str, hub_code: str) -> Tuple[Dict, Dict]:
        """Extract per-city capture rates for a specific carrier/hub."""
        city_captures = defaultdict(float)
        city_routes = defaultdict(list)
        for rl, data in bidir.items():
            parts = rl.split('-')
            if len(parts) < 5:
                continue
            if parts[2] != hub_code:
                continue
            if carrier_code not in (parts[1], parts[3]):
                continue
            city = data['city']
            city_captures[city] += data['adj_share']
            city_routes[city].append({'route_label': rl, 'adj_share': data['adj_share']})
        return dict(city_captures), dict(city_routes)

    def run(self, schedule_provider: ScheduleProvider) -> Dict[str, float]:
        """
        Full QSI processing: load  score  aggregate  bidirectional  extract.
        
        Returns dict of city_code -> capture_rate for the configured carrier/hub.
        """
        self.log(f"\n{'='*60}")
        self.log(f"QSI ENGINE: {self.config.summary()}")
        self.log(f"{'='*60}")

        # Load itineraries from provider
        q1 = schedule_provider.get_itineraries('qsi1')
        q2 = schedule_provider.get_itineraries('qsi2')

        # Log provider stats
        q1_hubs = defaultdict(int)
        q1_origins = defaultdict(int)
        for it in q1:
            q1_hubs[it.cnx_airport] += 1
            q1_origins[it.dep_airport] += 1

        meta = schedule_provider.get_metadata()
        self.log(f"  Provider: {meta.get('provider_type', 'unknown')}")
        self.log(f"  QSI 1: {len(q1):,} itineraries, {len(q1_hubs)} hubs")
        self.log(f"  QSI 2: {len(q2):,} itineraries")
        self.log(f"  Origins: {dict(q1_origins)}")

        # Score
        q1s = self.score_itineraries(q1)
        q2s = self.score_itineraries(q2)

        # Aggregate
        sh1, cm1 = self.aggregate_shares(q1s)
        sh2, cm2 = self.aggregate_shares(q2s)

        # Bidirectional
        bidir = self.bidirectional_calc(sh1, sh2, cm1, cm2)
        n_incl = sum(1 for d in bidir.values() if d['rt_check'] == 'Include')
        self.log(f"  Route labels: {len(bidir)} ({n_incl} included)")

        # Extract carrier captures
        captures, routes = self.extract_carrier_captures(
            bidir, self.config.airline_code, self.config.home_airport_code)
        n_routes = sum(len(v) for v in routes.values())
        self.log(f"  {self.config.airline_code}-{self.config.home_airport_code} captures: "
                 f"{n_routes} routes, {len(captures)} cities")

        # Store for audit
        self.bidir = bidir
        self.city_routes = routes
        self.city_captures = captures

        return captures


# ============================================================================
# FORECAST ASSEMBLER  Combines P2P + connecting
# ============================================================================

class ForecastAssembler:
    """
    Assembles the complete route forecast from provider data + QSI captures.
    
    Consumes:
        - DemandProvider for P2P segments and connecting city demand
        - QSI captures (from QSIEngine or pre-computed)
        - RouteConfig for capacity and QSI parameters
    """

    def __init__(self, config: RouteConfig):
        self.config = config
        self.audit = []
        self.results = {}

    def log(self, msg):
        self.audit.append(msg)
        print(msg)

    def compute_p2p(self, demand_provider: DemandProvider) -> Tuple[float, List[Dict]]:
        """Compute P2P forecast from demand provider segments."""
        self.log(f"\n[P2P FORECAST]")
        segments = demand_provider.get_p2p_segments()
        total = 0.0
        details = []

        for seg in segments:
            if seg.subsegments:
                seg_total = 0.0
                for sub in seg.subsegments:
                    yrs = getattr(sub, 'growth_years', 1) or 1
                    grown = sub.base_demand * ((1 + sub.growth_rate) ** yrs)
                    stimulated = grown * sub.stimulation
                    forecast = stimulated * sub.capture_rate
                    seg_total += forecast
                    details.append({
                        'name': f"{seg.name}/{sub.name}",
                        'base': sub.base_demand, 'growth': sub.growth_rate,
                        'growth_years': yrs,
                        'stimulation': sub.stimulation, 'capture': sub.capture_rate,
                        'forecast': forecast,
                    })
                    self.log(f"  {seg.name}/{sub.name}: {sub.base_demand:,.0f} "
                             f"×(1+{sub.growth_rate:.0%})^{yrs}={grown:,.0f} "
                             f"×{sub.stimulation:.2f}×{sub.capture_rate:.0%} "
                             f"= {forecast:,.0f}")
                total += seg_total
            else:
                yrs = getattr(seg, 'growth_years', 1) or 1
                grown = seg.base_demand * ((1 + seg.growth_rate) ** yrs)
                stimulated = grown * seg.stimulation
                forecast = stimulated * seg.capture_rate
                total += forecast
                details.append({
                    'name': seg.name,
                    'base': seg.base_demand, 'growth': seg.growth_rate,
                    'growth_years': yrs,
                    'stimulation': seg.stimulation, 'capture': seg.capture_rate,
                    'forecast': forecast,
                })
                self.log(f"  {seg.name}: {seg.base_demand:,.0f} "
                         f"×(1+{seg.growth_rate:.0%})^{yrs}={grown:,.0f} "
                         f"×{seg.stimulation:.2f}×{seg.capture_rate:.0%} "
                         f"= {forecast:,.0f}")

        self.log(f"  TOTAL P2P: {total:,.0f}")
        return total, details

    def assemble_connecting(self, demand_provider: DemandProvider,
                             qsi_captures: Dict[str, float],
                             direction: str) -> Tuple[float, List[Dict]]:
        """Assemble connecting forecast for one direction."""
        label = f"Connecting @ {self.config.home_airport_code if direction == 'home' else self.config.dest_airport_code}"
        self.log(f"\n[{label}]")

        cities = demand_provider.get_connecting_cities(direction)
        results = []
        total = 0.0
        matched = unmatched = 0
        capped = 0

        # Apply QSI ceiling and adjustment factor
        ceiling = getattr(self.config, 'qsi_ceiling', 1.0)
        adjustment = getattr(self.config, 'qsi_adjustment', 1.0)

        # Direct service penalty: when a connecting city already has direct
        # service to the route endpoint, the proposed carrier's connecting
        # itinerary via hub captures far less traffic. The analyst methodology
        # typically assigns 1-2% capture for O&Ds with direct competition
        # vs 9-15% for O&Ds without. We apply a penalty factor to QSI captures
        # for cities with existing direct service.
        direct_penalty = getattr(self.config, 'direct_service_penalty', 0.15)

        for city in cities:
            raw_capture = qsi_captures.get(city.city_code, 0)
            # Apply ceiling first, then adjustment
            capped_capture = min(raw_capture, ceiling)
            capture = capped_capture * adjustment

            # Penalise cities with existing direct service
            if city.direct_service and direct_penalty < 1.0:
                capture *= direct_penalty

            if raw_capture > ceiling:
                capped += 1

            grown = city.base_demand * (1 + city.growth_rate)
            forecast = grown * capture
            results.append({
                'city': city.city_code, 'name': city.city_name,
                'country': city.country,
                'base_demand': city.base_demand, 'growth': city.growth_rate,
                'qsi_capture': capture,
                'raw_qsi_capture': raw_capture,
                'original_qsi': city.qsi_score,
                'forecast': forecast, 'direct': city.direct_service,
            })
            total += forecast
            if capture > 0:
                matched += 1
            else:
                unmatched += 1

        results.sort(key=lambda x: -x['forecast'])
        cap_note = f", {capped} capped at {ceiling:.0%}" if capped > 0 else ""
        adj_note = f", adj={adjustment:.3f}" if adjustment != 1.0 else ""
        direct_count = sum(1 for r in results if r.get('direct'))
        direct_pax = sum(r['forecast'] for r in results if r.get('direct'))
        indirect_pax = total - direct_pax
        direct_note = f" (direct comp: {direct_count} cities, {direct_pax:,.0f} pax; no direct: {indirect_pax:,.0f} pax)" if direct_count > 0 else ""
        self.log(f"  {matched} matched, {unmatched} unmatched, {total:,.0f} pax"
                 f"{cap_note}{adj_note}{direct_note}")
        return total, results

    def run(self, demand_provider: DemandProvider,
            lhr_captures: Dict[str, float],
            sjc_captures: Dict[str, float]) -> Dict[str, Any]:
        """Execute full forecast assembly."""
        self.log(f"\n{'#'*60}")
        self.log(f"# FORECAST ASSEMBLY  {self.config.summary()}")
        self.log(f"# {datetime.now():%Y-%m-%d %H:%M}")
        self.log(f"{'#'*60}")

        # P2P
        p2p_total, p2p_details = self.compute_p2p(demand_provider)

        # Connecting
        home_total, home_results = self.assemble_connecting(
            demand_provider, lhr_captures, 'home')
        dest_total, dest_results = self.assemble_connecting(
            demand_provider, sjc_captures, 'dest')

        # Grand total
        grand_total = p2p_total + home_total + dest_total
        capacity = self.config.annual_capacity
        load_factor = grand_total / capacity if capacity > 0 else 0

        self.log(f"\n{'='*60}")
        self.log(f"FORECAST RESULTS")
        self.log(f"{'='*60}")
        self.log(f"  P2P Total:         {p2p_total:>10,.0f}")
        self.log(f"  Connecting @ Home: {home_total:>10,.0f}")
        self.log(f"  Connecting @ Dest: {dest_total:>10,.0f}")
        self.log(f"  {''*29}")
        self.log(f"  GRAND TOTAL:       {grand_total:>10,.0f}")
        self.log(f"  Load Factor:       {load_factor:>10.1%}")

        self.results = {
            'p2p_total': p2p_total, 'p2p_details': p2p_details,
            'home_total': home_total, 'home_results': home_results,
            'dest_total': dest_total, 'dest_results': dest_results,
            'grand_total': grand_total, 'load_factor': load_factor,
        }
        return self.results


# ============================================================================
# VALIDATION
# ============================================================================

def validate(config: RouteConfig, results: Dict[str, Any]) -> bool:
    """Validate results against RouteConfig targets."""
    checks = [
        ("P2P", results['p2p_total'], config.target_p2p),
        ("Cnx @ Home", results['home_total'], config.target_cnx_home),
        ("Cnx @ Dest", results['dest_total'], config.target_cnx_dest),
        ("GRAND TOTAL", results['grand_total'], config.target_total),
    ]

    print(f"\n{'='*60}")
    print(f"VALIDATION vs TARGET")
    print(f"{'='*60}")
    all_pass = True
    for label, actual, target in checks:
        if target > 0:
            pct = abs(actual - target) / target
            ok = pct < 0.05  # 5% tolerance for closed-loop
            sym = "PASS" if ok else "MISS"
            all_pass &= ok
            print(f"  {sym} {label:20s}: {actual:>10,.0f}  target: {target:>10,}  var: {pct:>7.2%}")

    if config.target_total > 0:
        variance = abs(results['grand_total'] - config.target_total) / config.target_total
        print(f"\n  Overall variance: {variance:.2%}")
    else:
        variance = 0.0
        print(f"\n  No target set (new route assessment)")
    tgt_lf = getattr(config, 'target_load_factor', 0) or 0
    print(f"  Load factor: {results['load_factor']:.1%}" + (f" (target: {tgt_lf:.1%})" if tgt_lf > 0 else ""))

    return all_pass


# ============================================================================
# CALIBRATION ANALYSIS
# ============================================================================

def calibration_analysis(home_results: List[Dict], label: str = "LHR") -> Dict:
    """Analyse calibration factors between pipeline QSI and forecast QSI."""
    cal_factors = []
    for r in home_results:
        pipe = r.get('qsi_capture', 0)
        fcst = r.get('original_qsi', 0)
        if pipe > 0 and fcst > 0:
            cal_factors.append(fcst / pipe)

    if not cal_factors:
        return {}

    cal_factors.sort()
    n = len(cal_factors)
    avg = sum(cal_factors) / n
    med = cal_factors[n // 2]

    print(f"\n  Calibration factor analysis ({n} cities with both QSI):")
    print(f"    Mean:   {avg:.3f}")
    print(f"    Median: {med:.3f}")
    print(f"    Range:  {min(cal_factors):.3f}  {max(cal_factors):.3f}")

    return {'mean': avg, 'median': med, 'min': min(cal_factors),
            'max': max(cal_factors), 'count': n}


# ============================================================================
# OUTPUT WRITER
# ============================================================================

def write_output(config: RouteConfig, results: Dict, qsi_engine: QSIEngine,
                 outpath: str) -> str:
    """Generate comprehensive output workbook."""
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hfill = PatternFill('solid', fgColor='002060')
    df = Font(name='Arial', size=10)
    bf = Font(name='Arial', size=10, bold=True)
    tf = Font(name='Arial', size=14, bold=True, color='002060')

    def hdr(ws, cols, row=1):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row, c, v)
            cell.font = hf; cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')

    #  Summary 
    ws = wb.active; ws.title = 'Summary'
    ws.cell(1, 1, 'Avia Solutions  Closed-Loop Pipeline V2').font = tf
    ws.cell(2, 1, config.summary()).font = Font(name='Arial', size=11, bold=True, color='002060')
    ws.cell(3, 1, f'Generated: {datetime.now():%Y-%m-%d %H:%M}').font = df

    r = 5
    for label, val in [
        ('P2P Total', results['p2p_total']),
        ('Connecting @ Home', results['home_total']),
        ('Connecting @ Dest', results['dest_total']),
        ('GRAND TOTAL', results['grand_total']),
        ('Load Factor', results['load_factor']),
    ]:
        ws.cell(r, 1, label).font = bf if 'GRAND' in label else df
        c = ws.cell(r, 2)
        c.font = bf if 'GRAND' in label else df
        if isinstance(val, float) and val < 1:
            c.value = val; c.number_format = '0.0%'
        else:
            c.value = val; c.number_format = '#,##0'
        r += 1

    r += 1
    ws.cell(r, 1, f'Target: {config.target_total:,}').font = df
    if config.target_total > 0:
        var = abs(results['grand_total'] - config.target_total) / config.target_total
    else:
        var = 0.0
    ws.cell(r+1, 1, f'Variance: {var:.2%}').font = df
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    #  Connecting sheets 
    for sheet_label, data_key in [('Connecting Home', 'home_results'),
                                    ('Connecting Dest', 'dest_results')]:
        ws2 = wb.create_sheet(sheet_label)
        hdr(ws2, ['City', 'Name', 'Base Demand', 'Growth', 'Pipeline QSI',
                   'Forecast QSI', 'Pipeline Pax', 'Forecast Pax', 'Difference'])
        for i, rd in enumerate(results[data_key], 2):
            orig_pax = rd['base_demand'] * (1 + rd['growth']) * rd['original_qsi'] if rd['original_qsi'] else 0
            ws2.cell(i, 1, rd['city']).font = df
            ws2.cell(i, 2, rd['name']).font = df
            ws2.cell(i, 3, rd['base_demand']).font = df
            ws2.cell(i, 3).number_format = '#,##0'
            ws2.cell(i, 4, rd['growth']).font = df
            ws2.cell(i, 4).number_format = '0.0%'
            ws2.cell(i, 5, rd['qsi_capture']).font = df
            ws2.cell(i, 5).number_format = '0.000000'
            ws2.cell(i, 6, rd['original_qsi']).font = df
            ws2.cell(i, 6).number_format = '0.000000'
            ws2.cell(i, 7, rd['forecast']).font = df
            ws2.cell(i, 7).number_format = '#,##0'
            ws2.cell(i, 8, orig_pax).font = df
            ws2.cell(i, 8).number_format = '#,##0'
            ws2.cell(i, 9, rd['forecast'] - orig_pax).font = df
            ws2.cell(i, 9).number_format = '#,##0'
        for c in range(1, 10):
            ws2.column_dimensions[get_column_letter(c)].width = 14

    #  Audit Trail 
    ws_audit = wb.create_sheet('Audit Trail')
    ws_audit.cell(1, 1, 'Pipeline Audit Trail').font = tf
    all_audit = qsi_engine.audit + ['', '--- FORECAST ASSEMBLY ---', ''] + \
                (results.get('assembler_audit', []))
    for i, line in enumerate(all_audit, 3):
        ws_audit.cell(i, 1, str(line)).font = Font(name='Consolas', size=8)
    ws_audit.column_dimensions['A'].width = 80

    wb.save(outpath)
    return outpath


# ============================================================================
# MAIN PIPELINE  Orchestrates everything via providers
# ============================================================================

def run_pipeline(config: RouteConfig, output_path: Optional[str] = None) -> Dict:
    """
    Run the complete closed-loop pipeline for a given RouteConfig.
    
    1. QSI Engine scores itineraries from ScheduleProvider
    2. ForecastAssembler combines P2P + connecting from DemandProvider
    3. Validation checks against targets
    4. Output workbook generated
    """
    if not config.schedule_provider:
        raise ValueError("RouteConfig has no schedule_provider set")
    if not config.demand_provider:
        raise ValueError("RouteConfig has no demand_provider set")

    # Log current QSI parameters (so closed-loop changes are visible)
    adj = getattr(config, 'qsi_adjustment', 1.0)
    ceil = getattr(config, 'qsi_ceiling', 1.0)
    if adj != 1.0 or ceil != 1.0:
        print(f"  [Pipeline] QSI params: adjustment={adj:.3f}, ceiling={ceil:.2f}")

    # Step 1: QSI scoring for home hub (QSILHR has both QSI 1 and QSI 2 for LHR perspective)
    qsi_home = QSIEngine(config)
    lhr_captures = qsi_home.run(config.schedule_provider)

    # Step 2: QSI scoring for destination (QSISJC has both QSI 1 and QSI 2 for SJC perspective)
    # Create a dest config where SJC is treated as the hub for carrier extraction
    dest_config = RouteConfig()
    dest_config.airline_code = config.airline_code
    dest_config.home_airport_code = config.dest_airport_code  # SJC is "home" for dest QSI
    dest_config.dest_airport_code = config.home_airport_code
    dest_config.aircraft_type = config.aircraft_type
    dest_config.seats = config.seats
    dest_config.frequency = config.frequency
    dest_config.online_coeff = config.online_coeff
    dest_config.alliance_coeff = config.alliance_coeff
    dest_config.interline_coeff = config.interline_coeff
    dest_config.et_decay_factor = config.et_decay_factor
    dest_config.et_decay_interval = config.et_decay_interval

    # The dest ScheduleProvider reads QSI 1 + QSI 2 from the DEST file (e.g., QSISJC)
    # Both directions are in the same file  NOT swapped
    from providers import ExcelScheduleProvider
    if hasattr(config.schedule_provider, 'qsi2_file'):
        dest_schedule = ExcelScheduleProvider(
            qsi1_file=config.schedule_provider.qsi2_file,  # QSISJC
            qsi2_file=config.schedule_provider.qsi2_file,  # QSISJC (same file, both sheets)
        )
    else:
        dest_schedule = config.schedule_provider

    qsi_dest = QSIEngine(dest_config)
    sjc_captures = qsi_dest.run(dest_schedule)

    # Step 3: Forecast assembly
    assembler = ForecastAssembler(config)
    results = assembler.run(config.demand_provider, lhr_captures, sjc_captures)
    results['assembler_audit'] = assembler.audit

    # Step 4: Calibration analysis
    cal = calibration_analysis(results['home_results'])
    results['calibration'] = cal

    # Step 5: Validation
    passed = validate(config, results)
    results['validation_passed'] = passed

    # Step 6: Output
    if output_path:
        write_output(config, results, qsi_home, output_path)
        print(f"\nOutput: {output_path}")

    return results


# ============================================================================
# CLI
# ============================================================================

def main():
    """Run BA LHR-SJC regression test."""
    print("=" * 60)
    print("CLOSED-LOOP PIPELINE V2  Provider Architecture")
    print("=" * 60)

    config = RouteConfig.ba_lhr_sjc()
    ensure_output_dir()
    outpath = str(OUTPUT_DIR / 'Closed_Loop_V2_BA_LHR_SJC.xlsx')
    results = run_pipeline(config, outpath)

    return results


if __name__ == '__main__':
    main()
