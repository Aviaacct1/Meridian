#!/usr/bin/env python3
"""Build the consolidated Genoa-New York business-case workbook (Avia Solutions)."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CASE = json.load(open(sys.argv[1])) if len(sys.argv) > 1 else json.load(open("genoa_nyc_case.json"))
OUT = sys.argv[2] if len(sys.argv) > 2 else "Genoa_NYC_business_case.xlsx"

NAVY = "002060"; LIGHT = "EAEEF6"; BLUEF = "0000FF"
hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
title = Font(name="Arial", bold=True, color=NAVY, size=15)
h2 = Font(name="Arial", bold=True, color=NAVY, size=12)
lab = Font(name="Arial", size=10)
labb = Font(name="Arial", bold=True, size=10)
inp = Font(name="Arial", color=BLUEF, size=10)          # blue = user input
muted = Font(name="Arial", italic=True, color="777777", size=9)
navy_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor=LIGHT)
USD = '$#,##0;($#,##0);"-"'; USD2 = '$#,##0.00'; PCT = '0.0%'; NUM = '#,##0'; MULT = '0.00'

wb = Workbook()
wb.properties.creator = "Avia Solutions"
wb.properties.lastModifiedBy = "Avia Solutions"

def band(ws, row, text, span=4):
    ws.cell(row, 1, text).font = hf
    for c in range(1, span + 1):
        ws.cell(row, c).fill = navy_fill
    return row + 1

def kv(ws, row, label, value, fmt=None, font=lab, note=None):
    ws.cell(row, 1, label).font = font
    c = ws.cell(row, 2, value)
    if fmt: c.number_format = fmt
    if note: ws.cell(row, 4, note).font = muted
    return row + 1

# ----------------------------------------------------------------- Inputs
wi = wb.active; wi.title = "Inputs"
wi.cell(1, 1, "Genoa - New York  |  inputs (blue = change me)").font = title
r = band(wi, 3, "Commercial inputs")
I = {}
inputs = [("Capture of leaked catchment", CASE["capture"], PCT, "capture"),
          ("Frequency (each way / week)", CASE.get("frequency", 7), "0", "freq"),
          ("Planning load-factor cap", 0.85, PCT, "plf"),
          ("Economy fare, one-way ($)", 345, USD, "ef"),
          ("Business fare, one-way ($)", 1400, USD, "bf"),
          ("Jet fuel price ($/kg)", 0.90, USD2, "fp"),
          ("Economy share of demand", 0.80, PCT, "es"),
          ("Market adjustment (1.0 = neutral)", 1.00, MULT, "mkt")]
r0 = r
for label, val, fmt, key in inputs:
    wi.cell(r, 1, label).font = lab
    c = wi.cell(r, 2, val); c.font = inp; c.number_format = fmt; c.fill = light_fill
    I[key] = f"Inputs!$B${r}"
    r += 1
r += 1
r = band(wi, r, "Catchment & demand (from the QSI chain)")
fixed = [("Catchment population", CASE["population"], NUM, "pop"),
         ("New York O&D market (Sabre 2024)", CASE["dest_od_total"], NUM, "od"),
         ("Genoa natural NY catchment", CASE["natural"], NUM, "nat"),
         ("Carried by Genoa today", CASE["current"], NUM, "cur")]
for label, val, fmt, key in fixed:
    wi.cell(r, 1, label).font = lab
    c = wi.cell(r, 2, val); c.font = lab; c.number_format = fmt
    I[key] = f"Inputs!$B${r}"; r += 1
r += 1
r = band(wi, r, "Aircraft & cost constants (A321XLR, validated module)")
rp = CASE["route_pnl"]
consts = [("Economy seats", 162, "0", "eseat"), ("Business seats", 20, "0", "bseat"),
          ("Fuel burn (kg / block-hour)", 2500, NUM, "burn"), ("Block hours / turnaround", 18, "0", "bh"),
          ("Annual utilisation (block-hr)", rp["eff_util"], NUM, "util"),
          ("Blended lease ($/aircraft/month)", round(rp["ownership"]/rp["block_hours_turn"]*rp["eff_util"]/12), USD, "lease"),
          ("Maintenance ($/turn)", rp["maintenance"], USD, "maint"),
          ("Ownership ($/turn)", rp["ownership"], USD, "own"),
          ("Crew ($/turn)", rp["crew"], USD, "crew"),
          ("Insurance ($/turn)", rp["insurance"], USD, "ins"),
          ("Landing, both ends ($/turn)", rp["landing"], USD, "land"),
          ("Ground handling ($/turn)", rp["handling"], USD, "hand"),
          ("En-route navigation ($/turn)", rp["nav"], USD, "nav"),
          ("Cargo revenue ($/turn)", rp["cargo_rev"], USD, "cargo"),
          ("Catering ($/pax)", round(rp["catering"]/rp["pax_turn"], 3), USD2, "catr"),
          ("Pax airport charge ($/pax)", round(rp["per_pax"]/rp["pax_turn"], 3), USD2, "ppr"),
          ("Charges recovery ($/pax)", round(rp["charges_recovery"]/rp["pax_turn"], 3), USD2, "recr")]
for label, val, fmt, key in consts:
    wi.cell(r, 1, label).font = lab
    c = wi.cell(r, 2, val); c.font = lab; c.number_format = fmt
    I[key] = f"Inputs!$B${r}"; r += 1
wi.column_dimensions["A"].width = 36; wi.column_dimensions["B"].width = 16; wi.column_dimensions["D"].width = 30

# ----------------------------------------------------------------- Route P&L
wp = wb.create_sheet("Route P&L")
wp.cell(1, 1, "Route P&L  |  per turnaround and annual (formulas off Inputs)").font = title
r = band(wp, 3, "Demand & load", 3)
def f(ws, row, label, formula, fmt=USD, font=lab):
    ws.cell(row, 1, label).font = font
    c = ws.cell(row, 2, formula); c.number_format = fmt
    return row + 1
P = {}
def put(ws, row, key, label, formula, fmt=USD, font=lab):
    ws.cell(row, 1, label).font = font
    c = ws.cell(row, 2, formula); c.number_format = fmt
    P[key] = f"'Route P&L'!$B${row}"
    return row + 1
r = put(wp, r, "dir", "Directional demand (each way)", f"=({I['cur']}+({I['nat']}-{I['cur']})*{I['capture']})*{I['mkt']}", NUM)
r = put(wp, r, "elf", "Economy load factor", f"=MIN({P['dir']}*{I['es']}/({I['eseat']}*{I['freq']}*52),{I['plf']})", PCT)
r = put(wp, r, "blf", "Business load factor", f"=MIN({P['dir']}*(1-{I['es']})/({I['bseat']}*{I['freq']}*52),{I['plf']})", PCT)
r = put(wp, r, "pax", "Passengers per turnaround", f"=2*({I['eseat']}*{P['elf']}+{I['bseat']}*{P['blf']})", NUM)
r += 1
r = band(wp, r, "Revenue ($/turn)", 3)
r = put(wp, r, "erev", "Economy", f"=2*{I['eseat']}*{P['elf']}*{I['ef']}")
r = put(wp, r, "brev", "Business", f"=2*{I['bseat']}*{P['blf']}*{I['bf']}")
r = put(wp, r, "crev", "Cargo", f"={I['cargo']}")
r = put(wp, r, "net", "Net revenue", f"={P['erev']}+{P['brev']}+{P['crev']}", USD, labb)
r = put(wp, r, "rec", "Charges recovery", f"={I['recr']}*{P['pax']}")
r = put(wp, r, "gross", "Gross revenue", f"={P['net']}+{P['rec']}", USD, labb)
r += 1
r = band(wp, r, "Cost ($/turn)", 3)
r = put(wp, r, "fuel", "Fuel", f"={I['burn']}*{I['bh']}*{I['fp']}")
r = put(wp, r, "maint", "Maintenance", f"={I['maint']}")
r = put(wp, r, "cat", "Catering", f"={I['catr']}*{P['pax']}")
r = put(wp, r, "land", "Landing", f"={I['land']}")
r = put(wp, r, "pp", "Pax airport charges", f"={I['ppr']}*{P['pax']}")
r = put(wp, r, "nav", "En-route navigation", f"={I['nav']}")
r = put(wp, r, "hand", "Ground handling", f"={I['hand']}")
r = put(wp, r, "var", "Variable cost", f"={P['fuel']}+{P['maint']}+{P['cat']}+{P['land']}+{P['pp']}+{P['nav']}+{P['hand']}", USD, labb)
r = put(wp, r, "own", "Ownership", f"={I['own']}")
r = put(wp, r, "ins", "Insurance", f"={I['ins']}")
r = put(wp, r, "crew", "Crew", f"={I['crew']}")
r = put(wp, r, "dfix", "Direct fixed", f"={P['own']}+{P['ins']}+{P['crew']}", USD, labb)
r = put(wp, r, "admin", "Admin (5% net)", f"=0.05*{P['net']}")
r = put(wp, r, "sales", "Sales (5% net)", f"=0.05*{P['net']}")
r = put(wp, r, "ind", "Indirect fixed", f"={P['admin']}+{P['sales']}", USD, labb)
r = put(wp, r, "tot", "Total cost", f"={P['var']}+{P['dfix']}+{P['ind']}", USD, labb)
r += 1
r = band(wp, r, "Result", 3)
r = put(wp, r, "prof", "Profit per turnaround", f"={P['gross']}-{P['tot']}", USD, labb)
r = put(wp, r, "marg", "Margin", f"={P['prof']}/{P['gross']}", PCT, labb)
pax_var = f"({P['cat']}+{P['pp']}+{P['ind']})"
r = put(wp, r, "be", "Breakeven load factor", f"=(({I['eseat']}*{P['elf']}+{I['bseat']}*{P['blf']})/({I['eseat']}+{I['bseat']}))*(({P['tot']}-{pax_var})/({P['gross']}-{pax_var}))", PCT, labb)
r = put(wp, r, "ann", "Annual profit", f"={P['prof']}*{I['freq']}*52", USD, Font(name='Arial', bold=True, color=NAVY, size=12))
r = put(wp, r, "annpax", "Annual passengers", f"={P['pax']}*{I['freq']}*52", NUM)
wp.column_dimensions["A"].width = 30; wp.column_dimensions["B"].width = 16

# ----------------------------------------------------------------- Fleet
wf = wb.create_sheet("Fleet")
wf.cell(1, 1, "Fleet utilisation & redeployment").font = title
r = band(wf, 3, "The headline charges ownership at a fully-utilised rate", 3)
F = {}
def putf(row, key, label, formula, fmt=USD, font=lab):
    wf.cell(row, 1, label).font = font
    c = wf.cell(row, 2, formula); c.number_format = fmt
    F[key] = f"Fleet!$B${row}"; return row + 1
r = putf(r, "tbh", "Annual block hours", f"={I['freq']}*52*{I['bh']}", NUM)
r = putf(r, "ac", "Aircraft required", f"=CEILING({F['tbh']}/{I['util']},1)", "0", labb)
r = putf(r, "perac", "Per-aircraft utilisation", f"={F['tbh']}/{F['ac']}", NUM)
r = putf(r, "utilpc", "Utilisation vs full", f"={F['perac']}/{I['util']}", PCT)
r = putf(r, "spare", "Spare block hours / year", f"={F['ac']}*{I['util']}-{F['tbh']}", NUM)
r = putf(r, "sparerot", "Spare rotations / year", f"={F['spare']}/{I['bh']}", NUM)
r += 1
r = band(wf, r, "Ownership recovery", 3)
r = putf(r, "modown", "Ownership charged (annual)", f"=({I['own']}/{I['bh']})*{F['tbh']}", USD)
r = putf(r, "dedown", "True cost if dedicated", f"={F['ac']}*{I['lease']}*12", USD)
r = putf(r, "gap", "Standalone ownership gap", f"={F['dedown']}-{F['modown']}", USD, labb)
r += 1
r = band(wf, r, "What the headline assumes", 3)
r = putf(r, "net", "Network headline (fleet busy)", f"='Route P&L'!{P['ann'].split('!')[1]}", USD, labb)
r = putf(r, "stand", "Standalone (Genoa-only fleet)", f"={F['net']}-{F['gap']}", USD, labb)
wf.column_dimensions["A"].width = 32; wf.column_dimensions["B"].width = 16

# ----------------------------------------------------------------- Seasonality
ws = wb.create_sheet("Seasonality")
ws.cell(1, 1, "Seasonality  |  leisure profile (assumption pending monthly Sabre pull)").font = title
r = 3
ws.cell(r, 1, "Month").font = labb; ws.cell(r, 2, "Days").font = labb
ws.cell(r, 3, "Demand index").font = labb; ws.cell(r, 4, "Monthly demand").font = labb
ws.cell(r, 5, "Flat-daily LF").font = labb; ws.cell(r, 6, "Seasonal turns").font = labb
for c in range(1, 7): ws.cell(r, c).fill = light_fill
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
days = [31,28,31,30,31,30,31,31,30,31,30,31]
base = [0.527,0.448,0.701,1.173,1.128,1.264,1.381,1.637,1.260,1.047,0.689,0.747]
first = r + 1
for m in range(12):
    rr = first + m
    ws.cell(rr, 1, months[m]).font = lab
    ws.cell(rr, 2, days[m]).font = lab
    ci = ws.cell(rr, 3, base[m]); ci.font = inp; ci.number_format = MULT; ci.fill = light_fill
    seats_turn = f"({I['eseat']}+{I['bseat']})"
    dm = ws.cell(rr, 4, f"={P['dir'].replace(chr(39),chr(39))}*(B{rr}/365)*C{rr}"); dm.number_format = NUM
    lf = ws.cell(rr, 5, f"=MIN(D{rr}/({seats_turn}*B{rr}),{I['plf']})"); lf.number_format = PCT
    tc = ws.cell(rr, 6, f"=MAX(MIN(B{rr},CEILING(D{rr}/({seats_turn}*{I['plf']}),1)),0.5*B{rr})"); tc.number_format = "0"
last = first + 11

def seasonal_abc():
    NAT=CASE["natural"]; CUR=CASE["current"]; directional=CUR+(NAT-CUR)*CASE["capture"]
    es=162; bs=20; seats=182; plf=0.85
    prof=[1+1.0*(b-1) for b in base]; sp=sum(prof); prof=[p*12/sp for p in prof]
    def tp(lf):
        net=2*es*lf*345+2*bs*lf*1400+rp["cargo_rev"]; pax=2*(es+bs)*lf
        rec=(rp["charges_recovery"]/rp["pax_turn"])*pax; gross=net+rec
        fuel=2500*18*0.90; cat=(rp["catering"]/rp["pax_turn"])*pax; pp=(rp["per_pax"]/rp["pax_turn"])*pax
        var=fuel+rp["maintenance"]+cat+rp["landing"]+pp+rp["nav"]+rp["handling"]
        return gross-(var+rp["ownership"]+rp["insurance"]+rp["crew"]+0.10*net)
    import math
    A=B=C=0; totd=sum(days)
    for m in range(12):
        dem=directional*(days[m]/totd)*prof[m]
        A+=days[m]*tp(plf)
        B+=days[m]*tp(min(dem/(seats*days[m]),plf))
        tc=max(min(days[m],math.ceil(dem/(seats*plf))),0.5*days[m])
        C+=tc*tp(min(dem/(seats*tc),plf))
    return A,B,C
A,B,C = seasonal_abc()
r = last + 2
ws.cell(r, 1, "Annual profit at the central case (leisure profile, daily summer):").font = labb; r += 1
for label, val in [("A  nominal flat (plan LF every month)", A),
                   ("B  flat daily, real monthly demand", B),
                   ("C  seasonal schedule (trim winter)", C)]:
    ws.cell(r, 1, label).font = lab
    ws.cell(r, 2, round(val)).number_format = USD; r += 1
ws.cell(r+1, 1, "Snapshot from seasonality_check.py at the central inputs. A daily-all-year schedule under-fills winter at full cost (B); a seasonal schedule recovers most of it (C). Adjust the demand-index column above and re-run seasonality_check.py for other profiles.").font = muted
ws.column_dimensions["A"].width = 42
for col in "BCDEF": ws.column_dimensions[col].width = 15

# ----------------------------------------------------------------- Scenarios (snapshot)
wsc = wb.create_sheet("Scenarios")
wsc.cell(1, 1, "Scenarios  |  annual profit ($m), computed from the model").font = title
def model(capture, freq, plf, ef, bf, fp, es, mkt):
    NAT=CASE["natural"]; CUR=CASE["current"]
    directional=(CUR+(NAT-CUR)*capture)*mkt
    esy=162*freq*52; bsy=20*freq*52
    elf=min(directional*es/esy,plf); blf=min(directional*(1-es)/bsy,plf)
    pax=2*(162*elf+20*blf)
    net=2*162*elf*ef+2*20*blf*bf+rp["cargo_rev"]
    rec=(rp["charges_recovery"]/rp["pax_turn"])*pax; gross=net+rec
    fuel=2500*18*fp; cat=(rp["catering"]/rp["pax_turn"])*pax; pp=(rp["per_pax"]/rp["pax_turn"])*pax
    var=fuel+rp["maintenance"]+cat+rp["landing"]+pp+rp["nav"]+rp["handling"]
    direct=rp["ownership"]+rp["insurance"]+rp["crew"]; ind=0.10*net
    profit=gross-(var+direct+ind)
    return profit*freq*52
r = 3
wsc.cell(r,1,"Fuel price down the side, capture across (freq 7, plan LF 85%, fares 345/1400, market neutral)").font = lab
r += 1
caps=[0.50,0.58,0.65,0.75]; fuels=[0.68,0.80,0.90,1.00,1.10]
wsc.cell(r,1,"fuel \\ capture").font=labb
for j,cp in enumerate(caps):
    c=wsc.cell(r,2+j,cp); c.font=labb; c.number_format=PCT; c.fill=light_fill
for i,fl in enumerate(fuels):
    rr=r+1+i
    cf=wsc.cell(rr,1,fl); cf.font=labb; cf.number_format=USD2; cf.fill=light_fill
    for j,cp in enumerate(caps):
        v=model(cp,7,0.85,345,1400,fl,0.80,1.0)/1e6
        cc=wsc.cell(rr,2+j,round(v,1)); cc.number_format='0.0'
        if abs(fl-0.90)<1e-9 and abs(cp-0.65)<1e-9:
            cc.fill=PatternFill("solid",fgColor="FFF2CC")
r = r+1+len(fuels)+2
wsc.cell(r,1,"Market adjustment sensitivity (central case otherwise)").font=labb; r+=1
wsc.cell(r,1,"market").font=labb; wsc.cell(r,2,"annual profit ($m)").font=labb
for c in (1,2): wsc.cell(r,c).fill=light_fill
for k,mk in enumerate([1.0,0.93,0.86,0.80]):
    rr=r+1+k
    wsc.cell(rr,1,mk).number_format=MULT
    wsc.cell(rr,2,round(model(0.65,7,0.85,345,1400,0.90,0.80,mk)/1e6,1)).number_format='0.0'
wsc.cell(r+6,1,"Highlighted cell = central planning case (fuel $0.90, capture 65%). Values are a model snapshot; change the Inputs sheet and recalculate to explore other points.").font=muted
wsc.column_dimensions["A"].width = 22
for col in "BCDE": wsc.column_dimensions[col].width = 13

# ----------------------------------------------------------------- Summary (first)
sm = wb.create_sheet("Summary", 0)
sm.cell(1, 1, "Genoa - New York  |  A321XLR business case").font = Font(name="Arial", bold=True, color=NAVY, size=18)
sm.cell(2, 1, "Avia Solutions  |  catchment - demand - economics - seasonality - fleet").font = h2
r = band(sm, 4, "Headline (central planning case)", 3)
r = kv(sm, r, "Annual profit", f"='Route P&L'!{P['ann'].split('!')[1]}", USD, labb)
r = kv(sm, r, "Route margin", f"='Route P&L'!{P['marg'].split('!')[1]}", PCT, labb)
r = kv(sm, r, "Breakeven load factor", f"='Route P&L'!{P['be'].split('!')[1]}", PCT)
r = kv(sm, r, "Directional demand (each way)", f"='Route P&L'!{P['dir'].split('!')[1]}", NUM)
r += 1
r = band(sm, r, "How it is operated changes the number", 3)
r = kv(sm, r, "Standalone (Genoa-only fleet)", f"=Fleet!{F['stand'].split('!')[1]}", USD)
r = kv(sm, r, "Network (fleet kept busy)", f"=Fleet!{F['net'].split('!')[1]}", USD)
sm.cell(r, 1, "The headline charges ownership at a fully-utilised rate, so it is a network number; a single daily Genoa route uses the aircraft 60% and the standalone figure is lower. A seasonal schedule plus winter redeployment is what earns the headline.").font = muted
r += 2
r = band(sm, r, "The case in one line", 3)
for line in ["Genoa's New York travellers leak to Milan today; a Genoa nonstop repatriates its own catchment.",
             "Fares one-way (2024 Sabre); business fare $1,400 the swing from the old $750; fuel at a $0.90/kg planning price.",
             "Viable across the band; honest planning is circa 15% margin, with fuel, seasonality and fleet the watchpoints."]:
    sm.cell(r, 1, line).font = lab; r += 1
r += 1
sm.cell(r, 1, "Indicative, for directional guidance only. Built on generic published assumptions, not any airline's actual LOPA, MTOW, contract terms or internal P&L. Catchment, demand and fares anchored to the validated Genoa-New York case; actual results will differ.").font = muted
sm.column_dimensions["A"].width = 40; sm.column_dimensions["B"].width = 16

# ----------------------------------------------------------------- Assumptions
wa = wb.create_sheet("Assumptions")
wa.cell(1, 1, "Assumptions & sources").font = title
rows = [("Catchment & demand", ""),
        ("GeoNames population, OSRM road times, catchment calibrated to the Sabre point-of-origin split (SSE 0.0009)", "validated"),
        ("New York O&D split and fares", "Sabre ODPOO 2024 (validated)"),
        ("Fares are one-way; economy $345 entrant yield, business $1,400", "John's entrant-yield judgement, 28 Jun 2026"),
        ("Economics", ""),
        ("Cost stack anchored to Project Maverick (E190 validated 0.3%); A321XLR extrapolated", "directional"),
        ("Jet fuel $0.90/kg through-cycle planning (spot ~$1.10 June 2026, IATA)", "planning assumption"),
        ("Maintenance from Airbus 2024-25 reserve curve", "validated"),
        ("Ownership from appraiser lease/value, A321neo ~$460k/mo (IBA 2025)", "directional"),
        ("Crew $1,200/block-hour (LCC)", "judgement, pending citation sweep"),
        ("GOA / JFK airport charges", "INDICATIVE placeholders, not validated tariffs"),
        ("Seasonality profile (Aug/Feb demand 2.5x)", "ASSUMPTION pending monthly Sabre pull"),
        ("Market", ""),
        ("2026 transatlantic bookings down ~14% Europe-US (Cirium)", "demand watchpoint")]
r = 3
for text, tag in rows:
    if tag == "":
        wa.cell(r, 1, text).font = h2
    else:
        wa.cell(r, 1, text).font = lab
        cell = wa.cell(r, 2, tag)
        cell.font = Font(name="Arial", size=9, color=("C00000" if ("INDICATIVE" in tag or "ASSUMPTION" in tag or "watchpoint" in tag) else "1F7A1F" if "validated" in tag else "777777"))
    r += 1
wa.column_dimensions["A"].width = 74; wa.column_dimensions["B"].width = 32

wb.save(OUT)
print("saved", OUT)
