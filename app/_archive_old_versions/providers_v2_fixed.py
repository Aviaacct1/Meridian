#!/usr/bin/env python3
"""
Avia Solutions  Provider Interfaces (Chat 12)
================================================
Abstract data providers that decouple the QSI engine from data sources.

Reconstructed for Chat 13 from the interfaces defined by:
    - closed_loop_pipeline_v2.py (consumer)
    - test_regression.py (contract tests)
    - route_config.py (factory methods)
"""

from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any

import openpyxl


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class Itinerary:
    """Single routing option from QSI sheet."""
    city: str
    airport: str
    route_label: str
    dep_airport: str
    cnx_airport: str
    carrier_l1: str
    carrier_l2: str
    freq: int
    elapsed: int  # minutes
    cnx_type: str  # ONLINE, ALLIANCE, INTERLINING
    # Scored fields (set by QSIEngine)
    et_coeff: float = 0.0
    cnx_coeff: float = 0.0
    qsi: float = 0.0
    is_proposed: bool = False


@dataclass
class P2PSubsegmentData:
    """Subsegment within a P2P segment."""
    name: str
    base_demand: float
    growth_rate: float
    seasonality: float = 1.0
    stimulation: float = 1.0
    capture_rate: float = 0.0
    growth_years: int = 1


@dataclass
class P2PSegmentData:
    """P2P demand segment (may contain subsegments)."""
    name: str
    base_demand: float
    growth_rate: float
    seasonality: float = 1.0
    stimulation: float = 1.0
    capture_rate: float = 0.0
    subsegments: List[P2PSubsegmentData] = field(default_factory=list)
    growth_years: int = 1


@dataclass
class ConnectingCityData:
    """Connecting city demand data."""
    city_code: str
    city_name: str
    country: str
    base_demand: float
    growth_rate: float
    qsi_score: float = 0.0
    direct_service: bool = False


# ============================================================================
# ABSTRACT PROVIDERS
# ============================================================================

class ScheduleProvider(ABC):
    """Provides itinerary data for QSI scoring."""

    @abstractmethod
    def get_itineraries(self, direction: str) -> List[Itinerary]:
        """Get itineraries for qsi1 or qsi2."""
        ...

    @abstractmethod
    def get_metadata(self) -> Dict[str, Any]:
        """Return provider metadata."""
        ...


class DemandProvider(ABC):
    """Provides P2P and connecting city demand data."""

    @abstractmethod
    def get_p2p_segments(self) -> List[P2PSegmentData]:
        """Get P2P demand segments."""
        ...

    @abstractmethod
    def get_connecting_cities(self, direction: str) -> List[ConnectingCityData]:
        """Get connecting cities for 'home' or 'dest'."""
        ...

    @abstractmethod
    def get_metadata(self) -> Dict[str, Any]:
        """Return provider metadata."""
        ...


class QSICaptureProvider(ABC):
    """Provides pre-computed QSI capture rates (optional)."""

    @abstractmethod
    def get_captures(self, hub_code: str) -> Dict[str, float]:
        """Get city_code -> capture_rate dict for a hub."""
        ...


# ============================================================================
# EXCEL IMPLEMENTATIONS
# ============================================================================

class ExcelScheduleProvider(ScheduleProvider):
    """Reads itineraries from QSI Excel files (QSILHR, QSISJC, etc.)."""

    def __init__(self, qsi1_file: str, qsi2_file: str = None):
        self.qsi1_file = qsi1_file
        self.qsi2_file = qsi2_file or qsi1_file
        self._cache = {}

    def get_itineraries(self, direction: str) -> List[Itinerary]:
        if direction in self._cache:
            return self._cache[direction]

        if direction == 'qsi1':
            data = self._load_sheet(self.qsi1_file, 'QSI 1')
        elif direction == 'qsi2':
            data = self._load_sheet(self.qsi1_file, 'QSI 2')
        else:
            data = []

        self._cache[direction] = data
        return data

    def _load_sheet(self, filepath: str, sheet_name: str) -> List[Itinerary]:
        """Load itineraries from a QSI sheet.
        
        QSILHR/QSISJC column layout (header at row 4):
            Col 0: City Label          Col 1: Airport Label
            Col 2: Route Label         Col 3: Dep. Aprt.
            Col 5: Carrier (leg 1)     Col 7: Conex Apt.
            Col 17: Carrier (leg 2)    Col 18: Freq.
            Col 19: Elapsed Time       Col 20: Conex Type
        """
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 6:
            return []

        # Find header row  look for 'Route Label' or 'City Label'
        header_row = None
        for i in range(min(10, len(rows))):
            row = rows[i]
            if row and any(str(v).strip() in ('Route Label', 'City Label', 'City Code')
                          for v in row if v):
                header_row = i
                break

        if header_row is None:
            return []

        header = rows[header_row]
        
        # Build column map from actual header names
        col_map = {}
        for i, val in enumerate(header):
            if val:
                key = str(val).strip().lower()
                col_map[key] = i

        # Map logical fields to actual column indices
        def find_col(*candidates):
            for c in candidates:
                cl = c.lower()
                if cl in col_map:
                    return col_map[cl]
                # Partial match
                for k, v in col_map.items():
                    if cl in k:
                        return v
            return None

        city_col = find_col('city label', 'city code', 'city')
        airport_col = find_col('airport label', 'airport code', 'airport')
        route_col = find_col('route label', 'route')
        dep_col = find_col('dep. aprt.', 'dep airport', 'dep. apt', 'origin')
        cnx_col = find_col('conex apt.', 'cnx airport', 'connection airport', 'conex apt')
        freq_col = find_col('freq.', 'frequency', 'freq')
        elapsed_col = find_col('elapsed time', 'elapsed')
        cnx_type_col = find_col('conex type', 'cnx type', 'connection type', 'type')

        # For carriers: leg 1 carrier is near col 5, leg 2 carrier near col 17
        # Handle the duplicate 'carrier' column name issue
        carrier_cols = [i for i, v in enumerate(header) if v and 'carrier' in str(v).lower()]
        carrier1_col = carrier_cols[0] if len(carrier_cols) >= 1 else find_col('carrier')
        carrier2_col = carrier_cols[1] if len(carrier_cols) >= 2 else carrier1_col

        if route_col is None or freq_col is None:
            return []

        itineraries = []
        for r in range(header_row + 1, len(rows)):
            row = rows[r]
            if not row or len(row) <= freq_col:
                continue

            def g(col_idx, default=''):
                if col_idx is not None and col_idx < len(row) and row[col_idx] is not None:
                    return row[col_idx]
                return default

            city = str(g(city_col, '')).strip()
            route_label = str(g(route_col, '')).strip()
            if not city or not route_label:
                continue

            airport = str(g(airport_col, '')).strip()
            dep_apt = str(g(dep_col, '')).strip()
            cnx_apt = str(g(cnx_col, '')).strip()
            carrier1 = str(g(carrier1_col, '')).strip()
            carrier2 = str(g(carrier2_col, '')).strip()

            # Frequency
            freq_val = g(freq_col, 0)
            try:
                freq = int(float(freq_val)) if freq_val else 0
            except (ValueError, TypeError):
                freq = 0

            # Elapsed time
            elapsed_val = g(elapsed_col, 0)
            try:
                elapsed = int(float(elapsed_val)) if elapsed_val else 0
            except (ValueError, TypeError):
                elapsed = 0

            # Connection type
            cnx_type = str(g(cnx_type_col, 'INTERLINING')).strip()

            if freq == 0:
                continue

            it = Itinerary(
                city=city, airport=airport, route_label=route_label,
                dep_airport=dep_apt, cnx_airport=cnx_apt,
                carrier_l1=carrier1, carrier_l2=carrier2,
                freq=freq, elapsed=elapsed, cnx_type=cnx_type,
            )
            itineraries.append(it)

        return itineraries

    def get_metadata(self) -> Dict[str, Any]:
        return {
            'provider_type': 'ExcelScheduleProvider',
            'qsi1_file': self.qsi1_file,
            'qsi2_file': self.qsi2_file,
        }


class ExcelDemandProvider(DemandProvider):
    """Reads demand data from forecast Excel files."""

    def __init__(self, forecast_file: str, p2p_config: Dict[str, Any],
                 home_growth: float = 0.10, dest_growth: float = 0.10):
        self.forecast_file = forecast_file
        self.p2p_config = p2p_config
        self.home_growth = home_growth
        self.dest_growth = dest_growth
        self._cnx_cache = {}

    def get_p2p_segments(self) -> List[P2PSegmentData]:
        """Build P2P segments from configuration."""
        segments = []
        for seg_cfg in self.p2p_config.get('segments', []):
            subsegments = []
            if 'subsegments' in seg_cfg:
                for sub_cfg in seg_cfg['subsegments']:
                    subsegments.append(P2PSubsegmentData(
                        name=sub_cfg['name'],
                        base_demand=sub_cfg['base_demand'],
                        growth_rate=sub_cfg['growth_rate'],
                        seasonality=sub_cfg.get('seasonality', 1.0),
                        stimulation=sub_cfg.get('stimulation', 1.0),
                        capture_rate=sub_cfg.get('capture_rate', 0.0),
                    ))

            segments.append(P2PSegmentData(
                name=seg_cfg['name'],
                base_demand=seg_cfg['base_demand'],
                growth_rate=seg_cfg['growth_rate'],
                seasonality=seg_cfg.get('seasonality', 1.0),
                stimulation=seg_cfg.get('stimulation', 1.0),
                capture_rate=seg_cfg.get('capture_rate', 0.0),
                subsegments=subsegments,
            ))
        return segments

    def get_connecting_cities(self, direction: str) -> List[ConnectingCityData]:
        """Load connecting city data from the forecast file."""
        if direction in self._cnx_cache:
            return self._cnx_cache[direction]

        cities = self._load_connecting(direction)
        self._cnx_cache[direction] = cities
        return cities

    def _load_connecting(self, direction: str) -> List[ConnectingCityData]:
        """Load connecting cities from forecast file sheets.
        
        Reads from 'Forecast Cnx @ Home Airport' or 'Forecast Cnx @ Dest. Airport'.
        These sheets have a fixed structure (header at row 4, data from row 5):
            Col 0: Nr
            Col 1: City Code
            Col 2: City Name
            Col 3: Country
            Col 4: Direct Service Flag
            Col 7: Base Total Demand
            Col 8: Compound Growth
            Col 12: QSI (expert-calibrated capture rate)
            Col 15: Adjusted QSI
        """
        wb = openpyxl.load_workbook(self.forecast_file, data_only=True, read_only=True)

        # Determine which sheet to read
        if direction == 'home':
            sheet_candidates = ['Forecast Cnx @ Home Airport', 'Forecast Cnx @ Home',
                                'Cnx at Home Hub TABLE', 'LHR Connecting', 'Connecting']
            default_growth = self.home_growth
        else:
            sheet_candidates = ['Forecast Cnx @ Dest. Airport', 'Forecast Cnx @ Dest',
                                'Cnx at Dest Apt TABLE', 'SJC Connecting', 'Dest Connecting']
            default_growth = self.dest_growth

        # Find matching sheet
        target_sheet = None
        for candidate in sheet_candidates:
            for sn in wb.sheetnames:
                if candidate.lower() == sn.lower() or candidate.lower() in sn.lower():
                    target_sheet = sn
                    break
            if target_sheet:
                break

        if not target_sheet:
            wb.close()
            return []

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 6:
            return []

        # Find header row by looking for 'City Code'
        header_row_idx = None
        for i, row in enumerate(rows):
            if row and len(row) > 1:
                for v in row:
                    if v and 'city code' in str(v).lower():
                        header_row_idx = i
                        break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            return []

        header = rows[header_row_idx]
        # Build column map
        col_map = {}
        for i, val in enumerate(header):
            if val:
                key = str(val).strip().lower()
                col_map[key] = i

        cities = []
        for r in range(header_row_idx + 1, len(rows)):
            row = rows[r]
            if not row or len(row) < 8:
                continue

            # City code  usually col 1
            city_idx = col_map.get('city code', 1)
            city_code = str(row[city_idx]).strip() if row[city_idx] else ''
            if not city_code:
                continue

            # Skip summary/total rows
            if city_code.lower() in ('total', 'totals', 'sum', 'grand total', '0'):
                continue

            # City name  usually col 2
            name_idx = col_map.get('city name', 2)
            city_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ''
            if not city_name or city_name == '-':
                continue

            # Country  usually col 3
            country_idx = col_map.get('country', 3)
            country = str(row[country_idx]).strip() if country_idx < len(row) and row[country_idx] else ''

            # Direct service flag  usually col 4
            direct_idx = col_map.get('direct service flag', 4)
            direct_val = row[direct_idx] if direct_idx < len(row) else None
            direct = False
            if direct_val and 'direct service' in str(direct_val).lower() and 'no' not in str(direct_val).lower():
                direct = True

            # Base total demand  usually col 7
            # Base demand: use INDIRECT demand when available.
            # For connecting traffic, analyst QSI (col 12) is calibrated
            # against indirect market. Using total demand for cities with
            # direct competition would overcount. For "No Direct Service"
            # cities, indirect == total anyway.
            indirect_idx = col_map.get('base indirect demand', 6)
            total_idx = col_map.get('base total demand', 7)
            try:
                indirect_dem = float(row[indirect_idx]) if indirect_idx is not None and indirect_idx < len(row) and row[indirect_idx] else 0
            except (ValueError, TypeError):
                indirect_dem = 0
            try:
                total_dem = float(row[total_idx]) if total_idx is not None and total_idx < len(row) and row[total_idx] else 0
            except (ValueError, TypeError):
                total_dem = 0
            base_demand = indirect_dem if indirect_dem > 0 else total_dem

            # Growth  usually col 8
            growth_idx = col_map.get('compound growth', 8)
            try:
                growth = float(row[growth_idx]) if growth_idx < len(row) and row[growth_idx] else default_growth
            except (ValueError, TypeError):
                growth = default_growth

            # QSI (expert-calibrated)  col 12 or 'Adjusted QSI' col 15
            qsi_idx = col_map.get('qsi', col_map.get('adjusted qsi', 12))
            try:
                qsi_score = float(row[qsi_idx]) if qsi_idx < len(row) and row[qsi_idx] else 0
            except (ValueError, TypeError):
                qsi_score = 0

            if base_demand == 0:
                continue

            cities.append(ConnectingCityData(
                city_code=city_code,
                city_name=city_name,
                country=country,
                base_demand=base_demand,
                growth_rate=growth,
                qsi_score=qsi_score,
                direct_service=direct,
            ))

        return cities

    def get_metadata(self) -> Dict[str, Any]:
        return {
            'provider_type': 'ExcelDemandProvider',
            'forecast_file': self.forecast_file,
            'home_growth': self.home_growth,
            'dest_growth': self.dest_growth,
        }


# ============================================================================
# IN-MEMORY IMPLEMENTATIONS (for testing)
# ============================================================================

class InMemoryScheduleProvider(ScheduleProvider):
    """In-memory schedule provider for unit tests."""

    def __init__(self, qsi1_data: List[Itinerary] = None,
                 qsi2_data: List[Itinerary] = None):
        self._data = {
            'qsi1': qsi1_data or [],
            'qsi2': qsi2_data or [],
        }

    def get_itineraries(self, direction: str) -> List[Itinerary]:
        return self._data.get(direction, [])

    def get_metadata(self) -> Dict[str, Any]:
        return {
            'provider_type': 'InMemoryScheduleProvider',
            'qsi1_count': len(self._data.get('qsi1', [])),
            'qsi2_count': len(self._data.get('qsi2', [])),
        }


class InMemoryDemandProvider(DemandProvider):
    """In-memory demand provider for unit tests."""

    def __init__(self, p2p_segments: List[P2PSegmentData] = None,
                 home_cities: List[ConnectingCityData] = None,
                 dest_cities: List[ConnectingCityData] = None):
        self._p2p = p2p_segments or []
        self._home = home_cities or []
        self._dest = dest_cities or []

    def get_p2p_segments(self) -> List[P2PSegmentData]:
        return self._p2p

    def get_connecting_cities(self, direction: str) -> List[ConnectingCityData]:
        return self._home if direction == 'home' else self._dest

    def get_metadata(self) -> Dict[str, Any]:
        return {
            'provider_type': 'InMemoryDemandProvider',
            'p2p_segments': len(self._p2p),
            'home_cities': len(self._home),
            'dest_cities': len(self._dest),
        }
