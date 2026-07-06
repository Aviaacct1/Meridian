#!/usr/bin/env python3
"""
Module VI: Assembly Loop / Pipeline Orchestrator
=================================================
Avia Solutions Proprietary Tool

Orchestrates the full forecast pipeline:
    1. Parse inputs (route config, demand data, OAG schedules, MCTs)
    2. Run Connection Builder (with and without new service)
    3. Score QSI (capture rates per connecting city)
    4. Assemble forecast (P2P + connecting at home + connecting at dest)
    5. Optionally run Departure Time Optimiser
    6. Apply reasonableness bounds checks
    7. Generate output workbook with full audit trail

Validation Target:
    BA LHR-SJC (without India): 129,162 total passengers
    - P2P: 78,110 (UK Visitors 51,462 + US Residents 26,648)
    - Connecting at LHR: 48,115
    - Connecting at SJC: 2,937
    - Load Factor: 82.9% (787-800, 214 seats, 7x weekly)
"""

import os
import sys
import math
from datetime import datetime, timedelta, time as dtime
from collections import defaultdict, OrderedDict
from typing import Dict, List, Tuple, Optional, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl --break-system-packages")
    sys.exit(1)


# ============================================================================
# ROUTE CONFIGURATION
# ============================================================================

class RouteConfig:
    """Defines all parameters for a route assessment."""
    
    def __init__(self):
        # Airline info
        self.airline_name = ""
        self.airline_code = ""
        self.home_city = ""
        self.home_city_code = ""
        self.home_airport = ""
        self.home_airport_code = ""
        self.home_country = ""
        self.dest_city = ""
        self.dest_city_code = ""
        self.dest_airport = ""
        self.dest_airport_code = ""
        self.dest_country = ""
        
        # Schedule
        self.frequency = 7
        self.aircraft_type = ""
        self.seats = 0
        self.outbound_dep = None  # time object
        self.outbound_arr = None
        self.return_dep = None
        self.return_arr = None
        self.flight_time_hrs = 0
        self.op_days = "1234567"
        
        # Forecast parameters
        self.forecast_year = 2016
        self.service_start = None
        self.service_end = None
        
        # Growth rates (by segment)
        self.growth_rates = {}  # segment_name -> rate
        
        # QSI parameters
        self.qsi_ceiling = 1.0
        self.qsi_adjustment = 1.0
        self.online_coeff = 1.0
        self.alliance_coeff = 0.615
        self.interline_coeff = 0.25
        
        # Demand segmentation
        self.p2p_segments = []  # list of P2PSegment
        self.connecting_home = []  # list of ConnectingCity
        self.connecting_dest = []  # list of ConnectingCity
    
    @classmethod
    def from_ba_lhr_sjc(cls):
        """Factory method for the BA LHR-SJC validation case."""
        cfg = cls()
        cfg.airline_name = "British Airways"
        cfg.airline_code = "BA"
        cfg.home_city = "London"
        cfg.home_city_code = "LON"
        cfg.home_airport = "London Heathrow"
        cfg.home_airport_code = "LHR"
        cfg.home_country = "United Kingdom"
        cfg.dest_city = "San Jose"
        cfg.dest_city_code = "SJC"
        cfg.dest_airport = "San Jose Intl"
        cfg.dest_airport_code = "SJC"
        cfg.dest_country = "USA"
        cfg.frequency = 7
        cfg.aircraft_type = "787-800"
        cfg.seats = 214
        cfg.outbound_dep = dtime(15, 30)  # LHR dep
        cfg.outbound_arr = dtime(18, 30)  # SJC arr
        cfg.return_dep = dtime(21, 30)    # SJC dep
        cfg.return_arr = dtime(15, 55)    # LHR arr (next day)
        cfg.flight_time_hrs = 11.0
        cfg.forecast_year = 2016
        cfg.qsi_ceiling = 1.0
        cfg.qsi_adjustment = 1.0
        return cfg


class P2PSegment:
    """A point-to-point demand segment."""
    
    def __init__(self, name, base_demand, growth_rate, seasonality=1.0,
                 stimulation=1.0, capture_rate=0.0, subsegments=None):
        self.name = name
        self.base_demand = base_demand
        self.growth_rate = growth_rate
        self.seasonality = seasonality
        self.stimulation = stimulation
        self.capture_rate = capture_rate
        self.subsegments = subsegments or []
        
        # Computed
        self.demand_forecast_year = 0
        self.demand_after_seasonality = 0
        self.demand_after_stimulation = 0
        self.forecast = 0
        self.ppdew = 0


class P2PSubsegment:
    """A subsegment within a P2P segment (e.g., Primary/Secondary/Contested within Leisure)."""
    
    def __init__(self, name, base_demand, growth_rate, seasonality=1.0,
                 stimulation=1.0, capture_rate=0.0):
        self.name = name
        self.base_demand = base_demand
        self.growth_rate = growth_rate
        self.seasonality = seasonality
        self.stimulation = stimulation
        self.capture_rate = capture_rate
        
        self.demand_forecast_year = 0
        self.demand_after_seasonality = 0
        self.demand_after_stimulation = 0
        self.forecast = 0
        self.ppdew = 0


class ConnectingCity:
    """A connecting city with demand and QSI score."""
    
    def __init__(self, city_code, city_name, country, base_demand,
                 growth_rate, qsi_score, direct_service=False,
                 seasonality=1.0, stimulation=1.0):
        self.city_code = city_code
        self.city_name = city_name
        self.country = country
        self.base_demand = base_demand
        self.growth_rate = growth_rate
        self.qsi_score = qsi_score
        self.direct_service = direct_service
        self.seasonality = seasonality
        self.stimulation = stimulation
        
        # Computed
        self.demand_forecast_year = 0
        self.forecast = 0
        self.adjusted_qsi = 0
        self.adjusted_forecast = 0
        self.pptew = 0


# ============================================================================
# DEMAND DATA LOADER
# ============================================================================

class DemandLoader:
    """Loads demand data from the project's Excel files."""
    
    @staticmethod
    def load_ba_lhr_sjc_p2p(project_dir):
        """Load P2P demand segments for BA LHR-SJC from the forecast file."""
        # These are the exact values from the validated forecast file
        segments = []
        
        # UK Visitors
        uk_business = P2PSegment(
            name="UK Business",
            base_demand=71441.55,
            growth_rate=0.10,
            seasonality=1.0,
            stimulation=1.15,
            capture_rate=0.40
        )
        
        uk_leisure = P2PSegment(
            name="UK Leisure/VFR",
            base_demand=0,  # computed from subsegments
            growth_rate=0.10,
            seasonality=1.0,
            stimulation=1.0,
            capture_rate=0.0,
            subsegments=[
                P2PSubsegment("Primary", 36385.76, 0.10, 1.0, 1.0, 0.25),
                P2PSubsegment("Secondary", 17448.74, 0.10, 1.0, 1.0, 0.25),
                P2PSubsegment("Contested", 4617.68, 0.10, 1.0, 1.0, 0.10),
            ]
        )
        
        # US Residents
        us_business = P2PSegment(
            name="US Business",
            base_demand=65946.05,
            growth_rate=0.10,
            seasonality=1.0,
            stimulation=1.15,
            capture_rate=0.15
        )
        
        us_leisure = P2PSegment(
            name="US Leisure/VFR",
            base_demand=0,
            growth_rate=0.10,
            seasonality=1.0,
            stimulation=1.0,
            capture_rate=0.0,
            subsegments=[
                P2PSubsegment("Primary", 33586.86, 0.10, 1.0, 1.0, 0.25),
                P2PSubsegment("Secondary", 16106.53, 0.10, 1.0, 1.0, 0.25),
                P2PSubsegment("Contested", 4262.47, 0.10, 1.0, 1.0, 0.10),
            ]
        )
        
        return [uk_business, uk_leisure, us_business, us_leisure]
    
    @staticmethod
    def load_ba_lhr_sjc_connecting(project_dir, direction='home'):
        """Load connecting demand from the forecast file."""
        fname = os.path.join(project_dir, 
                             'BA_Fcst_LHRSJC_JZ_23Feb2015_FINAL_without_INDIA.xlsm')
        wb = openpyxl.load_workbook(fname, data_only=True)
        
        if direction == 'home':
            ws = wb['Forecast Cnx @ Home Airport']
            growth = 0.09
        else:
            ws = wb['Forecast Cnx @ Dest. Airport']
            growth = 0.10
        
        cities = []
        for row in ws.iter_rows(min_row=5, max_row=200, max_col=19, values_only=True):
            if not row[1] or not isinstance(row[1], str) or len(row[1]) != 3:
                continue
            
            city_code = row[1]
            city_name = row[2] if row[2] else city_code
            country = row[3] if row[3] else ""
            direct_flag = (row[4] == "Direct Service") if row[4] else False
            base_direct = row[5] if row[5] else 0
            base_indirect = row[6] if row[6] else 0
            base_total = row[7] if row[7] else 0
            qsi = row[12] if row[12] else 0  # QSI score
            
            # KEY INSIGHT: For cities with direct service, QSI only applies
            # to the INDIRECT demand portion (passengers currently connecting).
            # Direct service passengers are already captured by the direct route.
            if direct_flag:
                base_demand = base_indirect  # Only indirect demand is capturable
            else:
                base_demand = base_total
            
            if base_demand > 0:
                cities.append(ConnectingCity(
                    city_code=city_code,
                    city_name=city_name,
                    country=country,
                    base_demand=base_demand,
                    growth_rate=growth,
                    qsi_score=qsi,
                    direct_service=direct_flag
                ))
        
        wb.close()
        return cities


# ============================================================================
# QSI SCORE LOADER  
# ============================================================================

class QSIScoreLoader:
    """Loads QSI scores from QSI model output files or forecast files."""
    
    @staticmethod
    def load_from_forecast(project_dir, direction='home'):
        """Load QSI scores directly from the forecast file's QSI sheet."""
        fname = os.path.join(project_dir,
                             'BA_Fcst_LHRSJC_JZ_23Feb2015_FINAL_without_INDIA.xlsm')
        wb = openpyxl.load_workbook(fname, data_only=True)
        
        if direction == 'home':
            ws = wb['QSI for Home Airport']
        else:
            ws = wb['QSI for Destination Airport']
        
        scores = {}
        for row in ws.iter_rows(min_row=5, max_row=200, max_col=17, values_only=True):
            if row[1] and isinstance(row[1], str) and len(row[1]) == 3:
                city_code = row[1]
                # The "Selected" QSI is in the last column (index 16)
                selected_qsi = row[16] if row[16] else 0
                # If no selected value, try column 7 (frequency=7 QSI)
                if selected_qsi == 0 and row[7]:
                    selected_qsi = row[7]
                scores[city_code] = selected_qsi
        
        wb.close()
        return scores


# ============================================================================
# FORECAST ENGINE
# ============================================================================

class ForecastEngine:
    """
    Assembles the complete route forecast from:
    - P2P demand segments with capture rates
    - Connecting demand at home hub with QSI-derived capture rates
    - Connecting demand at destination with QSI-derived capture rates
    """
    
    def __init__(self, config: RouteConfig):
        self.config = config
        self.p2p_segments = []
        self.connecting_home = []
        self.connecting_dest = []
        
        # Results
        self.total_p2p = 0
        self.total_cnx_home = 0
        self.total_cnx_dest = 0
        self.grand_total = 0
        self.annual_capacity = 0
        self.load_factor = 0
        self.ppdew = 0
        
        # Audit trail
        self.audit = []
    
    def load_demand(self, p2p_segments, connecting_home, connecting_dest):
        """Load all demand data."""
        self.p2p_segments = p2p_segments
        self.connecting_home = connecting_home
        self.connecting_dest = connecting_dest
    
    def _grow_demand(self, base, rate, years=1):
        """Apply compound growth."""
        return base * (1 + rate) ** years
    
    def run_p2p_forecast(self):
        """Compute P2P demand forecast for all segments."""
        self.audit.append("=" * 60)
        self.audit.append("P2P FORECAST COMPUTATION")
        self.audit.append("=" * 60)
        
        total_p2p = 0
        
        for seg in self.p2p_segments:
            if seg.subsegments:
                # Process subsegments
                seg_total = 0
                for sub in seg.subsegments:
                    sub.demand_forecast_year = self._grow_demand(
                        sub.base_demand, sub.growth_rate)
                    sub.demand_after_seasonality = sub.demand_forecast_year * sub.seasonality
                    sub.demand_after_stimulation = sub.demand_after_seasonality * sub.stimulation
                    sub.forecast = sub.demand_after_stimulation * sub.capture_rate
                    sub.ppdew = sub.forecast / 728 if sub.forecast else 0
                    seg_total += sub.forecast
                    
                    self.audit.append(
                        f"  {seg.name}/{sub.name}: Base={sub.base_demand:,.0f} "
                        f" Grown={sub.demand_forecast_year:,.0f} "
                        f" Stim={sub.demand_after_stimulation:,.0f} "
                        f" CR={sub.capture_rate:.0%} = {sub.forecast:,.0f}")
                
                seg.forecast = seg_total
                seg.ppdew = seg_total / 728 if seg_total else 0
            else:
                seg.demand_forecast_year = self._grow_demand(
                    seg.base_demand, seg.growth_rate)
                seg.demand_after_seasonality = seg.demand_forecast_year * seg.seasonality
                seg.demand_after_stimulation = seg.demand_after_seasonality * seg.stimulation
                seg.forecast = seg.demand_after_stimulation * seg.capture_rate
                seg.ppdew = seg.forecast / 728 if seg.forecast else 0
                
                self.audit.append(
                    f"  {seg.name}: Base={seg.base_demand:,.0f} "
                    f" Grown={seg.demand_forecast_year:,.0f} "
                    f" Stim={seg.demand_after_stimulation:,.0f} "
                    f" CR={seg.capture_rate:.0%} = {seg.forecast:,.0f}")
            
            total_p2p += seg.forecast
        
        self.total_p2p = total_p2p
        self.audit.append(f"\n  TOTAL P2P: {total_p2p:,.0f}")
        return total_p2p
    
    def run_connecting_forecast(self, cities, direction_label):
        """Compute connecting demand forecast."""
        self.audit.append(f"\n{'=' * 60}")
        self.audit.append(f"CONNECTING FORECAST: {direction_label}")
        self.audit.append("=" * 60)
        
        total_direct = 0
        total_no_direct = 0
        total_demand_direct = 0
        total_demand_no_direct = 0
        
        for city in cities:
            city.demand_forecast_year = self._grow_demand(
                city.base_demand, city.growth_rate)
            city.adjusted_qsi = city.qsi_score * self.config.qsi_adjustment
            
            # Apply ceiling
            if city.adjusted_qsi > self.config.qsi_ceiling:
                city.adjusted_qsi = self.config.qsi_ceiling
            
            city.forecast = city.demand_forecast_year * city.adjusted_qsi
            city.adjusted_forecast = city.forecast * city.seasonality
            city.pptew = city.adjusted_forecast / 728 if city.adjusted_forecast else 0
            
            if city.direct_service:
                total_direct += city.adjusted_forecast
                total_demand_direct += city.demand_forecast_year
            else:
                total_no_direct += city.adjusted_forecast
                total_demand_no_direct += city.demand_forecast_year
            
            if city.adjusted_forecast > 100:
                self.audit.append(
                    f"  {city.city_code} {city.city_name:20s}: "
                    f"Demand={city.demand_forecast_year:>10,.0f} "
                    f" QSI={city.adjusted_qsi:.4f} "
                    f"= {city.adjusted_forecast:>8,.0f}")
        
        total = total_direct + total_no_direct
        total_demand = total_demand_direct + total_demand_no_direct
        
        # Compute aggregate capture rate
        agg_capture = total / total_demand if total_demand > 0 else 0
        
        self.audit.append(f"\n  Direct competition: {total_direct:,.0f}")
        self.audit.append(f"  No direct competition: {total_no_direct:,.0f}")
        self.audit.append(f"  TOTAL {direction_label}: {total:,.0f}")
        self.audit.append(f"  Aggregate capture rate: {agg_capture:.4%}")
        
        return total, total_direct, total_no_direct, total_demand
    
    def run(self):
        """Execute the full forecast pipeline."""
        self.audit = []
        self.audit.append("AVIA SOLUTIONS ROUTE FORECAST")
        self.audit.append(f"Route: {self.config.home_airport_code}-{self.config.dest_airport_code}")
        self.audit.append(f"Airline: {self.config.airline_name} ({self.config.airline_code})")
        self.audit.append(f"Aircraft: {self.config.aircraft_type} ({self.config.seats} seats)")
        self.audit.append(f"Frequency: {self.config.frequency}x weekly")
        self.audit.append(f"Forecast Year: {self.config.forecast_year}")
        self.audit.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # 1. P2P forecast
        self.total_p2p = self.run_p2p_forecast()
        
        # 2. Connecting at home hub
        self.total_cnx_home, cnx_home_direct, cnx_home_no_direct, cnx_home_demand = \
            self.run_connecting_forecast(self.connecting_home, 
                                         f"Connecting at {self.config.home_airport_code}")
        
        # 3. Connecting at destination
        self.total_cnx_dest, cnx_dest_direct, cnx_dest_no_direct, cnx_dest_demand = \
            self.run_connecting_forecast(self.connecting_dest,
                                         f"Connecting at {self.config.dest_airport_code}")
        
        # 4. Grand total
        self.grand_total = self.total_p2p + self.total_cnx_home + self.total_cnx_dest
        
        # 5. Capacity and load factor
        weeks_per_year = 728 / self.config.frequency  # 728 = 52 weeks * 14 (7 days * 2 directions)
        # Actually: annual capacity = seats * frequency * 52 * 2 directions... 
        # No - in the BA case: 214 seats * 364 flights = 155,792 (one-way annual)
        # But pax is roundtrip counted once? Let me check...
        # From the output: Annual Capacity 155,792 = 214 * 728 flights
        # 728 = 7 freq * 52 weeks * 2 directions
        self.annual_capacity = self.config.seats * self.config.frequency * 52 * 2
        self.load_factor = self.grand_total / self.annual_capacity if self.annual_capacity > 0 else 0
        self.ppdew = self.grand_total / 728
        
        self.audit.append(f"\n{'=' * 60}")
        self.audit.append("SUMMARY")
        self.audit.append("=" * 60)
        self.audit.append(f"  P2P Total:                {self.total_p2p:>10,.0f}  ({self.total_p2p/self.grand_total:.1%})")
        self.audit.append(f"  Connecting at {self.config.home_airport_code}:      {self.total_cnx_home:>10,.0f}  ({self.total_cnx_home/self.grand_total:.1%})")
        self.audit.append(f"  Connecting at {self.config.dest_airport_code}:      {self.total_cnx_dest:>10,.0f}  ({self.total_cnx_dest/self.grand_total:.1%})")
        self.audit.append(f"  {'' * 40}")
        self.audit.append(f"  GRAND TOTAL:              {self.grand_total:>10,.0f}")
        self.audit.append(f"  Annual Capacity:          {self.annual_capacity:>10,}")
        self.audit.append(f"  Load Factor:              {self.load_factor:>10.1%}")
        self.audit.append(f"  PPDEW:                    {self.ppdew:>10.1f}")
        
        return self.grand_total
    
    def reasonableness_check(self):
        """Apply reasonableness bounds and flag issues."""
        issues = []
        
        # Load factor bounds
        if self.load_factor > 0.95:
            issues.append(f"WARNING: Load factor {self.load_factor:.1%} exceeds 95% - may be overconstrained")
        elif self.load_factor > 0.90:
            issues.append(f"CAUTION: Load factor {self.load_factor:.1%} is very high - limited upside")
        elif self.load_factor < 0.50:
            issues.append(f"WARNING: Load factor {self.load_factor:.1%} below 50% - route may not be viable")
        elif self.load_factor < 0.65:
            issues.append(f"CAUTION: Load factor {self.load_factor:.1%} is low - check assumptions")
        
        # P2P share
        p2p_share = self.total_p2p / self.grand_total if self.grand_total > 0 else 0
        if p2p_share < 0.20:
            issues.append(f"CAUTION: P2P share {p2p_share:.0%} is very low - heavily dependent on connections")
        
        # Connecting concentration
        if self.connecting_home:
            sorted_cities = sorted(self.connecting_home, key=lambda c: c.adjusted_forecast, reverse=True)
            if sorted_cities:
                top_city_share = sorted_cities[0].adjusted_forecast / self.total_cnx_home if self.total_cnx_home > 0 else 0
                if top_city_share > 0.20:
                    issues.append(f"CAUTION: Top connecting city ({sorted_cities[0].city_code}) represents "
                                 f"{top_city_share:.0%} of hub connections  concentration risk")
        
        # QSI score reasonableness
        for city in self.connecting_home:
            if city.adjusted_qsi > 0.30 and city.demand_forecast_year > 5000:
                issues.append(f"CHECK: {city.city_code} QSI={city.adjusted_qsi:.1%} seems high for "
                             f"a market with {city.demand_forecast_year:,.0f} demand")
        
        self.audit.append(f"\n{'=' * 60}")
        self.audit.append("REASONABLENESS CHECKS")
        self.audit.append("=" * 60)
        if issues:
            for issue in issues:
                self.audit.append(f"   {issue}")
        else:
            self.audit.append("   All checks passed")
        
        return issues


# ============================================================================
# OUTPUT WRITER
# ============================================================================

class OutputWriter:
    """Generates the output Excel workbook."""
    
    # Colour scheme
    HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="002060")
    SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="002060")
    DATA_FONT = Font(name="Calibri", size=10)
    TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
    TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    THIN_BORDER = Border(
        bottom=Side(style='thin', color='808080')
    )
    
    def __init__(self, engine: ForecastEngine):
        self.engine = engine
        self.wb = openpyxl.Workbook()
    
    def _set_col_widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    
    def _write_header_row(self, ws, row, headers, start_col=1):
        for j, h in enumerate(headers, start_col):
            cell = ws.cell(row=row, column=j, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    def write_forecast_table(self):
        """Write the main forecast table - matches the BA output format."""
        ws = self.wb.active
        ws.title = "Forecast TABLE"
        
        cfg = self.engine.config
        
        # Title
        title = (f"{cfg.airline_name}'s {cfg.home_airport_code} - {cfg.dest_airport_code} "
                f"Traffic Forecast ({cfg.frequency}x Weekly Service)")
        ws.cell(row=1, column=1, value=title).font = self.TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        
        # Headers
        headers = [
            "Market",
            "Base Annual\nDemand (000s)",
            "Compound\nGrowth Rate",
            "Annual Demand\nYE " + str(cfg.forecast_year) + "\n(000s)",
            "% of Annual\nDemand Jan-Dec",
            "Demand\nJan-Dec\n(000s)",
            "Stimulation\nDue to Direct\nService",
            "Demand After\nStimulation\n(000s)",
            f"{cfg.airline_code}'s\nCapture Rate",
            "Forecast\n(000s)",
            "PDEW"
        ]
        self._write_header_row(ws, 3, headers)
        self._set_col_widths(ws, [28, 14, 12, 14, 12, 14, 12, 14, 12, 12, 10])
        
        row = 4
        grand_total_forecast = 0
        grand_total_demand = 0
        grand_total_stim_demand = 0
        
        # Group segments by direction
        uk_segs = self.engine.p2p_segments[:2]  # UK Business + UK Leisure
        us_segs = self.engine.p2p_segments[2:]  # US Business + US Leisure
        
        for group_name, segs in [("UK Visitors", uk_segs), ("US Residents", us_segs)]:
            # Group header
            ws.cell(row=row, column=1, value=group_name).font = self.SUBTITLE_FONT
            row += 1
            group_total_base = 0
            group_total_demand = 0
            group_total_stim = 0
            group_total_forecast = 0
            
            for seg in segs:
                if seg.subsegments:
                    # Write subsegment header
                    ws.cell(row=row, column=1, value=seg.name).font = Font(name="Calibri", size=10, italic=True)
                    row += 1
                    
                    for sub in seg.subsegments:
                        ws.cell(row=row, column=1, value=f"    {sub.name}").font = self.DATA_FONT
                        ws.cell(row=row, column=2, value=sub.base_demand/1000).font = self.DATA_FONT
                        ws.cell(row=row, column=2).number_format = '#,##0.00'
                        ws.cell(row=row, column=3, value=sub.growth_rate).font = self.DATA_FONT
                        ws.cell(row=row, column=3).number_format = '0%'
                        ws.cell(row=row, column=4, value=sub.demand_forecast_year/1000).font = self.DATA_FONT
                        ws.cell(row=row, column=4).number_format = '#,##0.00'
                        ws.cell(row=row, column=5, value=sub.seasonality).font = self.DATA_FONT
                        ws.cell(row=row, column=5).number_format = '0%'
                        ws.cell(row=row, column=6, value=sub.demand_after_seasonality/1000).font = self.DATA_FONT
                        ws.cell(row=row, column=6).number_format = '#,##0.00'
                        ws.cell(row=row, column=7, value=sub.stimulation).font = self.DATA_FONT
                        ws.cell(row=row, column=7).number_format = '0.00'
                        ws.cell(row=row, column=8, value=sub.demand_after_stimulation/1000).font = self.DATA_FONT
                        ws.cell(row=row, column=8).number_format = '#,##0.00'
                        ws.cell(row=row, column=9, value=sub.capture_rate).font = self.DATA_FONT
                        ws.cell(row=row, column=9).number_format = '0%'
                        ws.cell(row=row, column=10, value=sub.forecast/1000).font = self.DATA_FONT
                        ws.cell(row=row, column=10).number_format = '#,##0.00'
                        ws.cell(row=row, column=11, value=sub.ppdew).font = self.DATA_FONT
                        ws.cell(row=row, column=11).number_format = '#,##0.0'
                        
                        group_total_base += sub.base_demand
                        group_total_demand += sub.demand_forecast_year
                        group_total_stim += sub.demand_after_stimulation
                        group_total_forecast += sub.forecast
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=seg.name).font = self.DATA_FONT
                    ws.cell(row=row, column=2, value=seg.base_demand/1000).font = self.DATA_FONT
                    ws.cell(row=row, column=2).number_format = '#,##0.00'
                    ws.cell(row=row, column=3, value=seg.growth_rate).font = self.DATA_FONT
                    ws.cell(row=row, column=3).number_format = '0%'
                    ws.cell(row=row, column=4, value=seg.demand_forecast_year/1000).font = self.DATA_FONT
                    ws.cell(row=row, column=4).number_format = '#,##0.00'
                    ws.cell(row=row, column=5, value=seg.seasonality).font = self.DATA_FONT
                    ws.cell(row=row, column=5).number_format = '0%'
                    ws.cell(row=row, column=6, value=seg.demand_after_seasonality/1000).font = self.DATA_FONT
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                    ws.cell(row=row, column=7, value=seg.stimulation).font = self.DATA_FONT
                    ws.cell(row=row, column=7).number_format = '0.00'
                    ws.cell(row=row, column=8, value=seg.demand_after_stimulation/1000).font = self.DATA_FONT
                    ws.cell(row=row, column=8).number_format = '#,##0.00'
                    ws.cell(row=row, column=9, value=seg.capture_rate).font = self.DATA_FONT
                    ws.cell(row=row, column=9).number_format = '0%'
                    ws.cell(row=row, column=10, value=seg.forecast/1000).font = self.DATA_FONT
                    ws.cell(row=row, column=10).number_format = '#,##0.00'
                    ws.cell(row=row, column=11, value=seg.ppdew).font = self.DATA_FONT
                    ws.cell(row=row, column=11).number_format = '#,##0.0'
                    
                    group_total_base += seg.base_demand
                    group_total_demand += seg.demand_forecast_year
                    group_total_stim += seg.demand_after_stimulation
                    group_total_forecast += seg.forecast
                    row += 1
            
            # Group total row
            ws.cell(row=row, column=1, value=f"Total {group_name.split()[0]}").font = self.TOTAL_FONT
            ws.cell(row=row, column=2, value=group_total_base/1000).font = self.TOTAL_FONT
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=group_total_demand/1000).font = self.TOTAL_FONT
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=group_total_stim/1000).font = self.TOTAL_FONT
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            capture = group_total_forecast / group_total_stim if group_total_stim > 0 else 0
            ws.cell(row=row, column=9, value=capture).font = self.TOTAL_FONT
            ws.cell(row=row, column=9).number_format = '0.0%'
            ws.cell(row=row, column=10, value=group_total_forecast/1000).font = self.TOTAL_FONT
            ws.cell(row=row, column=10).number_format = '#,##0.00'
            ws.cell(row=row, column=11, value=group_total_forecast/728).font = self.TOTAL_FONT
            ws.cell(row=row, column=11).number_format = '#,##0.0'
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = self.TOTAL_FILL
            
            grand_total_forecast += group_total_forecast
            grand_total_stim_demand += group_total_stim
            row += 2
        
        # P2P total
        ws.cell(row=row, column=1, value="Total Point to Point").font = self.TOTAL_FONT
        ws.cell(row=row, column=10, value=self.engine.total_p2p/1000).font = self.TOTAL_FONT
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=self.engine.total_p2p/728).font = self.TOTAL_FONT
        ws.cell(row=row, column=11).number_format = '#,##0.0'
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = self.TOTAL_FILL
        row += 2
        
        # Connecting at home hub
        ws.cell(row=row, column=1, 
                value=f"Connecting at {cfg.home_airport_code}").font = self.SUBTITLE_FONT
        row += 1
        
        cnx_home_direct = sum(c.adjusted_forecast for c in self.engine.connecting_home if c.direct_service)
        cnx_home_no_direct = sum(c.adjusted_forecast for c in self.engine.connecting_home if not c.direct_service)
        demand_home_direct = sum(c.demand_forecast_year for c in self.engine.connecting_home if c.direct_service)
        demand_home_no_direct = sum(c.demand_forecast_year for c in self.engine.connecting_home if not c.direct_service)
        
        for label, fcst, dem in [
            ("O&Ds with Direct Competition", cnx_home_direct, demand_home_direct),
            ("O&Ds with No Direct Competition", cnx_home_no_direct, demand_home_no_direct)
        ]:
            ws.cell(row=row, column=1, value=f"    {label}").font = self.DATA_FONT
            base_dem = dem / (1 + 0.09) if dem > 0 else 0  # back-calculate base
            ws.cell(row=row, column=2, value=base_dem/1000).font = self.DATA_FONT
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=0.09).font = self.DATA_FONT
            ws.cell(row=row, column=3).number_format = '0%'
            ws.cell(row=row, column=4, value=dem/1000).font = self.DATA_FONT
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            capture = fcst / dem if dem > 0 else 0
            ws.cell(row=row, column=9, value=capture).font = self.DATA_FONT
            ws.cell(row=row, column=9).number_format = '0.000%'
            ws.cell(row=row, column=10, value=fcst/1000).font = self.DATA_FONT
            ws.cell(row=row, column=10).number_format = '#,##0.00'
            row += 1
        
        # Home hub total
        ws.cell(row=row, column=1, value="Total").font = self.TOTAL_FONT
        ws.cell(row=row, column=10, value=self.engine.total_cnx_home/1000).font = self.TOTAL_FONT
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=self.engine.total_cnx_home/728).font = self.TOTAL_FONT
        ws.cell(row=row, column=11).number_format = '#,##0.0'
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = self.TOTAL_FILL
        row += 2
        
        # Connecting at destination
        ws.cell(row=row, column=1,
                value=f"Connecting at {cfg.dest_airport_code}").font = self.SUBTITLE_FONT
        row += 1
        
        cnx_dest_direct = sum(c.adjusted_forecast for c in self.engine.connecting_dest if c.direct_service)
        cnx_dest_no_direct = sum(c.adjusted_forecast for c in self.engine.connecting_dest if not c.direct_service)
        demand_dest_direct = sum(c.demand_forecast_year for c in self.engine.connecting_dest if c.direct_service)
        demand_dest_no_direct = sum(c.demand_forecast_year for c in self.engine.connecting_dest if not c.direct_service)
        
        for label, fcst, dem in [
            ("O&Ds with Direct Competition", cnx_dest_direct, demand_dest_direct),
            ("O&Ds with No Direct Competition", cnx_dest_no_direct, demand_dest_no_direct)
        ]:
            ws.cell(row=row, column=1, value=f"    {label}").font = self.DATA_FONT
            base_dem = dem / (1 + 0.10) if dem > 0 else 0
            ws.cell(row=row, column=2, value=base_dem/1000).font = self.DATA_FONT
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=0.10).font = self.DATA_FONT
            ws.cell(row=row, column=3).number_format = '0%'
            ws.cell(row=row, column=4, value=dem/1000).font = self.DATA_FONT
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            capture = fcst / dem if dem > 0 else 0
            ws.cell(row=row, column=9, value=capture).font = self.DATA_FONT
            ws.cell(row=row, column=9).number_format = '0.000%'
            ws.cell(row=row, column=10, value=fcst/1000).font = self.DATA_FONT
            ws.cell(row=row, column=10).number_format = '#,##0.00'
            row += 1
        
        # Dest total
        ws.cell(row=row, column=1, value="Total").font = self.TOTAL_FONT
        ws.cell(row=row, column=10, value=self.engine.total_cnx_dest/1000).font = self.TOTAL_FONT
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=self.engine.total_cnx_dest/728).font = self.TOTAL_FONT
        ws.cell(row=row, column=11).number_format = '#,##0.0'
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = self.TOTAL_FILL
        row += 2
        
        # GRAND TOTAL
        ws.cell(row=row, column=1, value="Grand Total").font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=row, column=10, value=self.engine.grand_total/1000).font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=self.engine.ppdew).font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=row, column=11).number_format = '#,##0.0'
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            ws.cell(row=row, column=c).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        
        row += 2
        # Load factor
        ws.cell(row=row, column=1, value="Aircraft").font = self.DATA_FONT
        ws.cell(row=row, column=2, value=cfg.aircraft_type).font = self.DATA_FONT
        row += 1
        ws.cell(row=row, column=1, value="Seats").font = self.DATA_FONT
        ws.cell(row=row, column=2, value=cfg.seats).font = self.DATA_FONT
        row += 1
        ws.cell(row=row, column=1, value="Annual Capacity").font = self.DATA_FONT
        ws.cell(row=row, column=2, value=self.engine.annual_capacity).font = self.DATA_FONT
        ws.cell(row=row, column=2).number_format = '#,##0'
        row += 1
        ws.cell(row=row, column=1, value="Seat Factor").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=2, value=self.engine.load_factor).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=2).number_format = '0.0%'
        
        return ws
    
    def write_connecting_table(self, cities, direction, sheet_name):
        """Write a connecting cities detail table."""
        ws = self.wb.create_sheet(sheet_name)
        
        cfg = self.engine.config
        airport = cfg.home_airport_code if direction == 'home' else cfg.dest_airport_code
        
        ws.cell(row=1, column=1, 
                value=f"Forecast Number of Passengers Connecting at {airport}").font = self.TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        
        headers = ["Nr", "City Code", "City Name", "Country",
                   f"Annual Demand\nYE {cfg.forecast_year}",
                   f"{cfg.airline_code}'s\nAnnual Share",
                   "Annual\nForecast",
                   "Seasonality\nAdj. Forecast",
                   "PDEW"]
        self._write_header_row(ws, 3, headers)
        self._set_col_widths(ws, [6, 10, 20, 20, 14, 12, 12, 14, 10])
        
        # Sort by forecast descending
        sorted_cities = sorted(cities, key=lambda c: c.adjusted_forecast, reverse=True)
        
        total_demand = 0
        total_forecast = 0
        
        for i, city in enumerate(sorted_cities, 1):
            r = i + 3
            ws.cell(row=r, column=1, value=i).font = self.DATA_FONT
            ws.cell(row=r, column=2, value=city.city_code).font = self.DATA_FONT
            ws.cell(row=r, column=3, value=city.city_name).font = self.DATA_FONT
            ws.cell(row=r, column=4, value=city.country).font = self.DATA_FONT
            ws.cell(row=r, column=5, value=city.demand_forecast_year).font = self.DATA_FONT
            ws.cell(row=r, column=5).number_format = '#,##0'
            ws.cell(row=r, column=6, value=city.adjusted_qsi).font = self.DATA_FONT
            ws.cell(row=r, column=6).number_format = '0.000%'
            ws.cell(row=r, column=7, value=city.adjusted_forecast).font = self.DATA_FONT
            ws.cell(row=r, column=7).number_format = '#,##0'
            ws.cell(row=r, column=8, value=city.adjusted_forecast).font = self.DATA_FONT
            ws.cell(row=r, column=8).number_format = '#,##0'
            ws.cell(row=r, column=9, value=city.pptew).font = self.DATA_FONT
            ws.cell(row=r, column=9).number_format = '#,##0.00'
            
            total_demand += city.demand_forecast_year
            total_forecast += city.adjusted_forecast
        
        # Total row
        r = len(sorted_cities) + 4
        ws.cell(row=r, column=1, value="Total").font = self.TOTAL_FONT
        ws.cell(row=r, column=5, value=total_demand).font = self.TOTAL_FONT
        ws.cell(row=r, column=5).number_format = '#,##0'
        agg_capture = total_forecast / total_demand if total_demand > 0 else 0
        ws.cell(row=r, column=6, value=agg_capture).font = self.TOTAL_FONT
        ws.cell(row=r, column=6).number_format = '0.000%'
        ws.cell(row=r, column=7, value=total_forecast).font = self.TOTAL_FONT
        ws.cell(row=r, column=7).number_format = '#,##0'
        ws.cell(row=r, column=9, value=total_forecast / 728).font = self.TOTAL_FONT
        ws.cell(row=r, column=9).number_format = '#,##0.00'
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = self.TOTAL_FILL
        
        # Notes
        r += 2
        ws.cell(row=r, column=1, value="Notes:").font = Font(name="Calibri", size=9, italic=True)
        ws.cell(row=r+1, column=1, value="1) Demand on double connections excluded.").font = Font(name="Calibri", size=9)
        ws.cell(row=r+2, column=1, value="2) Based on QSI model.").font = Font(name="Calibri", size=9)
        ws.cell(row=r+3, column=1, value="AviaSolutions analysis").font = Font(name="Calibri", size=9, italic=True)
        
        return ws
    
    def write_audit_trail(self):
        """Write the full audit trail."""
        ws = self.wb.create_sheet("Audit Trail")
        ws.cell(row=1, column=1, value="Forecast Audit Trail").font = self.TITLE_FONT
        ws.column_dimensions['A'].width = 100
        
        for i, line in enumerate(self.engine.audit, 2):
            ws.cell(row=i, column=1, value=line).font = Font(name="Consolas", size=9)
    
    def write_schedule_table(self):
        """Write the schedule options table."""
        ws = self.wb.create_sheet("Scheduled TABLE")
        cfg = self.engine.config
        
        ws.cell(row=1, column=1, value=f"Schedule Options: {cfg.aircraft_type}").font = self.TITLE_FONT
        
        headers = ["Sector", "Dep. Time", "Arr. Time", "Op. Days", "Aircraft",
                   "Seats", "Annual Seats", "Annual Pax", "PDEW", "Seat Factor"]
        self._write_header_row(ws, 3, headers)
        self._set_col_widths(ws, [14, 10, 10, 10, 12, 8, 14, 14, 10, 12])
        
        # Outbound
        r = 4
        ws.cell(row=r, column=1, value=f"{cfg.home_airport_code}-{cfg.dest_airport_code}").font = self.DATA_FONT
        if cfg.outbound_dep:
            ws.cell(row=r, column=2, value=cfg.outbound_dep.strftime("%H:%M")).font = self.DATA_FONT
        if cfg.outbound_arr:
            ws.cell(row=r, column=3, value=cfg.outbound_arr.strftime("%H:%M")).font = self.DATA_FONT
        ws.cell(row=r, column=4, value=cfg.op_days).font = self.DATA_FONT
        ws.cell(row=r, column=5, value=cfg.aircraft_type).font = self.DATA_FONT
        ws.cell(row=r, column=6, value=cfg.seats).font = self.DATA_FONT
        ws.cell(row=r, column=7, value=self.engine.annual_capacity).font = self.DATA_FONT
        ws.cell(row=r, column=7).number_format = '#,##0'
        ws.cell(row=r, column=8, value=self.engine.grand_total).font = self.DATA_FONT
        ws.cell(row=r, column=8).number_format = '#,##0'
        ws.cell(row=r, column=9, value=self.engine.ppdew).font = self.DATA_FONT
        ws.cell(row=r, column=9).number_format = '#,##0.0'
        ws.cell(row=r, column=10, value=self.engine.load_factor).font = self.DATA_FONT
        ws.cell(row=r, column=10).number_format = '0.0%'
        
        # Return
        r = 5
        ws.cell(row=r, column=1, value=f"{cfg.dest_airport_code}-{cfg.home_airport_code}").font = self.DATA_FONT
        if cfg.return_dep:
            ws.cell(row=r, column=2, value=cfg.return_dep.strftime("%H:%M")).font = self.DATA_FONT
        if cfg.return_arr:
            ws.cell(row=r, column=3, value=cfg.return_arr.strftime("%H:%M")).font = self.DATA_FONT
        ws.cell(row=r, column=4, value=cfg.op_days).font = self.DATA_FONT
        ws.cell(row=r, column=5, value=cfg.aircraft_type).font = self.DATA_FONT
        ws.cell(row=r, column=6, value=cfg.seats).font = self.DATA_FONT
        
        r = 7
        ws.cell(row=r, column=1, 
                value="Note: AviaSolutions analysis. Source for aircraft configuration is airline website").font = \
            Font(name="Calibri", size=9, italic=True)
        
        return ws
    
    def save(self, output_path):
        """Generate all sheets and save."""
        self.write_forecast_table()
        self.write_connecting_table(self.engine.connecting_home, 'home', 
                                     f"Cnx at {self.engine.config.home_airport_code}")
        self.write_connecting_table(self.engine.connecting_dest, 'dest',
                                     f"Cnx at {self.engine.config.dest_airport_code}")
        self.write_schedule_table()
        self.write_audit_trail()
        self.wb.save(output_path)
        return output_path


# ============================================================================
# PIPELINE ORCHESTRATOR
# ============================================================================

class Pipeline:
    """
    Orchestrates the full forecast pipeline.
    
    For now, this uses pre-computed QSI scores from the forecast files.
    In the full implementation, it would:
    1. Call OAG Parser to parse raw schedules
    2. Call Connection Builder to enumerate itineraries
    3. Call QSI Scorer to compute capture rates
    4. Optionally call Time Optimiser
    5. Then assemble the forecast
    
    This version validates the assembly logic by confirming that
    when fed the same inputs as the Excel model, it produces the
    same 129,162 pax output.
    """
    
    def __init__(self, project_dir):
        self.project_dir = project_dir
        self.config = None
        self.engine = None
        self.issues = []
    
    def run_ba_lhr_sjc_validation(self):
        """Run the full pipeline on BA LHR-SJC and validate against known output."""
        print("=" * 70)
        print("MODULE VI: ASSEMBLY LOOP  BA LHR-SJC VALIDATION")
        print("=" * 70)
        
        # 1. Configure route
        print("\n[1] Configuring route: BA LHR-SJC...")
        self.config = RouteConfig.from_ba_lhr_sjc()
        
        # 2. Load P2P demand
        print("[2] Loading P2P demand segments...")
        p2p_segments = DemandLoader.load_ba_lhr_sjc_p2p(self.project_dir)
        
        # 3. Load connecting demand + QSI scores
        print("[3] Loading connecting demand at LHR...")
        connecting_home = DemandLoader.load_ba_lhr_sjc_connecting(
            self.project_dir, direction='home')
        print(f"    Loaded {len(connecting_home)} connecting cities at LHR")
        
        print("[4] Loading connecting demand at SJC...")
        connecting_dest = DemandLoader.load_ba_lhr_sjc_connecting(
            self.project_dir, direction='dest')
        print(f"    Loaded {len(connecting_dest)} connecting cities at SJC")
        
        # 4. Run forecast engine
        print("[5] Running forecast engine...")
        self.engine = ForecastEngine(self.config)
        self.engine.load_demand(p2p_segments, connecting_home, connecting_dest)
        total = self.engine.run()
        
        # 5. Reasonableness checks
        print("[6] Running reasonableness checks...")
        self.issues = self.engine.reasonableness_check()
        
        # 6. Validate against target
        print("\n" + "=" * 70)
        print("VALIDATION RESULTS")
        print("=" * 70)
        
        target_total = 129162
        target_p2p = 78110  # UK 51,462 + US 26,648
        target_cnx_lhr = 48115
        target_cnx_sjc = 2937
        target_lf = 0.829
        
        def validate(label, actual, target, tolerance=0.005):
            pct_diff = abs(actual - target) / target if target > 0 else 0
            status = "" if pct_diff < tolerance else ""
            print(f"  {status} {label:30s}: {actual:>10,.0f}  (target: {target:>10,.0f}  diff: {pct_diff:.2%})")
            return pct_diff < tolerance
        
        all_pass = True
        all_pass &= validate("P2P Total", self.engine.total_p2p, target_p2p)
        all_pass &= validate("Connecting at LHR", self.engine.total_cnx_home, target_cnx_lhr)
        all_pass &= validate("Connecting at SJC", self.engine.total_cnx_dest, target_cnx_sjc)
        all_pass &= validate("Grand Total", self.engine.grand_total, target_total)
        
        print(f"\n  Load Factor: {self.engine.load_factor:.1%} (target: {target_lf:.1%})")
        print(f"  PPDEW: {self.engine.ppdew:.1f}")
        
        if all_pass:
            print(f"\n   ALL VALIDATION TARGETS MET ")
        else:
            print(f"\n   SOME TARGETS MISSED  investigating...")
        
        if self.issues:
            print(f"\n  Reasonableness issues:")
            for issue in self.issues:
                print(f"     {issue}")
        
        return all_pass, total
    
    def generate_output(self, output_path):
        """Generate the output Excel workbook."""
        if not self.engine:
            raise ValueError("Must run pipeline first")
        
        writer = OutputWriter(self.engine)
        writer.save(output_path)
        print(f"\n  Output saved: {output_path}")
        return output_path


# ============================================================================
# MAIN
# ============================================================================

def main():
    project_dir = '/mnt/project'
    output_path = '/mnt/user-data/outputs/BA_LHR_SJC_Forecast_Validated.xlsx'
    audit_path = '/mnt/user-data/outputs/BA_LHR_SJC_Audit_Trail.txt'
    
    # Run validation
    pipeline = Pipeline(project_dir)
    passed, total = pipeline.run_ba_lhr_sjc_validation()
    
    # Generate output
    pipeline.generate_output(output_path)
    
    # Save audit trail as text
    with open(audit_path, 'w') as f:
        for line in pipeline.engine.audit:
            f.write(line + '\n')
    print(f"  Audit trail saved: {audit_path}")
    
    return passed


if __name__ == '__main__':
    main()
