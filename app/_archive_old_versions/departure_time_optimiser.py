#!/usr/bin/env python3
"""
Module V: Departure Time Optimiser
Avia Solutions QSI Pipeline

Compares QSI model outputs across different departure times to identify
the schedule that maximises connecting traffic capture (and therefore
total passengers / load factor).

This module operates in two modes:

MODE 1  File Comparison (current implementation):
  Reads pre-computed QSI files at multiple departure times and compares
  their capture rates. This matches the current manual workflow where
  Jonathan runs the QSI model at different times and compares results.

MODE 2  Grid Search (future, requires Modules III+IV integration):
  Treats departure time as a continuous variable, re-running the
  Connection Builder and QSI Scorer at each candidate time across a
  30-minute coarse grid, then 5-minute fine grid around the top 3.

Validated against BA LHR-SJC case study:
  - Original time (15:30/21:30): 129,162 pax, 82.9% LF
  - Optimised time (17:00/22:00): 139,302 pax, 89.4% LF
  - Delta: +10,140 pax (+7.9%), confirmed 17:00 is optimal

Author: Avia Solutions / QSI Automation Project
"""

import sys
import os
import argparse
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl")
    sys.exit(1)


# 
# DATA STRUCTURES
# 

@dataclass
class TimeScenario:
    """A single departure time scenario with its QSI results."""
    label: str                    # e.g. "17:00 LHR dep / 22:00 SJC dep"
    dep_time_origin: str          # Departure from origin (e.g. "2200" for SJC)
    arr_time_hub: str             # Arrival at hub (e.g. "1625" for LHR)
    dep_time_hub: str             # Departure from hub (e.g. "1700" for LHR)
    arr_time_origin: str          # Arrival at origin (e.g. "2000" for SJC)
    qsi_file_dir1: str = ""       # QSI file for direction 1 (beyond hub)
    qsi_file_dir2: str = ""       # QSI file for direction 2 (beyond origin)
    city_captures_dir1: Dict[str, float] = field(default_factory=dict)
    city_captures_dir2: Dict[str, float] = field(default_factory=dict)
    total_capture_dir1: float = 0.0
    total_capture_dir2: float = 0.0
    n_cities_dir1: int = 0
    n_cities_dir2: int = 0


@dataclass
class OptimiserResult:
    """Complete optimisation output."""
    scenarios: List[TimeScenario] = field(default_factory=list)
    best_scenario: Optional[TimeScenario] = None
    origin: str = ""
    hub: str = ""
    airline: str = ""
    sensitivity_band: Dict = field(default_factory=dict)


# 
# QSI FILE READER
# 

def parse_time_hhmm(val):
    """Parse HHMM time to minutes since midnight."""
    if val is None:
        return None
    s = str(int(float(val))).zfill(4)
    return int(s[:2]) * 60 + int(s[2:])


def format_time(hhmm_str):
    """Format HHMM string as HH:MM."""
    if not hhmm_str:
        return "?"
    s = str(hhmm_str).zfill(4)
    return f"{s[:2]}:{s[2:]}"


def read_qsi_calc(filepath):
    """
    Read QSI Calc sheet and extract per-city adjusted average fair shares.
    
    Returns:
        dict: {city_code: adjusted_avg_fair_share}
        Also returns metadata dict with origin, hub, airline codes.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['QSI Calc']
    rows = list(ws.iter_rows(values_only=True))
    
    # Extract metadata from header rows
    metadata = {}
    for row in rows[:10]:
        if row and row[1]:
            key = str(row[1]).strip()
            val = row[2] if len(row) > 2 else None
            if 'Origin' in key:
                metadata['origin'] = str(val).strip() if val else ''
            elif 'Hub' in key:
                metadata['hub'] = str(val).strip() if val else ''
            elif 'Airline' in key:
                metadata['airline'] = str(val).strip() if val else ''
    
    # Find header row (contains 'Dest')
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(v and 'dest' in str(v).lower() for v in row[:5]):
            header_idx = i
            break
    
    if header_idx is None:
        wb.close()
        raise ValueError(f"Could not find header row in QSI Calc: {filepath}")
    
    # Read city-level data
    # Columns: [blank, Dest, Route Label, QSI1, Total QSI1, Fair Share1,
    #           QSI2, Total QSI2, Fair Share2, Avg Fair Share, RT Check, Adj Avg Share]
    # NOTE: Cities appear MULTIPLE times in QSI Calc (one row per hub routing).
    # We take the MAXIMUM non-zero value per city across all routings.
    cities = {}
    for row in rows[header_idx + 1:]:
        if not row or not row[1]:
            continue
        city = str(row[1]).strip()
        if not city:
            continue
        
        # Column 9 = Average Fair Share, Column 11 = Adjusted Average Share
        avg_share = row[9] if len(row) > 9 and isinstance(row[9], (int, float)) else None
        adj_share = row[11] if len(row) > 11 and isinstance(row[11], (int, float)) else None
        
        # Use adj_share if available, else avg_share, else skip this row
        value = adj_share if adj_share is not None else avg_share
        if value is None:
            continue
        
        rt_check = str(row[10]).strip() if len(row) > 10 and row[10] else 'Include'
        
        # Only include cities that pass roundtrip check
        if rt_check == 'Exclude':
            continue
        
        # Keep the maximum value across all routings for this city
        if city not in cities or value > cities[city]:
            cities[city] = value
    
    wb.close()
    return cities, metadata


def read_service_times(filepath):
    """
    Extract the proposed service departure/arrival times from Leg 1.1.
    Returns dict with dep_time, arr_time for the proposed carrier.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    
    # Read metadata to get airline code
    ws_calc = wb['QSI Calc']
    calc_rows = list(ws_calc.iter_rows(max_row=10, values_only=True))
    airline = ''
    origin = ''
    hub = ''
    for row in calc_rows:
        if row and row[1]:
            if 'Airline' in str(row[1]):
                airline = str(row[2]).strip() if row[2] else ''
            elif 'Origin' in str(row[1]):
                origin = str(row[2]).strip() if row[2] else ''
            elif 'Hub' in str(row[1]):
                hub = str(row[2]).strip() if row[2] else ''
    
    times = {}
    
    # Read Leg 1.1 (arrivals at hub) - find our carrier departing from origin
    ws = wb['Leg 1.1']
    rows = list(ws.iter_rows(values_only=True))
    header = None
    for i, row in enumerate(rows):
        if row and row[0] and 'carrier' in str(row[0]).lower():
            header = [str(v).lower().strip() if v else '' for v in row]
            col_map = {h: j for j, h in enumerate(header)}
            
            carrier_col = col_map.get('carrier1', col_map.get('carrier', None))
            dep_col = col_map.get('depairport', None)
            dep_time_col = col_map.get('localdeptime', None)
            arr_time_col = col_map.get('localarrtime', None)
            
            for drow in rows[i+1:]:
                if not drow or not drow[0]:
                    continue
                c = str(drow[carrier_col]) if carrier_col is not None else ''
                d = str(drow[dep_col]) if dep_col is not None else ''
                if c == airline and d == origin:
                    times['dep_time_origin'] = str(int(float(drow[dep_time_col]))) if dep_time_col and drow[dep_time_col] else ''
                    times['arr_time_hub'] = str(int(float(drow[arr_time_col]))) if arr_time_col and drow[arr_time_col] else ''
                    break
            break
    
    # Read Leg 1.2 (arrivals at origin from hub direction) to get return times
    if 'Leg 1.2' in wb.sheetnames:
        ws2 = wb['Leg 1.2']
        rows2 = list(ws2.iter_rows(values_only=True))
        for i, row in enumerate(rows2):
            if row and row[0] and 'carrier' in str(row[0]).lower():
                header2 = [str(v).lower().strip() if v else '' for v in row]
                col_map2 = {h: j for j, h in enumerate(header2)}
                
                carrier_col2 = col_map2.get('carrier1', col_map2.get('carrier', None))
                dep_col2 = col_map2.get('depairport', None)
                dep_time_col2 = col_map2.get('localdeptime', None)
                arr_time_col2 = col_map2.get('localarrtime', None)
                
                for drow in rows2[i+1:]:
                    if not drow or not drow[0]:
                        continue
                    c = str(drow[carrier_col2]) if carrier_col2 is not None else ''
                    d = str(drow[dep_col2]) if dep_col2 is not None else ''
                    if c == airline and d == hub:
                        times['dep_time_hub'] = str(int(float(drow[dep_time_col2]))) if dep_time_col2 and drow[dep_time_col2] else ''
                        times['arr_time_origin'] = str(int(float(drow[arr_time_col2]))) if arr_time_col2 and drow[arr_time_col2] else ''
                        break
                break
    
    wb.close()
    return times, airline, origin, hub


# 
# SCENARIO BUILDER
# 

def build_scenario_from_files(label, dir1_file, dir2_file=None):
    """
    Build a TimeScenario from one or two QSI files.
    
    dir1_file: QSI file for direction 1 (e.g., QSILHR  beyond hub)
    dir2_file: QSI file for direction 2 (e.g., QSISJC  beyond origin)
    """
    scenario = TimeScenario(label=label, dep_time_origin='', arr_time_hub='',
                            dep_time_hub='', arr_time_origin='')
    
    # Read direction 1
    cities1, meta1 = read_qsi_calc(dir1_file)
    times1, airline, origin, hub = read_service_times(dir1_file)
    
    scenario.qsi_file_dir1 = dir1_file
    scenario.city_captures_dir1 = cities1
    scenario.total_capture_dir1 = sum(cities1.values())
    scenario.n_cities_dir1 = len(cities1)
    
    # Apply service times
    scenario.dep_time_origin = times1.get('dep_time_origin', '')
    scenario.arr_time_hub = times1.get('arr_time_hub', '')
    scenario.dep_time_hub = times1.get('dep_time_hub', '')
    scenario.arr_time_origin = times1.get('arr_time_origin', '')
    
    # Read direction 2 if provided
    if dir2_file:
        cities2, meta2 = read_qsi_calc(dir2_file)
        times2, _, _, _ = read_service_times(dir2_file)
        scenario.qsi_file_dir2 = dir2_file
        scenario.city_captures_dir2 = cities2
        scenario.total_capture_dir2 = sum(cities2.values())
        scenario.n_cities_dir2 = len(cities2)
        
        # Dir2 file is from opposite perspective (origin=hub, hub=origin)
        # so its dep_time_origin = hub departure, arr_time_hub = origin arrival
        if not scenario.dep_time_hub and times2.get('dep_time_origin'):
            scenario.dep_time_hub = times2['dep_time_origin']
        if not scenario.arr_time_origin and times2.get('arr_time_hub'):
            scenario.arr_time_origin = times2['arr_time_hub']
    
    return scenario, origin, hub, airline


# 
# OPTIMISER
# 

class DepartureTimeOptimiser:
    """
    Compares multiple departure time scenarios and identifies the optimal schedule.
    """
    
    def __init__(self):
        self.scenarios: List[TimeScenario] = []
        self.origin = ''
        self.hub = ''
        self.airline = ''
    
    def add_scenario(self, label, dir1_file, dir2_file=None):
        """Add a departure time scenario from QSI files."""
        scenario, origin, hub, airline = build_scenario_from_files(label, dir1_file, dir2_file)
        self.scenarios.append(scenario)
        if origin:
            self.origin = origin
        if hub:
            self.hub = hub
        if airline:
            self.airline = airline
        return scenario
    
    def run(self) -> OptimiserResult:
        """
        Compare all scenarios and identify the best.
        
        "Best" is defined as highest total capture rate (sum of adjusted
        fair shares across all cities in both directions). This is a proxy
        for total connecting passengers  higher total capture means more
        connecting pax at every demand level.
        """
        if not self.scenarios:
            raise ValueError("No scenarios loaded. Add scenarios first.")
        
        # Calculate combined scores
        for s in self.scenarios:
            s._combined_capture = s.total_capture_dir1 + s.total_capture_dir2
        
        # Sort by combined capture
        ranked = sorted(self.scenarios, key=lambda s: s._combined_capture, reverse=True)
        best = ranked[0]
        
        # Calculate sensitivity band
        sensitivity = {}
        if len(ranked) >= 2:
            best_score = best._combined_capture
            for s in ranked:
                diff = s._combined_capture - best_score
                pct = (diff / best_score * 100) if best_score else 0
                sensitivity[s.label] = {
                    'total_capture': s._combined_capture,
                    'dir1_capture': s.total_capture_dir1,
                    'dir2_capture': s.total_capture_dir2,
                    'delta_vs_best': diff,
                    'pct_vs_best': pct,
                    'dep_origin': format_time(s.dep_time_origin),
                    'dep_hub': format_time(s.dep_time_hub),
                }
        
        result = OptimiserResult(
            scenarios=ranked,
            best_scenario=best,
            origin=self.origin,
            hub=self.hub,
            airline=self.airline,
            sensitivity_band=sensitivity,
        )
        
        return result
    
    def city_comparison(self, top_n=30) -> List[Dict]:
        """
        Compare capture rates for each city across all scenarios.
        Returns list of dicts showing which cities gain/lose at each time.
        """
        all_cities = set()
        for s in self.scenarios:
            all_cities.update(s.city_captures_dir1.keys())
        
        comparison = []
        for city in sorted(all_cities):
            row = {'city': city}
            captures = {}
            for s in self.scenarios:
                cap = s.city_captures_dir1.get(city, 0)
                captures[s.label] = cap
                row[s.label] = cap
            
            # Which scenario is best for this city?
            if captures:
                best_label = max(captures, key=lambda l: captures[l])
                worst_label = min(captures, key=lambda l: captures[l])
                best_val = captures[best_label]
                worst_val = captures[worst_label]
                row['best_scenario'] = best_label
                row['range'] = best_val - worst_val
                row['range_pct'] = ((best_val - worst_val) / worst_val * 100) if worst_val > 0 else 0
            
            comparison.append(row)
        
        # Sort by range (biggest sensitivity first)
        comparison.sort(key=lambda r: -r.get('range', 0))
        return comparison[:top_n] if top_n else comparison


# 
# EXCEL OUTPUT
# 

HEADER_FONT = Font(name='Arial', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='003366')
DATA_FONT = Font(name='Arial', size=10)
GOOD_FILL = PatternFill('solid', fgColor='C6EFCE')
BAD_FILL = PatternFill('solid', fgColor='FFC7CE')
NEUTRAL_FILL = PatternFill('solid', fgColor='FFEB9C')
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)


def write_output(result: OptimiserResult, city_comparison: List[Dict], output_path: str):
    """Write optimiser results to Excel."""
    wb = openpyxl.Workbook()
    
    #  Sheet 1: Summary 
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_properties.tabColor = '003366'
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Departure Time Optimiser  {result.airline} {result.hub}{result.origin}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='003366')
    
    ws['A3'] = 'Route:'
    ws['B3'] = f'{result.hub}  {result.origin}'
    ws['A4'] = 'Airline:'
    ws['B4'] = result.airline
    ws['A5'] = 'Scenarios:'
    ws['B5'] = len(result.scenarios)
    for r in range(3, 6):
        ws[f'A{r}'].font = Font(name='Arial', bold=True, size=10)
        ws[f'B{r}'].font = DATA_FONT
    
    # Scenario comparison table
    row = 8
    headers = ['Rank', 'Scenario', f'Dep {result.origin}', f'Arr {result.hub}',
               f'Dep {result.hub}', f'Arr {result.origin}',
               f'Capture Dir1 ({result.hub})', f'Capture Dir2 ({result.origin})',
               'Total Capture', ' vs Best', ' %']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    for rank, s in enumerate(result.scenarios, 1):
        row += 1
        combined = s._combined_capture
        best_combined = result.best_scenario._combined_capture
        delta = combined - best_combined
        pct = (delta / best_combined * 100) if best_combined else 0
        
        vals = [rank, s.label,
                format_time(s.dep_time_origin), format_time(s.arr_time_hub),
                format_time(s.dep_time_hub), format_time(s.arr_time_origin),
                s.total_capture_dir1, s.total_capture_dir2,
                combined, delta, pct]
        
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if c in (7, 8, 9):
                cell.number_format = '0.0000'
            elif c == 10:
                cell.number_format = '+0.0000;-0.0000;0'
            elif c == 11:
                cell.number_format = '+0.0%;-0.0%;0%'
        
        # Highlight best
        if rank == 1:
            for c in range(1, len(vals) + 1):
                ws.cell(row=row, column=c).fill = GOOD_FILL
    
    # Recommendations
    row += 3
    ws.cell(row=row, column=1, value='RECOMMENDATION').font = Font(name='Arial', bold=True, size=12, color='003366')
    row += 1
    best = result.best_scenario
    ws.cell(row=row, column=1, value=f'Optimal departure: {format_time(best.dep_time_hub)} from {result.hub}, '
            f'{format_time(best.dep_time_origin)} from {result.origin}').font = DATA_FONT
    
    if len(result.scenarios) >= 2:
        worst = result.scenarios[-1]
        total_diff = best._combined_capture - worst._combined_capture
        pct_diff = (total_diff / worst._combined_capture * 100) if worst._combined_capture else 0
        row += 1
        ws.cell(row=row, column=1,
                value=f'Capture rate advantage over worst scenario: +{total_diff:.4f} ({pct_diff:.1f}%)').font = DATA_FONT
    
    # Column widths
    for c, w in enumerate([8, 30, 12, 12, 12, 12, 14, 14, 14, 12, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    
    #  Sheet 2: City Comparison 
    ws2 = wb.create_sheet('City Comparison')
    ws2.sheet_properties.tabColor = '336699'
    
    ws2.merge_cells('A1:E1')
    ws2['A1'] = f'Per-City Capture Rate Comparison  Direction 1 (Beyond {result.hub})'
    ws2['A1'].font = Font(name='Arial', bold=True, size=12, color='003366')
    
    row = 3
    headers2 = ['City'] + [s.label for s in result.scenarios] + ['Best Scenario', 'Range', 'Range %']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    for comp in city_comparison:
        row += 1
        ws2.cell(row=row, column=1, value=comp['city']).font = DATA_FONT
        ws2.cell(row=row, column=1).border = THIN_BORDER
        
        col = 2
        for s in result.scenarios:
            val = comp.get(s.label, 0)
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.number_format = '0.000000'
            cell.border = THIN_BORDER
            col += 1
        
        ws2.cell(row=row, column=col, value=comp.get('best_scenario', '')).font = DATA_FONT
        ws2.cell(row=row, column=col).border = THIN_BORDER
        col += 1
        
        cell = ws2.cell(row=row, column=col, value=comp.get('range', 0))
        cell.font = DATA_FONT
        cell.number_format = '0.000000'
        cell.border = THIN_BORDER
        col += 1
        
        cell = ws2.cell(row=row, column=col, value=comp.get('range_pct', 0) / 100)
        cell.font = DATA_FONT
        cell.number_format = '0.0%'
        cell.border = THIN_BORDER
    
    for c in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 16
    ws2.column_dimensions['A'].width = 8
    
    #  Sheet 3: Methodology 
    ws3 = wb.create_sheet('Methodology')
    ws3.sheet_properties.tabColor = '666666'
    
    notes = [
        'DEPARTURE TIME OPTIMISER  METHODOLOGY',
        '',
        'Purpose:',
        'This module compares QSI model outputs at different departure times to identify',
        'the schedule that maximises BA\'s connecting traffic capture through the hub.',
        '',
        'How it works:',
        '1. For each candidate departure time, the QSI model is run separately',
        '   (currently manually, future: automated via Modules III+IV)',
        '2. Different departure times create different connection windows at the hub,',
        '   making some connecting itineraries viable and others not.',
        '3. This changes the elapsed time for each connection, which through the',
        '   steep decay curve, creates significant capture rate differences.',
        '4. The optimiser compares total capture rates across all scenarios.',
        '',
        'Metric: Total Adjusted Average Fair Share',
        'This is the sum of BA\'s adjusted average fair share across all connecting',
        'cities in both directions. Higher total capture = more connecting passengers',
        'at any given demand level.',
        '',
        'Sensitivity:',
        'MCT thresholds create step-changes  a 5-minute schedule shift can add or',
        'remove entire connection markets. The sensitivity band shows how much the',
        'forecast changes within 30 minutes of the optimum.',
        '',
        'Limitations:',
        '- Current mode compares pre-computed QSI files (requires manual model runs)',
        '- Future grid search mode will automate: 30-min coarse grid  5-min fine grid',
        '- Optimises for connecting capture only; P2P demand is time-independent',
        '- Does not account for crew scheduling, aircraft rotation, or curfew constraints',
        '- Return service timing is read from files; future mode will search 2D grid',
    ]
    
    for i, line in enumerate(notes, 1):
        cell = ws3.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(name='Arial', bold=True, size=12, color='003366')
        elif line.endswith(':') and not line.startswith(' '):
            cell.font = Font(name='Arial', bold=True, size=10)
        else:
            cell.font = Font(name='Arial', size=10)
    
    ws3.column_dimensions['A'].width = 80
    
    # Save
    wb.save(output_path)
    return output_path


# 
# VALIDATION
# 

def validate_against_ba_lhr_sjc():
    """
    Validate against the BA LHR-SJC case study.
    
    Known results from the project files:
    - Original (15:30 LHR dep / 21:30 SJC dep): Lower QSI capture
    - New time (17:00 LHR dep / 22:00 SJC dep): Higher QSI capture
    - The optimiser should rank 22:00 SJC above 21:30 SJC.
    
    Forecast outputs confirm:
    - Original: 129,162 pax at 82.9% LF
    - New time: 139,302 pax at 89.4% LF
    - Delta: +10,140 pax (+7.9%)
    """
    print("=" * 60)
    print("VALIDATION: BA LHR-SJC Departure Time Optimisation")
    print("=" * 60)
    
    project_dir = '/mnt/project'
    
    # LHR perspective files (Direction 1  beyond LHR)
    lhr_files = {
        'Original (21:30 SJC)': os.path.join(project_dir, 'QSILHR_v1_OS_JZ_original_time_17Feb15.xlsx'),
        'New time (22:00 SJC)': os.path.join(project_dir, 'QSILHR_v1_OS_JZ_new_time_05Mar15.xlsx'),
        '5pm dep (17:00 SJC)': os.path.join(project_dir, 'QSILHR_v1_OS_JZ_5pm_dep_SJC_10Jun15.xlsx'),
    }
    
    # SJC perspective files (Direction 2  beyond SJC)
    sjc_files = {
        'Original (21:30 SJC)': os.path.join(project_dir, 'QSISJC.xlsx'),
        'New time (22:00 SJC)': os.path.join(project_dir, 'QSISJC_v1_new_time.xlsx'),
        '5pm dep (17:00 SJC)': os.path.join(project_dir, 'QSISJC_v1_5pm_dep_SJC.xlsx'),
    }
    
    # Check all files exist
    for label in lhr_files:
        if not os.path.exists(lhr_files[label]):
            print(f"WARNING: Missing {lhr_files[label]}")
            return False
        if not os.path.exists(sjc_files[label]):
            print(f"WARNING: Missing {sjc_files[label]}")
            return False
    
    # Build optimiser
    opt = DepartureTimeOptimiser()
    
    for label in lhr_files:
        print(f"\nLoading: {label}")
        scenario = opt.add_scenario(label, lhr_files[label], sjc_files[label])
        print(f"  {opt.hub} direction: {scenario.n_cities_dir1} cities, "
              f"total capture = {scenario.total_capture_dir1:.4f}")
        print(f"  {opt.origin} direction: {scenario.n_cities_dir2} cities, "
              f"total capture = {scenario.total_capture_dir2:.4f}")
        print(f"  Combined: {scenario.total_capture_dir1 + scenario.total_capture_dir2:.4f}")
        print(f"  Times: dep {opt.origin}={format_time(scenario.dep_time_origin)}, "
              f"arr {opt.hub}={format_time(scenario.arr_time_hub)}, "
              f"dep {opt.hub}={format_time(scenario.dep_time_hub)}, "
              f"arr {opt.origin}={format_time(scenario.arr_time_origin)}")
    
    # Run optimisation
    result = opt.run()
    
    print(f"\n{'='*60}")
    print("RESULTS")
    print(f"{'='*60}")
    
    for rank, s in enumerate(result.scenarios, 1):
        combined = s._combined_capture
        delta = combined - result.best_scenario._combined_capture
        pct = (delta / result.best_scenario._combined_capture * 100) if result.best_scenario._combined_capture else 0
        star = '  BEST' if rank == 1 else ''
        print(f"\n  #{rank}: {s.label}{star}")
        print(f"      Dep {result.origin}: {format_time(s.dep_time_origin)}, "
              f"Dep {result.hub}: {format_time(s.dep_time_hub)}")
        print(f"      Dir1 ({result.hub}): {s.total_capture_dir1:.4f}")
        print(f"      Dir2 ({result.origin}): {s.total_capture_dir2:.4f}")
        print(f"      Combined: {combined:.4f} ({pct:+.2f}% vs best)")
    
    # City sensitivity comparison
    comparison = opt.city_comparison(top_n=15)
    
    print(f"\n{'='*60}")
    print("TOP 15 MOST TIME-SENSITIVE CITIES (Dir1  Beyond LHR)")
    print(f"{'='*60}")
    print(f"{'City':6s}  ", end='')
    for s in result.scenarios:
        short = s.label[:15]
        print(f"{short:>15s}  ", end='')
    print(f"{'Range':>8s}  {'Range%':>8s}")
    print("-" * (6 + 17 * len(result.scenarios) + 20))
    
    for comp in comparison:
        print(f"{comp['city']:6s}  ", end='')
        for s in result.scenarios:
            val = comp.get(s.label, 0)
            print(f"{val:15.6f}  ", end='')
        print(f"{comp.get('range', 0):8.6f}  {comp.get('range_pct', 0):7.1f}%")
    
    # Validation check
    print(f"\n{'='*60}")
    print("VALIDATION CHECK")
    print(f"{'='*60}")
    
    # The known result: the "New time (22:00)" or similar should beat "Original (21:30)"
    orig_scenario = None
    new_scenario = None
    for s in result.scenarios:
        if '21:30' in s.label or 'Original' in s.label:
            orig_scenario = s
        if '22:00' in s.label or 'New time' in s.label:
            new_scenario = s
    
    if orig_scenario and new_scenario:
        orig_combined = orig_scenario._combined_capture
        new_combined = new_scenario._combined_capture
        improvement = new_combined - orig_combined
        pct_improvement = (improvement / orig_combined * 100) if orig_combined else 0
        
        passed = new_combined > orig_combined
        print(f"\n  Original (21:30): {orig_combined:.4f}")
        print(f"  New time (22:00): {new_combined:.4f}")
        print(f"  Improvement: {improvement:+.4f} ({pct_improvement:+.1f}%)")
        print(f"\n  Expected: New time > Original time")
        print(f"  Result:   {' PASS' if passed else ' FAIL'}")
        
        # Cross-reference with known forecast outputs
        print(f"\n  Known forecast results:")
        print(f"    Original  129,162 pax at 82.9% LF")
        print(f"    New time  139,302 pax at 89.4% LF (unconstrained w/o India)")
        print(f"    Delta: +10,140 pax (+7.9%)")
    
    # Write output
    output_path = '/mnt/user-data/outputs/DepartureTimeOptimiser_BA_LHR_SJC.xlsx'
    write_output(result, comparison, output_path)
    print(f"\n  Output: {output_path}")
    
    return True


# 
# CLI
# 

def main():
    parser = argparse.ArgumentParser(
        description='Module V: Departure Time Optimiser  Avia Solutions QSI Pipeline',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Validate against BA LHR-SJC
  python departure_time_optimiser.py --validate

  # Compare two time scenarios
  python departure_time_optimiser.py \\
      --dir1 QSILHR_original.xlsx --dir2 QSISJC_original.xlsx --label "Original" \\
      --dir1 QSILHR_new_time.xlsx --dir2 QSISJC_new_time.xlsx --label "New Time"
        """
    )
    parser.add_argument('--validate', action='store_true',
                        help='Run validation against BA LHR-SJC case study')
    parser.add_argument('--dir1', action='append', default=[],
                        help='QSI file for Direction 1 (beyond hub). Repeat for each scenario.')
    parser.add_argument('--dir2', action='append', default=[],
                        help='QSI file for Direction 2 (beyond origin). Repeat for each scenario.')
    parser.add_argument('--label', action='append', default=[],
                        help='Label for each scenario. Repeat for each scenario.')
    parser.add_argument('-o', '--output', default=None,
                        help='Output file path')
    
    args = parser.parse_args()
    
    if args.validate:
        validate_against_ba_lhr_sjc()
        return
    
    if not args.dir1:
        parser.error("Provide --dir1 files (one per scenario) or use --validate")
    
    if len(args.dir1) < 2:
        parser.error("Need at least 2 scenarios to compare")
    
    # Pad labels and dir2
    while len(args.label) < len(args.dir1):
        args.label.append(f"Scenario {len(args.label) + 1}")
    while len(args.dir2) < len(args.dir1):
        args.dir2.append(None)
    
    opt = DepartureTimeOptimiser()
    for i in range(len(args.dir1)):
        print(f"Loading: {args.label[i]}")
        opt.add_scenario(args.label[i], args.dir1[i], args.dir2[i])
    
    result = opt.run()
    comparison = opt.city_comparison()
    
    output = args.output or '/mnt/user-data/outputs/DepartureTimeOptimiser_output.xlsx'
    write_output(result, comparison, output)
    
    print(f"\nBest: {result.best_scenario.label}")
    print(f"Output: {output}")


if __name__ == '__main__':
    main()
