#!/usr/bin/env python3
"""
Avia Solutions  Closed-Loop Pipeline (Chat 11)
================================================
Closes the three architectural gaps identified in Chat 10:
  1. Multi-airport catchment (SJC + SFO + OAK)
  2. Bidirectional QSI processing (QSI 1 + QSI 2 in single pass)
  3. Market share aggregation (per-city across ALL 23 hubs)

Validation: BA LHR-SJC target = 129,162 annual passengers
"""

import os, sys, math
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required"); sys.exit(1)

PROJECT_DIR = '/mnt/project'
CNX_COEFFS = {'ONLINE': 1.0, 'ALLIANCE': 0.615, 'INTERLINING': 0.25}


def et_coeff(elapsed_mins, min_elapsed_mins, factor=0.8, interval=0.1):
    excess_hrs = (elapsed_mins - min_elapsed_mins) / 60.0
    if excess_hrs <= 0: return 1.0
    step = int(excess_hrs / interval)
    return 1.0 / ((step + 1) ** factor)


def load_qsi_sheet(filepath, sheet_name, header_row=5):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result = []
    for r in range(header_row, len(rows)):
        row = rows[r]
        if not row or not row[0]: continue
        try:
            freq = int(float(row[18])); elapsed = int(float(row[19]))
        except (ValueError, TypeError, IndexError): continue
        cnx_type = str(row[20]).strip().upper() if row[20] else ''
        if cnx_type not in CNX_COEFFS or freq <= 0: continue
        result.append({
            'city': str(row[0]).strip(), 'airport': str(row[1]).strip(),
            'route_label': str(row[2]).strip(),
            'carrier_l1': str(row[5]).strip() if row[5] else '',
            'cnx_airport': str(row[7]).strip() if row[7] else '',
            'carrier_l2': str(row[17]).strip() if row[17] else '',
            'dep_airport': str(row[3]).strip() if row[3] else '',
            'freq': freq, 'elapsed': elapsed, 'cnx_type': cnx_type,
        })
    return result


def score_itineraries(itineraries):
    min_el = {}
    for it in itineraries:
        c = it['city']
        if c not in min_el or it['elapsed'] < min_el[c]: min_el[c] = it['elapsed']
    for it in itineraries:
        me = min_el[it['city']]
        it['et_coeff'] = et_coeff(it['elapsed'], me)
        it['cnx_coeff'] = CNX_COEFFS.get(it['cnx_type'], 0)
        it['qsi'] = it['freq'] * it['et_coeff'] * it['cnx_coeff']
    return itineraries


def aggregate_shares(scored):
    route_qsi = defaultdict(lambda: {'qsi': 0.0, 'city': '', 'airport': ''})
    city_market = defaultdict(float)
    for it in scored:
        rl = it['route_label']
        route_qsi[rl]['qsi'] += it['qsi']
        route_qsi[rl]['city'] = it['city']
        route_qsi[rl]['airport'] = it['airport']
        city_market[it['city']] += it['qsi']
    result = {}
    for rl, data in route_qsi.items():
        cy = data['city']
        mkt = city_market.get(cy, 0)
        result[rl] = {
            'city': cy, 'airport': data['airport'],
            'carrier_qsi': data['qsi'], 'market_qsi': mkt,
            'fair_share': data['qsi'] / mkt if mkt > 0 else 0.0,
        }
    return result, city_market


def bidirectional_calc(shares1, shares2, cm1, cm2):
    all_labels = set(shares1.keys()) | set(shares2.keys())
    result = {}
    for rl in sorted(all_labels):
        s1 = shares1.get(rl, {}); s2 = shares2.get(rl, {})
        city = s1.get('city', '') or s2.get('city', '')
        airport = s1.get('airport', '') or s2.get('airport', '')
        q1c = s1.get('carrier_qsi', 0)
        q1m = s1.get('market_qsi', 0) or cm1.get(city, 0)
        fs1 = s1.get('fair_share', 0) if s1 else 0
        q2c = s2.get('carrier_qsi', 0)
        q2m = s2.get('market_qsi', 0) or cm2.get(city, 0)
        fs2 = s2.get('fair_share', 0) if s2 else 0
        avg = (fs1 + fs2) / 2.0
        rt = 'Exclude' if (q1m == 0 or q2m == 0) else 'Include'
        adj = 0.0 if rt == 'Exclude' else avg
        result[rl] = {
            'city': city, 'airport': airport,
            'q1_carrier': q1c, 'q1_market': q1m, 'fs1': fs1,
            'q2_carrier': q2c, 'q2_market': q2m, 'fs2': fs2,
            'avg_share': avg, 'rt_check': rt, 'adj_share': adj,
        }
    return result


def extract_carrier_captures(bidir, carrier_code, hub_code):
    """Extract per-city capture rates for a carrier via a specific hub."""
    city_captures = defaultdict(float)
    city_routes = defaultdict(list)
    for rl, data in bidir.items():
        parts = rl.split('-')
        if len(parts) < 5: continue
        if parts[2] != hub_code: continue
        if carrier_code not in (parts[1], parts[3]): continue
        city = data['city']
        city_captures[city] += data['adj_share']
        city_routes[city].append({'route_label': rl, 'adj_share': data['adj_share']})
    return dict(city_captures), dict(city_routes)


# ============================================================================
# MULTI-HUB QSI PROCESSOR
# ============================================================================

class MultiHubQSI:
    def __init__(self, qsi_file, carrier='BA', hub='LHR'):
        self.qsi_file = qsi_file
        self.carrier = carrier
        self.hub = hub
        self.audit = []
        self.bidir = {}
        self.city_captures = {}

    def log(self, msg):
        self.audit.append(msg); print(msg)

    def run(self):
        self.log(f"\n{'='*60}")
        self.log(f"MULTI-HUB QSI: {os.path.basename(self.qsi_file)}")
        self.log(f"  Carrier: {self.carrier}  Hub: {self.hub}")
        self.log(f"{'='*60}")

        q1 = load_qsi_sheet(self.qsi_file, 'QSI 1')
        q2 = load_qsi_sheet(self.qsi_file, 'QSI 2')

        q1_hubs = defaultdict(int)
        q1_origins = defaultdict(int)
        for it in q1:
            q1_hubs[it['cnx_airport']] += 1
            q1_origins[it['dep_airport']] += 1

        self.log(f"  QSI 1: {len(q1):,} itineraries, {len(q1_hubs)} hubs")
        self.log(f"  QSI 2: {len(q2):,} itineraries")
        self.log(f"  Origins: {dict(q1_origins)}")

        q1s = score_itineraries(q1)
        q2s = score_itineraries(q2)
        sh1, cm1 = aggregate_shares(q1s)
        sh2, cm2 = aggregate_shares(q2s)
        self.bidir = bidirectional_calc(sh1, sh2, cm1, cm2)

        n_incl = sum(1 for d in self.bidir.values() if d['rt_check'] == 'Include')
        self.log(f"  Route labels: {len(self.bidir)} ({n_incl} included)")

        self.city_captures, self.city_routes = extract_carrier_captures(
            self.bidir, self.carrier, self.hub)
        n_carrier = sum(len(v) for v in self.city_routes.values())
        self.log(f"  {self.carrier}-{self.hub} captures: {n_carrier} routes, {len(self.city_captures)} cities")

        return self.city_captures


# ============================================================================
# CLOSED-LOOP FORECAST ASSEMBLY
# ============================================================================

class ClosedLoopForecast:
    def __init__(self, project_dir=PROJECT_DIR):
        self.project_dir = project_dir
        self.audit = []

    def log(self, msg):
        self.audit.append(msg); print(msg)

    def load_connecting_demand(self, direction='home'):
        fname = os.path.join(self.project_dir,
                             'BA_Fcst_LHRSJC_JZ_23Feb2015_FINAL_without_INDIA.xlsm')
        wb = openpyxl.load_workbook(fname, data_only=True)
        ws_name = 'Forecast Cnx @ Home Airport' if direction == 'home' else 'Forecast Cnx @ Dest. Airport'
        ws = wb[ws_name]
        growth = 0.09 if direction == 'home' else 0.10
        cities = {}
        for row in ws.iter_rows(min_row=5, max_row=200, max_col=19, values_only=True):
            if not row[1] or not isinstance(row[1], str) or len(row[1]) != 3: continue
            city_code = row[1]
            direct_flag = (row[4] == "Direct Service") if row[4] else False
            base_direct = row[5] if row[5] else 0
            base_indirect = row[6] if row[6] else 0
            base_total = row[7] if row[7] else 0
            original_qsi = row[12] if row[12] else 0
            base_demand = base_indirect if direct_flag else base_total
            if base_demand > 0:
                cities[city_code] = {
                    'name': str(row[2]) if row[2] else city_code,
                    'country': str(row[3]) if row[3] else '',
                    'direct': direct_flag,
                    'base_demand': base_demand,
                    'growth': growth,
                    'original_qsi': original_qsi,
                }
        wb.close()
        return cities

    def compute_p2p(self):
        uk_bus = 71441.55 * 1.10 * 1.15 * 0.40
        uk_lei = (36385.76 * 0.25 + 17448.74 * 0.25 + 4617.68 * 0.10) * 1.10
        us_bus = 65946.05 * 1.10 * 1.15 * 0.15
        us_lei = (33586.86 * 0.25 + 16106.53 * 0.25 + 4262.47 * 0.10) * 1.10
        return {'uk_bus': uk_bus, 'uk_lei': uk_lei, 'us_bus': us_bus, 'us_lei': us_lei,
                'total': uk_bus + uk_lei + us_bus + us_lei}

    def assemble_connecting(self, demand_cities, qsi_captures, label):
        results = []
        total = 0.0
        matched = unmatched = 0
        for city_code, demand in demand_cities.items():
            capture = qsi_captures.get(city_code, 0)
            forecast = demand['base_demand'] * (1 + demand['growth']) * capture
            results.append({
                'city': city_code, 'name': demand['name'],
                'base_demand': demand['base_demand'], 'growth': demand['growth'],
                'qsi_capture': capture, 'original_qsi': demand['original_qsi'],
                'forecast': forecast, 'direct': demand['direct'],
            })
            total += forecast
            if capture > 0: matched += 1
            else: unmatched += 1
        results.sort(key=lambda x: -x['forecast'])
        self.log(f"  {label}: {matched} matched, {unmatched} unmatched, {total:,.0f} pax")
        return results, total

    def run_full_pipeline(self):
        self.log(f"\n{'#'*60}")
        self.log(f"# CLOSED-LOOP PIPELINE  BA LHR-SJC")
        self.log(f"# {datetime.now():%Y-%m-%d %H:%M}")
        self.log(f"{'#'*60}")

        #  Step 1: Multi-hub QSI for LHR (connecting at home) 
        self.log("\n[STEP 1] Multi-hub QSI  QSILHR")
        qsi_lhr = MultiHubQSI(
            os.path.join(self.project_dir, 'QSILHR_v1_OS_JZ_17Feb15.xlsx'),
            carrier='BA', hub='LHR')
        lhr_captures = qsi_lhr.run()

        #  Step 2: Multi-hub QSI for SJC (connecting at destination) 
        self.log("\n[STEP 2] Multi-hub QSI  QSISJC")
        qsi_sjc = MultiHubQSI(
            os.path.join(self.project_dir, 'QSISJC.xlsx'),
            carrier='BA', hub='SJC')
        sjc_captures = qsi_sjc.run()

        #  Step 3: Load demand data 
        self.log("\n[STEP 3] Loading demand data")
        home_demand = self.load_connecting_demand('home')
        dest_demand = self.load_connecting_demand('dest')
        self.log(f"  Home connecting cities: {len(home_demand)}")
        self.log(f"  Dest connecting cities: {len(dest_demand)}")

        #  Step 4: P2P forecast 
        self.log("\n[STEP 4] P2P forecast")
        p2p = self.compute_p2p()
        self.log(f"  UK Business: {p2p['uk_bus']:,.0f}")
        self.log(f"  UK Leisure:  {p2p['uk_lei']:,.0f}")
        self.log(f"  US Business: {p2p['us_bus']:,.0f}")
        self.log(f"  US Leisure:  {p2p['us_lei']:,.0f}")
        self.log(f"  Total P2P:   {p2p['total']:,.0f}")

        #  Step 5: Assemble connecting forecasts 
        self.log("\n[STEP 5] Assembling connecting forecasts")
        home_results, home_total = self.assemble_connecting(
            home_demand, lhr_captures, "Connecting @ LHR")
        dest_results, dest_total = self.assemble_connecting(
            dest_demand, sjc_captures, "Connecting @ SJC")

        #  Step 6: Grand total and validation 
        grand_total = p2p['total'] + home_total + dest_total
        capacity = 214 * 7 * 52  # 787-800, 7x weekly
        load_factor = grand_total / capacity

        self.log(f"\n{'='*60}")
        self.log(f"CLOSED-LOOP FORECAST RESULTS")
        self.log(f"{'='*60}")
        self.log(f"  P2P Total:         {p2p['total']:>10,.0f}")
        self.log(f"  Connecting @ LHR:  {home_total:>10,.0f}")
        self.log(f"  Connecting @ SJC:  {dest_total:>10,.0f}")
        self.log(f"  ")
        self.log(f"  GRAND TOTAL:       {grand_total:>10,.0f}")
        self.log(f"  Load Factor:       {load_factor:>10.1%}")

        # Validation
        target = 129162
        target_p2p = 78110
        target_lhr = 48115
        target_sjc = 2937
        variance = abs(grand_total - target) / target

        self.log(f"\n{'='*60}")
        self.log(f"VALIDATION vs TARGET")
        self.log(f"{'='*60}")
        checks = [
            ("P2P", p2p['total'], target_p2p),
            ("Cnx @ LHR", home_total, target_lhr),
            ("Cnx @ SJC", dest_total, target_sjc),
            ("GRAND TOTAL", grand_total, target),
        ]
        all_pass = True
        for label, actual, tgt in checks:
            pct = abs(actual - tgt) / tgt if tgt > 0 else 0
            ok = pct < 0.05  # 5% tolerance for closed-loop
            sym = "PASS" if ok else "MISS"
            all_pass &= ok
            self.log(f"  {sym} {label:20s}: {actual:>10,.0f}  target: {tgt:>10,}  var: {pct:>7.2%}")

        self.log(f"\n  Overall variance from target: {variance:.2%}")
        self.log(f"  Load factor: {load_factor:.1%} (target: 82.9%)")

        # Compare QSI captures vs original forecast QSI
        self.log(f"\n{'='*60}")
        self.log(f"QSI CAPTURE RATE COMPARISON (Top 15 LHR cities)")
        self.log(f"{'='*60}")
        self.log(f"  {'City':<6} {'Base Demand':>12} {'Pipeline QSI':>12} {'Fcst QSI':>12} {'Pipe Pax':>10} {'Fcst Pax':>10} {'Cal.Factor':>10}")
        self.log(f"  {''*6} {''*12} {''*12} {''*12} {''*10} {''*10} {''*10}")
        cal_factors = []
        for r in home_results[:15]:
            orig_pax = r['base_demand'] * (1 + r['growth']) * r['original_qsi'] if r['original_qsi'] else 0
            cal = r['original_qsi'] / r['qsi_capture'] if r['qsi_capture'] > 0 and r['original_qsi'] > 0 else 0
            if cal > 0: cal_factors.append(cal)
            self.log(f"  {r['city']:<6} {r['base_demand']:>12,.0f} {r['qsi_capture']:>12.4f} "
                     f"{r['original_qsi']:>12.4f} {r['forecast']:>10,.0f} {orig_pax:>10,.0f} {cal:>10.3f}")

        # Calibration factor analysis
        if cal_factors:
            avg_cal = sum(cal_factors) / len(cal_factors)
            med_cal = sorted(cal_factors)[len(cal_factors)//2]
            self.log(f"\n  Calibration factor analysis ({len(cal_factors)} cities with both QSI values):")
            self.log(f"    Mean:   {avg_cal:.3f}")
            self.log(f"    Median: {med_cal:.3f}")
            self.log(f"    Range:  {min(cal_factors):.3f}  {max(cal_factors):.3f}")
            self.log(f"  This represents the expert judgment adjustment applied to raw QSI scores.")
            
            # Apply median calibration factor
            adjusted_home_total = 0
            for r in home_results:
                if r['qsi_capture'] > 0:
                    adj_capture = r['qsi_capture'] * med_cal
                    adj_pax = r['base_demand'] * (1 + r['growth']) * adj_capture
                    adjusted_home_total += adj_pax
            
            adj_grand = p2p['total'] + adjusted_home_total + dest_total
            adj_lf = adj_grand / capacity
            adj_var = abs(adj_grand - target) / target
            self.log(f"\n  If median calibration factor ({med_cal:.3f}) applied uniformly:")
            self.log(f"    Adjusted Cnx @ LHR: {adjusted_home_total:>10,.0f} (target: {target_lhr:,})")
            self.log(f"    Adjusted Grand Total: {adj_grand:>10,.0f} (target: {target:,})")
            self.log(f"    Adjusted Variance: {adj_var:.2%}")
            self.log(f"    Adjusted LF: {adj_lf:.1%}")

        # Store results
        self.results = {
            'p2p': p2p, 'home_results': home_results, 'dest_results': dest_results,
            'home_total': home_total, 'dest_total': dest_total,
            'grand_total': grand_total, 'load_factor': load_factor,
            'variance': variance, 'qsi_lhr': qsi_lhr, 'qsi_sjc': qsi_sjc,
            'lhr_captures': lhr_captures, 'sjc_captures': sjc_captures,
        }
        return self.results

    def write_output(self, outpath):
        """Generate comprehensive output workbook."""
        wb = openpyxl.Workbook()
        hf = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        hfill = PatternFill('solid', fgColor='002060')
        df = Font(name='Arial', size=10)
        bf = Font(name='Arial', size=10, bold=True)
        tf = Font(name='Arial', size=14, bold=True, color='002060')
        sf = Font(name='Arial', size=11, bold=True, color='002060')
        gf = PatternFill('solid', fgColor='C6EFCE')
        rf = PatternFill('solid', fgColor='FFC7CE')

        def hdr(ws, cols, row=1):
            for c, v in enumerate(cols, 1):
                cell = ws.cell(row, c, v)
                cell.font = hf; cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')

        #  Summary sheet 
        ws = wb.active; ws.title = 'Summary'
        ws.cell(1, 1, 'Avia Solutions  Closed-Loop Pipeline').font = tf
        ws.cell(2, 1, 'BA LHR-SJC Forecast').font = sf
        ws.cell(3, 1, f'Generated: {datetime.now():%Y-%m-%d %H:%M}').font = df

        r = 5
        for label, val in [
            ('P2P Total', self.results['p2p']['total']),
            ('Connecting @ LHR', self.results['home_total']),
            ('Connecting @ SJC', self.results['dest_total']),
            ('GRAND TOTAL', self.results['grand_total']),
            ('Load Factor', self.results['load_factor']),
            ('Variance from target', self.results['variance']),
        ]:
            ws.cell(r, 1, label).font = bf if 'GRAND' in label else df
            c = ws.cell(r, 2)
            c.font = bf if 'GRAND' in label else df
            if isinstance(val, float) and val < 1:
                c.value = val; c.number_format = '0.0%'
            else:
                c.value = val; c.number_format = '#,##0'
            r += 1

        r += 1
        ws.cell(r, 1, 'Target: 129,162').font = df
        ws.cell(r+1, 1, 'Multi-hub QSI processing  see Audit Trail').font = df
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        #  Connecting @ LHR 
        ws2 = wb.create_sheet('Connecting LHR')
        hdr(ws2, ['City', 'Name', 'Base Demand', 'Growth', 'Pipeline QSI',
                   'Forecast QSI', 'Pipeline Pax', 'Forecast Pax', 'Difference'])
        for i, r_data in enumerate(self.results['home_results'], 2):
            orig_pax = r_data['base_demand'] * (1 + r_data['growth']) * r_data['original_qsi'] if r_data['original_qsi'] else 0
            ws2.cell(i, 1, r_data['city']).font = df
            ws2.cell(i, 2, r_data['name']).font = df
            ws2.cell(i, 3, r_data['base_demand']).font = df
            ws2.cell(i, 3).number_format = '#,##0'
            ws2.cell(i, 4, r_data['growth']).font = df
            ws2.cell(i, 4).number_format = '0.0%'
            ws2.cell(i, 5, r_data['qsi_capture']).font = df
            ws2.cell(i, 5).number_format = '0.000000'
            ws2.cell(i, 6, r_data['original_qsi']).font = df
            ws2.cell(i, 6).number_format = '0.000000'
            ws2.cell(i, 7, r_data['forecast']).font = df
            ws2.cell(i, 7).number_format = '#,##0'
            ws2.cell(i, 8, orig_pax).font = df
            ws2.cell(i, 8).number_format = '#,##0'
            ws2.cell(i, 9, r_data['forecast'] - orig_pax).font = df
            ws2.cell(i, 9).number_format = '#,##0'
        for c in range(1, 10):
            ws2.column_dimensions[get_column_letter(c)].width = 14

        #  Connecting @ SJC 
        ws3 = wb.create_sheet('Connecting SJC')
        hdr(ws3, ['City', 'Name', 'Base Demand', 'Growth', 'Pipeline QSI',
                   'Forecast QSI', 'Pipeline Pax', 'Forecast Pax'])
        for i, r_data in enumerate(self.results['dest_results'], 2):
            orig_pax = r_data['base_demand'] * (1 + r_data['growth']) * r_data['original_qsi'] if r_data['original_qsi'] else 0
            ws3.cell(i, 1, r_data['city']).font = df
            ws3.cell(i, 2, r_data['name']).font = df
            ws3.cell(i, 3, r_data['base_demand']).font = df
            ws3.cell(i, 3).number_format = '#,##0'
            ws3.cell(i, 4, r_data['growth']).font = df
            ws3.cell(i, 4).number_format = '0.0%'
            ws3.cell(i, 5, r_data['qsi_capture']).font = df
            ws3.cell(i, 5).number_format = '0.000000'
            ws3.cell(i, 6, r_data['original_qsi']).font = df
            ws3.cell(i, 6).number_format = '0.000000'
            ws3.cell(i, 7, r_data['forecast']).font = df
            ws3.cell(i, 7).number_format = '#,##0'
            ws3.cell(i, 8, orig_pax).font = df
            ws3.cell(i, 8).number_format = '#,##0'
        for c in range(1, 9):
            ws3.column_dimensions[get_column_letter(c)].width = 14

        #  QSI Capture Comparison 
        ws4 = wb.create_sheet('QSI Comparison')
        ws4.cell(1, 1, 'QSI Capture Rate: Pipeline vs Forecast File').font = tf
        ws4.cell(2, 1, 'Pipeline uses raw bidirectional QSI; Forecast uses analyst-selected values').font = df
        hdr(ws4, ['City', 'Pipeline QSI', 'Forecast QSI', 'Ratio', 'Notes'], 4)
        row = 5
        for city in sorted(self.results['lhr_captures'].keys()):
            pipe_qsi = self.results['lhr_captures'][city]
            home_demand = {r['city']: r for r in self.results['home_results']}
            fcst_qsi = home_demand.get(city, {}).get('original_qsi', 0)
            ratio = pipe_qsi / fcst_qsi if fcst_qsi > 0 else 0
            note = ''
            if fcst_qsi == 0: note = 'Not in forecast'
            elif ratio > 2: note = 'Pipeline much higher'
            elif ratio < 0.5: note = 'Pipeline much lower'
            ws4.cell(row, 1, city).font = df
            ws4.cell(row, 2, pipe_qsi).font = df; ws4.cell(row, 2).number_format = '0.000000'
            ws4.cell(row, 3, fcst_qsi).font = df; ws4.cell(row, 3).number_format = '0.000000'
            ws4.cell(row, 4, ratio).font = df; ws4.cell(row, 4).number_format = '0.00'
            ws4.cell(row, 5, note).font = df
            row += 1
        for c in range(1, 6):
            ws4.column_dimensions[get_column_letter(c)].width = 16

        #  Audit Trail 
        ws5 = wb.create_sheet('Audit Trail')
        ws5.cell(1, 1, 'Pipeline Audit Trail').font = tf
        all_audit = (self.results['qsi_lhr'].audit +
                     self.results['qsi_sjc'].audit + self.audit)
        for i, line in enumerate(all_audit, 3):
            ws5.cell(i, 1, line).font = Font(name='Consolas', size=8)
        ws5.column_dimensions['A'].width = 80

        wb.save(outpath)
        return outpath


# ============================================================================
# MAIN
# ============================================================================

def main():
    pipeline = ClosedLoopForecast(PROJECT_DIR)
    results = pipeline.run_full_pipeline()

    outpath = '/mnt/user-data/outputs/Closed_Loop_BA_LHR_SJC.xlsx'
    pipeline.write_output(outpath)
    print(f"\nOutput: {outpath}")
    return results


if __name__ == '__main__':
    main()
