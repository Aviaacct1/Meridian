#!/usr/bin/env python3
"""
Module IV: QSI Scorer
Avia Solutions  QSI Model Automation Pipeline

KEY INSIGHT: All grouping (min_elapsed, market_qsi) uses CITY CODE not AIRPORT CODE.
Cities like BFS have two airports (BHD, BFS), EAP has two (BSL, MLH).
The Excel SUMIF groups by city-level market totals.

Formula: QSI = frequency * ET_coeff * CnxType_coeff
  ET_coeff = 1 / (step+1)^factor, step = floor(excess_hrs / interval)
  CnxType: ONLINE=1.0, ALLIANCE=0.615, INTERLINING=0.25
  Market QSI = sum ALL itinerary QSI for that CITY
  Carrier QSI = sum itinerary QSI for that route_label (SUMIF match)
  Fair Share = Carrier QSI / Market QSI
  Bidirectional = AVERAGE(FS1, FS2), exclude if either market = 0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

CNX_COEFFS = {'ONLINE': 1.0, 'ALLIANCE': 0.615, 'INTERLINING': 0.25}
DEFAULT_FACTOR = 0.8
DEFAULT_INTERVAL = 0.1  # hours


def et_coeff(elapsed_mins, min_elapsed_mins, factor=DEFAULT_FACTOR, interval=DEFAULT_INTERVAL):
    excess_hrs = (elapsed_mins - min_elapsed_mins) / 60.0
    if excess_hrs <= 0:
        return 1.0
    step = int(excess_hrs / interval)
    return 1.0 / ((step + 1) ** factor)


def load_qsi_sheet(filepath, sheet_name, header_row=5):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result = []
    for r in range(header_row, len(rows)):
        row = rows[r]
        if not row or not row[0]:
            continue
        try:
            freq = int(float(row[18]))
            elapsed = int(float(row[19]))
        except (ValueError, TypeError, IndexError):
            continue
        cnx_type = str(row[20]).strip().upper() if row[20] else ''
        if cnx_type not in CNX_COEFFS or freq <= 0:
            continue
        result.append({
            'city': str(row[0]).strip(),
            'airport': str(row[1]).strip(),
            'route_label': str(row[2]).strip(),
            'carrier_l1': str(row[5]).strip() if row[5] else '',
            'cnx_airport': str(row[7]).strip() if row[7] else '',
            'carrier_l2': str(row[17]).strip() if row[17] else '',
            'freq': freq, 'elapsed': elapsed, 'cnx_type': cnx_type,
        })
    return result


def score_sheet(itineraries, factor=DEFAULT_FACTOR, interval=DEFAULT_INTERVAL):
    """Score itineraries. Min elapsed is per CITY (not airport)."""
    min_el = {}
    for it in itineraries:
        c = it['city']  # GROUP BY CITY
        if c not in min_el or it['elapsed'] < min_el[c]:
            min_el[c] = it['elapsed']
    for it in itineraries:
        me = min_el[it['city']]
        it['min_elapsed'] = me
        it['excess_hrs'] = (it['elapsed'] - me) / 60.0
        it['et_coeff'] = et_coeff(it['elapsed'], me, factor, interval)
        it['cnx_coeff'] = CNX_COEFFS.get(it['cnx_type'], 0)
        it['qsi'] = it['freq'] * it['et_coeff'] * it['cnx_coeff']
    return itineraries


def aggregate_shares(scored):
    """Aggregate by route_label. Market total is per CITY."""
    route_qsi = defaultdict(lambda: {'qsi': 0.0, 'city': '', 'airport': ''})
    city_market = defaultdict(float)  # KEY CHANGE: city not airport
    for it in scored:
        rl = it['route_label']
        route_qsi[rl]['qsi'] += it['qsi']
        route_qsi[rl]['city'] = it['city']
        route_qsi[rl]['airport'] = it['airport']
        city_market[it['city']] += it['qsi']  # GROUP BY CITY
    result = {}
    for rl, data in route_qsi.items():
        cy = data['city']
        mkt = city_market.get(cy, 0)
        result[rl] = {
            'city': cy, 'airport': data['airport'],
            'carrier_qsi': data['qsi'], 'market_qsi': mkt,
            'fair_share': data['qsi'] / mkt if mkt > 0 else 0.0,
        }
    return result, city_market


def bidirectional_calc(shares1, shares2, cm1, cm2):
    """Bidirectional average. Missing route labels get carrier=0, market=city total."""
    all_labels = set(shares1.keys()) | set(shares2.keys())
    result = {}
    for rl in sorted(all_labels):
        s1 = shares1.get(rl, {})
        s2 = shares2.get(rl, {})
        city = s1.get('city', '') or s2.get('city', '')
        airport = s1.get('airport', '') or s2.get('airport', '')
        q1c = s1.get('carrier_qsi', 0)
        q1m = s1.get('market_qsi', 0) or cm1.get(city, 0)
        fs1 = s1.get('fair_share', 0) if s1 else 0
        q2c = s2.get('carrier_qsi', 0)
        q2m = s2.get('market_qsi', 0) or cm2.get(city, 0)
        fs2 = s2.get('fair_share', 0) if s2 else 0
        avg = (fs1 + fs2) / 2.0
        rt = 'Exclude' if (q1m == 0 or q2m == 0) else 'Include'
        adj = 0.0 if rt == 'Exclude' else avg
        result[rl] = {
            'city': city, 'airport': airport,
            'q1_carrier': q1c, 'q1_market': q1m, 'fs1': fs1,
            'q2_carrier': q2c, 'q2_market': q2m, 'fs2': fs2,
            'avg_share': avg, 'rt_check': rt, 'adj_share': adj,
        }
    return result


def load_actual(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['QSI Calc']
    actual = {}
    for row in ws.iter_rows(min_row=12, max_col=12, values_only=True):
        if not row or not row[2]:
            continue
        rl = str(row[2]).strip()
        if rl in ('Code', ''):
            continue
        try:
            actual[rl] = {
                'city': str(row[1]).strip() if row[1] else '',
                'q1c': float(row[3]) if row[3] else 0,
                'q1m': float(row[4]) if row[4] else 0,
                'fs1': float(row[5]) if row[5] else 0,
                'q2c': float(row[6]) if row[6] else 0,
                'q2m': float(row[7]) if row[7] else 0,
                'fs2': float(row[8]) if row[8] else 0,
                'avg': float(row[9]) if row[9] else 0,
                'adj': float(row[11]) if row[11] else 0,
            }
        except (ValueError, TypeError):
            continue
    wb.close()
    return actual


def validate(filepath='/mnt/project/QSILHR_v1_OS_JZ_17Feb15.xlsx'):
    q1 = load_qsi_sheet(filepath, 'QSI 1')
    q2 = load_qsi_sheet(filepath, 'QSI 2')
    q1s = score_sheet(q1)
    q2s = score_sheet(q2)
    sh1, cm1 = aggregate_shares(q1s)
    sh2, cm2 = aggregate_shares(q2s)
    bidir = bidirectional_calc(sh1, sh2, cm1, cm2)
    actual = load_actual(filepath)

    exact = close = off = missing = 0
    max_err = 0
    worst = []
    for rl in sorted(actual.keys()):
        a = actual[rl]
        if rl not in bidir:
            missing += 1
            continue
        b = bidir[rl]
        err = abs(b['adj_share'] - a['adj'])
        max_err = max(max_err, err)
        if err < 1e-10:
            exact += 1
        elif err < 0.0005:
            close += 1
        else:
            off += 1
            worst.append((err, rl, b['adj_share'], a['adj']))

    extra = len(set(bidir.keys()) - set(actual.keys()))
    total = exact + close + off
    worst.sort(reverse=True)

    print(f"Itineraries: QSI1={len(q1)}, QSI2={len(q2)}")
    print(f"Route labels: {len(actual)} actual, {len(bidir)} computed")
    print(f"Match: {total}/{len(actual)} | Exact={exact} Close={close} Off={off} Missing={missing} Extra={extra}")
    print(f"Max error: {max_err:.10f}")
    if worst[:10]:
        print("Worst mismatches:")
        for e, rl, ours, act in worst[:10]:
            print(f"  {rl}: ours={ours:.6f} actual={act:.6f} err={e:.6f}")

    return bidir, q1s, q2s


def write_output(bidir, q1s, q2s, metadata, outpath):
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    def hdr(ws, cols):
        ws.append(cols)
        for c in range(1, len(cols)+1):
            ws.cell(1,c).font = hf
            ws.cell(1,c).fill = hfill
            ws.cell(1,c).alignment = Alignment(horizontal='center')

    ws = wb.active; ws.title = 'Metadata'
    ws.append(['QSI Scorer  Module IV']); ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    for k,v in metadata.items(): ws.append([k, str(v)])
    ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 60

    ws2 = wb.create_sheet('QSI_Calc')
    hdr(ws2, ['City','Route Label','QSI1 Carrier','QSI1 Market','FS1',
        'QSI2 Carrier','QSI2 Market','FS2','Avg FS','RT Check','Adj Share'])
    for rl in sorted(bidir.keys()):
        d = bidir[rl]
        ws2.append([d['city'], rl, d['q1_carrier'], d['q1_market'], d['fs1'],
            d['q2_carrier'], d['q2_market'], d['fs2'],
            d['avg_share'], d['rt_check'], d['adj_share']])
    for r in range(2, ws2.max_row+1):
        for c in [5,8,9,11]: ws2.cell(r,c).number_format = '0.000000'
        for c in [3,4,6,7]: ws2.cell(r,c).number_format = '#,##0.000000'

    ws3 = wb.create_sheet('City_Summary')
    hdr(ws3, ['City','N Routes','Sum Adj Share','Best Route'])
    ca = defaultdict(lambda: {'n':0,'sum':0,'best':0,'brl':''})
    for rl, d in bidir.items():
        c = d['city']
        ca[c]['n'] += 1; ca[c]['sum'] += d['adj_share']
        if d['adj_share'] > ca[c]['best']:
            ca[c]['best'] = d['adj_share']; ca[c]['brl'] = rl
    for city in sorted(ca.keys()):
        a = ca[city]
        ws3.append([city, a['n'], a['sum'], a['brl']])

    dcols = ['City','Airport','Route Label','Carrier L1','Cnx Apt','Carrier L2',
        'Freq','Elapsed','Cnx Type','Min Elapsed','Excess Hrs','ET Coeff','Cnx Coeff','QSI']
    for name, scored in [('QSI1_Detail', q1s), ('QSI2_Detail', q2s)]:
        ws4 = wb.create_sheet(name)
        hdr(ws4, dcols)
        for it in sorted(scored, key=lambda x: (x['city'], x['route_label'])):
            ws4.append([it['city'], it['airport'], it['route_label'],
                it['carrier_l1'], it['cnx_airport'], it['carrier_l2'],
                it['freq'], it['elapsed'], it['cnx_type'], it['min_elapsed'],
                round(it['excess_hrs'],4), it['et_coeff'], it['cnx_coeff'], it['qsi']])

    ws5 = wb.create_sheet('Summary')
    ws5.append(['QSI Scorer Summary']); ws5['A1'].font = Font(bold=True, size=14)
    ws5.append([])
    n_incl = sum(1 for d in bidir.values() if d['rt_check']=='Include')
    ws5.append(['Total route labels', len(bidir)])
    ws5.append(['Unique cities', len(ca)])
    ws5.append(['Included (RT pass)', n_incl])
    ws5.append(['Excluded', len(bidir) - n_incl])
    ws5.append(['QSI 1 itineraries', len(q1s)])
    ws5.append(['QSI 2 itineraries', len(q2s)])

    wb.save(outpath)
    print(f"Output: {outpath}")
    return outpath


if __name__ == '__main__':
    fp = '/mnt/project/QSILHR_v1_OS_JZ_17Feb15.xlsx'
    bidir, q1s, q2s = validate(fp)
    meta = {'Source': fp, 'Route': 'LHR-SJC', 'Airline': 'BA',
            'Factor': DEFAULT_FACTOR, 'Interval': f'{DEFAULT_INTERVAL} hrs',
            'Coefficients': 'ONLINE=1.0, ALLIANCE=0.615, INTERLINING=0.25',
            'Grouping': 'City-level (not airport-level)'}
    write_output(bidir, q1s, q2s, meta, '/home/claude/QSI_Scorer_Validated.xlsx')
