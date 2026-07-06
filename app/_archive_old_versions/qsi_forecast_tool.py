#!/usr/bin/env python3
"""
================================================================================
  AVIA SOLUTIONS  QSI FORECAST TOOL v2.0
  All-in-one: Engine + Reader + Calibrator + Writer
================================================================================

  Reads a populated workbook template  runs forecast engine  calibrates 
  writes standardised output workbook with sensitivity analysis.

  Validated against three historical routes:
    BA LHR-SJC:     129,162 pax  (0.000% variance)
    KLM AMS-TPA:     79,771 pax  (0.002% variance)
    JX TPE-SJC:      69,736 pax  (0.000% variance)

  Usage:
    python3 qsi_forecast_tool.py                           # Run validation suite
    python3 qsi_forecast_tool.py input.xlsx                # Read & forecast
    python3 qsi_forecast_tool.py input.xlsx --target-lf 85 # Calibrate to 85% LF
    python3 qsi_forecast_tool.py input.xlsx --business-case # Business case test
    python3 qsi_forecast_tool.py --generate-template       # Create blank template
================================================================================
"""
import os, sys, argparse
from dataclasses import dataclass, field
from typing import Optional, List
from copy import deepcopy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 
# PART 1: ENGINE  Core calculation dataclasses
# 

@dataclass
class P2PSegment:
    """Single P2P demand segment. Formula: base  (1+growth)  seasonality  stimulation  capture"""
    name: str
    base_demand: float
    growth: float
    seasonality: float = 1.0
    stimulation: float = 1.0
    capture_rate: float = 0.0

    @property
    def grown_demand(self): return self.base_demand * (1 + self.growth)
    @property
    def stimulated_demand(self): return self.grown_demand * self.seasonality * self.stimulation
    @property
    def forecast(self): return self.stimulated_demand * self.capture_rate
    @property
    def ptew(self): return self.forecast / 52 / 2 if self.forecast > 0 else 0


@dataclass
class ConnectingAggregate:
    """Aggregate connecting traffic through a hub."""
    hub_code: str
    direct_base: float = 0;    direct_growth: float = 0;   direct_capture: float = 0
    indirect_base: float = 0;  indirect_growth: float = 0;  indirect_capture: float = 0
    stimulation: float = 1.0

    @property
    def direct_grown(self): return self.direct_base * (1 + self.direct_growth)
    @property
    def indirect_grown(self): return self.indirect_base * (1 + self.indirect_growth)
    @property
    def direct_forecast(self): return self.direct_grown * self.stimulation * self.direct_capture
    @property
    def indirect_forecast(self): return self.indirect_grown * self.stimulation * self.indirect_capture
    @property
    def total_forecast(self): return self.direct_forecast + self.indirect_forecast
    @property
    def total_base(self): return self.direct_base + self.indirect_base


@dataclass
class RouteConfig:
    mode: str = "Forecast";       analyst: str = "";         date: str = ""
    forecast_year: str = ""
    origin: str = "";             origin_city: str = "";     origin_country: str = ""
    destination: str = "";        destination_city: str = "";  destination_country: str = ""
    carrier: str = "";            carrier_iata: str = ""
    route_type: str = "Long-haul"; carrier_type: str = "Full-service"
    market_maturity: str = "New route"
    frequency: int = 0;           aircraft: str = "";        seats: int = 0
    annual_seats: int = 0;        service_weeks: float = 52.0


@dataclass
class ForecastResult:
    p2p_segments: list = field(default_factory=list)
    connecting: list = field(default_factory=list)
    config: Optional[RouteConfig] = None

    @property
    def p2p_forecast(self): return sum(s.forecast for s in self.p2p_segments)
    @property
    def cnx_forecast(self): return sum(c.total_forecast for c in self.connecting)
    @property
    def total_forecast(self): return self.p2p_forecast + self.cnx_forecast
    @property
    def annual_seats(self): return self.config.annual_seats if self.config else 0
    @property
    def load_factor(self): return self.total_forecast / self.annual_seats if self.annual_seats > 0 else 0

    def print_summary(self):
        c = self.config
        print(f"\n{''*70}")
        print(f"  {c.carrier} {c.origin}-{c.destination}  |  {c.frequency}x {c.aircraft}")
        print(f"{''*70}")
        print(f"\n  {'SEGMENT':<40s} {'BASE':>10s} {'GROWN':>10s} {'STIM':>6s} {'CAP':>6s} {'FCST':>10s}")
        print(f"  {'-'*82}")
        for s in self.p2p_segments:
            print(f"  {s.name:<40s} {s.base_demand:>10,.0f} {s.grown_demand:>10,.0f} "
                  f"{s.stimulation:>6.2f} {s.capture_rate:>5.1%} {s.forecast:>10,.0f}")
        if self.p2p_segments:
            print(f"  {'P2P TOTAL':<40s} {'':<10s} {'':<10s} {'':<6s} {'':<6s} {self.p2p_forecast:>10,.0f}")
        for cx in self.connecting:
            print(f"  {f'Cnx @ {cx.hub_code} (direct comp)':<40s} {cx.direct_base:>10,.0f} {cx.direct_grown:>10,.0f} "
                  f"{cx.stimulation:>6.2f} {cx.direct_capture:>5.1%} {cx.direct_forecast:>10,.0f}")
            print(f"  {f'Cnx @ {cx.hub_code} (no direct)':<40s} {cx.indirect_base:>10,.0f} {cx.indirect_grown:>10,.0f} "
                  f"{cx.stimulation:>6.2f} {cx.indirect_capture:>5.1%} {cx.indirect_forecast:>10,.0f}")
        print(f"  {'-'*82}")
        print(f"  {'GRAND TOTAL':<40s} {'':<10s} {'':<10s} {'':<6s} {'':<6s} {self.total_forecast:>10,.0f}")
        print(f"\n  Seats: {self.annual_seats:,d}  |  Pax: {self.total_forecast:,.0f}  |  LF: {self.load_factor:.1%}")


# 
# PART 2: READER  Parse workbook into engine objects
# 

def read_workbook(path: str) -> ForecastResult:
    """Read a standardised Avia Solutions workbook and return ForecastResult."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # Read Cover Page
    ws_cp = wb['Cover Page']
    params = {}
    for r in range(1, 100):
        lbl = ws_cp.cell(row=r, column=2).value
        val = ws_cp.cell(row=r, column=3).value
        if lbl and val is not None:
            params[str(lbl).strip()] = val

    config = RouteConfig(
        mode=str(params.get('Mode', 'Forecast')),
        analyst=str(params.get('Analyst', '')),
        date=str(params.get('Date', '')),
        forecast_year=str(params.get('Forecast Year', '')),
        origin=str(params.get('Origin Airport (IATA)', '')),
        origin_city=str(params.get('Origin City', '')),
        origin_country=str(params.get('Origin Country', '')),
        destination=str(params.get('Destination Airport (IATA)', '')),
        destination_city=str(params.get('Destination City', '')),
        destination_country=str(params.get('Destination Country', '')),
        carrier=str(params.get('Carrier', '')),
        carrier_iata=str(params.get('Carrier IATA', '')),
        frequency=int(params.get('Frequency (per week)', 0)),
        aircraft=str(params.get('Aircraft Type', '')),
        seats=int(params.get('Seats per Flight', 0)),
        annual_seats=int(params.get('Annual Seats', 0)),
    )

    # Read Forecast sheet
    ws = wb['Forecast']
    p2p_segments = []
    connecting = []
    cnx_hub = None

    # Scan for data rows  column layout:
    # A=label, B=base, C=growth, D=grown (formula), E=stim, F=after stim, G=capture, H=forecast, I=ptew
    for r in range(1, 200):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value  # base demand
        c_val = ws.cell(row=r, column=3).value  # growth
        e_val = ws.cell(row=r, column=5).value  # stimulation
        g_val = ws.cell(row=r, column=7).value  # capture rate
        h_val = ws.cell(row=r, column=8).value  # forecast

        if a_val is None:
            continue
        a_str = str(a_val).strip()

        # Skip headers and totals
        if a_str in ('Market', 'SEGMENT', 'P2P TOTAL', 'GRAND TOTAL', 'Annual Seats',
                      'Annual Pax', 'Load Factor', 'Capacity', '') or 'total' in a_str.lower():
            continue

        # Detect connecting section headers
        if 'Connecting' in a_str or 'Cnx @' in a_str or 'Beyond' in a_str:
            # Extract hub code
            for token in a_str.replace('(', ' ').replace(')', ' ').split():
                if len(token) == 3 and token.isupper():
                    cnx_hub = token
                    break
            continue

        # Skip sub-headers (no numeric base demand)
        if b_val is None or not isinstance(b_val, (int, float)):
            continue

        # Auto-detect scale: if base demand < 500, assume thousands
        base = float(b_val)
        if base > 0 and base < 500:
            base *= 1000

        growth = float(c_val) if isinstance(c_val, (int, float)) else 0.0
        stim = float(e_val) if isinstance(e_val, (int, float)) else 1.0
        capture = float(g_val) if isinstance(g_val, (int, float)) else 0.0

        if cnx_hub:
            # This is a connecting row  direct or indirect
            is_direct = 'direct' in a_str.lower() and 'no' not in a_str.lower()
            # Find or create the aggregate
            agg = None
            for c in connecting:
                if c.hub_code == cnx_hub:
                    agg = c
                    break
            if agg is None:
                agg = ConnectingAggregate(hub_code=cnx_hub)
                connecting.append(agg)

            if is_direct:
                agg.direct_base = base
                agg.direct_growth = growth
                agg.direct_capture = capture
            else:
                agg.indirect_base = base
                agg.indirect_growth = growth
                agg.indirect_capture = capture
            agg.stimulation = stim
        else:
            p2p_segments.append(P2PSegment(
                name=a_str, base_demand=base, growth=growth,
                seasonality=1.0, stimulation=stim, capture_rate=capture))

    wb.close()
    return ForecastResult(p2p_segments=p2p_segments, connecting=connecting, config=config)


# 
# PART 3: CALIBRATOR  Adjust parameters to hit targets
# 

def calibrate_capture(result: ForecastResult, target_lf: float = 0,
                      target_pax: float = 0) -> ForecastResult:
    """Adjust all capture rates proportionally to hit target LF or pax."""
    r = deepcopy(result)
    target = target_pax if target_pax > 0 else (target_lf * r.annual_seats if target_lf > 0 else 0)
    if target <= 0:
        return r

    current = r.total_forecast
    if current <= 0:
        return r
    ratio = target / current

    for s in r.p2p_segments:
        s.capture_rate = min(s.capture_rate * ratio, 1.0)
    for c in r.connecting:
        c.direct_capture = min(c.direct_capture * ratio, 1.0)
        c.indirect_capture = min(c.indirect_capture * ratio, 1.0)
    return r


def sensitivity_analysis(result: ForecastResult, param: str = 'capture',
                         range_pct: float = 0.30, steps: int = 5) -> dict:
    """Run sensitivity on a parameter. Returns {factor: {total_pax, load_factor, delta_pct}}"""
    base_total = result.total_forecast
    factors = [1.0 + (i - steps) * range_pct / steps for i in range(2 * steps + 1)]
    out = {}
    for f in factors:
        r = deepcopy(result)
        if param == 'capture':
            for s in r.p2p_segments: s.capture_rate *= f
            for c in r.connecting: c.direct_capture *= f; c.indirect_capture *= f
        elif param == 'stimulation':
            for s in r.p2p_segments: s.stimulation *= f
            for c in r.connecting: c.stimulation *= f
        elif param == 'growth':
            for s in r.p2p_segments: s.growth *= f
            for c in r.connecting: c.direct_growth *= f; c.indirect_growth *= f
        total = r.total_forecast
        out[f"{f:.2f}"] = {
            'total_pax': total,
            'load_factor': total / r.annual_seats if r.annual_seats else 0,
            'delta_pct': (total - base_total) / base_total * 100 if base_total else 0
        }
    return out


def business_case_test(result: ForecastResult, year1_lf: float = 0.70,
                       mature_lf: float = 0.82, ramp_years: int = 3) -> dict:
    """Test whether business case targets are achievable."""
    yr1_target = year1_lf * result.annual_seats
    mature_target = mature_lf * result.annual_seats
    current = result.total_forecast
    yr1_ok = current >= yr1_target
    if current > 0 and mature_target > current:
        cagr = (mature_target / current) ** (1.0 / ramp_years) - 1
    else:
        cagr = 0
    return {
        'forecast': current, 'load_factor': result.load_factor,
        'yr1_target': yr1_target, 'yr1_lf': year1_lf, 'yr1_ok': yr1_ok,
        'mature_target': mature_target, 'mature_lf': mature_lf,
        'cagr': cagr, 'cagr_ok': cagr < 0.10, 'ramp_years': ramp_years,
    }


# 
# PART 4: WRITER  Generate professional Excel output
# 

# Styles
BLUE = '003366'; MID = '4472C4'; LIGHT = 'D6E4F0'
HDR_FILL = PatternFill('solid', fgColor=BLUE)
SEC_FILL = PatternFill('solid', fgColor=MID)
INP_FILL = PatternFill('solid', fgColor='FFF2CC')
DER_FILL = PatternFill('solid', fgColor='E2EFDA')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
LBL_FONT = Font(name='Arial', size=10, color='333333')
INP_FONT = Font(name='Arial', size=10, color='0000FF')
DER_FONT = Font(name='Arial', size=10, color='006100')
BOLD_F = Font(name='Arial', size=10, bold=True)
NORM_F = Font(name='Arial', size=10)
TITLE_F = Font(name='Arial', bold=True, color=BLUE, size=14)
THIN = Border(left=Side('thin',color='B4B4B4'), right=Side('thin',color='B4B4B4'),
              top=Side('thin',color='B4B4B4'), bottom=Side('thin',color='B4B4B4'))
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _style(ws, r, c, val=None, font=NORM_F, fill=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font; cell.border = THIN; cell.alignment = CTR
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    return cell


def write_output(result: ForecastResult, path: str, sens_data=None, bc_data=None):
    """Write standardised output workbook."""
    wb = openpyxl.Workbook()
    c = result.config

    #  Cover Page 
    ws = wb.active; ws.title = 'Cover Page'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35

    _style(ws, 1, 2, 'AVIA SOLUTIONS  QSI FORECAST OUTPUT', font=TITLE_F)
    ws.merge_cells('B1:C1')

    cover_data = [
        ('', ''), ('ENGAGEMENT', ''),
        ('Mode', c.mode), ('Analyst', c.analyst), ('Date', c.date), ('Forecast Year', c.forecast_year),
        ('', ''), ('ROUTE', ''),
        ('Origin Airport (IATA)', c.origin), ('Origin City', c.origin_city), ('Origin Country', c.origin_country),
        ('Destination Airport (IATA)', c.destination), ('Destination City', c.destination_city),
        ('Destination Country', c.destination_country),
        ('', ''), ('CARRIER', ''),
        ('Carrier', c.carrier), ('Carrier IATA', c.carrier_iata),
        ('', ''), ('SERVICE', ''),
        ('Frequency (per week)', c.frequency), ('Aircraft Type', c.aircraft),
        ('Seats per Flight', c.seats), ('Annual Seats', c.annual_seats),
        ('', ''), ('RESULTS', ''),
        ('Total Forecast (pax)', round(result.total_forecast)),
        ('Load Factor', result.load_factor),
    ]

    row = 3
    for label, val in cover_data:
        if not label:
            if val:
                _style(ws, row, 2, val, font=Font(name='Arial', bold=True, color='FFFFFF', size=10), fill=SEC_FILL)
                _style(ws, row, 3, '', fill=SEC_FILL)
            row += 1
            continue
        _style(ws, row, 2, label, font=LBL_FONT)
        is_input = label not in ('Total Forecast (pax)', 'Load Factor', 'Annual Seats')
        _style(ws, row, 3, val, font=INP_FONT if is_input else DER_FONT,
               fill=INP_FILL if is_input else DER_FILL,
               fmt='0.0%' if 'Factor' in label or 'Load' in label else '#,##0' if isinstance(val, (int, float)) and 'IATA' not in label else None)
        row += 1

    #  Forecast Sheet 
    ws2 = wb.create_sheet('Forecast')
    headers = ['Market', 'Base Demand', 'Growth', 'Grown Demand', 'Stimulation',
               'After Stimulation', 'Capture Rate', 'Forecast', 'PTEW']
    widths = [40, 14, 10, 14, 12, 16, 12, 14, 10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _style(ws2, 1, i, h, font=HDR_FONT, fill=HDR_FILL)
        ws2.column_dimensions[chr(64+i) if i < 10 else 'I'].width = w

    def write_seg(ws, row, name, base, growth, stim, capture):
        _style(ws, row, 1, name, font=LBL_FONT); ws.cell(row=row, column=1).alignment = LFT
        _style(ws, row, 2, base, font=INP_FONT, fill=INP_FILL, fmt='#,##0')
        _style(ws, row, 3, growth, font=INP_FONT, fill=INP_FILL, fmt='0.0%')
        grown = base * (1 + growth)
        _style(ws, row, 4, grown, font=DER_FONT, fill=DER_FILL, fmt='#,##0')
        _style(ws, row, 5, stim, font=INP_FONT, fill=INP_FILL, fmt='0.00')
        after_stim = grown * stim
        _style(ws, row, 6, after_stim, font=DER_FONT, fill=DER_FILL, fmt='#,##0')
        _style(ws, row, 7, capture, font=INP_FONT, fill=INP_FILL, fmt='0.0%')
        fcst = after_stim * capture
        _style(ws, row, 8, fcst, font=DER_FONT, fill=DER_FILL, fmt='#,##0')
        _style(ws, row, 9, fcst / 52 / 2 if fcst > 0 else 0, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
        return row + 1

    r = 2
    # P2P segments
    for seg in result.p2p_segments:
        r = write_seg(ws2, r, seg.name, seg.base_demand, seg.growth, seg.stimulation, seg.capture_rate)

    # P2P total
    _style(ws2, r, 1, 'P2P TOTAL', font=BOLD_F); ws2.cell(row=r, column=1).alignment = LFT
    _style(ws2, r, 8, result.p2p_forecast, font=BOLD_F, fmt='#,##0')
    r += 1

    # Connecting
    for cx in result.connecting:
        _style(ws2, r, 1, f'Connecting @ {cx.hub_code}', font=Font(name='Arial', bold=True, color=MID, size=10))
        _style(ws2, r, 1, f'Connecting @ {cx.hub_code}'); ws2.cell(row=r, column=1).font = Font(name='Arial', bold=True, color=MID, size=10)
        r += 1
        r = write_seg(ws2, r, f'  Direct competition cities', cx.direct_base, cx.direct_growth, cx.stimulation, cx.direct_capture)
        r = write_seg(ws2, r, f'  No direct competition cities', cx.indirect_base, cx.indirect_growth, cx.stimulation, cx.indirect_capture)

    # Grand total
    r += 1
    _style(ws2, r, 1, 'GRAND TOTAL', font=Font(name='Arial', bold=True, color=BLUE, size=11)); ws2.cell(row=r, column=1).alignment = LFT
    _style(ws2, r, 8, result.total_forecast, font=Font(name='Arial', bold=True, color=BLUE, size=11), fmt='#,##0')
    r += 2
    _style(ws2, r, 1, 'Annual Seats', font=BOLD_F); _style(ws2, r, 8, result.annual_seats, font=BOLD_F, fmt='#,##0')
    r += 1
    _style(ws2, r, 1, 'Annual Pax', font=BOLD_F); _style(ws2, r, 8, round(result.total_forecast), font=BOLD_F, fmt='#,##0')
    r += 1
    _style(ws2, r, 1, 'Load Factor', font=BOLD_F); _style(ws2, r, 8, result.load_factor, font=BOLD_F, fmt='0.0%')

    #  Sensitivity Sheet 
    if sens_data:
        ws3 = wb.create_sheet('Sensitivity')
        sr = 1
        for param_name, data in sens_data.items():
            _style(ws3, sr, 1, f'Sensitivity: {param_name.title()} Rate (30%)',
                   font=Font(name='Arial', bold=True, color=BLUE, size=12))
            sr += 1
            for i, h in enumerate(['Factor', 'Pax', 'Load Factor', ' from Base'], 1):
                _style(ws3, sr, i, h, font=HDR_FONT, fill=HDR_FILL)
            sr += 1
            for factor_str, vals in data.items():
                f = float(factor_str)
                is_base = abs(f - 1.0) < 0.001
                font = BOLD_F if is_base else NORM_F
                _style(ws3, sr, 1, f, font=font, fmt='0%')
                _style(ws3, sr, 2, vals['total_pax'], font=font, fmt='#,##0')
                _style(ws3, sr, 3, vals['load_factor'], font=font, fmt='0.0%')
                _style(ws3, sr, 4, vals['delta_pct'] / 100, font=font, fmt='+0.0%;-0.0%')
                sr += 1
            sr += 2

        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 14
        ws3.column_dimensions['C'].width = 14
        ws3.column_dimensions['D'].width = 14

    #  Business Case Sheet 
    if bc_data:
        ws4 = wb.create_sheet('Business Case')
        _style(ws4, 1, 1, 'Business Case Assessment', font=TITLE_F)
        ws4.merge_cells('A1:D1')
        bc_rows = [
            ('Current Forecast', bc_data['forecast'], '#,##0'),
            ('Current Load Factor', bc_data['load_factor'], '0.0%'),
            ('Annual Seats', result.annual_seats, '#,##0'),
            ('', '', ''),
            ('Year 1 Target LF', bc_data['yr1_lf'], '0.0%'),
            ('Year 1 Target Pax', bc_data['yr1_target'], '#,##0'),
            ('Year 1 Achievable?', 'YES ' if bc_data['yr1_ok'] else 'NO ', ''),
            ('', '', ''),
            ('Maturity Target LF', bc_data['mature_lf'], '0.0%'),
            ('Maturity Target Pax', bc_data['mature_target'], '#,##0'),
            ('CAGR Required', bc_data['cagr'], '0.0%'),
            ('CAGR Realistic? (<10%)', 'YES ' if bc_data['cagr_ok'] else 'AGGRESSIVE ', ''),
        ]
        for i, (label, val, fmt) in enumerate(bc_rows, 3):
            _style(ws4, i, 1, label, font=LBL_FONT)
            _style(ws4, i, 2, val, font=BOLD_F, fmt=fmt if fmt else None)
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 20

    #  Assumptions Log 
    ws5 = wb.create_sheet('Assumptions Log')
    _style(ws5, 1, 1, 'ASSUMPTIONS LOG', font=TITLE_F); ws5.merge_cells('A1:D1')
    _style(ws5, 2, 1, f'{c.carrier} {c.origin}-{c.destination}  |  {c.date}', font=LBL_FONT)
    log_headers = ['Parameter', 'Value', 'Source / Justification', 'Confidence']
    for i, h in enumerate(log_headers, 1):
        _style(ws5, 4, i, h, font=HDR_FONT, fill=HDR_FILL)
    lr = 5
    for seg in result.p2p_segments:
        _style(ws5, lr, 1, f'{seg.name}  Growth', font=LBL_FONT)
        _style(ws5, lr, 2, seg.growth, font=INP_FONT, fmt='0.0%')
        lr += 1
        _style(ws5, lr, 1, f'{seg.name}  Stimulation', font=LBL_FONT)
        _style(ws5, lr, 2, seg.stimulation, font=INP_FONT, fmt='0.00')
        lr += 1
        _style(ws5, lr, 1, f'{seg.name}  Capture Rate', font=LBL_FONT)
        _style(ws5, lr, 2, seg.capture_rate, font=INP_FONT, fmt='0.0%')
        lr += 1
    for cx in result.connecting:
        _style(ws5, lr, 1, f'Cnx {cx.hub_code}  Direct Capture', font=LBL_FONT)
        _style(ws5, lr, 2, cx.direct_capture, font=INP_FONT, fmt='0.00%')
        lr += 1
        _style(ws5, lr, 1, f'Cnx {cx.hub_code}  Indirect Capture', font=LBL_FONT)
        _style(ws5, lr, 2, cx.indirect_capture, font=INP_FONT, fmt='0.00%')
        lr += 1
    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 14
    ws5.column_dimensions['C'].width = 40
    ws5.column_dimensions['D'].width = 14

    wb.save(path)
    return path


# 
# PART 5: VALIDATED TEST CASES
# 

def build_ba_lhr_sjc():
    cfg = RouteConfig(mode="Forecast", analyst="JZ/OS", date="23-Feb-2015", forecast_year="2016",
        origin="LHR", origin_city="London", origin_country="United Kingdom",
        destination="SJC", destination_city="San Jose", destination_country="USA",
        carrier="British Airways", carrier_iata="BA", frequency=7, aircraft="787-800",
        seats=214, annual_seats=155792, service_weeks=52)
    # Growth is 10% (0.10) not 2.5%  confirmed from BA_Fcst_LHRSJC_JZ_23Feb2015_FINAL.xlsm
    p2p = [
        P2PSegment("UK Visitors Business",         71441.55, 0.10, 1.0, 1.15, 0.40),
        P2PSegment("UK Visitors Leisure Primary",   36385.76, 0.10, 1.0, 1.0,  0.25),
        P2PSegment("UK Visitors Leisure Secondary",  17448.74, 0.10, 1.0, 1.0, 0.25),
        P2PSegment("UK Visitors Leisure Contested",   4617.68, 0.10, 1.0, 1.0, 0.10),
        P2PSegment("US Residents Business",         65946.05, 0.10, 1.0, 1.15, 0.15),
        P2PSegment("US Residents Leisure Primary",  33586.86, 0.10, 1.0, 1.0,  0.25),
        P2PSegment("US Residents Leisure Secondary", 16106.53, 0.10, 1.0, 1.0, 0.25),
        P2PSegment("US Residents Leisure Contested",  4262.47, 0.10, 1.0, 1.0, 0.10),
    ]
    # LHR connecting: ZERO direct competition base, 1,226,387 indirect only
    # SJC connecting: 977,784 direct, 129,411 indirect
    cnx = [
        ConnectingAggregate("LHR", 0, 0.09, 0.0, 1226387.35, 0.09, 0.035994),
        ConnectingAggregate("SJC", 977784, 0.10, 0.001017, 129411, 0.10, 0.012950),
    ]
    return ForecastResult(p2p, cnx, cfg)

def build_klm_ams_tpa():
    cfg = RouteConfig(mode="Forecast", analyst="JZ/OS/RN/JK", date="Jul 2023", forecast_year="2025",
        origin="AMS", origin_city="Amsterdam", origin_country="Netherlands",
        destination="TPA", destination_city="Tampa", destination_country="USA",
        carrier="KLM", carrier_iata="KL", frequency=3, aircraft="A330-300",
        seats=292, annual_seats=91104, service_weeks=52)
    p2p = [P2PSegment("Point to Point", 38617.37, 0.1528, 1.0, 1.0, 0.60)]
    cnx = [ConnectingAggregate("AMS", 290897, 0.1528, 0.003266, 450441, 0.16350, 0.09915)]
    return ForecastResult(p2p, cnx, cfg)

def build_jx_tpe_sjc():
    cfg = RouteConfig(mode="Forecast", analyst="AviaSolutions", date="09-Aug-2022", forecast_year="2024",
        origin="TPE", origin_city="Taipei", origin_country="Taiwan",
        destination="SJC", destination_city="San Jose", destination_country="USA",
        carrier="Starlux Airlines", carrier_iata="JX", frequency=3, aircraft="A350",
        seats=304, annual_seats=158080, service_weeks=52)
    p2p = [
        P2PSegment("Taiwan Business",          45599.68, 0.125, 1.0, 1.10, 0.40),
        P2PSegment("Taiwan Leisure Primary",     9333.19, 0.097, 1.0, 1.10, 0.40),
        P2PSegment("Taiwan Leisure Secondary",   1286.93, 0.097, 1.0, 1.05, 0.40),
        P2PSegment("Taiwan Leisure Contested",    779.80, 0.097, 1.0, 1.05, 0.25),
        P2PSegment("US Business",               68399.52, 0.125, 1.0, 1.10, 0.40),
        P2PSegment("US Leisure Primary",        13999.78, 0.097, 1.0, 1.10, 0.40),
        P2PSegment("US Leisure Secondary",       1930.40, 0.097, 1.0, 1.05, 0.40),
        P2PSegment("US Leisure Contested",       1169.70, 0.097, 1.0, 1.05, 0.25),
    ]
    return ForecastResult(p2p, [], cfg)


# 
# PART 6: TEMPLATE GENERATOR
# 

def generate_blank_template(output_path: str):
    blank = ForecastResult(
        p2p_segments=[P2PSegment("[Segment 1 Name]", 0, 0, 1.0, 1.0, 0)],
        connecting=[ConnectingAggregate("[HUB]")],
        config=RouteConfig(mode="Forecast", analyst="[Analyst]", date="[Date]",
                           forecast_year="[Year]", origin="[XXX]", destination="[YYY]",
                           carrier="[Airline]", carrier_iata="[XX]"),
    )
    write_output(blank, output_path)
    return output_path


# 
# PART 7: MAIN  CLI entry point
# 

def run_validation():
    print("" * 80)
    print("  AVIA SOLUTIONS  QSI FORECAST TOOL v2.0")
    print("  Engine Validation Suite")
    print("" * 80)

    tests = [
        ("BA LHR-SJC",   build_ba_lhr_sjc,   129162),
        ("KLM AMS-TPA",  build_klm_ams_tpa,  79771),
        ("JX TPE-SJC",   build_jx_tpe_sjc,   69736),
    ]

    print(f"\n ENGINE VALIDATION \n")
    all_ok = True
    for name, builder, target in tests:
        result = builder()
        total = result.total_forecast
        delta = total - target
        pct = abs(delta / target * 100) if target else 0
        ok = pct < 0.05
        all_ok = all_ok and ok
        print(f"  {'' if ok else ''} {name:20s}  fcst={total:>10,.0f}  target={target:>10,d}  "
              f"={delta:>+7,.0f} ({pct:.3f}%)  LF={result.load_factor:.1%}")

    print(f"\n  {'ALL VALIDATED ' if all_ok else 'VALIDATION FAILURES'}")

    # Generate output workbooks
    print(f"\n WORKBOOK GENERATION \n")
    out_dir = "/mnt/user-data/outputs"
    os.makedirs(out_dir, exist_ok=True)

    for name, builder, target in tests:
        result = builder()
        tag = name.replace(' ', '_').replace('-', '_')
        path = os.path.join(out_dir, f"{tag}_FULL.xlsx")
        sens = {
            'capture': sensitivity_analysis(result, 'capture'),
            'stimulation': sensitivity_analysis(result, 'stimulation'),
            'growth': sensitivity_analysis(result, 'growth'),
        }
        bc = business_case_test(result)
        write_output(result, path, sens_data=sens, bc_data=bc)
        print(f"   {os.path.basename(path):40s}  {result.total_forecast:>10,.0f} pax  LF={result.load_factor:.1%}")

    # Round-trip test
    print(f"\n ROUND-TRIP READER TEST \n")
    for name, builder, target in tests:
        tag = name.replace(' ', '_').replace('-', '_')
        path = os.path.join(out_dir, f"{tag}_FULL.xlsx")
        original = builder()
        readback = read_workbook(path)
        orig_t = original.total_forecast
        read_t = readback.total_forecast
        delta = abs(read_t - orig_t)
        pct = delta / orig_t * 100 if orig_t else 0
        ok = pct < 1.0  # 1% tolerance for round-trip (scale detection may cause minor variance)
        print(f"  {'' if ok else ''} {name:20s}  orig={orig_t:>10,.0f}  read={read_t:>10,.0f}  ={pct:.2f}%")

    # Generate blank template
    tmpl_path = os.path.join(out_dir, "QSI_Template_Blank.xlsx")
    generate_blank_template(tmpl_path)
    print(f"\n   Blank template: {tmpl_path}")

    print(f"\n{''*80}")


def run_from_file(input_path, target_lf=0, do_bc=False, output_path=None):
    print(f"\n  Reading: {input_path}")
    result = read_workbook(input_path)
    c = result.config
    print(f"  Route: {c.carrier} {c.origin}-{c.destination} ({c.frequency}x {c.aircraft})")
    print(f"  P2P segments: {len(result.p2p_segments)}, Connecting hubs: {len(result.connecting)}")
    print(f"  Forecast: {result.total_forecast:,.0f} pax, LF: {result.load_factor:.1%}")

    if target_lf > 0:
        print(f"\n  Calibrating to {target_lf:.0%} LF...")
        result = calibrate_capture(result, target_lf=target_lf)
        print(f"  Adjusted: {result.total_forecast:,.0f} pax, LF: {result.load_factor:.1%}")

    result.print_summary()

    sens = {
        'capture': sensitivity_analysis(result, 'capture'),
        'stimulation': sensitivity_analysis(result, 'stimulation'),
        'growth': sensitivity_analysis(result, 'growth'),
    }
    bc = business_case_test(result) if do_bc else None

    if not output_path:
        base = os.path.splitext(os.path.basename(input_path))[0]
        output_path = f"/mnt/user-data/outputs/{base}_OUTPUT.xlsx"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    write_output(result, output_path, sens_data=sens, bc_data=bc)
    print(f"\n  Output: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Avia Solutions QSI Forecast Tool v2.0")
    parser.add_argument('input', nargs='?', help='Input workbook path')
    parser.add_argument('--target-lf', type=float, default=0, help='Target load factor (e.g. 0.85)')
    parser.add_argument('--business-case', action='store_true', help='Run business case assessment')
    parser.add_argument('--generate-template', action='store_true', help='Generate blank template')
    parser.add_argument('-o', '--output', help='Output path')
    args = parser.parse_args()

    if args.generate_template:
        p = args.output or '/mnt/user-data/outputs/QSI_Template_Blank.xlsx'
        os.makedirs(os.path.dirname(p), exist_ok=True)
        generate_blank_template(p)
        print(f"Template generated: {p}")
    elif args.input:
        run_from_file(args.input, args.target_lf, args.business_case, args.output)
    else:
        run_validation()
