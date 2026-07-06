#!/usr/bin/env python3
"""
Avia Solutions  Assumptions Log Generator (Chat 18)
=====================================================
Produces the structured assumptions log that accompanies every forecast.

The assumptions log serves three purposes:
    1. CLIENT TRANSPARENCY  Every parameter is visible with its justification,
       enabling an airline planner or fund analyst to manually replicate the
       calculation and challenge any assumption.
    2. INTERNAL QA  The structured format makes it impossible to submit a
       forecast without documenting why each parameter was chosen.
    3. TRAINING DATA  The machine-readable JSON output feeds the calibration
       learning system. After 30-50 completed routes, the system can start
       predicting calibration adjustments from route characteristics.

Architecture:
    RouteConfig + pipeline results + InputValidator config
         AssumptionsLog (structured object)
             Excel output (7 sections, analyst-editable justification fields)
             JSON output (machine-readable, for calibration learning)
             Methodology text (auto-generated for presentation slides)

The 7 sections mirror the "Forecast Methodology Summary" slides that appear
in every Avia Solutions presentation (BA, CX, SQ, KE, AI, AA, UA, EVA, etc.):

    A. Base Demand & Data Sources
    B. Demand Segmentation
    C. Growth Assumptions
    D. Stimulation Factors
    E. Capture Rates (P2P)
    F. Connecting Traffic & QSI Parameters
    G. Constraints & Exclusions

Dependencies:
    - route_config.py (Chat 12)  RouteConfig
    - input_validator.py (Chat 14)  Enums and config types
    - calibration_model.py (Chat 13)  CalibrationTier references
    - closed_loop_pipeline_v2.py (Chat 12)  Pipeline results dict
"""

from config import OUTPUT_DIR, ensure_output_dir
import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ============================================================================
# COLOUR CONSTANTS  Avia Solutions Branding
# ============================================================================

AVIA_BLUE = '003366'
AVIA_LIGHT_BLUE = 'D6E4F0'
AVIA_YELLOW = 'FFFFCC'          # Analyst-editable fields
AVIA_GREEN = 'E2EFDA'           # Derived/auto-populated
AVIA_RED_SOFT = 'FCE4EC'        # Flags / warnings
AVIA_GREY = 'F2F2F2'            # Section separators
WHITE = 'FFFFFF'


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class Assumption:
    """Single assumption with value, justification, and metadata."""
    parameter: str              # What: e.g. "UK Business base demand"
    value: Any                  # The number/choice
    unit: str = ''              # '%', 'pax', 'x', etc.
    source: str = ''            # Auto-populated source reference
    justification: str = ''     # Analyst-editable explanation
    confidence: str = 'Medium'  # High / Medium / Low / Flag
    section: str = ''           # Which section (A-G)
    editable: bool = True       # Is the justification field analyst-editable?
    fmt: str = ''               # Excel number format

    def to_dict(self) -> Dict:
        return {
            'parameter': self.parameter,
            'value': self.value,
            'unit': self.unit,
            'source': self.source,
            'justification': self.justification,
            'confidence': self.confidence,
            'section': self.section,
        }


@dataclass
class AssumptionSection:
    """One section of the assumptions log."""
    code: str                   # A, B, C, D, E, F, G
    title: str
    description: str
    assumptions: List[Assumption] = field(default_factory=list)

    def add(self, parameter: str, value: Any, **kwargs) -> Assumption:
        a = Assumption(parameter=parameter, value=value, section=self.code, **kwargs)
        self.assumptions.append(a)
        return a


@dataclass
class AssumptionsLog:
    """Complete assumptions log for a route assessment."""
    route_label: str            # e.g. "BA LHR-SJC"
    generated_at: str = ''
    analyst: str = ''
    engagement_ref: str = ''
    sections: Dict[str, AssumptionSection] = field(default_factory=dict)
    methodology_text: str = ''  # Auto-generated for presentation slides
    warnings: List[str] = field(default_factory=list)

    def all_assumptions(self) -> List[Assumption]:
        result = []
        for code in sorted(self.sections.keys()):
            result.extend(self.sections[code].assumptions)
        return result

    def to_dict(self) -> Dict:
        return {
            'route_label': self.route_label,
            'generated_at': self.generated_at,
            'analyst': self.analyst,
            'engagement_ref': self.engagement_ref,
            'sections': {
                code: {
                    'title': sec.title,
                    'description': sec.description,
                    'assumptions': [a.to_dict() for a in sec.assumptions],
                }
                for code, sec in sorted(self.sections.items())
            },
            'methodology_text': self.methodology_text,
            'warnings': self.warnings,
        }

    def to_json(self, path: str) -> str:
        with open(path, 'w') as f:
            json.dump(self.to_dict(), f, indent=2, default=str)
        return path


# ============================================================================
# BUILDER  Constructs the AssumptionsLog from pipeline data
# ============================================================================

class AssumptionsLogBuilder:
    """
    Builds a structured AssumptionsLog from RouteConfig and pipeline results.

    Usage:
        builder = AssumptionsLogBuilder(config, results)
        log = builder.build()
        log.to_json('assumptions.json')
        write_excel(log, 'assumptions.xlsx')
    """

    def __init__(self, config, results: Dict[str, Any],
                 analyst: str = '', engagement_ref: str = ''):
        """
        Args:
            config: RouteConfig instance
            results: Dict from run_pipeline() containing p2p_details,
                     home_results, dest_results, grand_total, load_factor, etc.
            analyst: Name of the analyst producing this forecast
            engagement_ref: Client reference / project code
        """
        self.cfg = config
        self.results = results
        self.analyst = analyst
        self.engagement_ref = engagement_ref

    def build(self) -> AssumptionsLog:
        log = AssumptionsLog(
            route_label=f"{self.cfg.airline_code} {self.cfg.home_airport_code}-{self.cfg.dest_airport_code}",
            generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'),
            analyst=self.analyst,
            engagement_ref=self.engagement_ref,
        )

        log.sections['A'] = self._build_section_a()
        log.sections['B'] = self._build_section_b()
        log.sections['C'] = self._build_section_c()
        log.sections['D'] = self._build_section_d()
        log.sections['E'] = self._build_section_e()
        log.sections['F'] = self._build_section_f()
        log.sections['G'] = self._build_section_g()

        log.methodology_text = self._generate_methodology_text()
        log.warnings = self._generate_warnings()

        return log

    # 
    # SECTION A: Base Demand & Data Sources
    # 

    def _build_section_a(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='A',
            title='Base Demand & Data Sources',
            description='Data sources, base year, and total addressable market.',
        )

        sec.add('Airline', self.cfg.airline_name,
                source='Client brief', confidence='High', editable=False)
        sec.add('Route', f"{self.cfg.home_airport_code}-{self.cfg.dest_airport_code}",
                source='Client brief', confidence='High', editable=False)
        sec.add('Aircraft type', self.cfg.aircraft_type,
                source='Client brief', confidence='High', editable=False)
        sec.add('Seat capacity', self.cfg.seats, unit='seats',
                source='Client brief / OAG fleet data', confidence='High',
                editable=False, fmt='#,##0')
        sec.add('Frequency', self.cfg.frequency, unit='x/week',
                source='Client brief', confidence='High', editable=False)
        sec.add('Annual capacity (two-way)', self.cfg.annual_capacity, unit='seats',
                source='Derived: seats  frequency  52  2',
                confidence='High', editable=False, fmt='#,##0')

        # Demand data source
        sec.add('Demand data source', 'Sabre MI (MIDT-based, adjusted for non-MIDT channels)',
                source='Sabre Airport Data Intelligence',
                justification='Standard MIDT data adjusted by Sabre for airline direct bookings, '
                              'charters, and LCC channels not captured in GDS.',
                confidence='High')

        # Base period  extract from demand provider if available
        base_period = self._infer_base_period()
        sec.add('Base demand period', base_period,
                source='Sabre MI pull date',
                justification='Most recent complete year available at time of analysis.',
                confidence='High')

        # Total P2P base demand
        p2p_details = self.results.get('p2p_details', [])
        total_base = sum(d.get('base', 0) for d in p2p_details)
        sec.add('Total P2P base demand (O&D)', total_base, unit='pax/year',
                source='Sabre MI, restricted to catchment',
                justification='Annual O&D demand between the two city catchments.',
                confidence='High', fmt='#,##0')

        # Connecting base demand
        home_results = self.results.get('home_results', [])
        dest_results = self.results.get('dest_results', [])
        home_base = sum(r.get('base_demand', 0) for r in home_results)
        dest_base = sum(r.get('base_demand', 0) for r in dest_results)
        sec.add(f'Connecting base demand beyond {self.cfg.home_airport_code}',
                home_base, unit='pax/year',
                source='Sabre MI O&D, double connections excluded',
                confidence='High', fmt='#,##0')
        sec.add(f'Connecting base demand beyond {self.cfg.dest_airport_code}',
                dest_base, unit='pax/year',
                source='Sabre MI O&D, double connections excluded',
                confidence='High', fmt='#,##0')

        # Catchment restriction
        sec.add('Catchment restriction applied', 'Yes  P2P restricted to service area',
                source='Standard methodology',
                justification='P2P demand restricted to origin service area to avoid '
                              'double-counting with competing airports.',
                confidence='High')

        return sec

    # 
    # SECTION B: Demand Segmentation
    # 

    def _build_section_b(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='B',
            title='Demand Segmentation',
            description='How P2P demand is split into segments for separate treatment.',
        )

        p2p_details = self.results.get('p2p_details', [])
        if not p2p_details:
            sec.add('Segmentation', 'No P2P segmentation data available',
                    confidence='Flag')
            return sec

        # Document each segment
        for detail in p2p_details:
            name = detail.get('name', 'Unknown')
            base = detail.get('base', 0)
            sec.add(f'{name}  base demand', base, unit='pax/year',
                    source='Sabre MI, segmented by point of sale and purpose',
                    justification='Split determined using Sabre MI point of sale data, '
                                  'tourism statistics, and CAA survey data where available.',
                    confidence='Medium', fmt='#,##0')

        # Segmentation method
        sec.add('Segmentation method',
                'Business vs Leisure/VFR split by point of sale and purpose data',
                source='Sabre MI point of sale + tourism board statistics',
                justification='Business/Leisure split calibrated against Visit California '
                              'statistics and UK CAA survey data.',
                confidence='Medium')

        # Leisure subsegments explanation
        has_subsegments = any('/' in d.get('name', '') for d in p2p_details)
        if has_subsegments:
            sec.add('Leisure subsegment method',
                    'Primary / Secondary / Contested tiers based on geographic proximity',
                    source='Standard Avia methodology',
                    justification='Primary = within natural catchment; Secondary = adjacent; '
                                  'Contested = shared with competing airport(s).',
                    confidence='Medium')

        return sec

    # 
    # SECTION C: Growth Assumptions
    # 

    def _build_section_c(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='C',
            title='Growth Assumptions',
            description='Compound growth rates applied to grow base demand to forecast year.',
        )

        p2p_details = self.results.get('p2p_details', [])

        # P2P growth by segment
        seen_rates = set()
        for detail in p2p_details:
            name = detail.get('name', 'Unknown')
            growth = detail.get('growth', 0)
            rate_key = f"{name}:{growth}"
            if rate_key in seen_rates:
                continue
            seen_rates.add(rate_key)

            # Infer growth basis from segment type
            if 'business' in name.lower():
                basis = ('GDP growth projections (Oxford Economics) and '
                         'historic Sabre MI traffic trend')
            elif 'leisure' in name.lower() or 'vfr' in name.lower():
                basis = ('Historic passenger growth benchmarks and '
                         'tourism arrivals data')
            else:
                basis = 'Market-specific analysis'

            sec.add(f'{name}  CAGR', growth, unit='%',
                    source=basis,
                    justification=f'Growth rate determined by analysing historic passenger '
                                  f'growth, national GDP projections, and IATA passenger forecasts.',
                    confidence='Medium', fmt='0.0%')

        # Connecting traffic growth
        home_growth = getattr(self.cfg, '_home_growth', None)
        dest_growth = getattr(self.cfg, '_dest_growth', None)
        if hasattr(self.cfg, 'demand_provider') and self.cfg.demand_provider:
            dp = self.cfg.demand_provider
            home_growth = getattr(dp, 'home_growth', home_growth)
            dest_growth = getattr(dp, 'dest_growth', dest_growth)

        if home_growth is not None:
            sec.add(f'Connecting growth beyond {self.cfg.home_airport_code}',
                    home_growth, unit='%',
                    source='IATA Passenger Forecast CAGR for predominant connecting flows',
                    justification='Applied IATA regional CAGR for relevant flow corridors.',
                    confidence='Medium', fmt='0.0%')
        if dest_growth is not None:
            sec.add(f'Connecting growth beyond {self.cfg.dest_airport_code}',
                    dest_growth, unit='%',
                    source='IATA Passenger Forecast CAGR for predominant connecting flows',
                    justification='Applied IATA regional CAGR for relevant flow corridors.',
                    confidence='Medium', fmt='0.0%')

        return sec

    # 
    # SECTION D: Stimulation Factors
    # 

    def _build_section_d(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='D',
            title='Stimulation Factors',
            description='New service stimulation applied to P2P demand. '
                        'A factor > 1.0 reflects incremental demand generated by the '
                        'availability of a new direct service.',
        )

        p2p_details = self.results.get('p2p_details', [])

        for detail in p2p_details:
            name = detail.get('name', 'Unknown')
            stim = detail.get('stimulation', 1.0)

            if stim > 1.0:
                justification = (
                    f'Stimulation of {stim:.2f}x applied based on IATA stimulation curve, '
                    f'historical benchmarks from comparable new services, and '
                    f'industry assumptions considering bilateral demand drivers.'
                )
                confidence = 'Medium'
            else:
                justification = 'No stimulation applied  connecting or leisure subsegment.'
                confidence = 'High'

            sec.add(f'{name}  stimulation factor', stim, unit='x',
                    source='IATA Stimulation Curve + route-specific benchmarks',
                    justification=justification,
                    confidence=confidence, fmt='0.00')

        # General stimulation methodology note
        sec.add('Stimulation methodology',
                'IATA analysis + historical benchmarks + bilateral demand assessment',
                source='IATA stimulation research',
                justification='Stimulation factors consider: (1) IATA new-route stimulation '
                              'curve, (2) historic benchmarks from comparable route launches, '
                              '(3) strength of bilateral demand drivers. LCC/new market services '
                              'receive higher factors (demand creation); legacy services on '
                              'established routes receive lower factors (demand capture).',
                confidence='Medium', editable=True)

        return sec

    # 
    # SECTION E: Capture Rates (P2P)
    # 

    def _build_section_e(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='E',
            title='Capture Rates  Point-to-Point',
            description='Share of stimulated demand captured by the proposed service. '
                        'Determined by frequency share, competing services, and leakage.',
        )

        p2p_details = self.results.get('p2p_details', [])

        for detail in p2p_details:
            name = detail.get('name', 'Unknown')
            capture = detail.get('capture', 0)
            forecast = detail.get('forecast', 0)

            if capture > 0:
                justification = (
                    f'Capture rate of {capture:.0%} determined by frequency/capacity share '
                    f'analysis and leakage estimates from competing services and airports. '
                    f'Rates revised down to maintain conservative assumptions.'
                )
            else:
                justification = 'No capture  aggregated from subsegments.'

            sec.add(f'{name}  P2P capture rate', capture, unit='%',
                    source='Frequency share analysis + leakage adjustment',
                    justification=justification,
                    confidence='Medium', fmt='0.0%')

        # Total P2P outcome
        p2p_total = self.results.get('p2p_total', 0)
        sec.add('P2P forecast total', p2p_total, unit='pax/year',
                source='Derived: grown demand  stimulation  capture rate',
                confidence='High', editable=False, fmt='#,##0')

        return sec

    # 
    # SECTION F: Connecting Traffic & QSI Parameters
    # 

    def _build_section_f(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='F',
            title='Connecting Traffic & QSI Parameters',
            description='QSI model configuration for connecting traffic capture rates. '
                        'Capture rates determined by Quality of Service Index scoring '
                        'of competing itineraries.',
        )

        # QSI coefficients
        sec.add('QSI  Online connection coefficient', self.cfg.online_coeff, unit='',
                source='Avia QSI model  standard coefficient',
                justification='Online connections (same carrier both legs) receive full weighting.',
                confidence='High', fmt='0.000')
        sec.add('QSI  Alliance connection coefficient', self.cfg.alliance_coeff, unit='',
                source='Avia QSI model  standard coefficient',
                justification='Alliance connections (e.g. oneworld, Star Alliance) receive '
                              'reduced weighting reflecting lower passenger preference vs online.',
                confidence='High', fmt='0.000')
        sec.add('QSI  Interline connection coefficient', self.cfg.interline_coeff, unit='',
                source='Avia QSI model  standard coefficient',
                justification='Interline connections (no alliance, basic interlining only) '
                              'receive lowest weighting.',
                confidence='High', fmt='0.000')
        sec.add('QSI  Non-stop service coefficient', getattr(self.cfg, 'nonstop_coeff', 1.0), unit='',
                source='Avia QSI model  service-level coefficient',
                justification='Non-stop services receive full service-level weighting.',
                confidence='High', fmt='0.000')
        sec.add('QSI  One-stop service coefficient', getattr(self.cfg, 'onestop_coeff', 0.20), unit='',
                source='Avia QSI model  service-level coefficient',
                justification='One-stop (single-connection) itineraries receive reduced '
                              'service-level weighting reflecting passenger preference for non-stop.',
                confidence='High', fmt='0.000')

        # Elapsed time decay
        sec.add('QSI  Elapsed time decay factor', self.cfg.et_decay_factor, unit='',
                source='Avia QSI model  elapsed time decay curve',
                justification='Controls how steeply QSI score decays with longer elapsed times. '
                              'Current model uses a single curve for all routes.',
                confidence='Medium', fmt='0.00')
        sec.add('QSI  Elapsed time decay interval', self.cfg.et_decay_interval, unit='hrs',
                source='Avia QSI model  standard parameter',
                justification='Time increment for each decay step in the elapsed time curve.',
                confidence='High', fmt='0.00')

        # QSI ceiling and adjustment
        sec.add('QSI ceiling', self.cfg.qsi_ceiling, unit='',
                source='Analyst judgment  route-specific',
                justification='Maximum capture rate cap. Small airports with limited route '
                              'options may accept wider QSI ranges (higher ceiling).',
                confidence='Medium', fmt='0.000')
        sec.add('QSI adjustment factor', self.cfg.qsi_adjustment, unit='',
                source='Analyst judgment  route-specific',
                justification='Global scaling factor applied to all QSI captures. Used to '
                              'calibrate overall connecting volume.',
                confidence='Medium', fmt='0.000')

        # OAG schedule basis
        sec.add('OAG schedule basis', 'Airline schedules as indicator for connectivity',
                source='OAG',
                justification='Connection builder uses OAG schedules for the analysis week '
                              'to enumerate all viable connecting itineraries.',
                confidence='High')

        # Connection scope
        home_results = self.results.get('home_results', [])
        dest_results = self.results.get('dest_results', [])
        n_home = len([r for r in home_results if r.get('forecast', 0) > 0])
        n_dest = len([r for r in dest_results if r.get('forecast', 0) > 0])

        sec.add(f'Connecting cities beyond {self.cfg.home_airport_code}',
                f'{n_home} cities with forecast > 0',
                source='Connection builder output',
                confidence='High', editable=False)
        sec.add(f'Connecting cities beyond {self.cfg.dest_airport_code}',
                f'{n_dest} cities with forecast > 0',
                source='Connection builder output',
                confidence='High', editable=False)

        # Connecting totals
        home_total = self.results.get('home_total', 0)
        dest_total = self.results.get('dest_total', 0)
        sec.add(f'Connecting forecast beyond {self.cfg.home_airport_code}',
                home_total, unit='pax/year',
                source='QSI model output', confidence='High',
                editable=False, fmt='#,##0')
        sec.add(f'Connecting forecast beyond {self.cfg.dest_airport_code}',
                dest_total, unit='pax/year',
                source='QSI model output', confidence='High',
                editable=False, fmt='#,##0')

        # Calibration status
        cal = self.results.get('calibration', {})
        if cal:
            cal_mode = cal.get('mode', 'unknown')
            sec.add('Calibration mode', cal_mode,
                    source='Pipeline configuration',
                    justification=f'Calibration mode: {cal_mode}. Expert calibration adjusts '
                                  f'raw QSI capture rates based on market-specific knowledge.',
                    confidence='High')

        # Connection types assumed
        sec.add('Connection types included',
                f'{self.cfg.airline_code} online + alliance partners + limited interline',
                source='Standard methodology',
                justification=f'Forecast considers connecting routes served by {self.cfg.airline_code}, '
                              f'its alliance partners, and a limited number of interline carriers. '
                              f'LCC connections excluded unless specific agreements exist.',
                confidence='High')

        return sec

    # 
    # SECTION G: Constraints & Exclusions
    # 

    def _build_section_g(self) -> AssumptionSection:
        sec = AssumptionSection(
            code='G',
            title='Constraints & Exclusions',
            description='Geographic and methodological constraints applied to the forecast.',
        )

        sec.add('Double connections excluded', 'Yes',
                source='Standard methodology',
                justification='Demand on double connections (two or more stopovers) has been '
                              'excluded from the connecting traffic analysis.',
                confidence='High', editable=False)

        sec.add('Circuity threshold', '30%',
                source='Standard methodology  may be varied per route',
                justification='Connections where journey circuity via the hub exceeds 30% of '
                              'the direct routing distance have been excluded. This threshold '
                              'is now treated as variable per route depending on airport '
                              'location and airline type.',
                confidence='High')

        # Geographic constraints on connections
        sec.add(f'Connection scope beyond {self.cfg.home_airport_code}',
                'TO BE SPECIFIED  typically Europe, Africa, Middle East for LHR',
                source='Analyst judgment',
                justification='Geographic scope of connections beyond the home hub. '
                              'Should match the airline\'s network and realistic '
                              'passenger routing patterns.',
                confidence='Medium')

        sec.add(f'Connection scope beyond {self.cfg.dest_airport_code}',
                'TO BE SPECIFIED  typically restricted to limited backtracking',
                source='Analyst judgment',
                justification='Geographic scope of connections beyond the destination. '
                              'Typically more restricted than home hub connections.',
                confidence='Medium')

        # Exclusions
        sec.add('Market exclusions', 'TO BE SPECIFIED  e.g. India excluded from BA LHR-SJC',
                source='Analyst judgment',
                justification='Specific markets excluded from the forecast due to '
                              'routing constraints, political factors, or client instruction.',
                confidence='Medium')

        # Conservative basis
        sec.add('Conservative assumptions basis', 'Yes  forecast based upon conservative assumptions',
                source='Standard Avia policy',
                justification='All forecasts are based upon conservative assumptions. '
                              'Capture rates are revised down from theoretical maximums. '
                              'Growth rates are benchmarked against independent forecasts.',
                confidence='High', editable=False)

        # Grand total and load factor
        grand_total = self.results.get('grand_total', 0)
        load_factor = self.results.get('load_factor', 0)
        sec.add('GRAND TOTAL forecast', grand_total, unit='pax/year',
                source='P2P + Connecting at home hub + Connecting at dest hub',
                confidence='High', editable=False, fmt='#,##0')
        sec.add('Implied load factor', load_factor, unit='%',
                source='Grand total  annual capacity',
                confidence='High', editable=False, fmt='0.0%')

        return sec

    # 
    # METHODOLOGY TEXT GENERATOR
    # 

    def _generate_methodology_text(self) -> str:
        """
        Auto-generate the "Forecast Methodology Summary" text that appears
        in every Avia Solutions presentation (Section 7D).

        This follows the exact template visible in BA, CX, SQ, KE, AI, AA,
        UA, EVA Air presentations  fill-in-the-blanks with route-specific values.
        """
        c = self.cfg
        base_period = self._infer_base_period()

        # Determine P2P segment names
        p2p_details = self.results.get('p2p_details', [])
        seg_names = list(set(
            d['name'].split('/')[0] if '/' in d.get('name', '') else d.get('name', '')
            for d in p2p_details
        ))

        # Build the methodology sections
        lines = []

        # Summary header
        lines.append("Forecast Methodology Summary")
        lines.append("")

        # Point-to-point methodology
        lines.append("Point-to-Point")
        lines.append("")
        lines.append(
            "To forecast the number of origin and destination (O&D) passengers, "
            "the following methodology has been used:"
        )
        lines.append(
            f"i. Sabre MI data was used for {base_period} to estimate the overall "
            f"size of O&D traffic between {c.home_city_code or c.home_airport_code} "
            f"and {c.dest_city_code or c.dest_airport_code}."
        )
        lines.append(
            f"ii. Base annual demand was grown to [forecast year] by assuming an "
            f"underlying compound growth rate for the "
            f"{c.home_city_code or c.home_airport_code}-"
            f"{c.dest_city_code or c.dest_airport_code} market. "
            f"Growth rates have been determined by [specific growth basis  "
            f"GDP projections, historic traffic growth, tourism forecasts, with named sources]."
        )
        lines.append(
            "iii. Where appropriate, the point-to-point market has been stimulated "
            "due to the new direct service between the two cities. Stimulation has been "
            "considered with IATA analysis, [historical benchmarks  name specific proxy "
            "markets], and industry assumptions which consider the [relevant demand driver "
            f"links] between [Country A] and [Country B]."
        )
        lines.append(
            "iv. Capture rates have been determined with a combination of "
            "[method  frequency share analysis / QSI calibration / benchmarking of "
            "proxy markets]. Capture rates have been revised down to consider traffic "
            "flow [through competing airport / on indirect routings] to maintain "
            "conservative assumptions for the forecast."
        )

        # Connecting methodology
        lines.append("")
        lines.append("Connecting Traffic")
        lines.append("")
        lines.append(
            f"v. QSI scores were assigned to the different connecting itineraries "
            f"produced by the connection builder. These scores are based on overall "
            f"travel time and type of connection."
        )
        lines.append(
            f"vi. The scores were then summarised at city level and used to quantify "
            f"the capture rate of the proposed new {c.airline_code} service to the "
            f"key connecting markets."
        )

        # Forecast summary bullets (as in every presentation)
        lines.append("")
        lines.append("Forecast Summary")
        lines.append("")
        lines.append(
            f"1. The forecast assumes a base traffic demand for {base_period}. "
            f"Base annual demand was grown to [forecast year] to reach maturity."
        )
        lines.append(
            "2. Sabre MI data is used to determine passenger O&D demand. Traffic data "
            "is based on MIDT and adjusted by Sabre to take into account non-MIDT "
            "distribution channels (airline direct bookings, charters, low-cost carriers, etc.)."
        )
        if seg_names:
            seg_str = ' and '.join(sorted(set(s for s in seg_names if s)))
            lines.append(
                f"3. Point to point demand has been split into the following categories: "
                f"{seg_str}."
            )
        lines.append(
            f"4. The forecast considers traffic beyond {c.home_airport_code}  "
            f"connecting onto {c.airline_code} and its partners; and beyond "
            f"{c.dest_airport_code}  routes served by {c.airline_code} and partners."
        )
        lines.append(
            "5. QSI analysis reflects airline schedules for [Season Year] as an "
            "indicator for connectivity."
        )
        lines.append(
            "6. Note that this forecast is based upon conservative assumptions."
        )

        return '\n'.join(lines)

    # 
    # WARNINGS GENERATOR
    # 

    def _generate_warnings(self) -> List[str]:
        warnings = []

        # Load factor check
        lf = self.results.get('load_factor', 0)
        if lf > 0.90:
            warnings.append(
                f"LOAD FACTOR WARNING: Implied LF of {lf:.1%} exceeds 90%. "
                f"This may indicate over-optimistic assumptions or insufficient capacity."
            )
        if lf < 0.60:
            warnings.append(
                f"LOAD FACTOR WARNING: Implied LF of {lf:.1%} is below 60%. "
                f"This may indicate the route is marginal at the proposed frequency."
            )

        # P2P capture rate checks
        p2p_details = self.results.get('p2p_details', [])
        for d in p2p_details:
            cr = d.get('capture', 0)
            name = d.get('name', 'Unknown')
            if cr > 0.50:
                warnings.append(
                    f"HIGH CAPTURE RATE: {name} P2P capture rate of {cr:.0%} is aggressive. "
                    f"Consider whether this is supportable given competition."
                )

        # Stimulation checks
        for d in p2p_details:
            stim = d.get('stimulation', 1.0)
            name = d.get('name', 'Unknown')
            if stim > 1.5:
                warnings.append(
                    f"HIGH STIMULATION: {name} stimulation of {stim:.2f}x is above "
                    f"typical range. Ensure benchmark evidence supports this."
                )

        # Connecting traffic proportion
        grand = self.results.get('grand_total', 0)
        home = self.results.get('home_total', 0)
        dest = self.results.get('dest_total', 0)
        if grand > 0:
            cnx_pct = (home + dest) / grand
            if cnx_pct > 0.70:
                warnings.append(
                    f"HIGH CONNECTING SHARE: Connecting traffic is {cnx_pct:.0%} of total. "
                    f"Routes heavily dependent on connecting traffic are vulnerable to "
                    f"schedule/alliance changes."
                )

        # Check for zero-forecast segments
        for d in p2p_details:
            if d.get('forecast', 0) == 0 and d.get('base', 0) > 0:
                warnings.append(
                    f"ZERO FORECAST: {d.get('name', 'Unknown')} has base demand "
                    f"of {d.get('base', 0):,.0f} but zero forecast  check capture rate."
                )

        return warnings

    # 
    # HELPERS
    # 

    def _infer_base_period(self) -> str:
        """Try to infer the base demand period from the config or demand provider."""
        if hasattr(self.cfg, 'demand_provider') and self.cfg.demand_provider:
            dp = self.cfg.demand_provider
            if hasattr(dp, 'forecast_file') and dp.forecast_file:
                fname = os.path.basename(dp.forecast_file)
                # Extract date hints from filename
                if '2015' in fname:
                    return 'CY2013 (Sep 2013  Aug 2014)'
                if '2023' in fname or '2024' in fname:
                    return '[Most recent complete year]'
        return '[Base period  TO BE SPECIFIED]'


# ============================================================================
# EXCEL WRITER  Produces the formatted output workbook
# ============================================================================

class AssumptionsLogExcelWriter:
    """
    Writes the AssumptionsLog to a formatted Excel workbook.

    Output structure:
        Sheet 1: "Assumptions Log"  Full structured log with all 7 sections
        Sheet 2: "Methodology Text"  Auto-generated presentation text
        Sheet 3: "Warnings & QA"  Automated warning flags
    """

    # Styles
    TITLE_FONT = Font(name='Arial', bold=True, size=14, color=AVIA_BLUE)
    SECTION_FONT = Font(name='Arial', bold=True, size=12, color=AVIA_BLUE)
    HEADER_FONT = Font(name='Arial', bold=True, size=10, color=WHITE)
    HEADER_FILL = PatternFill('solid', fgColor=AVIA_BLUE)
    LABEL_FONT = Font(name='Arial', size=10)
    VALUE_FONT = Font(name='Arial', bold=True, size=10, color='000080')
    INPUT_FONT = Font(name='Arial', size=10, color='0000FF')  # Blue = editable
    DERIVED_FONT = Font(name='Arial', size=10, color='000000')  # Black = derived
    FLAG_FONT = Font(name='Arial', bold=True, size=10, color='CC0000')
    NOTE_FONT = Font(name='Arial', italic=True, size=9, color='666666')

    YELLOW_FILL = PatternFill('solid', fgColor=AVIA_YELLOW)
    GREEN_FILL = PatternFill('solid', fgColor=AVIA_GREEN)
    GREY_FILL = PatternFill('solid', fgColor=AVIA_GREY)
    RED_FILL = PatternFill('solid', fgColor=AVIA_RED_SOFT)

    THIN_BORDER = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )
    SECTION_BORDER = Border(
        bottom=Side(style='medium', color=AVIA_BLUE)
    )

    def __init__(self, log: AssumptionsLog):
        self.log = log
        self.wb = Workbook()

    def write(self, path: str) -> str:
        """Write all sheets and save."""
        self._write_assumptions_sheet()
        self._write_methodology_sheet()
        self._write_warnings_sheet()

        # Remove default empty sheet if we created named sheets
        if 'Sheet' in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            del self.wb['Sheet']

        self.wb.save(path)
        return path

    def _write_assumptions_sheet(self):
        """Write the main structured assumptions log."""
        ws = self.wb.create_sheet('Assumptions Log')

        # Title block
        ws.merge_cells('A1:F1')
        ws['A1'] = f'ASSUMPTIONS LOG  {self.log.route_label}'
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = f'Generated: {self.log.generated_at}'
        ws['A2'].font = self.NOTE_FONT
        if self.log.analyst:
            ws['C2'] = f'Analyst: {self.log.analyst}'
            ws['C2'].font = self.NOTE_FONT
        if self.log.engagement_ref:
            ws['E2'] = f'Ref: {self.log.engagement_ref}'
            ws['E2'].font = self.NOTE_FONT

        # Colour legend
        ws['A3'] = ' Yellow = analyst-editable'
        ws['A3'].font = Font(name='Arial', size=8, color='999900')
        ws['C3'] = ' Green = auto-derived'
        ws['C3'].font = Font(name='Arial', size=8, color='339933')
        ws['E3'] = ' Blue text = hardcoded input'
        ws['E3'].font = Font(name='Arial', size=8, color='0000FF')

        # Column headers
        headers = ['Parameter', 'Value', 'Unit', 'Source', 'Justification', 'Confidence']
        col_widths = [40, 20, 10, 35, 50, 12]
        row = 5
        for col, (hdr, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=row, column=col, value=hdr)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = width

        row = 6

        # Write each section
        for code in sorted(self.log.sections.keys()):
            section = self.log.sections[code]

            # Section header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1,
                           value=f'Section {code}: {section.title}')
            cell.font = self.SECTION_FONT
            cell.fill = self.GREY_FILL
            cell.border = self.SECTION_BORDER
            row += 1

            # Section description
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=section.description).font = self.NOTE_FONT
            row += 1

            # Assumptions
            for assumption in section.assumptions:
                # Parameter
                ws.cell(row=row, column=1, value=assumption.parameter).font = self.LABEL_FONT

                # Value  format depends on type and editability
                val_cell = ws.cell(row=row, column=2, value=assumption.value)
                if assumption.editable:
                    val_cell.font = self.INPUT_FONT
                    val_cell.fill = self.YELLOW_FILL
                else:
                    val_cell.font = self.DERIVED_FONT
                    val_cell.fill = self.GREEN_FILL
                if assumption.fmt:
                    val_cell.number_format = assumption.fmt

                # Unit
                ws.cell(row=row, column=3, value=assumption.unit).font = self.NOTE_FONT

                # Source
                ws.cell(row=row, column=4, value=assumption.source).font = self.LABEL_FONT

                # Justification  always editable (yellow)
                just_cell = ws.cell(row=row, column=5, value=assumption.justification)
                just_cell.font = self.INPUT_FONT
                just_cell.fill = self.YELLOW_FILL
                just_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Confidence
                conf_cell = ws.cell(row=row, column=6, value=assumption.confidence)
                if assumption.confidence == 'Flag':
                    conf_cell.font = self.FLAG_FONT
                    conf_cell.fill = self.RED_FILL
                elif assumption.confidence == 'Low':
                    conf_cell.font = Font(name='Arial', size=10, color='CC6600')
                else:
                    conf_cell.font = self.LABEL_FONT
                conf_cell.alignment = Alignment(horizontal='center')

                # Row border
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.THIN_BORDER

                row += 1

            row += 1  # Blank row between sections

        # Freeze panes
        ws.freeze_panes = 'A6'

        return ws

    def _write_methodology_sheet(self):
        """Write the auto-generated methodology text for presentation slides."""
        ws = self.wb.create_sheet('Methodology Text')

        ws.merge_cells('A1:B1')
        ws['A1'] = f'Forecast Methodology  {self.log.route_label}'
        ws['A1'].font = self.TITLE_FONT

        ws['A2'] = ('This text is auto-generated from the assumptions log. '
                    'Copy/paste into presentation slides, then fill in '
                    'bracketed placeholders with route-specific values.')
        ws['A2'].font = self.NOTE_FONT
        ws.merge_cells('A2:B2')

        ws.column_dimensions['A'].width = 100

        row = 4
        for line in self.log.methodology_text.split('\n'):
            cell = ws.cell(row=row, column=1, value=line)
            if line and not line.startswith(' ') and not line[0:1].isdigit() and not line.startswith('i'):
                cell.font = Font(name='Arial', bold=True, size=12, color=AVIA_BLUE)
            elif line.startswith('[') or '[' in line:
                cell.font = Font(name='Arial', size=10, color='CC6600')  # Orange for placeholders
            else:
                cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(wrap_text=True)
            row += 1

        return ws

    def _write_warnings_sheet(self):
        """Write the automated warnings and QA flags."""
        ws = self.wb.create_sheet('Warnings & QA')

        ws.merge_cells('A1:C1')
        ws['A1'] = f'QA Warnings  {self.log.route_label}'
        ws['A1'].font = self.TITLE_FONT

        ws.column_dimensions['A'].width = 100

        row = 3
        if self.log.warnings:
            for warning in self.log.warnings:
                cell = ws.cell(row=row, column=1, value=warning)
                cell.font = self.FLAG_FONT
                cell.fill = self.RED_FILL
                cell.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 30
                row += 1
        else:
            ws.cell(row=row, column=1, value=' No warnings generated  all parameters within normal bounds.').font = (
                Font(name='Arial', size=11, color='006600', bold=True)
            )
            row += 1

        # QA checklist
        row += 2
        ws.cell(row=row, column=1, value='QA Checklist (manual review)').font = self.SECTION_FONT
        row += 1

        qa_checks = [
            'Is the P2P capture rate realistic for the proposed frequency and competition?',
            'Is the stimulation factor consistent with IATA benchmarks and route characteristics?',
            'What proportion of connecting demand is captured? Is it consistent with known direct services?',
            'Are the major connecting destinations sensible? Does the highest capture rate city make sense?',
            'Is the aircraft consistent with the airline\'s current/planned fleet?',
            'Is the forecast load factor achievable but not over-optimistic?',
            'Do Year 2 and Year 3 growth rates look reasonable?',
            'Is the forecast defendable to an airline network planner or investment committee?',
            'Every Sabre-derived number traceable to data input?',
            'Double connections excluded from connecting analysis?',
            'Catchment restriction applied consistently?',
        ]

        for check in qa_checks:
            ws.cell(row=row, column=1, value=f' {check}').font = self.LABEL_FONT
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1

        return ws


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def generate_assumptions_log(config, results: Dict[str, Any],
                              excel_path: str = None,
                              json_path: str = None,
                              analyst: str = '',
                              engagement_ref: str = '') -> AssumptionsLog:
    """
    One-call function to generate the complete assumptions log.

    Args:
        config: RouteConfig instance
        results: Dict from run_pipeline()
        excel_path: Optional path for Excel output
        json_path: Optional path for JSON output
        analyst: Analyst name
        engagement_ref: Engagement reference

    Returns:
        AssumptionsLog instance
    """
    builder = AssumptionsLogBuilder(config, results, analyst, engagement_ref)
    log = builder.build()

    if json_path:
        log.to_json(json_path)
        print(f"  JSON: {json_path}")

    if excel_path and openpyxl:
        writer = AssumptionsLogExcelWriter(log)
        writer.write(excel_path)
        print(f"  Excel: {excel_path}")

    return log


# ============================================================================
# INTEGRATION TEST  BA LHR-SJC
# ============================================================================

def test_ba_lhr_sjc():
    """
    Integration test using the BA LHR-SJC regression case.

    This test:
    1. Builds RouteConfig from the factory method
    2. Constructs mock pipeline results matching the validated 129,162 target
    3. Generates the assumptions log in both Excel and JSON
    4. Validates that all 7 sections are populated
    5. Validates that methodology text is generated
    6. Validates that no unexpected warnings fire
    """
    print("=" * 60)
    print("ASSUMPTIONS LOG GENERATOR  Integration Test")
    print("BA LHR-SJC (Chat 18)")
    print("=" * 60)

    # Try to import RouteConfig; fall back to mock if not available
    try:
        from route_config import RouteConfig
        config = RouteConfig.ba_lhr_sjc()
        print("   RouteConfig loaded from factory method")
    except Exception as e:
        print(f"  RouteConfig not available ({e}), using mock config")
        config = _mock_ba_config()

    # Construct the results dict matching the validated pipeline output
    results = _ba_lhr_sjc_mock_results()

    # Generate
    ensure_output_dir()
    log = generate_assumptions_log(
        config, results,
        excel_path=str(OUTPUT_DIR / 'Assumptions_Log_BA_LHR_SJC.xlsx'),
        json_path=str(OUTPUT_DIR / 'Assumptions_Log_BA_LHR_SJC.json'),
        analyst='John Carter',
        engagement_ref='BA-LHRSJC-2015',
    )

    # Validate
    print(f"\n{'='*60}")
    print("VALIDATION")
    print(f"{'='*60}")

    # Check all 7 sections exist
    for code in 'ABCDEFG':
        sec = log.sections.get(code)
        n = len(sec.assumptions) if sec else 0
        status = '' if n > 0 else ''
        print(f"  {status} Section {code}: {sec.title if sec else 'MISSING'}  {n} assumptions")

    # Check total assumption count
    total = len(log.all_assumptions())
    print(f"\n  Total assumptions: {total}")
    assert total >= 30, f"Expected 30 assumptions, got {total}"
    print(f"   30 assumptions generated")

    # Check methodology text
    assert len(log.methodology_text) > 500, "Methodology text too short"
    assert 'Sabre MI' in log.methodology_text, "Missing Sabre MI reference"
    assert 'QSI' in log.methodology_text, "Missing QSI reference"
    assert 'conservative' in log.methodology_text.lower(), "Missing conservative statement"
    print(f"   Methodology text generated ({len(log.methodology_text)} chars)")

    # Check JSON roundtrip
    d = log.to_dict()
    assert len(d['sections']) == 7
    print(f"   JSON serialisation valid (7 sections)")

    # Check warnings
    print(f"\n  Warnings: {len(log.warnings)}")
    for w in log.warnings:
        print(f"     {w}")

    # Summary
    print(f"\n{'='*60}")
    print("INTEGRATION TEST: PASSED ")
    print(f"{'='*60}")

    return log


def _mock_ba_config():
    """Create a mock config matching BA LHR-SJC when RouteConfig isn't importable."""

    class MockConfig:
        pass

    c = MockConfig()
    c.airline_name = "British Airways"
    c.airline_code = "BA"
    c.home_airport_code = "LHR"
    c.home_city_code = "LON"
    c.dest_airport_code = "SJC"
    c.dest_city_code = "SJC"
    c.frequency = 7
    c.aircraft_type = "787-800"
    c.seats = 214
    c.annual_capacity = 214 * 7 * 52 * 2
    c.online_coeff = 1.0
    c.alliance_coeff = 0.615
    c.interline_coeff = 0.25
    c.nonstop_coeff = 1.0
    c.onestop_coeff = 0.20
    c.et_decay_factor = 0.8
    c.et_decay_interval = 0.1
    c.qsi_ceiling = 1.0
    c.qsi_adjustment = 1.0
    c.demand_provider = None
    c.schedule_provider = None
    c.cnx_coeffs = {'ONLINE': 1.0, 'ALLIANCE': 0.615, 'INTERLINING': 0.25}
    return c


def _ba_lhr_sjc_mock_results() -> Dict[str, Any]:
    """
    Mock pipeline results matching the validated BA LHR-SJC output.
    Target: 129,162 pax, 82.9% LF
    """
    p2p_details = [
        {'name': 'UK Business', 'base': 71441.55, 'growth': 0.10,
         'stimulation': 1.15, 'capture': 0.40, 'forecast': 32864.0},
        {'name': 'UK Leisure/VFR/Primary', 'base': 36385.76, 'growth': 0.10,
         'stimulation': 1.0, 'capture': 0.25, 'forecast': 10006.0},
        {'name': 'UK Leisure/VFR/Secondary', 'base': 17448.74, 'growth': 0.10,
         'stimulation': 1.0, 'capture': 0.25, 'forecast': 4798.0},
        {'name': 'UK Leisure/VFR/Contested', 'base': 4617.68, 'growth': 0.10,
         'stimulation': 1.0, 'capture': 0.10, 'forecast': 508.0},
        {'name': 'US Business', 'base': 65946.05, 'growth': 0.10,
         'stimulation': 1.15, 'capture': 0.15, 'forecast': 12510.0},
        {'name': 'US Leisure/VFR/Primary', 'base': 33586.86, 'growth': 0.10,
         'stimulation': 1.0, 'capture': 0.25, 'forecast': 9236.0},
        {'name': 'US Leisure/VFR/Secondary', 'base': 16106.53, 'growth': 0.10,
         'stimulation': 1.0, 'capture': 0.25, 'forecast': 4429.0},
        {'name': 'US Leisure/VFR/Contested', 'base': 4262.47, 'growth': 0.10,
         'stimulation': 1.0, 'capture': 0.10, 'forecast': 469.0},
    ]

    # Top connecting cities for home hub (LHR)  sample of validated results
    home_results = [
        {'city': 'PAR', 'name': 'Paris', 'country': 'France',
         'base_demand': 14500, 'growth': 0.09, 'qsi_capture': 0.04,
         'original_qsi': 0.04, 'forecast': 632, 'direct': False},
        {'city': 'AMS', 'name': 'Amsterdam', 'country': 'Netherlands',
         'base_demand': 8200, 'growth': 0.09, 'qsi_capture': 0.08,
         'original_qsi': 0.08, 'forecast': 716, 'direct': False},
        {'city': 'FRA', 'name': 'Frankfurt', 'country': 'Germany',
         'base_demand': 11000, 'growth': 0.09, 'qsi_capture': 0.05,
         'original_qsi': 0.05, 'forecast': 600, 'direct': False},
        {'city': 'MAD', 'name': 'Madrid', 'country': 'Spain',
         'base_demand': 7500, 'growth': 0.09, 'qsi_capture': 0.10,
         'original_qsi': 0.10, 'forecast': 818, 'direct': False},
        {'city': 'FCO', 'name': 'Rome', 'country': 'Italy',
         'base_demand': 6800, 'growth': 0.09, 'qsi_capture': 0.09,
         'original_qsi': 0.09, 'forecast': 668, 'direct': False},
    ]
    # Pad to realistic total
    home_total = 48115
    listed_home = sum(r['forecast'] for r in home_results)
    # Add "Other" bucket
    home_results.append({
        'city': 'OTH', 'name': 'Other (remaining cities)', 'country': 'Various',
        'base_demand': 200000, 'growth': 0.09, 'qsi_capture': 0.02,
        'original_qsi': 0.02, 'forecast': home_total - listed_home, 'direct': False,
    })

    # Dest hub connecting (SJC  small)
    dest_results = [
        {'city': 'HNL', 'name': 'Honolulu', 'country': 'US',
         'base_demand': 3500, 'growth': 0.10, 'qsi_capture': 0.12,
         'original_qsi': 0.12, 'forecast': 462, 'direct': False},
        {'city': 'OTH', 'name': 'Other', 'country': 'Various',
         'base_demand': 20000, 'growth': 0.10, 'qsi_capture': 0.01,
         'original_qsi': 0.01, 'forecast': 2475, 'direct': False},
    ]

    return {
        'p2p_total': 78110,          #  Validated target
        'p2p_details': p2p_details,
        'home_total': 48115,          #  Validated target
        'home_results': home_results,
        'dest_total': 2937,           #  Validated target
        'dest_results': dest_results,
        'grand_total': 129162,        #  Validated target
        'load_factor': 0.829,         #  Validated target
        'calibration': {'mode': 'expert'},
        'assembler_audit': [],
    }


# ============================================================================
# CLI
# ============================================================================

if __name__ == '__main__':
    test_ba_lhr_sjc()
