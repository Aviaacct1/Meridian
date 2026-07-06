#!/usr/bin/env python3
"""
Avia Solutions  MIDTDemandProvider (Chat 25)
==============================================
DemandProvider that reads raw Sabre/MIDT data extracts and aggregates
transaction-level booking records into city-level demand totals suitable
for the QSI forecast pipeline.

Data Sources (Sabre Extracts):
  - Connecting Demand files: LONSJCXXX.xlsx, SJCLONXXX.xlsx etc.
    Sheet: "Connecting Demand"  transaction-level records with connect points
  - P2P files: P2P_LONBAY_AREA_2013.xlsx, NH_P2P_Demand_OS_02Mar15.xlsx etc.
    Sheet: "Point to Point"  transaction-level P2P records
  - Both types share the same 20-column Sabre extract format

Sabre Extract Column Layout (header row):
  Col 0: Mod Org City        Col 1: Mod Dest City
  Col 2: Mod Dest City Name  Col 3: Mod Dest Country
  Col 4: Org City            Col 5: Dest City
  Col 6: Direct/Indirect     Col 7: Origin (airport)
  Col 8: Destination (airport)  Col 9: OperatingAirline
  Col 10: ConnectPoint1      Col 11: ConnectPoint2
  Col 12: ConnectPoint3      Col 13: Segment1Airline
  Col 14: Segment2Airline    Col 15: Segment3Airline
  Col 16: Segment4Airline    Col 17: Passengers
  Col 18: RevenueInUSD       Col 19: AvgFareInUSD / AvgOneWayFareInUSD

Aggregation Logic:
  1. Parse all transaction records from raw MIDT files
  2. Apply catchment area mapping (e.g., SFO/OAK/SJC  SJC service area)
  3. Factor up Sabre data (Sabre captures ~85% of bookings, varies by market)
  4. Split Direct vs Indirect demand
  5. Aggregate to city-level totals
  6. Apply growth rate and premium adjustments

Implements the DemandProvider interface from providers.py.
"""

import os
import math
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any, Tuple

import openpyxl

# Import from providers.py (the interface and data classes)
# When running standalone, we define them locally
try:
    from providers import (
        DemandProvider, P2PSegmentData, P2PSubsegmentData,
        ConnectingCityData
    )
except ImportError:
    from abc import ABC, abstractmethod

    @dataclass
    class P2PSubsegmentData:
        name: str
        base_demand: float
        growth_rate: float
        seasonality: float = 1.0
        stimulation: float = 1.0
        capture_rate: float = 0.0
        growth_years: int = 1  # Number of years to compound growth

    @dataclass
    class P2PSegmentData:
        name: str
        base_demand: float
        growth_rate: float
        seasonality: float = 1.0
        stimulation: float = 1.0
        capture_rate: float = 0.0
        subsegments: list = field(default_factory=list)
        growth_years: int = 1  # Number of years to compound growth

    @dataclass
    class ConnectingCityData:
        city_code: str
        city_name: str
        country: str
        base_demand: float
        growth_rate: float
        qsi_score: float = 0.0
        direct_service: bool = False

    class DemandProvider(ABC):
        @abstractmethod
        def get_p2p_segments(self) -> list: ...
        @abstractmethod
        def get_connecting_cities(self, direction: str) -> list: ...
        @abstractmethod
        def get_metadata(self) -> dict: ...


# ============================================================================
# RAW MIDT RECORD
# ============================================================================

@dataclass
class MIDTRecord:
    """Single transaction record from a Sabre/MIDT extract."""
    mod_org_city: str
    mod_dest_city: str
    dest_city_name: str
    dest_country: str
    org_city: str       # Raw origin city
    dest_city: str      # Raw dest city
    direct_indirect: str  # 'Direct' or 'Indirect'
    origin_apt: str     # Origin airport IATA
    dest_apt: str       # Destination airport IATA
    operating_airline: str
    connect_point1: str
    connect_point2: str
    connect_point3: str
    seg1_airline: str
    seg2_airline: str
    seg3_airline: str
    seg4_airline: str
    passengers: float
    revenue_usd: float
    avg_fare_usd: float


# ============================================================================
# CITY-LEVEL AGGREGATION
# ============================================================================

# ============================================================================
# PIVOT-P2P PARSER (VS-era Sabre format)
# ============================================================================
# These files have sheets: 'Pivot-P2P', 'Pivot-CNX', 'Data', 'City Lookup'
# Pivot-P2P layout:
#   Row 1: 'Mod Org City', <origin>
#   Row 3: 'Sum of Passengers', 'Column Labels'
#   Row 4: 'Row Labels', 'Direct', 'Indirect', 'Grand Total'
#   Row 5+: <city_code>, <direct>, <indirect>, <grand_total>
#   Last row: 'Grand Total', <total_direct>, <total_indirect>, <grand>

def parse_pivot_p2p(filepath: str) -> Dict[str, 'CityDemand']:
    """
    Parse a Sabre file with 'Pivot-P2P' sheet into city-level demand.
    Returns dict: city_code -> CityDemand.
    Also handles single-pair P2P files where Pivot-P2P has just one total row.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Try Pivot-P2P sheet
    pivot_sheet = None
    for sn in wb.sheetnames:
        if 'pivot' in sn.lower() and 'p2p' in sn.lower():
            pivot_sheet = sn
            break

    if not pivot_sheet:
        wb.close()
        return {}

    ws = wb[pivot_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        return {}

    # Find the header row with 'Row Labels' or column headers
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] and 'row label' in str(row[0]).lower():
            header_idx = i
            break
        if row and row[0] and 'sum of passengers' in str(row[0]).lower():
            # Header is next row
            if i + 1 < len(rows):
                header_idx = i + 1
                break

    if header_idx is None:
        return {}

    header = rows[header_idx]

    # Determine column positions for Direct/Indirect/Grand Total
    col_direct = None
    col_indirect = None
    col_grand = None
    for ci, val in enumerate(header):
        if val:
            vl = str(val).lower().strip()
            if vl == 'direct':
                col_direct = ci
            elif vl == 'indirect':
                col_indirect = ci
            elif 'grand' in vl and 'total' in vl:
                col_grand = ci

    # If only 'Indirect' + 'Grand Total' (no direct service), adjust
    if col_indirect is None and col_grand is None:
        # Try: header might just say 'Indirect', 'Grand Total'
        for ci, val in enumerate(header):
            if val and ci > 0:
                vl = str(val).lower().strip()
                if 'indirect' in vl:
                    col_indirect = ci
                elif 'grand' in vl:
                    col_grand = ci

    result = {}
    consecutive_blanks = 0
    total_row_demand = None  # Capture 'Total' row for single-pair P2P files
    origin_city = None
    dest_city = None

    # First scan the header area for Mod Org/Dest City
    for r_idx in range(0, min(header_idx, len(rows))):
        row = rows[r_idx]
        if row and row[0]:
            label = str(row[0]).lower().strip()
            val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if 'mod org' in label:
                origin_city = val
            elif 'mod dest' in label:
                dest_city = val

    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or not row[0]:
            consecutive_blanks += 1
            if consecutive_blanks >= 2:
                break  # End of this pivot table
            continue
        consecutive_blanks = 0

        city = str(row[0]).strip()

        # Capture 'Total' row values for single-pair P2P files
        if city.lower() in ('grand total', 'total'):
            # Save the demand from the total row
            t_direct = 0.0
            t_indirect = 0.0
            t_grand = 0.0
            if col_direct is not None and col_direct < len(row) and row[col_direct]:
                try: t_direct = float(row[col_direct])
                except (ValueError, TypeError): pass
            if col_indirect is not None and col_indirect < len(row) and row[col_indirect]:
                try: t_indirect = float(row[col_indirect])
                except (ValueError, TypeError): pass
            if col_grand is not None and col_grand < len(row) and row[col_grand]:
                try: t_grand = float(row[col_grand])
                except (ValueError, TypeError): pass
            if t_grand > 0 or t_direct > 0 or t_indirect > 0:
                total_row_demand = (t_direct, t_indirect, t_grand)
            continue

        if city.lower() == '(blank)':
            continue
        # Valid city codes are 3-letter IATA codes (uppercase alpha)
        if not (2 <= len(city) <= 4 and city.isalpha() and city == city.upper()):
            continue

        direct_pax = 0.0
        indirect_pax = 0.0
        grand_total = 0.0

        if col_direct is not None and col_direct < len(row) and row[col_direct]:
            try:
                direct_pax = float(row[col_direct])
            except (ValueError, TypeError):
                pass

        if col_indirect is not None and col_indirect < len(row) and row[col_indirect]:
            try:
                indirect_pax = float(row[col_indirect])
            except (ValueError, TypeError):
                pass

        if col_grand is not None and col_grand < len(row) and row[col_grand]:
            try:
                grand_total = float(row[col_grand])
            except (ValueError, TypeError):
                pass

        # If no grand total column, compute it
        if grand_total == 0 and (direct_pax > 0 or indirect_pax > 0):
            grand_total = direct_pax + indirect_pax

        if grand_total <= 0:
            continue

        result[city] = CityDemand(
            city_code=city,
            city_name='',
            country='',
            direct_pax=direct_pax,
            indirect_pax=indirect_pax,
            total_pax=grand_total,
        )

    # If no city-level rows found but we have a Total row and a specific dest city,
    # this is a single city-pair P2P file - use the total as the P2P demand
    if not result and total_row_demand and dest_city:
        t_direct, t_indirect, t_grand = total_row_demand
        if t_grand == 0:
            t_grand = t_direct + t_indirect
        if t_grand > 0:
            # Use dest city as the key (this is P2P demand to that city)
            result[dest_city] = CityDemand(
                city_code=dest_city,
                city_name='',
                country='',
                direct_pax=t_direct,
                indirect_pax=t_indirect,
                total_pax=t_grand,
            )

    return result

@dataclass
class CityDemand:
    """Aggregated demand for a single city."""
    city_code: str
    city_name: str
    country: str
    direct_pax: float = 0.0
    indirect_pax: float = 0.0
    total_pax: float = 0.0
    direct_revenue: float = 0.0
    indirect_revenue: float = 0.0
    total_revenue: float = 0.0
    record_count: int = 0
    airlines: Dict[str, float] = field(default_factory=lambda: defaultdict(float))

    @property
    def avg_fare(self) -> float:
        return self.total_revenue / self.total_pax if self.total_pax > 0 else 0

    @property
    def has_direct_service(self) -> bool:
        return self.direct_pax > 0


# ============================================================================
# SABRE FACTOR-UP LOGIC
# ============================================================================

# Sabre captures approximately 85% of total bookings globally,
# but the factor varies by market size and region.
# The BA LHR-SJC case uses banded factor-up rates:
DEFAULT_FACTOR_BANDS = [
    (100_000, 0.35),   # > 100,000 pax: Sabre captures 35%  factor = 1/0.35
    (50_000,  0.35),   # 50,000-100,000: same
    (5_000,   0.65),   # 5,000-50,000: Sabre captures 65%
    (0,       0.75),   # < 5,000: Sabre captures 75%
]


def get_sabre_factor(pax: float, factor_bands: List[Tuple[float, float]] = None) -> float:
    """
    Get the Sabre factor-up multiplier for a given passenger volume.
    
    Returns the multiplier to apply to Sabre data to estimate true market size.
    E.g., if Sabre captures 35% of bookings, multiplier = 1/0.35 = 2.857
    
    The bands match the approach in the BA forecast file's "Demand settings" sheet.
    """
    bands = factor_bands or DEFAULT_FACTOR_BANDS
    for threshold, share in bands:
        if pax >= threshold:
            return 1.0 / share if share > 0 else 1.0
    return 1.0


# ============================================================================
# CATCHMENT AREA MAPPING
# ============================================================================

# Standard catchment area groupings used by Avia Solutions
# Maps individual airports to the "service area" city code
DEFAULT_CATCHMENT_MAPS = {
    # Bay Area  SJC catchment
    'SJC': ['SJC', 'SFO', 'OAK'],
    # London  LON catchment (or LHR for hub-specific)
    'LON': ['LHR', 'LGW', 'STN', 'LTN', 'LCY', 'SEN'],
    'LHR': ['LHR', 'LGW', 'STN', 'LTN', 'LCY', 'SEN'],
    # New York area
    'NYC': ['JFK', 'EWR', 'LGA'],
    # Paris
    'PAR': ['CDG', 'ORY'],
    # Milan
    'MIL': ['MXP', 'LIN', 'BGY'],
    # Tokyo
    'TYO': ['NRT', 'HND'],
    # Stockholm
    'STO': ['ARN', 'BMA', 'NYO'],
    # Washington
    'WAS': ['IAD', 'DCA', 'BWI'],
    # Chicago
    'CHI': ['ORD', 'MDW'],
    # Moscow
    'MOW': ['SVO', 'DME', 'VKO'],
    # Rome
    'ROM': ['FCO', 'CIA'],
    # Berlin
    'BER': ['BER', 'TXL', 'SXF'],
    # Bucharest
    'BUH': ['OTP', 'BBU'],
    # Buenos Aires
    'BUE': ['EZE', 'AEP'],
    # Basel
    'EAP': ['BSL', 'MLH', 'EAP'],
    # So Paulo
    'SAO': ['GRU', 'CGH', 'VCP'],
}


def build_airport_to_city_map(catchment_map: Dict[str, List[str]],
                               city_lookup_file: str = None) -> Dict[str, str]:
    """
    Build airportcity mapping from catchment definitions and optional lookup file.
    
    Returns dict mapping airport IATA codes to city codes.
    """
    apt_to_city = {}

    # From catchment definitions
    for city_code, airports in catchment_map.items():
        for apt in airports:
            apt_to_city[apt] = city_code

    # From OAG city lookup file
    if city_lookup_file and os.path.exists(city_lookup_file):
        try:
            wb = openpyxl.load_workbook(city_lookup_file, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 3 and row[1] and row[2]:
                    apt = str(row[1]).strip().upper()
                    city = str(row[2]).strip().upper()
                    if len(apt) == 3 and len(city) == 3:
                        if apt not in apt_to_city:
                            apt_to_city[apt] = city
            wb.close()
        except Exception as e:
            pass  # Silently handle lookup file issues

    return apt_to_city


# ============================================================================
# MIDT FILE PARSER
# ============================================================================

def parse_midt_file(filepath: str,
                    data_sheet: str = None,
                    header_keyword: str = 'Mod Org City') -> List[MIDTRecord]:
    """
    Parse a Sabre/MIDT extract file and return transaction records.
    
    Auto-detects the data sheet if not specified. Handles both
    'Connecting Demand' and 'Point to Point' sheet types.
    
    Args:
        filepath: Path to the .xlsx file
        data_sheet: Sheet name containing raw data (auto-detected if None)
        header_keyword: Text to identify the header row
    
    Returns:
        List of MIDTRecord objects
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Auto-detect data sheet
    if data_sheet is None:
        candidates = ['Connecting Demand', 'Point to Point']
        for candidate in candidates:
            if candidate in wb.sheetnames:
                data_sheet = candidate
                break
        if data_sheet is None:
            # Try any sheet with enough rows
            for sn in wb.sheetnames:
                if sn.endswith('>>') or sn == 'City Lookup':
                    continue
                data_sheet = sn
                break

    if data_sheet not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[data_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        return []

    # Find header row
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and any(header_keyword.lower() in str(v).lower()
                       for v in row if v):
            header_row_idx = i
            break

    if header_row_idx is None:
        return []

    header = rows[header_row_idx]

    # Build column map
    col_map = {}
    for i, val in enumerate(header):
        if val:
            key = str(val).strip().lower()
            col_map[key] = i

    def find_col(*candidates, default=None):
        for c in candidates:
            cl = c.lower()
            if cl in col_map:
                return col_map[cl]
            for k, v in col_map.items():
                if cl in k:
                    return v
        return default

    # Map columns
    mod_org_col = find_col('mod org city', default=0)
    mod_dest_col = find_col('mod dest city', default=1)
    dest_name_col = find_col('mod dest city name', default=2)
    dest_country_col = find_col('mod dest country', default=3)
    org_city_col = find_col('org city', default=4)
    dest_city_col = find_col('dest city', default=5)
    di_col = find_col('direct/indirect', default=6)
    origin_col = find_col('origin', default=7)
    dest_col = find_col('destination', default=8)
    opair_col = find_col('operatingairline', 'operating airline', default=9)
    cp1_col = find_col('connectpoint1', 'connect point 1', default=10)
    cp2_col = find_col('connectpoint2', 'connect point 2', default=11)
    cp3_col = find_col('connectpoint3', 'connect point 3', default=12)
    s1_col = find_col('segment1airline', default=13)
    s2_col = find_col('segment2airline', default=14)
    s3_col = find_col('segment3airline', default=15)
    s4_col = find_col('segment4airline', default=16)
    pax_col = find_col('passengers', default=17)
    rev_col = find_col('revenueinusd', 'revenue', default=18)
    fare_col = find_col('avgfareinusd', 'avgonewayfareinusd', 'avg fare', default=19)

    def g(row, col, default=''):
        if col is not None and col < len(row) and row[col] is not None:
            return row[col]
        return default

    records = []
    for r_idx in range(header_row_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or len(row) < 10:
            continue

        mod_org = str(g(row, mod_org_col, '')).strip().upper()
        mod_dest = str(g(row, mod_dest_col, '')).strip().upper()
        if not mod_org or not mod_dest:
            continue

        # Parse passengers
        try:
            pax = float(g(row, pax_col, 0))
        except (ValueError, TypeError):
            pax = 0
        if pax <= 0:
            continue

        # Parse revenue
        try:
            rev = float(g(row, rev_col, 0))
        except (ValueError, TypeError):
            rev = 0

        # Parse fare
        try:
            fare = float(g(row, fare_col, 0))
        except (ValueError, TypeError):
            fare = 0

        records.append(MIDTRecord(
            mod_org_city=mod_org,
            mod_dest_city=mod_dest,
            dest_city_name=str(g(row, dest_name_col, '')).strip(),
            dest_country=str(g(row, dest_country_col, '')).strip(),
            org_city=str(g(row, org_city_col, '')).strip().upper(),
            dest_city=str(g(row, dest_city_col, '')).strip().upper(),
            direct_indirect=str(g(row, di_col, 'Indirect')).strip(),
            origin_apt=str(g(row, origin_col, '')).strip().upper(),
            dest_apt=str(g(row, dest_col, '')).strip().upper(),
            operating_airline=str(g(row, opair_col, '')).strip().upper(),
            connect_point1=str(g(row, cp1_col, '')).strip().upper(),
            connect_point2=str(g(row, cp2_col, '')).strip().upper(),
            connect_point3=str(g(row, cp3_col, '')).strip().upper(),
            seg1_airline=str(g(row, s1_col, '')).strip().upper(),
            seg2_airline=str(g(row, s2_col, '')).strip().upper(),
            seg3_airline=str(g(row, s3_col, '')).strip().upper(),
            seg4_airline=str(g(row, s4_col, '')).strip().upper(),
            passengers=pax,
            revenue_usd=rev,
            avg_fare_usd=fare,
        ))

    return records


def parse_city_lookup(filepath: str) -> Dict[str, Tuple[str, str]]:
    """
    Parse a City Lookup sheet from a Sabre extract file.
    
    Returns dict: airport_code  (city_name, country_name)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    
    lookup = {}
    if 'City Lookup' in wb.sheetnames:
        ws = wb['City Lookup']
        rows = list(ws.iter_rows(values_only=True))
        
        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and any('airport' in str(v).lower() for v in row if v):
                header_idx = i
                break
        
        if header_idx is not None:
            for r in range(header_idx + 1, len(rows)):
                row = rows[r]
                if not row or len(row) < 5:
                    continue
                apt = str(row[1]).strip().upper() if row[1] else ''
                city = str(row[2]).strip().upper() if row[2] else ''
                name = str(row[3]).strip() if row[3] else ''
                country = str(row[4]).strip() if row[4] else ''
                if apt and len(apt) == 3:
                    lookup[apt] = (name, country)
                if city and len(city) == 3 and city not in lookup:
                    lookup[city] = (name, country)
    
    wb.close()
    return lookup


# ============================================================================
# DEMAND AGGREGATION ENGINE
# ============================================================================

def aggregate_to_city_demand(records: List[MIDTRecord],
                              city_lookup: Dict[str, Tuple[str, str]] = None,
                              ) -> Dict[str, CityDemand]:
    """
    Aggregate MIDT transaction records to city-level demand totals.
    
    Groups by mod_dest_city (the modified/mapped destination city code)
    and separates direct vs indirect traffic.
    
    Returns dict: city_code  CityDemand
    """
    city_lookup = city_lookup or {}
    cities: Dict[str, CityDemand] = {}

    for rec in records:
        city_code = rec.mod_dest_city
        if not city_code:
            continue

        if city_code not in cities:
            # Resolve city name and country
            name = rec.dest_city_name
            country = rec.dest_country
            if not name and city_code in city_lookup:
                name, country = city_lookup[city_code]
            
            cities[city_code] = CityDemand(
                city_code=city_code,
                city_name=name if name and name != '#N/A' else city_code,
                country=country if country and country != '#N/A' else '',
            )

        cd = cities[city_code]
        cd.record_count += 1

        is_direct = rec.direct_indirect.lower() == 'direct'
        if is_direct:
            cd.direct_pax += rec.passengers
            cd.direct_revenue += rec.revenue_usd
        else:
            cd.indirect_pax += rec.passengers
            cd.indirect_revenue += rec.revenue_usd

        cd.total_pax += rec.passengers
        cd.total_revenue += rec.revenue_usd

        if rec.operating_airline:
            cd.airlines[rec.operating_airline] += rec.passengers

    return cities


def apply_sabre_factor(cities: Dict[str, CityDemand],
                        factor_bands: List[Tuple[float, float]] = None,
                        ) -> Dict[str, CityDemand]:
    """
    Apply Sabre factor-up to city demand totals.
    
    The factor is applied to indirect demand based on volume bands.
    Direct demand typically doesn't need factoring as it's well-captured.
    """
    for city_code, cd in cities.items():
        factor = get_sabre_factor(cd.indirect_pax, factor_bands)
        cd.indirect_pax *= factor
        cd.indirect_revenue *= factor
        cd.total_pax = cd.direct_pax + cd.indirect_pax
        cd.total_revenue = cd.direct_revenue + cd.indirect_revenue

    return cities


def apply_catchment_share(cities: Dict[str, CityDemand],
                           catchment_share: float = 1.0,
                           ) -> Dict[str, CityDemand]:
    """
    Apply catchment area share factor.
    
    For example, if the service area airport captures only a share of
    the broader city pair demand (SJC captures X% of Bay Area - LON demand),
    this factor adjusts accordingly.
    """
    if catchment_share >= 1.0:
        return cities

    for city_code, cd in cities.items():
        cd.direct_pax *= catchment_share
        cd.indirect_pax *= catchment_share
        cd.total_pax *= catchment_share
        cd.direct_revenue *= catchment_share
        cd.indirect_revenue *= catchment_share
        cd.total_revenue *= catchment_share

    return cities


# ============================================================================
# MIDT DEMAND PROVIDER
# ============================================================================

class MIDTDemandProvider(DemandProvider):
    """
    DemandProvider that reads raw Sabre/MIDT extracts and aggregates
    to city-level demand for the QSI pipeline.
    
    Typical usage for BA LHR-SJC:
    
        provider = MIDTDemandProvider(
            home_cnx_files=['LONSJCXXX_2013_data.xlsx'],
            dest_cnx_files=['SJCLONXXX__2013_CUT_4_data.xlsx'],
            p2p_files=['P2P_LONBAY_AREA_2013.xlsx'],
            home_growth=0.09,
            dest_growth=0.10,
            p2p_config={...},  # P2P segment definitions
        )
    
    Args:
        home_cnx_files: Sabre extract files for home hub connecting demand
            (e.g., LONSJCBeyond connecting traffic via LHR)
        dest_cnx_files: Sabre extract files for destination connecting demand
            (e.g., SJCLONBeyond connecting traffic via SJC)
        p2p_files: Sabre extract files for P2P demand
        p2p_config: P2P segment configuration (same format as ExcelDemandProvider)
        home_growth: Compound annual growth rate for home connecting
        dest_growth: Compound annual growth rate for dest connecting
        sabre_factor_bands: Custom Sabre factor-up bands (default: BA bands)
        catchment_share_home: Share of broader market captured at home hub
        catchment_share_dest: Share of broader market captured at dest
        city_lookup_file: OAG/Sabre city lookup file for name resolution
        min_demand_threshold: Minimum pax to include a city (default: 0)
    """

    def __init__(self,
                 home_cnx_files: List[str] = None,
                 dest_cnx_files: List[str] = None,
                 p2p_files: List[str] = None,
                 forecast_files: List[str] = None,
                 p2p_config: Dict[str, Any] = None,
                 home_growth: float = 0.09,
                 dest_growth: float = 0.10,
                 sabre_factor_bands: List[Tuple[float, float]] = None,
                 catchment_share_home: float = 1.0,
                 catchment_share_dest: float = 1.0,
                 city_lookup_file: str = None,
                 min_demand_threshold: float = 0,
                 ):
        self.home_cnx_files = home_cnx_files or []
        self.dest_cnx_files = dest_cnx_files or []
        self.p2p_files = p2p_files or []
        self.forecast_files = forecast_files or []
        self.p2p_config = p2p_config or {}
        self.home_growth = home_growth
        self.dest_growth = dest_growth
        self.sabre_factor_bands = sabre_factor_bands
        self.catchment_share_home = catchment_share_home
        self.catchment_share_dest = catchment_share_dest
        self.city_lookup_file = city_lookup_file
        self.min_demand_threshold = min_demand_threshold

        # Caches
        self._cnx_cache: Dict[str, List[ConnectingCityData]] = {}
        self._p2p_cache: Optional[List[P2PSegmentData]] = None
        self._raw_stats: Dict[str, Any] = {}
        self._city_lookup: Dict[str, Tuple[str, str]] = {}

        # Load city lookup from first available file
        all_files = self.home_cnx_files + self.dest_cnx_files + self.p2p_files
        for f in all_files:
            if os.path.exists(f):
                self._city_lookup.update(parse_city_lookup(f))
                break

    #  DemandProvider interface 

    def get_p2p_segments(self) -> List[P2PSegmentData]:
        """
        Build P2P segments from configuration.
        
        If p2p_config has 'segments' key (same as ExcelDemandProvider),
        uses that directly. Otherwise, if p2p_files are provided,
        aggregates from raw MIDT data.
        """
        if self._p2p_cache is not None:
            return self._p2p_cache

        if 'segments' in self.p2p_config:
            # Pre-configured segments (analyst has already defined them)
            segments = self._build_from_config()
        elif self.p2p_files or self.forecast_files:
            # Auto-detect from MIDT data or forecast workbook
            segments = self._build_from_midt_p2p()
        else:
            segments = []

        self._p2p_cache = segments
        return segments

    def get_connecting_cities(self, direction: str) -> List[ConnectingCityData]:
        """
        Get connecting cities for 'home' or 'dest' direction.
        
        Reads raw MIDT data, aggregates to city level, applies factor-up
        and catchment share, then returns as ConnectingCityData list.
        """
        if direction in self._cnx_cache:
            return self._cnx_cache[direction]

        if direction == 'home':
            files = self.home_cnx_files
            growth = self.home_growth
            catchment_share = self.catchment_share_home
        else:
            files = self.dest_cnx_files
            growth = self.dest_growth
            catchment_share = self.catchment_share_dest

        cities = self._aggregate_connecting(files, growth, catchment_share, direction)
        self._cnx_cache[direction] = cities
        return cities

    def get_metadata(self) -> Dict[str, Any]:
        return {
            'provider_type': 'MIDTDemandProvider',
            'home_cnx_files': [os.path.basename(f) for f in self.home_cnx_files],
            'dest_cnx_files': [os.path.basename(f) for f in self.dest_cnx_files],
            'p2p_files': [os.path.basename(f) for f in self.p2p_files],
            'home_growth': self.home_growth,
            'dest_growth': self.dest_growth,
            'catchment_share_home': self.catchment_share_home,
            'catchment_share_dest': self.catchment_share_dest,
            'stats': self._raw_stats,
        }

    #  Internal methods 

    def _aggregate_connecting(self, files: List[str], growth: float,
                               catchment_share: float,
                               direction: str) -> List[ConnectingCityData]:
        """Aggregate MIDT files into connecting city demand.
        
        Tries two parsing strategies:
          1. Old-format: 'Connecting Demand' sheet with transaction records
          2. New-format: 'Pivot-P2P' sheet with pre-aggregated city totals
        """
        all_records = []
        pivot_cities = {}

        for f in files:
            if not os.path.exists(f):
                self._raw_stats[f'{direction}_error'] = f"File not found: {f}"
                continue

            # Strategy 1: Old format (Connecting Demand sheet)
            try:
                records = parse_midt_file(f, data_sheet='Connecting Demand')
                if not records:
                    records = parse_midt_file(f)
            except Exception as e:
                records = []
                self._raw_stats[f'{direction}_parse_error'] = str(e)

            if records:
                all_records.extend(records)
            else:
                # Strategy 2: New format (Pivot-P2P sheet)
                try:
                    pv = parse_pivot_p2p(f)
                except Exception as e:
                    pv = {}
                    self._raw_stats[f'{direction}_pivot_error'] = str(e)
                    
                if pv:
                    for code, cd in pv.items():
                        if code in pivot_cities:
                            pivot_cities[code].total_pax += cd.total_pax
                            pivot_cities[code].direct_pax += cd.direct_pax
                            pivot_cities[code].indirect_pax += cd.indirect_pax
                        else:
                            pivot_cities[code] = cd

        self._raw_stats[f'{direction}_raw_records'] = len(all_records)
        self._raw_stats[f'{direction}_pivot_cities'] = len(pivot_cities)

        # If we got old-format records, use the standard aggregation pipeline
        if all_records:
            city_demand = aggregate_to_city_demand(all_records, self._city_lookup)
            self._raw_stats[f'{direction}_raw_cities'] = len(city_demand)
            city_demand = apply_sabre_factor(city_demand, self.sabre_factor_bands)
            city_demand = apply_catchment_share(city_demand, catchment_share)

        elif pivot_cities:
            # Use pivot data directly (already city-aggregated)
            city_demand = pivot_cities
            self._raw_stats[f'{direction}_raw_cities'] = len(city_demand)
            # Apply Sabre factor-up and catchment share to pivot data too
            city_demand = apply_sabre_factor(city_demand, self.sabre_factor_bands)
            city_demand = apply_catchment_share(city_demand, catchment_share)

        else:
            return []

        # Convert to ConnectingCityData list
        result = []
        for city_code, cd in sorted(city_demand.items(),
                                     key=lambda x: x[1].total_pax,
                                     reverse=True):
            # Use indirect demand as the base for connecting forecasts
            # (direct passengers aren't candidates for connections)
            # But if all traffic is indirect (no direct service), use total
            base_demand = cd.indirect_pax if cd.indirect_pax > 0 else cd.total_pax
            if base_demand < self.min_demand_threshold:
                continue

            result.append(ConnectingCityData(
                city_code=city_code,
                city_name=cd.city_name,
                country=cd.country,
                base_demand=cd.total_pax,  # Total demand (direct + indirect)
                growth_rate=growth,
                qsi_score=0.0,  # Will be set by QSI engine
                direct_service=cd.has_direct_service,
            ))

        self._raw_stats[f'{direction}_output_cities'] = len(result)
        self._raw_stats[f'{direction}_total_demand'] = sum(
            c.base_demand for c in result)

        return result

    def _extract_p2p_from_forecast(self) -> Tuple[float, float]:
        """
        Strategy 3: Extract P2P demand from forecast workbook's 'P2P Demand' sheet.
        
        Avia forecast workbooks contain a 'P2P Demand' sheet with columns:
        Nr | City Code | Direct | Indirect | Implied Factor Up | Premium | Factored + Premium
        
        Returns (total_direct, total_indirect).
        """
        import openpyxl
        
        for f in self.forecast_files:
            if not os.path.exists(f):
                continue
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                if 'P2P Demand' not in wb.sheetnames:
                    wb.close()
                    continue
                    
                ws = wb['P2P Demand']
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                
                # Find header row with "Direct" and "Indirect" columns
                header_row = None
                direct_col = None
                indirect_col = None
                for i, r in enumerate(rows):
                    vals = [str(v).strip() if v else '' for v in r]
                    for j, v in enumerate(vals):
                        if v.lower() == 'direct':
                            direct_col = j
                        if v.lower() == 'indirect':
                            indirect_col = j
                    if direct_col is not None or indirect_col is not None:
                        header_row = i
                        break
                
                if header_row is None:
                    continue
                
                # Look for "Total" row
                total_direct = 0.0
                total_indirect = 0.0
                for r in rows[header_row + 1:]:
                    vals = list(r)
                    first_val = str(vals[0]).strip().lower() if vals[0] else ''
                    if first_val == 'total':
                        if direct_col is not None and direct_col < len(vals):
                            v = vals[direct_col]
                            if isinstance(v, (int, float)):
                                total_direct = float(v)
                        if indirect_col is not None and indirect_col < len(vals):
                            v = vals[indirect_col]
                            if isinstance(v, (int, float)):
                                total_indirect = float(v)
                        break
                    # Also accumulate from data rows (city-level)
                    if isinstance(vals[0], (int, float)) or (vals[1] and len(str(vals[1]).strip()) == 3):
                        if direct_col is not None and direct_col < len(vals):
                            v = vals[direct_col]
                            if isinstance(v, (int, float)):
                                total_direct += float(v)
                        if indirect_col is not None and indirect_col < len(vals):
                            v = vals[indirect_col]
                            if isinstance(v, (int, float)):
                                total_indirect += float(v)
                
                if total_direct > 0 or total_indirect > 0:
                    self._raw_stats['p2p_source'] = f'forecast:{os.path.basename(f)}'
                    return (total_direct, total_indirect)
                    
            except Exception as e:
                self._raw_stats['p2p_forecast_error'] = str(e)
                continue
        
        return (0.0, 0.0)

    def _build_from_config(self) -> List[P2PSegmentData]:
        """Build P2P segments from pre-defined configuration."""
        segments = []
        for seg_cfg in self.p2p_config.get('segments', []):
            subsegments = []
            if 'subsegments' in seg_cfg:
                for sub_cfg in seg_cfg['subsegments']:
                    subsegments.append(P2PSubsegmentData(
                        name=sub_cfg['name'],
                        base_demand=sub_cfg['base_demand'],
                        growth_rate=sub_cfg['growth_rate'],
                        seasonality=sub_cfg.get('seasonality', 1.0),
                        stimulation=sub_cfg.get('stimulation', 1.0),
                        capture_rate=sub_cfg.get('capture_rate', 0.0),
                    ))

            segments.append(P2PSegmentData(
                name=seg_cfg['name'],
                base_demand=seg_cfg['base_demand'],
                growth_rate=seg_cfg['growth_rate'],
                seasonality=seg_cfg.get('seasonality', 1.0),
                stimulation=seg_cfg.get('stimulation', 1.0),
                capture_rate=seg_cfg.get('capture_rate', 0.0),
                subsegments=subsegments,
            ))
        return segments

    def _build_from_midt_p2p(self) -> List[P2PSegmentData]:
        """
        Auto-generate P2P segments from raw MIDT P2P data.
        
        Tries two strategies:
          1. Old-format: 'Point to Point' sheet with transaction records
          2. New-format: 'Pivot-P2P' sheet with pre-aggregated totals
        
        Creates segments based on direct vs indirect traffic split.
        """
        all_records = []
        pivot_total_direct = 0.0
        pivot_total_indirect = 0.0

        for f in self.p2p_files:
            if not os.path.exists(f):
                continue

            # Strategy 1: Old format
            records = parse_midt_file(f, data_sheet='Point to Point')
            if not records:
                records = parse_midt_file(f)
            if records:
                all_records.extend(records)
            else:
                # Strategy 2: New format (Pivot-P2P)
                pv = parse_pivot_p2p(f)
                if pv:
                    for code, cd in pv.items():
                        pivot_total_direct += cd.direct_pax
                        pivot_total_indirect += cd.indirect_pax

        self._raw_stats['p2p_raw_records'] = len(all_records)

        # Determine totals from whichever strategy worked
        if all_records:
            total_direct = sum(r.passengers for r in all_records
                               if r.direct_indirect.lower() == 'direct')
            total_indirect = sum(r.passengers for r in all_records
                                 if r.direct_indirect.lower() != 'direct')
        elif pivot_total_direct > 0 or pivot_total_indirect > 0:
            total_direct = pivot_total_direct
            total_indirect = pivot_total_indirect
        else:
            # Strategy 3: Extract from forecast workbook 'P2P Demand' sheet
            total_direct, total_indirect = self._extract_p2p_from_forecast()
            if total_direct == 0 and total_indirect == 0:
                return []

        total = total_direct + total_indirect

        self._raw_stats['p2p_direct_pax'] = total_direct
        self._raw_stats['p2p_indirect_pax'] = total_indirect
        self._raw_stats['p2p_total_pax'] = total

        # Detect monopoly status: if minimal direct P2P pax exist in Sabre data,
        # no carrier currently operates meaningful direct service on this O&D pair.
        # Threshold of 100 pax filters out Sabre noise/misroutes.
        # The portal can override via p2p_config['_is_monopoly']
        if '_is_monopoly' not in self.p2p_config:
            self.p2p_config['_is_monopoly'] = (total_direct < 100)

        # Use default growth from config or reasonable default
        p2p_growth = self.p2p_config.get('default_growth', 0.04)
        capture = self.p2p_config.get('default_capture', 0.25)

        # Growth years: how many years to compound from Sabre base to forecast
        # default_base_year: the year of the Sabre data (e.g., 2022)
        # default_forecast_year: the target forecast year (e.g., 2026)
        base_year = self.p2p_config.get('default_base_year', None)
        forecast_year = self.p2p_config.get('default_forecast_year', None)
        if base_year and forecast_year:
            growth_years = max(1, int(forecast_year) - int(base_year))
        else:
            growth_years = self.p2p_config.get('default_growth_years', 1)

        # Stimulation: use IATA curve if 'auto' or not specified
        # IATA formula: y = -0.491 * ln(x) + 6.8124 (from Avia presentation)
        # where x = existing indirect passengers
        stim_config = self.p2p_config.get('default_stimulation', 'auto')
        if stim_config == 'auto' and total_indirect > 0:
            import math
            stimulation = max(1.0, -0.491 * math.log(total_indirect) + 6.8124)
            stimulation = round(stimulation, 2)
        elif isinstance(stim_config, (int, float)):
            stimulation = float(stim_config)
        else:
            stimulation = 1.15  # fallback

        self._raw_stats['p2p_stimulation_used'] = stimulation
        self._raw_stats['p2p_growth_years'] = growth_years

        # Adjust capture rate based on indirect demand pool size
        # Pattern from analyst cases:
        #   Small pool (<20k indirect): 50% capture
        #     (less corporate demand, more price-sensitive leisure pax)
        #   Medium pool (20-50k): 55% capture
        #   Medium-large pool (25-100k): 60% capture (default)
        #   Very large pool (>100k): 55% (more hub competition)
        #
        # Only adjust if capture came from default (not PCE or manual).
        # PCE-derived capture is already informed by comparable cases.
        capture_source = self.p2p_config.get('_capture_source', 'default')
        if capture_source == 'default' and total_indirect > 0:
            if total_indirect < 15000:
                capture = min(capture, 0.50)
                self._raw_stats['p2p_capture_note'] = f'Reduced to {capture:.0%} (small market: {total_indirect:,.0f} indirect pax)'
            elif total_indirect < 25000:
                capture = min(capture, 0.55)
                self._raw_stats['p2p_capture_note'] = f'Reduced to {capture:.0%} (moderate market: {total_indirect:,.0f} indirect pax)'
            elif total_indirect > 100000:
                capture = min(capture, 0.55)
                self._raw_stats['p2p_capture_note'] = f'Reduced to {capture:.0%} (very large market: {total_indirect:,.0f} indirect pax — more hub competition)'

        self._raw_stats['p2p_capture_used'] = capture

        segments = []

        if total_direct > 0:
            segments.append(P2PSegmentData(
                name='P2P Direct',
                base_demand=total_direct,
                growth_rate=p2p_growth,
                stimulation=stimulation,
                capture_rate=capture,
                growth_years=growth_years,
            ))

        if total_indirect > 0:
            segments.append(P2PSegmentData(
                name='P2P Indirect',
                base_demand=total_indirect,
                growth_rate=p2p_growth,
                stimulation=stimulation,
                capture_rate=capture,
                growth_years=growth_years,
            ))

        return segments

    #  Diagnostic methods 

    def get_raw_city_demand(self, direction: str) -> Dict[str, CityDemand]:
        """
        Get the raw (pre-factor) city demand for diagnostic purposes.
        
        Returns the full CityDemand objects before Sabre factor-up,
        useful for validation against the forecast file.
        """
        if direction == 'home':
            files = self.home_cnx_files
        else:
            files = self.dest_cnx_files

        all_records = []
        for f in files:
            if os.path.exists(f):
                records = parse_midt_file(f, data_sheet='Connecting Demand')
                if not records:
                    records = parse_midt_file(f)
                all_records.extend(records)

        return aggregate_to_city_demand(all_records, self._city_lookup)

    def get_airline_breakdown(self, direction: str) -> Dict[str, Dict[str, float]]:
        """
        Get airline breakdown per city for diagnostic purposes.
        
        Returns dict: city_code  {airline_code: passengers}
        """
        raw = self.get_raw_city_demand(direction)
        return {code: dict(cd.airlines) for code, cd in raw.items()}

    def print_summary(self, direction: str = None):
        """Print a summary of the demand data."""
        directions = [direction] if direction else ['home', 'dest']
        for d in directions:
            cities = self.get_connecting_cities(d)
            print(f"\n{'='*60}")
            print(f"  {d.upper()} CONNECTING DEMAND")
            print(f"{'='*60}")
            print(f"  Cities: {len(cities)}")
            total = sum(c.base_demand for c in cities)
            print(f"  Total demand: {total:,.0f}")
            direct_count = sum(1 for c in cities if c.direct_service)
            print(f"  Cities with direct service: {direct_count}")
            print(f"  Growth rate: {self.home_growth if d == 'home' else self.dest_growth:.1%}")
            print(f"\n  Top 20 cities:")
            print(f"  {'City':<6} {'Name':<20} {'Country':<15} {'Demand':>10} {'Direct':>7}")
            print(f"  {'-'*60}")
            for c in cities[:20]:
                ds = 'YES' if c.direct_service else 'no'
                print(f"  {c.city_code:<6} {c.city_name[:20]:<20} "
                      f"{c.country[:15]:<15} {c.base_demand:>10,.0f} {ds:>7}")


# ============================================================================
# VALIDATION: Compare against the BA forecast file
# ============================================================================

def validate_against_forecast(provider: MIDTDemandProvider,
                               forecast_file: str,
                               direction: str = 'home',
                               ) -> Dict[str, Any]:
    """
    Compare MIDTDemandProvider output against the processed demand
    in a forecast file's 'Home Airport Cnx Demand' sheet.
    
    Returns comparison metrics.
    """
    # Get MIDT provider cities
    midt_cities = provider.get_connecting_cities(direction)
    midt_map = {c.city_code: c for c in midt_cities}

    # Load forecast file demand
    wb = openpyxl.load_workbook(forecast_file, data_only=True, read_only=True)

    if direction == 'home':
        sheet_name = 'Home Airport Cnx Demand'
    else:
        sheet_name = 'Destination Airport Cnx Demand'

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {'error': f'Sheet {sheet_name} not found'}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header
    header_idx = None
    for i, row in enumerate(rows):
        if row and any('city code' in str(v).lower() for v in row if v):
            header_idx = i
            break

    if header_idx is None:
        return {'error': 'Header not found'}

    # Parse forecast cities
    forecast_cities = {}
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        if not row or len(row) < 11:
            continue
        city_code = str(row[1]).strip().upper() if row[1] else ''
        if not city_code or city_code.lower() == 'total':
            continue
        try:
            # Col 9 = Base Indirect Demand, Col 10 = Base Total Demand
            base_indirect = float(row[9]) if row[9] else 0
            base_total = float(row[10]) if row[10] else 0
        except (ValueError, TypeError):
            base_indirect = 0
            base_total = 0
        if base_total > 0:
            forecast_cities[city_code] = {
                'base_indirect': base_indirect,
                'base_total': base_total,
                'has_direct': base_indirect != base_total,
            }

    # Compare
    both = set(midt_map.keys()) & set(forecast_cities.keys())
    midt_only = set(midt_map.keys()) - set(forecast_cities.keys())
    forecast_only = set(forecast_cities.keys()) - set(midt_map.keys())

    comparisons = []
    for city in sorted(both):
        midt_demand = midt_map[city].base_demand
        fcst_demand = forecast_cities[city]['base_total']
        ratio = midt_demand / fcst_demand if fcst_demand > 0 else 0
        comparisons.append({
            'city': city,
            'midt_demand': midt_demand,
            'forecast_demand': fcst_demand,
            'ratio': ratio,
            'diff_pct': (ratio - 1) * 100,
        })

    return {
        'direction': direction,
        'midt_cities': len(midt_map),
        'forecast_cities': len(forecast_cities),
        'cities_both': len(both),
        'cities_midt_only': len(midt_only),
        'cities_forecast_only': len(forecast_only),
        'midt_only_codes': sorted(midt_only)[:20],
        'forecast_only_codes': sorted(forecast_only)[:20],
        'comparisons': sorted(comparisons, key=lambda x: -x['forecast_demand']),
        'total_midt': sum(midt_map[c].base_demand for c in both),
        'total_forecast': sum(forecast_cities[c]['base_total'] for c in both),
    }


# ============================================================================
# CLI / TESTING
# ============================================================================

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='MIDT Demand Provider')
    parser.add_argument('--home-cnx', nargs='+', help='Home connecting demand files')
    parser.add_argument('--dest-cnx', nargs='+', help='Dest connecting demand files')
    parser.add_argument('--p2p', nargs='+', help='P2P demand files')
    parser.add_argument('--home-growth', type=float, default=0.09)
    parser.add_argument('--dest-growth', type=float, default=0.10)
    parser.add_argument('--validate', help='Forecast file to validate against')
    parser.add_argument('--verbose', action='store_true')
    args = parser.parse_args()

    provider = MIDTDemandProvider(
        home_cnx_files=args.home_cnx or [],
        dest_cnx_files=args.dest_cnx or [],
        p2p_files=args.p2p or [],
        home_growth=args.home_growth,
        dest_growth=args.dest_growth,
    )

    provider.print_summary()

    if args.validate:
        for d in ['home', 'dest']:
            result = validate_against_forecast(provider, args.validate, d)
            print(f"\n{'='*60}")
            print(f"  VALIDATION: {d.upper()}")
            print(f"{'='*60}")
            print(f"  MIDT cities: {result['midt_cities']}")
            print(f"  Forecast cities: {result['forecast_cities']}")
            print(f"  Overlap: {result['cities_both']}")
            if result.get('comparisons'):
                print(f"\n  Top 10 by forecast demand:")
                for c in result['comparisons'][:10]:
                    print(f"    {c['city']}: MIDT={c['midt_demand']:,.0f} "
                          f"Fcst={c['forecast_demand']:,.0f} "
                          f"Ratio={c['ratio']:.3f} ({c['diff_pct']:+.1f}%)")
