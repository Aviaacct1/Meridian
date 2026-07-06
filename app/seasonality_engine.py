#!/usr/bin/env python3
"""
Avia Solutions  Seasonality Engine
====================================
Distributes annual passenger and revenue forecasts into monthly profiles
using seasonal indices. Supports multiple profile types (route-specific from
Sabre data, regional defaults, business/leisure splits) and produces monthly
load factors, spill analysis, and presentation-ready seasonality charts.

Integration points:
  - run_pipeline() results dict  monthly breakdown
  - revenue_forecast.py  monthly revenue timing
  - output_workbook.py  monthly forecast sheets
  - Presentation template Section 6B  seasonality chart data

Key design principles:
  1. Annual total is ALWAYS preserved  monthly values sum exactly to annual
  2. Seasonality indices are normalised to mean 1.0 (sum = 12.0)
  3. Multiple profile sources with priority: route-specific > Sabre-derived > regional default
  4. Business and leisure segments can have different profiles
  5. Connecting traffic can use different profiles from P2P

Validated against:
  - BA LHR-SJC quarterly demand data: Q1=0.610, Q2=1.112, Q3=1.492, Q4=0.786
  - KLM AMS-TPA: seasonality=1.0 (year-round service, no seasonal adjustment)
"""

from config import OUTPUT_DIR, ensure_output_dir
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any
import math
import copy

# ============================================================================
# MONTH NAMES AND CONSTANTS
# ============================================================================

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

QUARTERS = {'Q1': [0, 1, 2], 'Q2': [3, 4, 5], 'Q3': [6, 7, 8], 'Q4': [9, 10, 11]}

DAYS_IN_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # non-leap
DAYS_IN_YEAR = 365


# ============================================================================
# SEASONAL PROFILE DATACLASS
# ============================================================================

@dataclass
class SeasonalProfile:
    """A 12-month seasonality index profile.

    Indices are normalised so mean = 1.0 (sum = 12.0).
    An index of 1.3 means 30% above the monthly average.
    An index of 0.7 means 30% below the monthly average.
    """
    name: str
    indices: List[float]  # 12 values, one per month
    source: str = ''      # Where this profile came from
    confidence: str = 'Medium'  # High / Medium / Low

    def __post_init__(self):
        if len(self.indices) != 12:
            raise ValueError(f"Profile '{self.name}' must have exactly 12 indices, got {len(self.indices)}")
        # Normalise to mean 1.0
        total = sum(self.indices)
        if total > 0 and abs(total - 12.0) > 0.001:
            factor = 12.0 / total
            self.indices = [v * factor for v in self.indices]

    @property
    def peak_month(self) -> str:
        idx = self.indices.index(max(self.indices))
        return MONTHS[idx]

    @property
    def trough_month(self) -> str:
        idx = self.indices.index(min(self.indices))
        return MONTHS[idx]

    @property
    def peak_trough_ratio(self) -> float:
        mn = min(self.indices)
        return max(self.indices) / mn if mn > 0 else float('inf')

    @property
    def coefficient_of_variation(self) -> float:
        mean = sum(self.indices) / 12
        variance = sum((x - mean) ** 2 for x in self.indices) / 12
        return math.sqrt(variance) / mean if mean > 0 else 0

    def quarterly_indices(self) -> Dict[str, float]:
        """Return quarterly indices (normalised to mean 1.0 over 4 quarters)."""
        result = {}
        for qname, months in QUARTERS.items():
            result[qname] = sum(self.indices[m] for m in months) / 3
        return result

    def to_dict(self) -> Dict[str, Any]:
        return {
            'name': self.name,
            'indices': dict(zip(MONTHS, self.indices)),
            'source': self.source,
            'confidence': self.confidence,
            'peak_month': self.peak_month,
            'trough_month': self.trough_month,
            'peak_trough_ratio': round(self.peak_trough_ratio, 2),
            'cv': round(self.coefficient_of_variation, 3),
        }


# ============================================================================
# DEFAULT SEASONAL PROFILE LIBRARY
# ============================================================================

# Regional default profiles derived from industry data and Avia experience.
# These are starting points  route-specific Sabre data should replace them.

PROFILE_LIBRARY = {

    # Transatlantic long-haul: strong summer peak, winter trough
    'transatlantic': SeasonalProfile(
        name='Transatlantic Long-Haul',
        indices=[0.70, 0.65, 0.80, 0.95, 1.10, 1.20, 1.35, 1.35, 1.15, 0.95, 0.80, 0.70],
        source='Avia Solutions regional default  Transatlantic',
        confidence='Medium',
    ),

    # Europe-Asia long-haul: flatter profile, slight summer peak
    'europe_asia': SeasonalProfile(
        name='EuropeAsia Long-Haul',
        indices=[0.85, 0.80, 0.90, 0.95, 1.00, 1.10, 1.15, 1.15, 1.05, 1.00, 0.95, 0.85],
        source='Avia Solutions regional default  Europe-Asia',
        confidence='Medium',
    ),

    # US-Asia Pacific: slight summer peak, Chinese New Year bump
    'us_asia': SeasonalProfile(
        name='USAsia Pacific',
        indices=[0.95, 0.90, 0.90, 0.90, 0.95, 1.10, 1.20, 1.20, 1.05, 0.95, 0.90, 0.85],
        source='Avia Solutions regional default  US-Asia',
        confidence='Medium',
    ),

    # Intra-European short-haul: very strong summer, weak winter
    'intra_europe': SeasonalProfile(
        name='Intra-European Short-Haul',
        indices=[0.55, 0.55, 0.70, 0.90, 1.10, 1.25, 1.45, 1.45, 1.20, 0.95, 0.65, 0.55],
        source='Avia Solutions regional default  Intra-Europe',
        confidence='Medium',
    ),

    # Middle East hub: winter peak (opposite to European pattern)
    'middle_east': SeasonalProfile(
        name='Middle East Hub',
        indices=[1.10, 1.05, 1.15, 1.00, 0.85, 0.75, 0.80, 0.85, 0.90, 1.05, 1.15, 1.25],
        source='Avia Solutions regional default  Middle East',
        confidence='Medium',
    ),

    # Year-round flat (for routes with no meaningful seasonality)
    'flat': SeasonalProfile(
        name='Year-Round Flat',
        indices=[1.0] * 12,
        source='Flat profile  no seasonal adjustment',
        confidence='High',
    ),

    # Business-heavy route: flatter with mild summer dip
    'business_heavy': SeasonalProfile(
        name='Business-Heavy Route',
        indices=[0.95, 0.95, 1.05, 1.05, 1.05, 1.00, 0.85, 0.80, 1.05, 1.10, 1.05, 0.90],
        source='Avia Solutions default  business-heavy routes',
        confidence='Medium',
    ),

    # Leisure-heavy route: extreme summer peak
    'leisure_heavy': SeasonalProfile(
        name='Leisure-Heavy Route',
        indices=[0.50, 0.50, 0.65, 0.90, 1.15, 1.30, 1.55, 1.55, 1.20, 0.90, 0.60, 0.50],
        source='Avia Solutions default  leisure-heavy routes',
        confidence='Medium',
    ),

    # BA LHR-SJC actual (from LONSJC_Base_Demand_Assumptions quarterly data)
    'ba_lhr_sjc': SeasonalProfile(
        name='BA LHR-SJC (Actual)',
        # Derived from quarterly: Q1=0.610, Q2=1.112, Q3=1.492, Q4=0.786
        # Interpolated to monthly using standard transatlantic shape within quarters
        indices=[0.58, 0.56, 0.69, 0.98, 1.13, 1.23, 1.42, 1.58, 1.46, 0.89, 0.76, 0.72],
        source='Sabre MI 2013 quarterly data interpolated to monthly',
        confidence='High',
    ),
}


# ============================================================================
# PROFILE SELECTION LOGIC
# ============================================================================

def select_profile(
    route_type: str = 'transatlantic',
    demand_split: str = 'mixed',
    custom_profile: Optional[SeasonalProfile] = None,
    custom_indices: Optional[List[float]] = None,
) -> SeasonalProfile:
    """Select the appropriate seasonal profile for a route.

    Priority:
      1. custom_profile (if provided directly)
      2. custom_indices (builds a profile from 12 values)
      3. demand_split-specific profile if available
      4. route_type from library
      5. 'flat' fallback

    Args:
        route_type: Key into PROFILE_LIBRARY ('transatlantic', 'europe_asia', etc.)
        demand_split: 'business' / 'leisure' / 'mixed'  selects sub-profile
        custom_profile: A pre-built SeasonalProfile object
        custom_indices: Raw 12-value list to build a custom profile
    """
    if custom_profile is not None:
        return custom_profile

    if custom_indices is not None:
        return SeasonalProfile(
            name=f'Custom ({route_type})',
            indices=custom_indices,
            source='User-specified custom indices',
            confidence='High',
        )

    # Demand-split specific overrides
    if demand_split == 'business' and 'business_heavy' in PROFILE_LIBRARY:
        return copy.deepcopy(PROFILE_LIBRARY['business_heavy'])
    if demand_split == 'leisure' and 'leisure_heavy' in PROFILE_LIBRARY:
        return copy.deepcopy(PROFILE_LIBRARY['leisure_heavy'])

    if route_type in PROFILE_LIBRARY:
        return copy.deepcopy(PROFILE_LIBRARY[route_type])

    return copy.deepcopy(PROFILE_LIBRARY['flat'])


# Season definitions matching the OAG two-pull split (summer Apr-mid Oct, winter mid Oct-end Mar).
# October is split half/half. Month indices 0=Jan .. 11=Dec.
_SUMMER_FULL = (3, 4, 5, 6, 7, 8)     # Apr-Sep
_WINTER_FULL = (10, 11, 0, 1, 2)      # Nov-Mar


def season_shares(indices):
    """Given 12 monthly indices, return (summer_share, winter_share) of annual demand on the OAG season
    split (summer Apr-mid Oct, winter mid Oct-end Mar; October split half/half). The two shares sum to 1,
    so a summer service captures its summer share of the annual O&D, not half of it."""
    s = sum(indices[m] for m in _SUMMER_FULL) + 0.5 * indices[9]
    w = sum(indices[m] for m in _WINTER_FULL) + 0.5 * indices[9]
    tot = s + w
    return (s / tot, w / tot) if tot else (0.5, 0.5)


def season_share_for(season, route_type='transatlantic', demand_split='mixed', custom_indices=None):
    """The demand share for one season ('summer'/'winter'; 'annual' -> 1.0) from the selected profile.
    A seasonal forecast multiplies the annual demand by this before capping at the season's capacity."""
    if not season or season == 'annual':
        return 1.0
    prof = select_profile(route_type=route_type, demand_split=demand_split, custom_indices=custom_indices)
    ss, ws = season_shares(prof.indices)
    return ss if season == 'summer' else ws


def blend_profiles(
    profiles: List[Tuple[SeasonalProfile, float]],
) -> SeasonalProfile:
    """Blend multiple profiles by weight.

    Args:
        profiles: List of (profile, weight) tuples. Weights are normalised internally.

    Returns:
        Blended SeasonalProfile
    """
    if not profiles:
        return copy.deepcopy(PROFILE_LIBRARY['flat'])

    if len(profiles) == 1:
        return copy.deepcopy(profiles[0][0])

    total_weight = sum(w for _, w in profiles)
    if total_weight == 0:
        return copy.deepcopy(PROFILE_LIBRARY['flat'])

    blended = [0.0] * 12
    names = []
    for profile, weight in profiles:
        norm_w = weight / total_weight
        names.append(f"{profile.name} ({norm_w:.0%})")
        for m in range(12):
            blended[m] += profile.indices[m] * norm_w

    return SeasonalProfile(
        name='Blended: ' + ' + '.join(names),
        indices=blended,
        source='Weighted blend of component profiles',
        confidence='Medium',
    )


def from_quarterly(
    q1: float, q2: float, q3: float, q4: float,
    name: str = 'Sabre-Derived',
    source: str = 'Sabre MI quarterly data',
    interpolation: str = 'smooth',
) -> SeasonalProfile:
    """Build a monthly profile from quarterly data.

    Quarterly values can be absolute (passengers) or indices.
    They are normalised to quarterly mean = 1.0 then interpolated to monthly.

    Args:
        q1-q4: Quarterly values (absolute or index)
        interpolation: 'flat' (3 months same) or 'smooth' (linear interpolation within quarter)
    """
    # Normalise to quarterly mean = 1.0
    q_mean = (q1 + q2 + q3 + q4) / 4
    if q_mean == 0:
        return copy.deepcopy(PROFILE_LIBRARY['flat'])

    q_idx = [q1 / q_mean, q2 / q_mean, q3 / q_mean, q4 / q_mean]

    if interpolation == 'flat':
        # Each month in the quarter gets the same index
        indices = []
        for qi in q_idx:
            indices.extend([qi, qi, qi])
        return SeasonalProfile(name=name, indices=indices, source=source, confidence='High')

    # Smooth interpolation: distribute within each quarter using a shape
    # that preserves the quarterly total (3 months sum to 3  q_index)
    # while creating smooth transitions between quarters.
    #
    # Within each quarter, ramp linearly from the boundary with the previous
    # quarter to the boundary with the next quarter, anchored so the 3-month
    # average equals the quarterly index.
    indices = []
    for qi in range(4):
        prev_q = q_idx[(qi - 1) % 4]
        curr_q = q_idx[qi]
        next_q = q_idx[(qi + 1) % 4]

        # Edge values: average of adjacent quarters
        start_val = (prev_q + curr_q) / 2
        end_val = (curr_q + next_q) / 2

        # Linear ramp: values at positions 0, 1, 2 within the quarter
        # v(t) = start_val + t * (end_val - start_val) / 2  where t = 0, 1, 2
        # But we need mean of 3 values = curr_q
        # v0 = a, v1 = a + d, v2 = a + 2d
        # mean = a + d = curr_q  a = curr_q - d
        # slope d = (end_val - start_val) / 2
        d = (end_val - start_val) / 2
        a = curr_q - d
        indices.extend([a, a + d, a + 2 * d])

    return SeasonalProfile(name=name, indices=indices, source=source, confidence='High')


def from_monthly_pax(
    monthly_pax: List[float],
    name: str = 'Route-Specific',
    source: str = 'Sabre MI monthly data',
) -> SeasonalProfile:
    """Build a seasonal profile from actual monthly passenger counts.

    The raw monthly counts are converted to indices relative to the monthly average.
    """
    if len(monthly_pax) != 12:
        raise ValueError(f"Need 12 monthly values, got {len(monthly_pax)}")

    total = sum(monthly_pax)
    if total == 0:
        return copy.deepcopy(PROFILE_LIBRARY['flat'])

    avg = total / 12
    indices = [m / avg for m in monthly_pax]

    return SeasonalProfile(name=name, indices=indices, source=source, confidence='High')


# ============================================================================
# MONTHLY DISTRIBUTION ENGINE
# ============================================================================

@dataclass
class MonthlyForecast:
    """Monthly breakdown of an annual forecast."""
    annual_total: float
    monthly_pax: List[float]  # 12 values
    monthly_capacity: List[float]  # 12 values
    monthly_load_factor: List[float]  # 12 values
    profile: SeasonalProfile
    spill_months: List[str] = field(default_factory=list)  # months where LF > 100%

    @property
    def peak_lf(self) -> float:
        return max(self.monthly_load_factor)

    @property
    def trough_lf(self) -> float:
        return min(self.monthly_load_factor)

    @property
    def peak_month(self) -> str:
        idx = self.monthly_load_factor.index(self.peak_lf)
        return MONTHS[idx]

    @property
    def trough_month(self) -> str:
        idx = self.monthly_load_factor.index(self.trough_lf)
        return MONTHS[idx]

    def quarterly_summary(self) -> Dict[str, Dict[str, float]]:
        result = {}
        for qname, months in QUARTERS.items():
            q_pax = sum(self.monthly_pax[m] for m in months)
            q_cap = sum(self.monthly_capacity[m] for m in months)
            result[qname] = {
                'pax': round(q_pax),
                'capacity': round(q_cap),
                'load_factor': q_pax / q_cap if q_cap > 0 else 0,
            }
        return result

    def to_dict(self) -> Dict[str, Any]:
        return {
            'annual_total': round(self.annual_total),
            'monthly': {
                MONTHS[m]: {
                    'pax': round(self.monthly_pax[m]),
                    'capacity': round(self.monthly_capacity[m]),
                    'load_factor': round(self.monthly_load_factor[m], 4),
                    'index': round(self.profile.indices[m], 3),
                }
                for m in range(12)
            },
            'quarterly': self.quarterly_summary(),
            'peak_lf': round(self.peak_lf, 4),
            'trough_lf': round(self.trough_lf, 4),
            'peak_month': self.peak_month,
            'trough_month': self.trough_month,
            'spill_months': self.spill_months,
            'profile': self.profile.to_dict(),
        }


def distribute_annual(
    annual_pax: float,
    annual_capacity: float,
    profile: SeasonalProfile,
    frequency: int = 7,
    seats: int = 214,
    capacity_profile: Optional[SeasonalProfile] = None,
) -> MonthlyForecast:
    """Distribute annual passenger forecast into monthly values.

    Annual total is always preserved exactly (no rounding drift).

    Args:
        annual_pax: Total annual passengers
        annual_capacity: Total annual seat capacity
        profile: Demand seasonality profile
        frequency: Weekly frequency (used if capacity_profile is None)
        seats: Seats per flight (used if capacity_profile is None)
        capacity_profile: Optional separate supply seasonality
            (if None, capacity is distributed proportional to days in month)
    """
    # Distribute demand using indices weighted by days-in-month
    # This ensures that a flat profile (all 1.0) distributes proportionally
    # to days in each month, matching the capacity distribution.
    monthly_pax = [0.0] * 12
    weighted = [profile.indices[m] * DAYS_IN_MONTH[m] for m in range(12)]
    w_sum = sum(weighted)
    if w_sum > 0:
        for m in range(12):
            monthly_pax[m] = annual_pax * weighted[m] / w_sum
    else:
        for m in range(12):
            monthly_pax[m] = annual_pax / 12

    # Verify sum preservation
    pax_diff = annual_pax - sum(monthly_pax)
    if abs(pax_diff) > 0.01:
        # Distribute rounding residual to peak month
        peak_idx = profile.indices.index(max(profile.indices))
        monthly_pax[peak_idx] += pax_diff

    # Distribute capacity
    monthly_capacity = [0.0] * 12
    if capacity_profile is not None:
        cap_sum = sum(capacity_profile.indices)
        for m in range(12):
            monthly_capacity[m] = annual_capacity * capacity_profile.indices[m] / cap_sum
    else:
        # Default: proportional to days in month (uniform daily frequency)
        for m in range(12):
            monthly_capacity[m] = annual_capacity * DAYS_IN_MONTH[m] / DAYS_IN_YEAR

    # Calculate load factors
    monthly_lf = [0.0] * 12
    spill_months = []
    for m in range(12):
        if monthly_capacity[m] > 0:
            lf = monthly_pax[m] / monthly_capacity[m]
            monthly_lf[m] = lf
            if lf > 1.0:
                spill_months.append(MONTHS[m])
        else:
            monthly_lf[m] = 0.0

    return MonthlyForecast(
        annual_total=annual_pax,
        monthly_pax=monthly_pax,
        monthly_capacity=monthly_capacity,
        monthly_load_factor=monthly_lf,
        profile=profile,
        spill_months=spill_months,
    )


# ============================================================================
# SPILL-ADJUSTED MONTHLY FORECAST
# ============================================================================

def distribute_with_spill(
    annual_pax_unconstrained: float,
    annual_capacity: float,
    profile: SeasonalProfile,
    c_factor: float = 1.5,
    frequency: int = 7,
    seats: int = 214,
) -> Dict[str, Any]:
    """Distribute annual forecast with monthly spill calculation.

    Unlike the pipeline-level spill (which applies a single annual C-factor),
    this distributes unconstrained demand monthly and applies spill per-month.
    This reveals which months are capacity-constrained and how much traffic
    is lost in peak periods.

    Args:
        annual_pax_unconstrained: Demand before any spill
        annual_capacity: Total annual seats
        profile: Demand seasonality
        c_factor: Spill curve exponent (1.5 typical)
    """
    unconstrained = distribute_annual(
        annual_pax_unconstrained, annual_capacity, profile, frequency, seats,
    )

    constrained_pax = [0.0] * 12
    spill_pax = [0.0] * 12
    monthly_cap = unconstrained.monthly_capacity

    for m in range(12):
        cap = monthly_cap[m]
        demand = unconstrained.monthly_pax[m]
        if cap > 0 and demand > 0:
            lf_unconstrained = demand / cap
            if lf_unconstrained <= 1.0:
                # No spill in this month
                constrained_pax[m] = demand
                spill_pax[m] = 0
            else:
                # Apply Boeing spill curve: achieved_LF = 1 - (1-LF_offered)^c
                # where LF_offered is what we'd achieve without constraints
                # Approximate: cap the demand at capacity and calculate spill
                constrained_pax[m] = cap  # Can't carry more than capacity
                spill_pax[m] = demand - cap
        else:
            constrained_pax[m] = demand

    constrained_annual = sum(constrained_pax)
    total_spill = sum(spill_pax)
    constrained_lf = [
        constrained_pax[m] / monthly_cap[m] if monthly_cap[m] > 0 else 0
        for m in range(12)
    ]

    return {
        'unconstrained': unconstrained.to_dict(),
        'constrained_annual': round(constrained_annual),
        'total_spill': round(total_spill),
        'spill_rate': total_spill / annual_pax_unconstrained if annual_pax_unconstrained > 0 else 0,
        'monthly_detail': {
            MONTHS[m]: {
                'unconstrained_pax': round(unconstrained.monthly_pax[m]),
                'constrained_pax': round(constrained_pax[m]),
                'spill': round(spill_pax[m]),
                'capacity': round(monthly_cap[m]),
                'unconstrained_lf': round(unconstrained.monthly_load_factor[m], 4),
                'constrained_lf': round(constrained_lf[m], 4),
            }
            for m in range(12)
        },
        'spill_months': [MONTHS[m] for m in range(12) if spill_pax[m] > 0],
        'profile': profile.to_dict(),
    }


# ============================================================================
# REVENUE SEASONALITY
# ============================================================================

def monthly_revenue(
    annual_pax_revenue: float,
    annual_cargo_revenue: float,
    annual_ancillary_revenue: float,
    pax_profile: SeasonalProfile,
    cargo_profile: Optional[SeasonalProfile] = None,
) -> Dict[str, Any]:
    """Distribute annual revenue into monthly values.

    Pax revenue follows the pax seasonality profile.
    Cargo can have a different profile (e.g., Q4 peak for holiday shipping).
    Ancillary follows pax profile.

    Args:
        annual_pax_revenue: Total annual passenger revenue
        annual_cargo_revenue: Total annual cargo revenue
        annual_ancillary_revenue: Total annual ancillary revenue
        pax_profile: Passenger seasonality profile
        cargo_profile: Cargo seasonality (defaults to flat if not provided)
    """
    if cargo_profile is None:
        cargo_profile = copy.deepcopy(PROFILE_LIBRARY['flat'])

    pax_sum = sum(pax_profile.indices)
    cargo_sum = sum(cargo_profile.indices)

    monthly = {}
    for m in range(12):
        pax_rev = annual_pax_revenue * pax_profile.indices[m] / pax_sum
        cargo_rev = annual_cargo_revenue * cargo_profile.indices[m] / cargo_sum
        anc_rev = annual_ancillary_revenue * pax_profile.indices[m] / pax_sum

        monthly[MONTHS[m]] = {
            'pax_revenue': round(pax_rev, 2),
            'cargo_revenue': round(cargo_rev, 2),
            'ancillary_revenue': round(anc_rev, 2),
            'total_revenue': round(pax_rev + cargo_rev + anc_rev, 2),
        }

    return {
        'monthly': monthly,
        'annual': {
            'pax_revenue': round(annual_pax_revenue, 2),
            'cargo_revenue': round(annual_cargo_revenue, 2),
            'ancillary_revenue': round(annual_ancillary_revenue, 2),
            'total_revenue': round(
                annual_pax_revenue + annual_cargo_revenue + annual_ancillary_revenue, 2
            ),
        },
    }


# ============================================================================
# PIPELINE INTEGRATION
# ============================================================================

def seasonalise_pipeline_output(
    results: Dict[str, Any],
    profile: Optional[SeasonalProfile] = None,
    profile_key: str = 'transatlantic',
    custom_indices: Optional[List[float]] = None,
    p2p_profile: Optional[SeasonalProfile] = None,
    cnx_profile: Optional[SeasonalProfile] = None,
) -> Dict[str, Any]:
    """Apply seasonality to pipeline results.

    Takes the output of run_pipeline() and adds monthly breakdowns for:
      - Total forecast (grand_total)
      - P2P subtotal
      - Connecting at home hub
      - Connecting at dest airport
      - Individual P2P segments (if detail available)

    Args:
        results: Output dict from run_pipeline()
        profile: Override profile for all segments
        profile_key: Library key if no profile provided
        custom_indices: Custom 12-value indices
        p2p_profile: Separate profile for P2P traffic (optional)
        cnx_profile: Separate profile for connecting traffic (optional)

    Returns:
        Enhanced results dict with 'seasonality' key added
    """
    grand_total = results.get('grand_total', 0)
    annual_capacity = results.get('annual_capacity', 0)
    p2p_total = results.get('p2p_total', 0)
    cnx_home = results.get('cnx_home_total', 0)
    cnx_dest = results.get('cnx_dest_total', 0)
    frequency = results.get('frequency', 7)
    seats = results.get('seats', 214)

    # Select profiles
    main_profile = profile or select_profile(
        route_type=profile_key,
        custom_indices=custom_indices,
    )

    # If separate P2P / CNX profiles not provided, use main for everything
    eff_p2p_profile = p2p_profile or main_profile
    eff_cnx_profile = cnx_profile or main_profile

    # Blend P2P and CNX profiles weighted by their share of total
    if grand_total > 0:
        p2p_weight = p2p_total / grand_total
        cnx_weight = (cnx_home + cnx_dest) / grand_total
    else:
        p2p_weight = 0.5
        cnx_weight = 0.5

    blended = blend_profiles([
        (eff_p2p_profile, p2p_weight),
        (eff_cnx_profile, cnx_weight),
    ])

    # Generate monthly distributions
    total_monthly = distribute_annual(
        grand_total, annual_capacity, blended, frequency, seats,
    )

    p2p_monthly = distribute_annual(
        p2p_total, annual_capacity * (p2p_total / grand_total if grand_total > 0 else 1),
        eff_p2p_profile, frequency, seats,
    ) if p2p_total > 0 else None

    cnx_home_monthly = distribute_annual(
        cnx_home, annual_capacity * (cnx_home / grand_total if grand_total > 0 else 0),
        eff_cnx_profile, frequency, seats,
    ) if cnx_home > 0 else None

    cnx_dest_monthly = distribute_annual(
        cnx_dest, annual_capacity * (cnx_dest / grand_total if grand_total > 0 else 0),
        eff_cnx_profile, frequency, seats,
    ) if cnx_dest > 0 else None

    seasonality_result = {
        'total': total_monthly.to_dict(),
        'blended_profile': blended.to_dict(),
    }

    if p2p_monthly:
        seasonality_result['p2p'] = p2p_monthly.to_dict()
    if cnx_home_monthly:
        seasonality_result['cnx_home'] = cnx_home_monthly.to_dict()
    if cnx_dest_monthly:
        seasonality_result['cnx_dest'] = cnx_dest_monthly.to_dict()

    # Add spill analysis if load factor > 85%
    annual_lf = grand_total / annual_capacity if annual_capacity > 0 else 0
    if annual_lf > 0.85 and total_monthly.spill_months:
        seasonality_result['spill_warning'] = {
            'annual_lf': round(annual_lf, 4),
            'months_over_100pct': total_monthly.spill_months,
            'peak_monthly_lf': round(total_monthly.peak_lf, 4),
            'peak_month': total_monthly.peak_month,
            'recommendation': (
                f"Peak month ({total_monthly.peak_month}) LF is "
                f"{total_monthly.peak_lf:.1%}. Consider frequency increase "
                f"or larger aircraft for summer peak."
            ),
        }

    # Add to results
    enhanced = dict(results)
    enhanced['seasonality'] = seasonality_result
    return enhanced


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

def write_seasonality_sheet(ws, monthly_forecast: MonthlyForecast, title: str = 'Monthly Forecast'):
    """Write monthly forecast data to an openpyxl worksheet.

    Args:
        ws: openpyxl worksheet
        monthly_forecast: MonthlyForecast object
        title: Sheet title
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill('solid', fgColor='003366')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    DATA_FONT = Font(size=10, name='Arial')
    BOLD_FONT = Font(bold=True, size=10, name='Arial')
    PCT_FMT = '0.0%'
    NUM_FMT = '#,##0'
    IDX_FMT = '0.000'
    THIN_BORDER = Border(
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Title
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, name='Arial')

    # Headers
    headers = ['Month', 'Seasonal Index', 'Passengers', 'Capacity', 'Load Factor']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Monthly data
    mf = monthly_forecast
    for m in range(12):
        row = m + 4
        ws.cell(row=row, column=1, value=MONTHS[m]).font = DATA_FONT
        ws.cell(row=row, column=2, value=round(mf.profile.indices[m], 3)).font = DATA_FONT
        ws.cell(row=row, column=2).number_format = IDX_FMT
        ws.cell(row=row, column=3, value=round(mf.monthly_pax[m])).font = DATA_FONT
        ws.cell(row=row, column=3).number_format = NUM_FMT
        ws.cell(row=row, column=4, value=round(mf.monthly_capacity[m])).font = DATA_FONT
        ws.cell(row=row, column=4).number_format = NUM_FMT
        ws.cell(row=row, column=5, value=round(mf.monthly_load_factor[m], 4)).font = DATA_FONT
        ws.cell(row=row, column=5).number_format = PCT_FMT

        # Highlight spill months in red
        if mf.monthly_load_factor[m] > 1.0:
            ws.cell(row=row, column=5).font = Font(bold=True, color='FF0000', size=10, name='Arial')

        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Total row
    total_row = 16
    ws.cell(row=total_row, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=total_row, column=2, value=12.0).font = BOLD_FONT
    ws.cell(row=total_row, column=2).number_format = IDX_FMT
    ws.cell(row=total_row, column=3, value=round(mf.annual_total)).font = BOLD_FONT
    ws.cell(row=total_row, column=3).number_format = NUM_FMT
    ws.cell(row=total_row, column=4, value=round(sum(mf.monthly_capacity))).font = BOLD_FONT
    ws.cell(row=total_row, column=4).number_format = NUM_FMT
    annual_lf = mf.annual_total / sum(mf.monthly_capacity) if sum(mf.monthly_capacity) > 0 else 0
    ws.cell(row=total_row, column=5, value=round(annual_lf, 4)).font = BOLD_FONT
    ws.cell(row=total_row, column=5).number_format = PCT_FMT

    # Summary stats
    ws.cell(row=18, column=1, value='Profile:').font = BOLD_FONT
    ws.cell(row=18, column=2, value=mf.profile.name).font = DATA_FONT
    ws.cell(row=19, column=1, value='Peak Month:').font = BOLD_FONT
    ws.cell(row=19, column=2, value=f"{mf.peak_month} ({mf.peak_lf:.1%})").font = DATA_FONT
    ws.cell(row=20, column=1, value='Trough Month:').font = BOLD_FONT
    ws.cell(row=20, column=2, value=f"{mf.trough_month} ({mf.trough_lf:.1%})").font = DATA_FONT
    ws.cell(row=21, column=1, value='Peak/Trough Ratio:').font = BOLD_FONT
    ws.cell(row=21, column=2, value=round(mf.profile.peak_trough_ratio, 2)).font = DATA_FONT

    if mf.spill_months:
        ws.cell(row=22, column=1, value=' Spill Months:').font = Font(bold=True, color='FF0000', size=10, name='Arial')
        ws.cell(row=22, column=2, value=', '.join(mf.spill_months)).font = Font(color='FF0000', size=10, name='Arial')

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14


# ============================================================================
# VALIDATION
# ============================================================================

def validate_ba_lhr_sjc():
    """Validate against BA LHR-SJC known data.

    Quarterly demand from LONSJC_Base_Demand_Assumptions_02Apr15.xlsx:
      Q1: 52,820  Q2: 96,310  Q3: 129,150  Q4: 68,020  Total: 346,300

    Pipeline output: 129,162 pax, 155,792 capacity, 82.9% LF
    """
    print("=" * 70)
    print("VALIDATION: BA LHR-SJC Seasonality Engine")
    print("=" * 70)

    # Test 1: Quarterly profile creation
    profile = from_quarterly(52820, 96310, 129150, 68020,
                             name='BA LHR-SJC (Sabre 2013)', source='LONSJC Base Demand Assumptions')
    q_idx = profile.quarterly_indices()
    print(f"\nTest 1  Quarterly profile from Sabre data:")
    print(f"  Q1: {q_idx['Q1']:.3f} (expected ~0.610)")
    print(f"  Q2: {q_idx['Q2']:.3f} (expected ~1.112)")
    print(f"  Q3: {q_idx['Q3']:.3f} (expected ~1.492)")
    print(f"  Q4: {q_idx['Q4']:.3f} (expected ~0.786)")
    print(f"  Peak: {profile.peak_month}, Trough: {profile.trough_month}")
    print(f"  Peak/Trough ratio: {profile.peak_trough_ratio:.2f}")

    # Verify quarterly indices match known values
    expected = {'Q1': 0.610, 'Q2': 1.112, 'Q3': 1.492, 'Q4': 0.786}
    all_ok = True
    for q, exp in expected.items():
        if abs(q_idx[q] - exp) > 0.002:
            print(f"   {q}: {q_idx[q]:.3f} vs expected {exp:.3f}")
            all_ok = False
    print(f"  {'' if all_ok else ''} Quarterly indices: {'PASS' if all_ok else 'FAIL'}")

    # Test 2: Annual distribution preserves total
    monthly = distribute_annual(129162, 155792, profile, frequency=7, seats=214)
    pax_sum = sum(monthly.monthly_pax)
    sum_ok = abs(pax_sum - 129162) < 1
    print(f"\nTest 2  Annual total preservation:")
    print(f"  Sum of monthly pax: {pax_sum:.0f} (target: 129,162)")
    print(f"  {'' if sum_ok else ''} Sum preservation: {'PASS' if sum_ok else 'FAIL'}")

    # Test 3: Load factor profile
    print(f"\nTest 3  Monthly load factor profile:")
    print(f"  Annual LF: {129162/155792:.1%}")
    print(f"  Peak LF: {monthly.peak_lf:.1%} ({monthly.peak_month})")
    print(f"  Trough LF: {monthly.trough_lf:.1%} ({monthly.trough_month})")
    lf_ok = monthly.peak_lf > monthly.trough_lf  # Basic sanity
    print(f"  {'' if lf_ok else ''} Peak > Trough: {'PASS' if lf_ok else 'FAIL'}")

    # Test 4: Flat profile produces uniform distribution (proportional to days)
    flat_profile = PROFILE_LIBRARY['flat']
    flat_monthly = distribute_annual(120000, 150000, flat_profile)
    # With flat demand indices, monthly pax should follow days-in-month pattern
    flat_sum = sum(flat_monthly.monthly_pax)
    jan_expected = 120000 * 31 / 365  # days-proportional
    jan_ok = abs(flat_monthly.monthly_pax[0] - jan_expected) < 1
    flat_sum_ok = abs(flat_sum - 120000) < 1
    print(f"\nTest 4  Flat profile uniformity:")
    print(f"  Jan pax: {flat_monthly.monthly_pax[0]:.0f} (expected: {jan_expected:.0f})")
    print(f"  Sum: {flat_sum:.0f} (expected: 120,000)")
    print(f"  {'' if jan_ok and flat_sum_ok else ''} Flat distribution: {'PASS' if jan_ok and flat_sum_ok else 'FAIL'}")

    # Test 5: Profile blending
    biz = PROFILE_LIBRARY['business_heavy']
    lei = PROFILE_LIBRARY['leisure_heavy']
    blended = blend_profiles([(biz, 0.55), (lei, 0.45)])
    blend_sum = sum(blended.indices)
    blend_ok = abs(blend_sum - 12.0) < 0.001
    print(f"\nTest 5  Profile blending (55% business + 45% leisure):")
    print(f"  Index sum: {blend_sum:.3f} (expected: 12.000)")
    print(f"  {'' if blend_ok else ''} Blend normalisation: {'PASS' if blend_ok else 'FAIL'}")

    # Test 6: Pipeline integration
    mock_results = {
        'grand_total': 129162,
        'annual_capacity': 155792,
        'p2p_total': 78110,
        'cnx_home_total': 48115,
        'cnx_dest_total': 2937,
        'frequency': 7,
        'seats': 214,
    }
    enhanced = seasonalise_pipeline_output(mock_results, profile_key='transatlantic')
    has_seasonality = 'seasonality' in enhanced
    has_total = 'total' in enhanced.get('seasonality', {})
    print(f"\nTest 6  Pipeline integration:")
    print(f"  {'' if has_seasonality else ''} seasonality key present: {'PASS' if has_seasonality else 'FAIL'}")
    print(f"  {'' if has_total else ''} total monthly breakdown: {'PASS' if has_total else 'FAIL'}")

    # Test 7: KLM AMS-TPA (year-round, flat profile)
    klm_monthly = distribute_annual(79771, 91104, PROFILE_LIBRARY['flat'])
    klm_sum = sum(klm_monthly.monthly_pax)
    klm_ok = abs(klm_sum - 79771) < 1
    print(f"\nTest 7  KLM AMS-TPA flat profile:")
    print(f"  Sum: {klm_sum:.0f} (target: 79,771)")
    print(f"  LF range: {klm_monthly.trough_lf:.1%}  {klm_monthly.peak_lf:.1%}")
    # With flat profile, LF should be nearly identical across months
    lf_range = klm_monthly.peak_lf - klm_monthly.trough_lf
    klm_flat_ok = lf_range < 0.005  # < 0.5% variation  flat demand, days-proportional capacity
    print(f"  LF variation: {lf_range:.3%} (should be <0.5% for flat demand + days-proportional capacity)")
    print(f"  {'' if klm_ok and klm_flat_ok else ''} KLM flat test: {'PASS' if klm_ok and klm_flat_ok else 'FAIL'}")

    # Test 8: Excel output
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Monthly Forecast'
        write_seasonality_sheet(ws, monthly, title='BA LHR-SJC Monthly Forecast')
        ensure_output_dir()
        test_path = str(OUTPUT_DIR / 'test_seasonality.xlsx')
        wb.save(test_path)
        print(f"\nTest 8  Excel output:  PASS (saved to {test_path})")
    except Exception as e:
        print(f"\nTest 8  Excel output:  FAIL ({e})")

    # Summary
    tests_passed = sum([all_ok, sum_ok, lf_ok, jan_ok, blend_ok, has_seasonality, has_total, klm_ok and klm_flat_ok])
    total_tests = 8
    print(f"\n{'=' * 70}")
    print(f"RESULT: {tests_passed}/{total_tests} tests passed")
    print(f"{'=' * 70}")

    return tests_passed == total_tests


if __name__ == '__main__':
    success = validate_ba_lhr_sjc()
    exit(0 if success else 1)
