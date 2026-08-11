#!/usr/bin/env python3
r"""Avia Solutions - the workbook to populate if airport charges are to be held for the top airports.

    py -3.12 airport_charges_template.py [n] [output.xlsx]      default n = 300

WHY. Airport and handling charges are the largest line in a short-haul route P&L and the tool holds
real figures for five airports. Everything else falls back to a placeholder that is not that airport.
This writes the table that would have to be filled to close that, sized on measured traffic so the
work is aimed where it pays: the top 300 airports carry 72% of world departures and the top 100 carry
45%, against 4,111 airports with any scheduled service at all.

THE COLUMN THAT MAKES THIS TRACTABLE is landing per tonne. Landing fees are almost always weight
based, so one rate per airport serves every aircraft type and the table is three hundred rows rather
than three hundred times twenty-five cells. Enter landing_per_turn ONLY where an airport genuinely
charges a flat fee, or where the only figure available is a worked example for one type, in which
case name the type in the notes.

WHAT THIS TABLE IS NOT. Published charges are a CEILING. Most carriers negotiate below them and many
airports waive them entirely for a new route for a period. Nothing entered here is what a carrier
actually pays, which is why the tool treats charges as a declared plug the client can override and
names the provenance in the payload. Populating this improves the starting point; it does not turn
the figure into a measurement of what is paid.

Output feeds app/airport_charges.json via the Populate sheet's own instructions.

Avia Solutions Limited. All rights reserved.
"""
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)


def traffic(n=300, period="2025-%"):
    """Top n airports by scheduled passenger departures, with the traffic that justifies the work."""
    import duckdb
    import capacity_frame as CF
    db = CF._oag()
    if not db:
        sys.exit("no OAG store found. Set AVIA_OAG_DUCKDB or AVIA_LOCAL_CACHE.")
    con = duckdb.connect(db, read_only=True)
    con.execute("SET memory_limit='4GB'; SET threads=3; SET enable_progress_bar=false")
    try:
        rows = con.execute("""
          SELECT dep_airport, any_value(dep_city) city, any_value(dep_country) country,
                 count(*) deps, sum(try_cast(seats AS DOUBLE)) seats,
                 count(DISTINCT carrier) carriers,
                 median(try_cast(gcd_km AS DOUBLE)) med_km
          FROM oag WHERE service_type='J' AND week LIKE ? AND try_cast(stops AS INT)=0
            AND dep_airport IS NOT NULL AND trim(dep_airport) <> ''
          GROUP BY 1 ORDER BY deps DESC
        """, [period]).fetchall()
    finally:
        con.close()
    total = sum(r[3] for r in rows) or 1
    return rows[:n], total, len(rows)


def build(path, n=300):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from aircraft_economics import AIRPORTS

    rows, total, n_all = traffic(n)
    ARIAL = "Arial"
    hdr = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F3864")
    body = Font(name=ARIAL, size=10)
    blue = Font(name=ARIAL, size=10, color="0000FF")
    note = Font(name=ARIAL, size=9, italic=True, color="595959")
    fillin = PatternFill("solid", fgColor="FFFF00")
    seeded = PatternFill("solid", fgColor="E2EFDA")
    thin = Border(bottom=Side(style="thin", color="BFBFBF"))

    wb = Workbook()

    # ---------------------------------------------------------------- legend
    ws = wb.active; ws.title = "How to use"
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 104
    lines = [
        ("Airport charges: the table to populate", None),
        ("", None),
        ("Scope", "The top %d airports by scheduled passenger departures in 2025. They carry %.0f%% "
                  "of all departures, against %s airports with any scheduled service. The top 100 "
                  "carry 45%%, so if the work has to be split, do it in rank order."
                  % (n, sum(r[3] for r in rows) / total * 100, f"{n_all:,}")),
        ("Fill the yellow", "One row an airport. Green rows are already held in the module and are "
                            "shown so they can be checked rather than re-entered."),
        ("landing_per_tonne", "PREFERRED. Landing fee per tonne of maximum take-off weight, in USD. "
                              "One rate serves every aircraft type, which is what keeps this to a "
                              "few hundred rows instead of a cell per airport per type."),
        ("landing_per_turn", "Only where the airport genuinely charges a flat fee per movement, or "
                             "where the only figure available is a worked example for one aircraft. "
                             "If it is a worked example, name the type in the notes. A per-tonne "
                             "figure always wins where both are entered."),
        ("pax_charge_per_pax", "Departing passenger charge in USD per departing passenger."),
        ("ground_handling_per_turn", "Handling per turnaround in USD, at this airport only. The "
                                     "tool adds both ends."),
        ("recovery_per_pax", "Charges and taxes RECOVERED from the passenger and passed through, in "
                             "USD per passenger. This is revenue, not cost. Omitting it is what made "
                             "LCY-EDI look loss-making: 21.42 USD a passenger of APD recovery was "
                             "simply absent."),
        ("currency and rate", "Enter the figures in USD. Record the local currency and the rate used "
                              "so the conversion can be checked and refreshed."),
        ("source and date", "Required, not optional. RDC AirportCharges, the airport's own published "
                            "schedule, or a regulator's determination. A row without a source cannot "
                            "be defended and should stay blank."),
        ("", None),
        ("What this is not", "PUBLISHED CHARGES ARE A CEILING. Most carriers negotiate below them "
                             "and many airports waive them for a new route for a period. Populating "
                             "this improves the starting point; it does not make the figure what a "
                             "carrier actually pays. The tool treats charges as a declared plug the "
                             "client can override, and names the provenance in the output."),
        ("", None),
        ("Loading it", "Save the Populate sheet as app/airport_charges.json in the shape "
                       '{\"airports\": {\"LHR\": {\"landing_per_tonne\": 0.0, ...}}} and the tool '
                       "picks it up. airport_charges.py resolves measured, then held, then generic, "
                       "and says which it used."),
        ("Source", "Traffic columns: OAG schedules 2025, service_type J, nonstop. Generated by "
                   "app/airport_charges_template.py."),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=a).font = (Font(name=ARIAL, bold=True, size=12) if i == 1
                                                  else Font(name=ARIAL, bold=True, size=10))
        if b:
            c = ws.cell(row=i, column=2, value=b); c.font = body
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 44

    # ---------------------------------------------------------------- populate
    heads = ["rank", "IATA", "city", "country", "2025 departures", "% of world", "cumulative %",
             "carriers", "median sector km",
             "landing_per_tonne", "landing_per_turn", "pax_charge_per_pax",
             "ground_handling_per_turn", "recovery_per_pax",
             "local currency", "FX rate to USD", "source", "source date", "notes"]
    widths = [6, 7, 20, 18, 15, 10, 12, 9, 15, 15, 15, 17, 22, 16, 13, 13, 40, 12, 34]
    ws = wb.create_sheet("Populate")
    for j, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = hdr; c.fill = hfill
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    ws.freeze_panes = "C2"; ws.row_dimensions[1].height = 32

    cum = 0
    for i, (iata, city, country, deps, seats, carriers, med) in enumerate(rows, 1):
        r = i + 1
        cum += deps
        held = AIRPORTS.get(iata)
        for j, v in ((1, i), (2, iata), (3, (city or "").strip()), (4, (country or "").strip()),
                     (5, int(deps)), (6, deps / total), (7, cum / total),
                     (8, int(carriers)), (9, int(med or 0))):
            c = ws.cell(row=r, column=j, value=v); c.font = body; c.border = thin
            if j in (6, 7):
                c.number_format = "0.00%"
        for j in range(10, 20):
            c = ws.cell(row=r, column=j); c.font = blue; c.border = thin
            c.fill = seeded if held else fillin
        if held:
            ws.cell(row=r, column=11, value=held.get("landing_per_turn"))
            ws.cell(row=r, column=12, value=held.get("pax_charge_per_pax"))
            ws.cell(row=r, column=13, value=held.get("ground_handling_per_turn"))
            ws.cell(row=r, column=14, value=held.get("recovery_per_pax"))
            ws.cell(row=r, column=17, value="already held in aircraft_economics.AIRPORTS: CHECK, "
                                            "do not assume it is current or sourced")
    last = len(rows) + 1
    ws.cell(row=last + 2, column=1, value=(
        "Yellow rows are to fill. Green rows are already in the module and should be checked against "
        "a current source rather than trusted. A row with no source should be left blank: the tool "
        "falls back to a placeholder and says so, which is safer than an unsourced figure that looks "
        "measured.")).font = note
    ws.cell(row=last + 3, column=1, value=(
        "Landing per tonne is the column that matters. With it, one row serves every aircraft type; "
        "without it the table has to be redone every time a new type appears.")).font = note

    # ---------------------------------------------------------------- coverage maths
    ws = wb.create_sheet("Coverage")
    for j, h in enumerate(["airports populated", "share of world departures covered",
                           "cumulative departures", "note"], 1):
        c = ws.cell(row=1, column=j, value=h); c.font = hdr; c.fill = hfill
        ws.column_dimensions[get_column_letter(j)].width = [22, 32, 22, 60][j - 1]
    marks = [10, 25, 50, 100, 150, 200, 250, 300]
    for i, m in enumerate(marks, 2):
        if m > len(rows):
            continue
        cov = sum(r[3] for r in rows[:m])
        for j, v in ((1, m), (2, cov / total), (3, int(cov))):
            c = ws.cell(row=i, column=j, value=v); c.font = body; c.border = thin
            if j == 2:
                c.number_format = "0.0%"
        ws.cell(row=i, column=4, value=("a route study touches two airports, so coverage of pairs "
                                        "is roughly the square of this" if m == 100 else "")).font = note

    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    wb.properties.title = "Airport charges: the table to populate"
    wb.save(path)
    return path, len(rows), sum(r[3] for r in rows) / total


if __name__ == "__main__":
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 300
    out = sys.argv[2] if len(sys.argv) > 2 else "Airport charges - to populate.xlsx"
    p, k, share = build(out, n)
    print("written: %s  (%d airports, %.1f%% of 2025 departures)" % (p, k, share * 100))
