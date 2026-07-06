#!/usr/bin/env python3
"""
Module III: Connection Builder
Avia Solutions  QSI Forecast Tool Pipeline

Replicates the Connection Builder VBA macro (v4.0, Dec 2014).
Takes OAG schedule data (leg1/leg2), MCT data, and alliance memberships,
then enumerates all viable two-leg connecting itineraries.

Pipeline position: OAG Parser (Module II)  Connection Builder (Module III)  QSI Scorer (Module IV)

Usage:
    python connection_builder.py <oag_parsed_file.xlsx> [options]

    --origin LHR              Origin airport code
    --destination SJC         Destination airport code
    --carrier BA              Proposed carrier code
    --freq 7                  Proposed weekly frequency
    --dep-time 1555           Proposed departure time (HHMM from origin)
    --arr-time 1810           Proposed arrival time (at destination, HHMM local)
    --dep-days 1234567        Days of operation
    --flying-time 660         Flying time in minutes (origindest)
    --with-service             Include proposed new service (default: without)
    --output FILE              Output xlsx path
    --mct-file FILE            Optional MCT file path (xlsx/xls)
    --alliance-file FILE       Optional alliance membership file
    --lcc-list FILE            Optional LCC carrier list (one code per line)
    --default-mct 90           Default MCT when not specified (minutes)
    --min-connect 20           Minimum connection window (minutes)
    --max-connect 720          Maximum connection window (minutes)
"""

import argparse
import sys
import os
import math
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 
# DEFAULT ALLIANCE MEMBERSHIPS (as of 2024, updatable via file)
# 

ONEWORLD = {
    'AA', 'BA', 'CX', 'AY', 'IB', 'JL', 'LA', 'MH', 'QF', 'QR',
    'RJ', 'S7', 'UL', 'AT', 'FJ', 'OA',  # recent additions
}
STAR_ALLIANCE = {
    'AC', 'AI', 'NZ', 'NH', 'OZ', 'OS', 'AV', 'CM', 'OU', 'MS',
    'ET', 'BR', 'LO', 'LH', 'SK', 'SQ', 'SA', 'LX', 'TP', 'TG',
    'TK', 'UA', 'SN',
}
SKYTEAM = {
    'SU', 'AM', 'UX', 'AF', 'AZ', 'CI', 'MU', 'CZ', 'OK', 'DL',
    'GA', 'KQ', 'KL', 'KE', 'ME', 'SV', 'RO', 'VN', 'AR', 'MF',
}

# Known LCC carriers (excluded from connecting itineraries per methodology)
DEFAULT_LCC_LIST = {
    'FR', 'U2', 'W6', 'VY', 'EW', 'DY', 'NK', 'F9', 'G4', 'WN',
    'AS', 'B6', 'HA', 'SY', 'XR', 'QS', 'TO', 'LS', 'MT', 'BY',
    'ZT', '5O', 'PC', 'XQ', 'HV', 'WW', 'AK', 'FD', 'QZ', 'Z2',
    'G9', 'FZ', '6E', 'SG', 'I5', 'AP', 'H9', 'JQ', 'TT', 'BL',
    'VJ', 'DD', 'SL', 'OD',
}


def _parse_duration_mins(val):
    """Parse a duration to minutes. Handles 'HH:MM:SS', 'HH:MM', 'HHMM'/int."""
    if val is None or val == '':
        return 0
    s = str(val).strip()
    if ':' in s:
        p = s.split(':')
        try:
            return int(p[0]) * 60 + int(p[1])
        except (ValueError, IndexError):
            return 0
    try:
        s = str(int(float(s))).zfill(4)
        return int(s[:2]) * 60 + int(s[2:])
    except (ValueError, TypeError):
        return 0


def parse_time_hhmm(val):
    """Parse HHMM time value (int or string) into minutes since midnight."""
    if val is None or val == '':
        return None
    s = str(int(val)).zfill(4)
    h, m = int(s[:2]), int(s[2:])
    return h * 60 + m


def minutes_to_hhmm(mins):
    """Convert minutes since midnight to HHMM integer."""
    mins = int(mins) % 1440
    return (mins // 60) * 100 + (mins % 60)


def parse_days_string(days_val):
    """Parse days-of-operation string/int into set of day numbers {1..7}."""
    if days_val is None:
        return set()
    s = str(days_val).strip()
    return {int(c) for c in s if c.isdigit() and 1 <= int(c) <= 7}


def classify_connection(carrier1, carrier2, alliances):
    """Classify connection type: ONLINE, ALLIANCE, or INTERLINING."""
    if carrier1 == carrier2:
        return 'ONLINE'
    for alliance_set in alliances:
        if carrier1 in alliance_set and carrier2 in alliance_set:
            return 'ALLIANCE'
    return 'INTERLINING'


def get_dom_int(country1, country2):
    """Determine DOM/INT transfer type."""
    if country1 and country2 and country1 == country2:
        return 'DOM'
    return 'INT'


# 
# MCT LOADING
# 

def load_mct_data(mct_file=None, default_mct=90):
    """Load MCT data from file. Returns dict: (airport, term_arr, term_dep, dom_int)  minutes."""
    mct = {}
    if mct_file and os.path.exists(mct_file):
        try:
            if mct_file.endswith('.xls'):
                import xlrd
                book = xlrd.open_workbook(mct_file)
                for sheet in book.sheets():
                    for r in range(1, sheet.nrows):
                        row = [sheet.cell_value(r, c) for c in range(min(5, sheet.ncols))]
                        if len(row) >= 5 and row[0]:
                            apt = str(row[0]).strip()
                            t_o = str(row[1]).strip() if row[1] else ''
                            t_d = str(row[2]).strip() if row[2] else ''
                            di = str(row[3]).strip() if row[3] else ''
                            try:
                                mins = int(float(row[4]))
                            except (ValueError, TypeError):
                                continue
                            key = (apt, t_o, t_d, di)
                            mct[key] = mins
                            # Also store without terminals for fallback
                            key_no_term = (apt, '', '', di)
                            if key_no_term not in mct:
                                mct[key_no_term] = mins
            else:
                wb = openpyxl.load_workbook(mct_file, data_only=True, read_only=True)
                for ws in wb:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            apt = str(row[0]).strip()
                            t_o = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                            t_d = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                            di = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                            try:
                                mins = int(float(row[4])) if len(row) > 4 and row[4] else default_mct
                            except (ValueError, TypeError):
                                continue
                            key = (apt, t_o, t_d, di)
                            mct[key] = mins
                            key_no_term = (apt, '', '', di)
                            if key_no_term not in mct:
                                mct[key_no_term] = mins
                wb.close()
        except Exception as e:
            print(f"Warning: Could not load MCT file {mct_file}: {e}")
    return mct


def lookup_mct(mct_data, airport, term_arr, term_dep, dom_int, default_mct=90):
    """Look up MCT with cascading fallback: exact  no-terminal  airport-only  default."""
    term_arr = str(term_arr).strip() if term_arr else ''
    term_dep = str(term_dep).strip() if term_dep else ''
    
    # Exact match
    key = (airport, term_arr, term_dep, dom_int)
    if key in mct_data:
        return mct_data[key]
    # Without terminals
    key = (airport, '', '', dom_int)
    if key in mct_data:
        return mct_data[key]
    # Any entry for this airport
    for k, v in mct_data.items():
        if k[0] == airport:
            return v
    return default_mct


# 
# OAG DATA LOADING
# 

_HEADER_MARKERS = {'Carrier1', 'DepAirport', 'ArrAirport',
                   'Carrier Code', 'Dep Airport Code', 'Arr Airport Code'}


def _has_header(row):
    return bool({str(v).strip() for v in row if v is not None} & _HEADER_MARKERS)


def _load_sheet_rows(filepath, sheet_name=None):
    """Return all rows (header included) of the OAG sheet.

    Prefers python-calamine, which loads the large multi-hub pulls in seconds;
    openpyxl streaming times out on 75MB+ files. Falls back to openpyxl if calamine
    is not installed. Resolves the sheet by name, else the single sheet, else the
    first sheet whose top rows carry an OAG header (OAG names sheets by job id).
    """
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(filepath)
        names = wb.sheet_names
        resolved = None
        if sheet_name:
            resolved = next((s for s in names if s.lower() == sheet_name.lower()), None)
        if resolved is None and len(names) == 1:
            resolved = names[0]
        if resolved is None:
            for s in names:
                data = wb.get_sheet_by_name(s).to_python()
                if data and any(_has_header(r) for r in data[:6]):
                    return data
            return []
        return wb.get_sheet_by_name(resolved).to_python()
    except ImportError:
        pass
    except Exception:
        pass
    # Fallback: openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        resolved = None
        if sheet_name:
            resolved = next((s for s in wb.sheetnames if s.lower() == sheet_name.lower()), None)
        if resolved is None:
            for s in wb.sheetnames:
                for row in wb[s].iter_rows(min_row=1, max_row=6, values_only=True):
                    if _has_header(row):
                        resolved = s
                        break
                if resolved:
                    break
            wb.close()
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if resolved is None:
            return []
        return list(wb[resolved].iter_rows(values_only=True))
    finally:
        wb.close()


def load_oag_legs(filepath, sheet_name=None):
    """Load leg data from OAG Parser output or raw OAG format.

    Returns list of dicts with standardised keys.
    Handles both the OAG Parser output format and the raw Connection Builder format.
    The OAG Analyser names its sheet by job id, so when the named sheet is absent
    the loader auto-detects the sheet whose top rows carry an OAG header.
    """
    legs = []
    rows = _load_sheet_rows(filepath, sheet_name)
    if len(rows) < 2:
        return legs
    
    # Find header row (first row with recognizable column names)
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = [str(v).strip() if v else '' for v in row]
        if any(h in row_str for h in ['Carrier1', 'DepAirport', 'ArrAirport', 'carrier', 'dep_airport', 'Carrier Code', 'Dep Airport Code', 'Arr Airport Code']):
            header_row = row_str
            header_idx = i
            break
    
    if header_row is None:
        # Try raw Connection Builder format (columns by position)
        header_row = rows[0]
        header_idx = 0
        if header_row and str(header_row[0]).strip() in ('ID', 'id'):
            # Connection Builder format - columns mapped by VBA positions
            col_map = {
                0: 'id', 1: 'carrier', 2: 'carrier_name', 3: 'flight_no',
                11: 'dep_airport', 12: 'dep_airport_name', 13: 'dep_terminal',
                14: 'dep_city', 18: 'dep_country',
                24: 'arr_airport', 25: 'arr_airport_name', 26: 'arr_terminal',
                27: 'arr_city', 31: 'arr_country',
                37: 'dep_time', 38: 'arr_time', 40: 'dep_days', 41: 'arr_days',
                56: 'elapsed_time', 57: 'flying_time',
            }
        else:
            return legs
    
    # Build column index from header
    col_idx = {}
    header_lower = [str(h).lower().strip() if h else '' for h in header_row]
    
    # Standard OAG column mappings
    # Aliases cover both machine names (Carrier1, DepAirport) and OAG Analyser
    # spaced headers (Carrier Code, Dep Airport Code) - matched case-insensitively.
    mappings = {
        'carrier': ['carrier1', 'carrier', 'airline', 'carrier code'],
        'flight_no': ['flightno1', 'flight_no', 'flight', 'flight no'],
        'dep_airport': ['depairport', 'dep_airport', 'origin', 'dep airport code'],
        'dep_terminal': ['depterminal', 'dep_terminal', 'dep terminal'],
        'dep_city': ['depcity', 'dep_city', 'dep city code'],
        'dep_country': ['depiatactry', 'dep_country', 'dep iata country code'],
        'arr_airport': ['arrairport', 'arr_airport', 'destination', 'arr airport code'],
        'arr_terminal': ['arrterminal', 'arr_terminal', 'arr terminal'],
        'arr_city': ['arrcity', 'arr_city', 'arr city code'],
        'arr_country': ['arriatactry', 'arr_country', 'arr iata country code'],
        'dep_time': ['localdeptime', 'dep_time', 'departure', 'local dep time'],
        'arr_time': ['localarrtime', 'arr_time', 'arrival', 'local arr time'],
        'dep_days': ['localdaysofop', 'dep_days', 'days_of_op', 'local days of op'],
        'arr_days': ['arrdaysofop', 'arr_days', 'local days of op arr'],
        'flying_time': ['flyingtime', 'flying_time', 'flying time'],
        'elapsed_time': ['elapsedtime', 'elapsed_time', 'elapsed time'],
        'id': ['id'],
        'seats': ['seats', 'seats (total)'],
        # Newer OAG pulls (optional, absent in older files)
        'alliance': ['carrier1alliance', 'carrier alliance', 'alliance'],
        'carrier_category': ['mainline/low cost', 'carrier category'],
        'dup_marker': ['dup marker', 'dupmarker'],
        'gcd_km': ['gcd (km)', 'gcd km', 'distkm'],
        'gcd_mi': ['gcd (m)', 'gcd (mi)', 'diststmiles'],
        'ask': ['asks', 'ask'],
        'pass_class': ['pass class', 'passclass'],
    }
    
    for field, aliases in mappings.items():
        for alias in aliases:
            if alias in header_lower:
                col_idx[field] = header_lower.index(alias)
                break
    
    # Parse data rows
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if row is None or all(v is None for v in row):
            continue
        
        rec = {}
        for field, idx in col_idx.items():
            if idx < len(row):
                rec[field] = row[idx]
        
        # Skip if no carrier or airports
        if not rec.get('carrier') or not rec.get('dep_airport'):
            continue
        
        # Clean values
        rec['carrier'] = str(rec['carrier']).strip()
        rec['dep_airport'] = str(rec['dep_airport']).strip()
        rec['arr_airport'] = str(rec.get('arr_airport', '')).strip()
        rec['dep_terminal'] = str(rec.get('dep_terminal', '')).strip()
        rec['arr_terminal'] = str(rec.get('arr_terminal', '')).strip()
        rec['dep_country'] = str(rec.get('dep_country', '')).strip()
        rec['arr_country'] = str(rec.get('arr_country', '')).strip()
        rec['dep_city'] = str(rec.get('dep_city', '')).strip()
        rec['arr_city'] = str(rec.get('arr_city', '')).strip()
        rec['flight_no'] = str(rec.get('flight_no', '')).strip()
        
        # Parse times
        rec['arr_time_mins'] = parse_time_hhmm(rec.get('arr_time'))
        rec['dep_time_mins'] = parse_time_hhmm(rec.get('dep_time'))
        
        # Parse flying time (handles 'HH:MM:SS', 'HH:MM' and 'HHMM'/int formats)
        rec['flying_mins'] = _parse_duration_mins(rec.get('flying_time') or rec.get('elapsed_time'))
        
        # Parse days of operation
        rec['dep_day_set'] = parse_days_string(rec.get('dep_days'))
        rec['arr_day_set'] = parse_days_string(rec.get('arr_days'))
        
        # Assign ID if missing
        if 'id' not in rec or rec['id'] is None:
            rec['id'] = i
        else:
            try:
                rec['id'] = int(float(rec['id']))
            except (ValueError, TypeError):
                rec['id'] = i
        
        # DOM/INT for the leg
        rec['dom_int'] = get_dom_int(rec['dep_country'], rec['arr_country'])

        # Optional enrichment fields (present in newer OAG pulls; absent OK)
        rec['alliance'] = str(rec.get('alliance', '') or '').strip()
        rec['carrier_category'] = str(rec.get('carrier_category', '') or '').strip()
        rec['dup_marker'] = str(rec.get('dup_marker', '') or '').strip()
        rec['pass_class'] = str(rec.get('pass_class', '') or '').strip()
        try:
            rec['gcd_km'] = float(rec['gcd_km']) if rec.get('gcd_km') not in (None, '') else None
        except (ValueError, TypeError):
            rec['gcd_km'] = None

        legs.append(rec)
    
    return legs


def load_alliance_data(alliance_file=None):
    """Load alliance memberships. Returns list of sets."""
    if alliance_file and os.path.exists(alliance_file):
        try:
            alliances = [set(), set(), set()]  # oneworld, star, skyteam
            wb = openpyxl.load_workbook(alliance_file, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row:
                    if len(row) > 1 and row[1]:
                        alliances[0].add(str(row[1]).strip())
                    if len(row) > 3 and row[3]:
                        alliances[1].add(str(row[3]).strip())
                    if len(row) > 5 and row[5]:
                        alliances[2].add(str(row[5]).strip())
            wb.close()
            return alliances
        except Exception as e:
            print(f"Warning: Could not load alliance file: {e}")
    return [ONEWORLD, STAR_ALLIANCE, SKYTEAM]


def load_lcc_list(lcc_file=None):
    """Load LCC carrier list from file or use defaults."""
    if lcc_file and os.path.exists(lcc_file):
        try:
            with open(lcc_file) as f:
                return {line.strip() for line in f if line.strip()}
        except Exception:
            pass
    return DEFAULT_LCC_LIST


# 
# PROPOSED SERVICE INJECTION
# 

def create_proposed_service(origin, destination, carrier, freq, dep_time,
                           arr_time, dep_days, flying_time,
                           dep_terminal='', arr_terminal='',
                           dep_country='', arr_country='',
                           dep_city='', arr_city=''):
    """Create a proposed service record matching leg data format."""
    dep_mins = parse_time_hhmm(dep_time)
    arr_mins = parse_time_hhmm(arr_time)
    day_set = parse_days_string(dep_days)
    
    # For arrival days, calculate based on dep_time + flying_time
    # If arrival is next day, shift days
    total_mins = dep_mins + flying_time if dep_mins is not None else 0
    next_day = total_mins >= 1440
    
    if next_day:
        arr_day_set = {(d % 7) + 1 for d in day_set}  # shift by 1 day
    else:
        arr_day_set = day_set.copy()
    
    ft_hhmm = (flying_time // 60) * 100 + (flying_time % 60)
    
    return {
        'id': 99999,
        'carrier': carrier,
        'flight_no': 'NEW',
        'dep_airport': origin,
        'arr_airport': destination,
        'dep_terminal': dep_terminal,
        'arr_terminal': arr_terminal,
        'dep_city': dep_city or origin,
        'arr_city': arr_city or destination,
        'dep_country': dep_country,
        'arr_country': arr_country,
        'dep_time': dep_time,
        'arr_time': arr_time,
        'dep_time_mins': dep_mins,
        'arr_time_mins': arr_mins,
        'dep_days': dep_days,
        'arr_days': dep_days,  # simplified
        'dep_day_set': day_set,
        'arr_day_set': arr_day_set,
        'flying_time': ft_hhmm,
        'flying_mins': flying_time,
        'dom_int': get_dom_int(dep_country, arr_country),
        'seats': 0,
        'is_proposed': True,
    }


# 
# CORE CONNECTION BUILDER ALGORITHM
# Replicates VBA: match_bytime  match_byday  output  connection_check
# 

def build_connections(leg1_data, leg2_data, alliances, mct_data, lcc_set,
                      min_connect=20, max_connect=720, default_mct=90,
                      hub_airport=None):
    """
    Core connection building algorithm.
    
    leg1_data: list of dicts (flights arriving at hub  OriginHub)
    leg2_data: list of dicts (flights departing hub  HubBeyond)
    alliances: list of sets [oneworld, star, skyteam]
    mct_data: dict of MCT lookups
    lcc_set: set of LCC carrier codes to exclude
    
    Returns: list of connection dicts with all output fields
    """
    
    # Step 1: Group leg1 by arriving airport (connection point)
    leg1_by_hub = defaultdict(list)
    for leg in leg1_data:
        arr = leg['arr_airport']
        # Exclude LCC carriers from connections
        if leg['carrier'] in lcc_set:
            continue
        leg1_by_hub[arr].append(leg)
    
    # Step 2: Group leg2 by departing airport (connection point)
    leg2_by_hub = defaultdict(list)
    for leg in leg2_data:
        dep = leg['dep_airport']
        if leg['carrier'] in lcc_set:
            continue
        leg2_by_hub[dep].append(leg)
    
    # Step 3: Find common connection points
    if hub_airport:
        connection_airports = [hub_airport]
    else:
        connection_airports = sorted(set(leg1_by_hub.keys()) & set(leg2_by_hub.keys()))
    
    connections = []
    
    for cnx_apt in connection_airports:
        arrivals = leg1_by_hub.get(cnx_apt, [])
        departures = leg2_by_hub.get(cnx_apt, [])
        
        if not arrivals or not departures:
            continue
        
        # Step 4: match_bytime  find all time-feasible pairs
        time_matches = []
        for arr_leg in arrivals:
            arr_time = arr_leg['arr_time_mins']
            if arr_time is None:
                continue
            
            for dep_leg in departures:
                dep_time = dep_leg['dep_time_mins']
                if dep_time is None:
                    continue
                
                # Calculate connection time
                cnx_time = dep_time - arr_time
                if cnx_time < 0:
                    cnx_time = 1440 + cnx_time  # overnight connection
                
                # Check within connect window
                if min_connect < cnx_time < max_connect:
                    # Calculate total elapsed time
                    elapsed = arr_leg['flying_mins'] + cnx_time + dep_leg['flying_mins']
                    
                    # Connection type (DOM/INT)
                    cnx_dom_int = arr_leg['dom_int'] + dep_leg['dom_int']
                    
                    time_matches.append({
                        'leg1': arr_leg,
                        'leg2': dep_leg,
                        'cnx_time': cnx_time,
                        'elapsed': elapsed,
                        'cnx_dom_int': cnx_dom_int,
                        'is_overnight': dep_time < arr_time,
                    })
        
        # Step 5: match_byday  check day-of-week feasibility and count frequency
        for tm in time_matches:
            leg1 = tm['leg1']
            leg2 = tm['leg2']
            
            frequency = 0
            
            if tm['is_overnight']:
                # Overnight: leg1 arrives day N, leg2 departs day N+1
                for day in range(1, 8):
                    if day in leg1.get('arr_day_set', leg1.get('dep_day_set', set())):
                        next_day = (day % 7) + 1
                        if next_day in leg2.get('dep_day_set', set()):
                            frequency += 1
            else:
                # Same-day connection
                arr_days = leg1.get('arr_day_set', leg1.get('dep_day_set', set()))
                dep_days = leg2.get('dep_day_set', set())
                frequency = len(arr_days & dep_days)
            
            if frequency > 0:
                # Step 6: Classify connection type
                carrier1 = leg1['carrier']
                carrier2 = leg2['carrier']
                cnx_type = classify_connection(carrier1, carrier2, alliances)
                
                # Step 7: MCT check
                mct_val = lookup_mct(
                    mct_data, cnx_apt,
                    leg1.get('arr_terminal', ''),
                    leg2.get('dep_terminal', ''),
                    tm['cnx_dom_int'],
                    default_mct
                )
                
                mct_pass = tm['cnx_time'] >= mct_val
                
                # Build route label
                dest_apt = leg2['arr_airport']
                route_label = f"{dest_apt}-{carrier1}-{cnx_apt}-{carrier2}-{leg1['dep_airport']}"
                
                connections.append({
                    'city_label': leg2.get('arr_city', dest_apt),
                    'airport_label': dest_apt,
                    'route_label': route_label,
                    'dep_airport': leg1['dep_airport'],
                    'leg1_flight': leg1.get('flight_no', ''),
                    'leg1_carrier': carrier1,
                    'arr_time_at_cnx': minutes_to_hhmm(leg1['arr_time_mins']) if leg1['arr_time_mins'] is not None else '',
                    'cnx_airport': cnx_apt,
                    'dep_time_from_cnx': minutes_to_hhmm(leg2['dep_time_mins']) if leg2['dep_time_mins'] is not None else '',
                    'cnx_time': tm['cnx_time'],
                    'leg1_arr_days': leg1.get('arr_days', leg1.get('dep_days', '')),
                    'leg2_dep_days': leg2.get('dep_days', ''),
                    'leg1_arr_terminal': leg1.get('arr_terminal', ''),
                    'leg2_dep_terminal': leg2.get('dep_terminal', ''),
                    'dom_int_transfer': tm['cnx_dom_int'],
                    'arr_airport': dest_apt,
                    'leg2_flight': leg2.get('flight_no', ''),
                    'leg2_carrier': carrier2,
                    'frequency': frequency,
                    'elapsed_time': tm['elapsed'],
                    'cnx_type': cnx_type,
                    'mct': mct_val,
                    'mct_pass': mct_pass,
                    'leg1_id': leg1.get('id', ''),
                    'leg2_id': leg2.get('id', ''),
                    'leg1_flying': leg1['flying_mins'],
                    'leg2_flying': leg2['flying_mins'],
                    'leg1_is_proposed': leg1.get('is_proposed', False),
                    'leg2_is_proposed': leg2.get('is_proposed', False),
                })
    
    # Step 8: Filter out MCT failures (replicating connection_check in VBA)
    valid = [c for c in connections if c['mct_pass']]
    failed = [c for c in connections if not c['mct_pass']]
    
    return valid, failed


# 
# EXCEL OUTPUT
# 

HEADER_FONT = Font(name='Arial', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='003366')
DATA_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

# Output columns matching QSI template input format (QSI 1 sheet)
OUTPUT_COLUMNS = [
    ('City Label', 12),
    ('Airport Label', 8),
    ('Route Label', 30),
    ('Dep. Aprt.', 8),
    ('Flight No.', 10),
    ('Carrier', 8),
    ('Arr. Time at Connex', 14),
    ('Conex Apt.', 8),
    ('Dep. Time from Connex', 14),
    ('Conex Time', 10),
    ('Days of Arrival', 12),
    ('Days of Departure', 12),
    ('Arrival Terminal', 10),
    ('Departure Terminal', 10),
    ('DOM/INT Transfer', 10),
    ('Arr. Apt', 8),
    ('Flight No.', 10),
    ('Carrier', 8),
    ('Freq.', 6),
    ('Elapsed Time', 10),
    ('Conex Type', 12),
    ('MCT', 6),
    ('MCT Check', 8),
]


def write_output(connections, failed, metadata, output_path):
    """Write connection builder output to Excel."""
    wb = openpyxl.Workbook()
    
    #  Metadata sheet 
    ws_meta = wb.active
    ws_meta.title = 'Metadata'
    meta_items = [
        ('Connection Builder Output', ''),
        ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('', ''),
    ]
    for k, v in metadata.items():
        meta_items.append((k, v))
    meta_items.extend([
        ('', ''),
        ('Total Valid Connections', len(connections)),
        ('Total MCT Failures', len(failed)),
        ('Unique Destinations', len(set(c['airport_label'] for c in connections))),
        ('Unique Connection Airports', len(set(c['cnx_airport'] for c in connections))),
    ])
    for i, (k, v) in enumerate(meta_items, 1):
        ws_meta.cell(row=i, column=1, value=k).font = Font(name='Arial', bold=True, size=10)
        ws_meta.cell(row=i, column=2, value=v).font = DATA_FONT
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 40
    
    #  Main Outputs sheet (matches Connection Builder format) 
    ws = wb.create_sheet('Outputs')
    
    # Header rows
    ws.cell(row=1, column=1, value='Connection Builder Outputs').font = Font(name='Arial', bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Date: {datetime.now().strftime('%Y-%m-%d')}").font = DATA_FONT
    ws.cell(row=2, column=4, value=f"Total Connections: {len(connections)}").font = DATA_FONT
    ws.cell(row=2, column=7, value=f"Unique Routings: {len(set(c['airport_label'] for c in connections))}").font = DATA_FONT
    
    # Column headers (row 4)
    for col, (header, width) in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Data rows (from row 5)
    for i, c in enumerate(connections, 5):
        row_data = [
            c['city_label'],
            c['airport_label'],
            c['route_label'],
            c['dep_airport'],
            c['leg1_flight'],
            c['leg1_carrier'],
            c['arr_time_at_cnx'],
            c['cnx_airport'],
            c['dep_time_from_cnx'],
            c['cnx_time'],
            c['leg1_arr_days'],
            c['leg2_dep_days'],
            c['leg1_arr_terminal'],
            c['leg2_dep_terminal'],
            c['dom_int_transfer'],
            c['arr_airport'],
            c['leg2_flight'],
            c['leg2_carrier'],
            c['frequency'],
            c['elapsed_time'],
            c['cnx_type'],
            c['mct'],
            'OK' if c['mct_pass'] else 'FAIL',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
    
    #  QSI Ready sheet (formatted for direct paste into QSI template) 
    ws_qsi = wb.create_sheet('QSI_Ready')
    
    # QSI template header format
    ws_qsi.cell(row=2, column=4, value='Output from Connection Builder').font = Font(name='Arial', bold=True, size=11)
    
    qsi_headers = [
        'City Label', 'Airport Label', 'Route Label',
        'Dep. Aprt.', 'Flight No.', 'Carrier',
        'Arr. Time at Connex', 'Conex Apt.', 'Dep. Time at Dest',
        'Conex Time', 'Days of Arrival', 'Days of Departure',
        'Departure Terminal', 'Arrival Terminal', 'DOM/INT Transfer',
        'Arr. Apt', 'Flight No.', 'Carrier',
        'Freq.', 'Elapsed Time', 'Conex Type', 'MCT',
    ]
    for col, header in enumerate(qsi_headers, 1):
        cell = ws_qsi.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    for i, c in enumerate(connections, 6):
        row_data = [
            c['city_label'], c['airport_label'], c['route_label'],
            c['dep_airport'], c['leg1_flight'], c['leg1_carrier'],
            c['arr_time_at_cnx'], c['cnx_airport'], c['dep_time_from_cnx'],
            c['cnx_time'], c['leg1_arr_days'], c['leg2_dep_days'],
            c['leg1_arr_terminal'], c['leg2_dep_terminal'], c['dom_int_transfer'],
            c['arr_airport'], c['leg2_flight'], c['leg2_carrier'],
            c['frequency'], c['elapsed_time'], c['cnx_type'], c['mct'],
        ]
        for col, val in enumerate(row_data, 1):
            ws_qsi.cell(row=i, column=col, value=val).font = DATA_FONT
    
    #  Summary sheet (connections aggregated by destination city) 
    ws_sum = wb.create_sheet('Summary')
    
    # Aggregate by destination city
    city_stats = defaultdict(lambda: {'count': 0, 'freq_total': 0, 'min_elapsed': 99999,
                                       'online': 0, 'alliance': 0, 'interline': 0,
                                       'airports': set(), 'carriers': set()})
    for c in connections:
        city = c['city_label']
        cs = city_stats[city]
        cs['count'] += 1
        cs['freq_total'] += c['frequency']
        cs['min_elapsed'] = min(cs['min_elapsed'], c['elapsed_time'])
        cs['airports'].add(c['airport_label'])
        cs['carriers'].add(c['leg2_carrier'])
        if c['cnx_type'] == 'ONLINE':
            cs['online'] += c['frequency']
        elif c['cnx_type'] == 'ALLIANCE':
            cs['alliance'] += c['frequency']
        else:
            cs['interline'] += c['frequency']
    
    sum_headers = ['City', 'Airport(s)', 'Routings', 'Total Weekly Freq',
                   'Min Elapsed (mins)', 'Online Freq', 'Alliance Freq', 'Interline Freq',
                   'Carriers']
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    for i, (city, cs) in enumerate(sorted(city_stats.items()), 2):
        ws_sum.cell(row=i, column=1, value=city).font = DATA_FONT
        ws_sum.cell(row=i, column=2, value=', '.join(sorted(cs['airports']))).font = DATA_FONT
        ws_sum.cell(row=i, column=3, value=cs['count']).font = DATA_FONT
        ws_sum.cell(row=i, column=4, value=cs['freq_total']).font = DATA_FONT
        ws_sum.cell(row=i, column=5, value=cs['min_elapsed']).font = DATA_FONT
        ws_sum.cell(row=i, column=6, value=cs['online']).font = DATA_FONT
        ws_sum.cell(row=i, column=7, value=cs['alliance']).font = DATA_FONT
        ws_sum.cell(row=i, column=8, value=cs['interline']).font = DATA_FONT
        ws_sum.cell(row=i, column=9, value=', '.join(sorted(cs['carriers']))).font = DATA_FONT
    
    for col in range(1, 10):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    #  MCT Failures sheet 
    if failed:
        ws_fail = wb.create_sheet('MCT_Failures')
        ws_fail.cell(row=1, column=1, value='Connections removed due to MCT violation').font = Font(name='Arial', bold=True)
        fail_headers = ['Route Label', 'Cnx Airport', 'Cnx Time', 'MCT', 'Shortfall', 'Cnx Type']
        for col, h in enumerate(fail_headers, 1):
            ws_fail.cell(row=3, column=col, value=h).font = HEADER_FONT
            ws_fail.cell(row=3, column=col).fill = PatternFill('solid', fgColor='CC0000')
        for i, c in enumerate(failed, 4):
            ws_fail.cell(row=i, column=1, value=c['route_label'])
            ws_fail.cell(row=i, column=2, value=c['cnx_airport'])
            ws_fail.cell(row=i, column=3, value=c['cnx_time'])
            ws_fail.cell(row=i, column=4, value=c['mct'])
            ws_fail.cell(row=i, column=5, value=c['mct'] - c['cnx_time'])
            ws_fail.cell(row=i, column=6, value=c['cnx_type'])
    
    wb.save(output_path)
    return len(connections), len(failed), len(city_stats)


# 
# MAIN  CLI INTERFACE
# 

def main():
    parser = argparse.ArgumentParser(
        description='Connection Builder  Avia Solutions QSI Pipeline Module III',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Without new service (baseline)
  python connection_builder.py oag_parsed.xlsx --origin LHR --destination SJC

  # With proposed BA service
  python connection_builder.py oag_parsed.xlsx --origin LHR --destination SJC \\
      --carrier BA --freq 7 --dep-time 1555 --arr-time 1810 \\
      --dep-days 1234567 --flying-time 660 --with-service
        """
    )
    parser.add_argument('oag_file', help='OAG parsed data file (xlsx)')
    parser.add_argument('--origin', required=True, help='Origin airport IATA code')
    parser.add_argument('--destination', required=True, help='Destination airport IATA code')
    parser.add_argument('--carrier', default='', help='Proposed carrier code')
    parser.add_argument('--freq', type=int, default=7, help='Proposed weekly frequency')
    parser.add_argument('--dep-time', type=int, default=0, help='Proposed departure time HHMM')
    parser.add_argument('--arr-time', type=int, default=0, help='Proposed arrival time HHMM')
    parser.add_argument('--dep-days', default='1234567', help='Days of operation')
    parser.add_argument('--flying-time', type=int, default=0, help='Flying time in minutes')
    parser.add_argument('--with-service', action='store_true', help='Include proposed new service')
    parser.add_argument('--output', default=None, help='Output file path')
    parser.add_argument('--mct-file', default=None, help='MCT data file')
    parser.add_argument('--alliance-file', default=None, help='Alliance membership file')
    parser.add_argument('--lcc-list', default=None, help='LCC carrier list file')
    parser.add_argument('--default-mct', type=int, default=90, help='Default MCT (minutes)')
    parser.add_argument('--min-connect', type=int, default=20, help='Min connection window')
    parser.add_argument('--max-connect', type=int, default=720, help='Max connection window')
    parser.add_argument('--hub', default=None, help='Restrict to single hub airport')
    parser.add_argument('--leg1-sheet', default='leg1', help='Sheet name for leg1 data')
    parser.add_argument('--leg2-sheet', default='leg2', help='Sheet name for leg2 data')
    
    args = parser.parse_args()
    
    origin = args.origin.upper()
    destination = args.destination.upper()
    
    print(f"Connection Builder  {origin}  {destination}")
    print(f"{'='*50}")
    
    # Load data
    print("Loading OAG leg data...")
    leg1_data = load_oag_legs(args.oag_file, args.leg1_sheet)
    leg2_data = load_oag_legs(args.oag_file, args.leg2_sheet)
    print(f"  Leg 1 (arrivals at hub): {len(leg1_data)} flights")
    print(f"  Leg 2 (departures from hub): {len(leg2_data)} flights")
    
    # Load supporting data
    alliances = load_alliance_data(args.alliance_file)
    mct_data = load_mct_data(args.mct_file, args.default_mct)
    lcc_set = load_lcc_list(args.lcc_list)
    print(f"  MCT entries: {len(mct_data)}")
    print(f"  LCC carriers excluded: {len(lcc_set)}")
    
    # Add proposed service if requested
    if args.with_service and args.carrier:
        print(f"\nInjecting proposed service: {args.carrier} {origin}{destination}")
        print(f"  Dep: {args.dep_time}, Arr: {args.arr_time}, Freq: {args.freq}x/wk")
        
        # The proposed service appears as:
        # - A leg1 entry: arriving at destination (for connections BEYOND destination)
        # - A leg2 entry: departing from origin (for connections BEYOND origin)
        # - Also as leg1 arriving at origin if running from destination perspective
        
        # For origindestination: new service is leg1 arriving at destination
        proposed_leg1 = create_proposed_service(
            origin, destination, args.carrier, args.freq,
            args.dep_time, args.arr_time, args.dep_days, args.flying_time,
        )
        leg1_data.append(proposed_leg1)
        
        # Reverse: destinationorigin service for the other perspective
        # (arrival time at origin would need separate calculation)
        
        print(f"  Added to leg1 data (arrives at {destination})")
    
    # Build connections
    print(f"\nBuilding connections...")
    hub = args.hub.upper() if args.hub else None
    
    valid, failed = build_connections(
        leg1_data, leg2_data, alliances, mct_data, lcc_set,
        min_connect=args.min_connect,
        max_connect=args.max_connect,
        default_mct=args.default_mct,
        hub_airport=hub,
    )
    
    print(f"  Valid connections: {len(valid)}")
    print(f"  MCT failures: {len(failed)}")
    print(f"  Unique destinations: {len(set(c['airport_label'] for c in valid))}")
    
    # Connection type breakdown
    type_counts = defaultdict(int)
    for c in valid:
        type_counts[c['cnx_type']] += c['frequency']
    print(f"  Weekly freq by type: {dict(type_counts)}")
    
    # Output
    if args.output is None:
        svc = 'with' if args.with_service else 'without'
        args.output = f"CnxBuilder_{origin}_{destination}_{svc}_service.xlsx"
    
    metadata = {
        'Origin': origin,
        'Destination': destination,
        'Hub Filter': hub or 'All hubs',
        'Mode': 'With new service' if args.with_service else 'Without new service',
        'Carrier': args.carrier or 'N/A',
        'Proposed Frequency': f"{args.freq}x/wk" if args.with_service else 'N/A',
        'Proposed Dep Time': str(args.dep_time) if args.with_service else 'N/A',
        'Min Connect': f"{args.min_connect} min",
        'Max Connect': f"{args.max_connect} min",
        'Default MCT': f"{args.default_mct} min",
        'MCT File': args.mct_file or 'None (using defaults)',
        'LCC Exclusions': f"{len(lcc_set)} carriers",
        'OAG Source': args.oag_file,
    }
    
    n_valid, n_failed, n_cities = write_output(valid, failed, metadata, args.output)
    print(f"\nOutput: {args.output}")
    print(f"  {n_valid} connections to {n_cities} destination cities")
    if n_failed:
        print(f"  {n_failed} connections removed (MCT failure)")
    
    return valid, failed


if __name__ == '__main__':
    main()
