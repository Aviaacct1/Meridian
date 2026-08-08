#!/usr/bin/env python3
"""Offline test of the ACI loader. Builds a wide workbook shaped like the real
one, with its awkwardnesses, and checks the melt and the rules.

    py -3.12 test_load_aci.py

The fixture reproduces, deliberately:
  * airports in rows, months in columns, header not on row 1
  * month headers in three formats, as the real sheet has them after nearly
    twenty years of hand editing: real dates, "Jan-2012", "Jan 12"
  * BLANK cells, which mean the airport did not report and must never become 0
  * a dash and an "n/a", which mean the same thing
  * junk rows above the header and a total row with no IATA code
  * an airport that reports some years and not others
  * ONE AIRPORT ON TWO ROWS, complementary in time, which must merge
  * and, in a second fixture, one airport on two rows carrying the SAME month,
    which must stop the build rather than double that airport silently

Every number here is a TEST FIXTURE.

Avia Solutions Limited. All rights reserved.
"""
import datetime as dt
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import load_aci as LA

FAIL = []
CHECKS = 0


def check(name, cond, detail=""):
    global CHECKS
    CHECKS += 1
    print("%-58s %s %s" % (name, "PASS" if cond else "FAIL", detail))
    if not cond:
        FAIL.append(name)


def build_fixture(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1-Monthly"
    ws.append(["ACI monthly passengers", None, None])          # junk row 1
    ws.append([None, None, None])                              # junk row 2
    # header: region, country, city, IATA, then months in three formats.
    # THE COUNTRY NAME IS LONGER THAN THE CITY NAME. That is the real sheet's
    # shape and the defect the first build shipped: "United Kingdom" beats
    # "EDINBURGH" on length, so every British airport was named after its
    # country. The name has to be chosen as a COLUMN, on distinctness.
    ws.append(["REG", "COUNTRY", "CITY", "CODE",
               dt.datetime(2023, 11, 1), "Dec-2023", "Jan 24",
               "2024-02", "Mar-2024"])
    ws.append(["EUR", "United Kingdom", "EDINBURGH", "EDI",
               1000, 1100, 900, 950, 1200])
    # AUS did not report two months: one blank, one dash
    ws.append(["NAM", "United States of America", "AUSTIN TX", "AUS",
               2000, None, 2100, "-", 2300])
    # LHR uses "n/a" and a comma-formatted string
    ws.append(["EUR", "United Kingdom", "LONDON HEATHROW", "LHR",
               "6,000", "n/a", 6500, 6600, 6700])
    # more airports so the detection is genuinely exercised: the REG column
    # (EUR/NAM/ASP, also three upper-case letters) gets a chance to win the
    # naive count and must lose on distinctness, and the country column carries
    # enough distinct values to be told apart from the region
    for i, (reg, ctry, code) in enumerate([
            ("EUR", "France", "CDG"), ("EUR", "Netherlands", "AMS"),
            ("EUR", "Germany", "FRA"), ("EUR", "Spain", "MAD"),
            ("NAM", "United States of America", "JFK"),
            ("NAM", "United States of America", "LAX"),
            ("NAM", "United States of America", "ORD"),
            ("ASP", "Singapore", "SIN"), ("ASP", "Japan", "NRT"),
            ("ASP", "Australia", "SYD"), ("LAC", "Brazil", "GRU"),
            ("AFR", "South Africa", "JNB")]):
        ws.append([reg, ctry, "%s CITY" % code, code,
                   1000 + i, 1100 + i, 1200 + i, 1300 + i, 1400 + i])
    # OSL on TWO rows, split in time as a hand-maintained sheet splits an
    # airport across a rename. The rows do not overlap, so the store must carry
    # all five months once each, not two partial airports and not a double.
    ws.append(["EUR", "Norway", "OSLO GARDERMOEN", "OSL",
               500, 550, None, None, None])
    ws.append(["EUR", "Norway", "OSLO (from 2024)", "OSL",
               None, None, 600, 650, 700])
    ws.append([None, None, "TOTAL", None,
               9000, 7100, 9500, 7550, 10200])           # no IATA
    wb.save(path)


def build_colliding_fixture(path):
    """The dangerous shape: the same airport and the same month, twice."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1-Monthly"
    ws.append(["REG", "CITY", "CODE", "Jan-2024", "Feb-2024", "Mar-2024"])
    ws.append(["EUR", "EDINBURGH", "EDI", 1000, 1100, 1200])
    ws.append(["EUR", "AMSTERDAM", "AMS", 2000, 2100, 2200])
    ws.append(["EUR", "PARIS CDG", "CDG", 3000, 3100, 3200])
    ws.append(["EUR", "PARIS CDG duplicate line", "CDG", 3000, 3100, 3200])
    wb.save(path)


tmp_x = os.path.join(tempfile.gettempdir(), "aci_fixture.xlsx")
tmp_db = os.path.join(tempfile.gettempdir(), "aci_test.duckdb")
tmp_c = os.path.join(tempfile.gettempdir(), "aci_collide.xlsx")
tmp_cdb = os.path.join(tempfile.gettempdir(), "aci_collide.duckdb")
build_fixture(tmp_x)
build_colliding_fixture(tmp_c)

rows, rep = LA.read_sheet(tmp_x)

# --- 1. the melt ---------------------------------------------------------------
check("header row found below the junk", rep["header_row"] == 3, rep["header_row"])
check("all five month columns parsed", rep["month_columns"] == 5, rep["month_columns"])
check("the IATA column is found, not the region column",
      rep["iata_column"] == 4, rep["iata_column"])
check("month range read correctly",
      rep["first_month"] == "2023-11" and rep["last_month"] == "2024-03",
      (rep["first_month"], rep["last_month"]))

# --- 2. the three header formats all resolve -----------------------------------
check("a real date header parses", LA.parse_month(dt.datetime(2024, 6, 1)) == (2024, 6))
check("'Jan-2012' parses", LA.parse_month("Jan-2012") == (2012, 1))
check("'Jan 12' parses to 2012", LA.parse_month("Jan 12") == (2012, 1))
check("'2024-02' parses", LA.parse_month("2024-02") == (2024, 2))
check("a non-month header is rejected", LA.parse_month("CODE") is None)
check("month 13 is rejected", LA.parse_month("2024-13") is None)

# --- 3. A BLANK IS NOT A ZERO. The rule that matters most. ---------------------
aus = {ym: p for c, _n, _ct, ym, _y, _m, p in rows if c == "AUS"}
check("AUS reports only the three months it has", len(aus) == 3, sorted(aus))
check("the blank month is ABSENT, not zero",
      "2023-12" not in aus, sorted(aus))
check("the dash month is ABSENT, not zero",
      "2024-02" not in aus, sorted(aus))
check("no zero passenger row was invented anywhere",
      all(p != 0 for _c, _n, _ct, _ym, _y, _m, p in rows))
check("the not-reported cells are counted and reported",
      rep["cells_not_reported"] == 8, rep["cells_not_reported"])

# --- 4. number cleaning ---------------------------------------------------------
lhr = {ym: p for c, _n, _ct, ym, _y, _m, p in rows if c == "LHR"}
check("a comma-formatted number is read", lhr.get("2023-11") == 6000.0, lhr.get("2023-11"))
check("'n/a' is treated as not reported", "2023-12" not in lhr, sorted(lhr))

# --- 5. rows without a code are excluded and counted ---------------------------
check("the TOTAL row is not loaded as an airport",
      not any(c == "TOT" or _n == "TOTAL" for c, _n, *_ in rows))
check("and it is counted as skipped", rep["rows_without_iata"] >= 1,
      rep["rows_without_iata"])
names = {c: n for c, n, *_ in rows}
countries = {c: ct for c, _n, ct, *_ in rows}
check("airport names come through", "EDINBURGH" in (names.get("EDI") or "").upper(),
      names.get("EDI"))
check("and the region code is NOT used as the name",
      (names.get("EDI") or "") != "EUR" and (names.get("JFK") or "") != "NAM",
      (names.get("EDI"), names.get("JFK")))
# The shipped defect: the country name is longer than the city name, so a
# longest-text rule named every UK airport "United Kingdom".
check("the COUNTRY is not used as the name, even when it is longer",
      names.get("EDI") == "EDINBURGH" and names.get("LHR") == "LONDON HEATHROW",
      (names.get("EDI"), names.get("LHR")))
check("the name column is chosen once, not row by row",
      rep["name_column"] == 3, rep["name_column"])
check("the country is captured in its own right",
      countries.get("EDI") == "United Kingdom"
      and countries.get("AUS") == "United States of America",
      (countries.get("EDI"), countries.get("AUS")))
check("and the country column is the country, not the region",
      rep["country_column"] == 2 and "EUR" not in set(countries.values()),
      (rep["country_column"], sorted(set(countries.values()))[:4]))
# A sheet with no country column at all must hold no country rather than
# quietly promoting the six-value region column into one.
rows_nc, rep_nc = LA.read_sheet(tmp_c)
check("a sheet without a country column holds none",
      rep_nc["country_column"] == "none"
      and all(ct == "" for _c, _n, ct, *_ in rows_nc),
      rep_nc["country_column"])

# --- 5b. one airport on two rows ------------------------------------------------
# 1,736 data rows against 1,732 distinct codes in the real file is why this is
# here. Complementary rows merge; overlapping rows must stop the build.
osl = {ym: p for c, _n, _ct, ym, _y, _m, p in rows if c == "OSL"}
check("an airport split over two rows merges to one series",
      len(osl) == 5, sorted(osl))
check("and neither half is doubled",
      osl.get("2023-11") == 500.0 and osl.get("2024-03") == 700.0, osl)
check("the split is reported, not passed over in silence",
      "OSL" in rep["codes_on_more_than_one_row"], rep["codes_on_more_than_one_row"])
check("complementary rows are not a collision",
      rep["colliding_cells"] == 0, rep["colliding_cells"])

rows_c, rep_c = rows_nc, rep_nc
check("a repeated airport-month is detected", rep_c["colliding_cells"] == 3,
      rep_c["colliding_cells"])
check("and the offending airport is named",
      "CDG" in rep_c["_collisions"], list(rep_c["_collisions"]))
check("the colliding value is dropped, never summed",
      sum(p for c, _n, _ct, _ym, _y, _m, p in rows_c if c == "CDG") == 9300.0,
      sum(p for c, _n, _ct, _ym, _y, _m, p in rows_c if c == "CDG"))
try:
    LA.build(tmp_c, tmp_cdb)
    refused, msg = False, ""
except SystemExit as e:
    refused, msg = True, str(e)
check("a colliding workbook does NOT build by default", refused, msg[:50])
check("and the refusal says what to do about it",
      "--allow-collisions" in msg and "CDG" in msg, msg[:80])
rep_f = LA.build(tmp_c, tmp_cdb, allow_collisions=True)
check("it builds when the collision is accepted deliberately",
      rep_f["rows"] == 9, rep_f["rows"])

# --- 6. the store -----------------------------------------------------------------
rep2 = LA.build(tmp_x, tmp_db)
import duckdb
con = duckdb.connect(tmp_db, read_only=True)
n = con.execute("SELECT COUNT(*) FROM aci_monthly").fetchone()[0]
check("store built with every read row", n == len(rows), n)
cov = dict(con.execute("SELECT year, months_reported FROM aci_coverage "
                       "WHERE iata='AUS' ORDER BY year").fetchall())
check("coverage records months reported, not months in the file",
      cov.get(2024) == 2 and cov.get(2023) == 1, cov)
meta = dict(con.execute("SELECT key, value FROM _store_meta").fetchall())
check("the measure is stamped as throughput",
      "arrivals plus departures plus transit" in meta.get("measure", ""),
      meta.get("measure"))
check("the vintage is stamped", meta.get("source_file") == "aci_fixture.xlsx",
      meta.get("source_file"))
check("the blank rule is stamped in the store",
      "not zero" in meta.get("blanks", ""), meta.get("blanks"))
check("the grain warns against YTD and YE",
      "never" in meta.get("grain", "") and "YTD" in meta.get("grain", ""),
      meta.get("grain"))
con.close()

# --- 7. airport_profile can read what this produced ----------------------------
import airport_profile as AP

# The fixture holds five months, so every year in it is a part year. A part year
# must NOT reach a chart: ACI blanks mean no return was filed, not no traffic,
# so plotting eight months against twelve shows a collapse that never happened.
p = AP.pax_by_year("EDI", "GB", stores={"aci": tmp_db})
check("a part year is refused, not plotted short", p["series"] == [], p["series"])
check("and the refusal names the months actually reported",
      "reported" in " ".join(p["notes"]), p["notes"][:1])

# A complete twelve months, which is what 80% of the real store looks like.
full_x = os.path.join(tempfile.gettempdir(), "aci_full.xlsx")
full_db = os.path.join(tempfile.gettempdir(), "aci_full.duckdb")
import openpyxl
_wb = openpyxl.Workbook()
_ws = _wb.active
_ws.title = "1-Monthly"
_ws.append(["REG", "COUNTRY", "CITY", "CODE"]
           + ["%d-%02d" % (y, m) for y in (2023, 2024) for m in range(1, 13)])
_ws.append(["EUR", "United Kingdom", "EDINBURGH", "EDI"] + [1000] * 24)
_ws.append(["EUR", "France", "PARIS CDG", "CDG"] + [2000] * 24)
_ws.append(["NAM", "USA", "AUSTIN TX", "AUS"] + [3000] * 24)
# INV reports patchily, as small airports do, and must be told apart from EDI
_ws.append(["EUR", "United Kingdom", "INVERNESS", "INV"]
           + [400 if i % 4 else None for i in range(24)])
# enough distinct countries that the country column can be told apart from the
# six-value region column, which is the whole basis of the detection
for _i, (_rg, _ct, _cd) in enumerate([
        ("EUR", "Netherlands", "AMS"), ("EUR", "Germany", "FRA"),
        ("EUR", "Spain", "MAD"), ("EUR", "Italy", "FCO"),
        ("ASP", "Singapore", "SIN"), ("ASP", "Japan", "NRT"),
        ("LAC", "Brazil", "GRU"), ("AFR", "South Africa", "JNB")]):
    _ws.append([_rg, _ct, "%s CITY" % _cd, _cd] + [500 + _i] * 24)
_wb.save(full_x)
LA.build(full_x, full_db)

f = AP.pax_by_year("EDI", "GB", stores={"aci": full_db})
check("a complete year IS read", f["series"] == [(2023, 12000.0), (2024, 12000.0)],
      f["series"])
check("and it is labelled throughput, not onboard", f["kind"] == "throughput", f["kind"])
check("and attributed to ACI", f["label"] and "ACI" in f["label"], f["label"])
check("the patchy airport is refused and its months named",
      AP.pax_by_year("INV", "GB", stores={"aci": full_db})["series"] == [], "INV")
check("AUS is NOT served from ACI, because it is American",
      AP.pax_by_year("AUS", "US", stores={"aci": full_db, "t100": ""})["series"] == [])
# The store carries the country, so the deck need not be told one.
check("the country comes back from the store",
      AP.aci_country(full_db, "AUS") == "USA"
      and AP.aci_country(full_db, "EDI") == "United Kingdom",
      (AP.aci_country(full_db, "AUS"), AP.aci_country(full_db, "EDI")))
check("and an airport not in the store gives no country, not a guess",
      AP.aci_country(full_db, "ZZZ") == "", AP.aci_country(full_db, "ZZZ"))
# The trap the explicit reader exists for: aci_monthly has BOTH an `airport`
# column holding the name and an `iata` column holding the code.
check("the reader matches on the code column, not the name column",
      AP.read_aci(full_db, "CDG")[0] == [(2023, 24000.0), (2024, 24000.0)],
      AP.read_aci(full_db, "CDG")[0])
check("and an airport name passed as a code finds nothing, rather than a view",
      AP.read_aci(full_db, "PARIS CDG")[0] == [])
for _f in (full_x, full_db):
    if os.path.exists(_f):
        os.remove(_f)

for f in (tmp_x, tmp_db, tmp_c, tmp_cdb):
    if os.path.exists(f):
        os.remove(f)
print("\n%d checks, %d failed" % (CHECKS, len(FAIL)))
if FAIL:
    print("FAILED: %s" % ", ".join(FAIL))
sys.exit(1 if FAIL else 0)
