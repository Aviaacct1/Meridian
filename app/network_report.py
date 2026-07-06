#!/usr/bin/env python3
"""
Avia Solutions - Network P&L report (the "Airline P&L 2" layer).
Renders a network of routes that share a fleet into one Network P&L workbook:
per-route annual revenue / cost / profit / margin / frames, network totals, and the
shared-fleet aircraft count. Carries the standard disclaimer.

  from aircraft_economics import AnnualRoutePnL, network_pnl
  from network_report import write_network_workbook
  net = network_pnl([AnnualRoutePnL(rp1, 7), AnnualRoutePnL(rp2, 4, 30), ...])
  write_network_workbook(net, "Network_PnL.xlsx", title="Genoa base - summer network")
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = '1F3864'
GREEN = '006100'
AR = lambda **k: Font(name='Arial', **k)
HF = AR(bold=True, color='FFFFFF', size=10)
HFILL = PatternFill('solid', fgColor=NAVY)
DERFILL = PatternFill('solid', fgColor='E2EFDA')
THIN = Side('thin', color='B4B4B4')
BD = Border(THIN, THIN, THIN, THIN)
NOTE = AR(size=9, italic=True, color='666666')

try:
    from aircraft_economics import DISCLAIMER_FULL
except Exception:
    DISCLAIMER_FULL = ("Indicative, for directional guidance only. Built on generic published "
                       "assumptions, not any airline's actual costs.")

COLS = [('Route', 16), ('Aircraft', 10), ('Programme', 18), ('Annual pax', 12),
        ('Annual revenue', 15), ('Annual cost', 14), ('Annual profit', 14),
        ('Margin', 9), ('Frames', 9)]


def write_network_workbook(net, out_path, title=None, disclaimer=DISCLAIMER_FULL):
    """net = the dict returned by aircraft_economics.network_pnl()."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Network P&L'
    ws.sheet_properties.tabColor = NAVY

    ws.cell(1, 1, title or 'Network P&L').font = AR(bold=True, color=NAVY, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))

    hr = 3
    for j, (h, w) in enumerate(COLS, 1):
        c = ws.cell(hr, j, h); c.font = HF; c.fill = HFILL; c.border = BD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[chr(64 + j)].width = w

    r = hr + 1
    for (origin, dest, aircraft, c) in net.get('routes', []):
        prog = (f"{c.get('annual_turnarounds', 0):.0f} turns/yr")
        vals = [f"{origin}-{dest}", aircraft, prog, c.get('annual_pax', 0),
                c.get('annual_gross_rev', 0), c.get('annual_total_cost', 0),
                c.get('annual_profit', 0), c.get('margin', 0),
                c.get('aircraft_required_fractional', 0)]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v); cell.border = BD; cell.font = AR(size=10)
            if j in (4, 5, 6, 7): cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
            if j == 8: cell.number_format = '0.0%'; cell.alignment = Alignment(horizontal='right')
            if j == 9: cell.number_format = '0.00'; cell.alignment = Alignment(horizontal='right')
        r += 1

    # Network totals
    tot = [('Network total', '', '', net.get('annual_pax', 0), net.get('annual_gross_rev', 0),
            net.get('annual_total_cost', 0), net.get('annual_profit', 0), net.get('margin', 0),
            net.get('aircraft_required_fractional', 0))]
    for vals in tot:
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v); cell.border = BD; cell.fill = DERFILL
            cell.font = AR(bold=True, size=10, color=(GREEN if j == 7 else NAVY))
            if j in (4, 5, 6, 7): cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right')
            if j == 8: cell.number_format = '0.0%'; cell.alignment = Alignment(horizontal='right')
            if j == 9: cell.number_format = '0.00'; cell.alignment = Alignment(horizontal='right')
        r += 1

    # Fleet line
    ws.cell(r + 1, 1, f"Shared fleet required: {net.get('aircraft_required', 0)} aircraft "
            f"({net.get('aircraft_required_fractional', 0):.2f} frames of flying; "
            f"the rest is spare capacity across the network).").font = AR(size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=len(COLS))

    dc = ws.cell(r + 3, 1, disclaimer); dc.font = NOTE; dc.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r + 3, start_column=1, end_row=r + 3, end_column=len(COLS))
    ws.row_dimensions[r + 3].height = 80

    wb.properties.creator = 'Avia Solutions'
    wb.properties.lastModifiedBy = 'Avia Solutions'
    wb.properties.title = title or 'Network P&L'
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    wb.save(out_path)
    return out_path
