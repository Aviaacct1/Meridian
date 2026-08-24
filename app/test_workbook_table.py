#!/usr/bin/env python3
"""Offline end-to-end check of the workbook's Traffic forecast table (18 August
2026): builds a real workbook from a synthetic payload and reads the table back.
The three invariants a network planner checks in the room, plus the decomposition:

  1. the rows SUM to the grand total (the carried allocation),
  2. each row MULTIPLIES through (stimulated x capture = forecast),
  3. base year / growth / grown are decomposed with year-labelled headers,
  4. a steady-state run prints base = grown with growth 0, as before.

    py -3.12 test_workbook_table.py

Every figure here is a TEST FIXTURE.

Avia Solutions Limited. All rights reserved.
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import cortex_workbook as CWB
import openpyxl

FAIL = []
CHECKS = 0


def check(name, cond, detail=""):
    global CHECKS
    CHECKS += 1
    print("%-62s %s %s" % (name, "PASS" if cond else "FAIL", detail))
    if not cond:
        FAIL.append(name)


def _fc(grown=True):
    sch = {"forecast_year": 2028, "base_year": 2025, "growth_rate": 0.0223,
           "growth_years": 3, "growth_basis": "measured pre-COVID trend, tapered",
           "legs": []} if grown else {"legs": []}
    _freq, _weeks, _total = 4, 52, 55692
    # 23 August 2026 (Jol Kingham: "tab Cover has a different PDEW to Connecting
    # feed and to forecast"). The fixture used to hardcode pdew_total at an
    # arbitrary 152.6, which could never have caught this bug - it never agreed
    # with the Forecast tab's own PTEW because nothing tied the two together. The
    # real defect was upstream, in cortex_app.py's own pdew_total calculation
    # (divided by a flat 365 regardless of frequency; fixed the same day to
    # freq x weeks, matching the formula this sheet's Forecast tab already used).
    # Computing it here the SAME way the fixed engine now does is what lets
    # "cross-tab PTEW consistency" below actually test something.
    return {
        "ok": True, "origin": {"city": "San Jose", "iata": "SJC", "country": "US"},
        "dest": {"city": "Taipei", "iata": "TPE", "country": "TW"},
        "distance_nm": 5637, "block_min": 825, "week": "2026-05-25", "year": 2025,
        "carrier_type": "FSC", "schedule": sch,
        "demand": {"natural": 203400, "captured": 45400, "qsi_share": 0.251,
                   "stimulation": 1.15, "feed_behind": 7400, "feed_beyond": 23200,
                   "feed_behind_base": 198200, "feed_beyond_base": 768800,
                   "p2p_carried": 33300, "connecting_carried": 22392,
                   "total": _total, "total_demand": 76000,
                   "pdew_total": round(_total / (_freq * _weeks), 1),
                   "beyond_pdew": [], "behind_pdew": [],
                   "avg_fare_band": {"label": "950-1000"}},
        "capacity": {"carried": _total, "load": 0.875, "freq": _freq, "aircraft": "A359",
                     "annual_capacity": 63648, "recommendation": ""},
        # Deliberately chosen (24 August 2026, sub-row PTEW footing test) so that
        # round(direct/(freq*weeks)) + round(no_direct/(freq*weeks)) does NOT equal
        # round(leg/(freq*weeks)) if each is rounded independently - direct=525/4890 sums
        # exactly to the behind leg's 5415, but rounds independently to 3+24=27, not the
        # leg's own 26; beyond's 50/16927 sums exactly to 16977 but rounds to 0+81=81, not
        # 82. Proves the remainder fix actually does something on this fixture, not just
        # on the real CI/JX/BR files where it happened to already foot.
        "competition_split": {
            "behind": {"totals": {"direct": {"base": 0, "forecast": 525},
                                   "no_direct": {"base": 0, "forecast": 4890}}},
            "beyond": {"totals": {"direct": {"base": 0, "forecast": 50},
                                   "no_direct": {"base": 0, "forecast": 16927}}},
        },
        "catchment": {"observed_share": {}, "names": {}, "home": "SJC"},
        "economics": {"raw": {}, "econ_fare": 975, "seats": 306},
        "season": {"mode": "annual", "share": 1.0, "weeks": _weeks},
    }


def _table(fc, tmp, name):
    out = os.path.join(tmp, name)
    CWB.build_workbook(out, fc, {"airline_name": "CI", "analyst": "Avia Solutions",
                                 "date": "18 Aug 2026", "plan_lf": 0.875,
                                 "capture_basis": "measured", "econ_fare": 975})
    wb = openpyxl.load_workbook(out, data_only=True)
    # 22 August 2026 (EW/2-way pair): "Forecast" no longer exists as a bare sheet
    # name; "Forecast EW" carries the identical figures the old single tab did
    # (native each-way, un-multiplied), so every numeric assertion below is
    # unchanged - only the sheet name moved.
    fs = wb["Forecast EW"]
    return {str(r[0].value): [c.value for c in r]
            for r in fs.iter_rows(min_row=4) if r[0].value}


def main():
    with tempfile.TemporaryDirectory() as tmp:
        rows = _table(_fc(grown=True), tmp, "grown.xlsx")
        hdr, p2p, gt = rows["Market"], rows["Total point to point"], rows["GRAND TOTAL"]
        legs = [rows[k][7] for k in rows
                if k.startswith("Total ") and "point" not in k]
        check("year-labelled headers", "2025" in str(hdr[1]) and "2028" in str(hdr[3]))
        check("base decomposed from grown", abs(p2p[1] - 190.4) < 0.5, p2p[1])
        # 20 August 2026 (John, consistency with the deck's CAGR fix): the column now
        # shows the per-annum rate (the fixture's growth_rate, 0.0223), not the
        # cumulative (0.0684) that rate compounds to over the fixture's 3 years.
        check("CAGR printed, not the cumulative", abs(p2p[2] - 0.0223) < 0.0003, p2p[2])
        check("grown column carries the grown market", abs(p2p[3] - 203.4) < 0.1)
        check("row multiplies through (effective capture)",
              abs(p2p[5] * p2p[6] - p2p[7]) < 0.6)
        check("rows sum to the grand total",
              abs(p2p[7] + sum(legs) - gt[7]) < 0.15,
              "%s + %s = %s" % (p2p[7], legs, gt[7]))
        check("grand total base column summed", abs(gt[1] - 1095.5) < 0.5, gt[1])

        # GRAND TOTAL PTEW FOOTS TO THE DISPLAYED PARTS (24 August 2026, Jol Kingham:
        # "the two forecast tabs PTEW sum of parts (267) does not match the total
        # (268)"). Not a basis bug - p2p+behind+beyond equals tot exactly by
        # carried_split's own invariant, but each row's PTEW was independently rounded
        # to a whole number, so the parts as displayed did not sum to the total as
        # displayed. The GRAND TOTAL row's PTEW is now the sum of the three displayed
        # PTEW cells, so this must hold exactly, not just within tolerance.
        legs_ptew = [rows[k][8] for k in rows if k.startswith("Total ") and "point" not in k]
        check("GRAND TOTAL PTEW foots exactly to the displayed parts",
              p2p[8] + sum(legs_ptew) == gt[8],
              f"{p2p[8]} + {legs_ptew} = {p2p[8] + sum(legs_ptew)}, GRAND TOTAL={gt[8]}")

        # CROSS-TAB PTEW CONSISTENCY (23 August 2026, Jol Kingham's exact complaint:
        # Cover's PDEW/PTEW did not match Connecting feed's or Forecast's). Cover
        # prints dem["pdew_total"] verbatim, straight from the engine payload;
        # Forecast EW computes its own grand-total PTEW independently, from the
        # carried total and the route's freq x weeks. The two must read the same
        # figure for the same route, or a client sees exactly what Jol saw.
        wb_grown = openpyxl.load_workbook(os.path.join(tmp, "grown.xlsx"), data_only=True)
        cov = wb_grown["Cover"]
        # Label updated 24 Aug 2026 (Jol Kingham): Cover now says "Passenger Trip Each
        # Way (PTEW)" in full, spelling the term out on first use for the whole workbook.
        cov_pdew = next(c[1].value for c in cov.iter_rows(min_row=1)
                         if c[0].value == "Passenger Trip Each Way (PTEW)")
        # Tightened to an exact equality (24 August 2026, Jol Kingham's "148 v 147"
        # question): Cover's PTEW now comes from the identical footed ptew() calculation
        # the Forecast tab's GRAND TOTAL row uses, not a separately-sourced figure that
        # could only ever be close. A tolerance check would hide a real future regression.
        check("Cover PTEW matches Forecast tab's grand-total PTEW exactly",
              cov_pdew == gt[8], f"Cover={cov_pdew}, Forecast grand total={gt[8]}")

        # SUB-ROW PTEW FOOTING (24 August 2026, Jol Kingham's annotated screenshot: his own
        # sum of the two "O&Ds with/without direct competition" rows plus point to point
        # read 267 against a printed GRAND TOTAL of 268 - the footing fix above only forces
        # the three LEG rows to sum to the grand total; the two sub-rows one level down were
        # still independently rounded and did not reliably sum to their own parent leg row.
        # "with direct competition" keeps its own rounded figure; "without" is now the
        # remainder against the leg row, so the two must sum to their parent exactly, for
        # both legs. Read from the raw sheet, not the _table() dict, because both legs'
        # sub-rows share identical labels and would collide as dict keys.
        fs_ws = wb_grown["Forecast EW"]
        _leg_ptew = None
        _sub_ptews = []
        for r in fs_ws.iter_rows(min_row=5, max_row=12):
            label = str(r[0].value or "")
            if label.startswith("Total connecting"):
                if _leg_ptew is not None:
                    check(f"sub-rows foot exactly to their leg row ({_leg_label})",
                          sum(_sub_ptews) == _leg_ptew,
                          f"{_sub_ptews} = {sum(_sub_ptews)}, leg={_leg_ptew}")
                _leg_ptew = r[8].value; _leg_label = label; _sub_ptews = []
            elif label.strip().startswith("O&Ds"):
                _sub_ptews.append(r[8].value)
        if _leg_ptew is not None:
            check(f"sub-rows foot exactly to their leg row ({_leg_label})",
                  sum(_sub_ptews) == _leg_ptew,
                  f"{_sub_ptews} = {sum(_sub_ptews)}, leg={_leg_ptew}")

        rows = _table(_fc(grown=False), tmp, "steady.xlsx")
        p2p = rows["Total point to point"]
        check("steady state: base equals grown", p2p[1] == p2p[3])
        check("steady state: growth prints zero", (p2p[2] or 0) == 0)

        # the Connecting feed sheet (19 August 2026, Jol's review): the top-N city
        # rows plus an All-other row must TOTAL to the same carried leg the Cover
        # prints, and every figure column names the service year
        fc = _fc(grown=True)
        fc["demand"]["behind_pdew"] = [
            {"code": "SEA", "name": "Seattle", "country": "US", "base": 40000,
             "share": 0.06, "forecast": 5000, "pdew": 13.7}]
        fc["demand"]["beyond_pdew"] = [
            {"code": "MNL", "name": "Manila", "country": "PH", "base": 80000,
             "share": 0.03, "forecast": 9000, "pdew": 24.7},
            {"code": "BKK", "name": "Bangkok", "country": "TH", "base": 60000,
             "share": 0.03, "forecast": 6000, "pdew": 16.4}]
        # 20 August 2026 (John, checking the EVA pack): the demand column completes to
        # the full uncapped market the same way the forecast column completes to the
        # carried leg, using feed_behind_base/feed_beyond_base - the same source the
        # shown cities' own "base" figures are drawn from, so additive with them.
        fc["demand"]["feed_behind_base"] = 55000
        fc["demand"]["feed_beyond_base"] = 200000
        out = os.path.join(tmp, "feed.xlsx")
        CWB.build_workbook(out, fc, {"airline_name": "CI", "analyst": "A", "date": "d",
                                     "plan_lf": 0.875, "capture_basis": "m",
                                     "econ_fare": 975})
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Connecting feed EW"]  # 22 August 2026: same figures as the old bare tab
        txt = [[c.value for c in row] for row in ws.iter_rows(min_row=1, max_row=30)]
        flat = ["|".join(str(v) for v in row if v is not None) for row in txt]
        check("feed headers carry the service year",
              any("Market demand 2028" in s for s in flat)
              and any("forecast 2028" in s.lower() for s in flat))
        check("All-other rows drawn on both legs",
              sum(1 for s in flat if "All other connecting markets" in s) == 2)
        # carried_split legs from the fixture: behind 7400, beyond 23200 scaled to
        # connecting_carried 22392 -> behind 5414, beyond 16978
        totals = [row for row in txt if row[0] == "Total"]
        t_beh, t_bey = totals[0][6], totals[1][6]
        check("behind total equals the carried leg", abs(t_beh - 5414) <= 1, t_beh)
        check("beyond total equals the carried leg", abs(t_bey - 16978) <= 1, t_bey)
        t_beh_dem, t_bey_dem = totals[0][4], totals[1][4]
        check("behind demand completes to the full market", t_beh_dem == 55000, t_beh_dem)
        check("beyond demand completes to the full market", t_bey_dem == 200000, t_bey_dem)
        check("note states both columns now reconcile",
              any("foots exactly to the Total row" in s for s in flat))
        # CONNECTING FEED TOTAL PTEW (24 August 2026, Jol Kingham: "the PTEWs in tabs
        # 'Connecting feed EW' and 'Connecting feed 2-way' ... does not match Cover tab
        # row 24, nor the two Forecast tabs"). Two defects: the Total row's PTEW cell
        # was blank (a reader summing the city rows by hand was building their own,
        # rounding-drifted total), and the All-other row's own PTEW used a flat
        # weeks x 7.0 denominator (assumed daily service) instead of the route's real
        # freq x weeks - the same bug class as the six other PTEW instances fixed this
        # week, just not caught until now because this tab had no visible total to
        # check it against. Both fixed: the Total row now prints the leg's carried
        # total divided by freq x weeks directly, the same figure and formula Cover
        # and the Forecast tabs use for the same leg.
        freq, weeks = 4, 52
        exp_beh_ptew = round(t_beh / (freq * weeks), 1)
        exp_bey_ptew = round(t_bey / (freq * weeks), 1)
        # Tightened to exact equality (24 August 2026, Jol Kingham: "Connecting feed 2-way"
        # H22+H43 = 147.4 v "Forecast 2-way"/Cover showing 148 - the two tabs rounded the
        # same figure to different precisions). Both now use the identical freq x weeks,
        # 1dp calculation, so this must hold exactly.
        check("Connecting feed Total PTEW matches Forecast basis exactly (behind)",
              totals[0][7] == exp_beh_ptew, f"{totals[0][7]} vs {exp_beh_ptew}")
        check("Connecting feed Total PTEW matches Forecast basis exactly (beyond)",
              totals[1][7] == exp_bey_ptew, f"{totals[1][7]} vs {exp_bey_ptew}")

        # JOL'S EXACT SCENARIO: the sum of the two Connecting feed Total PTEWs (his own
        # "147.4") must equal the sum of the Forecast tab's own two leg PTEW rows, not
        # merely round to the same whole number.
        fs2 = wb["Forecast EW"]
        leg_ptews = [r[8].value for r in fs2.iter_rows(min_row=5, max_row=12)
                     if str(r[0].value or "").startswith("Total connecting")]
        check("Connecting feed legs sum to the same figure as Forecast's own leg PTEWs",
              (totals[0][7] + totals[1][7]) == sum(leg_ptews),
              f"feed: {totals[0][7]}+{totals[1][7]}={totals[0][7]+totals[1][7]}, "
              f"forecast legs: {leg_ptews}={sum(leg_ptews)}")

        # CITY ROWS + ALL-OTHER FOOT EXACTLY TO THE TOTAL (24 August 2026, Jol Kingham's
        # annotated screenshot: his own sum of the 15 SJC-behind city rows plus All-other
        # read 35.5 against a printed Total of 35.6). All-other's PTEW is now the remainder
        # against the Total, the same design already used for its demand and forecast
        # columns, so the full column - every listed city plus All-other - must sum to the
        # Total exactly, not just within rounding tolerance.
        def _leg_rows(start_row):
            out = []
            for r in ws.iter_rows(min_row=start_row, max_row=start_row + 20):
                if r[0].value == "Total":
                    break
                pv = r[7].value
                if pv is not None:
                    out.append(pv)
            return out
        _tpe_hdr_row = next(r[0].row for r in ws.iter_rows(min_row=1)
                             if r[0].value and "Connecting at" in str(r[0].value) and "TPE" in str(r[0].value))
        beh_city_ptews = _leg_rows(6)
        bey_city_ptews = _leg_rows(_tpe_hdr_row + 2)
        check("SJC behind: city rows + All-other foot exactly to the Total",
              sum(beh_city_ptews) == totals[0][7],
              f"{beh_city_ptews} = {sum(beh_city_ptews)}, Total={totals[0][7]}")
        check("TPE beyond: city rows + All-other foot exactly to the Total",
              sum(bey_city_ptews) == totals[1][7],
              f"{bey_city_ptews} = {sum(bey_city_ptews)}, Total={totals[1][7]}")

        # the Departure curve sheet (19 August 2026): raw curve + the dashboard's own
        # carried transform + a native chart; must reconcile with the headline at the
        # chosen departure and never exist without an optimiser curve
        fc = _fc(grown=True)
        fc["capacity"]["plan_cap"] = 0.875
        fc["schedule"].update({
            "outbound": {"sector": "SJC-TPE", "dep": "20:59", "arr": "02:44+2"},
            "inbound": {"sector": "TPE-SJC", "dep": "20:14", "arr": "17:59"},
            "indicative": False, "basis": "optimised",
            "optimised": {"unrestricted_dep": "00:30", "restricted": ["21:00-06:00"],
                          "curve": [
                              {"dep": 0, "hhmm": "00:00", "total": 30000, "beyond": 26000,
                               "behind": 4000, "permitted": False},
                              {"dep": 720, "hhmm": "12:00", "total": 12000, "beyond": 9000,
                               "behind": 3000, "permitted": True},
                              {"dep": 1259, "hhmm": "20:59", "total": 15000, "beyond": 11500,
                               "behind": 3500, "permitted": True},
                              {"dep": 1410, "hhmm": "23:30", "total": 26000, "beyond": 22500,
                               "behind": 3500, "permitted": False}]}})
        out = os.path.join(tmp, "curve.xlsx")
        CWB.build_workbook(out, fc, {"airline_name": "BR", "analyst": "A", "date": "d",
                                     "plan_lf": 0.875, "capture_basis": "m",
                                     "econ_fare": 975})
        wb = openpyxl.load_workbook(out)
        # 22 August 2026: the old bare "Departure curve" tab was always built two-way
        # (every figure already carried a x2), so "Departure curve 2-way" is the exact
        # same figures under the new name; the *2 in the check below is unchanged.
        check("departure curve sheet exists", "Departure curve 2-way" in wb.sheetnames
              and "Departure curve EW" in wb.sheetnames)
        ws = wb["Departure curve 2-way"]
        cr = {str(r[0].value): [c.value for c in r] for r in ws.iter_rows(min_row=5, max_row=8)}
        check("curve reconciles with the headline at the chosen departure",
              abs(cr["20:59"][3] - fc["demand"]["connecting_carried"] * 2) < 3
              and abs(cr["20:59"][6] - fc["demand"]["total"] * 2) < 3)
        check("permitted flags carried", cr["00:00"][1] == "no" and cr["12:00"][1] == "yes")
        check("native chart embedded", len(ws._charts) == 1)
        nt = " ".join(str(c.value) for row in ws.iter_rows(min_row=9) for c in row if c.value)
        check("curve note states ceiling and source",
              "aircraft ceiling" in nt and "Meridian analysis" in nt)
        wb2 = openpyxl.load_workbook(os.path.join(tmp, "grown.xlsx"))
        check("no curve, no sheet, never fabricated",
              "Departure curve EW" not in wb2.sheetnames
              and "Departure curve 2-way" not in wb2.sheetnames)

        # THE CURVE PICTURE (24 August 2026, John Carter: running a batch of EVA/CI/JX
        # forecasts and wanting the departure curve as a picture alongside each workbook,
        # not a chart he has to screenshot). Same "never fabricated" discipline as the
        # Excel sheet - a picture must exist when there is a curve to draw, and must not
        # exist when there is not, checked against the SAME two fixtures used above.
        png_path = os.path.join(tmp, "curve_curve.png")
        check("curve picture generated alongside the workbook when a curve exists",
              os.path.exists(png_path) and os.path.getsize(png_path) > 1000)
        no_png_path = os.path.join(tmp, "grown_curve.png")
        check("no curve picture when there is no optimiser curve, never fabricated",
              not os.path.exists(no_png_path))
    print("\n%d checks, %d failed%s" % (CHECKS, len(FAIL),
          ": " + ", ".join(FAIL) if FAIL else ""))
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
