#!/usr/bin/env python3
r"""Avia Solutions - what the economics module holds, what the schedule flies, and the gap.

    py -3.12 aircraft_gap_report.py [output.xlsx]

WHY. aircraft_economics.AIRCRAFT holds 25 types. The 2025 schedule flies 136 equipment codes above
5,000 sectors. Everything not held is invisible to the aircraft selection and the route P&L, and on
10 August 2026 that bit for real: EVA operates the 787-10 on sectors of SJC-TPE length and the
optimiser could not offer it, because there is no entry to cost it against.

This writes the gap out for filling in by hand: what is held, what needs adding, how every OAG code
maps to a type, and what has been left out on purpose. The OAG columns are MEASURED (2025 schedule,
nonstop passenger services) and are there as evidence beside each row, not as the answer: a seat
count is a fleet-wide median across carriers and configurations, and the economics wants a
representative configuration, which is a judgement.

FAMILY CODES ARE FLAGGED, NOT RESOLVED. OAG carries generic codes ("737", "777", "A330") that cover
several variants. Each is mapped to its commonest variant and marked so on the map sheet, because a
mapping nobody can see is how a 777-200 ends up costed as a 777-300ER.

Source for every OAG column: OAG schedules 2025, service_type J, nonstop.
Avia Solutions Limited. All rights reserved.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

# ---------------------------------------------------------------- the classification, stated once
# code -> (economics key, status). Status: HELD the key already exists; ADD a new entry is needed;
# FAMILY a generic OAG code resolved to its commonest variant, for review; EXCLUDE not a commercial
# air service type for Avia's purposes.
MAP = {
    # --- already held ---
    "320": ("A320", "HELD"), "32A": ("A320", "HELD"), "32S": ("A320", "FAMILY"),
    "32N": ("A20N", "HELD"), "321": ("A321", "HELD"), "32B": ("A321", "HELD"),
    "32Q": ("A21N", "HELD"), "319": ("A319", "HELD"),
    "738": ("B738", "HELD"), "73H": ("B738", "HELD"), "7S8": ("B738", "HELD"),
    "7M8": ("B38M", "HELD"), "752": ("B752", "HELD"), "75W": ("B752", "HELD"),
    "AT7": ("ATR72", "HELD"), "ATR": ("ATR72", "FAMILY"),
    "DH4": ("DH8D", "HELD"), "DH8": ("DH8D", "FAMILY"),
    "CR9": ("CRJ900", "HELD"), "E70": ("E170", "HELD"), "E90": ("E190", "HELD"),
    "E95": ("E195", "HELD"), "SF3": ("SF34", "HELD"),
    "763": ("B763", "HELD"), "76W": ("B763", "HELD"), "767": ("B763", "FAMILY"),
    "333": ("A333", "HELD"), "330": ("A333", "FAMILY"), "339": ("A339", "HELD"),
    "359": ("A359", "HELD"), "350": ("A359", "FAMILY"),
    "788": ("B788", "HELD"), "789": ("B789", "HELD"), "787": ("B789", "FAMILY"),
    "77W": ("B77W", "HELD"),
    "919": ("C919", "HELD"),
    "C27": ("C909", "HELD"), "C09": ("C909", "HELD"), "909": ("C909", "HELD"),
    # --- new entries needed ---
    "221": ("A221", "ADD"), "223": ("A223", "ADD"), "31N": ("A31N", "ADD"),
    "332": ("A332", "ADD"), "343": ("A343", "ADD"), "351": ("A35K", "ADD"),
    "388": ("A388", "ADD"),
    "717": ("B717", "ADD"), "733": ("B733", "ADD"), "734": ("B734", "ADD"),
    "735": ("B735", "ADD"), "73E": ("B735", "ADD"), "736": ("B736", "ADD"),
    "73W": ("B737", "ADD"), "73G": ("B737", "ADD"),
    "739": ("B739", "ADD"), "73J": ("B739", "ADD"), "7M9": ("B39M", "ADD"),
    "74H": ("B748", "ADD"),
    "772": ("B772", "ADD"), "777": ("B772", "FAMILY"), "77L": ("B77L", "ADD"),
    "773": ("B773", "ADD"), "781": ("B781", "ADD"),
    "753": ("B753", "ADD"), "764": ("B764", "ADD"),
    "E75": ("E175", "ADD"), "E7W": ("E175", "ADD"),
    "290": ("E290", "ADD"), "295": ("E295", "ADD"),
    "ER4": ("ERJ145", "ADD"), "ERJ": ("ERJ145", "FAMILY"), "ER3": ("ERJ135", "ADD"),
    "CR7": ("CRJ700", "ADD"), "CR2": ("CRJ200", "ADD"), "CRJ": ("CRJ200", "FAMILY"),
    "CR5": ("CRJ550", "ADD"), "CRK": ("CRJ1000", "ADD"),
    "AT4": ("AT42", "ADD"),
    "DH1": ("DH8A", "ADD"), "DH2": ("DH8B", "ADD"), "DH3": ("DH8C", "ADD"),
    "DHT": ("DHC6", "ADD"),
    "SU9": ("SU95", "ADD"), "100": ("F100", "ADD"), "M82": ("MD82", "ADD"),
    "AN4": ("AN24", "ADD"),
    "BEH": ("B190", "ADD"), "BE1": ("B190", "ADD"),
    "EM2": ("E120", "ADD"), "SWM": ("SW4", "ADD"),
    # --- deliberately out ---
    "CNC": (None, "EXCLUDE"), "CNA": (None, "EXCLUDE"), "CN1": (None, "EXCLUDE"),
    "CN2": (None, "EXCLUDE"), "CNF": (None, "EXCLUDE"), "PA2": (None, "EXCLUDE"),
    "PL2": (None, "EXCLUDE"), "T12": (None, "EXCLUDE"), "BNI": (None, "EXCLUDE"),
    "DHL": (None, "EXCLUDE"),
    # --- ambiguous, named rather than guessed ---
    "73M": (None, "REVIEW"), "73L": (None, "REVIEW"),
}

EXCLUDE_REASON = ("light aircraft or air taxi, not a scheduled commercial type Avia forecasts; "
                  "no economics entry needed")
REVIEW_REASON = ("mixed-configuration code covering several 737 variants; the seat count does not "
                 "identify one, so it is left for a decision rather than mapped")

# The 13 variables the economics module holds, in the order they will be written.
VARS = ["econ_seats", "bus_seats", "range_km", "category", "mtow_kg", "fuel_burn_kg_per_bh",
        "maint_per_bh", "crew_per_bh", "ownership_per_bh", "annual_util_bh", "price_usd",
        "cargo_cap_kg", "src"]


def census(min_sectors=5000, period="2025-%"):
    """Measured equipment census from OAG: sectors, median seats, median premium seats, sector
    lengths and how many carriers fly each code."""
    import duckdb
    import capacity_frame as CF
    db = CF._oag()
    if not db:
        sys.exit("no OAG store found. Set AVIA_OAG_DUCKDB or AVIA_LOCAL_CACHE.")
    con = duckdb.connect(db, read_only=True)
    con.execute("SET memory_limit='3GB'; SET threads=3; SET enable_progress_bar=false")
    try:
        rows = con.execute("""
          SELECT aircraft_code, any_value(aircraft_name) nm, count(*) sectors,
                 median(try_cast(seats_total AS DOUBLE)) seats,
                 median(try_cast(business_seats AS DOUBLE) + try_cast(first_seats AS DOUBLE)) prem,
                 median(try_cast(gcd_km AS DOUBLE)) gcd, max(try_cast(gcd_km AS DOUBLE)) gcd_max,
                 count(DISTINCT carrier) carriers
          FROM oag WHERE service_type='J' AND week LIKE ? AND try_cast(stops AS INT)=0
          GROUP BY 1 HAVING count(*) >= ? ORDER BY sectors DESC
        """, [period, min_sectors]).fetchall()
    finally:
        con.close()
    return [dict(code=r[0], name=(r[1] or "").strip(), sectors=int(r[2]), seats=int(r[3] or 0),
                 prem=int(r[4] or 0), gcd=int(r[5] or 0), gcd_max=int(r[6] or 0),
                 carriers=int(r[7])) for r in rows]


def build(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from aircraft_economics import AIRCRAFT

    cen = {r["code"]: r for r in census()}
    # aggregate the OAG evidence up to the economics key
    by_key = {}
    for code, r in cen.items():
        key, status = MAP.get(code, (None, "UNMAPPED"))
        if not key:
            continue
        b = by_key.setdefault(key, {"sectors": 0, "codes": [], "carriers": 0,
                                    "seats": [], "prem": [], "gcd": [], "gcd_max": 0})
        b["sectors"] += r["sectors"]; b["codes"].append(r["code"])
        b["carriers"] = max(b["carriers"], r["carriers"])
        b["seats"].append((r["sectors"], r["seats"])); b["prem"].append((r["sectors"], r["prem"]))
        b["gcd"].append((r["sectors"], r["gcd"])); b["gcd_max"] = max(b["gcd_max"], r["gcd_max"])
    def wavg(pairs):
        tot = sum(w for w, _ in pairs) or 1
        return round(sum(w * v for w, v in pairs) / tot)
    for k, b in by_key.items():
        b["seats"] = wavg(b["seats"]); b["prem"] = wavg(b["prem"]); b["gcd"] = wavg(b["gcd"])

    ARIAL = "Arial"
    hdr = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F3864")
    body = Font(name=ARIAL, size=10)
    blue = Font(name=ARIAL, size=10, color="0000FF")
    note = Font(name=ARIAL, size=9, italic=True, color="595959")
    title = Font(name=ARIAL, bold=True, size=12)
    fillin = PatternFill("solid", fgColor="FFFF00")
    thin = Border(bottom=Side(style="thin", color="BFBFBF"))

    wb = Workbook()

    def sheet(name, headers, widths):
        ws = wb.create_sheet(name)
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h); c.font = hdr; c.fill = hfill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        ws.freeze_panes = "B2"; ws.row_dimensions[1].height = 30
        return ws

    # ---------------------------------------------------------------- legend
    ws = wb.active; ws.title = "Legend"
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 104
    rows = [
        ("Aircraft economics: what we hold and what is missing", None),
        ("", None),
        ("Purpose", "aircraft_economics.AIRCRAFT holds 25 types. The 2025 OAG schedule flies 136 "
                    "equipment codes above 5,000 sectors. A type with no entry cannot be costed and "
                    "cannot be offered to an airline: EVA's 787-10 is one such, on sectors of "
                    "SJC-TPE length."),
        ("What to fill in", "The YELLOW cells on 'To add'. Blue text marks a hardcoded input. The "
                            "OAG columns are measured evidence beside each row, not the answer: a "
                            "seat count there is a fleet-wide median across every carrier and "
                            "configuration, and the economics wants one representative cabin."),
        ("Held", "The 25 types already in the module, all 13 variables, with the OAG-measured "
                 "configuration beside them so a held seat count that no longer matches the fleet "
                 "is visible."),
        ("To add", "43 types to create. Sorted by 2025 sectors flown, so the top of the sheet is "
                   "where the work pays."),
        ("Code map", "Every OAG equipment code above 5,000 sectors and the type it resolves to. "
                     "FAMILY marks a generic OAG code resolved to its commonest variant, which is a "
                     "judgement and is shown rather than hidden. REVIEW marks a code that could not "
                     "be resolved honestly."),
        ("Excluded", "Codes deliberately left out, with the reason."),
        ("", None),
        ("Units", "seats: count, one cabin each. range_km: km. mtow_kg: kg. fuel_burn_kg_per_bh: kg "
                  "per block hour. maint_per_bh, crew_per_bh, ownership_per_bh: USD per block hour. "
                  "annual_util_bh: block hours a year. price_usd: USD. cargo_cap_kg: kg."),
        ("category", "Narrowbody, Widebody or Regional. aircraft_select uses it to keep a widebody "
                     "off a sector a narrowbody can fly."),
        ("src", "Where the numbers came from. Required, not optional: every existing entry carries "
                "one and a row without it cannot be defended in a report."),
        ("", None),
        ("Source", "OAG schedules 2025, service_type J, nonstop. Held values from "
                   "app/aircraft_economics.py. Generated by app/aircraft_gap_report.py."),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=a).font = title if i == 1 else Font(name=ARIAL, bold=True, size=10)
        if b:
            c = ws.cell(row=i, column=2, value=b); c.font = body
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = 42

    # ---------------------------------------------------------------- held
    heads = ["key", "OAG codes", "2025 sectors", "carriers"] + VARS + \
            ["total seats (held)", "OAG median seats", "OAG median premium", "seat variance"]
    ws = sheet("Held", heads, [9, 14, 12, 9] + [11, 10, 10, 12, 12, 14, 12, 11, 14, 13, 14, 12, 46] +
               [13, 13, 14, 12])
    r = 2
    for key in sorted(AIRCRAFT, key=lambda k: -(by_key.get(k, {}).get("sectors", 0))):
        v = AIRCRAFT[key]; b = by_key.get(key, {})
        ws.cell(row=r, column=1, value=key).font = Font(name=ARIAL, bold=True, size=10)
        ws.cell(row=r, column=2, value=", ".join(sorted(b.get("codes", []))) or "not in the 2025 schedule").font = body
        ws.cell(row=r, column=3, value=b.get("sectors") or 0).font = body
        ws.cell(row=r, column=4, value=b.get("carriers") or 0).font = body
        for j, name in enumerate(VARS, 5):
            ws.cell(row=r, column=j, value=v.get(name)).font = blue
        ws.cell(row=r, column=18, value=f"=E{r}+F{r}").font = body
        ws.cell(row=r, column=19, value=b.get("seats") or None).font = body
        ws.cell(row=r, column=20, value=b.get("prem") or None).font = body
        ws.cell(row=r, column=21, value=f'=IFERROR(R{r}/S{r}-1,"")').font = body
        ws.cell(row=r, column=21).number_format = "0.0%"
        for j in range(1, 22):
            ws.cell(row=r, column=j).border = thin
        r += 1
    ws.cell(row=r + 1, column=1, value="seat variance = held total seats over the OAG fleet-wide "
            "median. A large figure is not necessarily wrong: the held cabin is one representative "
            "configuration and the median is across every operator.").font = note

    # ---------------------------------------------------------------- to add
    adds = {}
    for code, (key, status) in MAP.items():
        if status != "ADD" or code not in cen:
            continue
        adds.setdefault(key, []).append(code)
    heads = ["key", "type", "OAG codes", "2025 sectors", "carriers", "OAG median seats",
             "OAG median premium", "OAG median sector km", "OAG max sector km"] + VARS
    ws = sheet("To add", heads, [10, 30, 14, 12, 9, 13, 14, 15, 14] +
               [11, 10, 10, 12, 12, 14, 12, 11, 14, 13, 14, 12, 46])
    r = 2
    for key in sorted(adds, key=lambda k: -by_key.get(k, {}).get("sectors", 0)):
        b = by_key.get(key, {}); codes = sorted(adds[key])
        nm = cen[codes[0]]["name"]
        ws.cell(row=r, column=1, value=key).font = Font(name=ARIAL, bold=True, size=10)
        ws.cell(row=r, column=2, value=nm).font = body
        ws.cell(row=r, column=3, value=", ".join(codes)).font = body
        for j, val in ((4, b.get("sectors")), (5, b.get("carriers")), (6, b.get("seats")),
                       (7, b.get("prem")), (8, b.get("gcd")), (9, b.get("gcd_max"))):
            ws.cell(row=r, column=j, value=val or 0).font = body
        for j in range(10, 10 + len(VARS)):
            c = ws.cell(row=r, column=j); c.fill = fillin; c.font = blue; c.border = thin
        for j in range(1, 10):
            ws.cell(row=r, column=j).border = thin
        r += 1
    ws.cell(row=r + 1, column=1, value="Fill the yellow cells. Example, from the held A350-900 row: "
            "econ_seats 300, bus_seats 36, range_km 15000, category Widebody, mtow_kg 280000, "
            "fuel_burn_kg_per_bh 5800, maint_per_bh 1200, crew_per_bh 1700, ownership_per_bh 2700, "
            "annual_util_bh 4400, price_usd 150000000, cargo_cap_kg 16000, src 'A350 newest; FAA WB "
            "+ lease; burn published'.").font = note
    ws.cell(row=r + 2, column=1, value="The OAG columns describe what the world flies. The cabin you "
            "enter should be a representative configuration for the type, and where a route needs a "
            "specific carrier's cabin the engine already reads that from OAG per carrier.").font = note

    # ---------------------------------------------------------------- code map
    heads = ["OAG code", "OAG name", "resolves to", "status", "2025 sectors", "carriers",
             "OAG median seats", "OAG median premium", "OAG median sector km"]
    ws = sheet("Code map", heads, [10, 42, 12, 10, 12, 9, 14, 16, 18])
    r = 2
    for rec in sorted(cen.values(), key=lambda x: -x["sectors"]):
        key, status = MAP.get(rec["code"], (None, "UNMAPPED"))
        if status == "EXCLUDE":
            continue
        ws.cell(row=r, column=1, value=rec["code"]).font = Font(name=ARIAL, bold=True, size=10)
        ws.cell(row=r, column=2, value=rec["name"]).font = body
        ws.cell(row=r, column=3, value=key or "-").font = body
        c = ws.cell(row=r, column=4, value=status); c.font = body
        if status in ("FAMILY", "REVIEW", "UNMAPPED"):
            c.fill = fillin
        for j, v in ((5, rec["sectors"]), (6, rec["carriers"]), (7, rec["seats"]),
                     (8, rec["prem"]), (9, rec["gcd"])):
            ws.cell(row=r, column=j, value=v).font = body
        for j in range(1, 10):
            ws.cell(row=r, column=j).border = thin
        r += 1
    ws.cell(row=r + 1, column=1, value="HELD the type is already in the economics module. ADD a new "
            "entry is needed. FAMILY a generic OAG code covering several variants, resolved to its "
            "commonest and shown here for review. REVIEW could not be resolved honestly. UNMAPPED "
            "appeared in the schedule after this table was written and needs a decision.").font = note

    # ---------------------------------------------------------------- excluded
    ws = sheet("Excluded", ["OAG code", "OAG name", "2025 sectors", "carriers", "median seats", "reason"],
               [10, 42, 12, 9, 13, 70])
    r = 2
    for rec in sorted(cen.values(), key=lambda x: -x["sectors"]):
        key, status = MAP.get(rec["code"], (None, "UNMAPPED"))
        if status not in ("EXCLUDE", "REVIEW"):
            continue
        for j, v in ((1, rec["code"]), (2, rec["name"]), (3, rec["sectors"]),
                     (4, rec["carriers"]), (5, rec["seats"]),
                     (6, EXCLUDE_REASON if status == "EXCLUDE" else REVIEW_REASON)):
            c = ws.cell(row=r, column=j, value=v); c.font = body; c.border = thin
        r += 1

    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    wb.properties.title = "Aircraft economics: held, missing and the OAG evidence"
    wb.properties.company = "Avia Solutions"
    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "Aircraft economics - gap and fill-in.xlsx"
    print("written:", build(out))
