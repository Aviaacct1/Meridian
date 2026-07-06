#!/usr/bin/env python3
"""
Avia Solutions  OAG Schedule Parser (Module II)
================================================
Reads raw OAG Analyser Excel exports and produces structured network data
for the Connection Builder (Module III).

Usage:
  python3 oag_parser.py --validate   # Run against project test files
  python3 oag_parser.py --home LHR --dest SJC --home-oag FILE --dest-oag FILE [options]

Version: 1.0
"""
from config import REFERENCE_CASE_DIR, OUTPUT_DIR, AIRPORT_DB, CITY_LOOKUP, ensure_output_dir
import argparse, math, os, sys, datetime
from collections import defaultdict

# 
# PART 1: DATA STRUCTURES & UTILITIES
# 

class Flight:
    __slots__ = ['carrier','carrier_name','flight_no','codeshare_carrier',
        'codeshare_name','codeshare_flight',
        'dep_airport','dep_city','dep_country','dep_region','dep_terminal',
        'arr_airport','arr_city','arr_country','arr_region','arr_terminal',
        'dep_time','arr_time','arr_day_offset','days_of_op','frequency',
        'seats','fst_seats','bus_seats','eco_seats','eff_from','eff_to',
        'elapsed_time_str','flying_time_str','stops','service_type',
        'direction','dom_int_dep','dom_int_arr']
    def __init__(self, **kw):
        for s in self.__slots__:
            setattr(self, s, kw.get(s))
    @property
    def elapsed_minutes(self):
        return _parse_time_to_minutes(self.elapsed_time_str)
    @property
    def dep_minutes(self):
        return _parse_hhmm_to_minutes(self.dep_time)
    @property
    def arr_minutes(self):
        return _parse_hhmm_to_minutes(self.arr_time)

class MCTEntry:
    __slots__ = ['airport','terminal_in','terminal_out','dom_int','minutes',
                 'arr_carrier','arr_flights','dep_carrier','dep_flights']
    def __init__(self, **kw):
        for s in self.__slots__:
            setattr(self, s, kw.get(s))

def _parse_hhmm_to_minutes(val):
    if val is None: return None
    s = str(val).strip()
    if not s: return None
    s = s.zfill(4)
    try: return int(s[:2])*60 + int(s[2:])
    except ValueError: return None

def _parse_time_to_minutes(val):
    if val is None: return None
    if hasattr(val, 'hour'): return val.hour*60 + val.minute
    s = str(val).strip()
    if ':' in s:
        parts = s.split(':')
        try: return int(parts[0])*60 + int(parts[1])
        except ValueError: return None
    return _parse_hhmm_to_minutes(s)

def _count_frequency(days_str):
    if not days_str: return 0
    return sum(1 for c in str(days_str) if c.isdigit())

def haversine_nm(lat1, lon1, lat2, lon2):
    R = 3440.065
    lat1,lon1,lat2,lon2 = map(math.radians,[lat1,lon1,lat2,lon2])
    dlat,dlon = lat2-lat1, lon2-lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return 2*R*math.asin(math.sqrt(a))


# 
# PART 2: IATA  ICAO COORDINATE MAPPING
# 

IATA_TO_ICAO = {
    'LHR':'EGLL','LGW':'EGKK','STN':'EGSS','LTN':'EGGW','LCY':'EGLC',
    'MAN':'EGCC','BHX':'EGBB','EDI':'EGPH','GLA':'EGPF','BRS':'EGGD',
    'NCL':'EGNT','LBA':'EGNM','EMA':'EGNX','ABZ':'EGPD','BFS':'EGAA',
    'BHD':'EGAC','CWL':'EGFF','SOU':'EGHI','EXT':'EGTE','NWI':'EGSH',
    'INV':'EGPE','SJC':'KSJC','SFO':'KSFO','OAK':'KOAK','LAX':'KLAX',
    'JFK':'KJFK','EWR':'KEWR','ORD':'KORD','DFW':'KDFW','IAH':'KIAH',
    'ATL':'KATL','MIA':'KMIA','BOS':'KBOS','SEA':'KSEA','DEN':'KDEN',
    'PHX':'KPHX','IAD':'KIAD','DCA':'KDCA','MSP':'KMSP','DTW':'KDTW',
    'CLT':'KCLT','PHL':'KPHL','TPA':'KTPA','MCO':'KMCO','FLL':'KFLL',
    'SAN':'KSAN','PDX':'KPDX','SMF':'KSMF','SLC':'KSLC','AUS':'KAUS',
    'RDU':'KRDU','BNA':'KBNA','HNL':'PHNL','ANC':'PANC','LAS':'KLAS',
    'PIT':'KPIT','STL':'KSTL','MCI':'KMCI','IND':'KIND','CMH':'KCMH',
    'SNA':'KSNA','BUR':'KBUR','ONT':'KONT','OKC':'KOKC','SAT':'KSAT',
    'JAX':'KJAX','PBI':'KPBI','RNO':'KRNO','BOI':'KBOI','ABQ':'KABQ',
    'CDG':'LFPG','ORY':'LFPO','AMS':'EHAM','FRA':'EDDF','MUC':'EDDM',
    'BCN':'LEBL','MAD':'LEMD','FCO':'LIRF','MXP':'LIMC','ZRH':'LSZH',
    'VIE':'LOWW','BRU':'EBBR','CPH':'EKCH','OSL':'ENGM','ARN':'ESSA',
    'HEL':'EFHK','DUB':'EIDW','LIS':'LPPT','ATH':'LGAV','IST':'LTFM',
    'WAW':'EPWA','PRG':'LKPR','BUD':'LHBP','OTP':'LROP',
    'DUS':'EDDL','HAM':'EDDH','TXL':'EDDT','BER':'EDDB',
    'KEF':'BIKF','NOC':'EIKN','SNN':'EINN','ORK':'EICK',
    'PMI':'LEPA','AGP':'LEMG','ALC':'LEAL','TFS':'GCTS',
    'NCE':'LFMN','LYS':'LFLL','MRS':'LFML',
    'PSA':'LIRP','VCE':'LIPZ','NAP':'LIRN','BLQ':'LIPE',
    'GVA':'LSGG','BSL':'LFSB','GIG':'SBGL',
    'DXB':'OMDB','AUH':'OMAA','DOH':'OTHH','BAH':'OBBI',
    'JED':'OEJN','RUH':'OERK','TLV':'LLBG','AMM':'OJAI',
    'HKG':'VHHH','SIN':'WSSS','NRT':'RJAA','HND':'RJTT',
    'ICN':'RKSI','PEK':'ZBAA','PVG':'ZSPD','BKK':'VTBS',
    'KUL':'WMKK','DEL':'VIDP','BOM':'VABB','TPE':'RCTP',
    'MNL':'RPLL','CGK':'WIII','CAN':'ZGGG','CTU':'ZUUU',
    'JNB':'FAOR','CPT':'FACT','NBO':'HKJK','LOS':'DNMM',
    'ACC':'DGAA','ADD':'HAAB','CAI':'HECA','CMN':'GMMN','ABV':'DNAA',
    'SYD':'YSSY','MEL':'YMML','AKL':'NZAA',
    'YYZ':'CYYZ','YVR':'CYVR','YUL':'CYUL','YOW':'CYOW',
    'MEX':'MMMX','GRU':'SBGR','BOG':'SKBO','SCL':'SCEL',
    'LIM':'SPJC','PTY':'MPTO','CUN':'MMUN','GDL':'MMGL',
    'DMM':'OEDF','LKO':'VILK','HYD':'VOHS','MAA':'VOMM','CCU':'VECC',
    'DAC':'VGHS','KHI':'OPKC','ISB':'OPRN','LHE':'OPLA',
    'CMB':'VCBI','MLE':'VRMM','KTM':'VNKT',
    'LED':'ULLI','SVO':'UUEE','DME':'UUDD',
    'BEG':'LYBE','SOF':'LBSF','ZAG':'LDZA','LJU':'LJLJ',
    'SKP':'LWSK','TIA':'LATI',
}

def read_airport_database(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    db = {}
    for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
        ident = str(row[1]).strip() if row[1] else ''
        lat, lon = row[4], row[5]
        if ident and lat and lon:
            db[ident] = (float(lat), float(lon))
    wb.close()
    return db

def build_coordinate_lookup(airports_by_icao):
    coords = {}
    for iata, icao in IATA_TO_ICAO.items():
        if icao in airports_by_icao:
            coords[iata] = airports_by_icao[icao]
    for icao, ll in airports_by_icao.items():
        if icao.startswith('K') and len(icao)==4:
            iata = icao[1:]
            if iata not in coords:
                coords[iata] = ll
    return coords


# 
# PART 3: OAG FILE READER
# 

def read_oag_xlsx(filepath, home_airport=None):
    """Read OAG Analyser export. Returns (metadata, flights, beyond_dests)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    meta = {'file': os.path.basename(filepath), 'sheets': wb.sheetnames}

    # Find Raw Data sheet
    raw_name = None
    for s in wb.sheetnames:
        if 'raw' in s.lower() and 'day' not in s.lower():
            raw_name = s; break
    if not raw_name: raw_name = wb.sheetnames[0]
    ws = wb[raw_name]

    # Find header row (Carrier1)
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=5, values_only=False):
        for c in row:
            v = str(c.value) if c.value else ''
            if v.startswith('VERSION'): meta['version'] = v
            if v.startswith('Period:'): meta['period'] = v[7:].strip()
            if c.value == 'Carrier1':
                header_row = c.row; break
        if header_row: break
    if not header_row:
        wb.close()
        raise ValueError(f"No header row in {filepath}")

    col_map = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, max_col=70, values_only=False):
        for c in row:
            if c.value: col_map[str(c.value).strip()] = c.column
    meta['header_row'] = header_row
    meta['columns'] = len(col_map)

    flights = []
    def _get(cells, name):
        idx = col_map.get(name)
        if idx is None: return None
        for c in cells:
            try:
                if c.column == idx: return c.value
            except AttributeError:
                continue
        return None

    max_col = max(col_map.values()) + 1
    for row in ws.iter_rows(min_row=header_row+1, max_col=max_col, values_only=False):
        carrier = _get(row, 'Carrier1')
        if not carrier: continue
        f = Flight(
            carrier=str(carrier).strip(),
            carrier_name=_get(row,'Carrier1Name'),
            flight_no=_get(row,'FlightNo1'),
            codeshare_carrier=str(_get(row,'Carrier2') or ''),
            codeshare_name=str(_get(row,'Carrier2Name') or ''),
            codeshare_flight=str(_get(row,'FlightNo2') or ''),
            dep_airport=str(_get(row,'DepAirport') or '').strip(),
            dep_city=str(_get(row,'DepCity') or '').strip(),
            dep_country=str(_get(row,'DepIATACtryName') or '').strip(),
            dep_region=str(_get(row,'DepRegName') or '').strip(),
            dep_terminal=str(_get(row,'DepTerminal') or '').strip(),
            arr_airport=str(_get(row,'ArrAirport') or '').strip(),
            arr_city=str(_get(row,'ArrCity') or '').strip(),
            arr_country=str(_get(row,'ArrIATACtryName') or '').strip(),
            arr_region=str(_get(row,'ArrRegName') or '').strip(),
            arr_terminal=str(_get(row,'ArrTerminal') or '').strip(),
            dep_time=_get(row,'LocalDepTime'),
            arr_time=_get(row,'LocalArrTime'),
            arr_day_offset=_get(row,'LocalArrday') or '',
            days_of_op=str(_get(row,'LocaldaysOfOp') or ''),
            frequency=_count_frequency(_get(row,'LocaldaysOfOp')),
            seats=_get(row,'Seats') or _get(row,'Seats (total)') or 0,
            fst_seats=_get(row,'FstSeats') or 0,
            bus_seats=_get(row,'BusSeats') or 0,
            eco_seats=_get(row,'EcoSeats') or 0,
            eff_from=_get(row,'EffFrom'), eff_to=_get(row,'EffTo'),
            elapsed_time_str=_get(row,'ElapsedTime'),
            flying_time_str=_get(row,'FlyingTime'),
            stops=_get(row,'Stops') or 0,
            service_type=str(_get(row,'Service') or '').strip(),
            direction=str(_get(row,'DEP/ARR') or '').strip(),
            dom_int_dep=str(_get(row,'CarrDom1') or '').strip(),
            dom_int_arr=str(_get(row,'CarrDom2') or '').strip(),
        )
        flights.append(f)

    # Read Beyond Destinations
    beyond = []
    bsheet = None
    for s in wb.sheetnames:
        sl = s.lower()
        if 'dest' in sl and ('beyond' in sl or 'taken' not in sl) and 'lookup' not in sl:
            bsheet = s; break
    if bsheet:
        ws_b = wb[bsheet]
        bhr, bcm = None, {}
        for row in ws_b.iter_rows(min_row=1, max_row=10, max_col=15, values_only=False):
            for c in row:
                if c.value and str(c.value).strip() in ('ArrAirport','FROM','VIA','Include?'):
                    bhr = c.row; break
            if bhr: break
        if bhr:
            for row in ws_b.iter_rows(min_row=bhr, max_row=bhr, max_col=15, values_only=False):
                for c in row:
                    if c.value: bcm[str(c.value).strip()] = c.column
            for row in ws_b.iter_rows(min_row=bhr+1, max_col=max(bcm.values())+1, values_only=False):
                def _bg(n):
                    idx = bcm.get(n)
                    if idx is None: return None
                    for c in row:
                        try:
                            if c.column == idx: return c.value
                        except AttributeError:
                            continue
                    return None
                apt = _bg('ArrAirport')
                if not apt: continue
                beyond.append({
                    'from': str(_bg('FROM') or '').strip(),
                    'via': str(_bg('VIA') or '').strip(),
                    'airport': str(apt).strip(),
                    'city': str(_bg('ArrCity') or '').strip(),
                    'country': str(_bg('ArrIATACtryName') or '').strip(),
                    'region': str(_bg('ArrRegName') or '').strip(),
                    'include': str(_bg('Include?') or '').strip().upper(),
                })

    wb.close()
    meta['flights'] = len(flights)
    meta['beyond_destinations'] = len(beyond)
    meta['beyond_included'] = sum(1 for b in beyond if b['include']=='IN')
    return meta, flights, beyond


# 
# PART 4: MCT & CITY LOOKUP READERS
# 

def read_city_lookup(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    a2c, c2a = {}, defaultdict(list)
    ws = None
    for s in wb.sheetnames:
        if s.lower()=='data': ws=wb[s]; break
    if not ws: ws=wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        apt,city = str(row[0]).strip() if row[0] else '', str(row[1]).strip() if row[1] else ''
        if apt and city:
            a2c[apt]=city; c2a[city].append(apt)
    wb.close()
    return a2c, dict(c2a)

def read_mct_xls(filepath):
    try: import xlrd
    except ImportError:
        print("WARNING: xlrd not installed"); return []
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    entries = []
    hdr = None
    for r in range(ws.nrows):
        v = str(ws.cell(r,0).value).strip()
        if v == 'Arrival Airport Code': hdr=r; break
    if hdr is None:
        # Compact format (LHR style)
        for r in range(ws.nrows):
            if str(ws.cell(r,0).value).strip()=='MCT Time (min)':
                for r2 in range(r+2, ws.nrows):
                    rv = [ws.cell(r2,c).value for c in range(min(5,ws.ncols))]
                    if not any(rv): continue
                    try: mins=int(float(str(rv[0])))
                    except: continue
                    apts = str(rv[1]) if rv[1] else ''
                    di = str(rv[2]) if rv[2] else ''
                    parts = apts.split('/')
                    apt1 = parts[0].strip().upper() if parts else apts.strip().upper()
                    entries.append(MCTEntry(airport=apt1,terminal_in='',terminal_out='',
                        dom_int=di,minutes=mins,arr_carrier='',arr_flights='',
                        dep_carrier='',dep_flights=''))
                break
        return entries
    for r in range(hdr+1, ws.nrows):
        apt = str(ws.cell(r,0).value).strip() if ws.cell(r,0).value else ''
        tv = ws.cell(r,1).value
        di = str(ws.cell(r,2).value).strip() if ws.cell(r,2).value else ''
        if not apt or not tv: continue
        try: mins=int(float(str(tv)))
        except: continue
        ac = str(ws.cell(r,4).value).strip() if ws.ncols>4 and ws.cell(r,4).value else ''
        dc = str(ws.cell(r,6).value).strip() if ws.ncols>6 and ws.cell(r,6).value else ''
        entries.append(MCTEntry(airport=apt,terminal_in='',terminal_out='',
            dom_int=di,minutes=mins,arr_carrier=ac,arr_flights='',
            dep_carrier=dc,dep_flights=''))
    return entries

def build_mct_lookup(entries, airport_code):
    lookup = {}
    for e in entries:
        if e.airport.upper()==airport_code.upper():
            k = e.dom_int.upper().replace(' ','').replace('_','')
            for prefix,key in [('DOMESTIC TO DOMESTIC','DOMDOM'),('DOMESTIC TO INT','DOMINT'),
                               ('INTERNATIONAL TO DOMESTIC','INTDOM'),('INTERNATIONAL TO INT','INTINT')]:
                if k.startswith(prefix.replace(' ','')):
                    k=key; break
            if k not in lookup or e.minutes < lookup[k]:
                lookup[k] = e.minutes
    for k,v in {'DOMDOM':45,'DOMINT':60,'INTDOM':60,'INTINT':60}.items():
        if k not in lookup: lookup[k]=v
    return lookup


# 
# PART 5: NETWORK ANALYSIS
# 

def build_network_summary(flights, airport_code):
    net = defaultdict(lambda: {'carriers':set(),'total_freq':0,'total_seats':0,
                                'routes':[],'city':'','country':'','region':''})
    for f in flights:
        if f.dep_airport==airport_code:
            dest,city,ctry,reg = f.arr_airport,f.arr_city,f.arr_country,f.arr_region
        elif f.arr_airport==airport_code:
            dest,city,ctry,reg = f.dep_airport,f.dep_city,f.dep_country,f.dep_region
        else: continue
        if not dest: continue
        n = net[dest]
        n['carriers'].add(f.carrier)
        n['total_freq'] += f.frequency
        n['total_seats'] += (int(f.seats) if f.seats else 0) * max(1, f.frequency)
        n['city']=city or n['city']; n['country']=ctry or n['country']; n['region']=reg or n['region']
        n['routes'].append(f)
    return dict(net)

def filter_beyond_destinations(beyond_list, circ_thresh, coords, home_apt, dest_apt):
    if home_apt not in coords or dest_apt not in coords:
        return [b for b in beyond_list if b['include']=='IN']
    h_ll, d_ll = coords[home_apt], coords[dest_apt]
    filtered = []
    for b in beyond_list:
        apt = b['airport']
        if apt not in coords:
            b['circuity']=None; b['circuity_status']='NO_COORDS'
            if b['include']=='IN': filtered.append(b)
            continue
        b_ll = coords[apt]
        routing = haversine_nm(*h_ll, *d_ll) + haversine_nm(*d_ll, *b_ll)
        direct = haversine_nm(*h_ll, *b_ll)
        ratio = routing/direct if direct>0 else 1.0
        b['circuity']=round(ratio,3); b['direct_nm']=round(direct); b['routing_nm']=round(routing)
        if ratio <= circ_thresh:
            b['circuity_status']='PASS'; filtered.append(b)
        elif b['include']=='IN':
            b['circuity_status']='EXPERT_OVERRIDE'; filtered.append(b)
        else:
            b['circuity_status']='FAIL'
    return filtered

def build_wave_analysis(flights, airport_code, direction='DEP'):
    waves = defaultdict(lambda: {'count':0,'flights':[]})
    for f in flights:
        if direction=='DEP' and f.dep_airport==airport_code: tv=f.dep_time
        elif direction=='ARR' and f.arr_airport==airport_code: tv=f.arr_time
        else: continue
        mins = _parse_hhmm_to_minutes(tv)
        if mins is None: continue
        h = mins//60
        wk = f"{h:02d}-{(h+1)%24:02d}"
        waves[wk]['count'] += f.frequency
        waves[wk]['flights'].append(f)
    return dict(sorted(waves.items()))


# 
# PART 6: EXCEL OUTPUT WRITER
# 

def write_output(filepath, home_apt, dest_apt,
                 home_meta, home_flights, home_beyond,
                 dest_meta, dest_flights, dest_beyond,
                 home_network, dest_network,
                 coords, home_mct_lookup, dest_mct_lookup,
                 airport_to_city, circ_thresh,
                 filtered_home_beyond, filtered_dest_beyond):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    HF = Font(bold=True,size=11,name='Arial',color='FFFFFF')
    HFL = PatternFill('solid',fgColor='1F4E79')
    DF = Font(size=10,name='Arial')
    TF = Font(bold=True,size=14,name='Arial',color='1F4E79')
    SF = Font(bold=True,size=11,name='Arial',color='4472C4')
    GF = PatternFill('solid',fgColor='E2EFDA')
    RF = PatternFill('solid',fgColor='FCE4EC')
    YF = PatternFill('solid',fgColor='FFFFCC')

    def whdr(ws,r,hdrs):
        for c,h in enumerate(hdrs,1):
            cl=ws.cell(r,c,h); cl.font=HF; cl.fill=HFL; cl.alignment=Alignment(horizontal='center',wrap_text=True)

    def aw(ws,mx=None):
        for col in ws.columns:
            if mx and col[0].column>mx: break
            ml=max((len(str(c.value or '')) for c in col),default=0)
            ws.column_dimensions[col[0].column_letter].width=min(ml+3,30)

    #  Metadata 
    wm=wb.active; wm.title='Metadata'
    wm.cell(1,1,f'OAG Schedule Parser  {home_apt} to {dest_apt}').font=TF
    wm.cell(2,1,f'Generated: {datetime.datetime.now():%Y-%m-%d %H:%M}').font=DF
    params=[('Home Airport',home_apt),('Destination Airport',dest_apt),
        ('Home OAG File',home_meta.get('file','')),('Dest OAG File',dest_meta.get('file','')),
        ('OAG Period (Home)',home_meta.get('period','')),('OAG Period (Dest)',dest_meta.get('period','')),
        ('Home Flights Parsed',home_meta.get('flights',0)),('Dest Flights Parsed',dest_meta.get('flights',0)),
        ('Home Network Routes',len(home_network)),('Dest Network Routes',len(dest_network)),
        ('Circuity Threshold',f'{circ_thresh:.0%}'),
        ('Beyond Home (filtered)',len(filtered_home_beyond)),
        ('Beyond Dest (filtered)',len(filtered_dest_beyond))]
    whdr(wm,4,['Parameter','Value'])
    for i,(k,v) in enumerate(params):
        wm.cell(5+i,1,k).font=DF; wm.cell(5+i,2,v).font=DF
    aw(wm)

    #  Network sheets 
    nhdr=['Destination','City','Country','Region','Carriers','Freq/wk','Wkly Seats','Distance NM']
    for apt,net,label in [(home_apt,home_network,'Network_'+home_apt),(dest_apt,dest_network,'Network_'+dest_apt)]:
        ws=wb.create_sheet(label)
        ws.cell(1,1,f'Non-stop routes from {apt}').font=TF
        whdr(ws,3,nhdr); r=4
        for d,info in sorted(net.items()):
            dist=''
            if apt in coords and d in coords:
                dist=round(haversine_nm(*coords[apt],*coords[d]))
            ws.cell(r,1,d).font=DF; ws.cell(r,2,info['city']).font=DF
            ws.cell(r,3,info['country']).font=DF; ws.cell(r,4,info['region']).font=DF
            ws.cell(r,5,', '.join(sorted(info['carriers']))).font=DF
            ws.cell(r,6,info['total_freq']).font=DF
            ws.cell(r,7,info['total_seats']).font=DF
            ws.cell(r,8,dist).font=DF; r+=1
        aw(ws)

    #  Beyond Destinations 
    bhdr=['Airport','City','Country','Region','Include','Circuity','Direct NM','Via NM','Status']
    for label,blist,from_apt,via_apt in [
        (f'Beyond_{dest_apt}', filtered_dest_beyond, home_apt, dest_apt),
        (f'Beyond_{home_apt}', filtered_home_beyond, dest_apt, home_apt)]:
        ws=wb.create_sheet(label)
        ws.cell(1,1,f'Destinations beyond {via_apt} from {from_apt}').font=TF
        whdr(ws,3,bhdr); r=4
        for b in sorted(blist, key=lambda x:x.get('airport','')):
            ws.cell(r,1,b.get('airport','')).font=DF
            ws.cell(r,2,b.get('city','')).font=DF
            ws.cell(r,3,b.get('country','')).font=DF
            ws.cell(r,4,b.get('region','')).font=DF
            ws.cell(r,5,b.get('include','')).font=DF
            circ=b.get('circuity')
            ws.cell(r,6,f'{circ:.2f}' if circ else '').font=DF
            ws.cell(r,7,b.get('direct_nm','')).font=DF
            ws.cell(r,8,b.get('routing_nm','')).font=DF
            st=b.get('circuity_status','')
            cl=ws.cell(r,9,st); cl.font=DF
            if st=='PASS': cl.fill=GF
            elif st=='FAIL': cl.fill=RF
            elif st=='EXPERT_OVERRIDE': cl.fill=YF
            r+=1
        aw(ws)

    #  leg1 & leg2 (Connection Builder format) 
    lhdr=['ID','Carrier1','Carrier1Name','FlightNo1','CarrDom1','DepAirport',
          'DepTerminal','DepCity','ArrAirport','ArrTerminal','ArrCity',
          'LocalDepTime','LocalArrTime','LocaldaysOfOp','ElapsedTime',
          'Seats','FstSeats','BusSeats','EcoSeats']
    for label, apt, direction in [('leg1', home_apt, 'ARR'), ('leg2', home_apt, 'DEP')]:
        ws=wb.create_sheet(label)
        whdr(ws,1,lhdr); r=2; fid=1
        for f in home_flights:
            if (direction=='ARR' and f.arr_airport==apt) or (direction=='DEP' and f.dep_airport==apt):
                ws.cell(r,1,fid); ws.cell(r,2,f.carrier); ws.cell(r,3,f.carrier_name)
                ws.cell(r,4,f.flight_no); ws.cell(r,5,f.dom_int_dep)
                ws.cell(r,6,f.dep_airport); ws.cell(r,7,f.dep_terminal)
                ws.cell(r,8,f.dep_city); ws.cell(r,9,f.arr_airport)
                ws.cell(r,10,f.arr_terminal); ws.cell(r,11,f.arr_city)
                ws.cell(r,12,f.dep_time); ws.cell(r,13,f.arr_time)
                ws.cell(r,14,f.days_of_op); ws.cell(r,15,f.elapsed_time_str)
                ws.cell(r,16,f.seats); ws.cell(r,17,f.fst_seats)
                ws.cell(r,18,f.bus_seats); ws.cell(r,19,f.eco_seats)
                r+=1; fid+=1
        aw(ws,10)

    #  Wave Analysis 
    ws=wb.create_sheet('Wave_Analysis')
    ws.cell(1,1,f'Wave Analysis').font=TF; r=3
    for apt,fls in [(home_apt,home_flights),(dest_apt,dest_flights)]:
        for dirn,dl in [('DEP','Departures from'),('ARR','Arrivals to')]:
            ws.cell(r,1,f'{dl} {apt}').font=SF; r+=1
            whdr(ws,r,['Hour','Freq/wk']); r+=1
            waves=build_wave_analysis(fls,apt,dirn)
            for wk,info in waves.items():
                ws.cell(r,1,wk).font=DF; ws.cell(r,2,info['count']).font=DF; r+=1
            r+=1
    aw(ws)

    #  MCT 
    ws=wb.create_sheet('MCT')
    ws.cell(1,1,'Minimum Connection Times').font=TF
    whdr(ws,3,['Airport','Type','Minutes']); r=4
    for apt,lk in [(home_apt,home_mct_lookup),(dest_apt,dest_mct_lookup)]:
        for t in ['DOMDOM','DOMINT','INTDOM','INTINT']:
            ws.cell(r,1,apt).font=DF; ws.cell(r,2,t).font=DF; ws.cell(r,3,lk.get(t,60)).font=DF; r+=1
        r+=1
    aw(ws)

    wb.save(filepath)
    return filepath


# 
# PART 7: VALIDATION & CLI
# 

def validate():
    base = str(REFERENCE_CASE_DIR)
    print("="*60)
    print("OAG SCHEDULE PARSER  VALIDATION (LHR-SJC)")
    print("="*60)

    files = {
        'home_oag': os.path.join(base,'OAG__LHR__WORLD__LHR_AUG2014.xlsx'),
        'dest_oag': os.path.join(base,'OAG__SJC__WORLD__SJC_AUG2014.xlsx'),
        'airport_db': os.path.join(base,'Airport_Database.xlsx'),
        'home_mct': os.path.join(base,'LHR_MCTs.xls'),
        'dest_mct': os.path.join(base,'Minimum_Cnx_Times_SJC.xls'),
        'city_lookup': os.path.join(base,'OAG_Airport__City_Lookup_DS_25Feb11.xlsx'),
    }
    for n,p in files.items():
        print(f"  {'' if os.path.exists(p) else ''} {n}: {os.path.basename(p)}")
    if any(not os.path.exists(p) for p in files.values()):
        print("Missing files!"); return False

    print("\n Airport Database ")
    db = read_airport_database(files['airport_db'])
    coords = build_coordinate_lookup(db)
    print(f"  {len(db)} ICAO airports, {len(coords)} IATA coords")
    for a in ['LHR','SJC','SFO','JFK','CDG','FRA','AMS']:
        if a in coords: print(f"   {a}: {coords[a][0]:.4f}, {coords[a][1]:.4f}")
        else: print(f"   {a}: NOT FOUND")
    if 'LHR' in coords and 'SJC' in coords:
        d=haversine_nm(*coords['LHR'],*coords['SJC'])
        print(f"  LHR-SJC: {d:.0f} NM ({d*1.852:.0f} km)")

    print("\n City Lookup ")
    a2c,c2a = read_city_lookup(files['city_lookup'])
    print(f"  {len(a2c)} mappings")

    print("\n Home OAG (LHR) ")
    hm,hf,hb = read_oag_xlsx(files['home_oag'],'LHR')
    print(f"  {hm['flights']} flights, {hm['beyond_destinations']} beyond ({hm['beyond_included']} included)")

    print("\n Dest OAG (SJC) ")
    dm,df_,db_ = read_oag_xlsx(files['dest_oag'],'SJC')
    print(f"  {dm['flights']} flights, {dm['beyond_destinations']} beyond ({dm['beyond_included']} included)")

    print("\n MCTs ")
    hm_mct = read_mct_xls(files['home_mct'])
    dm_mct = read_mct_xls(files['dest_mct'])
    print(f"  LHR: {len(hm_mct)} entries, SJC: {len(dm_mct)} entries")
    hml = build_mct_lookup(hm_mct,'LHR'); dml = build_mct_lookup(dm_mct,'SJC')
    print(f"  LHR MCTs: {hml}")
    print(f"  SJC MCTs: {dml}")

    print("\n Networks ")
    hn = build_network_summary(hf,'LHR'); dn = build_network_summary(df_,'SJC')
    print(f"  LHR: {len(hn)} routes, SJC: {len(dn)} routes")
    print(f"  Top LHR:")
    for d,i in sorted(hn.items(),key=lambda x:-x[1]['total_freq'])[:10]:
        print(f"    {d} ({i['city']}): {i['total_freq']}x/wk [{', '.join(sorted(i['carriers']))}]")
    print(f"  Top SJC:")
    for d,i in sorted(dn.items(),key=lambda x:-x[1]['total_freq'])[:10]:
        print(f"    {d} ({i['city']}): {i['total_freq']}x/wk [{', '.join(sorted(i['carriers']))}]")

    print("\n Circuity Filter (30%) ")
    fh = filter_beyond_destinations(hb,1.30,coords,'SJC','LHR')
    fd = filter_beyond_destinations(db_,1.30,coords,'LHR','SJC')
    print(f"  Beyond LHR (SJC pax): {len(fh)}/{len(hb)}")
    print(f"  Beyond SJC (LHR pax): {len(fd)}/{len(db_)}")
    print(f"  Expected: ~171 beyond LHR, ~30 beyond SJC (from OAG sheets)")

    print(f"\n  Sample beyond SJC:")
    for b in sorted(fd,key=lambda x:x.get('airport',''))[:15]:
        c=f"{b.get('circuity',0):.2f}" if b.get('circuity') else 'N/A'
        print(f"    {b['airport']} ({b.get('city','')}) circ={c} {b.get('circuity_status','')}")

    print("\n Writing Output ")
    ensure_output_dir()
    out = str(OUTPUT_DIR / 'OAG_Parsed_LHR_SJC.xlsx')
    write_output(out,'LHR','SJC',hm,hf,hb,dm,df_,db_,hn,dn,coords,hml,dml,a2c,1.30,fh,fd)
    print(f"   {out}")

    import openpyxl as ox
    wb=ox.load_workbook(out,data_only=True,read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for s in wb.sheetnames:
        ws=wb[s]; rows=sum(1 for _ in ws.iter_rows(max_col=1,values_only=True))
        print(f"    {s}: {rows} rows")
    wb.close()

    print("\n" + "="*60)
    print("VALIDATION COMPLETE ")
    print("="*60)
    return True

def main():
    p = argparse.ArgumentParser(description='OAG Schedule Parser (Module II)')
    p.add_argument('--validate',action='store_true')
    p.add_argument('--home'); p.add_argument('--dest')
    p.add_argument('--home-oag'); p.add_argument('--dest-oag')
    p.add_argument('--airport-db'); p.add_argument('--home-mct'); p.add_argument('--dest-mct')
    p.add_argument('--city-lookup'); p.add_argument('--circuity',type=float,default=1.30)
    p.add_argument('--output')
    args = p.parse_args()
    if args.validate: validate(); return
    if not all([args.home,args.dest,args.home_oag,args.dest_oag]):
        print("Required: --home, --dest, --home-oag, --dest-oag"); p.print_help(); sys.exit(1)
    adb = args.airport_db or str(AIRPORT_DB)
    cl = args.city_lookup or str(CITY_LOOKUP)
    out = args.output or f'OAG_Parsed_{args.home}_{args.dest}.xlsx'
    db=read_airport_database(adb); coords=build_coordinate_lookup(db)
    a2c,c2a=read_city_lookup(cl)
    hm,hf,hb=read_oag_xlsx(args.home_oag,args.home)
    dm,df_,db_=read_oag_xlsx(args.dest_oag,args.dest)
    hmc,dmc=[],[]
    if args.home_mct: hmc=read_mct_xls(args.home_mct)
    if args.dest_mct: dmc=read_mct_xls(args.dest_mct)
    hml=build_mct_lookup(hmc,args.home); dml=build_mct_lookup(dmc,args.dest)
    hn=build_network_summary(hf,args.home); dn=build_network_summary(df_,args.dest)
    fh=filter_beyond_destinations(hb,args.circuity,coords,args.dest,args.home)
    fd=filter_beyond_destinations(db_,args.circuity,coords,args.home,args.dest)
    write_output(out,args.home,args.dest,hm,hf,hb,dm,df_,db_,hn,dn,coords,hml,dml,a2c,args.circuity,fh,fd)
    print(f" {out}  {args.home}:{len(hn)} routes, {args.dest}:{len(dn)} routes")

if __name__=='__main__':
    main()
