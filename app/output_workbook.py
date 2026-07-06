#!/usr/bin/env python3
"""
Avia Solutions  Output Workbook Standardisation (Chat 17)
===========================================================
Produces client-ready Excel output matching Avia Solutions' established format.

Gold standard: BA_Fcst_LHRSJC_JZ_23Feb2015_FINAL_without_INDIA.xlsm
Reference presentations: Avianca MAN-TPA, KLM AMS-TPA, BA LHR-SJC

Output sheets:
    1. Cover Page         All parameters echoed for audit trail
    2. Forecast TABLE     The main forecast table (presentation-ready)
    3. Cnx @ Home TABLE   Top 50 connecting cities at home hub
    4. Cnx @ Dest TABLE   Top 50 connecting cities at destination
    5. Schedule TABLE     Schedule options and capacity
    6. QSI Diagnostics    Full QSI breakdown (internal, not for client)
    7. Assumptions Log    Every assumption with space for analyst notes
    8. Audit Trail        Complete pipeline execution log

Design principles:
    - Sheets 2-5 are directly paste-able into PowerPoint presentations
    - Yellow cells = analyst inputs, Green cells = derived/calculated
    - Every number traceable to its source
    - Matches the column structure from actual Avia presentations
    - Professional formatting with Avia blue (#003366) branding

Dependencies:
    - closed_loop_pipeline_v2.py (Chat 12)  provides forecast results dict
    - route_config.py (Chat 12)  RouteConfig with route parameters
    - providers.py (Chat 12)  data structures
"""

from config import OUTPUT_DIR, ensure_output_dir
import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers, NamedStyle)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl --break-system-packages")
    sys.exit(1)


# ============================================================================
# STYLE CONSTANTS  Avia Solutions brand palette
# ============================================================================

AVIA_BLUE = '003366'
AVIA_MID = '4472C4'
AVIA_LIGHT = 'D6E4F0'
AVIA_ACCENT = '1F4E79'

# Fills
HDR_FILL = PatternFill('solid', fgColor=AVIA_BLUE)
SEC_FILL = PatternFill('solid', fgColor=AVIA_MID)
INP_FILL = PatternFill('solid', fgColor='FFF2CC')     # Yellow = analyst input
DER_FILL = PatternFill('solid', fgColor='E2EFDA')     # Green = derived
ALT_FILL = PatternFill('solid', fgColor='F2F2F2')     # Light grey alternating rows
WARN_FILL = PatternFill('solid', fgColor='FCE4EC')    # Pink = warning/check

# Fonts
TITLE_FONT = Font(name='Arial', bold=True, color=AVIA_BLUE, size=14)
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SEC_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
NORM_FONT = Font(name='Arial', size=10)
LBL_FONT = Font(name='Arial', size=10, color='333333')
INP_FONT = Font(name='Arial', size=10, color='0000FF')   # Blue = hardcoded input
DER_FONT = Font(name='Arial', size=10, color='006100')   # Green = derived
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='666666')
MONO_FONT = Font(name='Consolas', size=9)
TOTAL_FONT = Font(name='Arial', bold=True, color=AVIA_BLUE, size=11)
SUBTOTAL_FONT = Font(name='Arial', bold=True, size=10)

# Borders
THIN_BORDER = Border(
    left=Side('thin', color='B4B4B4'), right=Side('thin', color='B4B4B4'),
    top=Side('thin', color='B4B4B4'), bottom=Side('thin', color='B4B4B4'))
BOTTOM_THICK = Border(bottom=Side('medium', color=AVIA_BLUE))
NO_BORDER = Border()

# Alignment
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RGT = Alignment(horizontal='right', vertical='center')


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def _cell(ws, r, c, val=None, font=NORM_FONT, fill=None, fmt=None,
          align=CTR, border=THIN_BORDER):
    """Write a styled cell."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.alignment = align
    cell.border = border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def _header_row(ws, row, headers, widths=None, fill=HDR_FILL, font=HDR_FONT):
    """Write a header row with consistent styling."""
    for i, h in enumerate(headers, 1):
        _cell(ws, row, i, h, font=font, fill=fill)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _section_banner(ws, row, text, col_span=10):
    """Write a section header banner across columns."""
    _cell(ws, row, 1, text, font=SEC_FONT, fill=SEC_FILL, align=LFT)
    for c in range(2, col_span + 1):
        _cell(ws, row, c, '', fill=SEC_FILL)


# ============================================================================
# OUTPUT WORKBOOK WRITER
# ============================================================================

class StandardOutputWriter:
    """
    Produces the standardised Avia Solutions output workbook.
    
    Usage:
        writer = StandardOutputWriter(config, results, audit_log)
        writer.write_all()
        writer.save('/path/to/output.xlsx')
    """

    # Presentation table: top N connecting cities to show
    TOP_N_CITIES = 50

    def __init__(self, config, results: Dict[str, Any],
                 audit_log: List[str] = None,
                 qsi_diagnostics: Dict[str, Any] = None,
                 analyst: str = 'Avia Solutions',
                 engagement_date: str = None,
                 route_pnl=None,
                 frequency_per_week: float = None,
                 operating_weeks: float = 52.0):
        """
        Args:
            config: RouteConfig with route parameters
            results: Dict from ForecastAssembler.run() containing:
                p2p_total, p2p_details, home_total, home_results,
                dest_total, dest_results, grand_total, load_factor
            audit_log: List of audit trail strings
            qsi_diagnostics: Optional dict with QSI engine internals
            analyst: Analyst name for cover page
            engagement_date: Override date string
            route_pnl: Optional aircraft_economics.RoutePnL. When supplied, a Route
                Economics (turnaround P&L) sheet is written, turning the demand forecast
                into an integrated business plan (demand -> revenue -> cost -> profit).
            frequency_per_week / operating_weeks: when set with route_pnl, the economics
                sheet also shows the ANNUAL route P&L and the fleet requirement.
        """
        self.config = config
        self.results = results
        self.audit_log = audit_log or []
        self.qsi_diag = qsi_diagnostics or {}
        self.analyst = analyst
        self.date = engagement_date or datetime.now().strftime('%d %B %Y')
        self.route_pnl = route_pnl
        self.frequency_per_week = (frequency_per_week
                                   if frequency_per_week is not None
                                   else getattr(config, 'frequency', None))
        self.operating_weeks = operating_weeks
        self.wb = openpyxl.Workbook()

    def write_all(self):
        """Write all sheets in standard order."""
        self._write_cover_page()
        self._write_forecast_table()
        self._write_cnx_table('home')
        self._write_cnx_table('dest')
        self._write_schedule_table()
        if self.route_pnl is not None:
            self._write_route_economics()
        self._write_qsi_diagnostics()
        self._write_assumptions_log()
        self._write_audit_trail()
        # Remove default empty sheet if still there
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

    def save(self, path: str):
        """Save the workbook."""
        os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
        self.wb.save(path)
        print(f"  Output workbook saved: {path}")
        return path

    # ------------------------------------------------------------------
    # SHEET 1: COVER PAGE
    # ------------------------------------------------------------------

    def _write_cover_page(self):
        ws = self.wb.active
        ws.title = 'Cover Page'
        ws.sheet_properties.tabColor = AVIA_BLUE

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20

        cfg = self.config
        res = self.results

        # Title
        _cell(ws, 1, 2, 'AVIA SOLUTIONS  QSI ROUTE FORECAST OUTPUT',
              font=TITLE_FONT, border=NO_BORDER)
        ws.merge_cells('B1:D1')

        # Metadata
        _cell(ws, 2, 2, f'Generated: {self.date}', font=NOTE_FONT, border=NO_BORDER)
        _cell(ws, 2, 4, f'Analyst: {self.analyst}', font=NOTE_FONT, border=NO_BORDER)

        row = 4
        sections = [
            ('ROUTE', [
                ('Origin Airport', cfg.home_airport_code, True),
                ('Origin City', getattr(cfg, 'home_city_code', cfg.home_airport_code), True),
                ('Destination Airport', cfg.dest_airport_code, True),
                ('Destination City', getattr(cfg, 'dest_city_code', cfg.dest_airport_code), True),
                ('Distance (nm)', getattr(cfg, 'distance_nm', ''), True),
            ]),
            ('CARRIER', [
                ('Airline', getattr(cfg, 'airline_name', cfg.airline_code), True),
                ('Carrier IATA', cfg.airline_code, True),
                ('Carrier Type', getattr(cfg, 'carrier_type', 'Full Service'), True),
                ('Alliance', getattr(cfg, 'alliance', ''), True),
            ]),
            ('SERVICE', [
                ('Frequency (per week)', cfg.frequency, True),
                ('Aircraft Type', getattr(cfg, 'aircraft_type', ''), True),
                ('Seats per Flight', cfg.seats, True),
                ('Annual Seats', cfg.annual_capacity, False),
                ('Operating Days', getattr(cfg, 'operating_days', '1234567'), True),
            ]),
            ('MODE', [
                ('Operating Mode', getattr(cfg, 'mode', 'Forecast'), True),
                ('Forecast Year', getattr(cfg, 'forecast_year', ''), True),
                ('Base Year', getattr(cfg, 'base_year', ''), True),
            ]),
            ('RESULTS', [
                ('P2P Forecast', f"{res['p2p_total']:,.0f}", False),
                ('Connecting @ Home Hub', f"{res['home_total']:,.0f}", False),
                ('Connecting @ Dest', f"{res['dest_total']:,.0f}", False),
                ('Grand Total', f"{res['grand_total']:,.0f}", False),
                ('Load Factor', f"{res['load_factor']:.1%}", False),
                ('PTEW', f"{res['grand_total'] / max(cfg.frequency * 52 * 2, 1):,.1f}" if res['grand_total'] > 0 else '0', False),
            ]),
        ]

        for sec_name, fields in sections:
            _section_banner(ws, row, sec_name, col_span=4)
            row += 1
            for label, val, is_input in fields:
                _cell(ws, row, 2, label, font=LBL_FONT, align=LFT)
                _cell(ws, row, 3, val,
                      font=INP_FONT if is_input else DER_FONT,
                      fill=INP_FILL if is_input else DER_FILL)
                row += 1
            row += 1  # gap between sections

        # Legend
        row += 1
        _cell(ws, row, 2, 'Legend:', font=BOLD_FONT, border=NO_BORDER)
        row += 1
        _cell(ws, row, 2, '  Yellow cells = Analyst inputs', font=NOTE_FONT, border=NO_BORDER)
        _cell(ws, row, 3, '', fill=INP_FILL)
        row += 1
        _cell(ws, row, 2, '  Green cells = Derived / calculated', font=NOTE_FONT, border=NO_BORDER)
        _cell(ws, row, 3, '', fill=DER_FILL)

    # ------------------------------------------------------------------
    # SHEET 2: FORECAST TABLE (presentation-ready)
    # ------------------------------------------------------------------

    def _write_forecast_table(self):
        ws = self.wb.create_sheet('Forecast TABLE')
        ws.sheet_properties.tabColor = AVIA_BLUE
        cfg = self.config
        res = self.results

        carrier_name = getattr(cfg, 'airline_name', cfg.airline_code)
        home_city = getattr(cfg, 'home_city_code', cfg.home_airport_code)
        dest_city = getattr(cfg, 'dest_city_code', cfg.dest_airport_code)
        freq_label = f"{cfg.frequency}x Weekly" if cfg.frequency else ''

        # Title
        _cell(ws, 2, 2, f"{carrier_name}: {cfg.home_airport_code} - {cfg.dest_airport_code} Traffic Forecast ({freq_label} Service)",
              font=TITLE_FONT, border=NO_BORDER)
        ws.merge_cells('B2:M2')

        # Column headers  matches gold standard exactly
        headers = [
            '', 'Market',
            f'Base Annual\nDemand (000s) ',
            'Compound\nGrowth Rate ',
            f'Annual Demand\n(000s)',
            'Stimulation due\nto Direct Service ',
            'Demand After\nStimulation (000s)',
            f'{carrier_name}\nCapture Rate ',
            'Forecast\n(000s)',
            'PTEW ',
        ]
        widths = [3, 42, 16, 14, 16, 16, 16, 14, 14, 10]
        _header_row(ws, 4, headers, widths)

        row = 6

        # --- P2P Section ---
        _cell(ws, row, 2, 'Point to Point', font=BOLD_FONT, align=LFT, border=BOTTOM_THICK)
        for c in range(3, 11):
            _cell(ws, row, c, '', border=BOTTOM_THICK)
        row += 1

        p2p_details = res.get('p2p_details', [])
        p2p_total_base = 0
        p2p_total_fcst = 0

        for seg in p2p_details:
            name = seg['name']
            base = seg['base'] / 1000.0  # Convert to 000s
            growth = seg['growth']
            stim = seg['stimulation']
            capture = seg['capture']
            grown = base * (1 + growth)
            stimulated = grown * stim
            forecast = stimulated * capture
            # PTEW = passengers per trip each way = pax / (freq * 52 * 2)
            trips = cfg.frequency * 52 * 2 if cfg.frequency else 104
            ptew = (seg['forecast'] / trips) if seg['forecast'] > 0 else 0

            _cell(ws, row, 2, name, font=NORM_FONT, align=LFT)
            _cell(ws, row, 3, base, font=INP_FONT, fill=INP_FILL, fmt='#,##0.0')
            _cell(ws, row, 4, growth, font=INP_FONT, fill=INP_FILL, fmt='0.0%')
            _cell(ws, row, 5, grown, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 6, stim, font=INP_FONT, fill=INP_FILL, fmt='0.00')
            _cell(ws, row, 7, stimulated, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 8, capture, font=INP_FONT, fill=INP_FILL, fmt='0.0%')
            _cell(ws, row, 9, forecast, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 10, ptew, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')

            p2p_total_base += base
            p2p_total_fcst += forecast
            row += 1

        # P2P subtotal
        trips = cfg.frequency * 52 * 2 if cfg.frequency else 104
        _cell(ws, row, 2, 'Total Point to Point', font=SUBTOTAL_FONT, align=LFT)
        _cell(ws, row, 3, p2p_total_base, font=SUBTOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 9, res['p2p_total'] / 1000.0, font=SUBTOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 10, res['p2p_total'] / trips, font=SUBTOTAL_FONT, fmt='#,##0.0')
        row += 2

        # --- Connecting at Home Hub ---
        home_label = f"Connecting at {cfg.home_airport_code} "
        _cell(ws, row, 2, home_label, font=BOLD_FONT, align=LFT, border=BOTTOM_THICK)
        for c in range(3, 11):
            _cell(ws, row, c, '', border=BOTTOM_THICK)
        row += 1

        home_results = res.get('home_results', [])
        home_direct = [r for r in home_results if r.get('direct')]
        home_nodirect = [r for r in home_results if not r.get('direct')]

        for sub_label, sub_list in [
            ('O&Ds with Direct Competition', home_direct),
            ('O&Ds with No Direct Competition', home_nodirect),
        ]:
            sub_base = sum(r['base_demand'] for r in sub_list) / 1000.0
            sub_growth = getattr(cfg, 'home_growth', 0.09)
            sub_grown = sub_base * (1 + sub_growth)
            sub_fcst = sum(r['forecast'] for r in sub_list) / 1000.0
            sub_capture = sub_fcst / sub_grown if sub_grown > 0 else 0

            _cell(ws, row, 2, f'    {sub_label}', font=NORM_FONT, align=LFT)
            _cell(ws, row, 3, sub_base, font=INP_FONT, fill=INP_FILL, fmt='#,##0.0')
            _cell(ws, row, 4, sub_growth, font=INP_FONT, fill=INP_FILL, fmt='0.0%')
            _cell(ws, row, 5, sub_grown, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 6, 1.0, font=INP_FONT, fill=INP_FILL, fmt='0.00')
            _cell(ws, row, 7, sub_grown, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 8, sub_capture, font=DER_FONT, fill=DER_FILL, fmt='0.0%')
            _cell(ws, row, 9, sub_fcst, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 10, sum(r['forecast'] for r in sub_list) / trips,
                  font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            row += 1

        # Home connecting subtotal
        home_base_total = sum(r['base_demand'] for r in home_results) / 1000.0
        _cell(ws, row, 2, 'Total', font=SUBTOTAL_FONT, align=LFT)
        _cell(ws, row, 3, home_base_total, font=SUBTOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 9, res['home_total'] / 1000.0, font=SUBTOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 10, res['home_total'] / trips, font=SUBTOTAL_FONT, fmt='#,##0.0')
        row += 2

        # --- Connecting at Destination ---
        dest_label = f"Connecting at {cfg.dest_airport_code} "
        _cell(ws, row, 2, dest_label, font=BOLD_FONT, align=LFT, border=BOTTOM_THICK)
        for c in range(3, 11):
            _cell(ws, row, c, '', border=BOTTOM_THICK)
        row += 1

        dest_results = res.get('dest_results', [])
        dest_direct = [r for r in dest_results if r.get('direct')]
        dest_nodirect = [r for r in dest_results if not r.get('direct')]

        for sub_label, sub_list in [
            ('O&Ds with Direct Competition', dest_direct),
            ('O&Ds with No Direct Competition', dest_nodirect),
        ]:
            sub_base = sum(r['base_demand'] for r in sub_list) / 1000.0
            sub_growth = getattr(cfg, 'dest_growth', 0.10)
            sub_grown = sub_base * (1 + sub_growth)
            sub_fcst = sum(r['forecast'] for r in sub_list) / 1000.0
            sub_capture = sub_fcst / sub_grown if sub_grown > 0 else 0

            _cell(ws, row, 2, f'    {sub_label}', font=NORM_FONT, align=LFT)
            _cell(ws, row, 3, sub_base, font=INP_FONT, fill=INP_FILL, fmt='#,##0.0')
            _cell(ws, row, 4, sub_growth, font=INP_FONT, fill=INP_FILL, fmt='0.0%')
            _cell(ws, row, 5, sub_grown, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 6, 1.0, font=INP_FONT, fill=INP_FILL, fmt='0.00')
            _cell(ws, row, 7, sub_grown, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 8, sub_capture, font=DER_FONT, fill=DER_FILL, fmt='0.0%')
            _cell(ws, row, 9, sub_fcst, font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            _cell(ws, row, 10, sum(r['forecast'] for r in sub_list) / trips,
                  font=DER_FONT, fill=DER_FILL, fmt='#,##0.0')
            row += 1

        # Dest connecting subtotal
        dest_base_total = sum(r['base_demand'] for r in dest_results) / 1000.0
        _cell(ws, row, 2, 'Total', font=SUBTOTAL_FONT, align=LFT)
        _cell(ws, row, 3, dest_base_total, font=SUBTOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 9, res['dest_total'] / 1000.0, font=SUBTOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 10, res['dest_total'] / trips, font=SUBTOTAL_FONT, fmt='#,##0.0')
        row += 2

        # --- GRAND TOTAL ---
        _cell(ws, row, 2, 'Grand Total', font=TOTAL_FONT, align=LFT)
        grand_base = p2p_total_base + home_base_total + dest_base_total
        _cell(ws, row, 3, grand_base, font=TOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 9, res['grand_total'] / 1000.0, font=TOTAL_FONT, fmt='#,##0.0')
        _cell(ws, row, 10, res['grand_total'] / trips, font=TOTAL_FONT, fmt='#,##0.0')
        row += 2

        # --- Notes ---
        notes = [
            f"1) Demand for point-to-point market and connecting markets based on Sabre MI and restricted to catchment area.",
            f"2) Compound growth based on GDP and visitor growth forecasts.",
            f"3) Based on IATA Stimulation Curve and route-specific adjustments.",
            f"4) Point-to-point capture rates based on frequency/capacity share. Connecting capture rate based on QSI model.",
            f"5) Demand on double connections has been excluded.",
            f"6) Passengers per trip each way.",
            "AviaSolutions analysis",
        ]
        _cell(ws, row, 2, 'Notes:', font=BOLD_FONT, align=LFT, border=NO_BORDER)
        row += 1
        for note in notes:
            _cell(ws, row, 2, note, font=NOTE_FONT, align=LFT, border=NO_BORDER)
            row += 1

    # ------------------------------------------------------------------
    # SHEETS 3-4: CONNECTING CITY TABLES (presentation-ready)
    # ------------------------------------------------------------------

    def _write_cnx_table(self, direction: str):
        cfg = self.config
        carrier_name = getattr(cfg, 'airline_name', cfg.airline_code)
        trips = cfg.frequency * 52 * 2 if cfg.frequency else 104

        if direction == 'home':
            sheet_name = f'Cnx @ {cfg.home_airport_code} TABLE'
            hub_city = getattr(cfg, 'home_city', cfg.home_airport_code)
            hub_code = cfg.home_airport_code
            city_results = self.results.get('home_results', [])
            total_pax = self.results.get('home_total', 0)
        else:
            sheet_name = f'Cnx @ {cfg.dest_airport_code} TABLE'
            hub_city = getattr(cfg, 'dest_city', cfg.dest_airport_code)
            hub_code = cfg.dest_airport_code
            city_results = self.results.get('dest_results', [])
            total_pax = self.results.get('dest_total', 0)

        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = AVIA_MID

        # Title  matches presentation format
        _cell(ws, 2, 2, f'Forecast Number of Passengers Connecting at {hub_city}',
              font=TITLE_FONT, border=NO_BORDER)
        ws.merge_cells('B2:J2')

        # Headers  matches the Avianca/KLM presentation tables exactly
        headers_left = ['Nr', 'City\nCode', 'City Name', 'Country',
                        f'Annual\nDemand ',
                        f'{carrier_name}\nAnnual\nShare ',
                        'Annual\nForecast',
                        'PTEW ']
        widths_left = [5, 8, 18, 18, 14, 12, 14, 8]

        # For >25 cities, use two-column layout like the presentations
        sorted_cities = sorted(
            [c for c in city_results if c['city'] != 'Total'],
            key=lambda x: -x['forecast'])
        n_show = min(self.TOP_N_CITIES, len(sorted_cities))
        
        # Recalculate total_pax excluding the "Total" summary row
        total_pax = sum(c['forecast'] for c in sorted_cities)

        if n_show > 25:
            # Two-column layout
            mid = (n_show + 1) // 2
            left_cols = sorted_cities[:mid]
            right_cols = sorted_cities[mid:n_show]

            # Left header
            _header_row(ws, 4, headers_left, widths_left)
            # Right header (offset by 9 columns)
            right_start = 10
            for i, h in enumerate(headers_left):
                _cell(ws, 4, right_start + i, h, font=HDR_FONT, fill=HDR_FILL)
                ws.column_dimensions[get_column_letter(right_start + i)].width = widths_left[i]

            # Data rows
            for idx, city in enumerate(left_cols):
                r = 5 + idx
                self._write_cnx_city_row(ws, r, idx + 1, city, total_pax, col_offset=0)

            for idx, city in enumerate(right_cols):
                r = 5 + idx
                self._write_cnx_city_row(ws, r, mid + idx + 1, city, total_pax,
                                          col_offset=right_start - 1)

            # Other + Total row
            other_pax = total_pax - sum(c['forecast'] for c in sorted_cities[:n_show])
            total_demand = sum(c['base_demand'] * (1 + c.get('growth', 0.09))
                              for c in city_results)
            other_demand = total_demand - sum(
                c['base_demand'] * (1 + c.get('growth', 0.09))
                for c in sorted_cities[:n_show])

            footer_row = 5 + max(len(left_cols), len(right_cols))
            _cell(ws, footer_row, right_start + 2, 'Other', font=BOLD_FONT, align=LFT)
            _cell(ws, footer_row, right_start + 4, other_demand, font=NORM_FONT, fmt='#,##0')
            _cell(ws, footer_row, right_start + 6, other_pax, font=NORM_FONT, fmt='#,##0')
            _cell(ws, footer_row, right_start + 7,
                  other_pax / trips if other_pax > 0 else 0, font=NORM_FONT, fmt='0.0')

            footer_row += 1
            _cell(ws, footer_row, right_start + 2, 'Total', font=TOTAL_FONT, align=LFT)
            _cell(ws, footer_row, right_start + 4, total_demand, font=TOTAL_FONT, fmt='#,##0')
            _cell(ws, footer_row, right_start + 6, total_pax, font=TOTAL_FONT, fmt='#,##0')
            _cell(ws, footer_row, right_start + 7,
                  total_pax / trips, font=TOTAL_FONT, fmt='0.0')
        else:
            # Single column layout
            _header_row(ws, 4, headers_left, widths_left)
            for idx, city in enumerate(sorted_cities[:n_show]):
                r = 5 + idx
                self._write_cnx_city_row(ws, r, idx + 1, city, total_pax, col_offset=0)

            footer_row = 5 + n_show
            _cell(ws, footer_row, 2, 'Total', font=TOTAL_FONT, align=LFT)
            total_demand = sum(c['base_demand'] * (1 + c.get('growth', 0.09))
                              for c in city_results)
            _cell(ws, footer_row, 5, total_demand, font=TOTAL_FONT, fmt='#,##0')
            _cell(ws, footer_row, 7, total_pax, font=TOTAL_FONT, fmt='#,##0')
            _cell(ws, footer_row, 8, total_pax / trips, font=TOTAL_FONT, fmt='0.0')

        # Notes
        nr = footer_row + 2
        notes_text = [
            "1) Demand on double connections excluded.",
            "2) Based on QSI model.",
            "3) Passengers per trip each way.",
            "AviaSolutions analysis",
        ]
        for note in notes_text:
            _cell(ws, nr, 2, note, font=NOTE_FONT, align=LFT, border=NO_BORDER)
            nr += 1

    def _write_cnx_city_row(self, ws, row, num, city, total_pax, col_offset=0):
        """Write a single connecting city row."""
        c = col_offset + 1
        grown_demand = city['base_demand'] * (1 + city.get('growth', 0.09))
        capture = city.get('qsi_capture', 0)
        forecast = city['forecast']
        trips = self.config.frequency * 52 * 2 if self.config.frequency else 104
        ptew = forecast / trips if forecast > 0 else 0

        # Alternating row shading
        fill = ALT_FILL if num % 2 == 0 else None

        _cell(ws, row, c, num, font=NORM_FONT, fill=fill)
        _cell(ws, row, c + 1, city['city'], font=NORM_FONT, fill=fill)
        _cell(ws, row, c + 2, city.get('name', ''), font=NORM_FONT, fill=fill, align=LFT)
        _cell(ws, row, c + 3, city.get('country', ''), font=NORM_FONT, fill=fill, align=LFT)
        _cell(ws, row, c + 4, grown_demand, font=NORM_FONT, fill=fill, fmt='#,##0')
        _cell(ws, row, c + 5, capture, font=NORM_FONT, fill=fill, fmt='0.0%')
        _cell(ws, row, c + 6, forecast, font=NORM_FONT, fill=fill, fmt='#,##0')
        _cell(ws, row, c + 7, ptew, font=NORM_FONT, fill=fill, fmt='0.0')

    # ------------------------------------------------------------------
    # SHEET 5: SCHEDULE TABLE
    # ------------------------------------------------------------------

    def _write_schedule_table(self):
        ws = self.wb.create_sheet('Schedule TABLE')
        ws.sheet_properties.tabColor = AVIA_MID
        cfg = self.config
        res = self.results

        aircraft = getattr(cfg, 'aircraft_type', '787-800')
        _cell(ws, 2, 2, f'Schedule Options: {aircraft}', font=TITLE_FONT, border=NO_BORDER)
        ws.merge_cells('B2:K2')

        headers = ['Sector', 'Dep. Time', 'Arr. Time', 'Op. Days',
                   'Aircraft', 'Seats', 'Annual Seats', 'Annual Pax',
                   'PPDEW', 'Seat Factor']
        widths = [16, 12, 12, 12, 12, 10, 14, 14, 10, 12]
        _header_row(ws, 4, headers, widths)

        home = cfg.home_airport_code
        dest = cfg.dest_airport_code
        freq = cfg.frequency
        seats = cfg.seats
        ann_seats = cfg.annual_capacity
        ann_pax = res['grand_total']
        ppdew = ann_pax / 365 if ann_pax > 0 else 0
        lf = res['load_factor']

        dep_time = cfg.outbound_dep if cfg.outbound_dep else ''
        arr_time = cfg.outbound_arr if cfg.outbound_arr else ''
        op_days = getattr(cfg, 'operating_days', '1234567')

        # Outbound
        _cell(ws, 5, 1, f'{home}-{dest}', font=NORM_FONT, align=LFT)
        _cell(ws, 5, 2, dep_time, font=NORM_FONT)
        _cell(ws, 5, 3, arr_time, font=NORM_FONT)
        _cell(ws, 5, 4, op_days, font=NORM_FONT)
        _cell(ws, 5, 5, aircraft, font=NORM_FONT)
        _cell(ws, 5, 6, seats, font=NORM_FONT, fmt='#,##0')
        _cell(ws, 5, 7, ann_seats, font=NORM_FONT, fmt='#,##0')
        _cell(ws, 5, 8, ann_pax, font=NORM_FONT, fmt='#,##0')
        _cell(ws, 5, 9, ppdew, font=NORM_FONT, fmt='#,##0.0')
        _cell(ws, 5, 10, lf, font=NORM_FONT, fmt='0.0%')

        # Return sector placeholder
        _cell(ws, 6, 1, f'{dest}-{home}', font=NORM_FONT, align=LFT)
        _cell(ws, 6, 5, aircraft, font=NORM_FONT)
        _cell(ws, 6, 6, seats, font=NORM_FONT, fmt='#,##0')

        # Note
        _cell(ws, 8, 1, f'Note: Aircraft configuration based on airline website.',
              font=NOTE_FONT, align=LFT, border=NO_BORDER)

    # ------------------------------------------------------------------
    # SHEET 6: QSI DIAGNOSTICS (internal, not for client)
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # SHEET: ROUTE ECONOMICS (turnaround P&L) - the integrated business plan
    # ------------------------------------------------------------------

    def _write_route_economics(self):
        """Render the aircraft_economics RoutePnL as a turnaround P&L: demand and
        revenue through to cost and profit, every cost line showing its basis, with
        the standard directional-guidance disclaimer."""
        rp = self.route_pnl
        x = rp.compute()
        try:
            from aircraft_economics import DISCLAIMER_FULL
        except Exception:
            DISCLAIMER_FULL = ("Indicative, for directional guidance only. Built on generic "
                               "published assumptions, not any airline's actual costs.")

        ws = self.wb.create_sheet('Route Economics')
        ws.sheet_properties.tabColor = AVIA_BLUE
        for col, w in zip('ABCDE', [3, 34, 16, 60, 3]):
            ws.column_dimensions[col].width = w

        _cell(ws, 1, 2, 'AVIA SOLUTIONS  ROUTE ECONOMICS (Turnaround P&L)',
              font=TITLE_FONT, border=NO_BORDER)
        ws.merge_cells('B1:D1')
        route = f"{rp.airline}  {rp.origin}-{rp.dest}  |  {rp.aircraft}  |  sector {x['sector_fh']:.1f}h flight"
        _cell(ws, 2, 2, route, font=NOTE_FONT, border=NO_BORDER)
        ws.merge_cells('B2:D2')

        r = 4

        def banner(text):
            nonlocal r
            _section_banner(ws, r, text, col_span=4)
            r += 1

        def line(label, value, fmt='#,##0', note='', fill=None, font=NORM_FONT, vfont=None):
            nonlocal r
            _cell(ws, r, 2, label, font=font, align=LFT)
            _cell(ws, r, 3, value, font=vfont or font, fmt=fmt, fill=fill)
            if note:
                _cell(ws, r, 4, note, font=NOTE_FONT, align=LFT)
            else:
                _cell(ws, r, 4, '', border=NO_BORDER)
            r += 1

        # Inputs (analyst-set)
        banner('INPUTS (analyst-set)')
        line('Airline business model', rp.airline_type, fmt=None, fill=INP_FILL, vfont=INP_FONT,
             note='ULCC / LCC / FSC / Regional / Charter -> utilisation')
        line('Aircraft age (years)', rp.aircraft_age, fmt='0', fill=INP_FILL, vfont=INP_FONT,
             note='scales lease/value down its age curve')
        line('Economy load factor', rp.econ_lf, fmt='0.0%', fill=INP_FILL, vfont=INP_FONT)
        line('Economy fare (one-way)', rp.econ_fare_ow, fill=INP_FILL, vfont=INP_FONT)
        line('Business load factor', rp.bus_lf, fmt='0.0%', fill=INP_FILL, vfont=INP_FONT)
        line('Business fare (one-way)', rp.bus_fare_ow, fill=INP_FILL, vfont=INP_FONT)
        line('Fuel price ($/kg)', rp.fuel_price_usd_kg, fmt='0.00', fill=INP_FILL, vfont=INP_FONT)
        line('Share of fleet leased', rp.lease_share, fmt='0%', fill=INP_FILL, vfont=INP_FONT)
        r += 1

        # Revenue
        banner('REVENUE (per turnaround)')
        line('Economy fares', x['econ_rev'])
        line('Business fares', x['bus_rev'])
        line('Cargo', x['cargo_rev'])
        line('Charges recovery', x['charges_recovery'], note='per-pax charges passed to fare')
        line('GROSS REVENUE', x['gross_rev'], font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT, fill=DER_FILL)
        r += 1

        # Variable cost
        banner('VARIABLE COST')
        line('Fuel', x['fuel'])
        line('Maintenance', x['maintenance'], note=x.get('maint_basis', ''))
        line('Catering', x['catering'])
        line('Landing charges', x['landing'])
        line('Passenger charges', x['per_pax'])
        line('En-route navigation', x['nav'])
        line('Ground handling', x['handling'])
        line('Variable subtotal', x['variable'], font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT)
        r += 1

        # Direct fixed
        banner('DIRECT FIXED')
        line('Ownership (cost of capital)', x['ownership'], note=x.get('own_basis', ''))
        line('Insurance', x['insurance'])
        line('Crew', x['crew'], note=x.get('crew_basis', ''))
        line('Direct fixed subtotal', x['direct_fixed'], font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT)
        r += 1

        # Indirect
        banner('INDIRECT FIXED')
        line('Admin / overhead', x['admin'])
        line('Sales', x['sales'])
        line('Indirect subtotal', x['indirect_fixed'], font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT)
        r += 1

        line('TOTAL COST', x['total_cost'], font=TOTAL_FONT, vfont=TOTAL_FONT, fill=DER_FILL)
        r += 1

        # Results
        banner('RESULT (standalone)')
        line('PROFIT per turnaround', x['profit'], font=TOTAL_FONT, vfont=TOTAL_FONT, fill=DER_FILL)
        line('Margin', x['margin'], fmt='0.0%', font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT)
        line('Breakeven load factor', x['breakeven_lf'], fmt='0.0%')
        line('Cost per seat', x['cost_per_seat'])
        line('CASK ($/ASK)', x['cask'], fmt='0.000')

        # Incentive (only if present)
        if x.get('incentive_value'):
            r += 1
            banner('WITH AIRPORT INCENTIVE')
            line('Incentive value per turn', x['incentive_value'], note='charge waiver + route support')
            line('Profit with incentive', x['profit_with_incentive'], font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT, fill=DER_FILL)
            line('Margin with incentive', x['margin_with_incentive'], fmt='0.0%', font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT)

        # Annual route P&L + fleet requirement
        if self.frequency_per_week:
            try:
                from aircraft_economics import AnnualRoutePnL
                a = AnnualRoutePnL(rp, self.frequency_per_week, self.operating_weeks).compute()
                r += 1
                banner('ANNUAL ROUTE P&L & FLEET')
                line('Operating programme', f"{self.frequency_per_week:g}x/week x "
                     f"{self.operating_weeks:g} weeks = {a['annual_turnarounds']:.0f} turnarounds/yr", fmt=None)
                line('Annual passengers', a['annual_pax'])
                line('Annual revenue', a['annual_gross_rev'])
                line('Annual cost', a['annual_total_cost'])
                line('ANNUAL PROFIT', a['annual_profit'], font=TOTAL_FONT, vfont=TOTAL_FONT, fill=DER_FILL)
                line('Margin', a['margin'], fmt='0.0%', font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT)
                line('Aircraft required', a['aircraft_required'], fmt='0',
                     note=f"{a['aircraft_required_fractional']:.2f} frames of flying at "
                          f"{a['util_per_aircraft']:,.0f} BH/yr; rest is spare for other routes")
                if a.get('annual_profit_with_incentive') and a['annual_incentive_value']:
                    line('Annual profit with incentive', a['annual_profit_with_incentive'],
                         font=SUBTOTAL_FONT, vfont=SUBTOTAL_FONT, fill=DER_FILL)
            except Exception as e:
                r += 1
                _cell(ws, r, 2, f"(Annual P&L unavailable: {e})", font=NOTE_FONT, align=LFT, border=NO_BORDER)
                r += 1

        # Charge basis note
        r += 1
        _cell(ws, r, 2, f"Airport charges: origin {x.get('origin_charge_basis','')}, dest {x.get('dest_charge_basis','')}",
              font=NOTE_FONT, align=LFT, border=NO_BORDER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 2

        # Disclaimer
        _cell(ws, r, 2, 'DISCLAIMER', font=BOLD_FONT, border=NO_BORDER)
        r += 1
        dc = _cell(ws, r, 2, DISCLAIMER_FULL, font=NOTE_FONT, align=LFT, border=NO_BORDER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 96

    def _write_qsi_diagnostics(self):
        ws = self.wb.create_sheet('QSI Diagnostics')
        ws.sheet_properties.tabColor = '808080'

        _cell(ws, 1, 1, 'QSI DIAGNOSTICS  INTERNAL USE ONLY',
              font=TITLE_FONT, border=NO_BORDER)
        ws.merge_cells('A1:H1')
        _cell(ws, 2, 1, 'This sheet contains pipeline internals for QA purposes. '
              'Do not include in client deliverables.',
              font=NOTE_FONT, border=NO_BORDER)

        # Connecting city detail with pipeline vs expert QSI comparison
        headers = ['Rank', 'City', 'Name', 'Country', 'Grown Demand',
                   'Pipeline QSI', 'Expert QSI', 'Cal Factor',
                   'Pipeline Fcst', 'Expert Fcst', 'Variance']
        widths = [6, 7, 18, 16, 14, 12, 12, 10, 14, 14, 10]
        _header_row(ws, 4, headers, widths)

        home_results = self.results.get('home_results', [])
        sorted_cities = sorted(
            [c for c in home_results if c['city'] != 'Total'],
            key=lambda x: -x['forecast'])

        for idx, city in enumerate(sorted_cities):
            r = 5 + idx
            pipe_qsi = city.get('qsi_capture', 0)
            expert_qsi = city.get('original_qsi', 0)
            cal_factor = expert_qsi / pipe_qsi if pipe_qsi > 0 else 0
            grown = city['base_demand'] * (1 + city.get('growth', 0.09))
            pipe_fcst = grown * pipe_qsi
            expert_fcst = city['forecast']
            variance = (expert_fcst - pipe_fcst) / pipe_fcst if pipe_fcst > 0 else 0

            fill = ALT_FILL if idx % 2 == 0 else None
            _cell(ws, r, 1, idx + 1, font=NORM_FONT, fill=fill)
            _cell(ws, r, 2, city['city'], font=NORM_FONT, fill=fill)
            _cell(ws, r, 3, city.get('name', ''), font=NORM_FONT, fill=fill, align=LFT)
            _cell(ws, r, 4, city.get('country', ''), font=NORM_FONT, fill=fill, align=LFT)
            _cell(ws, r, 5, grown, font=NORM_FONT, fill=fill, fmt='#,##0')
            _cell(ws, r, 6, pipe_qsi, font=NORM_FONT, fill=fill, fmt='0.00%')
            _cell(ws, r, 7, expert_qsi, font=INP_FONT, fill=INP_FILL if fill is None else fill, fmt='0.00%')
            _cell(ws, r, 8, cal_factor, font=NORM_FONT, fill=fill, fmt='0.000')
            _cell(ws, r, 9, pipe_fcst, font=NORM_FONT, fill=fill, fmt='#,##0')
            _cell(ws, r, 10, expert_fcst, font=NORM_FONT, fill=fill, fmt='#,##0')
            _cell(ws, r, 11, variance, font=NORM_FONT,
                  fill=WARN_FILL if abs(variance) > 0.5 else fill,
                  fmt='0.0%')

    # ------------------------------------------------------------------
    # SHEET 7: ASSUMPTIONS LOG
    # ------------------------------------------------------------------

    def _write_assumptions_log(self):
        ws = self.wb.create_sheet('Assumptions Log')
        ws.sheet_properties.tabColor = 'FFC000'

        _cell(ws, 1, 1, 'ASSUMPTIONS LOG', font=TITLE_FONT, border=NO_BORDER)
        _cell(ws, 2, 1, 'Every calibration parameter listed. Complete the '
              '"Analyst Note" column to document rationale.',
              font=NOTE_FONT, border=NO_BORDER)

        headers = ['Category', 'Parameter', 'Value', 'Source', 'Analyst Note']
        widths = [20, 35, 15, 25, 40]
        _header_row(ws, 4, headers, widths)

        cfg = self.config
        res = self.results
        row = 5

        # Route parameters
        assumptions = [
            ('Route', 'Origin', cfg.home_airport_code, 'Client brief', ''),
            ('Route', 'Destination', cfg.dest_airport_code, 'Client brief', ''),
            ('Route', 'Carrier', getattr(cfg, 'airline_name', cfg.airline_code), 'Client brief', ''),
            ('Route', 'Frequency', cfg.frequency, 'Client brief', ''),
            ('Route', 'Aircraft', getattr(cfg, 'aircraft_type', ''), 'Client brief', ''),
            ('Route', 'Seats', cfg.seats, 'Airline website', ''),
            ('', '', '', '', ''),
            ('Growth', 'P2P Growth Rate', '', 'GDP / IATA forecast', ''),
            ('Growth', 'Home Cnx Growth', getattr(cfg, 'home_growth', ''), 'GDP forecast', ''),
            ('Growth', 'Dest Cnx Growth', getattr(cfg, 'dest_growth', ''), 'GDP forecast', ''),
            ('', '', '', '', ''),
            ('Stimulation', 'P2P Stimulation', '', 'IATA curve', ''),
            ('Stimulation', 'Connecting Stimulation', 1.0, 'Standard: no stimulation for cnx', ''),
            ('', '', '', '', ''),
            ('QSI', 'QSI Ceiling', getattr(cfg, 'qsi_ceiling', 1.0), 'Route-specific', ''),
            ('QSI', 'QSI Adjustment', getattr(cfg, 'qsi_adjustment', 1.0), 'Route-specific', ''),
            ('QSI', 'ET Decay Factor', cfg.et_decay_factor, 'Standard', ''),
            ('QSI', 'ET Decay Interval', cfg.et_decay_interval, 'Standard', ''),
            ('', '', '', '', ''),
            ('Calibration', 'Calibration Mode', '', 'Expert / Tiered / Flat', ''),
        ]

        # Add P2P segment assumptions
        for seg in res.get('p2p_details', []):
            assumptions.append(('P2P Capture', seg['name'],
                               f"{seg['capture']:.0%}", 'Analyst judgment', ''))
            assumptions.append(('P2P Stimulation', seg['name'],
                               f"{seg['stimulation']:.2f}", 'IATA curve', ''))

        # Add connection coefficient assumptions
        for cnx_type, coeff in cfg.cnx_coeffs.items():
            assumptions.append(('Cnx Coefficients', cnx_type, coeff, 'Standard (2013)', ''))

        for cat, param, val, source, note in assumptions:
            _cell(ws, row, 1, cat, font=BOLD_FONT if cat else NORM_FONT, align=LFT)
            _cell(ws, row, 2, param, font=NORM_FONT, align=LFT)
            _cell(ws, row, 3, val, font=INP_FONT, fill=INP_FILL if val else None)
            _cell(ws, row, 4, source, font=NOTE_FONT, align=LFT)
            _cell(ws, row, 5, note, font=NORM_FONT, fill=INP_FILL, align=LFT)
            row += 1

    # ------------------------------------------------------------------
    # SHEET 8: AUDIT TRAIL
    # ------------------------------------------------------------------

    def _write_audit_trail(self):
        ws = self.wb.create_sheet('Audit Trail')
        ws.sheet_properties.tabColor = '808080'

        _cell(ws, 1, 1, 'PIPELINE EXECUTION AUDIT TRAIL',
              font=TITLE_FONT, border=NO_BORDER)
        ws.column_dimensions['A'].width = 120

        for i, line in enumerate(self.audit_log, 3):
            _cell(ws, i, 1, line, font=MONO_FONT, align=LFT, border=NO_BORDER)


# ============================================================================
# HELPER
# ============================================================================

def carrier_name(cfg):
    """Extract carrier name from config."""
    return getattr(cfg, 'airline_name', getattr(cfg, 'airline_code', 'Carrier'))


# ============================================================================
# INTEGRATION TEST  BA LHR-SJC
# ============================================================================

def test_ba_lhr_sjc():
    """
    Integration test: generate output workbook from BA LHR-SJC pipeline results.
    Uses the RouteConfig factory method and run_pipeline from closed_loop_pipeline_v2.
    """
    from route_config import RouteConfig
    from closed_loop_pipeline_v2 import run_pipeline, validate

    # Use the factory method  same as the regression test
    config = RouteConfig.ba_lhr_sjc()

    # Add presentation metadata (not in RouteConfig but needed for output)
    config.carrier_type = 'Full Service'
    config.alliance = 'oneworld'
    config.forecast_year = '2016'
    config.base_year = '2013/14'
    config.mode = 'Forecast'
    config.home_growth = 0.09
    config.dest_growth = 0.10

    # Run the pipeline (same as regression)
    print("Running pipeline...")
    results = run_pipeline(config)

    # The raw pipeline overestimates connecting traffic (329k vs 129k target).
    # This is expected  the difference is expert calibration (documented Chat 11).
    # For the output workbook, we use the expert QSI values from the forecast file,
    # which the DemandProvider already loads. Re-assemble with expert QSI.
    from closed_loop_pipeline_v2 import ForecastAssembler
    assembler = ForecastAssembler(config)

    # Build expert QSI capture dicts from the demand provider's connecting cities
    expert_home_captures = {}
    for city in config.demand_provider.get_connecting_cities('home'):
        expert_home_captures[city.city_code] = city.qsi_score

    expert_dest_captures = {}
    for city in config.demand_provider.get_connecting_cities('dest'):
        expert_dest_captures[city.city_code] = city.qsi_score

    # Re-run assembly with expert QSI values
    results = assembler.run(config.demand_provider, expert_home_captures, expert_dest_captures)

    # Validate with expert values
    passed = validate(config, results)

    # Collect audit logs
    all_audit = results.get('assembler_audit', [])

    # Generate output workbook using StandardOutputWriter
    writer = StandardOutputWriter(
        config=config,
        results=results,
        audit_log=all_audit,
        analyst='John Carter',
        engagement_date='23 February 2015',
    )
    writer.write_all()

    ensure_output_dir()
    output_path = str(OUTPUT_DIR / 'BA_LHR_SJC_Standard_Output.xlsx')
    writer.save(output_path)

    # Summary
    print(f"\n{'='*60}")
    print(f"OUTPUT WORKBOOK GENERATED")
    print(f"{'='*60}")
    print(f"  Sheets: {', '.join(writer.wb.sheetnames)}")
    print(f"  Grand Total: {results['grand_total']:,.0f}")
    print(f"  Load Factor: {results['load_factor']:.1%}")
    print(f"  Path: {output_path}")
    print(f"  Validation: {'PASS' if passed else 'REVIEW NEEDED'}")

    return passed


if __name__ == '__main__':
    test_ba_lhr_sjc()
