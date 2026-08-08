#!/usr/bin/env python3
r"""Build aci.duckdb from the ACI monthly time series, so the tool can read it.

    py -3.12 load_aci.py --xlsx "<path to MONTH_ACI Monthly Time Series(NO).xlsx>"
    py -3.12 load_aci.py --xlsx ... --out C:\Avia\aci.duckdb --sheet "1-Monthly"
    py -3.12 load_aci.py --inspect --xlsx ...      # look first, load nothing

The source lives on Egnyte at
  /Shared/Company Data/02 Knowledge/3 Airports/_ACI/Analysis/ACI Time Series/
and is maintained by hand each month. This turns it into a store the engine and
the deck can query, on the same footing as oag.duckdb and sabre.duckdb.

WHAT THE SOURCE IS, from HowToUpdateTheMonthlyTimeSeries(ADF).docx
------------------------------------------------------------------
Three files exist and only ONE is loaded here:

    MONTH_...xlsx   passengers IN that month           <- this one
    YTD_...xlsx     year to date, a running total
    YE_...xlsx      year ending, a rolling twelve months

MONTH, YTD and YE describe the same traffic three ways. Loading more than one,
or summing across them, multiplies it. This is the same trap oag.duckdb sets
with its annual, monthly, half-year and weekly rows, and it is handled the same
way: pick the finest grain and never mix.

The sheet is WIDE. Airports are rows, months are columns, and a column is added
by hand every month. It is melted to one row per airport per month here, because
a wide sheet that grows a column a month is not a thing to query.

THE RULES THIS ENFORCES
-----------------------
1. A BLANK IS "DID NOT REPORT". IT IS NEVER ZERO. Not every airport reports to
   ACI, many report annually only, and the reporting set changes every month.
   Coercing a blank to zero would silently understate an airport's year, and the
   chart would show a collapse that never happened. Blanks are dropped and the
   coverage table records what was actually present.
2. COVERAGE IS RECORDED, NOT ASSUMED. `aci_coverage` gives months reported per
   airport per year, so a reader can apply the same completeness test the OAG
   reader applies rather than trusting a year to be whole.
3. THE VINTAGE IS STAMPED. Airports revise published figures, and the house
   procedure is to take an earlier year from a later file. `_store_meta` records
   which workbook this store was built from and when, so a number can always be
   traced to the file it came from.
4. THE MEASURE IS NAMED. ACI is TOTAL PASSENGER THROUGHPUT: arrivals plus
   departures plus transit, domestic and international together. It is not
   origin and destination and it is not one direction. `airport_profile.py`
   halves it before it meets a departing seat count, and refuses to treat it as
   origin and destination.

Avia Solutions Limited. All rights reserved.
"""

import argparse
import datetime as _dt
import os
import re
import sys

SHEET_DEFAULT = "1-Monthly"
MEASURE = "Total passengers, arrivals plus departures plus transit, dom + intl"
SOURCE = "ACI Worldwide Airport Traffic, monthly time series"

IATA_RE = re.compile(r"^[A-Z]{3}$")
# The month headers are written by hand over nearly twenty years, so they arrive
# as real dates, as "Jan-2012", as "Jan 12" and as text. All are accepted; a
# header that resolves to none of them is reported rather than guessed at.
_MONTHS = {m.lower(): i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}


def parse_month(v):
    """A column header to (year, month), or None if it is not a month at all."""
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.year, v.month
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4})[-/](\d{1,2})$", s)                 # 2024-06
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        return (y, mo) if 1 <= mo <= 12 else None
    m = re.match(r"^([A-Za-z]{3})[a-z]*[-\s/]+(\d{2,4})$", s)  # Jan-2012, Jan 12
    if m:
        mo = _MONTHS.get(m.group(1).lower())
        if not mo:
            return None
        y = int(m.group(2))
        if y < 100:
            y += 2000 if y < 70 else 1900
        return y, mo
    m = re.match(r"^(\d{1,2})[-/](\d{4})$", s)                 # 06-2024
    if m:
        mo, y = int(m.group(1)), int(m.group(2))
        return (y, mo) if 1 <= mo <= 12 else None
    return None


def _num(v):
    """A cell to a passenger count, or None. A blank is NEVER a zero."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None          # NaN guard
    s = str(v).strip().replace(",", "").replace(" ", "")
    if not s or s in ("-", "--", "n/a", "N/A", "na", "NA", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _text(row, col):
    """A text cell, or "" if the column was not found or the cell is empty."""
    if col is None or col >= len(row):
        return ""
    v = row[col]
    return str(v).strip() if isinstance(v, str) and str(v).strip() else ""


def read_sheet(xlsx, sheet=SHEET_DEFAULT, header_scan=12):
    """Melt the wide sheet to (iata, airport, country, ym, year, month, pax).

    Returns (rows, report). The report names every column that could not be read
    as a month and every row without a usable IATA code, because a loader that
    drops data quietly is worse than one that will not run.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit("sheet %r not in %s. Sheets: %s"
                         % (sheet, os.path.basename(xlsx), ", ".join(wb.sheetnames)))
    ws = wb[sheet]

    grid = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        grid.append(row)
        if i > 100000:
            break

    # Find the header row: the one with the most parseable month headers.
    best_i, best_cols = None, {}
    for i in range(min(header_scan, len(grid))):
        cols = {}
        for j, v in enumerate(grid[i]):
            ym = parse_month(v)
            if ym:
                cols[j] = ym
        if len(cols) > len(best_cols):
            best_i, best_cols = i, cols
    if not best_cols:
        raise SystemExit("no row in the first %d looked like month headers in %s"
                         % (header_scan, os.path.basename(xlsx)))

    # The IATA column, found on DISTINCTNESS rather than on count. The sheet's
    # first column is an ACI region, and EUR, NAM and ASP are three upper-case
    # letters too, so counting matches alone picks the region column and loads
    # six airports called EUR. Airport codes are near-unique down the sheet and
    # regions repeat, so the column with the most DISTINCT codes is the one.
    body = grid[best_i + 1:]
    ncol = max((len(r) for r in body[:500]), default=0)
    uniq = {}
    for j in range(min(ncol, 20)):
        if j in best_cols:
            continue
        seen = {r[j].strip().upper() for r in body[:5000]
                if j < len(r) and isinstance(r[j], str)
                and IATA_RE.match(r[j].strip().upper())}
        if seen:
            uniq[j] = len(seen)
    iata_col = max(uniq, key=uniq.get) if uniq else None
    if iata_col is None or uniq.get(iata_col, 0) < 3:
        raise SystemExit("could not find a column of IATA codes under the header "
                         "row in %s. Columns of three-letter codes and how many "
                         "distinct: %s" % (os.path.basename(xlsx),
                                           uniq or "none"))
    # The name and the country are picked as COLUMNS, on the same distinctness
    # test that found the code, not cell by cell. Taking the longest text on
    # each row looks reasonable and is wrong: to the left of a UK airport sits
    # "United Kingdom", which is longer than "Edinburgh", so every British
    # airport in the store came back named after its country. Down the sheet the
    # airport name is near-unique, the country repeats a few hundred times and
    # the ACI region repeats six times, so distinctness separates all three.
    text_cols = [j for j in range(iata_col) if j not in best_cols]
    text_uniq = {}
    for j in text_cols:
        vals = {str(r[j]).strip() for r in body[:5000]
                if j < len(r) and isinstance(r[j], str) and str(r[j]).strip()}
        if vals:
            text_uniq[j] = len(vals)
    name_col = max(text_uniq, key=text_uniq.get) if text_uniq else None
    rest = {j: n for j, n in text_uniq.items() if j != name_col}
    country_col = max(rest, key=rest.get) if rest else None
    # A worldwide country column holds of the order of a hundred values. ACI
    # has six regions. Eight is comfortably above the one and far below the
    # other, so a column thinner than that is the region and a region is not a
    # country: better to hold no country than a wrong one.
    if country_col is not None and rest.get(country_col, 0) < 8:
        country_col = None

    # An airport can appear on more than one row: the sheet is hand-maintained
    # and a rename or a re-coded airport gets a second line. Two rows that do
    # not overlap in time are complementary and summing them is right. Two rows
    # that both carry the same month DOUBLE that airport on every SUM, which no
    # reader downstream can detect. Both are counted here and the collision is
    # refused at build.
    rows, no_code, blanks = [], 0, 0
    code_rows, seen_cells, dup_cells, dup_detail = {}, set(), 0, {}
    for r in body:
        if iata_col >= len(r):
            continue
        code = (r[iata_col] or "")
        code = str(code).strip().upper()
        if not IATA_RE.match(code):
            if any(v is not None for v in r):
                no_code += 1
            continue
        name = _text(r, name_col)
        country = _text(r, country_col)
        code_rows[code] = code_rows.get(code, 0) + 1
        for j, (y, mo) in best_cols.items():
            if j >= len(r):
                continue
            pax = _num(r[j])
            if pax is None:            # DID NOT REPORT. Never written as zero.
                blanks += 1
                continue
            ym = "%04d-%02d" % (y, mo)
            if (code, ym) in seen_cells:
                dup_cells += 1
                dup_detail.setdefault(code, []).append(ym)
                continue
            seen_cells.add((code, ym))
            rows.append((code, name, country, ym, y, mo, pax))
    wb.close()
    repeated = sorted(c for c, n in code_rows.items() if n > 1)
    report = {"sheet": sheet, "header_row": best_i + 1, "month_columns": len(best_cols),
              "iata_column": iata_col + 1,
              "name_column": (name_col + 1) if name_col is not None else "none",
              "country_column": (country_col + 1) if country_col is not None else "none",
              "rows": len(rows),
              "rows_without_iata": no_code, "cells_not_reported": blanks,
              "codes_on_more_than_one_row": ", ".join(repeated) or "none",
              "colliding_cells": dup_cells,
              "first_month": "%04d-%02d" % min(best_cols.values()),
              "last_month": "%04d-%02d" % max(best_cols.values())}
    report["_collisions"] = {c: sorted(v) for c, v in dup_detail.items()}
    return rows, report


def build(xlsx, out, sheet=SHEET_DEFAULT, allow_collisions=False):
    import duckdb
    rows, report = read_sheet(xlsx, sheet)
    if not rows:
        raise SystemExit("no data rows read from %s" % xlsx)
    if report["colliding_cells"] and not allow_collisions:
        detail = "; ".join("%s: %s%s" % (c, ", ".join(v[:6]),
                                         " and %d more" % (len(v) - 6) if len(v) > 6 else "")
                           for c, v in sorted(report["_collisions"].items()))
        raise SystemExit(
            "%d cells appear twice for the same airport and month in %s.\n"
            "Two rows carry the same airport for the same month, so summing the\n"
            "airport doubles it. Resolve in the source, or re-run with\n"
            "--allow-collisions to keep the first value and drop the second.\n"
            "  %s" % (report["colliding_cells"], os.path.basename(xlsx), detail))
    folder = os.path.dirname(os.path.abspath(out))
    os.makedirs(folder, exist_ok=True)
    if os.path.exists(out):
        os.remove(out)
    con = duckdb.connect(out)
    con.execute("""CREATE TABLE aci_monthly(
        iata VARCHAR, airport VARCHAR, country VARCHAR, ym VARCHAR,
        year INTEGER, month INTEGER, passengers DOUBLE)""")
    con.executemany("INSERT INTO aci_monthly VALUES (?,?,?,?,?,?,?)", rows)
    # Coverage, so a reader can test a year for completeness instead of assuming.
    con.execute("""CREATE VIEW aci_coverage AS
        SELECT iata, year, COUNT(*) AS months_reported,
               SUM(passengers) AS passengers
        FROM aci_monthly GROUP BY iata, year""")
    con.execute("""CREATE TABLE _store_meta(
        key VARCHAR, value VARCHAR)""")
    meta = [("source", SOURCE), ("measure", MEASURE),
            ("source_file", os.path.basename(xlsx)),
            ("source_modified", _dt.datetime.fromtimestamp(
                os.path.getmtime(xlsx), _dt.timezone.utc).strftime("%Y-%m-%d")),
            ("built", _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("grain", "one row per airport per month; MONTH file only, never "
                      "YTD or YE, which restate the same traffic"),
            ("blanks", "a month with no figure is NOT in this table. A blank in "
                       "the source means the airport did not report, not zero"),
            ("rows", str(len(rows)))]
    meta += [(k, str(v)) for k, v in report.items() if not k.startswith("_")]
    con.executemany("INSERT INTO _store_meta VALUES (?,?)", meta)
    con.close()
    return report


def inspect(xlsx, sheet=SHEET_DEFAULT):
    rows, report = read_sheet(xlsx, sheet)
    print("READ  %s" % os.path.basename(xlsx))
    for k in ("sheet", "header_row", "iata_column", "month_columns",
              "first_month", "last_month", "rows", "rows_without_iata",
              "cells_not_reported", "codes_on_more_than_one_row",
              "colliding_cells"):
        print("   %-20s %s" % (k, report[k]))
    for code, months in sorted(report["_collisions"].items()):
        print("   COLLISION %s carries %d months twice: %s%s"
              % (code, len(months), ", ".join(months[:8]),
                 " ..." if len(months) > 8 else ""))
    apts = sorted({r[0] for r in rows})
    print("   %-20s %d" % ("airports", len(apts)))
    print("   %-20s %s" % ("sample", ", ".join(apts[:12])))
    for code in ("EDI", "AUS", "LHR"):
        got = sorted((r[3], r[5]) for r in rows if r[0] == code)
        if got:
            ys = {}
            for y, p in got:
                ys[y] = ys.get(y, 0) + p
            span = "%d-%d" % (min(ys), max(ys))
            print("   %-20s %s, %d months, %s, latest full-ish year %s = %s"
                  % (code, span, len(got), "reports",
                     max(ys), "{:,.0f}".format(ys[max(ys)])))
        else:
            print("   %-20s not present" % code)
    return report


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help="the MONTH_ ACI time series workbook")
    ap.add_argument("--sheet", default=SHEET_DEFAULT)
    ap.add_argument("--out", default="", help="output store; default aci.duckdb "
                                              "beside the other stores via config")
    ap.add_argument("--inspect", action="store_true",
                    help="report what would be read and write nothing")
    ap.add_argument("--allow-collisions", action="store_true",
                    help="build even where one airport carries the same month "
                         "twice, keeping the first value and dropping the second")
    a = ap.parse_args()
    if not os.path.exists(a.xlsx):
        raise SystemExit("not found: %s" % a.xlsx)
    if a.inspect:
        inspect(a.xlsx, a.sheet)
        return
    out, companions = a.out, []
    if not out:
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import config as CFG
            out = str(getattr(CFG, "ACI_DUCKDB", "")) or os.path.join(
                os.path.dirname(str(CFG.OAG_DUCKDB)), "aci.duckdb")
            companions = [str(CFG.OAG_DUCKDB), str(CFG.SABRE_DUCKDB)]
        except Exception:
            out = "aci.duckdb"
    rep = build(a.xlsx, out, a.sheet, allow_collisions=a.allow_collisions)
    print("BUILT %s" % out)
    for k, v in rep.items():
        if not k.startswith("_"):
            print("   %-20s %s" % (k, v))
    # The folder is created if absent, which is how a store ends up somewhere the
    # tool never reads. If the other stores are not beside it, say so plainly
    # rather than reporting a clean build into an empty folder.
    missing = [c for c in companions if not os.path.exists(c)]
    if missing:
        print("\nNOTE  the other stores are NOT in this folder:")
        for c in missing:
            print("        absent: %s" % c)
        print("      Either this machine holds its stores elsewhere, in which case"
              "\n      set AVIA_LOCAL_CACHE and rebuild, or aci.duckdb has just been"
              "\n      written where nothing will look for it.")
    elif companions:
        print("\n      beside oag.duckdb and sabre.duckdb, as intended.")


if __name__ == "__main__":
    main()
