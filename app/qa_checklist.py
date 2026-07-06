"""
QA Checklist Automation Module  Avia Solutions
================================================

Automates the Q&A Checklist templates (Q_A_Checklist_TEMPLATE.docx versions 1-3)
by evaluating pipeline results against Avia Solutions' quality control standards.

Three checklist sections are automated:
  1. USER CHECKLIST  12 workflow steps (file prep  pitch preparation)
     Tracks which pipeline steps have been completed vs still manual.
  2. TECHNICAL QA  8 expert-review checks (consistency, LF, capture rates, etc.)
     Auto-evaluates quantitative checks; flags qualitative ones for reviewer.
  3. REVIEWER FEEDBACK  Sign-off tracking for technical and grammar review.

Each check produces a QAResult with:
  - status: PASS / WARN / FAIL / MANUAL (needs human review)
  - detail: Explanation of finding
  - evidence: Specific numbers/data supporting the finding

Integration:
  from qa_checklist import run_qa_checklist, QAReport
  report = run_qa_checklist(config, pipeline_results, previous_forecast=None)
  report.summary()          # Print summary
  report.to_excel(path)     # Branded Excel output
  report.to_dict()          # Machine-readable
"""

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any, Tuple
from enum import Enum
from datetime import datetime
import math


# ============================================================================
# ENUMS AND DATA CLASSES
# ============================================================================

class CheckStatus(Enum):
    PASS = "PASS"
    WARN = "WARNING"
    FAIL = "FAIL"
    MANUAL = "MANUAL REVIEW"
    SKIP = "SKIPPED"


class CheckCategory(Enum):
    WORKFLOW = "User Checklist"
    TECHNICAL = "Technical QA"
    REVIEWER = "Reviewer Feedback"


@dataclass
class QAResult:
    """Single Q&A check result."""
    check_id: str
    category: CheckCategory
    title: str
    status: CheckStatus
    detail: str
    evidence: str = ""
    reviewer_notes: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            'check_id': self.check_id,
            'category': self.category.value,
            'title': self.title,
            'status': self.status.value,
            'detail': self.detail,
            'evidence': self.evidence,
            'reviewer_notes': self.reviewer_notes,
        }


@dataclass
class QAReport:
    """Complete Q&A checklist report."""
    route: str
    airline: str
    analyst: str
    date: str
    results: List[QAResult] = field(default_factory=list)
    overall_status: CheckStatus = CheckStatus.PASS

    def add(self, result: QAResult):
        self.results.append(result)
        # Escalate overall status
        if result.status == CheckStatus.FAIL:
            self.overall_status = CheckStatus.FAIL
        elif result.status == CheckStatus.WARN and self.overall_status != CheckStatus.FAIL:
            self.overall_status = CheckStatus.WARN
        elif result.status == CheckStatus.MANUAL and self.overall_status == CheckStatus.PASS:
            self.overall_status = CheckStatus.MANUAL

    @property
    def pass_count(self) -> int:
        return sum(1 for r in self.results if r.status == CheckStatus.PASS)

    @property
    def warn_count(self) -> int:
        return sum(1 for r in self.results if r.status == CheckStatus.WARN)

    @property
    def fail_count(self) -> int:
        return sum(1 for r in self.results if r.status == CheckStatus.FAIL)

    @property
    def manual_count(self) -> int:
        return sum(1 for r in self.results if r.status == CheckStatus.MANUAL)

    @property
    def skip_count(self) -> int:
        return sum(1 for r in self.results if r.status == CheckStatus.SKIP)

    def by_category(self, cat: CheckCategory) -> List[QAResult]:
        return [r for r in self.results if r.category == cat]

    def summary(self) -> str:
        lines = [
            f"QA CHECKLIST REPORT: {self.airline} {self.route}",
            f"Date: {self.date}  |  Analyst: {self.analyst}",
            f"Overall: {self.overall_status.value}",
            f"Results: {self.pass_count} PASS | {self.warn_count} WARN | "
            f"{self.fail_count} FAIL | {self.manual_count} MANUAL | {self.skip_count} SKIP",
            "",
        ]
        for cat in CheckCategory:
            items = self.by_category(cat)
            if items:
                lines.append(f"--- {cat.value} ---")
                for r in items:
                    icon = {"PASS": "", "WARNING": "", "FAIL": "",
                            "MANUAL REVIEW": "", "SKIPPED": ""}.get(r.status.value, "?")
                    lines.append(f"  {icon} [{r.check_id}] {r.title}: {r.status.value}")
                    if r.detail:
                        lines.append(f"      {r.detail}")
                lines.append("")
        return "\n".join(lines)

    def to_dict(self) -> Dict[str, Any]:
        return {
            'route': self.route,
            'airline': self.airline,
            'analyst': self.analyst,
            'date': self.date,
            'overall_status': self.overall_status.value,
            'counts': {
                'pass': self.pass_count,
                'warn': self.warn_count,
                'fail': self.fail_count,
                'manual': self.manual_count,
                'skip': self.skip_count,
            },
            'results': [r.to_dict() for r in self.results],
        }

    def to_excel(self, path: str) -> str:
        """Write branded Excel Q&A checklist report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl")

        wb = Workbook()

        # Colours
        AVIA_BLUE = '1B3A5C'
        AVIA_GOLD = 'C9A84C'
        PASS_GREEN = 'C6EFCE'
        WARN_AMBER = 'FFEB9C'
        FAIL_RED = 'FFC7CE'
        MANUAL_BLUE = 'D6E4F0'
        HEADER_FILL = PatternFill(start_color=AVIA_BLUE, end_color=AVIA_BLUE, fill_type='solid')
        HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        TITLE_FONT = Font(name='Calibri', size=14, bold=True, color=AVIA_BLUE)
        BODY_FONT = Font(name='Calibri', size=10)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        status_fills = {
            'PASS': PatternFill(start_color=PASS_GREEN, end_color=PASS_GREEN, fill_type='solid'),
            'WARNING': PatternFill(start_color=WARN_AMBER, end_color=WARN_AMBER, fill_type='solid'),
            'FAIL': PatternFill(start_color=FAIL_RED, end_color=FAIL_RED, fill_type='solid'),
            'MANUAL REVIEW': PatternFill(start_color=MANUAL_BLUE, end_color=MANUAL_BLUE, fill_type='solid'),
            'SKIPPED': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        }

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = 'QA Summary'
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 30

        ws.cell(row=1, column=1, value=f'QA Checklist: {self.airline} {self.route}').font = TITLE_FONT
        ws.cell(row=2, column=1, value=f'Date: {self.date}').font = BODY_FONT
        ws.cell(row=2, column=2, value=f'Analyst: {self.analyst}').font = BODY_FONT
        ws.cell(row=3, column=1, value=f'Overall: {self.overall_status.value}').font = Font(
            name='Calibri', size=12, bold=True,
            color='006100' if self.overall_status == CheckStatus.PASS else
            '9C5700' if self.overall_status == CheckStatus.WARN else
            '9C0006' if self.overall_status == CheckStatus.FAIL else AVIA_BLUE
        )

        row = 5
        headers = ['Check ID', 'Check', 'Status', 'Detail', 'Evidence']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal='center')

        for cat in CheckCategory:
            items = self.by_category(cat)
            if items:
                row += 1
                c = ws.cell(row=row, column=1, value=cat.value)
                c.font = Font(name='Calibri', size=11, bold=True, color=AVIA_BLUE)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

                for r in items:
                    row += 1
                    ws.cell(row=row, column=1, value=r.check_id).font = BODY_FONT
                    ws.cell(row=row, column=2, value=r.title).font = BODY_FONT
                    status_cell = ws.cell(row=row, column=3, value=r.status.value)
                    status_cell.font = Font(name='Calibri', size=10, bold=True)
                    status_cell.fill = status_fills.get(r.status.value, PatternFill())
                    status_cell.alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=4, value=r.detail).font = BODY_FONT
                    ws.cell(row=row, column=5, value=r.evidence).font = BODY_FONT
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).border = thin_border

        # --- Sheet 2: Reviewer Sign-off ---
        ws2 = wb.create_sheet('Reviewer Sign-off')
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 30
        ws2.column_dimensions['C'].width = 25

        ws2.cell(row=1, column=1, value='Reviewer Feedback').font = TITLE_FONT
        ws2.cell(row=3, column=1, value='Q&A Check').font = HEADER_FONT
        ws2.cell(row=3, column=1).fill = HEADER_FILL
        ws2.cell(row=3, column=2, value="Reviewer's Name").font = HEADER_FONT
        ws2.cell(row=3, column=2).fill = HEADER_FILL
        ws2.cell(row=3, column=3, value='Signed and Dated').font = HEADER_FONT
        ws2.cell(row=3, column=3).fill = HEADER_FILL

        ws2.cell(row=4, column=1, value='Technical Forecast Proof').font = BODY_FONT
        ws2.cell(row=5, column=1, value='General Spelling & Grammar Proof').font = BODY_FONT
        for r in range(4, 6):
            for c in range(1, 4):
                ws2.cell(row=r, column=c).border = thin_border

        wb.save(path)
        return path


# ============================================================================
# BOUNDS AND THRESHOLDS  calibrated from 22-case library
# ============================================================================

# Load factor bounds by route type
LF_BOUNDS = {
    'new_route': (0.55, 0.90),       # New routes: 55-90% acceptable
    'existing': (0.60, 0.95),         # Existing routes: 60-95%
    'default': (0.55, 0.95),
}

# P2P capture rate bounds (% of base P2P demand)
CAPTURE_BOUNDS = {
    'new_route_competitive': (0.15, 0.40),   # New route into competitive market
    'new_route_underserved': (0.20, 0.55),   # New route, underserved market
    'existing': (0.10, 0.35),                 # Existing route
    'default': (0.10, 0.55),
}

# Stimulation factor bounds
STIM_BOUNDS = {
    'lcc': (1.05, 1.50),           # LCCs drive new demand
    'legacy': (1.00, 1.30),        # Legacy carriers capture existing
    'zero_metro_direct': (1.30, 1.55),  # No hub-metro direct service
    'default': (1.00, 1.50),
}

# Growth rate bounds (annual %)
GROWTH_BOUNDS = {
    'mature': (0.01, 0.05),        # Mature market: 1-5%
    'developing': (0.03, 0.10),    # Developing market: 3-10%
    'default': (0.01, 0.10),
}

# Connecting traffic share bounds
CONNECTING_SHARE_BOUNDS = {
    'hub_route': (0.30, 0.80),     # Hub routes: 30-80% connecting
    'p2p_route': (0.00, 0.30),     # Point-to-point: 0-30%
    'default': (0.00, 0.80),
}

# QSI adjustment factor bounds (from calibration library)
QSI_ADJ_BOUNDS = {
    'new_route': (0.80, 1.20),     # New routes: near 1.0
    'existing': (0.15, 0.50),      # Existing routes: heavy calibration needed
    'default': (0.15, 1.20),
}


# ============================================================================
# WORKFLOW CHECKS (User Checklist  12 steps from templates)
# ============================================================================

def _check_workflow(config, results: Dict[str, Any],
                    data_sources: Optional[Dict[str, bool]] = None) -> List[QAResult]:
    """
    Evaluate the 12-step workflow checklist from Q_A_Checklist_TEMPLATE.
    Maps each manual step to its automated equivalent.
    """
    checks = []
    ds = data_sources or {}

    # W01: File/folder preparation
    checks.append(QAResult(
        check_id='W01', category=CheckCategory.WORKFLOW,
        title='File & Folder Preparation',
        status=CheckStatus.PASS,
        detail='Pipeline manages all file structures automatically.',
        evidence='RouteConfig + provider abstraction handles data paths.'
    ))

    # W02: Airline info
    airline = getattr(config, 'airline', None) or results.get('airline', '')
    carrier = getattr(config, 'carrier_type', None) or results.get('carrier_type', '')
    if airline and carrier:
        checks.append(QAResult(
            check_id='W02', category=CheckCategory.WORKFLOW,
            title='Airline Information',
            status=CheckStatus.PASS,
            detail=f'Airline: {airline}, Type: {carrier}',
            evidence=f'From RouteConfig: {config.__class__.__name__}'
        ))
    else:
        checks.append(QAResult(
            check_id='W02', category=CheckCategory.WORKFLOW,
            title='Airline Information',
            status=CheckStatus.WARN,
            detail='Airline or carrier type not specified in config.',
        ))

    # W03: Data preparation (OAG city lists, connecting markets)
    has_oag = ds.get('oag_schedule', False) or 'schedule_provider' in str(type(getattr(config, 'schedule_provider', '')))
    has_midt = ds.get('midt_demand', False) or results.get('p2p_demand') is not None
    if has_oag and has_midt:
        checks.append(QAResult(
            check_id='W03', category=CheckCategory.WORKFLOW,
            title='Data Preparation (OAG + MIDT/Sabre)',
            status=CheckStatus.PASS,
            detail='Both OAG schedule and MIDT demand data loaded.',
        ))
    elif has_oag or has_midt:
        missing = 'MIDT/Sabre demand' if not has_midt else 'OAG schedule'
        checks.append(QAResult(
            check_id='W03', category=CheckCategory.WORKFLOW,
            title='Data Preparation (OAG + MIDT/Sabre)',
            status=CheckStatus.WARN,
            detail=f'Missing: {missing}. Partial data only.',
        ))
    else:
        checks.append(QAResult(
            check_id='W03', category=CheckCategory.WORKFLOW,
            title='Data Preparation (OAG + MIDT/Sabre)',
            status=CheckStatus.MANUAL,
            detail='Data source status unknown  verify OAG and MIDT inputs.',
        ))

    # W04: Market demand (Sabre O&D, P2P, connecting)
    p2p = results.get('p2p_demand') or results.get('p2p_base')
    connecting = results.get('connecting_total') or results.get('total_connecting')
    if p2p is not None and connecting is not None:
        checks.append(QAResult(
            check_id='W04', category=CheckCategory.WORKFLOW,
            title='Market Demand Analysis',
            status=CheckStatus.PASS,
            detail=f'P2P demand: {p2p:,.0f}, Connecting: {connecting:,.0f}',
            evidence='From pipeline results.'
        ))
    else:
        checks.append(QAResult(
            check_id='W04', category=CheckCategory.WORKFLOW,
            title='Market Demand Analysis',
            status=CheckStatus.MANUAL,
            detail='P2P and/or connecting demand not found in results.',
        ))

    # W05: Proposed schedule
    freq = getattr(config, 'weekly_frequency', None)
    aircraft = getattr(config, 'aircraft_type', '') or getattr(config, 'aircraft', '')
    seats = getattr(config, 'seats', None) or getattr(config, 'seat_capacity', None)
    if freq and seats:
        checks.append(QAResult(
            check_id='W05', category=CheckCategory.WORKFLOW,
            title='Proposed Schedule',
            status=CheckStatus.PASS,
            detail=f'{freq}x weekly, {aircraft or "unspecified"} ({seats} seats)',
        ))
    else:
        checks.append(QAResult(
            check_id='W05', category=CheckCategory.WORKFLOW,
            title='Proposed Schedule',
            status=CheckStatus.WARN,
            detail='Frequency or seat capacity not fully specified.',
        ))

    # W06: OAG schedules data
    itineraries = results.get('itinerary_count') or results.get('total_itineraries')
    if itineraries:
        checks.append(QAResult(
            check_id='W06', category=CheckCategory.WORKFLOW,
            title='OAG Schedules Data',
            status=CheckStatus.PASS,
            detail=f'{itineraries:,} itineraries parsed from OAG data.',
        ))
    else:
        checks.append(QAResult(
            check_id='W06', category=CheckCategory.WORKFLOW,
            title='OAG Schedules Data',
            status=CheckStatus.MANUAL,
            detail='Itinerary count not available  verify OAG data loaded.',
        ))

    # W07: Connection builder
    connections = results.get('connections_built') or results.get('connection_count')
    if connections:
        checks.append(QAResult(
            check_id='W07', category=CheckCategory.WORKFLOW,
            title='Connection Builder',
            status=CheckStatus.PASS,
            detail=f'{connections:,} valid connections built.',
        ))
    else:
        checks.append(QAResult(
            check_id='W07', category=CheckCategory.WORKFLOW,
            title='Connection Builder',
            status=CheckStatus.MANUAL,
            detail='Connection count not available in results.',
        ))

    # W08: QSI model
    qsi_markets = results.get('qsi_market_count') or results.get('markets_scored')
    if qsi_markets:
        checks.append(QAResult(
            check_id='W08', category=CheckCategory.WORKFLOW,
            title='QSI Model',
            status=CheckStatus.PASS,
            detail=f'{qsi_markets} markets scored in QSI model.',
        ))
    else:
        checks.append(QAResult(
            check_id='W08', category=CheckCategory.WORKFLOW,
            title='QSI Model',
            status=CheckStatus.MANUAL,
            detail='QSI market count not available  verify QSI run.',
        ))

    # W09: Model outputs integration
    total_pax = results.get('total_annual_pax') or results.get('grand_total')
    if total_pax:
        checks.append(QAResult(
            check_id='W09', category=CheckCategory.WORKFLOW,
            title='Model Outputs',
            status=CheckStatus.PASS,
            detail=f'Total annual pax: {total_pax:,.0f}',
        ))
    else:
        checks.append(QAResult(
            check_id='W09', category=CheckCategory.WORKFLOW,
            title='Model Outputs',
            status=CheckStatus.WARN,
            detail='Total pax not found in results.',
        ))

    # W10: Forecast tables
    checks.append(QAResult(
        check_id='W10', category=CheckCategory.WORKFLOW,
        title='Forecast Tables',
        status=CheckStatus.PASS if total_pax else CheckStatus.MANUAL,
        detail='Output workbook generates standardised forecast tables.' if total_pax
               else 'Cannot generate forecast tables without total pax.',
    ))

    # W11: Calibration (for existing routes)
    calibrated = results.get('calibrated', False) or results.get('qsi_adj_factor') is not None
    checks.append(QAResult(
        check_id='W11', category=CheckCategory.WORKFLOW,
        title='Calibration',
        status=CheckStatus.PASS if calibrated else CheckStatus.MANUAL,
        detail='Pipeline calibration applied.' if calibrated
               else 'Calibration status unknown  check if route requires P2P calibration.',
    ))

    # W12: Pitch preparation
    checks.append(QAResult(
        check_id='W12', category=CheckCategory.WORKFLOW,
        title='Pitch Preparation',
        status=CheckStatus.MANUAL,
        detail='Presentation assembly requires manual review and image selection.',
        evidence='Automated output workbook available; pitch deck needs human curation.'
    ))

    return checks


# ============================================================================
# TECHNICAL QA CHECKS (8 checks from template)
# ============================================================================

def _check_technical(config, results: Dict[str, Any],
                     previous_forecast: Optional[Dict[str, Any]] = None,
                     calibration_library: Optional[Dict] = None) -> List[QAResult]:
    """
    Technical Q&A checks from Q_A_Checklist_TEMPLATE.
    8 checks matching the reviewer checklist.
    """
    checks = []

    #  T01: Consistency with previous forecast 
    if previous_forecast:
        prev_pax = previous_forecast.get('total_annual_pax', 0)
        curr_pax = results.get('total_annual_pax') or results.get('grand_total') or 0
        if prev_pax > 0 and curr_pax > 0:
            delta = (curr_pax - prev_pax) / prev_pax
            if abs(delta) < 0.05:
                checks.append(QAResult(
                    check_id='T01', category=CheckCategory.TECHNICAL,
                    title='Consistency with Previous Forecast',
                    status=CheckStatus.PASS,
                    detail=f'Within 5% of previous ({delta:+.1%}).',
                    evidence=f'Previous: {prev_pax:,.0f}, Current: {curr_pax:,.0f}'
                ))
            elif abs(delta) < 0.15:
                checks.append(QAResult(
                    check_id='T01', category=CheckCategory.TECHNICAL,
                    title='Consistency with Previous Forecast',
                    status=CheckStatus.WARN,
                    detail=f'Differs by {delta:+.1%} from previous  review assumptions.',
                    evidence=f'Previous: {prev_pax:,.0f}, Current: {curr_pax:,.0f}'
                ))
            else:
                checks.append(QAResult(
                    check_id='T01', category=CheckCategory.TECHNICAL,
                    title='Consistency with Previous Forecast',
                    status=CheckStatus.FAIL,
                    detail=f'Differs by {delta:+.1%} from previous  requires explanation.',
                    evidence=f'Previous: {prev_pax:,.0f}, Current: {curr_pax:,.0f}'
                ))
        else:
            checks.append(QAResult(
                check_id='T01', category=CheckCategory.TECHNICAL,
                title='Consistency with Previous Forecast',
                status=CheckStatus.MANUAL,
                detail='Previous forecast pax data incomplete  manual comparison needed.',
            ))
    else:
        checks.append(QAResult(
            check_id='T01', category=CheckCategory.TECHNICAL,
            title='Consistency with Previous Forecast',
            status=CheckStatus.SKIP,
            detail='No previous forecast provided for comparison.',
        ))

    #  T02: Charts and tables consistency 
    # Check internal consistency: total = P2P + connecting
    total = results.get('total_annual_pax') or results.get('grand_total')
    p2p = results.get('p2p_pax') or results.get('p2p_captured')
    cnx = results.get('connecting_pax') or results.get('total_connecting_pax')
    if total and p2p is not None and cnx is not None:
        computed_total = p2p + cnx
        diff = abs(total - computed_total)
        if diff < 2:  # Allow rounding
            checks.append(QAResult(
                check_id='T02', category=CheckCategory.TECHNICAL,
                title='Internal Consistency (P2P + Connecting = Total)',
                status=CheckStatus.PASS,
                detail=f'P2P ({p2p:,.0f}) + Connecting ({cnx:,.0f}) = {computed_total:,.0f}  Total ({total:,.0f})',
            ))
        else:
            checks.append(QAResult(
                check_id='T02', category=CheckCategory.TECHNICAL,
                title='Internal Consistency (P2P + Connecting = Total)',
                status=CheckStatus.FAIL,
                detail=f'Mismatch: P2P ({p2p:,.0f}) + Connecting ({cnx:,.0f}) = {computed_total:,.0f}  Total ({total:,.0f})',
                evidence=f'Difference: {diff:,.0f} pax'
            ))
    else:
        checks.append(QAResult(
            check_id='T02', category=CheckCategory.TECHNICAL,
            title='Internal Consistency (P2P + Connecting = Total)',
            status=CheckStatus.MANUAL,
            detail='Cannot verify  P2P and/or connecting breakdown not in results.',
        ))

    #  T03: Load factor sensibility 
    seats = getattr(config, 'seats', None) or getattr(config, 'seat_capacity', None) or 0
    freq = getattr(config, 'weekly_frequency', None) or 0
    if total and seats and freq:
        annual_seats = seats * freq * 52 * 2  # Both directions
        lf = total / annual_seats if annual_seats > 0 else 0
        route_type = 'new_route' if getattr(config, 'is_new_route', True) else 'existing'
        lo, hi = LF_BOUNDS.get(route_type, LF_BOUNDS['default'])

        if lo <= lf <= hi:
            checks.append(QAResult(
                check_id='T03', category=CheckCategory.TECHNICAL,
                title='Load Factor Sensibility',
                status=CheckStatus.PASS,
                detail=f'LF: {lf:.1%} (acceptable range: {lo:.0%}-{hi:.0%})',
                evidence=f'{total:,.0f} pax / {annual_seats:,.0f} annual seats'
            ))
        elif lf > hi:
            checks.append(QAResult(
                check_id='T03', category=CheckCategory.TECHNICAL,
                title='Load Factor Sensibility',
                status=CheckStatus.WARN if lf < 0.95 else CheckStatus.FAIL,
                detail=f'LF: {lf:.1%} exceeds upper bound ({hi:.0%}). Risk of capacity spill.',
                evidence=f'{total:,.0f} pax / {annual_seats:,.0f} annual seats'
            ))
        else:
            checks.append(QAResult(
                check_id='T03', category=CheckCategory.TECHNICAL,
                title='Load Factor Sensibility',
                status=CheckStatus.WARN,
                detail=f'LF: {lf:.1%} below lower bound ({lo:.0%}). May indicate weak demand case.',
                evidence=f'{total:,.0f} pax / {annual_seats:,.0f} annual seats'
            ))
    else:
        checks.append(QAResult(
            check_id='T03', category=CheckCategory.TECHNICAL,
            title='Load Factor Sensibility',
            status=CheckStatus.MANUAL,
            detail='Cannot compute LF  missing seats, frequency, or total pax.',
        ))

    #  T04: P2P capture rate and stimulation 
    p2p_base = results.get('p2p_demand') or results.get('p2p_base')
    capture_rate = results.get('capture_rate') or results.get('p2p_capture_rate')
    stim = results.get('stimulation') or results.get('stim_factor')

    if capture_rate is not None:
        lo, hi = CAPTURE_BOUNDS['default']
        if lo <= capture_rate <= hi:
            status = CheckStatus.PASS
            detail = f'P2P capture rate: {capture_rate:.1%} (range: {lo:.0%}-{hi:.0%})'
        elif capture_rate > hi:
            status = CheckStatus.WARN
            detail = f'P2P capture rate: {capture_rate:.1%} exceeds {hi:.0%}  may be aggressive.'
        else:
            status = CheckStatus.WARN
            detail = f'P2P capture rate: {capture_rate:.1%} below {lo:.0%}  may be conservative.'

        evidence_parts = []
        if p2p_base: evidence_parts.append(f'Base P2P: {p2p_base:,.0f}')
        if p2p: evidence_parts.append(f'Captured P2P: {p2p:,.0f}')
        if stim: evidence_parts.append(f'Stimulation: {stim:.2f}x')

        checks.append(QAResult(
            check_id='T04', category=CheckCategory.TECHNICAL,
            title='P2P Capture Rate & Stimulation',
            status=status, detail=detail,
            evidence=', '.join(evidence_parts)
        ))
    elif p2p_base and p2p:
        # Compute capture rate
        cr = p2p / p2p_base if p2p_base > 0 else 0
        checks.append(QAResult(
            check_id='T04', category=CheckCategory.TECHNICAL,
            title='P2P Capture Rate & Stimulation',
            status=CheckStatus.PASS if 0.10 <= cr <= 0.55 else CheckStatus.WARN,
            detail=f'Implied capture: {cr:.1%} ({p2p:,.0f} / {p2p_base:,.0f})',
            evidence=f'Stimulation: {stim:.2f}x' if stim else 'Stimulation factor not specified'
        ))
    else:
        checks.append(QAResult(
            check_id='T04', category=CheckCategory.TECHNICAL,
            title='P2P Capture Rate & Stimulation',
            status=CheckStatus.MANUAL,
            detail='Capture rate data not available  manual review required.',
        ))

    #  T05: Connecting demand proportion 
    if total and cnx is not None:
        cnx_share = cnx / total if total > 0 else 0
        lo, hi = CONNECTING_SHARE_BOUNDS['default']
        if lo <= cnx_share <= hi:
            status = CheckStatus.PASS
        elif cnx_share > hi:
            status = CheckStatus.WARN
        else:
            status = CheckStatus.PASS  # Low connecting is fine for P2P routes

        checks.append(QAResult(
            check_id='T05', category=CheckCategory.TECHNICAL,
            title='Connecting Demand Proportion',
            status=status,
            detail=f'Connecting share: {cnx_share:.1%} ({cnx:,.0f} of {total:,.0f})',
            evidence='Check consistency with known direct services, alliances, and codeshares.'
        ))
    else:
        checks.append(QAResult(
            check_id='T05', category=CheckCategory.TECHNICAL,
            title='Connecting Demand Proportion',
            status=CheckStatus.MANUAL,
            detail='Connecting traffic proportion not available.',
        ))

    #  T06: Major connection destinations 
    top_markets = results.get('top_connecting_markets') or results.get('top_markets')
    if top_markets and isinstance(top_markets, (list, dict)):
        if isinstance(top_markets, dict):
            top_list = sorted(top_markets.items(), key=lambda x: x[1], reverse=True)[:5]
            market_str = ', '.join(f'{m}: {v:,.0f}' for m, v in top_list)
        else:
            market_str = ', '.join(str(m) for m in top_markets[:5])

        checks.append(QAResult(
            check_id='T06', category=CheckCategory.TECHNICAL,
            title='Major Connection Destinations',
            status=CheckStatus.MANUAL,
            detail=f'Top markets: {market_str}',
            evidence='Review: does the ranking look sensible given network geography?'
        ))
    else:
        checks.append(QAResult(
            check_id='T06', category=CheckCategory.TECHNICAL,
            title='Major Connection Destinations',
            status=CheckStatus.MANUAL,
            detail='Top connecting markets not itemised in results.',
        ))

    #  T07: Aircraft consistency with fleet 
    aircraft = getattr(config, 'aircraft_type', '') or getattr(config, 'aircraft', '')
    airline_name = getattr(config, 'airline', '') or results.get('airline', '')
    if aircraft:
        checks.append(QAResult(
            check_id='T07', category=CheckCategory.TECHNICAL,
            title='Aircraft / Fleet Consistency',
            status=CheckStatus.MANUAL,
            detail=f'Aircraft: {aircraft} ({seats} seats) for {airline_name}',
            evidence='Reviewer: confirm aircraft is in current/planned fleet.'
        ))
    else:
        checks.append(QAResult(
            check_id='T07', category=CheckCategory.TECHNICAL,
            title='Aircraft / Fleet Consistency',
            status=CheckStatus.WARN,
            detail='Aircraft type not specified.',
        ))

    #  T08: Forecast defendable? 
    # This is always a manual expert judgment call
    fail_count = sum(1 for c in checks if c.status == CheckStatus.FAIL)
    warn_count = sum(1 for c in checks if c.status == CheckStatus.WARN)

    if fail_count > 0:
        detail = f'{fail_count} technical check(s) failed  forecast may not be defendable without resolution.'
    elif warn_count > 2:
        detail = f'{warn_count} warnings raised  reviewer should assess cumulative risk.'
    elif warn_count > 0:
        detail = f'{warn_count} minor warning(s)  generally acceptable with documented justification.'
    else:
        detail = 'All automated checks pass. Final defendability is a reviewer judgment.'

    checks.append(QAResult(
        check_id='T08', category=CheckCategory.TECHNICAL,
        title='Is the Forecast Defendable?',
        status=CheckStatus.MANUAL,
        detail=detail,
        evidence='This check always requires senior reviewer sign-off.'
    ))

    return checks


# ============================================================================
# ADDITIONAL AUTOMATED CHECKS (beyond template  added by pipeline)
# ============================================================================

def _check_extended(config, results: Dict[str, Any]) -> List[QAResult]:
    """
    Extended automated checks that go beyond the template.
    These are pipeline-specific quality gates.
    """
    checks = []

    # X01: Circuity threshold applied
    circuity = getattr(config, 'circuity_threshold', None) or results.get('circuity_threshold')
    if circuity is not None:
        checks.append(QAResult(
            check_id='X01', category=CheckCategory.TECHNICAL,
            title='Circuity Threshold',
            status=CheckStatus.PASS,
            detail=f'Circuity threshold: {circuity:.0%}' if isinstance(circuity, float) else f'Circuity: {circuity}',
            evidence='Connections exceeding this threshold are excluded from QSI.'
        ))

    # X02: QSI adjustment factor
    qsi_adj = results.get('qsi_adj_factor') or results.get('qsi_adjustment')
    if qsi_adj is not None:
        is_new = getattr(config, 'is_new_route', True)
        lo, hi = QSI_ADJ_BOUNDS.get('new_route' if is_new else 'existing', QSI_ADJ_BOUNDS['default'])
        if lo <= qsi_adj <= hi:
            status = CheckStatus.PASS
        else:
            status = CheckStatus.WARN
        checks.append(QAResult(
            check_id='X02', category=CheckCategory.TECHNICAL,
            title='QSI Adjustment Factor',
            status=status,
            detail=f'QSI adj: {qsi_adj:.3f} (expected: {lo:.2f}-{hi:.2f} for {"new" if is_new else "existing"} route)',
        ))

    # X03: Sabre factor-up applied?
    sabre_factor = results.get('sabre_factor_up') or results.get('sabre_factor')
    if sabre_factor is not None:
        if 1.0 <= sabre_factor <= 2.0:
            status = CheckStatus.PASS
        else:
            status = CheckStatus.WARN
        checks.append(QAResult(
            check_id='X03', category=CheckCategory.TECHNICAL,
            title='Sabre Factor-Up',
            status=status,
            detail=f'Sabre factor: {sabre_factor:.2f}x',
            evidence='Adjusts for Sabre undercount vs ACI/airport-reported figures.'
        ))

    # X04: Growth rate sanity
    growth = results.get('annual_growth') or results.get('growth_rate')
    if growth is not None:
        lo, hi = GROWTH_BOUNDS['default']
        if lo <= growth <= hi:
            status = CheckStatus.PASS
        elif growth > hi:
            status = CheckStatus.WARN
        else:
            status = CheckStatus.PASS  # Conservative growth is fine
        checks.append(QAResult(
            check_id='X04', category=CheckCategory.TECHNICAL,
            title='Annual Growth Rate',
            status=status,
            detail=f'Growth: {growth:.1%} (typical range: {lo:.0%}-{hi:.0%})',
        ))

    # X05: Before/beyond balance (hub routes)
    before_pax = results.get('before_pax') or results.get('connecting_home')
    beyond_pax = results.get('beyond_pax') or results.get('connecting_dest')
    if before_pax is not None and beyond_pax is not None and (before_pax + beyond_pax) > 0:
        total_cnx = before_pax + beyond_pax
        ratio = before_pax / total_cnx
        if 0.25 <= ratio <= 0.75:
            status = CheckStatus.PASS
            detail = f'Before/beyond split: {ratio:.0%} / {1-ratio:.0%}'
        else:
            status = CheckStatus.WARN
            heavier = 'before (home hub)' if ratio > 0.75 else 'beyond (dest hub)'
            detail = f'Heavily skewed toward {heavier}: {ratio:.0%} / {1-ratio:.0%}'
        checks.append(QAResult(
            check_id='X05', category=CheckCategory.TECHNICAL,
            title='Before/Beyond Balance',
            status=status, detail=detail,
            evidence=f'Before: {before_pax:,.0f}, Beyond: {beyond_pax:,.0f}'
        ))

    return checks


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def run_qa_checklist(config, results: Dict[str, Any],
                     previous_forecast: Optional[Dict[str, Any]] = None,
                     data_sources: Optional[Dict[str, bool]] = None,
                     calibration_library: Optional[Dict] = None,
                     analyst: str = "Pipeline") -> QAReport:
    """
    Run the full Q&A checklist against pipeline results.

    Args:
        config: RouteConfig or similar with route parameters.
        results: Dict from run_pipeline() with forecast outputs.
        previous_forecast: Optional dict with previous forecast for comparison.
        data_sources: Optional dict flagging which data sources were used.
        calibration_library: Optional calibration library data.
        analyst: Name of the analyst running the check.

    Returns:
        QAReport with all check results.
    """
    # Extract route info
    origin = getattr(config, 'origin', '') or results.get('origin', '???')
    dest = getattr(config, 'destination', '') or results.get('destination', '???')
    airline = getattr(config, 'airline', '') or results.get('airline', '???')

    report = QAReport(
        route=f'{origin}-{dest}',
        airline=airline,
        analyst=analyst,
        date=datetime.now().strftime('%Y-%m-%d %H:%M'),
    )

    # Run all check categories
    for result in _check_workflow(config, results, data_sources):
        report.add(result)

    for result in _check_technical(config, results, previous_forecast, calibration_library):
        report.add(result)

    for result in _check_extended(config, results):
        report.add(result)

    return report


# ============================================================================
# VALIDATION
# ============================================================================

def _test_ba_lhr_sjc():
    """Validate against BA LHR-SJC known case."""

    class MockConfig:
        origin = 'LHR'
        destination = 'SJC'
        airline = 'BA'
        carrier_type = 'Full-service'
        weekly_frequency = 7
        seats = 265
        seat_capacity = 265
        aircraft_type = 'A330-200'
        aircraft = 'A330-200'
        is_new_route = True
        circuity_threshold = 0.30
        schedule_provider = 'SingleExtractOAGProvider'

    results = {
        'total_annual_pax': 129162,
        'grand_total': 129162,
        'p2p_pax': 57729,
        'p2p_captured': 57729,
        'connecting_pax': 71433,
        'total_connecting_pax': 71433,
        'p2p_demand': 214500,
        'p2p_base': 214500,
        'connecting_total': 71433,
        'capture_rate': 0.269,
        'p2p_capture_rate': 0.269,
        'stimulation': 1.10,
        'stim_factor': 1.10,
        'total_itineraries': 4500,
        'itinerary_count': 4500,
        'connections_built': 3200,
        'connection_count': 3200,
        'qsi_market_count': 85,
        'markets_scored': 85,
        'calibrated': True,
        'qsi_adj_factor': 1.0,
        'sabre_factor_up': 1.15,
        'annual_growth': 0.025,
        'growth_rate': 0.025,
        'connecting_home': 38000,
        'connecting_dest': 33433,
        'before_pax': 38000,
        'beyond_pax': 33433,
        'airline': 'BA',
    }

    config = MockConfig()
    report = run_qa_checklist(config, results, analyst='Test')

    print(report.summary())

    # Assertions
    assert report.fail_count == 0, f"Expected 0 fails, got {report.fail_count}"
    assert report.pass_count >= 15, f"Expected 15 passes, got {report.pass_count}"
    assert len(report.results) >= 20, f"Expected 20 checks, got {len(report.results)}"

    # Test Excel output
    import tempfile, os
    path = os.path.join(tempfile.gettempdir(), 'qa_test.xlsx')
    report.to_excel(path)
    assert os.path.exists(path), "Excel output not created"
    size = os.path.getsize(path)
    assert size > 3000, f"Excel file too small: {size}"
    print(f"\nExcel written: {path} ({size:,} bytes)")

    # Test dict output
    d = report.to_dict()
    assert d['overall_status'] in ('PASS', 'WARNING', 'MANUAL REVIEW')
    assert d['counts']['fail'] == 0
    print(f"\nDict output: {d['counts']}")

    print("\n BA LHR-SJC Q&A validation PASSED")
    return report


def _test_minimal():
    """Test with minimal data to ensure graceful handling."""

    class MinConfig:
        origin = 'XXX'
        destination = 'YYY'
        airline = ''
        carrier_type = ''
        weekly_frequency = None
        seats = None
        seat_capacity = None
        aircraft_type = ''
        aircraft = ''
        is_new_route = True
        circuity_threshold = None

    report = run_qa_checklist(MinConfig(), {}, analyst='Test')
    print(f"\nMinimal test: {report.pass_count} pass, {report.warn_count} warn, "
          f"{report.fail_count} fail, {report.manual_count} manual, {report.skip_count} skip")
    assert report.fail_count == 0, "Minimal data should not cause failures"
    print(" Minimal data test PASSED")
    return report


if __name__ == '__main__':
    print("=" * 70)
    print("QA CHECKLIST MODULE  VALIDATION")
    print("=" * 70)
    _test_ba_lhr_sjc()
    print()
    _test_minimal()
    print("\n ALL TESTS PASSED")
