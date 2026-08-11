#!/usr/bin/env python3
r"""Avia Solutions - the aircraft economics fill-in pack, for someone outside this session.

    py -3.12 aircraft_fillin_workbook.py [output.xlsx]

WHY A SEPARATE PACK. aircraft_gap_report.py describes the gap for us. This is written to be handed to
someone who was not here: it carries what Avia already found on 10 August 2026 so the same ground is
not covered twice, states what is measured against what is scaled, and asks only for what is missing.

WHAT IT ASKS FOR, in priority order:
  1. The types the schedule flies that the economics module still cannot cost. Twenty-six of them,
     led by the E175 at 1.4m sectors a year, which is the aircraft an airport takes to a regional
     carrier and which the tool cannot currently put on a route.
  2. Market values and lease rates, for EVERY type including the ones already held. This is the
     largest hole in the module. Four independent searches on 10 August, three of them external,
     found no current type-and-age figure in free public form; appraiser data is subscription and
     may not be republished. Anything a subscription or an appraiser relationship can supply here is
     worth more than the rest of the pack combined.

THE RULES, and they matter more than coverage:
  - Every figure needs a source and a date. A figure without one cannot be defended in a report.
  - NOT FOUND is a valid and useful answer. Do not estimate, interpolate or scale from a similar
    type. A visible gap can be closed; a guess that looks like a fact cannot be caught.
  - BLOCK fuel includes taxi, TRIP fuel does not, CRUISE flow is neither. Say which the source gives.
  - Cargo means the hold WEIGHT limit. A manufacturer's maximum structural payload is passengers
    plus bags plus cargo and is two to four times the hold limit. Do not substitute it.

Avia Solutions Limited. All rights reserved.
"""
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

# Fields the module holds, and which of them research already closed for the types still to add.
ASK = ["econ_seats", "bus_seats", "range_km", "mtow_kg", "fuel_burn_kg_per_bh", "cargo_cap_kg",
       "maint_per_bh", "crew_per_bh", "annual_util_bh", "price_usd", "lease_usd_month"]


def build(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from aircraft_economics import AIRCRAFT, OWNERSHIP_PROVENANCE
    import aircraft_gap_report as GR
    try:
        from aircraft_research_data import RESEARCH, F41_2023, AC_BLOCK, NOT_RESEARCHED
    except Exception:
        RESEARCH, F41_2023, AC_BLOCK, NOT_RESEARCHED = {}, {}, {}, []

    cen = {r["code"]: r for r in GR.census()}
    by_key = {}
    for code, r in cen.items():
        key, declared = GR.MAP.get(code, (None, "UNMAPPED"))
        if not key or declared in ("EXCLUDE", "REVIEW", "UNMAPPED"):
            continue
        b = by_key.setdefault(key, {"sec": 0, "s": [], "p": [], "g": [], "gmax": 0, "car": 0,
                                    "codes": [], "name": r["name"]})
        b["sec"] += r["sectors"]; b["codes"].append(code)
        b["car"] = max(b["car"], r["carriers"]); b["gmax"] = max(b["gmax"], r["gcd_max"])
        b["s"].append((r["sectors"], r["seats"])); b["p"].append((r["sectors"], r["prem"]))
        b["g"].append((r["sectors"], r["gcd"]))
    w = lambda ps: round(sum(a * b for a, b in ps) / (sum(a for a, _ in ps) or 1))
    for b in by_key.values():
        b["seats"], b["prem"], b["gcd"] = w(b["s"]), w(b["p"]), w(b["g"])

    ARIAL = "Arial"
    hdr = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F3864")
    body = Font(name=ARIAL, size=10)
    blue = Font(name=ARIAL, size=10, color="0000FF")
    note = Font(name=ARIAL, size=9, italic=True, color="595959")
    fill_ask = PatternFill("solid", fgColor="FFFF00")     # to fill
    fill_have = PatternFill("solid", fgColor="E2EFDA")    # already found, check only
    thin = Border(bottom=Side(style="thin", color="BFBFBF"))

    wb = Workbook()

    def sheet(name, heads, widths, freeze="C2"):
        ws = wb.create_sheet(name)
        for j, h in enumerate(heads, 1):
            c = ws.cell(row=1, column=j, value=h); c.font = hdr; c.fill = hfill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
        ws.freeze_panes = freeze; ws.row_dimensions[1].height = 34
        return ws

    # ---------------------------------------------------------------- start here
    ws = wb.active; ws.title = "Start here"
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 108
    to_add = [k for k in by_key if k not in AIRCRAFT]
    lines = [
        ("Aircraft economics: what Avia holds and what is missing", None),
        ("", None),
        ("What this is", "The route profitability tool costs a route from a per-aircraft table. It "
                         "holds %d types. The 2025 schedule flies %d equipment codes above 5,000 "
                         "sectors, so anything absent cannot be put on a route at all."
                         % (len(AIRCRAFT), len(cen))),
        ("Sheet 1", "TYPES TO ADD. %d types the schedule flies and the tool cannot cost. Green "
                    "cells are what Avia already found on 10 August 2026, so please check rather "
                    "than repeat them. Yellow cells are the ask." % len(to_add)),
        ("Sheet 2", "VALUES AND LEASES. The largest hole, and it applies to types already held as "
                    "well as new ones. Four independent searches on 10 August, three of them "
                    "external, found no current type-and-age market value or lease rate in free "
                    "public form. Appraiser data is subscription and may not be republished, so "
                    "anything reachable here is worth more than the rest of the pack."),
        ("Sheet 3", "HELD TYPES. What the tool currently uses, so it can be challenged. The src "
                    "column says where each figure came from."),
        ("", None),
        ("Rule 1", "Every figure needs a source and a date. A figure without one cannot be defended "
                   "in a client report and will not be used."),
        ("Rule 2", "NOT FOUND is a valid and useful answer. Please do not estimate, interpolate or "
                   "scale from a similar type. A visible gap can be closed later; a guess that looks "
                   "like a fact cannot be caught."),
        ("Rule 3", "BLOCK fuel includes taxi. TRIP fuel does not. CRUISE flow is neither. Say which "
                   "the source gives, and do not convert between them."),
        ("Rule 4", "Cargo means the hold WEIGHT limit in kg. A manufacturer's maximum structural "
                   "payload is passengers plus bags plus cargo combined and runs two to four times "
                   "the hold limit. If only a volume is published, say so."),
        ("Rule 5", "Where a figure varies by variant, engine or weight option, say which one."),
        ("", None),
        ("Sources that worked", "Manufacturer airport planning documents; EASA and FAA type "
                                "certificate data sheets; EUROCONTROL; Aircraft Commerce; IBA; "
                                "Cirium; Airfinance Journal and Air Investor; SEC filings including "
                                "EETC prospectuses, which carry three-appraiser valuations and are "
                                "the most productive free route to values."),
        ("Sources to avoid", "Scribd, aerocorner, aviatorinsider, Simple Flying, planespotters, "
                             "flugzeuginfo, globalair, PPRuNe, Airliners.net and aircraft-data "
                             "aggregators. Wikipedia only to locate a primary document."),
        ("", None),
        ("Source", "Traffic columns: OAG schedules 2025, service_type J, nonstop. Generated by "
                   "app/aircraft_fillin_workbook.py."),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=a).font = (Font(name=ARIAL, bold=True, size=12) if i == 1
                                                  else Font(name=ARIAL, bold=True, size=10))
        if b:
            c = ws.cell(row=i, column=2, value=b); c.font = body
            c.alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[i].height = 46

    # ---------------------------------------------------------------- 1. types to add
    heads = ["priority", "key", "type", "OAG codes", "2025 sectors", "carriers",
             "OAG median seats", "OAG median premium", "OAG median km", "OAG max km",
             "range_km", "mtow_kg", "fuel_burn_kg_per_bh", "burn basis (block/trip/cruise)",
             "cargo_cap_kg", "price_usd", "lease_usd_month", "source", "source date", "notes"]
    widths = [8, 9, 30, 13, 12, 9, 12, 13, 12, 12, 12, 12, 16, 24, 13, 14, 15, 40, 12, 34]
    ws = sheet("1. Types to add", heads, widths)
    r = 2
    for i, key in enumerate(sorted(to_add, key=lambda k: -by_key[k]["sec"]), 1):
        b = by_key[key]; res = RESEARCH.get(key, {})
        get = lambda f: (res.get(f) or (None,))[0]
        known = {"range_km": get("range_km"), "mtow_kg": get("mtow_kg"),
                 "cargo_cap_kg": get("cargo_cap_kg"),
                 "fuel_burn_kg_per_bh": F41_2023.get(key, (None, None, None))[1]
                                        if key in F41_2023 else
                                        (round(sum(AC_BLOCK[key][:2]) / 2) if key in AC_BLOCK else None)}
        for j, v in ((1, i), (2, key), (3, b["name"][:38]), (4, ", ".join(sorted(b["codes"]))),
                     (5, b["sec"]), (6, b["car"]), (7, b["seats"]), (8, b["prem"]),
                     (9, b["gcd"]), (10, b["gmax"])):
            c = ws.cell(row=r, column=j, value=v); c.font = body; c.border = thin
        for j, f in ((11, "range_km"), (12, "mtow_kg"), (13, "fuel_burn_kg_per_bh"),
                     (15, "cargo_cap_kg")):
            c = ws.cell(row=r, column=j, value=known.get(f)); c.border = thin
            c.font = blue; c.fill = fill_have if known.get(f) else fill_ask
        for j in (14, 16, 17, 18, 19, 20):
            c = ws.cell(row=r, column=j); c.font = blue; c.fill = fill_ask; c.border = thin
        if known.get("fuel_burn_kg_per_bh"):
            ws.cell(row=r, column=14, value=("BLOCK, US DOT Form 41 2023 over the T-100 measured "
                                             "block ratio" if key in F41_2023 else
                                             "BLOCK, Aircraft Commerce published"))
            ws.cell(row=r, column=14).fill = fill_have
        r += 1
    ws.cell(row=r + 1, column=1, value=(
        "Green cells are what Avia found on 10 August 2026. Please check them rather than repeat "
        "them, and say if you disagree. Yellow is the ask. Priority is by sectors flown, so the top "
        "of the sheet is where the work pays: the E175 alone is 1.4m sectors a year.")).font = note

    # ---------------------------------------------------------------- 2. values and leases
    heads = ["priority", "key", "type", "held by Avia", "2025 sectors",
             "current market value USD", "value: age or build year", "monthly lease USD",
             "lease: age or build year", "condition assumed (half-life / full-life)",
             "source", "source date", "notes"]
    widths = [8, 9, 30, 12, 12, 20, 20, 18, 20, 26, 40, 12, 34]
    ws = sheet("2. Values and leases", heads, widths)
    allkeys = sorted(set(list(AIRCRAFT) + to_add),
                     key=lambda k: -(by_key.get(k, {}).get("sec", 0)))
    r = 2
    for i, key in enumerate(allkeys, 1):
        b = by_key.get(key, {})
        prov = OWNERSHIP_PROVENANCE.get(key, "none") if key in AIRCRAFT else "not held"
        for j, v in ((1, i), (2, key),
                     (3, (b.get("name") or "")[:38]),
                     (4, "yes, ownership %s" % prov if key in AIRCRAFT else "no"),
                     (5, b.get("sec", 0))):
            c = ws.cell(row=r, column=j, value=v); c.font = body; c.border = thin
        for j in range(6, 14):
            c = ws.cell(row=r, column=j); c.font = blue; c.border = thin
            c.fill = fill_have if (key in AIRCRAFT and prov == "citable") else fill_ask
        r += 1
    ws.cell(row=r + 1, column=1, value=(
        "Green means Avia already has a citable figure from the June 2026 register and only needs it "
        "refreshed. Everything yellow is unsourced today, including for types the tool already uses: "
        "of the types held, only eight have a citable ownership figure, three are proxies and the "
        "rest have no recorded origin at all.")).font = note
    ws.cell(row=r + 2, column=1, value=(
        "EETC prospectuses on SEC EDGAR carry three-appraiser valuations and are public. They were "
        "the single most productive free route found and are worth a systematic look.")).font = note

    # ---------------------------------------------------------------- 3. held reference
    heads = ["key", "2025 sectors", "econ_seats", "bus_seats", "mtow_kg", "range_km",
             "fuel_burn_kg_per_bh", "maint_per_bh", "crew_per_bh", "ownership_per_bh",
             "annual_util_bh", "price_usd", "cargo_cap_kg", "ownership provenance", "src"]
    ws = sheet("3. Held types", heads, [9, 12, 11, 10, 11, 11, 16, 12, 12, 15, 12, 14, 12, 18, 90],
               freeze="B2")
    r = 2
    for key in sorted(AIRCRAFT, key=lambda k: -(by_key.get(k, {}).get("sec", 0))):
        v = AIRCRAFT[key]
        vals = [key, by_key.get(key, {}).get("sec", 0), v["econ_seats"], v["bus_seats"],
                v["mtow_kg"], v["range_km"], v["fuel_burn_kg_per_bh"], v["maint_per_bh"],
                v["crew_per_bh"], v["ownership_per_bh"], v["annual_util_bh"], v["price_usd"],
                v["cargo_cap_kg"], OWNERSHIP_PROVENANCE.get(key, "none"), v["src"]]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=val); c.font = body; c.border = thin
            if j == 15:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    ws.cell(row=r + 1, column=1, value=(
        "The src column says where each figure came from. Where it says SCALED, the figure is "
        "derived from a neighbouring type and is not sourced: those are the ones most worth "
        "challenging.")).font = note

    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    wb.properties.title = "Aircraft economics: fill-in pack"
    wb.save(path)
    return path, len(to_add), len(AIRCRAFT)


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "Aircraft economics - fill-in pack.xlsx"
    p, n_add, n_held = build(out)
    print("written: %s  (%d types held, %d to add)" % (p, n_held, n_add))
