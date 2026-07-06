#!/usr/bin/env python3
"""
Avia Solutions - Learning Comparison Module
============================================
Compares the auto-generated forecast (from the pipeline) against an
analyst-completed forecast workbook to extract calibration learnings.

PURPOSE:
  This is the feedback loop that makes the system smarter over time.
  After the pipeline produces an auto-forecast, the analyst uploads
  their completed forecast. This module:

    1. Parses the analyst's forecast workbook (standard Avia format)
    2. Compares headline numbers (total, P2P, connecting, LF)
    3. Compares city-by-city connecting captures (where the real
       calibration intelligence lives)
    4. Extracts the implicit QSI adjustment factors the analyst applied
    5. Classifies the learnings into calibration patterns
    6. Produces a structured output the PCE can ingest

KEY INSIGHT:
  The difference between the raw QSI score and the analyst's final QSI
  value IS the expert calibration. For BA LHR-SJC, the pipeline might
  compute QSI=0.23 for Paris, but the analyst uses 0.04. That 5.8x
  reduction encodes judgment about BA's competitive position vs AF on
  LHR-CDG-SJC routings. Extracting these ratios across many routes
  builds the pattern database for predictive calibration.

ANALYST WORKBOOK FORMAT (standard Avia):
  The forecast workbook typically contains:
    - 'Forecast Cnx @ Home Airport' or similar: city-by-city connecting
      with columns for base demand, QSI capture, and forecast pax
    - 'Forecast Cnx @ Destination' or similar: same for dest side
    - 'Forecast TABLE' or 'Forecast Finalised': headline summary
    - 'P2P' or similar: point-to-point segment details

USAGE:
  from learning_comparison import LearningComparison, parse_analyst_workbook

  auto_results = st.session_state.results  # from pipeline
  analyst_data = parse_analyst_workbook('/path/to/analyst_forecast.xlsm')
  comparison = LearningComparison(auto_results, analyst_data, config)
  report = comparison.run()
  report.print_summary()

LAST UPDATED: February 2026
"""

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any
from enum import Enum
import copy
import json
from datetime import datetime


# =============================================================================
# ENUMS
# =============================================================================

class VarianceSeverity(Enum):
    EXACT = "exact"           # <1% variance
    CLOSE = "close"           # 1-5% variance
    MODERATE = "moderate"     # 5-15% variance
    SIGNIFICANT = "significant"  # 15-30% variance
    MAJOR = "major"           # >30% variance


class LearningType(Enum):
    QSI_OVERESTIMATE = "qsi_overestimate"     # Pipeline QSI > analyst QSI
    QSI_UNDERESTIMATE = "qsi_underestimate"   # Pipeline QSI < analyst QSI
    P2P_CAPTURE_HIGH = "p2p_capture_high"
    P2P_CAPTURE_LOW = "p2p_capture_low"
    STIM_HIGH = "stim_high"
    STIM_LOW = "stim_low"
    CITY_EXCLUDED = "city_excluded"           # Analyst excluded a city pipeline included
    CITY_ADDED = "city_added"                 # Analyst added a city pipeline missed
    DEMAND_BASE_DIFF = "demand_base_diff"     # Different base demand figures


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class AnalystData:
    """Parsed data from analyst forecast workbook."""
    # Headlines
    total_pax: float = 0
    p2p_total: float = 0
    home_cnx_total: float = 0
    dest_cnx_total: float = 0
    load_factor: float = 0.0

    # City-by-city connecting (list of dicts)
    home_cities: List[Dict] = field(default_factory=list)
    dest_cities: List[Dict] = field(default_factory=list)

    # P2P segments
    p2p_segments: List[Dict] = field(default_factory=list)

    # Metadata
    source_file: str = ""
    route_label: str = ""
    sheets_found: List[str] = field(default_factory=list)


@dataclass
class CityComparison:
    """Comparison of one connecting city between pipeline and analyst."""
    city_code: str
    city_name: str = ""
    side: str = ""  # 'home' or 'dest'

    # Pipeline values
    pipe_base_demand: float = 0
    pipe_qsi_capture: float = 0
    pipe_forecast: float = 0

    # Analyst values
    analyst_base_demand: float = 0
    analyst_qsi_capture: float = 0
    analyst_forecast: float = 0

    # Derived
    qsi_ratio: float = 0.0       # analyst_qsi / pipe_qsi
    forecast_variance: float = 0.0
    base_demand_variance: float = 0.0

    # Classification
    learning_type: str = ""
    severity: str = ""
    note: str = ""

    def to_dict(self) -> Dict:
        return {
            'city': self.city_code,
            'side': self.side,
            'pipe_qsi': self.pipe_qsi_capture,
            'analyst_qsi': self.analyst_qsi_capture,
            'qsi_ratio': self.qsi_ratio,
            'pipe_pax': self.pipe_forecast,
            'analyst_pax': self.analyst_forecast,
            'variance': self.forecast_variance,
            'severity': self.severity,
            'learning': self.learning_type,
        }


@dataclass
class HeadlineComparison:
    """Top-level comparison of pipeline vs analyst."""
    pipe_total: float = 0
    analyst_total: float = 0
    total_variance: float = 0.0

    pipe_p2p: float = 0
    analyst_p2p: float = 0
    p2p_variance: float = 0.0

    pipe_home: float = 0
    analyst_home: float = 0
    home_variance: float = 0.0

    pipe_dest: float = 0
    analyst_dest: float = 0
    dest_variance: float = 0.0

    pipe_lf: float = 0.0
    analyst_lf: float = 0.0
    lf_diff_pp: float = 0.0  # percentage points

    overall_severity: str = ""


@dataclass
class CalibrationLearning:
    """A single calibration learning extracted from the comparison."""
    learning_type: str
    parameter: str        # which parameter was off
    direction: str        # 'high' or 'low'
    magnitude: float      # how far off
    city_code: str = ""   # which city (if city-specific)
    side: str = ""        # 'home' or 'dest'
    pipe_value: float = 0.0
    analyst_value: float = 0.0
    recommendation: str = ""

    def to_dict(self) -> Dict:
        return {
            'type': self.learning_type,
            'parameter': self.parameter,
            'direction': self.direction,
            'magnitude': self.magnitude,
            'city': self.city_code,
            'side': self.side,
            'pipe_value': self.pipe_value,
            'analyst_value': self.analyst_value,
            'recommendation': self.recommendation,
        }


@dataclass
class LearningReport:
    """Complete output from the learning comparison."""
    # Route identity
    route_label: str = ""
    timestamp: str = ""

    # Headlines
    headline: Optional[HeadlineComparison] = None

    # City comparisons
    city_comparisons: List[CityComparison] = field(default_factory=list)

    # Extracted learnings
    learnings: List[CalibrationLearning] = field(default_factory=list)

    # Summary statistics
    median_qsi_ratio_home: float = 0.0
    median_qsi_ratio_dest: float = 0.0
    cities_overestimated: int = 0
    cities_underestimated: int = 0
    cities_matched: int = 0

    # Blended QSI adjustment implied by analyst
    implied_qsi_adjustment: float = 1.0

    # Summary text
    summary_lines: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict:
        return {
            'route': self.route_label,
            'timestamp': self.timestamp,
            'headline': {
                'total_var': self.headline.total_variance if self.headline else 0,
                'p2p_var': self.headline.p2p_variance if self.headline else 0,
                'home_var': self.headline.home_variance if self.headline else 0,
                'dest_var': self.headline.dest_variance if self.headline else 0,
                'severity': self.headline.overall_severity if self.headline else '',
            },
            'city_comparisons': [c.to_dict() for c in self.city_comparisons],
            'learnings': [l.to_dict() for l in self.learnings],
            'median_qsi_ratio_home': self.median_qsi_ratio_home,
            'median_qsi_ratio_dest': self.median_qsi_ratio_dest,
            'implied_qsi_adjustment': self.implied_qsi_adjustment,
        }

    def print_summary(self):
        for line in self.summary_lines:
            print(line)


# =============================================================================
# ANALYST WORKBOOK PARSER
# =============================================================================

def parse_analyst_workbook(filepath: str) -> AnalystData:
    """
    Parse an analyst-completed forecast workbook.
    Handles standard Avia forecast format with tolerance for variations.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required for workbook parsing")

    data = AnalystData(source_file=filepath)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        # Try xlrd for .xls files
        try:
            import xlrd
            xwb = xlrd.open_workbook(filepath)
            data.sheets_found = xwb.sheet_names()
            # Basic extraction from xlrd - limited but functional
            return data
        except Exception:
            return data

    data.sheets_found = wb.sheetnames

    # --- Find and parse connecting city sheets ---
    for sheet_name in wb.sheetnames:
        # Skip Chartsheet objects (pie charts, bar charts stored as sheets)
        try:
            ws_check = wb[sheet_name]
            if type(ws_check).__name__ == 'Chartsheet':
                continue
        except Exception:
            continue
        sn_lower = sheet_name.lower()

        # Home connecting
        if ('home' in sn_lower and 'cnx' in sn_lower) or \
           ('home' in sn_lower and 'connect' in sn_lower) or \
           ('cnx' in sn_lower and 'home' in sn_lower and 'airport' in sn_lower):
            data.home_cities = _parse_connecting_sheet(wb[sheet_name], 'home')

        # Destination connecting
        elif ('dest' in sn_lower and 'cnx' in sn_lower) or \
             ('dest' in sn_lower and 'connect' in sn_lower) or \
             ('cnx' in sn_lower and 'dest' in sn_lower and 'airport' in sn_lower):
            data.dest_cities = _parse_connecting_sheet(wb[sheet_name], 'dest')

        # Forecast table / finalised
        elif 'forecast' in sn_lower and ('table' in sn_lower or 'final' in sn_lower):
            _parse_headline_sheet(wb[sheet_name], data)

        # P2P
        elif 'p2p' in sn_lower:
            data.p2p_segments = _parse_p2p_sheet(wb[sheet_name])

    # Calculate totals if not found in headline sheet
    if data.home_cnx_total == 0 and data.home_cities:
        data.home_cnx_total = sum(c.get('forecast', 0) for c in data.home_cities)
    if data.dest_cnx_total == 0 and data.dest_cities:
        data.dest_cnx_total = sum(c.get('forecast', 0) for c in data.dest_cities)
    if data.total_pax == 0 and (data.p2p_total or data.home_cnx_total or data.dest_cnx_total):
        data.total_pax = data.p2p_total + data.home_cnx_total + data.dest_cnx_total

    wb.close()
    return data


def _parse_connecting_sheet(ws, side: str) -> List[Dict]:
    """Parse a connecting city sheet (Home or Dest side)."""
    cities = []

    # Scan first 10 rows to find header
    header_row = None
    col_map = {}

    for row_idx in range(1, 12):
        row_vals = []
        for col_idx in range(1, 25):
            cell = ws.cell(row_idx, col_idx)
            row_vals.append((col_idx, str(cell.value).lower().strip() if cell.value else ''))

        for col_idx, val in row_vals:
            if 'city' in val or 'iata' in val or 'airport' in val:
                col_map['city'] = col_idx
                header_row = row_idx
            if 'base' in val and ('demand' in val or 'indirect' in val or 'total' in val):
                col_map['base_demand'] = col_idx
            if 'qsi' in val and ('capture' in val or 'share' in val or 'adj' in val):
                col_map['qsi'] = col_idx
            if 'forecast' in val or ('captured' in val and 'pax' in val):
                col_map['forecast'] = col_idx
            # Also look for column with just 'QSI' as header
            if val == 'qsi' and 'qsi' not in col_map:
                col_map['qsi'] = col_idx

        if header_row:
            break

    if not header_row or 'city' not in col_map:
        # Try positional defaults (common Avia layout)
        # Col B=city, Col G/H=base demand, Col M=QSI, Col S/T=forecast
        header_row = 4
        col_map = {'city': 2, 'base_demand': 7, 'qsi': 13, 'forecast': 19}

    # Parse data rows
    for row_idx in range(header_row + 1, ws.max_row + 1 if ws.max_row else 500):
        city_cell = ws.cell(row_idx, col_map.get('city', 2))
        city = str(city_cell.value).strip() if city_cell.value else ''

        # Skip non-city rows
        if not city or len(city) != 3 or city.lower() in ('tot', 'total', 'sum', 'nan'):
            continue

        entry = {
            'city': city.upper(),
            'side': side,
            'base_demand': 0,
            'qsi_capture': 0.0,
            'forecast': 0,
        }

        if 'base_demand' in col_map:
            v = ws.cell(row_idx, col_map['base_demand']).value
            entry['base_demand'] = float(v) if v and isinstance(v, (int, float)) else 0

        if 'qsi' in col_map:
            v = ws.cell(row_idx, col_map['qsi']).value
            entry['qsi_capture'] = float(v) if v and isinstance(v, (int, float)) else 0.0

        if 'forecast' in col_map:
            v = ws.cell(row_idx, col_map['forecast']).value
            entry['forecast'] = float(v) if v and isinstance(v, (int, float)) else 0

        if entry['base_demand'] > 0 or entry['forecast'] > 0:
            cities.append(entry)

    return cities


def _parse_headline_sheet(ws, data: AnalystData):
    """Parse the Forecast TABLE / Finalised sheet for headline numbers."""
    for row in ws.iter_rows(min_row=1, max_row=80, max_col=20, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_lower = cell.value.lower().strip()
                # Look for labels and grab the value in the next column or nearby
                if 'grand total' in val_lower or 'total forecast' in val_lower:
                    for offset in [1, 2, 3]:
                        v = ws.cell(cell.row, cell.column + offset).value
                        if v and isinstance(v, (int, float)) and v > 1000:
                            data.total_pax = float(v)
                            break
                elif 'p2p' in val_lower and 'total' in val_lower:
                    for offset in [1, 2, 3]:
                        v = ws.cell(cell.row, cell.column + offset).value
                        if v and isinstance(v, (int, float)):
                            data.p2p_total = float(v)
                            break
                elif 'load factor' in val_lower:
                    for offset in [1, 2, 3]:
                        v = ws.cell(cell.row, cell.column + offset).value
                        if v and isinstance(v, (int, float)):
                            data.load_factor = float(v) if v > 1 else float(v) * 100
                            break


def _parse_p2p_sheet(ws) -> List[Dict]:
    """Parse P2P segment details."""
    segments = []
    for row in ws.iter_rows(min_row=2, max_row=30, max_col=10, values_only=True):
        if row[0] and isinstance(row[0], str) and len(row[0]) > 2:
            seg = {'name': str(row[0])}
            for i, key in enumerate(['base_demand', 'growth', 'stimulation', 'capture', 'forecast'], 1):
                if i < len(row) and row[i] is not None:
                    seg[key] = float(row[i]) if isinstance(row[i], (int, float)) else 0
            segments.append(seg)
    return segments


# =============================================================================
# LEARNING COMPARISON ENGINE
# =============================================================================

class LearningComparison:
    """
    Compares auto-generated pipeline results against analyst forecast.
    Extracts calibration learnings.
    """

    def __init__(self, auto_results: Dict, analyst_data: AnalystData,
                 config=None):
        self.auto = auto_results
        self.analyst = analyst_data
        self.config = config

    def _classify_variance(self, variance: float) -> str:
        av = abs(variance)
        if av < 0.01:
            return VarianceSeverity.EXACT.value
        elif av < 0.05:
            return VarianceSeverity.CLOSE.value
        elif av < 0.15:
            return VarianceSeverity.MODERATE.value
        elif av < 0.30:
            return VarianceSeverity.SIGNIFICANT.value
        else:
            return VarianceSeverity.MAJOR.value

    def _safe_var(self, pipe_val: float, analyst_val: float) -> float:
        if analyst_val == 0:
            return 0.0 if pipe_val == 0 else 1.0
        return (pipe_val - analyst_val) / analyst_val

    def _compare_headlines(self) -> HeadlineComparison:
        """Compare top-level numbers."""
        h = HeadlineComparison()
        h.pipe_total = self.auto.get('grand_total', 0)
        h.analyst_total = self.analyst.total_pax
        h.total_variance = self._safe_var(h.pipe_total, h.analyst_total)

        h.pipe_p2p = self.auto.get('p2p_total', 0)
        h.analyst_p2p = self.analyst.p2p_total
        h.p2p_variance = self._safe_var(h.pipe_p2p, h.analyst_p2p)

        h.pipe_home = self.auto.get('home_total', 0)
        h.analyst_home = self.analyst.home_cnx_total
        h.home_variance = self._safe_var(h.pipe_home, h.analyst_home)

        h.pipe_dest = self.auto.get('dest_total', 0)
        h.analyst_dest = self.analyst.dest_cnx_total
        h.dest_variance = self._safe_var(h.pipe_dest, h.analyst_dest)

        h.pipe_lf = self.auto.get('load_factor', 0)
        analyst_lf = self.analyst.load_factor
        if analyst_lf > 1:
            analyst_lf = analyst_lf / 100  # convert from percentage
        h.analyst_lf = analyst_lf
        h.lf_diff_pp = (h.pipe_lf - analyst_lf) * 100

        h.overall_severity = self._classify_variance(h.total_variance)
        return h

    def _compare_cities(self) -> List[CityComparison]:
        """Compare city-by-city connecting captures."""
        comparisons = []

        # Build lookup from pipeline results
        for side, pipe_key, analyst_cities in [
            ('home', 'home_results', self.analyst.home_cities),
            ('dest', 'dest_results', self.analyst.dest_cities),
        ]:
            pipe_cities = self.auto.get(pipe_key, [])

            # Index pipeline cities by code
            pipe_by_city = {}
            for pc in pipe_cities:
                if isinstance(pc, dict):
                    code = pc.get('city', pc.get('city_code', ''))
                    if code:
                        pipe_by_city[code.upper()] = pc

            # Compare each analyst city
            for ac in analyst_cities:
                city = ac.get('city', '').upper()
                if not city:
                    continue

                cc = CityComparison(city_code=city, side=side)
                cc.analyst_base_demand = ac.get('base_demand', 0)
                cc.analyst_qsi_capture = ac.get('qsi_capture', 0)
                cc.analyst_forecast = ac.get('forecast', 0)

                if city in pipe_by_city:
                    pc = pipe_by_city[city]
                    cc.pipe_base_demand = pc.get('base_demand', 0)
                    cc.pipe_qsi_capture = pc.get('qsi_capture', pc.get('capture_rate', 0))
                    cc.pipe_forecast = pc.get('forecast', 0)

                    # QSI ratio: how much did analyst scale the QSI?
                    if cc.pipe_qsi_capture > 0:
                        cc.qsi_ratio = cc.analyst_qsi_capture / cc.pipe_qsi_capture
                    elif cc.analyst_qsi_capture > 0:
                        cc.qsi_ratio = 0.01  # Pipeline had zero, analyst had something

                    cc.forecast_variance = self._safe_var(cc.pipe_forecast, cc.analyst_forecast)
                    cc.base_demand_variance = self._safe_var(cc.pipe_base_demand, cc.analyst_base_demand)

                    # Classify
                    if cc.qsi_ratio < 0.50:
                        cc.learning_type = LearningType.QSI_OVERESTIMATE.value
                        cc.note = f"Pipeline QSI {cc.qsi_ratio:.1f}x higher than analyst"
                    elif cc.qsi_ratio > 2.0:
                        cc.learning_type = LearningType.QSI_UNDERESTIMATE.value
                        cc.note = f"Pipeline QSI {1/cc.qsi_ratio:.1f}x lower than analyst"
                    else:
                        cc.learning_type = "aligned"
                        cc.note = "QSI within 2x range"

                else:
                    cc.learning_type = LearningType.CITY_ADDED.value
                    cc.note = "Analyst included city not in pipeline output"

                cc.severity = self._classify_variance(cc.forecast_variance)
                comparisons.append(cc)

            # Check for cities in pipeline but not in analyst
            analyst_codes = {ac.get('city', '').upper() for ac in analyst_cities}
            for code, pc in pipe_by_city.items():
                if code not in analyst_codes:
                    cc = CityComparison(city_code=code, side=side)
                    cc.pipe_base_demand = pc.get('base_demand', 0)
                    cc.pipe_qsi_capture = pc.get('qsi_capture', pc.get('capture_rate', 0))
                    cc.pipe_forecast = pc.get('forecast', 0)
                    cc.learning_type = LearningType.CITY_EXCLUDED.value
                    cc.note = "Pipeline included city that analyst excluded"
                    cc.severity = "moderate"
                    comparisons.append(cc)

        return comparisons

    def _extract_learnings(self, headline: HeadlineComparison,
                           city_comps: List[CityComparison]) -> List[CalibrationLearning]:
        """Extract structured calibration learnings."""
        learnings = []

        # Headline-level learnings
        if abs(headline.home_variance) > 0.15:
            learnings.append(CalibrationLearning(
                learning_type="connecting_home_variance",
                parameter="qsi_adjustment",
                direction="high" if headline.home_variance > 0 else "low",
                magnitude=abs(headline.home_variance),
                side="home",
                pipe_value=headline.pipe_home,
                analyst_value=headline.analyst_home,
                recommendation=f"Home connecting {'overestimated' if headline.home_variance > 0 else 'underestimated'} "
                               f"by {abs(headline.home_variance):.0%}. "
                               f"{'Reduce' if headline.home_variance > 0 else 'Increase'} QSI adjustment."
            ))

        if abs(headline.dest_variance) > 0.15:
            learnings.append(CalibrationLearning(
                learning_type="connecting_dest_variance",
                parameter="qsi_adjustment",
                direction="high" if headline.dest_variance > 0 else "low",
                magnitude=abs(headline.dest_variance),
                side="dest",
                pipe_value=headline.pipe_dest,
                analyst_value=headline.analyst_dest,
                recommendation=f"Dest connecting {'overestimated' if headline.dest_variance > 0 else 'underestimated'} "
                               f"by {abs(headline.dest_variance):.0%}."
            ))

        if abs(headline.p2p_variance) > 0.10:
            learnings.append(CalibrationLearning(
                learning_type="p2p_variance",
                parameter="p2p_capture_or_stimulation",
                direction="high" if headline.p2p_variance > 0 else "low",
                magnitude=abs(headline.p2p_variance),
                pipe_value=headline.pipe_p2p,
                analyst_value=headline.analyst_p2p,
                recommendation=f"P2P {'overestimated' if headline.p2p_variance > 0 else 'underestimated'} "
                               f"by {abs(headline.p2p_variance):.0%}. Review stimulation and capture rates."
            ))

        # City-level learnings (top outliers)
        qsi_outliers = [c for c in city_comps
                        if c.learning_type in (LearningType.QSI_OVERESTIMATE.value,
                                                LearningType.QSI_UNDERESTIMATE.value)
                        and abs(c.forecast_variance) > 0.20]

        for cc in sorted(qsi_outliers, key=lambda x: -abs(x.forecast_variance))[:10]:
            learnings.append(CalibrationLearning(
                learning_type=cc.learning_type,
                parameter="city_qsi_capture",
                direction="high" if cc.qsi_ratio < 1 else "low",
                magnitude=abs(1 - cc.qsi_ratio),
                city_code=cc.city_code,
                side=cc.side,
                pipe_value=cc.pipe_qsi_capture,
                analyst_value=cc.analyst_qsi_capture,
                recommendation=f"{cc.city_code}: analyst QSI {cc.analyst_qsi_capture:.4f} vs "
                               f"pipeline {cc.pipe_qsi_capture:.4f} (ratio {cc.qsi_ratio:.2f})"
            ))

        return learnings

    def _compute_statistics(self, city_comps: List[CityComparison]) -> Dict:
        """Compute summary statistics."""
        home_ratios = [c.qsi_ratio for c in city_comps
                       if c.side == 'home' and c.qsi_ratio > 0 and c.pipe_qsi_capture > 0]
        dest_ratios = [c.qsi_ratio for c in city_comps
                       if c.side == 'dest' and c.qsi_ratio > 0 and c.pipe_qsi_capture > 0]

        def median(lst):
            if not lst:
                return 0.0
            s = sorted(lst)
            n = len(s)
            return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2

        return {
            'median_home': median(home_ratios),
            'median_dest': median(dest_ratios),
            'overestimated': sum(1 for c in city_comps
                                 if c.learning_type == LearningType.QSI_OVERESTIMATE.value),
            'underestimated': sum(1 for c in city_comps
                                  if c.learning_type == LearningType.QSI_UNDERESTIMATE.value),
            'matched': sum(1 for c in city_comps if c.learning_type == 'aligned'),
            'all_ratios': home_ratios + dest_ratios,
        }

    def run(self) -> LearningReport:
        """Execute the full comparison and produce a learning report."""
        report = LearningReport()
        report.timestamp = datetime.now().isoformat()

        # Route label
        if self.config:
            ac = getattr(self.config, 'airline_code', '??')
            ho = getattr(self.config, 'home_airport_code', '???')
            de = getattr(self.config, 'dest_airport_code', '???')
            report.route_label = f"{ac} {ho}-{de}"
        else:
            report.route_label = self.analyst.route_label or "Unknown"

        # Headlines
        report.headline = self._compare_headlines()

        # City-by-city
        report.city_comparisons = self._compare_cities()

        # Learnings
        report.learnings = self._extract_learnings(report.headline, report.city_comparisons)

        # Statistics
        stats = self._compute_statistics(report.city_comparisons)
        report.median_qsi_ratio_home = stats['median_home']
        report.median_qsi_ratio_dest = stats['median_dest']
        report.cities_overestimated = stats['overestimated']
        report.cities_underestimated = stats['underestimated']
        report.cities_matched = stats['matched']

        # Implied QSI adjustment
        all_ratios = stats['all_ratios']
        if all_ratios:
            # The median ratio IS the implied QSI adjustment
            s = sorted(all_ratios)
            n = len(s)
            report.implied_qsi_adjustment = s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2
        else:
            report.implied_qsi_adjustment = 1.0

        # Build summary
        report.summary_lines = self._build_summary(report)

        return report

    def _build_summary(self, report: LearningReport) -> List[str]:
        lines = []
        lines.append(f"=== Learning Comparison Report ===")
        lines.append(f"Route: {report.route_label}")
        lines.append(f"Timestamp: {report.timestamp}")
        lines.append("")

        h = report.headline
        if h:
            lines.append(f"Headline: Pipeline {h.pipe_total:,.0f} vs Analyst {h.analyst_total:,.0f} "
                         f"({h.total_variance:+.1%}) [{h.overall_severity}]")
            lines.append(f"  P2P: {h.pipe_p2p:,.0f} vs {h.analyst_p2p:,.0f} ({h.p2p_variance:+.1%})")
            lines.append(f"  Home cnx: {h.pipe_home:,.0f} vs {h.analyst_home:,.0f} ({h.home_variance:+.1%})")
            lines.append(f"  Dest cnx: {h.pipe_dest:,.0f} vs {h.analyst_dest:,.0f} ({h.dest_variance:+.1%})")
            if h.analyst_lf > 0:
                lines.append(f"  LF: {h.pipe_lf:.1%} vs {h.analyst_lf:.1%} ({h.lf_diff_pp:+.1f} pp)")
        lines.append("")

        lines.append(f"City comparisons: {len(report.city_comparisons)} cities analysed")
        lines.append(f"  Matched (QSI within 2x): {report.cities_matched}")
        lines.append(f"  Overestimated: {report.cities_overestimated}")
        lines.append(f"  Underestimated: {report.cities_underestimated}")
        lines.append(f"  Median QSI ratio (home): {report.median_qsi_ratio_home:.3f}")
        lines.append(f"  Median QSI ratio (dest): {report.median_qsi_ratio_dest:.3f}")
        lines.append(f"  Implied QSI adjustment: {report.implied_qsi_adjustment:.3f}")
        lines.append("")

        if report.learnings:
            lines.append(f"Key learnings ({len(report.learnings)}):")
            for l in report.learnings[:8]:
                lines.append(f"  - {l.recommendation}")

        return lines


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def run_learning_comparison(auto_results: Dict, analyst_filepath: str,
                            config=None) -> LearningReport:
    """One-call convenience: parse analyst file and run comparison."""
    analyst_data = parse_analyst_workbook(analyst_filepath)
    comparison = LearningComparison(auto_results, analyst_data, config)
    return comparison.run()


# =============================================================================
# SELF-TEST
# =============================================================================

if __name__ == '__main__':
    print("Learning Comparison Module - Self Test")
    print("=" * 50)

    # Simulate pipeline results (BA LHR-SJC)
    auto = {
        'grand_total': 129162,
        'p2p_total': 78110,
        'home_total': 48115,
        'dest_total': 2937,
        'load_factor': 0.829,
        'home_results': [
            {'city': 'PAR', 'base_demand': 25000, 'qsi_capture': 0.2313, 'forecast': 5783},
            {'city': 'AMS', 'base_demand': 15000, 'qsi_capture': 0.2077, 'forecast': 3116},
            {'city': 'MAN', 'base_demand': 8000, 'qsi_capture': 0.3990, 'forecast': 3192},
        ],
        'dest_results': [
            {'city': 'LAX', 'base_demand': 5000, 'qsi_capture': 0.05, 'forecast': 250},
        ],
    }

    # Simulate analyst data
    analyst = AnalystData(
        total_pax=129162,
        p2p_total=78110,
        home_cnx_total=48115,
        dest_cnx_total=2937,
        load_factor=82.9,
        home_cities=[
            {'city': 'PAR', 'base_demand': 25000, 'qsi_capture': 0.0400, 'forecast': 1000, 'side': 'home'},
            {'city': 'AMS', 'base_demand': 15000, 'qsi_capture': 0.0155, 'forecast': 233, 'side': 'home'},
            {'city': 'MAN', 'base_demand': 8000, 'qsi_capture': 0.0318, 'forecast': 254, 'side': 'home'},
        ],
        dest_cities=[
            {'city': 'LAX', 'base_demand': 5000, 'qsi_capture': 0.05, 'forecast': 250, 'side': 'dest'},
        ],
    )

    comparison = LearningComparison(auto, analyst)
    report = comparison.run()
    report.print_summary()

    print(f"\nImplied QSI adjustment: {report.implied_qsi_adjustment:.3f}")
    print(f"Learnings extracted: {len(report.learnings)}")
    print("\n=== Self-test complete ===")
