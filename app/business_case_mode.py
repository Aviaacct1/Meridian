#!/usr/bin/env python3
"""
Avia Solutions  Business Case Mode (Chat 19)
===============================================
Goal-seek engine for airport route pitches.

Business Case Mode works BACKWARDS from airline targets to test whether
they're achievable given the market data. This is the core differentiator
between the airport sales tool and the honest forecast.

Two fundamentally different modes:
  Forecast Mode   "Here's what the data says"
  Business Case   "Can the airline hit 70% LF in Year 1 and 82% at maturity?
                    If so, what assumptions are required? If not, where does it break?"

This module does NOT fabricate demand. It:
  1. Runs the standard forecast pipeline to get the base case
  2. Tests target load factors against capacity
  3. Identifies which parameters would need to change to hit targets
  4. Runs multi-parameter sensitivity to map the achievability surface
  5. Reports a clear verdict with the assumptions required
  6. Generates an airline-ready sensitivity & assumptions summary

The output tells the airport route development team:
  - Whether targets are achievable (YES / MARGINAL / NO)
  - The most sensitive parameters (where the airline will push back)
  - What "best honest case" assumptions look like
  - Specific risk flags an airline network planner would raise

Integration:
  from business_case_mode import BusinessCaseEngine
  engine = BusinessCaseEngine(config, pipeline_results)
  bc_results = engine.run()

Dependencies:
  - route_config.py (Chat 12)
  - closed_loop_pipeline_v2.py (Chat 12)
  - input_validator.py (Chat 14)
  - assumptions_log.py (Chat 18)
"""

from config import OUTPUT_DIR, ensure_output_dir
import os
import math
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ============================================================================
# CONSTANTS
# ============================================================================

# Verdict thresholds
VERDICT_YES_MARGIN = 0.03        # Within 3% of target = YES
VERDICT_MARGINAL_MARGIN = 0.10   # Within 10% = MARGINAL
# Beyond 10% gap = NO

# Ramp profiles (year-over-year growth from Year 1 to maturity)
RAMP_PROFILES = {
    'conservative': {
        'label': 'Conservative Ramp',
        'description': 'Slow build typical of competitive long-haul markets',
        'yr1_fraction': 0.75,   # Y1 achieves 75% of mature forecast
        'yr2_fraction': 0.88,
        'yr3_fraction': 0.95,
        'yr4_fraction': 1.00,
    },
    'standard': {
        'label': 'Standard Ramp',
        'description': 'Typical new route ramp-up on underserved market',
        'yr1_fraction': 0.82,
        'yr2_fraction': 0.92,
        'yr3_fraction': 1.00,
    },
    'aggressive': {
        'label': 'Aggressive Ramp',
        'description': 'Fast fill on highly stimulated or VFR-heavy route',
        'yr1_fraction': 0.90,
        'yr2_fraction': 0.97,
        'yr3_fraction': 1.00,
    },
    'hub_carrier': {
        'label': 'Hub Carrier Ramp',
        'description': 'Hub feed route with immediate connecting traffic base',
        'yr1_fraction': 0.88,
        'yr2_fraction': 0.95,
        'yr3_fraction': 1.00,
    },
}

# Sensitivity parameter ranges
SENSITIVITY_PARAMS = {
    'capture_rate': {
        'label': 'P2P Capture Rate',
        'description': 'Share of addressable market captured by the new service',
        'range': (-0.30, 0.30),
        'steps': 7,
        'airline_concern': 'High  airlines scrutinise capture rate assumptions heavily',
    },
    'stimulation': {
        'label': 'Demand Stimulation',
        'description': 'New demand generated beyond existing market size',
        'range': (-0.30, 0.30),
        'steps': 7,
        'airline_concern': 'Very High  most disputed assumption in route pitches',
    },
    'growth': {
        'label': 'Base Demand Growth',
        'description': 'Annual growth applied to base year demand',
        'range': (-0.50, 0.50),
        'steps': 7,
        'airline_concern': 'Medium  depends on economic outlook and data vintage',
    },
    'connecting_share': {
        'label': 'Connecting Traffic Share',
        'description': 'Proportion of total from connecting vs P2P',
        'range': (-0.30, 0.30),
        'steps': 7,
        'airline_concern': 'High for hub routes  airlines have own hub data',
    },
    'frequency': {
        'label': 'Weekly Frequency',
        'description': 'Service frequency impact on demand and costs',
        'discrete_values': [3, 4, 5, 6, 7, 10, 14],
        'airline_concern': 'Critical  determines aircraft utilisation and crew costs',
    },
}


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class TargetSet:
    """Business case targets from the airport/client."""
    load_factor_y1: float = 0.70
    load_factor_mature: float = 0.82
    ramp_years: int = 3
    ramp_profile: str = 'standard'
    p2p_split_target: Optional[float] = None       # P2P as fraction of total
    cnx_balance_target: Optional[float] = None      # Home hub cnx / total cnx
    min_frequency: Optional[int] = None
    max_frequency: Optional[int] = None

    @property
    def ramp(self) -> dict:
        return RAMP_PROFILES.get(self.ramp_profile, RAMP_PROFILES['standard'])


@dataclass
class SensitivityPoint:
    """Single point in a sensitivity sweep."""
    param_name: str
    factor: float
    total_pax: float
    load_factor: float
    p2p_pax: float
    cnx_pax: float
    delta_pct: float
    hits_y1_target: bool
    hits_mature_target: bool


@dataclass
class ParameterGap:
    """How much a parameter needs to change to hit a target."""
    param_name: str
    current_value: float
    required_value: float
    change_pct: float
    achievable: bool
    risk_level: str       # 'low', 'medium', 'high', 'very_high'
    analyst_note: str


@dataclass
class BusinessCaseVerdict:
    """The final verdict on whether targets are achievable."""
    verdict: str          # 'YES', 'MARGINAL', 'NO'
    confidence: str       # 'high', 'medium', 'low'
    headline: str
    summary: str
    base_forecast: float
    base_load_factor: float
    y1_target_pax: float
    mature_target_pax: float
    annual_capacity: float
    gap_to_y1_pct: float
    gap_to_mature_pct: float
    ramp_profile_used: str
    parameter_gaps: List[ParameterGap]
    sensitivity_results: Dict[str, List[SensitivityPoint]]
    risk_flags: List[str]
    best_case_assumptions: Dict[str, Any]
    worst_case_assumptions: Dict[str, Any]
    airline_pushback_areas: List[str]
    year_by_year: List[Dict]


# ============================================================================
# BUSINESS CASE ENGINE
# ============================================================================

class BusinessCaseEngine:
    """
    Goal-seek engine that tests whether airline targets are achievable.

    Usage:
        engine = BusinessCaseEngine(config, pipeline_results)
        verdict = engine.run(targets)
    """

    def __init__(self, config, pipeline_results: Dict):
        """
        config: RouteConfig from pipeline
        pipeline_results: dict from run_pipeline() containing
            p2p_total, p2p_details, home_total, home_results,
            dest_total, dest_results, grand_total, load_factor
        """
        self.config = config
        self.results = pipeline_results
        self.audit = []

    def log(self, msg: str):
        self.audit.append(msg)

    def run(self, targets: Optional[TargetSet] = None) -> BusinessCaseVerdict:
        """Execute the full business case analysis."""
        if targets is None:
            targets = TargetSet()

        self.log(f"{'='*60}")
        self.log(f"BUSINESS CASE ANALYSIS")
        self.log(f"{'='*60}")

        # Step 1: Base case metrics
        base = self._extract_base_case()
        self.log(f"Base forecast: {base['grand_total']:,.0f} pax, "
                 f"LF: {base['load_factor']:.1%}")

        # Step 2: Compute targets
        capacity = base['annual_capacity']
        y1_target_pax = targets.load_factor_y1 * capacity
        mature_target_pax = targets.load_factor_mature * capacity

        # Apply ramp to get Y1 forecast from mature forecast
        ramp = targets.ramp
        y1_fraction = ramp['yr1_fraction']
        # The base forecast is the "mature" (unconstrained) forecast
        y1_forecast = base['grand_total'] * y1_fraction
        mature_forecast = base['grand_total']

        self.log(f"Capacity: {capacity:,.0f} seats/year")
        self.log(f"Y1 target: {y1_target_pax:,.0f} ({targets.load_factor_y1:.0%} LF)")
        self.log(f"Y1 forecast (with {ramp['label']} ramp): {y1_forecast:,.0f}")
        self.log(f"Mature target: {mature_target_pax:,.0f} ({targets.load_factor_mature:.0%} LF)")
        self.log(f"Mature forecast: {mature_forecast:,.0f}")

        # Step 3: Gap analysis
        gap_y1 = (y1_target_pax - y1_forecast) / y1_target_pax if y1_target_pax > 0 else 0
        gap_mature = (mature_target_pax - mature_forecast) / mature_target_pax if mature_target_pax > 0 else 0

        self.log(f"Gap to Y1 target: {gap_y1:+.1%}")
        self.log(f"Gap to mature target: {gap_mature:+.1%}")

        # Step 4: Sensitivity analysis
        sensitivity = self._run_sensitivity(base, targets)

        # Step 5: Parameter gap analysis (what needs to change)
        param_gaps = self._compute_parameter_gaps(base, targets, sensitivity)

        # Step 6: Risk flags
        risk_flags = self._assess_risks(base, targets, param_gaps)

        # Step 7: Airline pushback areas
        pushback = self._identify_pushback_areas(base, targets, param_gaps)

        # Step 8: Year-by-year build
        year_by_year = self._build_year_by_year(base, targets)

        # Step 9: Best/worst case
        best_case = self._compute_best_case(base, sensitivity)
        worst_case = self._compute_worst_case(base, sensitivity)

        # Step 10: Verdict
        verdict, confidence, headline, summary = self._determine_verdict(
            gap_y1, gap_mature, param_gaps, risk_flags, targets)

        self.log(f"\nVERDICT: {verdict} ({confidence} confidence)")
        self.log(f"  {headline}")

        return BusinessCaseVerdict(
            verdict=verdict,
            confidence=confidence,
            headline=headline,
            summary=summary,
            base_forecast=base['grand_total'],
            base_load_factor=base['load_factor'],
            y1_target_pax=y1_target_pax,
            mature_target_pax=mature_target_pax,
            annual_capacity=capacity,
            gap_to_y1_pct=gap_y1,
            gap_to_mature_pct=gap_mature,
            ramp_profile_used=targets.ramp_profile,
            parameter_gaps=param_gaps,
            sensitivity_results=sensitivity,
            risk_flags=risk_flags,
            best_case_assumptions=best_case,
            worst_case_assumptions=worst_case,
            airline_pushback_areas=pushback,
            year_by_year=year_by_year,
        )

    # 
    # STEP 1: Extract base case from pipeline results
    # 

    def _extract_base_case(self) -> Dict:
        """Extract key metrics from pipeline results."""
        r = self.results
        p2p = r.get('p2p_total', 0)
        home = r.get('home_total', 0)
        dest = r.get('dest_total', 0)
        grand = r.get('grand_total', 0)

        cfg = self.config
        capacity = cfg.annual_capacity if hasattr(cfg, 'annual_capacity') else (
            cfg.seats * cfg.frequency * 52 * 2 if hasattr(cfg, 'seats') else 0)

        lf = grand / capacity if capacity > 0 else 0

        # Extract P2P details for sensitivity
        p2p_details = r.get('p2p_details', [])
        home_results = r.get('home_results', [])
        dest_results = r.get('dest_results', [])

        return {
            'p2p_total': p2p,
            'home_total': home,
            'dest_total': dest,
            'cnx_total': home + dest,
            'grand_total': grand,
            'load_factor': lf,
            'annual_capacity': capacity,
            'p2p_share': p2p / grand if grand > 0 else 0,
            'cnx_share': (home + dest) / grand if grand > 0 else 0,
            'home_cnx_share': home / (home + dest) if (home + dest) > 0 else 0,
            'p2p_details': p2p_details,
            'home_results': home_results,
            'dest_results': dest_results,
            'frequency': cfg.frequency if hasattr(cfg, 'frequency') else 7,
            'seats': cfg.seats if hasattr(cfg, 'seats') else 0,
        }

    # 
    # STEP 4: Multi-parameter sensitivity
    # 

    def _run_sensitivity(self, base: Dict, targets: TargetSet) -> Dict[str, List[SensitivityPoint]]:
        """Run sensitivity on each key parameter."""
        results = {}

        capacity = base['annual_capacity']
        y1_target = targets.load_factor_y1 * capacity
        mature_target = targets.load_factor_mature * capacity
        ramp = targets.ramp

        # Capture rate sensitivity
        results['capture_rate'] = self._sweep_multiplier(
            base, 'capture_rate', targets,
            lo=-0.30, hi=0.30, steps=7)

        # Stimulation sensitivity
        results['stimulation'] = self._sweep_multiplier(
            base, 'stimulation', targets,
            lo=-0.30, hi=0.30, steps=7)

        # Growth sensitivity
        results['growth'] = self._sweep_multiplier(
            base, 'growth', targets,
            lo=-0.50, hi=0.50, steps=7)

        # Connecting share sensitivity (scale connecting up/down)
        results['connecting_share'] = self._sweep_multiplier(
            base, 'connecting_share', targets,
            lo=-0.30, hi=0.30, steps=7)

        # Frequency sensitivity (discrete)
        results['frequency'] = self._sweep_frequency(base, targets)

        return results

    def _sweep_multiplier(self, base: Dict, param: str, targets: TargetSet,
                          lo: float, hi: float, steps: int) -> List[SensitivityPoint]:
        """Sweep a multiplier across a parameter range."""
        points = []
        capacity = base['annual_capacity']
        y1_target = targets.load_factor_y1 * capacity
        mature_target = targets.load_factor_mature * capacity
        ramp = targets.ramp

        for i in range(steps):
            factor = 1.0 + lo + (hi - lo) * i / (steps - 1) if steps > 1 else 1.0

            if param == 'capture_rate':
                # Scale P2P capture rates
                p2p_adj = base['p2p_total'] * factor
                cnx_adj = base['cnx_total']
            elif param == 'stimulation':
                # Scale both P2P and connecting (stimulation affects addressable market)
                p2p_adj = base['p2p_total'] * factor
                cnx_adj = base['cnx_total'] * factor
            elif param == 'growth':
                # Growth affects base demand before capture
                # Approximate: if growth changes by X%, total changes proportionally
                p2p_adj = base['p2p_total'] * factor
                cnx_adj = base['cnx_total'] * factor
            elif param == 'connecting_share':
                # Scale connecting only, keep P2P fixed
                p2p_adj = base['p2p_total']
                cnx_adj = base['cnx_total'] * factor
            else:
                p2p_adj = base['p2p_total'] * factor
                cnx_adj = base['cnx_total'] * factor

            total = p2p_adj + cnx_adj
            lf = total / capacity if capacity > 0 else 0
            delta = (total - base['grand_total']) / base['grand_total'] * 100 if base['grand_total'] > 0 else 0
            y1_est = total * ramp['yr1_fraction']

            points.append(SensitivityPoint(
                param_name=param,
                factor=factor,
                total_pax=total,
                load_factor=lf,
                p2p_pax=p2p_adj,
                cnx_pax=cnx_adj,
                delta_pct=delta,
                hits_y1_target=y1_est >= y1_target,
                hits_mature_target=total >= mature_target,
            ))

        return points

    def _sweep_frequency(self, base: Dict, targets: TargetSet) -> List[SensitivityPoint]:
        """Discrete frequency sweep."""
        points = []
        base_freq = base['frequency']
        base_seats = base['seats']
        ramp = targets.ramp

        for freq in SENSITIVITY_PARAMS['frequency']['discrete_values']:
            # Capacity changes with frequency
            new_capacity = base_seats * freq * 52 * 2
            # Demand scales sub-linearly with frequency (S-curve)
            # Rough approximation: demand ~ freq^0.6 (industry rule of thumb)
            freq_factor = (freq / base_freq) ** 0.6 if base_freq > 0 else 1.0
            new_total = base['grand_total'] * freq_factor
            new_p2p = base['p2p_total'] * freq_factor
            new_cnx = base['cnx_total'] * freq_factor
            new_lf = new_total / new_capacity if new_capacity > 0 else 0
            delta = (new_total - base['grand_total']) / base['grand_total'] * 100 if base['grand_total'] > 0 else 0

            y1_target = targets.load_factor_y1 * new_capacity
            mature_target = targets.load_factor_mature * new_capacity
            y1_est = new_total * ramp['yr1_fraction']

            points.append(SensitivityPoint(
                param_name='frequency',
                factor=freq,  # Actual frequency value, not a multiplier
                total_pax=new_total,
                load_factor=new_lf,
                p2p_pax=new_p2p,
                cnx_pax=new_cnx,
                delta_pct=delta,
                hits_y1_target=y1_est >= y1_target,
                hits_mature_target=new_total >= mature_target,
            ))

        return points

    # 
    # STEP 5: Parameter gap analysis
    # 

    def _compute_parameter_gaps(self, base: Dict, targets: TargetSet,
                                 sensitivity: Dict) -> List[ParameterGap]:
        """Determine what each parameter needs to change to hit targets."""
        gaps = []
        capacity = base['annual_capacity']
        mature_target = targets.load_factor_mature * capacity

        if base['grand_total'] >= mature_target:
            # Already hitting target  no gaps needed
            return gaps

        required_multiplier = mature_target / base['grand_total'] if base['grand_total'] > 0 else float('inf')

        # For each continuous parameter, what multiplier hits the target?
        for param in ['capture_rate', 'stimulation', 'growth', 'connecting_share']:
            points = sensitivity.get(param, [])
            if not points:
                continue

            # Find the factor that gets closest to target
            best_point = min(points, key=lambda p: abs(p.total_pax - mature_target))
            hits_target = any(p.hits_mature_target for p in points)

            # Compute required change
            if param == 'capture_rate':
                current_label = f"{base['p2p_share']:.0%} P2P capture"
                required_change = (required_multiplier - 1.0)
            elif param == 'stimulation':
                current_label = "1.0x base stimulation"
                required_change = (required_multiplier - 1.0)
            elif param == 'growth':
                current_label = "base growth rate"
                required_change = (required_multiplier - 1.0)
            elif param == 'connecting_share':
                current_label = f"{base['cnx_share']:.0%} connecting"
                required_change = (required_multiplier - 1.0)
            else:
                current_label = param
                required_change = 0

            # Risk assessment
            if abs(required_change) < 0.05:
                risk = 'low'
                note = 'Minor adjustment  well within normal range'
            elif abs(required_change) < 0.15:
                risk = 'medium'
                note = 'Moderate adjustment  requires justification'
            elif abs(required_change) < 0.25:
                risk = 'high'
                note = 'Significant adjustment  airline will challenge this'
            else:
                risk = 'very_high'
                note = 'Extreme adjustment  unlikely to be accepted'

            gaps.append(ParameterGap(
                param_name=param,
                current_value=1.0,
                required_value=required_multiplier if param != 'connecting_share' else (
                    1.0 + required_change),
                change_pct=required_change * 100,
                achievable=hits_target,
                risk_level=risk,
                analyst_note=note,
            ))

        # Frequency gap
        freq_points = sensitivity.get('frequency', [])
        if freq_points:
            hitting_freqs = [p for p in freq_points if p.hits_mature_target]
            if hitting_freqs:
                best_freq = min(hitting_freqs, key=lambda p: p.factor)
                change = (best_freq.factor - base['frequency']) / base['frequency'] * 100
                risk = 'low' if abs(change) < 20 else 'medium' if abs(change) < 50 else 'high'
                gaps.append(ParameterGap(
                    param_name='frequency',
                    current_value=base['frequency'],
                    required_value=best_freq.factor,
                    change_pct=change,
                    achievable=True,
                    risk_level=risk,
                    analyst_note=f"Increase from {base['frequency']}x to {int(best_freq.factor)}x weekly",
                ))
            else:
                gaps.append(ParameterGap(
                    param_name='frequency',
                    current_value=base['frequency'],
                    required_value=14,
                    change_pct=100,
                    achievable=False,
                    risk_level='very_high',
                    analyst_note='Target not achievable through frequency alone',
                ))

        return gaps

    # 
    # STEP 6: Risk assessment
    # 

    def _assess_risks(self, base: Dict, targets: TargetSet,
                      param_gaps: List[ParameterGap]) -> List[str]:
        """Identify specific risk flags."""
        flags = []

        # Load factor risks
        if targets.load_factor_y1 > 0.80:
            flags.append("Y1 target LF >80% is aggressive for a new route  "
                         "most airlines expect 65-75% in Year 1")
        if targets.load_factor_mature > 0.90:
            flags.append("Mature LF >90% leaves no room for seasonal variation or "
                         "competitive response  airlines will flag this")

        # Connecting traffic risks
        if base['cnx_share'] > 0.50:
            flags.append(f"Connecting traffic is {base['cnx_share']:.0%} of total  "
                         f"high dependency on hub connectivity makes forecast "
                         f"vulnerable to schedule changes")

        # P2P concentration
        if base['p2p_share'] > 0.85:
            flags.append(f"P2P is {base['p2p_share']:.0%} of total  "
                         "route is highly sensitive to local demand; "
                         "limited upside from connecting")

        # Stimulation dependency
        p2p_details = base.get('p2p_details', [])
        high_stim = [d for d in p2p_details
                     if d.get('stimulation', 1.0) > 1.20]
        if high_stim:
            names = ', '.join(d['name'] for d in high_stim[:3])
            flags.append(f"Stimulation >1.20x assumed for {names}  "
                         "airline will ask for evidence of demand generation")

        # Parameter gap risks
        very_high_risks = [g for g in param_gaps if g.risk_level == 'very_high']
        if very_high_risks:
            names = ', '.join(g.param_name for g in very_high_risks)
            flags.append(f"Extreme parameter adjustments needed ({names})  "
                         "business case requires fundamental reassessment")

        # Ramp profile
        if targets.ramp_profile == 'aggressive':
            flags.append("Aggressive ramp profile assumed  "
                         "Y1 at 90% of mature. Most new routes take 2-3 years "
                         "to reach maturity")

        # Gap analysis
        if base['grand_total'] < targets.load_factor_y1 * base['annual_capacity'] * 0.8:
            flags.append("Forecast is >20% below Y1 target  "
                         "significant assumption changes needed. "
                         "Consider whether the route is viable at the proposed capacity")

        return flags

    # 
    # STEP 7: Airline pushback areas
    # 

    def _identify_pushback_areas(self, base: Dict, targets: TargetSet,
                                  param_gaps: List[ParameterGap]) -> List[str]:
        """Identify where an airline network planner will push back."""
        areas = []

        # High capture rates
        p2p_details = base.get('p2p_details', [])
        high_capture = [d for d in p2p_details if d.get('capture', 0) > 0.35]
        if high_capture:
            names = ', '.join(d['name'] for d in high_capture[:3])
            areas.append(f"Capture rates above 35% for {names}  "
                         "airline will benchmark against their own route data")

        # Stimulation assumptions
        stim_segments = [d for d in p2p_details if d.get('stimulation', 1.0) > 1.10]
        if stim_segments:
            areas.append("Demand stimulation assumptions  "
                         "airline will want evidence from proxy markets "
                         "or IATA stimulation curves")

        # Connecting traffic methodology
        if base['cnx_share'] > 0.30:
            areas.append("Connecting traffic forecast  "
                         "airline will compare against their own hub feed data "
                         "and challenge QSI assumptions for their network")

        # Growth assumptions
        areas.append("Base demand growth rate  "
                     "airline will reference their own traffic data "
                     "and economic outlook for the market")

        # Load factor trajectory
        if targets.load_factor_mature > 0.85:
            areas.append(f"Mature load factor target of {targets.load_factor_mature:.0%}  "
                         "airline will compare against route portfolio average")

        # Source data vintage
        areas.append("Data vintage  "
                     "airline will ask when demand data was sourced "
                     "and whether post-COVID recovery is reflected")

        return areas

    # 
    # STEP 8: Year-by-year build
    # 

    def _build_year_by_year(self, base: Dict, targets: TargetSet) -> List[Dict]:
        """Build year-by-year ramp forecast."""
        ramp = targets.ramp
        capacity = base['annual_capacity']
        mature_forecast = base['grand_total']
        years = []

        # Build all years from Y1 to maturity
        fractions = []
        for key in sorted(ramp.keys()):
            if key.startswith('yr') and key.endswith('_fraction'):
                fractions.append(ramp[key])

        for yr, fraction in enumerate(fractions, 1):
            pax = mature_forecast * fraction
            lf = pax / capacity if capacity > 0 else 0
            p2p = base['p2p_total'] * fraction
            cnx = base['cnx_total'] * fraction

            target_lf = targets.load_factor_y1 if yr == 1 else (
                targets.load_factor_mature if yr >= len(fractions) else
                targets.load_factor_y1 + (targets.load_factor_mature - targets.load_factor_y1) *
                (yr - 1) / (len(fractions) - 1)
            )
            target_pax = target_lf * capacity
            gap = pax - target_pax

            years.append({
                'year': yr,
                'ramp_fraction': fraction,
                'forecast_pax': round(pax),
                'forecast_lf': lf,
                'target_lf': target_lf,
                'target_pax': round(target_pax),
                'gap_pax': round(gap),
                'gap_pct': gap / target_pax if target_pax > 0 else 0,
                'p2p_pax': round(p2p),
                'cnx_pax': round(cnx),
                'verdict': 'ABOVE' if gap >= 0 else 'BELOW',
            })

        return years

    # 
    # STEP 9: Best/worst case scenarios
    # 

    def _compute_best_case(self, base: Dict, sensitivity: Dict) -> Dict:
        """Best plausible case  optimistic but defensible."""
        best = {
            'capture_factor': 1.15,
            'stimulation_factor': 1.15,
            'growth_factor': 1.20,
            'cnx_factor': 1.10,
            'label': 'Optimistic but defensible upside',
            'description': 'Higher capture (strong brand/schedule), '
                          'above-average stimulation, strong economic growth',
        }
        # Apply all upside factors
        p2p_up = base['p2p_total'] * best['capture_factor'] * best['stimulation_factor'] * best['growth_factor']
        cnx_up = base['cnx_total'] * best['stimulation_factor'] * best['growth_factor'] * best['cnx_factor']
        total = p2p_up + cnx_up
        lf = total / base['annual_capacity'] if base['annual_capacity'] > 0 else 0
        best['total_pax'] = round(total)
        best['load_factor'] = lf
        return best

    def _compute_worst_case(self, base: Dict, sensitivity: Dict) -> Dict:
        """Worst plausible case  conservative stress test."""
        worst = {
            'capture_factor': 0.80,
            'stimulation_factor': 0.90,
            'growth_factor': 0.85,
            'cnx_factor': 0.85,
            'label': 'Downside stress test',
            'description': 'Competitive response reduces capture, '
                          'lower stimulation realised, economic weakness',
        }
        p2p_dn = base['p2p_total'] * worst['capture_factor'] * worst['stimulation_factor'] * worst['growth_factor']
        cnx_dn = base['cnx_total'] * worst['stimulation_factor'] * worst['growth_factor'] * worst['cnx_factor']
        total = p2p_dn + cnx_dn
        lf = total / base['annual_capacity'] if base['annual_capacity'] > 0 else 0
        worst['total_pax'] = round(total)
        worst['load_factor'] = lf
        return worst

    # 
    # STEP 10: Final verdict
    # 

    def _determine_verdict(self, gap_y1: float, gap_mature: float,
                           param_gaps: List[ParameterGap],
                           risk_flags: List[str],
                           targets: TargetSet) -> Tuple[str, str, str, str]:
        """Determine YES / MARGINAL / NO verdict."""

        # Count high/very_high risk parameters
        high_risks = sum(1 for g in param_gaps if g.risk_level in ('high', 'very_high'))
        very_high_risks = sum(1 for g in param_gaps if g.risk_level == 'very_high')

        # Determine verdict
        if gap_mature <= VERDICT_YES_MARGIN and gap_y1 <= VERDICT_YES_MARGIN:
            # Forecast meets or exceeds targets
            verdict = 'YES'
            if high_risks == 0:
                confidence = 'high'
                headline = "Targets achievable  forecast supports the business case"
                summary = (
                    f"The base forecast of {self.results['grand_total']:,.0f} passengers "
                    f"({self.results['load_factor']:.1%} LF) meets the target load factor "
                    f"of {targets.load_factor_mature:.0%} at maturity. "
                    f"The Year 1 ramp ({targets.ramp['label']}) also achieves the "
                    f"{targets.load_factor_y1:.0%} Y1 target. "
                    f"No aggressive assumptions are required."
                )
            else:
                confidence = 'medium'
                headline = "Targets achievable but with noted risks"
                summary = (
                    f"The base forecast meets targets, but {high_risks} parameter(s) "
                    f"carry elevated risk. The airline will likely challenge these. "
                    f"Prepare supporting evidence for: "
                    f"{', '.join(g.param_name for g in param_gaps if g.risk_level in ('high', 'very_high'))}."
                )

        elif gap_mature <= VERDICT_MARGINAL_MARGIN:
            verdict = 'MARGINAL'
            confidence = 'medium' if gap_mature <= 0.05 else 'low'
            headline = "Targets achievable with moderate assumption adjustments"
            summary = (
                f"The base forecast is {abs(gap_mature):.1%} below the mature target. "
                f"Adjustments to {', '.join(g.param_name for g in param_gaps[:3])} "
                f"could close the gap. This is a plausible but not conservative case. "
                f"The sensitivity analysis shows which parameters are most impactful."
            )

        else:
            # Target not achievable
            verdict = 'NO'
            if very_high_risks > 1:
                confidence = 'high'  # High confidence that it DOESN'T work
                headline = "Targets not supportable  route may not be viable at proposed capacity"
                summary = (
                    f"The base forecast of {self.results['grand_total']:,.0f} passengers "
                    f"({self.results['load_factor']:.1%} LF) is {abs(gap_mature):.0%} below "
                    f"the {targets.load_factor_mature:.0%} mature target. "
                    f"Multiple parameters would require extreme adjustments to close the gap. "
                    f"Consider: reduced capacity (smaller aircraft or lower frequency), "
                    f"revised targets, or alternative route options."
                )
            else:
                confidence = 'medium'
                headline = "Targets not achievable with standard assumptions"
                summary = (
                    f"The base forecast falls {abs(gap_mature):.0%} short of the mature target. "
                    f"While individual parameter adjustments could narrow the gap, "
                    f"the combination required would be difficult to defend to an airline. "
                    f"Review the sensitivity analysis for the most promising adjustments."
                )

        return verdict, confidence, headline, summary


# ============================================================================
# OUTPUT WRITER  Business Case Excel workbook
# ============================================================================

class BusinessCaseWriter:
    """Produces a client-ready Business Case workbook."""

    # Avia Solutions branding
    BLUE = '003366'
    MID_BLUE = '4472C4'
    LIGHT_BLUE = 'D6E4F0'
    GREEN = '006100'
    RED = 'C00000'
    YELLOW = 'FFF2CC'
    LIGHT_GREEN = 'E2EFDA'

    def __init__(self):
        self.hf = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        self.hfill = PatternFill('solid', fgColor=self.BLUE)
        self.tf = Font(name='Arial', bold=True, color=self.BLUE, size=14)
        self.sf = Font(name='Arial', bold=True, color=self.BLUE, size=12)
        self.df = Font(name='Arial', size=10)
        self.bf = Font(name='Arial', size=10, bold=True)
        self.inp = Font(name='Arial', size=10, color='0000FF')
        self.green_f = Font(name='Arial', size=10, color=self.GREEN)
        self.red_f = Font(name='Arial', size=10, color=self.RED)
        self.verdict_yes = Font(name='Arial', size=14, bold=True, color=self.GREEN)
        self.verdict_no = Font(name='Arial', size=14, bold=True, color=self.RED)
        self.verdict_marginal = Font(name='Arial', size=14, bold=True, color='FF8C00')
        self.warn_fill = PatternFill('solid', fgColor='FFF2CC')
        self.green_fill = PatternFill('solid', fgColor=self.LIGHT_GREEN)
        self.light_fill = PatternFill('solid', fgColor=self.LIGHT_BLUE)
        self.thin = Border(
            left=Side('thin', color='B4B4B4'),
            right=Side('thin', color='B4B4B4'),
            top=Side('thin', color='B4B4B4'),
            bottom=Side('thin', color='B4B4B4'))

    def _hdr(self, ws, row, cols):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row, c, v)
            cell.font = self.hf
            cell.fill = self.hfill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.thin

    def _cell(self, ws, r, c, val, font=None, fill=None, fmt=None):
        cell = ws.cell(r, c, val)
        cell.font = font or self.df
        cell.border = self.thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        return cell

    def write(self, config, verdict: BusinessCaseVerdict, path: str):
        """Write the complete business case workbook."""
        if openpyxl is None:
            raise ImportError("openpyxl required")

        wb = openpyxl.Workbook()

        self._write_verdict_sheet(wb, config, verdict)
        self._write_year_by_year(wb, config, verdict)
        self._write_sensitivity(wb, config, verdict)
        self._write_parameter_gaps(wb, config, verdict)
        self._write_risk_flags(wb, config, verdict)
        self._write_scenarios(wb, config, verdict)

        os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
        wb.save(path)
        return path

    #  Sheet 1: Verdict 

    def _write_verdict_sheet(self, wb, config, v: BusinessCaseVerdict):
        ws = wb.active
        ws.title = 'Verdict'
        ws.sheet_properties.tabColor = self.BLUE

        # Title
        ws.merge_cells('A1:F1')
        ws.cell(1, 1, 'BUSINESS CASE ASSESSMENT').font = self.tf
        ws.cell(2, 1, f'{config.airline_name} {config.home_airport_code}-'
                       f'{config.dest_airport_code}').font = self.sf
        ws.cell(3, 1, f'Generated: {datetime.now():%Y-%m-%d %H:%M}').font = self.df

        # Verdict box
        r = 5
        verdict_font = {
            'YES': self.verdict_yes,
            'MARGINAL': self.verdict_marginal,
            'NO': self.verdict_no,
        }.get(v.verdict, self.bf)

        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(r, 1, f'VERDICT: {v.verdict}').font = verdict_font
        r += 1
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(r, 1, v.headline).font = self.bf
        r += 1
        ws.merge_cells(f'A{r}:F{r+1}')
        ws.cell(r, 1, v.summary).font = self.df
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
        r += 3

        # Key metrics table
        metrics = [
            ('Annual Capacity', f'{v.annual_capacity:,.0f}', 'seats'),
            ('Base Forecast', f'{v.base_forecast:,.0f}', 'passengers'),
            ('Base Load Factor', f'{v.base_load_factor:.1%}', ''),
            ('', '', ''),
            ('Year 1 Target LF', f'{v.y1_target_pax / v.annual_capacity:.0%}'
             if v.annual_capacity else 'N/A', ''),
            ('Year 1 Target Pax', f'{v.y1_target_pax:,.0f}', ''),
            ('Gap to Y1', f'{v.gap_to_y1_pct:+.1%}',
             '' if v.gap_to_y1_pct <= 0 else ''),
            ('', '', ''),
            ('Mature Target LF', f'{v.mature_target_pax / v.annual_capacity:.0%}'
             if v.annual_capacity else 'N/A', ''),
            ('Mature Target Pax', f'{v.mature_target_pax:,.0f}', ''),
            ('Gap to Mature', f'{v.gap_to_mature_pct:+.1%}',
             '' if v.gap_to_mature_pct <= 0 else ''),
            ('', '', ''),
            ('Ramp Profile', v.ramp_profile_used.replace('_', ' ').title(), ''),
            ('Confidence', v.confidence.title(), ''),
        ]

        for label, val, note in metrics:
            if not label:
                r += 1
                continue
            self._cell(ws, r, 1, label, font=self.bf)
            self._cell(ws, r, 2, val, font=self.inp)
            self._cell(ws, r, 3, note, font=self.df)
            r += 1

        for c in [1, 2, 3, 4, 5, 6]:
            ws.column_dimensions[get_column_letter(c)].width = [30, 18, 12, 12, 12, 12][c-1]

    #  Sheet 2: Year-by-Year 

    def _write_year_by_year(self, wb, config, v: BusinessCaseVerdict):
        ws = wb.create_sheet('Year-by-Year Ramp')
        ws.sheet_properties.tabColor = self.MID_BLUE

        ws.cell(1, 1, 'Year-by-Year Forecast vs Target').font = self.sf
        ws.cell(2, 1, f'Ramp profile: {v.ramp_profile_used.replace("_", " ").title()}').font = self.df

        self._hdr(ws, 4, ['Year', 'Ramp %', 'Forecast Pax', 'Forecast LF',
                          'Target LF', 'Target Pax', 'Gap', 'Gap %', 'Verdict',
                          'P2P', 'Connecting'])

        for i, yr in enumerate(v.year_by_year, 5):
            gap_font = self.green_f if yr['gap_pax'] >= 0 else self.red_f
            gap_fill = self.green_fill if yr['gap_pax'] >= 0 else self.warn_fill
            self._cell(ws, i, 1, f"Year {yr['year']}", font=self.bf)
            self._cell(ws, i, 2, yr['ramp_fraction'], fmt='0%')
            self._cell(ws, i, 3, yr['forecast_pax'], fmt='#,##0')
            self._cell(ws, i, 4, yr['forecast_lf'], fmt='0.0%')
            self._cell(ws, i, 5, yr['target_lf'], fmt='0.0%', font=self.inp)
            self._cell(ws, i, 6, yr['target_pax'], fmt='#,##0', font=self.inp)
            self._cell(ws, i, 7, yr['gap_pax'], fmt='#,##0', font=gap_font, fill=gap_fill)
            self._cell(ws, i, 8, yr['gap_pct'], fmt='+0.0%;-0.0%', font=gap_font)
            self._cell(ws, i, 9, yr['verdict'], font=gap_font)
            self._cell(ws, i, 10, yr['p2p_pax'], fmt='#,##0')
            self._cell(ws, i, 11, yr['cnx_pax'], fmt='#,##0')

        for c in range(1, 12):
            ws.column_dimensions[get_column_letter(c)].width = 14

    #  Sheet 3: Sensitivity 

    def _write_sensitivity(self, wb, config, v: BusinessCaseVerdict):
        ws = wb.create_sheet('Sensitivity Analysis')
        ws.sheet_properties.tabColor = self.MID_BLUE

        ws.cell(1, 1, 'Multi-Parameter Sensitivity Analysis').font = self.sf

        r = 3
        for param_name, points in v.sensitivity_results.items():
            param_info = SENSITIVITY_PARAMS.get(param_name, {})
            label = param_info.get('label', param_name)
            concern = param_info.get('airline_concern', '')

            ws.cell(r, 1, label).font = self.bf
            ws.cell(r, 2, f'Airline concern: {concern}').font = Font(
                name='Arial', size=9, italic=True, color='666666')
            r += 1

            is_freq = param_name == 'frequency'
            cols = (['Frequency' if is_freq else 'Factor',
                     'Total Pax', 'Load Factor', 'P2P', 'Connecting',
                     ' from Base', 'Hits Y1?', 'Hits Mature?'])
            self._hdr(ws, r, cols)
            r += 1

            for pt in points:
                is_base = (abs(pt.factor - 1.0) < 0.001) if not is_freq else (
                    pt.factor == v.base_forecast)  # won't match, but base row highlighted differently
                font = self.bf if (not is_freq and abs(pt.factor - 1.0) < 0.001) else self.df
                fill = self.light_fill if (not is_freq and abs(pt.factor - 1.0) < 0.001) else None

                self._cell(ws, r, 1, int(pt.factor) if is_freq else pt.factor,
                          fmt='#,##0' if is_freq else '0.00', font=font, fill=fill)
                self._cell(ws, r, 2, pt.total_pax, fmt='#,##0', font=font, fill=fill)
                self._cell(ws, r, 3, pt.load_factor, fmt='0.0%', font=font, fill=fill)
                self._cell(ws, r, 4, pt.p2p_pax, fmt='#,##0', font=font, fill=fill)
                self._cell(ws, r, 5, pt.cnx_pax, fmt='#,##0', font=font, fill=fill)
                self._cell(ws, r, 6, pt.delta_pct / 100, fmt='+0.0%;-0.0%', font=font, fill=fill)
                y1_font = self.green_f if pt.hits_y1_target else self.red_f
                mat_font = self.green_f if pt.hits_mature_target else self.red_f
                self._cell(ws, r, 7, '' if pt.hits_y1_target else '', font=y1_font)
                self._cell(ws, r, 8, '' if pt.hits_mature_target else '', font=mat_font)
                r += 1

            r += 2  # gap between parameter sections

        for c in range(1, 9):
            ws.column_dimensions[get_column_letter(c)].width = 14

    #  Sheet 4: Parameter Gaps 

    def _write_parameter_gaps(self, wb, config, v: BusinessCaseVerdict):
        ws = wb.create_sheet('Parameter Gaps')
        ws.sheet_properties.tabColor = self.MID_BLUE

        ws.cell(1, 1, 'What Needs to Change to Hit Targets').font = self.sf
        ws.cell(2, 1, 'Each row shows how much one parameter must change, '
                       'holding others constant').font = Font(
            name='Arial', size=9, italic=True, color='666666')

        self._hdr(ws, 4, ['Parameter', 'Current', 'Required', 'Change %',
                          'Achievable?', 'Risk Level', 'Analyst Note'])

        risk_fills = {
            'low': PatternFill('solid', fgColor='E2EFDA'),
            'medium': PatternFill('solid', fgColor='FFF2CC'),
            'high': PatternFill('solid', fgColor='FCE4D6'),
            'very_high': PatternFill('solid', fgColor='F4CCCC'),
        }

        for i, gap in enumerate(v.parameter_gaps, 5):
            label = SENSITIVITY_PARAMS.get(gap.param_name, {}).get('label', gap.param_name)
            rfill = risk_fills.get(gap.risk_level, None)

            self._cell(ws, i, 1, label, font=self.bf)
            self._cell(ws, i, 2, gap.current_value, fmt='0.00')
            self._cell(ws, i, 3, gap.required_value, fmt='0.00')
            self._cell(ws, i, 4, gap.change_pct / 100, fmt='+0.0%;-0.0%')
            self._cell(ws, i, 5, '' if gap.achievable else '',
                      font=self.green_f if gap.achievable else self.red_f)
            self._cell(ws, i, 6, gap.risk_level.replace('_', ' ').title(),
                      fill=rfill)
            self._cell(ws, i, 7, gap.analyst_note, font=self.df)
            ws.cell(i, 7).alignment = Alignment(wrap_text=True)

        for c, w in enumerate([22, 12, 12, 12, 12, 14, 45], 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    #  Sheet 5: Risk Flags 

    def _write_risk_flags(self, wb, config, v: BusinessCaseVerdict):
        ws = wb.create_sheet('Risk & Pushback')
        ws.sheet_properties.tabColor = 'C00000'

        ws.cell(1, 1, 'Risk Flags & Anticipated Airline Pushback').font = self.sf

        r = 3
        ws.cell(r, 1, 'RISK FLAGS').font = self.bf
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='FCE4D6')
        r += 1
        if v.risk_flags:
            for flag in v.risk_flags:
                ws.cell(r, 1, '').font = Font(name='Arial', size=12, color='C00000')
                ws.cell(r, 2, flag).font = self.df
                ws.cell(r, 2).alignment = Alignment(wrap_text=True)
                r += 1
        else:
            ws.cell(r, 1, '').font = self.df
            ws.cell(r, 2, 'No significant risk flags identified').font = self.green_f
            r += 1

        r += 2
        ws.cell(r, 1, 'AIRLINE PUSHBACK AREAS').font = self.bf
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='FFF2CC')
        r += 1
        for i, area in enumerate(v.airline_pushback_areas, 1):
            ws.cell(r, 1, str(i)).font = self.bf
            ws.cell(r, 2, area).font = self.df
            ws.cell(r, 2).alignment = Alignment(wrap_text=True)
            r += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 80

    #  Sheet 6: Scenarios 

    def _write_scenarios(self, wb, config, v: BusinessCaseVerdict):
        ws = wb.create_sheet('Scenarios')
        ws.sheet_properties.tabColor = self.MID_BLUE

        ws.cell(1, 1, 'Best / Base / Worst Case Scenarios').font = self.sf

        self._hdr(ws, 3, ['Scenario', 'Total Pax', 'Load Factor',
                          'Capture', 'Stimulation', 'Growth', 'Connecting', 'Notes'])

        bc = v.best_case_assumptions
        wc = v.worst_case_assumptions

        scenarios = [
            ('Best Case', bc.get('total_pax', 0), bc.get('load_factor', 0),
             bc.get('capture_factor', 1), bc.get('stimulation_factor', 1),
             bc.get('growth_factor', 1), bc.get('cnx_factor', 1),
             bc.get('description', '')),
            ('Base Case', v.base_forecast, v.base_load_factor,
             1.0, 1.0, 1.0, 1.0, 'Standard forecast assumptions'),
            ('Worst Case', wc.get('total_pax', 0), wc.get('load_factor', 0),
             wc.get('capture_factor', 1), wc.get('stimulation_factor', 1),
             wc.get('growth_factor', 1), wc.get('cnx_factor', 1),
             wc.get('description', '')),
        ]

        for i, (label, pax, lf, cap, stim, grw, cnx, notes) in enumerate(scenarios, 4):
            font = self.bf if label == 'Base Case' else self.df
            fill = self.light_fill if label == 'Base Case' else (
                self.green_fill if label == 'Best Case' else self.warn_fill)
            self._cell(ws, i, 1, label, font=self.bf, fill=fill)
            self._cell(ws, i, 2, pax, fmt='#,##0', font=font)
            self._cell(ws, i, 3, lf, fmt='0.0%', font=font)
            self._cell(ws, i, 4, cap, fmt='0.00x', font=font)
            self._cell(ws, i, 5, stim, fmt='0.00x', font=font)
            self._cell(ws, i, 6, grw, fmt='0.00x', font=font)
            self._cell(ws, i, 7, cnx, fmt='0.00x', font=font)
            self._cell(ws, i, 8, notes, font=self.df)
            ws.cell(i, 8).alignment = Alignment(wrap_text=True)

        for c, w in enumerate([14, 14, 12, 10, 12, 10, 12, 45], 1):
            ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# INTEGRATION: Run business case from pipeline results
# ============================================================================

def run_business_case(config, pipeline_results: Dict,
                      targets: Optional[TargetSet] = None,
                      output_path: Optional[str] = None) -> BusinessCaseVerdict:
    """
    Convenience function to run business case analysis.

    Usage:
        from business_case_mode import run_business_case, TargetSet
        from closed_loop_pipeline_v2 import run_pipeline
        from route_config import RouteConfig

        config = RouteConfig.ba_lhr_sjc()
        results = run_pipeline(config)

        targets = TargetSet(
            load_factor_y1=0.70,
            load_factor_mature=0.82,
            ramp_profile='standard',
        )
        verdict = run_business_case(config, results, targets,
                                     output_path='BusinessCase_BA_LHR_SJC.xlsx')
    """
    engine = BusinessCaseEngine(config, pipeline_results)
    verdict = engine.run(targets)

    if output_path:
        writer = BusinessCaseWriter()
        writer.write(config, verdict, output_path)
        print(f"Business case workbook: {output_path}")

    return verdict


# ============================================================================
# VALIDATION: Test against BA LHR-SJC
# ============================================================================

def validate_ba_lhr_sjc():
    """
    Validate Business Case Mode against the BA LHR-SJC dataset.

    The raw pipeline produces 329,047 pax (211% LF) because connecting
    traffic uses uncalibrated QSI. The CALIBRATED result is 129,162 pax
    (82.9% LF). Business Case Mode must work with BOTH, but our tests
    use a synthetic "calibrated results" dict matching the known target.

    Tests:
    1. With calibrated results (82.9% LF): 70/82 targets  YES
    2. With calibrated results: 85/90 targets  MARGINAL or NO
    3. With calibrated results: 90/95 targets  NO
    4. Full output workbook
    """
    import sys
    from route_config import RouteConfig

    print("=" * 60)
    print("BUSINESS CASE MODE  VALIDATION")
    print("=" * 60)

    config = RouteConfig.ba_lhr_sjc()
    capacity = config.annual_capacity

    # Use calibrated results (the expert-adjusted forecast)
    # This is what the pipeline produces AFTER calibration (Chat 16 integration)
    calibrated_results = {
        'p2p_total': 78110,
        'p2p_details': [
            {'name': 'UK Business', 'base': 71442, 'growth': 0.10,
             'stimulation': 1.15, 'capture': 0.40, 'forecast': 36149},
            {'name': 'UK Leisure/VFR/Primary', 'base': 36386, 'growth': 0.10,
             'stimulation': 1.0, 'capture': 0.25, 'forecast': 10006},
            {'name': 'UK Leisure/VFR/Secondary', 'base': 17449, 'growth': 0.10,
             'stimulation': 1.0, 'capture': 0.25, 'forecast': 4798},
            {'name': 'UK Leisure/VFR/Contested', 'base': 4618, 'growth': 0.10,
             'stimulation': 1.0, 'capture': 0.10, 'forecast': 508},
            {'name': 'US Business', 'base': 65946, 'growth': 0.10,
             'stimulation': 1.15, 'capture': 0.15, 'forecast': 12513},
            {'name': 'US Leisure/VFR/Primary', 'base': 33587, 'growth': 0.10,
             'stimulation': 1.0, 'capture': 0.25, 'forecast': 9236},
            {'name': 'US Leisure/VFR/Secondary', 'base': 16107, 'growth': 0.10,
             'stimulation': 1.0, 'capture': 0.25, 'forecast': 4429},
            {'name': 'US Leisure/VFR/Contested', 'base': 4262, 'growth': 0.10,
             'stimulation': 1.0, 'capture': 0.10, 'forecast': 469},
        ],
        'home_total': 48115,
        'home_results': [],
        'dest_total': 2937,
        'dest_results': [],
        'grand_total': 129162,
        'load_factor': 129162 / capacity,
    }

    base_lf = calibrated_results['grand_total'] / capacity
    print(f"\nCalibrated base: {calibrated_results['grand_total']:,.0f} pax, "
          f"{base_lf:.1%} LF, capacity {capacity:,.0f}")

    # Test 1: Targets forecast should meet (82.9% LF vs 82% target)
    print("\n" + "" * 60)
    print("TEST 1: Achievable targets (70% Y1, 82% mature)")
    print("" * 60)
    t1 = TargetSet(load_factor_y1=0.70, load_factor_mature=0.82,
                   ramp_profile='standard')
    v1 = run_business_case(config, calibrated_results, t1)
    print(f"  Verdict: {v1.verdict} ({v1.confidence} confidence)")
    print(f"  {v1.headline}")
    assert v1.verdict == 'YES', f"Expected YES, got {v1.verdict}"
    print("   PASS  correct verdict")

    # Test 2: Tight targets (82.9% base vs 90% mature = ~8.5% gap)
    print("\n" + "" * 60)
    print("TEST 2: Tight targets (85% Y1, 90% mature)")
    print("" * 60)
    t2 = TargetSet(load_factor_y1=0.85, load_factor_mature=0.90,
                   ramp_profile='standard')
    v2 = run_business_case(config, calibrated_results, t2)
    print(f"  Verdict: {v2.verdict} ({v2.confidence} confidence)")
    print(f"  {v2.headline}")
    assert v2.verdict in ('MARGINAL', 'NO'), f"Expected MARGINAL/NO, got {v2.verdict}"
    print(f"   PASS  correct verdict ({v2.verdict})")

    # Test 3: Unachievable targets
    print("\n" + "" * 60)
    print("TEST 3: Unachievable targets (90% Y1, 95% mature)")
    print("" * 60)
    t3 = TargetSet(load_factor_y1=0.90, load_factor_mature=0.95,
                   ramp_profile='conservative')
    v3 = run_business_case(config, calibrated_results, t3)
    print(f"  Verdict: {v3.verdict} ({v3.confidence} confidence)")
    print(f"  {v3.headline}")
    assert v3.verdict == 'NO', f"Expected NO, got {v3.verdict}"
    print("   PASS  correct verdict")

    # Test 4: Write full output for achievable case
    print("\n" + "" * 60)
    print("TEST 4: Full output workbook")
    print("" * 60)
    ensure_output_dir()
    out_path = str(OUTPUT_DIR / 'BusinessCase_BA_LHR_SJC.xlsx')
    v4 = run_business_case(config, calibrated_results, t1, output_path=out_path)
    assert os.path.exists(out_path), "Output file not created"
    print(f"   Workbook written: {out_path}")

    # Summary
    print("\n" + "=" * 60)
    print("VALIDATION SUMMARY")
    print("=" * 60)
    print(f"  Test 1 (achievable):   {v1.verdict} ")
    print(f"  Test 2 (tight):        {v2.verdict} ")
    print(f"  Test 3 (impossible):   {v3.verdict} ")
    print(f"  Test 4 (output):       Written ")
    print(f"\n  Year-by-year ramp (Test 1):")
    for yr in v1.year_by_year:
        print(f"    Y{yr['year']}: {yr['forecast_pax']:>8,} pax, "
              f"{yr['forecast_lf']:.1%} LF vs {yr['target_lf']:.0%} target "
              f" {yr['verdict']}")
    print(f"\n  Risk flags (Test 1): {len(v1.risk_flags)}")
    for f in v1.risk_flags:
        print(f"     {f}")
    print(f"\n  Sensitivity parameters tested: {len(v1.sensitivity_results)}")
    for param, pts in v1.sensitivity_results.items():
        label = SENSITIVITY_PARAMS.get(param, {}).get('label', param)
        hitting = sum(1 for p in pts if p.hits_mature_target)
        print(f"    {label}: {hitting}/{len(pts)} scenarios hit mature target")
    print(f"\n  Airline pushback areas: {len(v1.airline_pushback_areas)}")
    for area in v1.airline_pushback_areas:
        print(f"     {area[:80]}...")

    print("\n  ALL TESTS PASSED ")
    return v1, v2, v3


if __name__ == '__main__':
    validate_ba_lhr_sjc()
