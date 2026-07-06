#!/usr/bin/env python3
"""
Avia Solutions  Market Research Executor (Module 30)
=====================================================
Executes the research plan from market_research_module.py and produces
structured, fully-cited research outputs for city pair presentations.

Two execution modes:
    1. CLAUDE MODE  designed to be invoked within a Claude conversation
       where Claude has web_search access. The module structures the
       research workflow, and Claude populates findings from live searches.
    
    2. MANUAL MODE  generates a research execution checklist with
       pre-formatted query strings that an analyst can run manually,
       plus empty structured templates for data entry.

Output formats:
    - Markdown research report (presentation-ready, fully cited)
    - Excel workbook with structured data per block
    - JSON findings file (for pipeline integration)

Maps to City Pair Presentation Template sections:
    Section 4: Why This Route? (economic context feeds into strategic rationale)
    Section 5A-5H: Bilateral Links / Demand Drivers (10 research blocks)
    Section 8: Airport Overview

Citation standard:
    Every factual claim includes source name, publication date, and URL.
    Preferred sources hierarchy enforced per block.
    Single-source claims flagged. Conflicting sources presented side-by-side.
"""

import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Dict, Optional, Any, Tuple
from enum import Enum


# ============================================================================
# DATA CLASSES  RESEARCH FINDINGS
# ============================================================================

@dataclass
class Citation:
    """A single source citation meeting Avia credibility standards."""
    source_name: str          # e.g., "US Census Bureau"
    title: str                # Article/page title
    date: str                 # Publication date (YYYY-MM-DD or "2024" or "accessed Feb 2026")
    url: str                  # Full URL
    source_type: str = ""     # "government", "industry", "news", "academic", "airline"
    confidence: str = "high"  # "high", "medium", "low"
    notes: str = ""           # Any caveats


@dataclass
class Finding:
    """A single research finding with citation."""
    claim: str                # The factual statement
    value: str = ""           # Numeric value if applicable (e.g., "4.2 million")
    unit: str = ""            # Unit if applicable (e.g., "visitors", "USD", "%")
    year: str = ""            # Reference year for the data
    citations: List[Citation] = field(default_factory=list)
    is_single_source: bool = False   # Flag if only one source found
    conflicting_info: str = ""       # Note if sources disagree
    relevance_to_case: str = ""      # How this supports the route case


@dataclass
class BlockFindings:
    """All findings for one research block."""
    block_id: str
    block_name: str
    section_ref: str
    relevance: str              # ESSENTIAL / INCLUDE / OPTIONAL
    findings: List[Finding] = field(default_factory=list)
    summary: str = ""           # 2-3 sentence summary for the block
    presentation_text: str = "" # Draft text ready for presentation slide
    data_gaps: List[str] = field(default_factory=list)  # What couldn't be found
    queries_executed: int = 0
    queries_with_results: int = 0


@dataclass
class ResearchOutput:
    """Complete research output for a city pair."""
    route: str                  # e.g., "LHR-SJC"
    airline: str
    origin_city: str
    destination_city: str
    origin_country: str
    destination_country: str
    demand_profile: str
    route_type: str
    buyer_type: str
    execution_date: str
    blocks: List[BlockFindings] = field(default_factory=list)
    executive_summary: str = ""
    key_statistics: Dict[str, str] = field(default_factory=dict)
    data_quality_score: float = 0.0  # 0-100
    total_citations: int = 0
    total_data_gaps: int = 0


# ============================================================================
# FINDING BUILDERS  helpers for creating structured findings
# ============================================================================

def make_finding(claim: str, value: str = "", unit: str = "", year: str = "",
                 source_name: str = "", source_title: str = "", source_date: str = "",
                 source_url: str = "", source_type: str = "", confidence: str = "high",
                 relevance: str = "") -> Finding:
    """Quick builder for a single-source finding."""
    citations = []
    if source_name:
        citations.append(Citation(
            source_name=source_name,
            title=source_title,
            date=source_date,
            url=source_url,
            source_type=source_type,
            confidence=confidence,
        ))
    return Finding(
        claim=claim,
        value=value,
        unit=unit,
        year=year,
        citations=citations,
        is_single_source=(len(citations) == 1),
        relevance_to_case=relevance,
    )


def add_citation(finding: Finding, source_name: str, title: str = "",
                 date: str = "", url: str = "", source_type: str = "",
                 confidence: str = "high") -> Finding:
    """Add an additional citation to an existing finding."""
    finding.citations.append(Citation(
        source_name=source_name, title=title, date=date,
        url=url, source_type=source_type, confidence=confidence,
    ))
    finding.is_single_source = (len(finding.citations) == 1)
    return finding


# ============================================================================
# RESEARCH EXECUTION FRAMEWORK
# ============================================================================

class ResearchExecutor:
    """
    Manages the research execution workflow.
    
    In Claude mode, this structures the conversation flow:
    1. Generate query plan from market_research_module
    2. For each ESSENTIAL block, execute priority-1 queries first
    3. Structure findings into BlockFindings
    4. Generate presentation-ready text
    5. Identify data gaps
    6. Produce final output document
    
    The executor doesn't run web searches itself  it provides the
    framework for Claude (or an analyst) to populate findings.
    """
    
    def __init__(self, origin: str, destination: str, origin_city: str,
                 destination_city: str, origin_country: str, destination_country: str,
                 airline: str, demand_profile: str = "mixed",
                 route_type: str = "hub_longhaul", buyer_type: str = "airport",
                 hub_airport: str = "", origin_region: str = "",
                 destination_region: str = ""):
        self.origin = origin
        self.destination = destination
        self.origin_city = origin_city
        self.destination_city = destination_city
        self.origin_country = origin_country
        self.destination_country = destination_country
        self.airline = airline
        self.demand_profile = demand_profile
        self.route_type = route_type
        self.buyer_type = buyer_type
        self.hub_airport = hub_airport
        self.origin_region = origin_region
        self.destination_region = destination_region
        
        self.output = ResearchOutput(
            route=f"{origin}-{destination}",
            airline=airline,
            origin_city=origin_city,
            destination_city=destination_city,
            origin_country=origin_country,
            destination_country=destination_country,
            demand_profile=demand_profile,
            route_type=route_type,
            buyer_type=buyer_type,
            execution_date=datetime.now().strftime("%Y-%m-%d"),
        )
        
        self.blocks: Dict[str, BlockFindings] = {}
    
    def init_block(self, block_id: str, block_name: str, section_ref: str,
                   relevance: str) -> BlockFindings:
        """Initialise a research block for findings collection."""
        bf = BlockFindings(
            block_id=block_id,
            block_name=block_name,
            section_ref=section_ref,
            relevance=relevance,
        )
        self.blocks[block_id] = bf
        return bf
    
    def add_finding(self, block_id: str, finding: Finding) -> None:
        """Add a finding to a specific block."""
        if block_id not in self.blocks:
            raise ValueError(f"Block {block_id} not initialised. Call init_block first.")
        self.blocks[block_id].findings.append(finding)
    
    def add_data_gap(self, block_id: str, gap_description: str) -> None:
        """Record that a piece of data couldn't be found."""
        if block_id in self.blocks:
            self.blocks[block_id].data_gaps.append(gap_description)
    
    def set_block_summary(self, block_id: str, summary: str,
                          presentation_text: str = "") -> None:
        """Set the summary and presentation text for a block."""
        if block_id in self.blocks:
            self.blocks[block_id].summary = summary
            if presentation_text:
                self.blocks[block_id].presentation_text = presentation_text
    
    def finalise(self, executive_summary: str = "",
                 key_statistics: Dict[str, str] = None) -> ResearchOutput:
        """Finalise the research output, computing quality metrics."""
        self.output.blocks = list(self.blocks.values())
        self.output.executive_summary = executive_summary
        if key_statistics:
            self.output.key_statistics = key_statistics
        
        # Compute quality metrics
        total_cites = 0
        total_gaps = 0
        blocks_with_findings = 0
        essential_blocks_complete = 0
        essential_blocks_total = 0
        
        for bf in self.output.blocks:
            total_cites += sum(len(f.citations) for f in bf.findings)
            total_gaps += len(bf.data_gaps)
            if bf.findings:
                blocks_with_findings += 1
            if bf.relevance == "ESSENTIAL":
                essential_blocks_total += 1
                if bf.findings and not bf.data_gaps:
                    essential_blocks_complete += 1
        
        self.output.total_citations = total_cites
        self.output.total_data_gaps = total_gaps
        
        # Quality score: essential completeness (60%) + citation density (20%) + gap penalty (20%)
        if essential_blocks_total > 0:
            essential_score = (essential_blocks_complete / essential_blocks_total) * 60
        else:
            essential_score = 60
        
        total_blocks = len(self.output.blocks)
        if total_blocks > 0:
            citation_score = min(20, (total_cites / (total_blocks * 2)) * 20)
        else:
            citation_score = 0
        
        gap_penalty = min(20, total_gaps * 3)
        self.output.data_quality_score = max(0, essential_score + citation_score - gap_penalty)
        
        return self.output
    
    # ====================================================================
    # CLAUDE EXECUTION PROMPTS  structured prompts for each block
    # ====================================================================
    
    def get_execution_prompt(self, block_id: str) -> str:
        """
        Generate a structured prompt that tells Claude exactly what to
        search for, what data to extract, and how to structure findings.
        
        This is used when Claude is executing research within a conversation.
        """
        prompts = {
            "economic_context": self._prompt_economic_context(),
            "corporate_links": self._prompt_corporate_links(),
            "tourism": self._prompt_tourism(),
            "trade": self._prompt_trade(),
            "education": self._prompt_education(),
            "diaspora": self._prompt_diaspora(),
            "passenger_profile": self._prompt_passenger_profile(),
            "non_cannibalization": self._prompt_non_cannibalization(),
            "case_study": self._prompt_case_study(),
            "airport_overview": self._prompt_airport_overview(),
        }
        return prompts.get(block_id, f"No execution prompt defined for block: {block_id}")
    
    def _prompt_economic_context(self) -> str:
        return f"""RESEARCH BLOCK: Economic Context (Section 4 background)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})
Airline: {self.airline}

SEARCH FOR:
1. {self.origin_city} metro GDP, median household income, and economic growth rate
   - Prefer: Census Bureau, BEA, or local economic development authority
   - Need: Current value + year-on-year change
2. {self.destination_city}/{self.destination_country} GDP and growth forecast
   - Prefer: IMF World Economic Outlook, World Bank, or national statistics office
   - Need: Current GDP + forecast growth rate for next 2-3 years
3. {self.origin_country}-{self.destination_country} bilateral air service agreement status
   - Prefer: ICAO, government aviation authority
   - Need: Whether open skies or restricted, any recent changes
4. {self.origin_region or self.origin_city} key economic indicators
   - If tech region: AI/semiconductor/fintech sector size and growth
   - If leisure: tourism contribution to GDP

EXTRACT AND STRUCTURE:
- Origin GDP per capita: $[value], [year], [source + URL]
- Origin household income: $[value], [year], [source + URL]  
- Destination GDP: $[value], growth [X]%, [year], [source + URL]
- Bilateral agreement: [status], [source + URL]
- Key sector data: [description], [source + URL]

FLAG: Any data older than 2023. Any single-source claims."""

    def _prompt_corporate_links(self) -> str:
        return f"""RESEARCH BLOCK: Corporate / Technology Links (Section 5A)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})
Region context: {self.origin_region or 'N/A'}  {self.destination_region or 'N/A'}

SEARCH FOR:
1. Major {self.destination_country} companies with offices/operations in {self.origin_city}/{self.origin_region or ''}
   - Need: Company names, what they do there, approximate employee count if available
   - Prefer: Company websites, Chamber of Commerce, government trade bodies
2. Major {self.origin_country}/{self.origin_region or self.origin_city} companies with offices in {self.destination_city}
   - Same format as above
3. Recent corporate investments or office openings (last 2 years) between the markets
   - Prefer: Press releases, financial press
4. Key sectors driving business travel between these markets
   - e.g., technology, financial services, pharmaceuticals, automotive

EXTRACT AND STRUCTURE:
- Companies {self.destination_country}{self.origin_city}: [list with details]
- Companies {self.origin_city}{self.destination_country}: [list with details]
- Total company count estimate: [number] companies with cross-border presence
- Key sectors: [list]
- Recent investments: [notable examples with dates and sources]
- Headline for presentation: "[N] {self.destination_country} companies have offices in {self.origin_region or self.origin_city}"

FLAG: Distinguish between companies with substantial local operations vs token presence.
PRESENTATION ANGLE: Business travel generators  these companies drive regular air travel between the markets."""

    def _prompt_tourism(self) -> str:
        return f"""RESEARCH BLOCK: Tourism / Visitor Data (Section 5B)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})

SEARCH FOR:
1. {self.origin_country} visitors to {self.destination_country}: annual arrivals, recent trend
   - Prefer: National tourism board, UNWTO, government statistics
   - Need: Annual figure + 3-5 year trend + purpose of visit breakdown
2. {self.destination_country} visitors to {self.origin_country}: same data
3. Visitor spending per capita for both directions
   - Prefer: Tourism board annual report, Mastercard destination insights
4. Tourism strategy/growth targets for {self.destination_city} or {self.destination_country}
   - Any government targets for visitor number growth?
5. Key attractions and tourism propositions relevant to air demand

EXTRACT AND STRUCTURE:
- {self.origin_country}{self.destination_country} visitors: [annual], [year], [growth %], [source + URL]
- {self.destination_country}{self.origin_country} visitors: [annual], [year], [growth %], [source + URL]
- Purpose breakdown: Business [X]%, Holiday [Y]%, VFR [Z]%, [source]
- Visitor spending: $[per capita], [year], [source]
- Tourism growth target: [description], [source]

FLAG: Pre-COVID vs post-COVID recovery  note whether figures have returned to 2019 levels.
PRESENTATION ANGLE: Leisure demand generators  visitor flows create baseline air travel demand."""

    def _prompt_trade(self) -> str:
        return f"""RESEARCH BLOCK: Trade and Investment (Section 5C)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})

SEARCH FOR:
1. {self.origin_country}-{self.destination_country} bilateral trade volume
   - Prefer: Dept of Commerce, ONS, national statistics, World Bank
   - Need: Total trade + exports + imports + trend
2. {self.origin_city} metro area exports to {self.destination_country}
   - Prefer: Bureau of Economic Analysis, metro trade statistics
   - This is a powerful data point for business route cases
3. Foreign direct investment flows between the countries
   - Prefer: UNCTAD, national investment body
   - Need: Both directions
4. Notable trade sectors and recent trade developments

EXTRACT AND STRUCTURE:
- Bilateral trade total: $[value], [year], [source + URL]
- {self.origin_country} exports to {self.destination_country}: $[value]
- {self.destination_country} exports to {self.origin_country}: $[value]
- Metro exports ({self.origin_city}{self.destination_country}): $[value], [year], [source]
- FDI {self.origin_country}{self.destination_country}: $[value]
- FDI {self.destination_country}{self.origin_country}: $[value]
- Key trade sectors: [list]

PRESENTATION ANGLE: Trade volumes underpin business travel  goods and services trade requires face-to-face meetings."""

    def _prompt_education(self) -> str:
        return f"""RESEARCH BLOCK: Education / Student Links (Section 5D)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})

SEARCH FOR:
1. {self.destination_country} students studying in {self.origin_country}
   - Prefer: IIE Open Doors, education ministry, HESA (UK)
   - Need: Annual enrollment + trend
2. {self.origin_country} students studying in {self.destination_country}
   - Same sources
3. University partnerships between {self.origin_city} and {self.destination_country} institutions
4. Notable universities in catchment areas of both airports

EXTRACT AND STRUCTURE:
- {self.destination_country} students in {self.origin_country}: [count], [year], [source + URL]
- {self.origin_country} students in {self.destination_country}: [count], [year], [source + URL]
- Key universities in {self.origin_city} catchment: [list]
- Key universities in {self.destination_city} catchment: [list]
- Notable partnerships: [description]

PRESENTATION ANGLE: Student flows generate term-time and vacation travel, plus VFR from visiting families."""

    def _prompt_diaspora(self) -> str:
        return f"""RESEARCH BLOCK: Diaspora / VFR Population (Section 5E)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})

SEARCH FOR:
1. {self.destination_country}-born population in {self.origin_city} metro area
   - Prefer: Census Bureau (US), ACS, national statistics
   - Need: Population count + metro area definition used
2. {self.origin_country}-born population in {self.destination_city} area
   - Same type of sources
3. Broader ancestry/heritage data if available
   - e.g., {self.destination_country} ancestry in {self.origin_city} state/region

EXTRACT AND STRUCTURE:
- {self.destination_country}-born in {self.origin_city} metro: [count], [year], [source + URL]
- {self.origin_country}-born in {self.destination_city} area: [count], [year], [source + URL]
- Broader heritage population: [count], [description], [source]

PRESENTATION ANGLE: Diaspora populations generate VFR (visiting friends and relatives) traffic  
a reliable, year-round demand source that is less price-sensitive and more recession-resistant than leisure."""

    def _prompt_passenger_profile(self) -> str:
        return f"""RESEARCH BLOCK: Passenger Profile (Section 5F)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})

SEARCH FOR:
1. Purpose of travel breakdown for this corridor
   - Prefer: CAA survey (UK routes), DOT, tourism board surveys
   - Need: Business % / Leisure % / VFR %
2. Passenger demographics if available
3. Booking patterns (advance purchase, direct vs agent)
4. Premium cabin demand indicators

EXTRACT AND STRUCTURE:
- Purpose split: Business [X]%, Leisure [Y]%, VFR [Z]%, [source + URL]
- Premium demand: [description of business/first class demand], [source]
- Booking patterns: [if available]

NOTE: Much of this data comes from Sabre MI (internal) rather than web research.
Focus web research on published survey data and tourism board reports.
PRESENTATION ANGLE: Passenger mix determines revenue quality  high business share supports premium fares."""

    def _prompt_non_cannibalization(self) -> str:
        return f"""RESEARCH BLOCK: Non-Cannibalization Evidence (Section 5G)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})
Airline: {self.airline}

SEARCH FOR:
1. Evidence that new air services between these markets stimulate demand
   rather than cannibalising existing services
   - Look for: New route launches at {self.origin} or {self.destination} and their market impact
   - Prefer: CAPA, Aviation Week, airline investor presentations
2. Examples where a competing carrier started service and total market grew
3. Any academic or industry research on route stimulation effects

EXTRACT AND STRUCTURE:
- Case studies: [route], [carrier], [year launched], [market impact], [source + URL]
- Market growth evidence: [before vs after new service data if available]
- Industry research: [any relevant studies on stimulation]

PRESENTATION ANGLE: New services grow the pie  evidence reduces airline concern about 
cannibalising their existing connecting traffic through other hubs."""

    def _prompt_case_study(self) -> str:
        return f"""RESEARCH BLOCK: Case Study / Comparable Route (Section 5H)
Route: {self.origin}-{self.destination} ({self.origin_city} to {self.destination_city})
Airline: {self.airline}

SEARCH FOR:
1. {self.airline}'s recent new long-haul route launches and their performance
   - Prefer: Airline press releases, investor presentations, CAPA
   - Need: Route, launch date, performance indicators (load factor, demand growth)
2. Other carriers' experience launching new routes to {self.destination_city}
   - Need: Which carriers, when, how they performed
3. Comparable route launches at {self.origin} (similar distance/market type)

EXTRACT AND STRUCTURE:
- {self.airline} comparable: [route], [launch year], [performance], [source + URL]
- Other carriers at {self.destination_city}: [carrier], [route], [performance], [source]
- Comparable at {self.origin}: [carrier], [route], [performance], [source]

PRESENTATION ANGLE: Successful precedents de-risk the opportunity for the target airline."""

    def _prompt_airport_overview(self) -> str:
        return f"""RESEARCH BLOCK: Airport Overview (Section 8)
Airport: {self.origin} ({self.origin_city})

SEARCH FOR:
1. {self.origin} airport annual passenger statistics (most recent year)
   - Prefer: Airport authority, ACI, FAA T-100
   - Need: Total pax, international pax, year-on-year growth
2. {self.origin} airport development and expansion plans
   - Terminal expansion, new facilities, runway development
   - Prefer: Airport master plan, planning authority documents
3. {self.origin} airport airline base and current international services
   - Which airlines operate, what international routes exist
4. {self.origin} airport incentive programme for new routes
   - Fee discounts, marketing support, revenue guarantees
5. {self.origin} airport catchment area demographics
   - Population within drive-time bands, household income

EXTRACT AND STRUCTURE:
- Annual passengers: [count], [year], [growth %], [source + URL]
- International passengers: [count], [year], [source]
- Current international routes: [count], key carriers: [list]
- Development plans: [summary], [source]
- Incentive programme: [summary if publicly available]
- Catchment: [population], [area definition], [source]

PRESENTATION ANGLE: Airport capability and commitment  demonstrates the airport can support 
the proposed service and is investing in growth."""


# ============================================================================
# OUTPUT GENERATORS
# ============================================================================

def generate_markdown_report(output: ResearchOutput) -> str:
    """
    Generate a presentation-ready markdown research report.
    
    This is the primary output  structured for direct use in
    populating city pair presentation sections.
    """
    lines = []
    
    # Header
    lines.append(f"# Market Research Report: {output.route}")
    lines.append(f"## {output.airline}  {output.origin_city} to {output.destination_city}")
    lines.append("")
    lines.append(f"**Prepared by:** Avia Solutions")
    lines.append(f"**Date:** {output.execution_date}")
    lines.append(f"**Route type:** {output.route_type}")
    lines.append(f"**Demand profile:** {output.demand_profile}")
    lines.append(f"**Buyer type:** {output.buyer_type}")
    lines.append(f"**Data quality score:** {output.data_quality_score:.0f}/100")
    lines.append(f"**Total citations:** {output.total_citations}")
    if output.total_data_gaps > 0:
        lines.append(f"**Data gaps:** {output.total_data_gaps} (see details below)")
    lines.append("")
    
    # Executive summary
    if output.executive_summary:
        lines.append("---")
        lines.append("## Executive Summary")
        lines.append("")
        lines.append(output.executive_summary)
        lines.append("")
    
    # Key statistics box
    if output.key_statistics:
        lines.append("---")
        lines.append("## Key Statistics at a Glance")
        lines.append("")
        lines.append("| Metric | Value | Source |")
        lines.append("|--------|-------|--------|")
        for metric, value in output.key_statistics.items():
            lines.append(f"| {metric} | {value} |  |")
        lines.append("")
    
    # Research blocks
    for bf in output.blocks:
        lines.append("---")
        relevance_marker = {"ESSENTIAL": "", "INCLUDE": "", "OPTIONAL": ""}.get(
            bf.relevance, "")
        lines.append(f"## {relevance_marker} {bf.block_name} ({bf.section_ref})  {bf.relevance}")
        lines.append("")
        
        if bf.summary:
            lines.append(f"**Summary:** {bf.summary}")
            lines.append("")
        
        if bf.presentation_text:
            lines.append("### Presentation Draft")
            lines.append("")
            lines.append(bf.presentation_text)
            lines.append("")
        
        if bf.findings:
            lines.append("### Findings")
            lines.append("")
            for i, f in enumerate(bf.findings, 1):
                # Main claim
                value_str = f""
                if f.value:
                    value_str = f" **{f.value}** {f.unit}"
                lines.append(f"**{i}.** {f.claim}{value_str}")
                if f.year:
                    lines.append(f"   - Data year: {f.year}")
                
                # Citations
                for c in f.citations:
                    cite_parts = [f"*{c.source_name}*"]
                    if c.title:
                        cite_parts.append(f'"{c.title}"')
                    if c.date:
                        cite_parts.append(f"({c.date})")
                    if c.url:
                        cite_parts.append(f"[Link]({c.url})")
                    if c.confidence != "high":
                        cite_parts.append(f"[confidence: {c.confidence}]")
                    lines.append(f"   - Source: {', '.join(cite_parts)}")
                
                # Flags
                if f.is_single_source:
                    lines.append(f"   -  **Single source**  verify independently")
                if f.conflicting_info:
                    lines.append(f"   -  **Conflicting info:** {f.conflicting_info}")
                if f.relevance_to_case:
                    lines.append(f"   -  *Route relevance:* {f.relevance_to_case}")
                lines.append("")
        
        if bf.data_gaps:
            lines.append("### Data Gaps")
            lines.append("")
            for gap in bf.data_gaps:
                lines.append(f"-  **DATA REQUIRED:** {gap}")
            lines.append("")
        
        lines.append(f"*Queries executed: {bf.queries_executed} | "
                     f"With results: {bf.queries_with_results}*")
        lines.append("")
    
    # Source quality summary
    lines.append("---")
    lines.append("## Source Quality Summary")
    lines.append("")
    
    all_citations = []
    for bf in output.blocks:
        for f in bf.findings:
            all_citations.extend(f.citations)
    
    if all_citations:
        # Count by source type
        type_counts = {}
        for c in all_citations:
            st = c.source_type or "unclassified"
            type_counts[st] = type_counts.get(st, 0) + 1
        
        lines.append("| Source Type | Count |")
        lines.append("|------------|-------|")
        for st, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            lines.append(f"| {st} | {count} |")
        lines.append("")
        
        # Confidence breakdown
        conf_counts = {}
        for c in all_citations:
            conf_counts[c.confidence] = conf_counts.get(c.confidence, 0) + 1
        
        lines.append("| Confidence | Count |")
        lines.append("|-----------|-------|")
        for conf, count in sorted(conf_counts.items()):
            lines.append(f"| {conf} | {count} |")
    
    lines.append("")
    lines.append("---")
    lines.append(f"*Report generated by Avia Solutions Market Research Executor "
                 f"on {output.execution_date}*")
    
    return "\n".join(lines)


def generate_json_output(output: ResearchOutput) -> str:
    """Generate JSON output for pipeline integration."""
    
    def citation_to_dict(c: Citation) -> dict:
        return {
            "source_name": c.source_name,
            "title": c.title,
            "date": c.date,
            "url": c.url,
            "source_type": c.source_type,
            "confidence": c.confidence,
            "notes": c.notes,
        }
    
    def finding_to_dict(f: Finding) -> dict:
        return {
            "claim": f.claim,
            "value": f.value,
            "unit": f.unit,
            "year": f.year,
            "citations": [citation_to_dict(c) for c in f.citations],
            "is_single_source": f.is_single_source,
            "conflicting_info": f.conflicting_info,
            "relevance_to_case": f.relevance_to_case,
        }
    
    data = {
        "route": output.route,
        "airline": output.airline,
        "origin_city": output.origin_city,
        "destination_city": output.destination_city,
        "demand_profile": output.demand_profile,
        "route_type": output.route_type,
        "buyer_type": output.buyer_type,
        "execution_date": output.execution_date,
        "executive_summary": output.executive_summary,
        "key_statistics": output.key_statistics,
        "data_quality_score": output.data_quality_score,
        "total_citations": output.total_citations,
        "total_data_gaps": output.total_data_gaps,
        "blocks": [
            {
                "block_id": bf.block_id,
                "block_name": bf.block_name,
                "section_ref": bf.section_ref,
                "relevance": bf.relevance,
                "summary": bf.summary,
                "presentation_text": bf.presentation_text,
                "findings": [finding_to_dict(f) for f in bf.findings],
                "data_gaps": bf.data_gaps,
                "queries_executed": bf.queries_executed,
                "queries_with_results": bf.queries_with_results,
            }
            for bf in output.blocks
        ],
    }
    
    return json.dumps(data, indent=2)


def generate_excel_output(output: ResearchOutput, filepath: str) -> str:
    """Generate Excel workbook with structured research data."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl not available  install with: pip install openpyxl"
    
    wb = openpyxl.Workbook()
    
    # ---- Cover sheet ----
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    header_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    label_font = Font(name='Arial', size=11, bold=True)
    value_font = Font(name='Arial', size=11)
    
    ws['A1'] = "AVIA SOLUTIONS"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E79')
    ws['A2'] = "Market Research Report"
    ws['A2'].font = Font(name='Arial', size=14, color='2E75B6')
    
    cover_data = [
        ("", ""),
        ("Route", output.route),
        ("Airline", output.airline),
        ("Origin", f"{output.origin_city} ({output.route.split('-')[0]})"),
        ("Destination", f"{output.destination_city} ({output.route.split('-')[1]})"),
        ("Demand Profile", output.demand_profile),
        ("Route Type", output.route_type),
        ("Buyer Type", output.buyer_type),
        ("Execution Date", output.execution_date),
        ("", ""),
        ("Quality Score", f"{output.data_quality_score:.0f}/100"),
        ("Total Citations", str(output.total_citations)),
        ("Data Gaps", str(output.total_data_gaps)),
    ]
    
    for i, (label, value) in enumerate(cover_data, 4):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = label_font
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = value_font
    
    # ---- Summary sheet ----
    ws_sum = wb.create_sheet("Key Statistics")
    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 30
    ws_sum.column_dimensions['C'].width = 40
    
    ws_sum['A1'] = "Key Statistics"
    ws_sum['A1'].font = header_font
    
    ws_sum['A3'] = "Metric"
    ws_sum['B3'] = "Value"
    ws_sum['C3'] = "Source"
    for cell in [ws_sum['A3'], ws_sum['B3'], ws_sum['C3']]:
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
    
    row = 4
    for metric, value in output.key_statistics.items():
        ws_sum[f'A{row}'] = metric
        ws_sum[f'B{row}'] = value
        row += 1
    
    # ---- Block sheets ----
    for bf in output.blocks:
        safe_name = bf.block_name[:28].replace("/", "-")
        ws_block = wb.create_sheet(safe_name)
        ws_block.column_dimensions['A'].width = 40
        ws_block.column_dimensions['B'].width = 20
        ws_block.column_dimensions['C'].width = 15
        ws_block.column_dimensions['D'].width = 40
        ws_block.column_dimensions['E'].width = 50
        
        ws_block['A1'] = bf.block_name
        ws_block['A1'].font = header_font
        ws_block['A2'] = f"Section: {bf.section_ref} | Relevance: {bf.relevance}"
        ws_block['A2'].font = Font(name='Arial', size=10, italic=True)
        
        if bf.summary:
            ws_block['A3'] = "Summary:"
            ws_block['A3'].font = label_font
            ws_block['B3'] = bf.summary
        
        # Findings table
        headers = ["Finding", "Value", "Year", "Source", "URL"]
        row = 5
        for col, h in enumerate(headers, 1):
            cell = ws_block.cell(row=row, column=col, value=h)
            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2E75B6')
        
        row = 6
        for f in bf.findings:
            ws_block.cell(row=row, column=1, value=f.claim)
            ws_block.cell(row=row, column=2, value=f"{f.value} {f.unit}".strip())
            ws_block.cell(row=row, column=3, value=f.year)
            if f.citations:
                ws_block.cell(row=row, column=4, value=f.citations[0].source_name)
                ws_block.cell(row=row, column=5, value=f.citations[0].url)
            if f.is_single_source:
                ws_block.cell(row=row, column=1).font = Font(name='Arial', size=10,
                                                              color='CC6600')
            row += 1
        
        # Data gaps
        if bf.data_gaps:
            row += 1
            ws_block.cell(row=row, column=1, value="DATA GAPS")
            ws_block.cell(row=row, column=1).font = Font(name='Arial', size=10,
                                                          bold=True, color='CC0000')
            row += 1
            for gap in bf.data_gaps:
                ws_block.cell(row=row, column=1, value=f" {gap}")
                ws_block.cell(row=row, column=1).font = Font(name='Arial', size=10,
                                                              color='CC0000')
                row += 1
    
    # ---- Citations sheet ----
    ws_cite = wb.create_sheet("All Citations")
    ws_cite.column_dimensions['A'].width = 15
    ws_cite.column_dimensions['B'].width = 25
    ws_cite.column_dimensions['C'].width = 40
    ws_cite.column_dimensions['D'].width = 15
    ws_cite.column_dimensions['E'].width = 50
    ws_cite.column_dimensions['F'].width = 15
    ws_cite.column_dimensions['G'].width = 12
    
    cite_headers = ["Block", "Source", "Title", "Date", "URL", "Type", "Confidence"]
    for col, h in enumerate(cite_headers, 1):
        cell = ws_cite.cell(row=1, column=col, value=h)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
    
    cite_row = 2
    for bf in output.blocks:
        for f in bf.findings:
            for c in f.citations:
                ws_cite.cell(row=cite_row, column=1, value=bf.block_name)
                ws_cite.cell(row=cite_row, column=2, value=c.source_name)
                ws_cite.cell(row=cite_row, column=3, value=c.title)
                ws_cite.cell(row=cite_row, column=4, value=c.date)
                ws_cite.cell(row=cite_row, column=5, value=c.url)
                ws_cite.cell(row=cite_row, column=6, value=c.source_type)
                ws_cite.cell(row=cite_row, column=7, value=c.confidence)
                cite_row += 1
    
    wb.save(filepath)
    return filepath


# ============================================================================
# EXECUTION ORCHESTRATOR  for Claude conversation flow
# ============================================================================

def get_execution_sequence(demand_profile: str, route_type: str,
                           buyer_type: str) -> List[Tuple[str, str]]:
    """
    Returns the optimal execution sequence for research blocks.
    
    ESSENTIAL blocks run first, then INCLUDE, then OPTIONAL.
    Within each tier, blocks are ordered by dependency:
    1. Economic context (background for everything)
    2. Corporate links / Tourism (core demand evidence)
    3. Trade / Education / Diaspora (supporting evidence)
    4. Passenger profile (synthesis)
    5. Non-cannibalization / Case study (de-risking)
    6. Airport overview (closing section)
    
    Returns list of (block_id, execution_priority) tuples.
    """
    # Build a simple relevance map based on profile
    profile_lower = demand_profile.lower()
    
    # Base sequence (always this order within a tier)
    base_sequence = [
        "economic_context",
        "corporate_links",
        "tourism",
        "trade",
        "education",
        "diaspora",
        "passenger_profile",
        "non_cannibalization",
        "case_study",
        "airport_overview",
    ]
    
    # Determine relevance per block
    relevance = {}
    if profile_lower in ("business", "mixed"):
        relevance["corporate_links"] = "ESSENTIAL"
        relevance["trade"] = "ESSENTIAL" if profile_lower == "business" else "INCLUDE"
        relevance["tourism"] = "INCLUDE" if profile_lower == "business" else "ESSENTIAL"
    elif profile_lower == "leisure":
        relevance["tourism"] = "ESSENTIAL"
        relevance["corporate_links"] = "OPTIONAL"
        relevance["trade"] = "OPTIONAL"
    elif profile_lower == "vfr_diaspora":
        relevance["diaspora"] = "ESSENTIAL"
        relevance["tourism"] = "INCLUDE"
    
    relevance.setdefault("economic_context", "ESSENTIAL" if route_type.lower() in ("hub_longhaul",) else "INCLUDE")
    relevance.setdefault("education", "INCLUDE" if profile_lower == "business" else "OPTIONAL")
    relevance.setdefault("diaspora", relevance.get("diaspora", "INCLUDE" if profile_lower == "mixed" else "OPTIONAL"))
    relevance.setdefault("passenger_profile", "ESSENTIAL" if buyer_type.lower() == "fund" else "INCLUDE")
    relevance.setdefault("airport_overview", "ESSENTIAL" if buyer_type.lower() == "airport" else "INCLUDE")
    relevance.setdefault("non_cannibalization", "INCLUDE" if buyer_type.lower() == "airport" else "OPTIONAL")
    relevance.setdefault("case_study", "INCLUDE")
    
    # Fill defaults
    for block in base_sequence:
        relevance.setdefault(block, "OPTIONAL")
    
    # Sort: ESSENTIAL first, then INCLUDE, then OPTIONAL, preserving base order within tier
    tier_order = {"ESSENTIAL": 0, "INCLUDE": 1, "OPTIONAL": 2}
    sorted_blocks = sorted(base_sequence,
                           key=lambda b: (tier_order.get(relevance[b], 2),
                                         base_sequence.index(b)))
    
    return [(block, relevance[block]) for block in sorted_blocks]


def generate_execution_plan(executor: ResearchExecutor) -> str:
    """
    Generate a complete execution plan as a structured prompt.
    
    This is what Claude uses to systematically execute the research,
    block by block, with full citation tracking.
    """
    sequence = get_execution_sequence(
        executor.demand_profile, executor.route_type, executor.buyer_type
    )
    
    lines = []
    lines.append(f"{'=' * 70}")
    lines.append(f"MARKET RESEARCH EXECUTION PLAN")
    lines.append(f"Route: {executor.origin}-{executor.destination}")
    lines.append(f"Airline: {executor.airline}")
    lines.append(f"Profile: {executor.demand_profile} | Type: {executor.route_type} | Buyer: {executor.buyer_type}")
    lines.append(f"{'=' * 70}")
    lines.append("")
    
    essential_count = sum(1 for _, r in sequence if r == "ESSENTIAL")
    include_count = sum(1 for _, r in sequence if r == "INCLUDE")
    optional_count = sum(1 for _, r in sequence if r == "OPTIONAL")
    
    lines.append(f"EXECUTION SUMMARY: {len(sequence)} blocks")
    lines.append(f"  ESSENTIAL: {essential_count} (must complete)")
    lines.append(f"  INCLUDE: {include_count} (should complete)")
    lines.append(f"  OPTIONAL: {optional_count} (if time permits)")
    lines.append("")
    
    for i, (block_id, relevance) in enumerate(sequence, 1):
        marker = {"ESSENTIAL": "", "INCLUDE": "", "OPTIONAL": ""}[relevance]
        prompt = executor.get_execution_prompt(block_id)
        
        lines.append(f"{'' * 70}")
        lines.append(f"STEP {i}/{len(sequence)}: {marker} [{relevance}] {block_id}")
        lines.append(f"{'' * 70}")
        lines.append(prompt)
        lines.append("")
    
    return "\n".join(lines)


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def quick_research(origin: str, destination: str, origin_city: str,
                   destination_city: str, origin_country: str,
                   destination_country: str, airline: str,
                   demand_profile: str = "mixed",
                   route_type: str = "hub_longhaul",
                   buyer_type: str = "airport",
                   hub_airport: str = "",
                   origin_region: str = "",
                   destination_region: str = "") -> Tuple[ResearchExecutor, str]:
    """
    Quick setup: create executor and generate execution plan.
    
    Returns (executor, execution_plan_text).
    The executor is ready for findings to be added as research proceeds.
    """
    executor = ResearchExecutor(
        origin=origin,
        destination=destination,
        origin_city=origin_city,
        destination_city=destination_city,
        origin_country=origin_country,
        destination_country=destination_country,
        airline=airline,
        demand_profile=demand_profile,
        route_type=route_type,
        buyer_type=buyer_type,
        hub_airport=hub_airport,
        origin_region=origin_region,
        destination_region=destination_region,
    )
    
    plan = generate_execution_plan(executor)
    return executor, plan


# ============================================================================
# TESTS
# ============================================================================

def test_executor_ba_lhr_sjc():
    """Test executor setup and plan generation for BA LHR-SJC."""
    
    executor, plan = quick_research(
        origin="SJC", destination="LHR",
        origin_city="San Jose", destination_city="London",
        origin_country="United States", destination_country="United Kingdom",
        airline="British Airways",
        demand_profile="mixed", route_type="hub_longhaul", buyer_type="airport",
        hub_airport="LHR", origin_region="Silicon Valley",
    )
    
    print(plan[:500])
    print(f"\n... ({len(plan)} chars total)")
    
    # Test execution sequence
    sequence = get_execution_sequence("mixed", "hub_longhaul", "airport")
    essential = [b for b, r in sequence if r == "ESSENTIAL"]
    
    assert "economic_context" in essential, "Economic context should be ESSENTIAL for hub longhaul"
    assert "corporate_links" in essential, "Corporate should be ESSENTIAL for mixed demand"
    assert "tourism" in essential, "Tourism should be ESSENTIAL for mixed demand"
    assert "airport_overview" in essential, "Airport should be ESSENTIAL for airport buyer"
    
    # Test finding creation
    executor.init_block("economic_context", "Economic Context", "S4", "ESSENTIAL")
    executor.add_finding("economic_context", make_finding(
        claim="San Jose metro area median household income is the highest in the US",
        value="$143,000", unit="USD", year="2023",
        source_name="US Census Bureau", source_title="ACS 2023",
        source_date="2024", source_url="https://census.gov/acs",
        source_type="government", confidence="high",
        relevance="Premium fare demand  highest income catchment supports business class revenue"
    ))
    
    executor.add_finding("economic_context", make_finding(
        claim="UK GDP growth forecast at 1.1% for 2025",
        value="1.1%", unit="growth", year="2025",
        source_name="IMF", source_title="World Economic Outlook",
        source_date="Oct 2024", source_url="https://imf.org/weo",
        source_type="international", confidence="high",
    ))
    
    executor.add_data_gap("economic_context", 
                          "US-UK bilateral air service agreement  open skies confirmed but specific terms not found")
    
    executor.set_block_summary("economic_context",
        summary="Silicon Valley remains the highest-income metro in the US with strong economic fundamentals. "
                "UK growth modest but stable, supporting baseline demand.",
        presentation_text="San Jose boasts the highest median household income in the United States at $143,000 "
                         "(US Census Bureau, 2023), underpinning strong premium cabin demand. The UK economy "
                         "is forecast to grow at 1.1% in 2025 (IMF WEO), providing stable baseline demand.")
    
    # Finalise
    output = executor.finalise(
        executive_summary="BA LHR-SJC route assessment supported by strong bilateral economic ties.",
        key_statistics={"Median household income (SJC)": "$143,000",
                        "UK GDP growth forecast": "1.1%"}
    )
    
    assert output.total_citations == 2
    assert output.total_data_gaps == 1
    assert output.data_quality_score > 0
    
    # Test markdown output
    md = generate_markdown_report(output)
    assert "Market Research Report" in md
    assert "$143,000" in md
    assert "DATA REQUIRED" in md
    
    # Test JSON output
    js = generate_json_output(output)
    parsed = json.loads(js)
    assert parsed["route"] == "SJC-LHR"
    assert len(parsed["blocks"]) == 1
    
    print(f"\n BA LHR-SJC executor test passed")
    print(f"  Citations: {output.total_citations}")
    print(f"  Data gaps: {output.total_data_gaps}")
    print(f"  Quality score: {output.data_quality_score:.0f}/100")
    print(f"  Markdown length: {len(md)} chars")
    
    return output


def test_execution_sequence():
    """Test that execution sequences vary correctly by route profile."""
    
    # Business route: corporate and trade should be ESSENTIAL
    seq_biz = get_execution_sequence("business", "hub_longhaul", "airline")
    biz_essential = {b for b, r in seq_biz if r == "ESSENTIAL"}
    assert "corporate_links" in biz_essential
    assert "trade" in biz_essential
    
    # Leisure route: tourism should be ESSENTIAL, corporate OPTIONAL
    seq_lei = get_execution_sequence("leisure", "hub_longhaul", "airport")
    lei_dict = {b: r for b, r in seq_lei}
    assert lei_dict["tourism"] == "ESSENTIAL"
    assert lei_dict["corporate_links"] == "OPTIONAL"
    
    # VFR route: diaspora should be ESSENTIAL
    seq_vfr = get_execution_sequence("vfr_diaspora", "hub_longhaul", "airport")
    vfr_essential = {b for b, r in seq_vfr if r == "ESSENTIAL"}
    assert "diaspora" in vfr_essential
    
    # Fund buyer: passenger profile should be ESSENTIAL
    seq_fund = get_execution_sequence("mixed", "hub_longhaul", "fund")
    fund_dict = {b: r for b, r in seq_fund}
    assert fund_dict["passenger_profile"] == "ESSENTIAL"
    
    print(" Execution sequence tests passed")
    print(f"  Business ESSENTIAL: {biz_essential}")
    print(f"  Leisure tourism: {lei_dict['tourism']}")
    print(f"  VFR ESSENTIAL: {vfr_essential}")
    print(f"  Fund pax profile: {fund_dict['passenger_profile']}")


if __name__ == "__main__":
    print("=" * 70)
    print("MARKET RESEARCH EXECUTOR  TEST SUITE")
    print("=" * 70)
    
    test_execution_sequence()
    print()
    test_executor_ba_lhr_sjc()
    
    print("\n" + "=" * 70)
    print("ALL TESTS PASSED")
    print("=" * 70)
