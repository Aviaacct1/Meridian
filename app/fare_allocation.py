#!/usr/bin/env python3
"""
Avia Solutions - Fare Allocation Module
=========================================
Replaces single-average-fare approach in revenue_forecast.py with
per-market fare data extracted from Sabre booking files.

Current problem:
  revenue_forecast.py uses ONE set of fares per segment.
  This loses huge fare variation between connecting markets.
  E.g., BA LHR-SJC connecting: AAL coach $291 vs DXB coach $1018.

This module:
  1. Parses Sabre fare extract files (OUTPUTS tab structure)
  2. Builds per-market fare & cabin-split lookups
  3. Computes weighted-average fares using pipeline forecast pax as weights
  4. Produces market-level revenue forecasts (not just segment totals)
  5. Integrates with existing RevenueEngine via enhanced RevenueConfig

Integration:
  from fare_allocation import parse_fare_files, compute_market_revenue
  allocator = parse_fare_files(home_file='...', dest_file='...')
  market_revenue = compute_market_revenue(allocator, pipeline_results)

Dependencies: openpyxl, revenue_forecast.py, closed_loop_pipeline_v2.py
"""

from config import REFERENCE_CASE_DIR, OUTPUT_DIR, ensure_output_dir
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Any, Tuple
from collections import defaultdict
import math

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================================
# CONSTANTS
# ============================================================================

SABRE_CABIN_MAP = {
    'First': 'F', 'Business': 'J', 'Discount Business': 'J',
    'Coach': 'PY', 'Discount Coach': 'Y', 'Premium Economy': 'PY',
}

SABRE_CABIN_MAP_3CLASS = {
    'First': 'F', 'Business': 'J', 'Discount Business': 'J',
    'Coach': 'Y', 'Discount Coach': 'Y',
}

MIN_PAX_THRESHOLD = 10

FARE_BOUNDS = {
    'Y': (80, 3000), 'PY': (150, 5000), 'J': (500, 15000), 'F': (1000, 25000),
}

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class MarketFareSet:
    """Fare and cabin mix data for a single connecting market."""
    city_code: str
    city_name: str = ''
    direction: str = ''
    revenue_by_cabin: Dict[str, float] = field(default_factory=dict)
    pax_by_cabin: Dict[str, float] = field(default_factory=dict)
    fares: Dict[str, float] = field(default_factory=dict)
    cabin_split: Dict[str, float] = field(default_factory=dict)
    total_pax: float = 0
    total_revenue: float = 0
    avg_ow_fare: float = 0
    low_sample: bool = False
    missing_cabins: List[str] = field(default_factory=list)
    fare_outliers: List[str] = field(default_factory=list)

    def compute_derived(self, cabin_map=None):
        """Compute Avia-standard fares and splits from raw Sabre data."""
        if cabin_map is None:
            cabin_map = SABRE_CABIN_MAP
        avia_rev = defaultdict(float)
        avia_pax = defaultdict(float)
        for sabre_cabin, rev in self.revenue_by_cabin.items():
            avia_cabin = cabin_map.get(sabre_cabin, 'Y')
            avia_rev[avia_cabin] += rev
            avia_pax[avia_cabin] += self.pax_by_cabin.get(sabre_cabin, 0)
        self.total_pax = sum(avia_pax.values())
        self.total_revenue = sum(avia_rev.values())
        if self.total_pax < MIN_PAX_THRESHOLD:
            self.low_sample = True
        for cabin in ['Y', 'PY', 'J', 'F']:
            pax = avia_pax.get(cabin, 0)
            rev = avia_rev.get(cabin, 0)
            if pax > 0:
                ow_fare = rev / (pax * 2)
                self.fares[cabin] = ow_fare
                self.cabin_split[cabin] = pax / self.total_pax if self.total_pax > 0 else 0
                lo, hi = FARE_BOUNDS.get(cabin, (0, 99999))
                if ow_fare < lo or ow_fare > hi:
                    self.fare_outliers.append(f"{cabin}: ${ow_fare:,.0f}")
            else:
                self.missing_cabins.append(cabin)
        self.avg_ow_fare = self.total_revenue / (self.total_pax * 2) if self.total_pax > 0 else 0
        split_total = sum(self.cabin_split.values())
        if split_total > 0 and abs(split_total - 1.0) > 0.01:
            for k in self.cabin_split:
                self.cabin_split[k] /= split_total


# ============================================================================
# FARE FILE PARSER
# ============================================================================

class FareParser:
    """Parses Sabre fare extract files in the Avia template format."""

    def __init__(self, filepath, direction='home', cabin_map=None):
        self.filepath = filepath
        self.direction = direction
        self.cabin_map = cabin_map or SABRE_CABIN_MAP
        self.audit = []
        self.direct_markets = {}
        self.connecting_markets = {}

    def _log(self, msg):
        self.audit.append(msg)

    def parse(self):
        """Parse the fare file. Returns (direct_markets, connecting_markets)."""
        if not HAS_OPENPYXL:
            return {}, {}
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        outputs_sheet = None
        for name in wb.sheetnames:
            if 'output' in name.lower():
                outputs_sheet = wb[name]
                break
        if outputs_sheet is None:
            self._log(f"ERROR: No outputs sheet in {self.filepath}")
            return {}, {}
        ws = outputs_sheet
        self._log(f"Parsing {self.filepath} -> {ws.title} ({ws.max_row}r x {ws.max_column}c)")
        all_rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
        direct_start = indirect_start = None
        for i, row in enumerate(all_rows):
            row_str = ' '.join(str(v) for v in row if v is not None).upper()
            if 'DIRECT' in row_str and 'INDIRECT' not in row_str and direct_start is None:
                direct_start = i
            elif 'INDIRECT' in row_str and indirect_start is None:
                indirect_start = i
        if direct_start is not None:
            end = indirect_start if indirect_start else len(all_rows)
            self.direct_markets = self._parse_section(all_rows, direct_start, end, 'direct')
            self._log(f"  Direct markets: {len(self.direct_markets)}")
        if indirect_start is not None:
            self.connecting_markets = self._parse_section(all_rows, indirect_start, len(all_rows), 'connecting')
            self._log(f"  Connecting markets: {len(self.connecting_markets)}")
        wb.close()
        return self.direct_markets, self.connecting_markets

    def _parse_section(self, all_rows, start, end, section_type):
        """Parse a DIRECT or INDIRECT section."""
        markets = {}
        header_row = header_idx = None
        for i in range(start, min(start + 8, end)):
            row_vals = [str(v).strip() if v else '' for v in all_rows[i]]
            if any('business' in v.lower() for v in row_vals):
                header_row = row_vals
                header_idx = i
                break
        if header_row is None:
            self._log(f"  WARNING: No header for {section_type}")
            return markets
        col_map = self._detect_columns(header_row)
        if not col_map:
            self._log(f"  WARNING: Could not detect columns for {section_type}")
            return markets
        for i in range(header_idx + 1, end):
            row = all_rows[i]
            if not row:
                continue
            city_code = None
            for ci in [1, 0]:
                if ci < len(row) and row[ci] is not None:
                    val = str(row[ci]).strip()
                    if len(val) == 3 and val.isalpha():
                        city_code = val.upper()
                        break
            if not city_code or city_code in ('MOD', 'CIT', 'NON', 'GRA', 'TOT', 'SUM'):
                continue
            mfs = MarketFareSet(city_code=city_code, direction=self.direction)
            self._extract_fare_data(row, col_map, mfs)
            if mfs.total_pax > 0:
                mfs.compute_derived(self.cabin_map)
                markets[city_code] = mfs
        return markets

    def _detect_columns(self, header_row):
        """Detect column layout from header row.
        
        The layout has 2-3 repeating blocks of cabin columns:
          Revenue block | Pax block | [Fare block]
        Each block has cabin-named columns. We detect block boundaries
        by looking for where the same cabin name repeats.
        
        Handles variable cabin counts (e.g., DIRECT has 4 cabins,
        INDIRECT has 5 with Discount Business).
        """
        cabin_names = ['Business', 'Coach', 'Discount Business', 'Discount Coach', 'First']
        positions = []
        for ci, val in enumerate(header_row):
            val_clean = val.strip() if val else ''
            for cn in cabin_names:
                if cn.lower() == val_clean.lower():
                    positions.append((ci, cn))
                    break
        if len(positions) < 4:
            return None
        
        # Strategy: find the first cabin that appears more than once.
        # Its second occurrence marks the start of the pax block.
        seen = {}
        pax_start = None
        for ci, cn in positions:
            if cn in seen:
                pax_start = ci
                break
            seen[cn] = ci
        
        if pax_start is None:
            return None
        
        # Check for a third occurrence (fare block)
        fare_start = None
        seen2 = {}
        for ci, cn in positions:
            if ci >= pax_start:
                if cn in seen2:
                    fare_start = ci
                    break
                seen2[cn] = ci
        
        col_map = {'rev_cols': {}, 'pax_cols': {}, 'fare_cols': {}}
        for ci, cn in positions:
            if fare_start and ci >= fare_start:
                col_map['fare_cols'][cn] = ci
            elif ci >= pax_start:
                col_map['pax_cols'][cn] = ci
            else:
                col_map['rev_cols'][cn] = ci
        return col_map if col_map['rev_cols'] and col_map['pax_cols'] else None

    def _extract_fare_data(self, row, col_map, mfs):
        """Extract revenue and pax data from a row."""
        total_pax = 0.0
        for cabin_name, ci in col_map['rev_cols'].items():
            if ci < len(row) and row[ci] is not None:
                try:
                    mfs.revenue_by_cabin[cabin_name] = float(row[ci])
                except (ValueError, TypeError):
                    pass
        for cabin_name, ci in col_map['pax_cols'].items():
            if ci < len(row) and row[ci] is not None:
                try:
                    pax = float(row[ci])
                    mfs.pax_by_cabin[cabin_name] = pax
                    total_pax += pax
                except (ValueError, TypeError):
                    pass
        mfs.total_pax = total_pax
        mfs.total_revenue = sum(mfs.revenue_by_cabin.values())


# ============================================================================
# FARE ALLOCATOR
# ============================================================================

class FareAllocator:
    """Manages fare data across files and provides per-market fare lookups."""

    def __init__(self, cabin_map=None, fare_weight=0.85):
        self.cabin_map = cabin_map or SABRE_CABIN_MAP
        self.fare_weight = fare_weight
        self.p2p_fares = {}
        self.cnx_home_fares = {}
        self.cnx_dest_fares = {}
        self.fallback_cnx_home = {}
        self.fallback_cnx_dest = {}
        self.fallback_p2p = {}
        self.fallback_cabin_split = {}
        self.audit = []
        self.parsers = []

    def _log(self, msg):
        self.audit.append(msg)

    def add_home_file(self, filepath):
        """Add home-direction fare file (e.g., LHR->SJC->XXX)."""
        parser = FareParser(filepath, direction='home', cabin_map=self.cabin_map)
        direct, connecting = parser.parse()
        self.parsers.append(parser)
        self.audit.extend(parser.audit)
        for code, mfs in direct.items():
            self.p2p_fares[code] = mfs
        for code, mfs in connecting.items():
            mfs.direction = 'dest'
            self.cnx_dest_fares[code] = mfs
        self._log(f"Home file: {len(direct)} P2P, {len(connecting)} cnx dest markets")

    def add_dest_file(self, filepath):
        """Add dest-direction fare file (e.g., SJC->LHR->XXX)."""
        parser = FareParser(filepath, direction='dest', cabin_map=self.cabin_map)
        direct, connecting = parser.parse()
        self.parsers.append(parser)
        self.audit.extend(parser.audit)
        for code, mfs in direct.items():
            if code not in self.p2p_fares:
                self.p2p_fares[code] = mfs
        for code, mfs in connecting.items():
            mfs.direction = 'home'
            self.cnx_home_fares[code] = mfs
        self._log(f"Dest file: {len(direct)} P2P, {len(connecting)} cnx home markets")

    def build(self):
        """Compute fallback averages and validate."""
        if self.cnx_home_fares:
            self.fallback_cnx_home, self.fallback_cabin_split = self._weighted_avg(self.cnx_home_fares)
            self._log(f"Fallback cnx home: Y=${self.fallback_cnx_home.get('Y', 0):,.0f}, J=${self.fallback_cnx_home.get('J', 0):,.0f}")
        if self.cnx_dest_fares:
            self.fallback_cnx_dest, _ = self._weighted_avg(self.cnx_dest_fares)
            self._log(f"Fallback cnx dest: Y=${self.fallback_cnx_dest.get('Y', 0):,.0f}, J=${self.fallback_cnx_dest.get('J', 0):,.0f}")
        if self.p2p_fares:
            self.fallback_p2p, split = self._weighted_avg(self.p2p_fares)
            if not self.fallback_cabin_split:
                self.fallback_cabin_split = split
        for label, markets in [('cnx_home', self.cnx_home_fares), ('cnx_dest', self.cnx_dest_fares)]:
            low = sum(1 for m in markets.values() if m.low_sample)
            if low > 0:
                self._log(f"  WARNING: {low}/{len(markets)} {label} markets below {MIN_PAX_THRESHOLD} pax")
            outliers = sum(1 for m in markets.values() if m.fare_outliers)
            if outliers > 0:
                self._log(f"  WARNING: {outliers} {label} markets with fare outliers")
        self._log(f"Fare allocator built: {len(self.cnx_home_fares)} home, {len(self.cnx_dest_fares)} dest, {len(self.p2p_fares)} P2P")

    def _weighted_avg(self, markets):
        """Compute pax-weighted average fares across markets."""
        cabin_rev = defaultdict(float)
        cabin_pax = defaultdict(float)
        total_pax = 0.0
        for mfs in markets.values():
            for cabin in ['Y', 'PY', 'J', 'F']:
                fare = mfs.fares.get(cabin, 0)
                split = mfs.cabin_split.get(cabin, 0)
                pax = mfs.total_pax * split
                if fare > 0 and pax > 0:
                    cabin_rev[cabin] += fare * pax
                    cabin_pax[cabin] += pax
            total_pax += mfs.total_pax
        avg_fares = {}
        avg_split = {}
        for cabin in ['Y', 'PY', 'J', 'F']:
            if cabin_pax[cabin] > 0:
                avg_fares[cabin] = cabin_rev[cabin] / cabin_pax[cabin]
                avg_split[cabin] = cabin_pax[cabin] / total_pax if total_pax > 0 else 0
        return avg_fares, avg_split

    def get_connecting_fare(self, city_code, direction):
        """Get fare and cabin split for a connecting market. Returns (fares, splits)."""
        markets = self.cnx_home_fares if direction == 'home' else self.cnx_dest_fares
        fallback = self.fallback_cnx_home if direction == 'home' else self.fallback_cnx_dest
        mfs = markets.get(city_code)
        if mfs and mfs.fares:
            return mfs.fares, mfs.cabin_split
        return fallback, self.fallback_cabin_split

    def get_p2p_fares(self):
        """Get blended P2P fares and cabin split."""
        return self.fallback_p2p, self.fallback_cabin_split

    def compute_market_revenue(self, pipeline_results, fare_weight=None, fare_weight_cnx=None):
        """Compute per-market revenue using pipeline forecast pax and market-specific fares."""
        fw = fare_weight if fare_weight is not None else self.fare_weight
        fw_cnx = fare_weight_cnx if fare_weight_cnx is not None else fw
        result = {
            'p2p': self._compute_p2p_revenue(pipeline_results, fw),
            'cnx_home': self._compute_cnx_revenue(pipeline_results.get('home_results', []), 'home', fw_cnx),
            'cnx_dest': self._compute_cnx_revenue(pipeline_results.get('dest_results', []), 'dest', fw_cnx),
        }
        total_pax = result['p2p']['total_pax'] + result['cnx_home']['total_pax'] + result['cnx_dest']['total_pax']
        total_rev = result['p2p']['total_revenue'] + result['cnx_home']['total_revenue'] + result['cnx_dest']['total_revenue']
        result['summary'] = {
            'grand_total_pax': total_pax,
            'grand_total_revenue': total_rev,
            'overall_avg_fare': total_rev / total_pax if total_pax > 0 else 0,
            'p2p_pct_revenue': result['p2p']['total_revenue'] / total_rev * 100 if total_rev > 0 else 0,
            'cnx_home_pct_revenue': result['cnx_home']['total_revenue'] / total_rev * 100 if total_rev > 0 else 0,
            'cnx_dest_pct_revenue': result['cnx_dest']['total_revenue'] / total_rev * 100 if total_rev > 0 else 0,
            'markets_with_specific_fares': result['cnx_home']['markets_with_fares'] + result['cnx_dest']['markets_with_fares'],
            'markets_using_fallback': result['cnx_home']['markets_using_fallback'] + result['cnx_dest']['markets_using_fallback'],
        }
        return result

    def _compute_p2p_revenue(self, pipeline_results, fare_weight):
        total_pax = pipeline_results.get('p2p_total', 0)
        fares = self.fallback_p2p
        splits = self.fallback_cabin_split
        by_cabin = {}
        total_rev = 0.0
        for cabin in ['Y', 'PY', 'J', 'F']:
            fare = fares.get(cabin, 0)
            split = splits.get(cabin, 0)
            if fare > 0 and split > 0:
                pax = total_pax * split
                rev = pax * fare * fare_weight
                by_cabin[cabin] = {'pax': pax, 'fare_ow': fare * fare_weight, 'revenue': rev}
                total_rev += rev
        return {'total_pax': total_pax, 'total_revenue': total_rev,
                'avg_fare': total_rev / total_pax if total_pax > 0 else 0,
                'by_cabin': by_cabin, 'fares_used': fares, 'fare_weight': fare_weight}

    def _compute_cnx_revenue(self, market_results, direction, fare_weight_cnx):
        markets_detail = []
        total_pax = total_rev = 0.0
        markets_with_fares = markets_using_fallback = 0
        for mr in market_results:
            city = mr.get('city', '')
            pax = mr.get('forecast', 0)
            if pax <= 0:
                continue
            fares, splits = self.get_connecting_fare(city, direction)
            specific = self.cnx_home_fares if direction == 'home' else self.cnx_dest_fares
            if city in specific:
                fare_source = 'specific'
                markets_with_fares += 1
            else:
                fare_source = 'fallback'
                markets_using_fallback += 1
            market_rev = 0.0
            cabin_detail = {}
            for cabin in ['Y', 'PY', 'J', 'F']:
                fare = fares.get(cabin, 0)
                split = splits.get(cabin, 0)
                if fare > 0 and split > 0:
                    cpax = pax * split
                    crev = cpax * fare * fare_weight_cnx
                    cabin_detail[cabin] = {'pax': cpax, 'fare_ow': fare * fare_weight_cnx, 'revenue': crev}
                    market_rev += crev
            markets_detail.append({
                'city': city, 'name': mr.get('name', ''), 'pax': pax,
                'revenue': market_rev, 'avg_fare': market_rev / pax if pax > 0 else 0,
                'fare_source': fare_source, 'by_cabin': cabin_detail,
            })
            total_pax += pax
            total_rev += market_rev
        markets_detail.sort(key=lambda x: -x['revenue'])
        return {'total_pax': total_pax, 'total_revenue': total_rev,
                'avg_fare': total_rev / total_pax if total_pax > 0 else 0,
                'markets': markets_detail, 'markets_with_fares': markets_with_fares,
                'markets_using_fallback': markets_using_fallback}

    def to_revenue_config(self, **overrides):
        """Generate a RevenueConfig using weighted-average fares."""
        from revenue_forecast import RevenueConfig
        base = {
            'fares_p2p': self.fallback_p2p or overrides.get('fares_p2p', {}),
            'fares_cnx_home': self.fallback_cnx_home or overrides.get('fares_cnx_home', {}),
            'fares_cnx_dest': self.fallback_cnx_dest or overrides.get('fares_cnx_dest', {}),
            'cabin_split_p2p': self.fallback_cabin_split or overrides.get('cabin_split_p2p', {}),
            'cabin_split_cnx': self.fallback_cabin_split or overrides.get('cabin_split_cnx', {}),
            'fare_weight': overrides.get('fare_weight', self.fare_weight),
        }
        for k, v in overrides.items():
            if k not in base:
                base[k] = v
        return RevenueConfig(**base)


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

class FareAllocationWorkbook:
    """Writes branded Excel workbook showing per-market fare allocation."""

    HEADER_FILL = PatternFill(start_color='002060', end_color='002060', fill_type='solid') if HAS_OPENPYXL else None
    HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF') if HAS_OPENPYXL else None
    DATA_FONT = Font(name='Calibri', size=10) if HAS_OPENPYXL else None
    TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='002060') if HAS_OPENPYXL else None

    def __init__(self, allocator, market_revenue, route_label='', carrier=''):
        self.allocator = allocator
        self.mr = market_revenue
        self.route = route_label
        self.carrier = carrier
        if HAS_OPENPYXL:
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)

    def _header(self, ws, row, headers, widths=None):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = self.HEADER_FONT
            c.fill = self.HEADER_FILL
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            if widths and ci - 1 < len(widths):
                ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]

    def write_all(self):
        if not HAS_OPENPYXL:
            return
        self._write_summary()
        self._write_cnx('cnx_home', 'Connecting @ Home')
        self._write_cnx('cnx_dest', 'Connecting @ Dest')
        self._write_comparison()
        self._write_audit()

    def _write_summary(self):
        ws = self.wb.create_sheet('Revenue Summary')
        s = self.mr.get('summary', {})
        r = 1
        ws.cell(r, 1, f'Fare Allocation - {self.carrier} {self.route}').font = self.TITLE_FONT
        r += 2
        for label, val in [
            ('Grand Total Passengers', f"{s.get('grand_total_pax', 0):,.0f}"),
            ('Grand Total Revenue (USD)', f"${s.get('grand_total_revenue', 0):,.0f}"),
            ('Overall Average OW Fare', f"${s.get('overall_avg_fare', 0):,.2f}"),
            ('', ''),
            ('P2P Revenue Share', f"{s.get('p2p_pct_revenue', 0):.1f}%"),
            ('Cnx @ Home Revenue Share', f"{s.get('cnx_home_pct_revenue', 0):.1f}%"),
            ('Cnx @ Dest Revenue Share', f"{s.get('cnx_dest_pct_revenue', 0):.1f}%"),
            ('', ''),
            ('Markets w/ Specific Fares', str(s.get('markets_with_specific_fares', 0))),
            ('Markets Using Fallback', str(s.get('markets_using_fallback', 0))),
        ]:
            ws.cell(r, 1, label).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(r, 2, val).font = self.DATA_FONT
            r += 1
        r += 1
        self._header(ws, r, ['Segment', 'Passengers', 'Revenue (USD)', 'Avg Fare'], [25, 15, 18, 15])
        r += 1
        for key, lab in [('p2p', 'Point-to-Point'), ('cnx_home', 'Connecting @ Home'), ('cnx_dest', 'Connecting @ Dest')]:
            seg = self.mr.get(key, {})
            ws.cell(r, 1, lab)
            ws.cell(r, 2, seg.get('total_pax', 0)).number_format = '#,##0'
            ws.cell(r, 3, seg.get('total_revenue', 0)).number_format = '#,##0'
            ws.cell(r, 4, seg.get('avg_fare', 0)).number_format = '#,##0.00'
            r += 1
        ws.cell(r, 1, 'TOTAL').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(r, 2, s.get('grand_total_pax', 0)).number_format = '#,##0'
        ws.cell(r, 3, s.get('grand_total_revenue', 0)).number_format = '#,##0'
        ws.cell(r, 4, s.get('overall_avg_fare', 0)).number_format = '#,##0.00'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 22

    def _write_cnx(self, seg_key, sheet_name):
        ws = self.wb.create_sheet(sheet_name)
        seg = self.mr.get(seg_key, {})
        r = 1
        ws.cell(r, 1, f'{sheet_name} - Per-Market Revenue').font = self.TITLE_FONT
        r += 2
        self._header(ws, r, ['City', 'Name', 'Pax', 'Revenue', 'Avg Fare', 'Y Fare', 'PY Fare', 'J Fare', 'Y%', 'PY%', 'J%', 'Source'],
                     [8, 18, 12, 16, 12, 10, 10, 10, 8, 8, 8, 10])
        r += 1
        for m in seg.get('markets', []):
            ws.cell(r, 1, m['city'])
            ws.cell(r, 2, m.get('name', ''))
            ws.cell(r, 3, m['pax']).number_format = '#,##0'
            ws.cell(r, 4, m['revenue']).number_format = '#,##0'
            ws.cell(r, 5, m['avg_fare']).number_format = '#,##0.00'
            bc = m.get('by_cabin', {})
            for ci, cab in enumerate(['Y', 'PY', 'J'], 6):
                if cab in bc:
                    ws.cell(r, ci, bc[cab]['fare_ow']).number_format = '#,##0.00'
            spec = self.allocator.cnx_home_fares if 'home' in seg_key else self.allocator.cnx_dest_fares
            mfs = spec.get(m['city'])
            if mfs:
                for ci, cab in enumerate(['Y', 'PY', 'J'], 9):
                    sp = mfs.cabin_split.get(cab, 0)
                    if sp > 0:
                        ws.cell(r, ci, sp).number_format = '0.0%'
            ws.cell(r, 12, m.get('fare_source', ''))
            if m.get('fare_source') == 'fallback':
                for ci in range(1, 13):
                    ws.cell(r, ci).font = Font(name='Calibri', size=10, color='808080')
            r += 1
        r += 1
        ws.cell(r, 1, 'TOTAL').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(r, 3, seg.get('total_pax', 0)).number_format = '#,##0'
        ws.cell(r, 4, seg.get('total_revenue', 0)).number_format = '#,##0'
        ws.cell(r, 5, seg.get('avg_fare', 0)).number_format = '#,##0.00'

    def _write_comparison(self):
        ws = self.wb.create_sheet('Fare Comparison')
        r = 1
        ws.cell(r, 1, 'Per-Market vs Simple-Average Fare Impact').font = self.TITLE_FONT
        r += 2
        self._header(ws, r, ['Segment', 'Simple Avg', 'Weighted Avg', 'Diff', '% Impact'], [25, 14, 14, 12, 10])
        r += 1
        for key, lab in [('cnx_home', 'Connecting @ Home'), ('cnx_dest', 'Connecting @ Dest')]:
            mks = self.mr.get(key, {}).get('markets', [])
            if not mks:
                continue
            fares = [m['avg_fare'] for m in mks if m.get('avg_fare', 0) > 0]
            simple = sum(fares) / len(fares) if fares else 0
            weighted = self.mr[key].get('avg_fare', 0)
            diff = weighted - simple
            pct = diff / simple if simple > 0 else 0
            ws.cell(r, 1, lab)
            ws.cell(r, 2, simple).number_format = '#,##0.00'
            ws.cell(r, 3, weighted).number_format = '#,##0.00'
            ws.cell(r, 4, diff).number_format = '#,##0.00'
            ws.cell(r, 5, pct).number_format = '+0.0%;-0.0%'
            r += 1
        ws.column_dimensions['A'].width = 25

    def _write_audit(self):
        ws = self.wb.create_sheet('Audit Trail')
        ws.cell(1, 1, 'Fare Allocation Audit Trail').font = self.TITLE_FONT
        for i, line in enumerate(self.allocator.audit, 3):
            ws.cell(i, 1, line).font = Font(name='Calibri', size=9, color='333333')
        ws.column_dimensions['A'].width = 80

    def save(self, path):
        self.wb.save(path)


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def parse_fare_files(home_file=None, dest_file=None, cabin_map=None, fare_weight=0.85):
    """One-call parser for Sabre fare files. Returns FareAllocator."""
    allocator = FareAllocator(cabin_map=cabin_map, fare_weight=fare_weight)
    if home_file:
        allocator.add_home_file(home_file)
    if dest_file:
        allocator.add_dest_file(dest_file)
    allocator.build()
    return allocator


def compute_market_revenue(allocator, pipeline_results, fare_weight=0.85,
                            fare_weight_cnx=None, output_path=None,
                            route_label='', carrier=''):
    """One-call market revenue computation with optional Excel output."""
    mr = allocator.compute_market_revenue(pipeline_results, fare_weight=fare_weight,
                                          fare_weight_cnx=fare_weight_cnx)
    if output_path and HAS_OPENPYXL:
        writer = FareAllocationWorkbook(allocator, mr, route_label=route_label, carrier=carrier)
        writer.write_all()
        writer.save(output_path)
        print(f"Fare allocation workbook: {output_path}")
    return mr


# ============================================================================
# VALIDATION: BA LHR-SJC
# ============================================================================

def validate_ba_lhr_sjc():
    """Validate fare parsing and allocation against BA LHR-SJC fare files."""
    import os
    base = str(REFERENCE_CASE_DIR)
    home_file = os.path.join(base, 'BALHR__SJC__XXX_Sep2013Aug2014_Fares_v2.xlsx')
    dest_file = os.path.join(base, 'BASJC__LHR__XXX_Sep2013Aug2014_FARES_v2.xlsx')

    print("=" * 60)
    print("VALIDATION: BA LHR-SJC Fare Allocation")
    print("=" * 60)

    allocator = parse_fare_files(home_file, dest_file, fare_weight=0.85)

    print(f"\n  Parsing results:")
    print(f"    P2P markets:       {len(allocator.p2p_fares)}")
    print(f"    Cnx @ Home (LHR):  {len(allocator.cnx_home_fares)}")
    print(f"    Cnx @ Dest (SJC):  {len(allocator.cnx_dest_fares)}")

    print(f"\n  Sample connecting fares at LHR (raw, before fare_weight):")
    for city in ['PAR', 'FRA', 'AMS', 'DXB', 'DEL', 'MUC']:
        fares, splits = allocator.get_connecting_fare(city, 'home')
        if fares:
            print(f"    {city}: Y=${fares.get('Y', 0):,.0f}, J=${fares.get('J', 0):,.0f}")

    print(f"\n  Sample connecting fares at SJC (raw):")
    for city in ['LAX', 'PHX', 'SEA']:
        fares, splits = allocator.get_connecting_fare(city, 'dest')
        if fares:
            print(f"    {city}: Y=${fares.get('Y', 0):,.0f}, J=${fares.get('J', 0):,.0f}")

    pipeline_results = {
        'p2p_total': 78110, 'home_total': 48115, 'dest_total': 2937,
        'grand_total': 129162, 'load_factor': 0.829,
        'home_results': [
            {'city': 'PAR', 'name': 'Paris', 'forecast': 5800},
            {'city': 'FRA', 'name': 'Frankfurt', 'forecast': 4200},
            {'city': 'AMS', 'name': 'Amsterdam', 'forecast': 3900},
            {'city': 'MUC', 'name': 'Munich', 'forecast': 2800},
            {'city': 'DXB', 'name': 'Dubai', 'forecast': 2500},
            {'city': 'DUB', 'name': 'Dublin', 'forecast': 2200},
            {'city': 'ZRH', 'name': 'Zurich', 'forecast': 2100},
            {'city': 'DEL', 'name': 'Delhi', 'forecast': 1900},
            {'city': 'BOM', 'name': 'Mumbai', 'forecast': 1700},
            {'city': 'BCN', 'name': 'Barcelona', 'forecast': 1800},
            {'city': 'MAD', 'name': 'Madrid', 'forecast': 1600},
            {'city': 'CPH', 'name': 'Copenhagen', 'forecast': 1400},
            {'city': 'OTHER', 'name': 'Other Markets', 'forecast': 16115},
        ],
        'dest_results': [
            {'city': 'LAX', 'name': 'Los Angeles', 'forecast': 1200},
            {'city': 'SEA', 'name': 'Seattle', 'forecast': 800},
            {'city': 'PHX', 'name': 'Phoenix', 'forecast': 500},
            {'city': 'OTHER', 'name': 'Other', 'forecast': 437},
        ],
    }

    ensure_output_dir()
    mr = compute_market_revenue(
        allocator, pipeline_results, fare_weight=0.85,
        output_path=str(OUTPUT_DIR / 'BA_LHR_SJC_Fare_Allocation.xlsx'),
        route_label='LHR-SJC', carrier='BA',
    )

    s = mr['summary']
    print(f"\n  Revenue Results (per-market allocation):")
    print(f"    Grand Total Pax:     {s['grand_total_pax']:>12,.0f}")
    print(f"    Grand Total Revenue: ${s['grand_total_revenue']:>12,.0f}")
    print(f"    Overall Avg Fare:    ${s['overall_avg_fare']:>12,.2f}")
    print(f"    P2P Rev Share:       {s['p2p_pct_revenue']:>12.1f}%")
    print(f"    Cnx Home Rev Share:  {s['cnx_home_pct_revenue']:>12.1f}%")
    print(f"    Markets w/ fares:    {s['markets_with_specific_fares']}")
    print(f"    Markets fallback:    {s['markets_using_fallback']}")

    print(f"\n  Top 10 Connecting @ LHR (by revenue):")
    for m in mr['cnx_home']['markets'][:10]:
        print(f"    {m['city']:5s} {m['name']:18s} {m['pax']:>7,.0f} pax  "
              f"${m['revenue']:>12,.0f}  avg ${m['avg_fare']:>7,.0f}  [{m['fare_source']}]")

    print(f"\n  Reasonableness:")
    for label, ok in [
        ("Overall avg fare $300-$2000", 300 < s['overall_avg_fare'] < 2000),
        ("Markets with fares > 5", s['markets_with_specific_fares'] > 5),
        ("Total pax matches pipeline", abs(s['grand_total_pax'] - 129162) < 1),
    ]:
        print(f"    {'OK' if ok else 'XX'} {label}")

    return mr


if __name__ == '__main__':
    validate_ba_lhr_sjc()
