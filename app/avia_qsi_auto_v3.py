#!/usr/bin/env python3
"""
Avia Solutions -- QSI Route Forecast Portal v1.0
==========================================================
Streamlit web portal with two data-source paths and 16 tabs:

Tab 1: Data Upload & Run (Path A: Pre-computed, Path B: New Route)
Tab 2: Results (forecast breakdown + CRE assessment)
Tab 3: Monthly Profile (seasonality engine)
Tab 4: Revenue Forecast (pax + cargo + ancillary)
Tab 5: Assumptions Log (72-parameter methodology summary)
Tab 6: Business Case (goal-seek + sensitivity)
Tab 7: Output Workbook (standardised Excel output)
Tab 8: Q&A Checklist (automated quality control)
Tab 9: Spill Analysis (capacity constraint tool)
Tab 10: Market Research (brief + executor with findings tracking)
Tab 11: Comparison (analyst vs pipeline)
Tab 12: Validation (cross-route regression suite)
Tab 13: Time Grid (departure time search - new route only)
Tab 14: Calibration Engine (predictive parameter suggestion)
Tab 15: Fare Allocation (per-market fares from Sabre data)
Tab 16: Presentation (PPTX generator with buyer-type variants)

v12 changes:
  - CRE (Commercial Reasonableness Engine) wired as post-pipeline step
    * Auto-runs after every pipeline execution
    * Confidence score badge (HIGH/MODERATE/LOW/VERY LOW)
    * Raw vs adjusted side-by-side when adjustments made
    * Expandable checks table, adjustment trail, frequency optimisation
    * Analyst review flag with reasons
  - All existing functionality preserved

Regression target: BA LHR-SJC = 129,162 passengers (Path A, unchanged).
"""

import os
import sys
import io
import tempfile
import traceback
from datetime import datetime, time as dtime
from typing import Dict, Optional, Any, List

import streamlit as st
import pandas as pd

#  Pipeline imports 
from providers import (
    ExcelScheduleProvider, ExcelDemandProvider,
    P2PSegmentData, P2PSubsegmentData, ConnectingCityData,
)
from single_extract_oag_provider import SingleExtractOAGProvider
from midt_demand_provider import MIDTDemandProvider
from route_config import RouteConfig
from closed_loop_pipeline_v2 import run_pipeline
from departure_time_grid import (
    TimeShiftProvider, TimeGridRunner, GridSearchResult,
    write_grid_output,
)
from input_validator import (
    KNOWN_AIRPORTS, AIRCRAFT_DB,
    RouteInput, InputValidator, ValidationResult as InputValidationResult,
    lookup_airport, compute_distance, classify_distance_band,
)
from seasonality_engine import (
    SeasonalProfile, PROFILE_LIBRARY, MONTHS, QUARTERS, DAYS_IN_MONTH, DAYS_IN_YEAR,
    select_profile, blend_profiles, from_quarterly, from_monthly_pax,
    distribute_annual, distribute_with_spill, monthly_revenue,
    seasonalise_pipeline_output, MonthlyForecast,
)
import copy
import math
import json

#  New module imports (Chat 46: wiring) 
from assumptions_log import (
    generate_assumptions_log, AssumptionsLog, AssumptionsLogExcelWriter,
    AssumptionsLogBuilder,
)
from business_case_mode import (
    BusinessCaseEngine, BusinessCaseWriter, TargetSet, BusinessCaseVerdict,
    run_business_case, RAMP_PROFILES,
)
from output_workbook import StandardOutputWriter
from cross_route_validator import (
    RouteValidationCase, ValidationResult,
    case_ba_lhr_sjc, case_ke_icn_sjc_7x, case_ke_icn_sjc_5x,
    case_sq_sin_sjc, case_cx_hkg_sjc, case_fi_kef_sjc,
    test_forecast_table_math, test_ptew_calculations,
    test_capacity_load_factor, test_connecting_city_aggregation,
    test_parameter_reasonableness, test_cross_route_patterns,
    run_all_validations, write_validation_workbook,
)

#  New module imports (Chat 47: Q&A + Market Research + Spill) 
from qa_checklist import run_qa_checklist, QAReport, CheckStatus, CheckCategory
from market_research_module import (
    RouteResearchConfig, DemandProfile, RouteType as MRRouteType,
    BuyerType, generate_queries, summarise_research_plan,
    get_relevance_matrix, Relevance,
)

#  Market Research Executor (Chat 51: wiring) 
from market_research_executor import (
    ResearchExecutor, ResearchOutput, BlockFindings, Finding, Citation,
    make_finding, add_citation,
    get_execution_sequence, generate_execution_plan,
    generate_markdown_report, generate_json_output, generate_excel_output,
    quick_research,
)


def _repair_truncated_json(text: str) -> str:
    """Attempt to repair truncated JSON from Claude hitting output limit.
    
    Strategy: walk backwards closing any open strings, arrays, objects.
    Returns repaired JSON string or empty string if unrepairable.
    """
    if not text or not text.strip():
        return ""
    
    s = text.rstrip()
    
    # If it already parses, return as-is
    try:
        json.loads(s)
        return s
    except Exception:
        pass
    
    # Find the last complete entry by looking for the last complete object/value
    # Strategy: truncate to last valid comma or closing bracket, then close all open brackets
    
    # Check if we're inside an unterminated string
    in_string = False
    escape = False
    last_good = 0
    
    for i, ch in enumerate(s):
        if escape:
            escape = False
            continue
        if ch == '\\':
            escape = True
            continue
        if ch == '"':
            in_string = not in_string
        if not in_string and ch in ',]}':
            last_good = i
    
    # Truncate to last complete value
    if last_good > 0:
        s = s[:last_good + 1]
    
    # Remove any trailing comma
    s = s.rstrip().rstrip(',')
    
    # Count open brackets and close them
    opens = 0
    open_sq = 0
    in_str = False
    esc = False
    for ch in s:
        if esc:
            esc = False
            continue
        if ch == '\\':
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
        if not in_str:
            if ch == '{': opens += 1
            elif ch == '}': opens -= 1
            elif ch == '[': open_sq += 1
            elif ch == ']': open_sq -= 1
    
    # Close any open brackets
    s += ']' * open_sq + '}' * opens
    
    # Verify it parses
    try:
        json.loads(s)
        return s
    except Exception:
        return ""

#  Predictive Calibration Engine (Chat 50: wiring) 
from predictive_calibration_engine import (
    PredictiveCalibrationEngine, RouteProfile, CalibrationSuggestion,
    SegmentSuggestion, MarketMaturity, HubStatus, CarrierType, Confidence,
    from_route_config, quick_predict, write_suggestion_excel,
    SEGMENT_NAMES,
)

# ---- Fare Allocation (Chat 52: wiring) ----
from fare_allocation import (
    FareAllocator, FareParser, MarketFareSet, FareAllocationWorkbook,
    parse_fare_files, compute_market_revenue,
)

# ---- Commercial Reasonableness Engine (CRE) ----
try:
    from commercial_reasonableness_engine import (
        CommercialReasonablenessEngine, CREResult, CommercialBounds,
        CheckStatus as CRECheckStatus, run_cre, get_bounds_for_route_type,
        FrequencyRecommendation, AdjustmentRecord, ReasonablenessCheck,
    )
    HAS_CRE = True
except ImportError:
    HAS_CRE = False

# ---- CRE-PCE Bridge (Informed Closed-Loop) ----
try:
    from cre_pce_bridge import (
        CREPCEBridge, InformedClosedLoopRunner, InformedClosedLoopResult,
        InformedRerunRecord, run_informed_closed_loop, apply_pce_to_config,
    )
    HAS_BRIDGE = True
except ImportError:
    HAS_BRIDGE = False

# ---- Learning Comparison (Step 4: analyst benchmark extraction) ----
try:
    from learning_comparison import (
        LearningComparison, LearningReport, AnalystData, CityComparison,
        HeadlineComparison, CalibrationLearning, parse_analyst_workbook,
        run_learning_comparison,
    )
    HAS_LEARNING = True
except ImportError:
    HAS_LEARNING = False

# ---- Narrative Generator (Chat 53: presentation automation) ----
try:
    from narrative_generator import NarrativeGenerator, generate_narrative
    HAS_NARRATIVE = True
except ImportError:
    HAS_NARRATIVE = False


# ============================================================================
# PAGE CONFIG
# ============================================================================

st.set_page_config(
    page_title="Avia Solutions  QSI Route Forecast",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================================
# CUSTOM STYLING
# ============================================================================

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    .stApp { font-family: 'DM Sans', sans-serif; }

    .portal-header {
        background: #ffffff;
        padding: 1.5rem 2rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        border-bottom: 3px solid #0d3b66;
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    .portal-header h1 {
        color: #0d3b66; font-family: 'DM Sans', sans-serif;
        font-weight: 700; font-size: 1.8rem; margin: 0; letter-spacing: -0.5px;
    }
    .portal-header p { color: #5a7a9a; font-size: 0.9rem; margin: 0.3rem 0 0 0; }
    .portal-header .header-brand {
        color: #0d3b66; font-family: 'DM Sans', sans-serif;
        font-weight: 600; font-size: 1.1rem; text-align: right; white-space: nowrap;
    }

    .metric-card {
        background: #f8f9fc; border: 1px solid #e2e8f0;
        border-radius: 10px; padding: 1.2rem; text-align: center;
    }
    .metric-card .label {
        font-size: 0.75rem; color: #6b7c93; text-transform: uppercase;
        letter-spacing: 0.5px; font-weight: 600;
    }
    .metric-card .value {
        font-size: 1.6rem; font-weight: 700; color: #0a1628;
        font-family: 'JetBrains Mono', monospace; margin-top: 0.3rem;
    }
    .metric-card .sub { font-size: 0.7rem; color: #8ba4c4; margin-top: 0.2rem; }

    .verdict-pass {
        background: #d4edda; color: #155724; padding: 0.4rem 1rem;
        border-radius: 6px; font-weight: 700; display: inline-block; font-size: 0.85rem;
    }
    .verdict-fail {
        background: #f8d7da; color: #721c24; padding: 0.4rem 1rem;
        border-radius: 6px; font-weight: 700; display: inline-block; font-size: 0.85rem;
    }
    .verdict-marginal {
        background: #fff3cd; color: #856404; padding: 0.4rem 1rem;
        border-radius: 6px; font-weight: 700; display: inline-block; font-size: 0.85rem;
    }

    .section-label {
        font-size: 0.7rem; font-weight: 600; color: #6b7c93;
        text-transform: uppercase; letter-spacing: 1px;
        margin-bottom: 0.5rem; padding-bottom: 0.3rem; border-bottom: 2px solid #e8a83e;
    }

    .path-card {
        border: 2px solid #e2e8f0; border-radius: 10px; padding: 1rem;
        margin-bottom: 0.5rem;
    }
    .path-card.active { border-color: #e8a83e; background: #fdf8ef; }
    .path-card h4 { margin: 0 0 0.3rem 0; color: #0a1628; font-size: 0.95rem; }
    .path-card p { margin: 0; color: #6b7c93; font-size: 0.8rem; }

    .provider-badge {
        display: inline-block; padding: 0.2rem 0.6rem; border-radius: 12px;
        font-size: 0.7rem; font-weight: 600; letter-spacing: 0.5px;
    }
    .provider-badge.precomputed { background: #e2efda; color: #155724; }
    .provider-badge.newroute { background: #cce5ff; color: #004085; }

    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ============================================================================
# HEADER
# ============================================================================

# Embed logo as base64 for inline display in header
import os as _os, base64 as _b64
_logo_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "Avia_QSI_only.png")
_logo_b64 = ""
if _os.path.exists(_logo_path):
    with open(_logo_path, "rb") as _lf:
        _logo_b64 = _b64.b64encode(_lf.read()).decode()

_logo_html = ""
if _logo_b64:
    _logo_html = (f'<img src="data:image/png;base64,{_logo_b64}" '
                  f'alt="Avia Solutions" style="height:60px; object-fit:contain;">')

st.markdown(f"""
<div class="portal-header">
    <div>
        <h1>QSI Route Forecast Portal</h1>
        <p>Avia Solutions Quality of Service Index Pipeline v8.0 &nbsp; | &nbsp;
        38 modules &nbsp; | &nbsp; Pre-computed + New Route Assessment &nbsp; | &nbsp; 16 tabs</p>
    </div>
    <div style="flex-shrink:0;">{_logo_html}</div>
</div>
""", unsafe_allow_html=True)


# ============================================================================
# SESSION STATE
# ============================================================================

defaults = {
    'results': None,
    'output_bytes': None,
    'run_log': [],
    'last_config_summary': "",
    'last_config': None,         # RouteConfig from last pipeline run
    'p2p_segments': [],
    'data_source': 'precomputed',
    'grid_results': [],
    'grid_output_bytes': None,
    'grid_search_result': None,
    'last_grid_config': None,
    'seasonality_result': None,
    'revenue_result': None,
    'assumptions_log': None,     # AssumptionsLog object
    'business_case_verdict': None,  # BusinessCaseVerdict
    'business_case_bytes': None,    # Excel download
    'validation_results': None,     # Cross-route validation results
    'validation_bytes': None,       # Validation Excel download
    'calib_suggestion': None,       # CalibrationSuggestion from predictive engine
    'calib_excel_bytes': None,      # Calibration suggestion Excel download
    'fare_allocator': None,         # FareAllocator object
    'fare_market_revenue': None,    # Per-market revenue dict
    'fare_excel_bytes': None,       # Fare allocation Excel download
    'cre_result': None,             # CREResult from Commercial Reasonableness Engine
    'closed_loop_result': None,     # InformedClosedLoopResult from CRE-PCE bridge
    'learning_report': None,        # LearningReport from analyst benchmark comparison
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ============================================================================
# HELPERS
# ============================================================================

def save_uploaded(uploaded_file) -> str:
    """Save an uploaded file to a temp location and return the path."""
    tmp_dir = tempfile.mkdtemp()
    path = os.path.join(tmp_dir, uploaded_file.name)
    with open(path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    return path


def save_multiple_uploads(uploaded_files) -> List[str]:
    """Save multiple uploaded files and return list of paths."""
    paths = []
    for uf in (uploaded_files or []):
        paths.append(save_uploaded(uf))
    return paths


def fmt(n, decimals=0):
    if n is None:
        return ""
    if isinstance(n, float) and 0 < n < 1:
        return f"{n:.1%}"
    if decimals > 0:
        return f"{n:,.{decimals}f}"
    return f"{int(n):,}"


def metric_card(label, value, sub=""):
    sub_html = f'<div class="sub">{sub}</div>' if sub else ""
    return f'<div class="metric-card"><div class="label">{label}</div><div class="value">{value}</div>{sub_html}</div>'


# ============================================================================
# SIDEBAR  ROUTE CONFIGURATION
# ============================================================================

with st.sidebar:
    st.markdown("###  Route Configuration")

    #  DATA SOURCE SELECTION 
    st.markdown('<div class="section-label">Data Source</div>', unsafe_allow_html=True)
    data_source = st.radio(
        "Assessment type",
        ["Pre-computed QSI Files", "New Route Assessment"],
        horizontal=True,
        help="Pre-computed: use existing QSI workbooks. New Route: build from raw OAG legs + MIDT demand.",
        label_visibility="collapsed",
    )
    st.session_state.data_source = 'precomputed' if data_source == "Pre-computed QSI Files" else 'newroute'

    if st.session_state.data_source == 'precomputed':
        st.markdown('<span class="provider-badge precomputed">VALIDATED PATH</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="provider-badge newroute">NEW ROUTE PATH</span>', unsafe_allow_html=True)

    #  Operating Mode 
    st.markdown('<div class="section-label">Operating Mode</div>', unsafe_allow_html=True)
    mode = st.selectbox("Mode", ["Forecast", "Business Case"])

    #  Business Case Targets (visible only in BC mode) 
    if mode == "Business Case":
        st.markdown('<div class="section-label">Business Case Targets</div>', unsafe_allow_html=True)
        bc_lf_y1 = st.number_input("Y1 Load Factor Target", min_value=0.30, max_value=0.95,
                                    value=0.70, step=0.05, format="%.2f",
                                    help="Minimum load factor the airline needs in Year 1")
        bc_lf_mature = st.number_input("Mature Load Factor Target", min_value=0.50, max_value=0.95,
                                        value=0.82, step=0.02, format="%.2f",
                                        help="Target load factor at maturity (Year 3+)")
        bc_ramp_years = st.number_input("Ramp-up Years", min_value=1, max_value=5, value=3)
        bc_ramp_profile = st.selectbox("Ramp Profile",
                                        list(RAMP_PROFILES.keys()),
                                        index=0,
                                        help="How quickly demand builds to maturity")
    else:
        bc_lf_y1 = 0.70
        bc_lf_mature = 0.82
        bc_ramp_years = 3
        bc_ramp_profile = 'standard'

    #  Route Identity 
    st.markdown('<div class="section-label">Route Identity</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        origin = st.text_input("Origin (IATA)", value="", max_chars=3, placeholder="LHR").upper().strip()
    with col2:
        destination = st.text_input("Dest (IATA)", value="", max_chars=3, placeholder="SJC").upper().strip()

    carrier_name = st.text_input("Airline Name", placeholder="British Airways")
    carrier_code = st.text_input("Airline Code", value="", max_chars=2, placeholder="BA").upper().strip()

    carrier_type = st.selectbox("Carrier Type",
        ["Full Service", "LCC", "Ultra LCC", "Hybrid", "Charter"])
    ct_map = {"Full Service": "full_service", "LCC": "lcc", "Ultra LCC": "ultra_lcc",
              "Hybrid": "hybrid", "Charter": "charter"}

    route_type = st.selectbox("Route Type",
        ["Long Haul", "Hub Feed", "LCC Point-to-Point", "Charter/Leisure", "Mixed"])

    market_maturity = st.selectbox("Market Maturity",
        ["New Route", "Existing Underserved", "Existing Competitive", "Mature"])

    #  Service Parameters 
    st.markdown('<div class="section-label">Service Parameters</div>', unsafe_allow_html=True)
    frequency = st.number_input("Weekly Frequency", min_value=1, max_value=28, value=7)

    ac_options = sorted(AIRCRAFT_DB.keys())
    ac_display = [f"{k}  {AIRCRAFT_DB[k].name} ({AIRCRAFT_DB[k].typical_seats}s)" for k in ac_options]
    ac_idx = st.selectbox("Aircraft Type", range(len(ac_options)),
                          format_func=lambda i: ac_display[i],
                          index=ac_options.index('787') if '787' in ac_options else 0)
    aircraft_type = ac_options[ac_idx]
    ac_spec = AIRCRAFT_DB[aircraft_type]
    seats = st.number_input("Seats", min_value=50, max_value=900, value=ac_spec.typical_seats)

    #  QSI Coefficients 
    st.markdown('<div class="section-label">QSI Coefficients</div>', unsafe_allow_html=True)
    is_lcc = ct_map[carrier_type] in ('lcc', 'ultra_lcc')
    online_coeff = st.number_input("Online", min_value=0.0, max_value=2.0, value=1.0, step=0.05, format="%.3f")
    alliance_coeff = st.number_input("Alliance", min_value=0.0, max_value=2.0,
                                     value=0.0 if is_lcc else 0.615, step=0.05, format="%.3f")
    interline_coeff = st.number_input("Interline", min_value=0.0, max_value=2.0,
                                      value=0.0 if is_lcc else 0.25, step=0.05, format="%.3f")
    qsi_ceiling = st.number_input("QSI Ceiling", min_value=0.1, max_value=2.0, value=1.0, step=0.05, format="%.2f")

    #  Growth Assumptions 
    st.markdown('<div class="section-label">Growth Assumptions</div>', unsafe_allow_html=True)
    home_growth = st.number_input("Home Cnx Growth", min_value=-0.10, max_value=0.30, value=0.04, step=0.01, format="%.2f",
                                  help="Annual growth for connecting traffic at home hub")
    dest_growth = st.number_input("Dest Cnx Growth", min_value=-0.10, max_value=0.30, value=0.04, step=0.01, format="%.2f",
                                  help="Annual growth for connecting traffic at destination hub")

    st.markdown('<div class="section-label">P2P Base & Forecast Year</div>', unsafe_allow_html=True)
    sabre_base_year = st.number_input("Sabre Base Year", min_value=2015, max_value=2030,
                                       value=2022, step=1,
                                       help="Year of Sabre/MIDT data (e.g., 2022). Growth is compounded from this year to forecast year.")
    p2p_forecast_year = st.number_input("Forecast Year", min_value=2020, max_value=2035,
                                         value=2026, step=1,
                                         help="Target forecast year. P2P growth compounds over (Forecast - Base) years.")

    #  New Route: MIDT settings 
    if st.session_state.data_source == 'newroute':
        st.markdown('<div class="section-label">MIDT / Sabre Settings</div>', unsafe_allow_html=True)
        catchment_share_home = st.number_input(
            "Catchment Share (Home)", min_value=0.0, max_value=1.0, value=1.0,
            step=0.05, format="%.2f",
            help="Share of broader catchment captured at home airport (e.g., SJC share of Bay Area)")
        catchment_share_dest = st.number_input(
            "Catchment Share (Dest)", min_value=0.0, max_value=1.0, value=1.0,
            step=0.05, format="%.2f")
        min_demand_threshold = st.number_input(
            "Min City Demand", min_value=0, max_value=1000, value=0,
            help="Minimum annual pax to include a connecting city")
    else:
        catchment_share_home = 1.0
        catchment_share_dest = 1.0
        min_demand_threshold = 0

    # -- Seasonality Profile --
    st.markdown('<div class="section-label">Seasonality Profile</div>', unsafe_allow_html=True)
    profile_options = list(PROFILE_LIBRARY.keys())
    profile_display = {
        'transatlantic': 'Transatlantic (summer peak)',
        'europe_asia': 'Europe-Asia (mild peak)',
        'us_asia': 'US-Asia Pacific',
        'intra_europe': 'Intra-European (strong peak)',
        'middle_east': 'Middle East (winter peak)',
        'flat': 'Year-Round Flat',
        'business_heavy': 'Business-Heavy',
        'leisure_heavy': 'Leisure-Heavy',
        'ba_lhr_sjc': 'BA LHR-SJC (actual)',
    }
    season_profile_key = st.selectbox(
        "Seasonal Profile",
        profile_options,
        format_func=lambda k: profile_display.get(k, k),
        index=0,
    )

    # -- Revenue Assumptions --
    st.markdown('<div class="section-label">Revenue Assumptions</div>', unsafe_allow_html=True)
    avg_ow_fare = st.number_input("Avg One-Way Fare ($)", min_value=50, max_value=5000,
                                   value=750, step=50,
                                   help="Blended average one-way fare across all cabins")
    fare_weight = st.number_input("Fare Weight", min_value=0.5, max_value=1.0,
                                   value=0.85, step=0.05, format="%.2f",
                                   help="Discount factor on Sabre raw fares (0.85 typical)")
    cargo_capacity_kg = st.number_input("Cargo Capacity (kg/flight)", min_value=0, max_value=50000,
                                         value=15000, step=1000)
    cargo_lf = st.number_input("Cargo Load Factor", min_value=0.0, max_value=1.0,
                                value=0.60, step=0.05, format="%.2f")
    cargo_yield = st.number_input("Cargo Yield ($/kg)", min_value=0.0, max_value=5.0,
                                   value=1.75, step=0.25, format="%.2f")
    ancillary_per_pax = st.number_input("Ancillary $/pax", min_value=0.0, max_value=100.0,
                                         value=20.0, step=5.0, format="%.1f")

    # -- Distance calculation --
    if origin and destination and len(origin) == 3 and len(destination) == 3:
        dist = compute_distance(origin, destination)
        if dist:
            band = classify_distance_band(dist)
            st.info(f" {origin}{destination}: {dist:,.0f} nm ({band.replace('_', ' ')})")


# ============================================================================
# MAIN TABS
# ============================================================================

if st.session_state.data_source == 'newroute':
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13, tab14, tab15, tab16 = st.tabs([
        "Data Upload & Run", "Results", "Monthly Profile", "Revenue",
        "Assumptions", "Business Case", "Output Workbook",
        "Q&A Checklist", "Spill Analysis", "Market Research",
        "Comparison", "Validation", "Time Grid", "Calibration Engine",
        "Fare Allocation", " Presentation"
    ])
else:
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab14, tab15, tab16 = st.tabs([
        "Data Upload & Run", "Results", "Monthly Profile", "Revenue",
        "Assumptions", "Business Case", "Output Workbook",
        "Q&A Checklist", "Spill Analysis", "Market Research",
        "Comparison", "Validation", "Calibration Engine",
        "Fare Allocation", " Presentation"
    ])
    tab13 = None


# ============================================================================
# TAB 1: UPLOAD & RUN
# ============================================================================

with tab1:

    # ================================================================
    # PATH A: PRE-COMPUTED QSI FILES
    # ================================================================
    if st.session_state.data_source == 'precomputed':
        st.markdown("### Upload Pre-computed QSI Files")
        st.caption("Upload the analyst-prepared QSI workbooks and forecast file. "
                    "This is the validated path matching Chat 22.")

        col_up1, col_up2, col_up3 = st.columns(3)
        with col_up1:
            st.markdown(f"**Home Hub QSI File** (QSI{origin or 'LHR'}.xlsx)")
            home_qsi_upload = st.file_uploader("Home QSI", type=['xlsx', 'xlsm', 'xls'],
                                               key='home_qsi', label_visibility="collapsed")
        with col_up2:
            st.markdown(f"**Destination QSI File** (QSI{destination or 'SJC'}.xlsx)")
            dest_qsi_upload = st.file_uploader("Dest QSI", type=['xlsx', 'xlsm', 'xls'],
                                               key='dest_qsi', label_visibility="collapsed")
        with col_up3:
            st.markdown("**Forecast / Demand File**")
            forecast_upload = st.file_uploader("Forecast", type=['xlsx', 'xlsm', 'xls'],
                                               key='forecast', label_visibility="collapsed")

    # ================================================================
    # PATH B: NEW ROUTE ASSESSMENT
    # ================================================================
    else:
        st.markdown("### New Route Assessment - Smart Upload")
        st.caption("Drop all your files at once. The classifier will auto-sort them into "
                    "QSI templates, MCT files, connecting demand, P2P demand, and city lookups.")

        all_uploads = st.file_uploader(
            "Upload all route data files",
            type=['xlsx', 'xlsm', 'xls'],
            key='smart_upload',
            accept_multiple_files=True,
            help="Upload everything - QSI template, MCT files, MIDT demand files, city lookups. "
                 "The system will classify each file automatically.")

        # Auto-classify uploads into buckets
        qsi_template_upload = None
        mct_uploads = []
        home_cnx_uploads = []
        dest_cnx_uploads = []
        p2p_uploads = []
        forecast_uploads = []
        city_lookup_upload = None

        if all_uploads:
            try:
                from file_classifier import classify_file, FileType
                has_classifier = True
            except ImportError:
                has_classifier = False

            if has_classifier:
                classified = {}
                for uf in all_uploads:
                    try:
                        uf.seek(0)
                        ft = classify_file(uf)
                        uf.seek(0)
                        classified[uf.name] = (ft, uf)
                    except Exception:
                        uf.seek(0)
                        classified[uf.name] = (FileType.UNKNOWN, uf)

                # Sort into buckets - TWO PASS to ensure QSI files take priority
                # Pass 1: Assign QSI template/scored files first
                for fname, (ft, uf) in classified.items():
                    if ft == FileType.QSI_TEMPLATE or ft == FileType.QSI_SCORED:
                        if qsi_template_upload is None:
                            qsi_template_upload = uf

                # Pass 2: Assign everything else
                for fname, (ft, uf) in classified.items():
                    fn_upper = fname.upper()
                    if ft == FileType.QSI_TEMPLATE or ft == FileType.QSI_SCORED:
                        continue  # Already handled in pass 1
                    elif ft == FileType.MCT:
                        mct_uploads.append(uf)
                    elif ft == FileType.MIDT_CONNECTING:
                        # Decide home vs dest from FILE CONTENT, not filename
                        # Read Pivot-P2P to find 'Mod Org City' value
                        routed = False
                        org = (origin or '').upper()
                        dst = (destination or '').upper()
                        if org or dst:
                            try:
                                uf.seek(0)
                                import openpyxl as _opx
                                _wb = _opx.load_workbook(uf, read_only=True, data_only=True)
                                for _sn in _wb.sheetnames:
                                    if 'pivot' in _sn.lower() and 'p2p' in _sn.lower():
                                        _ws = _wb[_sn]
                                        for _r in range(1, 10):
                                            _v0 = _ws.cell(_r, 1).value
                                            _v1 = _ws.cell(_r, 2).value
                                            if _v0 and 'mod org' in str(_v0).lower() and _v1:
                                                mod_org = str(_v1).strip().upper()
                                                # If origin city matches home airport -> home connecting
                                                # (traffic FROM home hub to cities beyond dest)
                                                # If origin city matches dest airport -> traffic
                                                # originates at dest, connects through home hub
                                                if mod_org == org:
                                                    # Origin is home airport - this is traffic
                                                    # FROM home connecting through dest hub
                                                    dest_cnx_uploads.append(uf)
                                                elif mod_org == dst:
                                                    # Origin is dest airport - this is traffic
                                                    # FROM dest connecting through home hub
                                                    home_cnx_uploads.append(uf)
                                                else:
                                                    # Can't determine - default to home
                                                    home_cnx_uploads.append(uf)
                                                routed = True
                                                break
                                        break
                                _wb.close()
                                uf.seek(0)
                            except Exception:
                                if hasattr(uf, 'seek'):
                                    uf.seek(0)
                        if not routed:
                            # Fallback: default to home connecting
                            home_cnx_uploads.append(uf)
                    elif ft == FileType.P2P_DEMAND:
                        p2p_uploads.append(uf)
                    elif ft == FileType.CITY_LOOKUP:
                        city_lookup_upload = uf
                    elif ft == FileType.OAG_SCHEDULE:
                        if qsi_template_upload is None:
                            qsi_template_upload = uf
                    # else: skip unknown/fare files
                    elif ft == FileType.FORECAST:
                        forecast_uploads.append(uf)

                # Display classification results
                st.markdown("**File Classification Results:**")
                status_icons = {
                    FileType.QSI_TEMPLATE: "QSI Template", FileType.QSI_SCORED: "QSI Scored",
                    FileType.MCT: "MCT", FileType.MIDT_CONNECTING: "Connecting Demand",
                    FileType.P2P_DEMAND: "P2P Demand", FileType.CITY_LOOKUP: "City Lookup",
                    FileType.OAG_SCHEDULE: "OAG Schedule",
                }
                for fname, (ft, uf) in classified.items():
                    label = status_icons.get(ft, ft.value)
                    st.caption(f"  {fname} -> {label}")

                # Summary
                parts = []
                if qsi_template_upload:
                    parts.append(f"1 QSI template")
                if mct_uploads:
                    parts.append(f"{len(mct_uploads)} MCT")
                if home_cnx_uploads:
                    parts.append(f"{len(home_cnx_uploads)} home connecting")
                if dest_cnx_uploads:
                    parts.append(f"{len(dest_cnx_uploads)} dest connecting")
                if p2p_uploads:
                    parts.append(f"{len(p2p_uploads)} P2P")
                if city_lookup_upload:
                    parts.append("1 city lookup")
                if parts:
                    st.success(f"Classified: {', '.join(parts)}")
            else:
                st.warning("File classifier not available - please install file_classifier.py")
                # Fallback: just use first file as QSI template
                if all_uploads:
                    qsi_template_upload = all_uploads[0]

    st.markdown("---")

    # ================================================================
    # P2P SEGMENTS  both paths need these
    # ================================================================
    st.markdown("### P2P Demand Segments")
    if st.session_state.data_source == 'precomputed':
        st.caption("Configure point-to-point demand. Connecting city data is read from the forecast file.")
    else:
        st.caption("Configure point-to-point demand segments. Connecting city demand comes from the MIDT files above.")

    # IATA Stimulation Quick Calculator
    with st.expander("IATA Stimulation Calculator (for new routes)", expanded=False):
        st.caption(
            "For new routes with no existing direct service, enter the Sabre indirect O&D demand "
            "between the two cities. The IATA stimulation formula estimates how much additional "
            "demand a new direct service would generate."
        )
        st.caption("Formula: Stimulation = -0.491 x ln(indirect_pax) + 6.8124  (Source: IATA)")
        ic1, ic2, ic3 = st.columns(3)
        with ic1:
            iata_indirect = st.number_input(
                "Indirect O&D Pax (annual, grown to start year)",
                min_value=0, value=0, step=1000, key="iata_indirect",
                help="Total indirect passengers between the two cities from Sabre, "
                     "grown to the service start year")
        with ic2:
            iata_capture = st.number_input(
                "P2P Capture Rate",
                min_value=0.0, max_value=1.0, value=0.60, step=0.05,
                key="iata_capture", format="%.2f",
                help="Typical: 55-65% for full-service carrier with no direct competition")
        with ic3:
            iata_growth = st.number_input(
                "Growth Rate (already applied)",
                min_value=0.0, max_value=0.50, value=0.0, step=0.01,
                key="iata_growth", format="%.2f",
                help="Set to 0 if indirect demand figure already includes growth to start year")

        if iata_indirect > 0:
            import math
            stim_factor = max(1.0, -0.491 * math.log(iata_indirect) + 6.8124)
            stimulated = iata_indirect * stim_factor
            forecast = stimulated * iata_capture
            st.info(
                f"Stimulation factor: **{stim_factor:.2f}** | "
                f"Stimulated demand: **{stimulated:,.0f}** | "
                f"Forecast P2P (at {iata_capture:.0%} capture): **{forecast:,.0f}** pax"
            )
            if st.button("Apply to P2P Segments"):
                st.session_state.p2p_segments = [{
                    'name': 'P2P (IATA stimulated)',
                    'base_demand': round(iata_indirect, 0),
                    'growth': iata_growth,
                    'stimulation': round(stim_factor, 2),
                    'capture_rate': iata_capture,
                    'subsegments': [],
                }]
                st.rerun()

    # Initialize with generic empty P2P segments - user must fill in their route-specific values
    if not st.session_state.p2p_segments:
        st.session_state.p2p_segments = [
            {'name': 'Segment 1', 'base_demand': 0, 'growth': 0.05,
             'stimulation': 1.0, 'capture_rate': 0.25, 'subsegments': []},
            {'name': 'Segment 2', 'base_demand': 0, 'growth': 0.05,
             'stimulation': 1.0, 'capture_rate': 0.25, 'subsegments': []},
        ]

    for i, seg in enumerate(st.session_state.p2p_segments):
        with st.expander(f"Segment {i+1}: {seg['name']}", expanded=False):
            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                seg['name'] = st.text_input("Name", value=seg['name'], key=f"sn_{i}")
            with c2:
                seg['base_demand'] = st.number_input("Base Demand", value=float(seg['base_demand']),
                                                     step=1000.0, key=f"sd_{i}", format="%.2f")
            with c3:
                seg['growth'] = st.number_input("Growth", value=float(seg['growth']),
                                                step=0.01, key=f"sg_{i}", format="%.2f")
            with c4:
                seg['stimulation'] = st.number_input("Stimulation", value=float(seg['stimulation']),
                                                     step=0.05, key=f"ss_{i}", format="%.2f")
            with c5:
                seg['capture_rate'] = st.number_input("Capture Rate", value=float(seg['capture_rate']),
                                                      step=0.05, key=f"sc_{i}", format="%.2f")
            if seg['subsegments']:
                st.markdown("**Subsegments:**")
                for j, sub in enumerate(seg['subsegments']):
                    sc1, sc2, sc3, sc4, sc5 = st.columns(5)
                    with sc1:
                        sub['name'] = st.text_input("Name", value=sub['name'], key=f"bn_{i}_{j}")
                    with sc2:
                        sub['base_demand'] = st.number_input("Base Demand", value=float(sub['base_demand']),
                                                             step=1000.0, key=f"bd_{i}_{j}", format="%.2f")
                    with sc3:
                        sub['growth'] = st.number_input("Growth", value=float(sub['growth']),
                                                        step=0.01, key=f"bg_{i}_{j}", format="%.2f")
                    with sc4:
                        sub['stimulation'] = st.number_input("Stimulation", value=float(sub['stimulation']),
                                                             step=0.05, key=f"bs_{i}_{j}", format="%.2f")
                    with sc5:
                        sub['capture_rate'] = st.number_input("Capture Rate", value=float(sub['capture_rate']),
                                                              step=0.05, key=f"bc_{i}_{j}", format="%.2f")

    btn1, btn2, _ = st.columns([1, 1, 4])
    with btn1:
        if st.button(" Add Segment"):
            st.session_state.p2p_segments.append({
                'name': f'Segment {len(st.session_state.p2p_segments)+1}',
                'base_demand': 0, 'growth': 0.05, 'stimulation': 1.0,
                'capture_rate': 0.10, 'subsegments': []
            })
            st.rerun()
    with btn2:
        if len(st.session_state.p2p_segments) > 1 and st.button(" Remove Last"):
            st.session_state.p2p_segments.pop()
            st.rerun()

    st.markdown("---")

    # ================================================================
    # READINESS CHECKS
    # ================================================================
    ready = True
    issues = []
    if not origin or len(origin) != 3:
        issues.append(" Origin airport required"); ready = False
    if not destination or len(destination) != 3:
        issues.append(" Destination airport required"); ready = False

    if st.session_state.data_source == 'precomputed':
        if not home_qsi_upload:
            issues.append(" Home Hub QSI file required"); ready = False
        if not dest_qsi_upload:
            issues.append(" Destination QSI file required"); ready = False
        if not forecast_upload:
            issues.append(" Forecast/Demand file required"); ready = False
    else:
        if not qsi_template_upload:
            issues.append(" QSI Template file required (with Leg sheets)"); ready = False
        if not home_cnx_uploads and not dest_cnx_uploads:
            issues.append(" No MIDT connecting demand files  connecting traffic will be zero")

    for iss in issues:
        st.warning(iss)

    # ---- Closed-loop mode toggle ----
    cl_col1, cl_col2 = st.columns([1, 3])
    with cl_col1:
        use_closed_loop = st.checkbox(
            "Closed-Loop Mode",
            value=True,
            help="When enabled, the pipeline runs iteratively: if the CRE detects "
                 "unreasonable results, it adjusts parameters using the Predictive "
                 "Calibration Engine and re-runs (max 3 iterations). This produces "
                 "forecasts with defensible assumptions, not just scaled outputs.",
            disabled=not (HAS_CRE and HAS_BRIDGE),
        )
    with cl_col2:
        if use_closed_loop:
            st.caption("Pipeline will auto-adjust QSI parameters if forecast "
                       "fails commercial reasonableness checks (max 3 re-runs)")
        elif not HAS_BRIDGE:
            st.caption("Closed-loop mode requires cre_pce_bridge module")

    # ================================================================
    # RUN PIPELINE
    # ================================================================
    if st.button("Run Pipeline", type="primary", disabled=not ready, use_container_width=True):
        with st.spinner("Running QSI pipeline  scoring itineraries across all hubs..."):
            try:
                log = []
                log.append(f"[{datetime.now():%H:%M:%S}] Pipeline started")
                log.append(f"  Route: {carrier_code} {origin}-{destination}")
                log.append(f"  Aircraft: {aircraft_type} ({seats}s), {frequency}x/wk")
                log.append(f"  Data source: {st.session_state.data_source}")

                # Build P2P config (shared by both paths)
                # If all segment base demands are zero, omit segments key
                # so the MIDT provider auto-detects from uploaded P2P files
                p2p_config = {'segments': [],
                              'default_base_year': sabre_base_year,
                              'default_forecast_year': p2p_forecast_year}
                has_nonzero_p2p = False
                for seg in st.session_state.p2p_segments:
                    seg_dict = {
                        'name': seg['name'], 'base_demand': seg['base_demand'],
                        'growth_rate': seg['growth'], 'seasonality': 1.0,
                        'stimulation': seg['stimulation'], 'capture_rate': seg['capture_rate'],
                    }
                    if seg['subsegments']:
                        seg_dict['subsegments'] = [{
                            'name': s['name'], 'base_demand': s['base_demand'],
                            'growth_rate': s['growth'], 'seasonality': 1.0,
                            'stimulation': s['stimulation'], 'capture_rate': s['capture_rate'],
                        } for s in seg['subsegments']]
                        if any(s['base_demand'] > 0 for s in seg['subsegments']):
                            has_nonzero_p2p = True
                    if seg['base_demand'] > 0:
                        has_nonzero_p2p = True
                    p2p_config['segments'].append(seg_dict)

                # If user left all P2P segments at zero, remove them so auto-detect works
                if not has_nonzero_p2p:
                    del p2p_config['segments']
                    log.append(f"  P2P: auto-detect from uploaded files (no manual values entered)")

                # Build RouteConfig
                cfg = RouteConfig()
                cfg.airline_name = carrier_name or carrier_code
                cfg.airline_code = carrier_code
                cfg.home_airport_code = origin
                cfg.dest_airport_code = destination
                o_info = lookup_airport(origin)
                d_info = lookup_airport(destination)
                cfg.home_city_code = o_info[0] if o_info else origin
                cfg.dest_city_code = d_info[0] if d_info else destination
                cfg.frequency = frequency
                cfg.aircraft_type = aircraft_type
                cfg.seats = seats
                cfg.qsi_ceiling = qsi_ceiling
                cfg.qsi_adjustment = 1.0
                cfg.online_coeff = online_coeff
                cfg.alliance_coeff = alliance_coeff
                cfg.interline_coeff = interline_coeff
                cfg.et_decay_factor = 0.8
                cfg.et_decay_interval = 0.1

                # ==== PATH A: Pre-computed ====
                if st.session_state.data_source == 'precomputed':
                    home_qsi_path = save_uploaded(home_qsi_upload)
                    dest_qsi_path = save_uploaded(dest_qsi_upload)
                    forecast_path = save_uploaded(forecast_upload)
                    log.append(f"  Files: {home_qsi_upload.name}, {dest_qsi_upload.name}, {forecast_upload.name}")

                    cfg.schedule_provider = ExcelScheduleProvider(
                        qsi1_file=home_qsi_path, qsi2_file=dest_qsi_path,
                    )
                    cfg.demand_provider = ExcelDemandProvider(
                        forecast_file=forecast_path, p2p_config=p2p_config,
                        home_growth=home_growth, dest_growth=dest_growth,
                    )

                # ==== PATH B: New Route ====
                else:
                    qsi_template_path = save_uploaded(qsi_template_upload)
                    log.append(f"  QSI Template: {qsi_template_upload.name}")

                    # MCT files
                    mct_file_map = {}
                    if mct_uploads:
                        for mf in mct_uploads:
                            mct_path = save_uploaded(mf)
                            # Try to extract airport code from filename
                            name_upper = mf.name.upper()
                            for apt_code in KNOWN_AIRPORTS:
                                if apt_code in name_upper:
                                    mct_file_map[apt_code] = mct_path
                                    break
                            else:
                                # Use generic key
                                mct_file_map[mf.name] = mct_path
                        log.append(f"  MCT files: {list(mct_file_map.keys())}")

                    # City lookup
                    city_lookup_path = None
                    if city_lookup_upload:
                        city_lookup_path = save_uploaded(city_lookup_upload)

                    # Build SingleExtractOAGProvider
                    cfg.schedule_provider = SingleExtractOAGProvider(
                        qsi_file=qsi_template_path,
                        origin_airport=origin,
                        dest_airport=destination,
                        proposed_carrier=carrier_code or 'XX',
                        use_city_codes=True,
                        city_lookup_file=city_lookup_path,
                        mct_files=mct_file_map,
                    )
                    log.append(f"  Schedule: SingleExtractOAGProvider")

                    # Save MIDT files
                    home_cnx_paths = save_multiple_uploads(home_cnx_uploads) if home_cnx_uploads else []
                    dest_cnx_paths = save_multiple_uploads(dest_cnx_uploads) if dest_cnx_uploads else []
                    p2p_paths = save_multiple_uploads(p2p_uploads) if p2p_uploads else []
                    forecast_paths = save_multiple_uploads(forecast_uploads) if forecast_uploads else []

                    if home_cnx_paths or dest_cnx_paths:
                        log.append(f"  MIDT home: {len(home_cnx_paths)} files, dest: {len(dest_cnx_paths)} files")

                    # P2P: check if forecast workbook can provide P2P
                    if not p2p_paths and not has_nonzero_p2p:
                        if forecast_paths:
                            log.append(f"  P2P: will attempt auto-extract from forecast workbook")
                        else:
                            log.append(f"  WARNING: No P2P demand - no P2P file uploaded and no manual values. "
                                       f"Use the IATA Stimulation Calculator or enter P2P segments manually.")

                    # Build MIDTDemandProvider
                    # Inject PCE capture rate + stimulation if available and user hasn't manually set P2P
                    if not has_nonzero_p2p:
                        pce_sug = st.session_state.get('calib_suggestion')
                        if pce_sug and hasattr(pce_sug, 'blended_capture') and pce_sug.blended_capture > 0:
                            p2p_config['default_capture'] = pce_sug.blended_capture
                            p2p_config['_capture_source'] = 'pce'
                            log.append(f"  P2P capture rate: {pce_sug.blended_capture:.0%} (from PCE)")
                        else:
                            # Use market maturity-aware defaults instead of flat 25%
                            _maturity = st.session_state.get('market_maturity', 'New Route')
                            _carrier_type = st.session_state.get('carrier_type', 'Full Service')
                            _default_captures = {
                                'New Route': 0.60,          # New monopoly route - high capture
                                'Existing Underserved': 0.40,
                                'Existing Competitive': 0.25,
                                'Mature': 0.20,
                            }
                            default_cr = _default_captures.get(_maturity, 0.40)
                            p2p_config['default_capture'] = default_cr
                            p2p_config['_capture_source'] = 'default'
                            log.append(f"  P2P capture rate: {default_cr:.0%} (default for {_maturity})")

                        if pce_sug and hasattr(pce_sug, 'blended_stimulation') and pce_sug.blended_stimulation > 1.0:
                            p2p_config['default_stimulation'] = pce_sug.blended_stimulation
                            log.append(f"  P2P stimulation: {pce_sug.blended_stimulation:.2f} (from PCE)")

                        _gyrs = max(1, p2p_forecast_year - sabre_base_year)
                        log.append(f"  P2P growth: {sabre_base_year} → {p2p_forecast_year} ({_gyrs} years compound at {p2p_config.get('default_growth', 0.09):.0%})")

                    cfg.demand_provider = MIDTDemandProvider(
                        home_cnx_files=home_cnx_paths,
                        dest_cnx_files=dest_cnx_paths,
                        p2p_files=p2p_paths,
                        forecast_files=forecast_paths,
                        p2p_config=p2p_config,
                        home_growth=home_growth,
                        dest_growth=dest_growth,
                        catchment_share_home=catchment_share_home,
                        catchment_share_dest=catchment_share_dest,
                        city_lookup_file=city_lookup_path,
                        min_demand_threshold=min_demand_threshold,
                    )
                    log.append(f"  Demand: MIDTDemandProvider")

                # Common config
                cfg.target_total = 0
                cfg.target_p2p = 0
                cfg.target_cnx_home = 0
                cfg.target_cnx_dest = 0
                cfg.target_load_factor = 0.0

                log.append(f"[{datetime.now():%H:%M:%S}] Config: {cfg.summary()}")
                st.session_state.last_config_summary = cfg.summary()

                # Capture stdout
                old_stdout = sys.stdout
                sys.stdout = captured = io.StringIO()
                try:
                    output_path = tempfile.mktemp(suffix='.xlsx')

                    if use_closed_loop and HAS_BRIDGE and HAS_CRE:
                        # ---- CLOSED-LOOP MODE ----
                        log.append(f"[{datetime.now():%H:%M:%S}] Mode: Informed Closed-Loop")

                        # Get PCE suggestion if available
                        pce_sug = st.session_state.get('calib_suggestion')

                        def _pipeline_fn(c, op):
                            return run_pipeline(c, op)

                        cl_result = run_informed_closed_loop(
                            config=cfg,
                            pipeline_fn=_pipeline_fn,
                            output_path_fn=lambda: tempfile.mktemp(suffix='.xlsx'),
                            log_fn=lambda msg: log.append(f"  CL: {msg}"),
                            pce_suggestion=pce_sug,
                        )
                        st.session_state.closed_loop_result = cl_result

                        # Use final results
                        results = cl_result.final_results or cl_result.initial_results
                        # Write final output workbook
                        final_output = tempfile.mktemp(suffix='.xlsx')
                        _ = run_pipeline(cfg, final_output)
                        with open(final_output, 'rb') as f:
                            st.session_state.output_bytes = f.read()

                        # Store CRE result from final iteration
                        st.session_state.cre_result = cl_result.final_cre

                    else:
                        # ---- SINGLE-RUN MODE ----
                        results = run_pipeline(cfg, output_path)
                        with open(output_path, 'rb') as f:
                            st.session_state.output_bytes = f.read()
                        st.session_state.closed_loop_result = None

                finally:
                    sys.stdout = old_stdout
                    for line in captured.getvalue().strip().split('\n'):
                        if line.strip():
                            log.append(f"  {line}")

                log.append(f"[{datetime.now():%H:%M:%S}] Complete: {results['grand_total']:,} pax, "
                           f"{results['load_factor']:.1%} LF")

                # Log provider metadata
                if hasattr(cfg.schedule_provider, 'get_metadata'):
                    meta = cfg.schedule_provider.get_metadata()
                    ptype = meta.get('provider_type', 'unknown')
                    if ptype == 'SingleExtractOAGProvider':
                        stats = meta.get('stats', {})
                        log.append(f"  Schedule stats: {stats.get('qsi1_itineraries', '?')} QSI1 + "
                                   f"{stats.get('qsi2_itineraries', '?')} QSI2 itineraries, "
                                   f"{stats.get('qsi1_hubs', '?')} hubs detected")

                if hasattr(cfg.demand_provider, 'get_metadata'):
                    meta = cfg.demand_provider.get_metadata()
                    ptype = meta.get('provider_type', 'unknown')
                    if ptype == 'MIDTDemandProvider':
                        stats = meta.get('stats', {})
                        home_cities = stats.get('home_output_cities', stats.get('home_final_cities', 0))
                        dest_cities = stats.get('dest_output_cities', stats.get('dest_final_cities', 0))
                        home_pivot = stats.get('home_pivot_cities', 0)
                        dest_pivot = stats.get('dest_pivot_cities', 0)
                        log.append(f"  MIDT stats: home={stats.get('home_raw_records', 0)} records / "
                                   f"{home_cities} cities"
                                   f"{f' ({home_pivot} pivot)' if home_pivot else ''}, "
                                   f"dest={stats.get('dest_raw_records', 0)} records / "
                                   f"{dest_cities} cities"
                                   f"{f' ({dest_pivot} pivot)' if dest_pivot else ''}")
                        # Log any parse errors
                        for k in ('home_error', 'dest_error', 'home_parse_error',
                                  'dest_parse_error', 'home_pivot_error', 'dest_pivot_error'):
                            if k in stats:
                                log.append(f"  MIDT WARNING: {k}: {stats[k]}")

                # Log closed-loop convergence info
                cl_res = st.session_state.get('closed_loop_result')
                if cl_res is not None:
                    log.append(f"[{datetime.now():%H:%M:%S}] Closed-loop: "
                               f"{cl_res.reruns_performed} re-runs, "
                               f"converged={cl_res.converged}")
                    if cl_res.config_changes:
                        for param, vals in cl_res.config_changes.items():
                            log.append(f"  Config change: {param} "
                                       f"{vals.get('original', '?')} -> {vals.get('final', '?')}")

                st.session_state.results = results
                st.session_state.last_config = cfg
                st.session_state.run_log = log
                st.success(f"Pipeline complete: **{results['grand_total']:,} passengers**, "
                           f"**{results['load_factor']:.1%}** load factor")

                # ---- CRE: post-pipeline (single-run mode only) ----
                if st.session_state.closed_loop_result is None:
                    if HAS_CRE:
                        try:
                            cre_result = run_cre(cfg, results)
                            st.session_state.cre_result = cre_result
                            log.append(f"[{datetime.now():%H:%M:%S}] CRE: confidence {cre_result.confidence_score}/100, "
                                       f"adjusted={cre_result.adjusted}")
                            if cre_result.adjusted:
                                st.info(f"CRE adjusted forecast: {cre_result.adjusted_total:,.0f} pax "
                                        f"({cre_result.adjusted_load_factor:.1%} LF) -- "
                                        f"see Results tab for details")
                        except Exception as cre_err:
                            log.append(f"[{datetime.now():%H:%M:%S}] CRE warning: {str(cre_err)}")
                            st.session_state.cre_result = None
                    else:
                        st.session_state.cre_result = None

            except Exception as e:
                log.append(f"[{datetime.now():%H:%M:%S}] ERROR: {str(e)}")
                st.session_state.run_log = log
                st.error(f"Pipeline failed: {str(e)}")
                with st.expander("Full traceback"):
                    st.code(traceback.format_exc())

    if st.session_state.run_log:
        with st.expander(" Pipeline Log", expanded=False):
            st.code('\n'.join(st.session_state.run_log), language='text')


# ============================================================================
# TAB 2: RESULTS
# ============================================================================

with tab2:
    results = st.session_state.results
    if results is None:
        st.info("Run the pipeline from the 'Data Upload & Run' tab to see results here.")
    else:
        path_badge = ("precomputed" if st.session_state.data_source == 'precomputed'
                       else "newroute")
        path_label = ("PRE-COMPUTED" if st.session_state.data_source == 'precomputed'
                       else "NEW ROUTE")
        st.markdown(f'### Results: {st.session_state.last_config_summary} '
                    f'<span class="provider-badge {path_badge}">{path_label}</span>',
                    unsafe_allow_html=True)

        m1, m2, m3, m4, m5 = st.columns(5)
        with m1:
            st.markdown(metric_card("Grand Total", fmt(results['grand_total']), "Annual pax"),
                        unsafe_allow_html=True)
        with m2:
            st.markdown(metric_card("Load Factor", f"{results['load_factor']:.1%}"),
                        unsafe_allow_html=True)
        with m3:
            pct = f"{results['p2p_total']/results['grand_total']:.0%}" if results['grand_total'] else ""
            st.markdown(metric_card("P2P", fmt(results['p2p_total']), pct),
                        unsafe_allow_html=True)
        with m4:
            pct = f"{results['home_total']/results['grand_total']:.0%}" if results['grand_total'] else ""
            st.markdown(metric_card("Cnx Home", fmt(results['home_total']), pct),
                        unsafe_allow_html=True)
        with m5:
            pct = f"{results['dest_total']/results['grand_total']:.0%}" if results['grand_total'] else ""
            st.markdown(metric_card("Cnx Dest", fmt(results['dest_total']), pct),
                        unsafe_allow_html=True)

        st.markdown("---")

        # PTEW
        denom = frequency * 52 * 2 if frequency > 0 else 1
        ptew = results['grand_total'] / denom
        pt1, pt2, pt3 = st.columns(3)
        with pt1:
            st.markdown(metric_card("PTEW Total", f"{ptew:.1f}", "Pax per trip each way"),
                        unsafe_allow_html=True)
        with pt2:
            st.markdown(metric_card("PTEW P2P", f"{results['p2p_total']/denom:.1f}"),
                        unsafe_allow_html=True)
        with pt3:
            cnx = results['home_total'] + results['dest_total']
            st.markdown(metric_card("PTEW Connecting", f"{cnx/denom:.1f}"),
                        unsafe_allow_html=True)

        st.markdown("---")

        # ---- CRE: Commercial Reasonableness Assessment ----
        cre_result = st.session_state.get('cre_result')
        if cre_result is not None and HAS_CRE:
            st.markdown("#### Commercial Reasonableness Assessment")

            # Confidence badge
            score = cre_result.confidence_score
            if score >= 80:
                badge_color, badge_label = "#22c55e", "HIGH"
            elif score >= 60:
                badge_color, badge_label = "#f59e0b", "MODERATE"
            elif score >= 40:
                badge_color, badge_label = "#ef4444", "LOW"
            else:
                badge_color, badge_label = "#991b1b", "VERY LOW"

            st.markdown(
                f'<div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">'
                f'<span style="background:{badge_color};color:white;padding:4px 12px;'
                f'border-radius:6px;font-weight:700;font-size:0.85rem;">'
                f'{badge_label} ({score}/100)</span>'
                f'<span style="color:#6b7c93;font-size:0.85rem;">'
                f'{"Forecast passed all commercial checks" if not cre_result.adjusted else "Forecast adjusted for commercial reasonableness"}'
                f'</span></div>', unsafe_allow_html=True)

            # If adjusted: show raw vs adjusted side by side
            if cre_result.adjusted:
                ac1, ac2 = st.columns(2)
                with ac1:
                    st.markdown("**Raw Pipeline Output**")
                    st.markdown(f"- Total: **{cre_result.raw_total:,.0f}** pax")
                    st.markdown(f"- Load Factor: **{cre_result.raw_load_factor:.1%}**")
                with ac2:
                    st.markdown("**CRE Adjusted Output**")
                    st.markdown(f"- Total: **{cre_result.adjusted_total:,.0f}** pax")
                    st.markdown(f"- Load Factor: **{cre_result.adjusted_load_factor:.1%}**")
                    delta_pct = ((cre_result.adjusted_total - cre_result.raw_total) / cre_result.raw_total * 100) if cre_result.raw_total else 0.0
                    st.caption(f"Delta: {delta_pct:+.1f}%")

            # Checks table
            if cre_result.checks:
                with st.expander(f"Reasonableness Checks ({len(cre_result.checks)})", expanded=not cre_result.adjusted):
                    check_rows = []
                    for c in cre_result.checks:
                        status_icon = {"pass": "OK", "warning": "!!", "fail": "XX", "critical": "!!!"}.get(
                            c.status.value, "?"
                        )
                        actual_str = f"{c.actual_value:.2f}" if isinstance(c.actual_value, (int, float)) else str(c.actual_value)
                        range_str = (f"{c.acceptable_range[0]:.2f} - {c.acceptable_range[1]:.2f}"
                                     if c.acceptable_range and c.acceptable_range != (0, 0) else "-")
                        check_rows.append({
                            'Status': f"{status_icon} {c.status.value.upper()}",
                            'Check': c.name,
                            'Value': actual_str,
                            'Range': range_str,
                            'Finding': c.message,
                        })
                    st.dataframe(check_rows, use_container_width=True, hide_index=True)

            # Adjustment trail
            if cre_result.adjustments:
                with st.expander(f"Adjustment Trail ({len(cre_result.adjustments)} steps)", expanded=False):
                    adj_rows = []
                    for a in cre_result.adjustments:
                        adj_rows.append({
                            'Iter': a.iteration,
                            'Type': a.adjustment_type.value.replace('_', ' ').title(),
                            'Parameter': a.parameter,
                            'Before': f"{a.old_value:,.0f}",
                            'After': f"{a.new_value:,.0f}",
                            'LF Before': f"{a.lf_before:.1%}",
                            'LF After': f"{a.lf_after:.1%}",
                            'Reason': a.reason,
                        })
                    st.dataframe(adj_rows, use_container_width=True, hide_index=True)

            # Frequency recommendation
            if cre_result.frequency_rec:
                fr = cre_result.frequency_rec
                with st.expander("Frequency Optimisation", expanded=False):
                    fc1, fc2, fc3 = st.columns(3)
                    with fc1:
                        st.markdown(metric_card("Min Viable Freq", f"{fr.min_viable_frequency}x/wk",
                                                "at ~65% LF"), unsafe_allow_html=True)
                    with fc2:
                        st.markdown(metric_card("Recommended Freq", f"{fr.recommended_frequency}x/wk",
                                                f"at {fr.recommended_lf:.0%} LF"), unsafe_allow_html=True)
                    with fc3:
                        st.markdown(metric_card("Max Supportable", f"{fr.max_supportable_frequency}x/wk",
                                                "at ~80% LF"), unsafe_allow_html=True)
                    if fr.reasoning:
                        st.caption(fr.reasoning)

            # Analyst review flag
            if cre_result.needs_analyst_review:
                reasons_text = " | ".join(cre_result.review_reasons[:3]) if cre_result.review_reasons else ""
                st.warning(f"Analyst review recommended: {reasons_text}")

            # Summary lines
            if cre_result.summary:
                with st.expander("CRE Summary", expanded=False):
                    for line in cre_result.summary:
                        st.markdown(f"- {line}")

            st.markdown("---")

        # ---- Closed-Loop Convergence History ----
        cl_result = st.session_state.get('closed_loop_result')
        if cl_result is not None and HAS_BRIDGE:
            st.markdown("#### Closed-Loop Parameter Convergence")

            # Status banner
            if cl_result.converged:
                st.success(f"Converged after {cl_result.reruns_performed} re-run(s). "
                           f"{cl_result.convergence_note}")
            elif cl_result.error:
                st.error(f"Error: {cl_result.error}")
            else:
                st.warning(f"Did not converge: {cl_result.convergence_note}")

            # PCE context
            if cl_result.pce_suggestion_available:
                st.caption(f"PCE provided parameter guidance (confidence: {cl_result.pce_confidence}, "
                           f"{cl_result.comparable_cases_count} comparable cases)")
            else:
                st.caption("No PCE suggestion available -- used proportional adjustments")

            # Config changes summary
            if cl_result.config_changes:
                cc_rows = []
                for param, vals in cl_result.config_changes.items():
                    cc_rows.append({
                        'Parameter': param,
                        'Original': f"{vals.get('original', 0):.3f}",
                        'Final': f"{vals.get('final', 0):.3f}",
                        'Delta': f"{(vals.get('final', 0) - vals.get('original', 0)):+.3f}",
                    })
                st.dataframe(cc_rows, use_container_width=True, hide_index=True)

            # Re-run history detail
            if cl_result.rerun_history:
                with st.expander(f"Re-run History ({len(cl_result.rerun_history)} iterations)",
                                 expanded=False):
                    rr_rows = []
                    for rr in cl_result.rerun_history:
                        rr_rows.append({
                            'Iter': rr.iteration,
                            'Category': rr.category.replace('_', ' ').title(),
                            'Parameter': rr.parameter_changed,
                            'Old': f"{rr.old_value:.3f}",
                            'New': f"{rr.new_value:.3f}",
                            'PCE Range': f"{rr.pce_range[0]:.3f}-{rr.pce_range[1]:.3f}",
                            'PCE Conf': rr.pce_confidence,
                            'Result Pax': f"{rr.result_total:,.0f}",
                            'Result LF': f"{rr.result_lf:.1%}",
                            'CRE Score': f"{rr.cre_confidence}/100",
                        })
                    st.dataframe(rr_rows, use_container_width=True, hide_index=True)

            # Initial vs final comparison
            if cl_result.reruns_performed > 0 and cl_result.initial_results and cl_result.final_results:
                with st.expander("Initial vs Final Comparison", expanded=False):
                    comp_c1, comp_c2 = st.columns(2)
                    with comp_c1:
                        st.markdown("**Initial (before closed-loop)**")
                        ini = cl_result.initial_results
                        st.markdown(f"- Total: {ini.get('grand_total', 0):,.0f}")
                        st.markdown(f"- LF: {ini.get('load_factor', 0):.1%}")
                        if cl_result.initial_cre:
                            st.markdown(f"- CRE: {cl_result.initial_cre.confidence_score}/100")
                    with comp_c2:
                        st.markdown("**Final (after closed-loop)**")
                        fin = cl_result.final_results
                        st.markdown(f"- Total: {fin.get('grand_total', 0):,.0f}")
                        st.markdown(f"- LF: {fin.get('load_factor', 0):.1%}")
                        if cl_result.final_cre:
                            st.markdown(f"- CRE: {cl_result.final_cre.confidence_score}/100")

            st.markdown("---")

        # P2P detail
        st.markdown("#### P2P Breakdown")
        if 'p2p_detail' in results:
            rows = []
            for seg in results['p2p_detail']:
                rows.append({
                    'Segment': seg.get('name', ''),
                    'Base Demand': f"{seg.get('base_demand', 0):,.0f}",
                    'Growth': f"{seg.get('growth', 0):.0%}",
                    'Stimulation': f"{seg.get('stimulation', 1.0):.2f}",
                    'Capture': f"{seg.get('capture_rate', 0):.0%}",
                    'Forecast': f"{seg.get('forecast', 0):,.0f}",
                })
            st.dataframe(rows, use_container_width=True, hide_index=True)

        # Connecting cities
        for label, key in [("Top Connecting  Home Hub", 'home_results'),
                           ("Top Connecting  Destination", 'dest_results')]:
            st.markdown(f"#### {label}")
            if key in results and results[key]:
                data = sorted(results[key], key=lambda x: -x.get('forecast', 0))[:25]
                rows = [{
                    'City': r.get('city', ''), 'Name': r.get('name', ''),
                    'Base Demand': f"{r.get('base_demand', 0):,.0f}",
                    'QSI Capture': f"{r.get('qsi_capture', 0):.4f}" if r.get('qsi_capture') else "",
                    'Expert QSI': f"{r.get('original_qsi', 0):.4f}" if r.get('original_qsi') else "",
                    'Cal Factor': f"{r.get('calibration_factor', 0):.3f}" if r.get('calibration_factor') else "",
                    'Forecast': f"{r.get('forecast', 0):,.0f}",
                } for r in data]
                st.dataframe(rows, use_container_width=True, hide_index=True)
            else:
                st.caption("No data available.")

        st.markdown("---")
        if st.session_state.output_bytes:
            fn = f"QSI_Forecast_{origin}_{destination}_{carrier_code}_{datetime.now():%Y%m%d}.xlsx"
            st.download_button(" Download Output Workbook", data=st.session_state.output_bytes,
                               file_name=fn, type="primary", use_container_width=True,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================================
# TAB 3: MONTHLY PROFILE (Seasonality)
# ============================================================================

with tab3:
    st.markdown("### Monthly Seasonality Profile")
    results = st.session_state.results
    if results is None:
        st.info("Run the pipeline from the 'Data Upload & Run' tab to see monthly breakdown here.")
    else:
        total_pax = results['grand_total']
        annual_capacity = frequency * 52 * 2 * seats

        # Get selected profile
        profile = PROFILE_LIBRARY.get(season_profile_key, PROFILE_LIBRARY['flat'])

        # Use the seasonality engine's distribute_annual
        mf = distribute_annual(
            annual_pax=total_pax,
            annual_capacity=annual_capacity,
            profile=profile,
            frequency=frequency,
            seats=seats,
        )

        st.markdown(f"**Profile:** {profile_display.get(season_profile_key, season_profile_key)} "
                    f"| **Annual:** {total_pax:,} pax | {mf.annual_total / annual_capacity:.1%} LF"
                    if annual_capacity > 0 else "")
        st.markdown("---")

        # Monthly data table
        rows = []
        for i, m in enumerate(MONTHS):
            pax = round(mf.monthly_pax[i])
            cap = round(mf.monthly_capacity[i])
            lf = mf.monthly_load_factor[i]
            idx = profile.indices[i]
            spill = max(0, pax - cap)
            rows.append({
                'Month': m,
                'Index': f"{idx:.3f}",
                'Passengers': f"{pax:,}",
                'Capacity': f"{cap:,}",
                'Load Factor': f"{lf:.1%}",
                'Spill': f"{spill:,}" if spill > 0 else "-",
            })
        st.dataframe(rows, use_container_width=True, hide_index=True)

        # Summary metrics
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            pk_idx = mf.monthly_pax.index(max(mf.monthly_pax))
            st.markdown(metric_card("Peak Month", MONTHS[pk_idx],
                        f"{round(mf.monthly_pax[pk_idx]):,} pax"), unsafe_allow_html=True)
        with c2:
            tr_idx = mf.monthly_pax.index(min(mf.monthly_pax))
            st.markdown(metric_card("Trough Month", MONTHS[tr_idx],
                        f"{round(mf.monthly_pax[tr_idx]):,} pax"), unsafe_allow_html=True)
        with c3:
            pk_pax = max(mf.monthly_pax)
            tr_pax = min(mf.monthly_pax)
            ratio = pk_pax / tr_pax if tr_pax > 0 else 0
            st.markdown(metric_card("Peak/Trough", f"{ratio:.2f}x"), unsafe_allow_html=True)
        with c4:
            cv = profile.coefficient_of_variation
            st.markdown(metric_card("CV", f"{cv:.3f}", "Coefficient of variation"),
                        unsafe_allow_html=True)

        # Spill warning
        if mf.spill_months:
            total_spill = sum(max(0, round(mf.monthly_pax[i]) - round(mf.monthly_capacity[i]))
                             for i in range(12))
            st.warning(f"Capacity constrained in {', '.join(mf.spill_months)}. "
                       f"Total annual spill: {total_spill:,} pax. "
                       f"Consider frequency increase or larger aircraft for peak months.")

        st.markdown("---")

        # Monthly pax chart
        st.markdown("#### Monthly Passenger Distribution")
        chart_df = pd.DataFrame({
            'Month': MONTHS,
            'Passengers': [round(mf.monthly_pax[i]) for i in range(12)],
            'Capacity': [round(mf.monthly_capacity[i]) for i in range(12)],
        }).set_index('Month')
        st.bar_chart(chart_df, color=['#1a73e8', '#e8eaed'])

        # Monthly LF chart
        st.markdown("#### Monthly Load Factor")
        lf_df = pd.DataFrame({
            'Month': MONTHS,
            'Load Factor (%)': [mf.monthly_load_factor[i] * 100 for i in range(12)],
        }).set_index('Month')
        st.line_chart(lf_df)

        # Quarterly summary
        st.markdown("#### Quarterly Summary")
        q_data = mf.quarterly_summary()
        q_rows = []
        for qname in ['Q1', 'Q2', 'Q3', 'Q4']:
            q = q_data[qname]
            q_share = q['pax'] / total_pax if total_pax > 0 else 0
            q_rows.append({
                'Quarter': qname,
                'Passengers': f"{q['pax']:,}",
                'Capacity': f"{q['capacity']:,}",
                'Load Factor': f"{q['load_factor']:.1%}",
                'Share of Annual': f"{q_share:.1%}",
            })
        st.dataframe(q_rows, use_container_width=True, hide_index=True)

        # Store for use by revenue tab
        st.session_state.seasonality_result = {
            'forecast': mf,
            'profile_key': season_profile_key,
        }


# ============================================================================
# TAB 4: REVENUE FORECAST
# ============================================================================

with tab4:
    st.markdown("### Revenue Forecast")
    results = st.session_state.results
    if results is None:
        st.info("Run the pipeline from the 'Data Upload & Run' tab to see revenue forecast here.")
    else:
        total_pax = results['grand_total']
        p2p_pax = results['p2p_total']
        cnx_home_pax = results['home_total']
        cnx_dest_pax = results['dest_total']

        # --- Compute Revenue ---
        effective_fare = avg_ow_fare * fare_weight

        # P2P revenue (2 one-way segments per roundtrip passenger)
        p2p_revenue = p2p_pax * effective_fare * 2
        cnx_home_revenue = cnx_home_pax * effective_fare * 2
        cnx_dest_revenue = cnx_dest_pax * effective_fare * 2
        total_pax_revenue = p2p_revenue + cnx_home_revenue + cnx_dest_revenue

        # Cargo revenue
        annual_flights = frequency * 52 * 2  # both directions
        cargo_revenue = annual_flights * cargo_capacity_kg * cargo_lf * cargo_yield

        # Ancillary revenue
        ancillary_revenue = total_pax * ancillary_per_pax

        # Total
        grand_revenue = total_pax_revenue + cargo_revenue + ancillary_revenue

        # --- Key Metrics ---
        annual_capacity = frequency * 52 * 2 * seats
        dist = compute_distance(origin, destination) if (origin and destination) else None
        dist_km = dist * 1.852 if dist else 0  # nm to km

        ask = annual_capacity * dist_km if dist_km > 0 else 0
        rpk = total_pax * dist_km if dist_km > 0 else 0

        yield_val = total_pax_revenue / rpk if rpk > 0 else 0  # $/RPK
        prask = total_pax_revenue / ask if ask > 0 else 0  # Pax rev / ASK
        trask = grand_revenue / ask if ask > 0 else 0  # Total rev / ASK

        avg_fare_actual = total_pax_revenue / (total_pax * 2) if total_pax > 0 else 0

        # --- Display ---
        st.markdown("---")

        # Revenue summary cards
        r1, r2, r3, r4 = st.columns(4)
        with r1:
            st.markdown(metric_card("Pax Revenue", f"${total_pax_revenue/1e6:.1f}M"),
                        unsafe_allow_html=True)
        with r2:
            st.markdown(metric_card("Cargo Revenue", f"${cargo_revenue/1e6:.1f}M"),
                        unsafe_allow_html=True)
        with r3:
            st.markdown(metric_card("Ancillary", f"${ancillary_revenue/1e6:.1f}M"),
                        unsafe_allow_html=True)
        with r4:
            st.markdown(metric_card("Total Revenue", f"${grand_revenue/1e6:.1f}M"),
                        unsafe_allow_html=True)

        st.markdown("---")

        # Revenue breakdown table
        st.markdown("#### Revenue by Segment")
        rev_rows = [
            {'Segment': 'Point-to-Point', 'Passengers': f"{p2p_pax:,}",
             'Avg OW Fare': f"${effective_fare:,.0f}", 'Revenue': f"${p2p_revenue:,.0f}",
             'Share': f"{p2p_revenue/grand_revenue:.1%}" if grand_revenue > 0 else "-"},
            {'Segment': f'Connecting @ {origin}', 'Passengers': f"{cnx_home_pax:,}",
             'Avg OW Fare': f"${effective_fare:,.0f}", 'Revenue': f"${cnx_home_revenue:,.0f}",
             'Share': f"{cnx_home_revenue/grand_revenue:.1%}" if grand_revenue > 0 else "-"},
            {'Segment': f'Connecting @ {destination}', 'Passengers': f"{cnx_dest_pax:,}",
             'Avg OW Fare': f"${effective_fare:,.0f}", 'Revenue': f"${cnx_dest_revenue:,.0f}",
             'Share': f"{cnx_dest_revenue/grand_revenue:.1%}" if grand_revenue > 0 else "-"},
            {'Segment': 'Cargo', 'Passengers': '-',
             'Avg OW Fare': f"${cargo_yield:.2f}/kg", 'Revenue': f"${cargo_revenue:,.0f}",
             'Share': f"{cargo_revenue/grand_revenue:.1%}" if grand_revenue > 0 else "-"},
            {'Segment': 'Ancillary', 'Passengers': '-',
             'Avg OW Fare': f"${ancillary_per_pax:.0f}/pax", 'Revenue': f"${ancillary_revenue:,.0f}",
             'Share': f"{ancillary_revenue/grand_revenue:.1%}" if grand_revenue > 0 else "-"},
            {'Segment': 'TOTAL', 'Passengers': f"{total_pax:,}",
             'Avg OW Fare': f"${avg_fare_actual:,.0f}", 'Revenue': f"${grand_revenue:,.0f}",
             'Share': '100.0%'},
        ]
        st.dataframe(rev_rows, use_container_width=True, hide_index=True)

        # Unit revenue metrics
        st.markdown("---")
        st.markdown("#### Unit Revenue Metrics")
        u1, u2, u3, u4 = st.columns(4)
        with u1:
            st.markdown(metric_card("Avg OW Fare", f"${avg_fare_actual:,.0f}",
                        f"Raw ${avg_ow_fare} x {fare_weight:.0%} weight"),
                        unsafe_allow_html=True)
        with u2:
            st.markdown(metric_card("Yield", f"{yield_val:.4f}" if yield_val > 0 else "N/A",
                        "$/RPK"), unsafe_allow_html=True)
        with u3:
            st.markdown(metric_card("PRASK", f"{prask:.4f}" if prask > 0 else "N/A",
                        "Pax rev / ASK"), unsafe_allow_html=True)
        with u4:
            st.markdown(metric_card("TRASK", f"{trask:.4f}" if trask > 0 else "N/A",
                        "Total rev / ASK"), unsafe_allow_html=True)

        # Cargo detail
        st.markdown("---")
        st.markdown("#### Cargo Assumptions")
        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            st.markdown(metric_card("Annual Flights", f"{annual_flights:,}",
                        f"{frequency}x weekly x 52 wks x 2 dirs"), unsafe_allow_html=True)
        with cc2:
            annual_cargo_kg = annual_flights * cargo_capacity_kg * cargo_lf
            st.markdown(metric_card("Annual Cargo", f"{annual_cargo_kg/1000:,.0f} tonnes",
                        f"{cargo_capacity_kg:,} kg x {cargo_lf:.0%} LF"), unsafe_allow_html=True)
        with cc3:
            st.markdown(metric_card("Cargo Yield", f"${cargo_yield:.2f}/kg",
                        f"Revenue: ${cargo_revenue:,.0f}"), unsafe_allow_html=True)

        # --- Multi-Year Projection ---
        st.markdown("---")
        st.markdown("#### 3-Year Revenue Projection")

        yr1_growth = st.number_input("Year 2 pax growth", min_value=-0.10, max_value=0.30,
                                      value=0.05, step=0.01, format="%.2f", key="rev_yr2g")
        yr2_growth = st.number_input("Year 3 pax growth", min_value=-0.10, max_value=0.30,
                                      value=0.04, step=0.01, format="%.2f", key="rev_yr3g")
        fare_growth = st.number_input("Annual fare growth", min_value=-0.05, max_value=0.15,
                                       value=0.02, step=0.01, format="%.2f", key="rev_fareg")

        yr_pax = [total_pax, int(total_pax * (1 + yr1_growth)),
                  int(total_pax * (1 + yr1_growth) * (1 + yr2_growth))]
        yr_fare = [effective_fare, effective_fare * (1 + fare_growth),
                   effective_fare * (1 + fare_growth)**2]
        yr_cargo_y = [cargo_yield, cargo_yield * 1.03, cargo_yield * 1.03**2]

        proj_rows = []
        for y in range(3):
            yr_pax_rev = yr_pax[y] * yr_fare[y] * 2
            yr_cargo_rev = annual_flights * cargo_capacity_kg * cargo_lf * yr_cargo_y[y]
            yr_anc_rev = yr_pax[y] * ancillary_per_pax
            yr_total = yr_pax_rev + yr_cargo_rev + yr_anc_rev
            yr_lf = yr_pax[y] / annual_capacity if annual_capacity > 0 else 0
            proj_rows.append({
                'Year': f"Year {y+1}",
                'Passengers': f"{yr_pax[y]:,}",
                'Load Factor': f"{yr_lf:.1%}",
                'Pax Revenue': f"${yr_pax_rev/1e6:.1f}M",
                'Cargo': f"${yr_cargo_rev/1e6:.1f}M",
                'Ancillary': f"${yr_anc_rev/1e6:.1f}M",
                'Total Revenue': f"${yr_total/1e6:.1f}M",
            })
        st.dataframe(proj_rows, use_container_width=True, hide_index=True)

        # --- Monthly Revenue (if seasonality computed) ---
        season_data = st.session_state.seasonality_result
        if season_data is not None:
            st.markdown("---")
            st.markdown("#### Monthly Revenue (Year 1)")
            mf = season_data['forecast']
            m_rev_rows = []
            total_monthly_rev = 0
            for i, m in enumerate(MONTHS):
                m_pax_val = round(mf.monthly_pax[i])
                m_pax_rev = m_pax_val * effective_fare * 2
                m_cargo_rev = cargo_revenue / 12  # simple even split for cargo
                m_anc = m_pax_val * ancillary_per_pax
                m_total = m_pax_rev + m_cargo_rev + m_anc
                total_monthly_rev += m_total
                m_rev_rows.append({
                    'Month': m,
                    'Passengers': f"{m_pax_val:,}",
                    'Pax Revenue': f"${m_pax_rev:,.0f}",
                    'Cargo': f"${m_cargo_rev:,.0f}",
                    'Ancillary': f"${m_anc:,.0f}",
                    'Total': f"${m_total:,.0f}",
                })
            m_rev_rows.append({
                'Month': 'TOTAL',
                'Passengers': f"{total_pax:,}",
                'Pax Revenue': f"${total_pax_revenue:,.0f}",
                'Cargo': f"${cargo_revenue:,.0f}",
                'Ancillary': f"${ancillary_revenue:,.0f}",
                'Total': f"${grand_revenue:,.0f}",
            })
            st.dataframe(m_rev_rows, use_container_width=True, hide_index=True)

        # Store revenue result
        st.session_state.revenue_result = {
            'pax_revenue': total_pax_revenue,
            'cargo_revenue': cargo_revenue,
            'ancillary_revenue': ancillary_revenue,
            'grand_revenue': grand_revenue,
            'avg_fare': avg_fare_actual,
            'yield': yield_val,
            'prask': prask,
            'trask': trask,
        }


# ============================================================================
# TAB 5: ASSUMPTIONS LOG
# ============================================================================

with tab5:
    st.markdown("###  Assumptions Log")
    st.markdown("Auto-generated 72-parameter methodology summary for client deliverables.")

    results = st.session_state.results
    cfg = st.session_state.last_config

    if results is None or cfg is None:
        st.info("Run the pipeline first to generate the assumptions log.")
    else:
        # Analyst name and engagement ref inputs
        c1, c2 = st.columns(2)
        with c1:
            al_analyst = st.text_input("Analyst Name", value="Avia Solutions", key="al_analyst")
        with c2:
            al_ref = st.text_input("Engagement Reference", value="", key="al_ref",
                                    placeholder="e.g. MAC-2025-017")

        if st.button(" Generate Assumptions Log", key="btn_assumptions"):
            with st.spinner("Building assumptions log..."):
                try:
                    log = generate_assumptions_log(cfg, results,
                                                    analyst=al_analyst,
                                                    engagement_ref=al_ref)
                    st.session_state.assumptions_log = log
                    st.success(f"Assumptions log generated: {len(log.all_assumptions())} parameters, "
                               f"{len(log.warnings)} warnings")
                except Exception as e:
                    st.error(f"Failed to generate assumptions log: {e}")
                    with st.expander("Traceback"):
                        st.code(traceback.format_exc())

        alog = st.session_state.assumptions_log
        if alog is not None:
            # Show sections
            for code in sorted(alog.sections.keys()):
                sec = alog.sections[code]
                with st.expander(f"Section {sec.code}: {sec.title} ({len(sec.assumptions)} parameters)",
                                  expanded=(code == 'A')):
                    if sec.description:
                        st.caption(sec.description)
                    rows = []
                    for a in sec.assumptions:
                        val_str = f"{a.value}" if not isinstance(a.value, float) else f"{a.value:,.4g}"
                        if a.unit:
                            val_str = f"{val_str} {a.unit}"
                        rows.append({
                            'Parameter': a.parameter,
                            'Value': val_str,
                            'Confidence': a.confidence,
                            'Source': a.source or '',
                            'Justification': a.justification or '',
                        })
                    if rows:
                        df = pd.DataFrame(rows)
                        st.dataframe(df, use_container_width=True, hide_index=True)

            # Warnings
            if alog.warnings:
                st.markdown("####  Warnings")
                for w in alog.warnings:
                    st.warning(w)
            else:
                st.success("No warnings  all parameters within normal bounds.")

            # Methodology text
            if alog.methodology_text:
                with st.expander(" Auto-generated Methodology Text (for presentations)"):
                    st.markdown(alog.methodology_text)

            # Download as Excel
            try:
                tmpf = tempfile.mktemp(suffix='.xlsx')
                writer = AssumptionsLogExcelWriter(alog)
                writer.write(tmpf)
                with open(tmpf, 'rb') as f:
                    al_bytes = f.read()
                fn = f"Assumptions_Log_{origin}_{destination}_{carrier_code}.xlsx"
                st.download_button(" Download Assumptions Log (Excel)", data=al_bytes,
                                    file_name=fn, type="primary",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_assumptions")
            except Exception as e:
                st.warning(f"Excel export not available: {e}")

            # Download as JSON
            try:
                json_str = json.dumps(alog.to_dict(), indent=2, default=str)
                fn2 = f"Assumptions_Log_{origin}_{destination}_{carrier_code}.json"
                st.download_button(" Download JSON", data=json_str,
                                    file_name=fn2,
                                    mime="application/json",
                                    key="dl_assumptions_json")
            except Exception:
                pass


# ============================================================================
# TAB 6: BUSINESS CASE
# ============================================================================

with tab6:
    st.markdown("###  Business Case Assessment")
    st.markdown("Goal-seek engine: tests whether airline targets are achievable given forecast data.")

    results = st.session_state.results
    cfg = st.session_state.last_config

    if results is None or cfg is None:
        st.info("Run the pipeline in Forecast mode first, then switch to Business Case to test targets.")
    else:
        st.markdown(f"**Base forecast:** {results['grand_total']:,} pax, "
                     f"{results['load_factor']:.1%} LF")

        st.markdown("---")
        st.markdown("##### Target Parameters")
        st.caption("Set targets in the sidebar under 'Business Case Targets', then click below.")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Y1 LF Target", f"{bc_lf_y1:.0%}")
        with c2:
            st.metric("Mature LF Target", f"{bc_lf_mature:.0%}")
        with c3:
            st.metric("Ramp Years", bc_ramp_years)
        with c4:
            st.metric("Ramp Profile", bc_ramp_profile)

        if st.button(" Run Business Case Analysis", key="btn_bc"):
            with st.spinner("Running goal-seek and sensitivity analysis..."):
                try:
                    targets = TargetSet(
                        load_factor_y1=bc_lf_y1,
                        load_factor_mature=bc_lf_mature,
                        ramp_years=bc_ramp_years,
                        ramp_profile=bc_ramp_profile,
                    )
                    engine = BusinessCaseEngine(cfg, results)
                    verdict = engine.run(targets)
                    st.session_state.business_case_verdict = verdict

                    # Generate Excel
                    tmpf = tempfile.mktemp(suffix='.xlsx')
                    writer = BusinessCaseWriter()
                    writer.write(cfg, verdict, tmpf)
                    with open(tmpf, 'rb') as f:
                        st.session_state.business_case_bytes = f.read()

                    st.success(f"Business case complete: **{verdict.verdict}** ({verdict.confidence} confidence)")
                except Exception as e:
                    st.error(f"Business case analysis failed: {e}")
                    with st.expander("Traceback"):
                        st.code(traceback.format_exc())

        verdict = st.session_state.business_case_verdict
        if verdict is not None:
            # Verdict banner
            if verdict.verdict == 'YES':
                st.markdown(f'<div style="background:#e2efda;padding:1rem;border-radius:8px;'
                            f'border-left:4px solid #006100;margin:1rem 0;">'
                            f'<h3 style="color:#006100;margin:0;"> {verdict.verdict}  {verdict.headline}</h3>'
                            f'<p style="margin:0.5rem 0 0;">{verdict.summary}</p></div>',
                            unsafe_allow_html=True)
            elif verdict.verdict == 'MARGINAL':
                st.markdown(f'<div style="background:#fff2cc;padding:1rem;border-radius:8px;'
                            f'border-left:4px solid #ff8c00;margin:1rem 0;">'
                            f'<h3 style="color:#ff8c00;margin:0;"> {verdict.verdict}  {verdict.headline}</h3>'
                            f'<p style="margin:0.5rem 0 0;">{verdict.summary}</p></div>',
                            unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="background:#f8d7da;padding:1rem;border-radius:8px;'
                            f'border-left:4px solid #c00000;margin:1rem 0;">'
                            f'<h3 style="color:#c00000;margin:0;"> {verdict.verdict}  {verdict.headline}</h3>'
                            f'<p style="margin:0.5rem 0 0;">{verdict.summary}</p></div>',
                            unsafe_allow_html=True)

            # Key metrics
            st.markdown("---")
            st.markdown("##### Key Metrics")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.metric("Base Forecast", f"{verdict.base_forecast:,.0f}",
                           delta=f"{verdict.gap_to_y1_pct:+.1f}% to Y1")
            with c2:
                st.metric("Base LF", f"{verdict.base_load_factor:.1%}")
            with c3:
                st.metric("Y1 Target Pax", f"{verdict.y1_target_pax:,.0f}")
            with c4:
                st.metric("Mature Target Pax", f"{verdict.mature_target_pax:,.0f}")

            # Year-by-year ramp
            if verdict.year_by_year:
                st.markdown("##### Year-by-Year Ramp")
                yby_rows = []
                for yr in verdict.year_by_year:
                    yby_rows.append({
                        'Year': yr.get('year', ''),
                        'Ramp %': f"{yr.get('ramp_fraction', 0):.0%}",
                        'Pax': f"{yr.get('pax', 0):,.0f}",
                        'LF': f"{yr.get('load_factor', 0):.1%}",
                        'Meets Target': '' if yr.get('meets_target', False) else '',
                    })
                st.dataframe(pd.DataFrame(yby_rows), use_container_width=True, hide_index=True)

            # Parameter gaps
            if verdict.parameter_gaps:
                st.markdown("##### Parameter Gap Analysis")
                st.caption("What needs to change to hit targets?")
                gap_rows = []
                for g in verdict.parameter_gaps:
                    gap_rows.append({
                        'Parameter': g.param_name,
                        'Current': f"{g.current_value:,.2f}",
                        'Required': f"{g.required_value:,.2f}",
                        'Change': f"{g.change_pct:+.1f}%",
                        'Risk': g.risk_level.upper(),
                        'Achievable': '' if g.achievable else '',
                        'Note': g.analyst_note,
                    })
                st.dataframe(pd.DataFrame(gap_rows), use_container_width=True, hide_index=True)

            # Sensitivity results
            if verdict.sensitivity_results:
                st.markdown("##### Sensitivity Analysis")
                for param, points in verdict.sensitivity_results.items():
                    with st.expander(f"Sensitivity: {param}"):
                        s_rows = []
                        for p in points:
                            s_rows.append({
                                'Factor': f"{p.factor:.2f}",
                                'Total Pax': f"{p.total_pax:,.0f}",
                                'LF': f"{p.load_factor:.1%}",
                                'Delta': f"{p.delta_pct:+.1f}%",
                                'Hits Y1': '' if p.hits_y1_target else '',
                                'Hits Mature': '' if p.hits_mature_target else '',
                            })
                        st.dataframe(pd.DataFrame(s_rows), use_container_width=True, hide_index=True)

            # Risk flags
            if verdict.risk_flags:
                st.markdown("#####  Risk Flags")
                for rf in verdict.risk_flags:
                    st.warning(rf)

            # Airline pushback areas
            if verdict.airline_pushback_areas:
                st.markdown("#####  Airline Pushback Areas")
                st.caption("Where the airline is most likely to challenge the assumptions:")
                for pb in verdict.airline_pushback_areas:
                    st.markdown(f"- {pb}")

            # Download
            if st.session_state.business_case_bytes:
                fn = f"BusinessCase_{origin}_{destination}_{carrier_code}.xlsx"
                st.download_button(" Download Business Case (Excel)",
                                    data=st.session_state.business_case_bytes,
                                    file_name=fn, type="primary",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_bc")


# ============================================================================
# TAB 7: OUTPUT WORKBOOK
# ============================================================================

with tab7:
    st.markdown("###  Standard Output Workbook")
    st.markdown("Generates the standardised Avia Solutions output workbook with all forecast tables.")

    results = st.session_state.results
    cfg = st.session_state.last_config

    if results is None or cfg is None:
        st.info("Run the pipeline first to generate the output workbook.")
    else:
        st.markdown(f"**Route:** {st.session_state.last_config_summary}")
        st.markdown(f"**Forecast:** {results['grand_total']:,} pax, {results['load_factor']:.1%} LF")

        c1, c2 = st.columns(2)
        with c1:
            ow_analyst = st.text_input("Analyst", value="Avia Solutions", key="ow_analyst")
        with c2:
            ow_date = st.text_input("Date", value=datetime.now().strftime('%d %B %Y'),
                                     key="ow_date")

        if st.button(" Generate Output Workbook", key="btn_ow"):
            with st.spinner("Building standardised output workbook..."):
                try:
                    tmpf = tempfile.mktemp(suffix='.xlsx')
                    writer = StandardOutputWriter(
                        cfg, results,
                        audit_log=st.session_state.run_log,
                        analyst=ow_analyst,
                        engagement_date=ow_date,
                    )
                    writer.write_all()
                    writer.save(tmpf)

                    with open(tmpf, 'rb') as f:
                        ow_bytes = f.read()

                    fn = f"Output_{origin}_{destination}_{carrier_code}_{datetime.now():%Y%m%d}.xlsx"
                    st.success("Output workbook generated successfully.")
                    st.download_button(" Download Output Workbook",
                                        data=ow_bytes,
                                        file_name=fn, type="primary",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="dl_ow")
                except Exception as e:
                    st.error(f"Output workbook generation failed: {e}")
                    with st.expander("Traceback"):
                        st.code(traceback.format_exc())


# ============================================================================
# TAB 8: Q&A CHECKLIST
# ============================================================================

with tab8:
    st.markdown("###  Q&A Checklist")
    st.caption("Automated quality control checklist based on Avia Solutions' QA templates. "
               "Evaluates workflow completeness, technical forecast validity, and flags items "
               "requiring manual reviewer sign-off.")

    results = st.session_state.results
    if results:
        try:
            config = st.session_state.get('last_config')

            # Optional: previous forecast for comparison
            with st.expander("Previous Forecast (optional comparison)", expanded=False):
                prev_pax = st.number_input("Previous forecast total pax (0 = skip)",
                                           min_value=0, value=0, step=1000, key='qa_prev_pax')
                previous_forecast = {'total_annual_pax': prev_pax} if prev_pax > 0 else None

            analyst_name = st.text_input("Analyst name", value="Pipeline", key='qa_analyst')

            if st.button("Run Q&A Checklist", type="primary", key='run_qa'):
                with st.spinner("Running quality checks..."):
                    report = run_qa_checklist(
                        config=config,
                        results=results,
                        previous_forecast=previous_forecast,
                        analyst=analyst_name,
                    )
                    st.session_state['qa_report'] = report

            if 'qa_report' in st.session_state and st.session_state['qa_report']:
                report = st.session_state['qa_report']

                # Overall status banner
                status_colors = {
                    'PASS': ('#006100', '#C6EFCE'),
                    'WARNING': ('#9C5700', '#FFEB9C'),
                    'FAIL': ('#9C0006', '#FFC7CE'),
                    'MANUAL REVIEW': ('#1B3A5C', '#D6E4F0'),
                }
                fg, bg = status_colors.get(report.overall_status.value, ('#333', '#eee'))
                st.markdown(
                    f'<div style="background:{bg};color:{fg};padding:1rem;border-radius:8px;'
                    f'text-align:center;font-size:1.2rem;font-weight:600;margin-bottom:1rem;">'
                    f'Overall: {report.overall_status.value} &nbsp;|&nbsp; '
                    f'{report.pass_count} Pass &nbsp;|&nbsp; {report.warn_count} Warn &nbsp;|&nbsp; '
                    f'{report.fail_count} Fail &nbsp;|&nbsp; {report.manual_count} Manual'
                    f'</div>', unsafe_allow_html=True
                )

                # Results by category
                for cat in CheckCategory:
                    items = report.by_category(cat)
                    if items:
                        with st.expander(f"{cat.value} ({len(items)} checks)", expanded=True):
                            rows = []
                            for r in items:
                                icon = {"PASS": "", "WARNING": "", "FAIL": "",
                                        "MANUAL REVIEW": "", "SKIPPED": ""}.get(r.status.value, "?")
                                rows.append({
                                    '': icon,
                                    'ID': r.check_id,
                                    'Check': r.title,
                                    'Status': r.status.value,
                                    'Detail': r.detail,
                                    'Evidence': r.evidence,
                                })
                            st.dataframe(rows, use_container_width=True, hide_index=True)

                # Download
                st.markdown("---")
                dl1, dl2, _ = st.columns([1, 1, 2])
                with dl1:
                    try:
                        qa_path = os.path.join(tempfile.gettempdir(), 'qa_checklist.xlsx')
                        report.to_excel(qa_path)
                        with open(qa_path, 'rb') as f:
                            qa_bytes = f.read()
                        origin = getattr(config, 'origin', 'XXX') if config else 'XXX'
                        dest = getattr(config, 'destination', 'YYY') if config else 'YYY'
                        st.download_button(
                            " Download Q&A Checklist",
                            data=qa_bytes,
                            file_name=f"QA_Checklist_{origin}_{dest}_{datetime.now():%Y%m%d}.xlsx",
                            type="primary",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"Excel export failed: {e}")
                with dl2:
                    st.download_button(
                        " Download as JSON",
                        data=json.dumps(report.to_dict(), indent=2),
                        file_name=f"QA_Checklist_{datetime.now():%Y%m%d}.json",
                        mime="application/json"
                    )

        except Exception as e:
            st.error(f"Q&A checklist failed: {e}")
            with st.expander("Traceback"):
                st.code(traceback.format_exc())
    else:
        st.info("Run the pipeline from the 'Data Upload & Run' tab to generate the Q&A checklist.")


# ============================================================================
# TAB 9: SPILL ANALYSIS (Capacity Constraint Tool)
# ============================================================================

with tab9:
    st.markdown("###  Spill & Capacity Constraint Analysis")
    st.caption("Interactive tool for analysing capacity spill  when demand exceeds "
               "available seats. Adjusts the monthly profile to show constrained vs "
               "unconstrained passengers.")

    results = st.session_state.results
    if results:
        config = st.session_state.get('last_config')
        total_pax = results.get('total_annual_pax') or results.get('grand_total') or 0
        seats = getattr(config, 'seats', 0) or getattr(config, 'seat_capacity', 0) or 0
        freq = getattr(config, 'weekly_frequency', 0) or 0

        if total_pax > 0 and seats > 0 and freq > 0:
            st.markdown("##### Capacity Parameters")
            col1, col2, col3 = st.columns(3)
            with col1:
                adj_seats = st.number_input("Seats per flight", value=int(seats),
                                            min_value=50, max_value=600, step=10, key='spill_seats')
            with col2:
                adj_freq = st.number_input("Weekly frequency", value=int(freq),
                                           min_value=1, max_value=28, step=1, key='spill_freq')
            with col3:
                c_factor = st.slider("C-factor (spill elasticity)", 0.5, 1.5, 1.0, 0.05,
                                     key='spill_cfactor',
                                     help="Lower = more demand lost to spill. 1.0 = standard.")

            # Seasonal profile
            profile_key = st.session_state.get('seasonal_profile', 'transatlantic')
            PROFILES = {
                'flat': [1.0]*12,
                'transatlantic': [0.75, 0.72, 0.85, 0.95, 1.05, 1.20, 1.30, 1.28, 1.10, 0.95, 0.80, 0.75],
                'summer_peak': [0.70, 0.65, 0.80, 0.90, 1.05, 1.25, 1.40, 1.35, 1.05, 0.85, 0.70, 0.65],
                'winter_sun': [1.20, 1.15, 1.00, 0.85, 0.75, 0.70, 0.80, 0.85, 0.90, 1.05, 1.15, 1.25],
                'vfr_diaspora': [0.90, 0.85, 0.95, 0.95, 1.00, 1.10, 1.15, 1.10, 1.00, 0.95, 0.95, 1.05],
            }
            profile = PROFILES.get(profile_key, PROFILES['transatlantic'])

            # Compute monthly spill
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

            # Normalise profile to preserve annual total
            total_weight = sum(p * d for p, d in zip(profile, days_in_month))
            annual_days = sum(days_in_month)

            spill_rows = []
            total_unconstrained = 0
            total_constrained = 0
            total_spilled = 0

            for i, (m, p, d) in enumerate(zip(months, profile, days_in_month)):
                # Monthly unconstrained demand
                monthly_pax = total_pax * (p * d) / total_weight
                # Monthly capacity (both directions)
                weekly_seats = adj_seats * adj_freq * 2
                monthly_capacity = weekly_seats * d / 7

                # Apply C-factor spill model
                if monthly_pax > monthly_capacity:
                    overflow = monthly_pax - monthly_capacity
                    spill = overflow * c_factor
                    constrained = monthly_pax - spill
                    lf = 1.0  # At capacity
                else:
                    spill = 0
                    constrained = monthly_pax
                    lf = monthly_pax / monthly_capacity if monthly_capacity > 0 else 0

                total_unconstrained += monthly_pax
                total_constrained += constrained
                total_spilled += spill

                spill_rows.append({
                    'Month': m,
                    'Unconstrained': f'{monthly_pax:,.0f}',
                    'Capacity': f'{monthly_capacity:,.0f}',
                    'Constrained': f'{constrained:,.0f}',
                    'Spill': f'{spill:,.0f}',
                    'Load Factor': f'{lf:.1%}',
                    'Status': ' SPILL' if spill > 0 else ' OK',
                })

            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Unconstrained", f"{total_unconstrained:,.0f}")
            col2.metric("Constrained", f"{total_constrained:,.0f}")
            col3.metric("Total Spill", f"{total_spilled:,.0f}",
                         delta=f"-{total_spilled/total_unconstrained:.1%}" if total_unconstrained > 0 else "")
            annual_cap = adj_seats * adj_freq * 52 * 2
            annual_lf = total_constrained / annual_cap if annual_cap > 0 else 0
            col4.metric("Annual LF", f"{annual_lf:.1%}")

            # Monthly table
            st.dataframe(spill_rows, use_container_width=True, hide_index=True)

            # Spill chart
            spill_df = pd.DataFrame({
                'Month': months,
                'Unconstrained': [total_pax * (p * d) / total_weight for p, d in zip(profile, days_in_month)],
                'Capacity': [adj_seats * adj_freq * 2 * d / 7 for d in days_in_month],
            })
            st.bar_chart(spill_df.set_index('Month')[['Unconstrained']], color='#336699',
                         use_container_width=True)
            st.line_chart(spill_df.set_index('Month')[['Capacity']], color='#cc3333',
                          use_container_width=True)

            # Frequency analysis: what frequency eliminates spill?
            with st.expander("Frequency to Eliminate Spill"):
                for test_freq in range(int(adj_freq), int(adj_freq) + 15):
                    any_spill = False
                    for p, d in zip(profile, days_in_month):
                        m_pax = total_pax * (p * d) / total_weight
                        m_cap = adj_seats * test_freq * 2 * d / 7
                        if m_pax > m_cap:
                            any_spill = True
                            break
                    if not any_spill:
                        st.success(f"**{test_freq}x weekly** eliminates all monthly spill "
                                   f"(up from {adj_freq}x)")
                        break
                else:
                    st.warning("Even at +14 frequency, spill remains in peak months.")

        else:
            st.warning("Need total pax, seats, and frequency to run spill analysis.")
    else:
        st.info("Run the pipeline from the 'Data Upload & Run' tab to see spill analysis.")



# ============================================================================
# TAB 10: MARKET RESEARCH (Brief Generator + Research Executor)
# ============================================================================

with tab10:
    st.markdown("###  Market Research")
    st.caption("Two-phase workflow: (1) generate a structured research brief with classified "
               "queries, then (2) track findings using the Research Executor for full citation "
               "management, quality scoring, and downloadable outputs.")

    # ----------------------------------------------------------------
    # Phase selector
    # ----------------------------------------------------------------
    mr_phase = st.radio(
        "Select phase",
        ["Research Brief", "Research Executor"],
        horizontal=True,
        key="mr_phase_select",
    )

    # ================================================================
    # PHASE 1: Research Brief (existing functionality, preserved)
    # ================================================================
    if mr_phase == "Research Brief":
        st.markdown("#### Research Brief Generator")
        st.caption("Generates classified queries for each of the 10 research blocks, "
                   "prioritised by route profile and buyer type.")

        # Auto-populate from sidebar route config if available
        _cfg = st.session_state.get('last_config')
        _def_origin = getattr(_cfg, 'home_airport_code', '') if _cfg else ''
        _def_dest = getattr(_cfg, 'dest_airport_code', '') if _cfg else ''
        _def_airline = getattr(_cfg, 'airline_name', '') if _cfg else ''
        _def_carrier_type = getattr(_cfg, 'carrier_type', 'Full Service') if _cfg else 'Full Service'
        _def_route_type = getattr(_cfg, 'route_type', 'Long Haul') if _cfg else 'Long Haul'

        # Map sidebar carrier/route type to research route type
        _rt_map_sidebar = {
            'Long Haul': 'HUB_LONGHAUL', 'Ultra Long Haul': 'HUB_LONGHAUL',
            'Hub Feed': 'HUB_LONGHAUL', 'LCC P2P': 'LCC_P2P',
            'Charter': 'CHARTER_LEISURE',
        }
        _def_rt_key = _rt_map_sidebar.get(str(_def_route_type), 'HUB_LONGHAUL')

        # City names from airport codes (basic mapping)
        _city_map = {
            'LHR': 'London', 'CDG': 'Paris', 'AMS': 'Amsterdam', 'FRA': 'Frankfurt',
            'TPA': 'Tampa', 'SJC': 'San Jose', 'SFO': 'San Francisco', 'LAX': 'Los Angeles',
            'JFK': 'New York', 'ORD': 'Chicago', 'DXB': 'Dubai', 'SIN': 'Singapore',
            'HKG': 'Hong Kong', 'ICN': 'Seoul', 'NRT': 'Tokyo', 'TPE': 'Taipei',
            'MAN': 'Manchester', 'DUB': 'Dublin', 'MUC': 'Munich', 'MAD': 'Madrid',
            'FCO': 'Rome', 'ATH': 'Athens', 'IST': 'Istanbul', 'DOH': 'Doha',
            'BKK': 'Bangkok', 'DEL': 'Delhi', 'BOM': 'Mumbai', 'CAN': 'Guangzhou',
            'MIA': 'Miami', 'ATL': 'Atlanta', 'BOS': 'Boston', 'SEA': 'Seattle',
        }
        _country_map = {
            'LHR': 'United Kingdom', 'CDG': 'France', 'AMS': 'Netherlands', 'FRA': 'Germany',
            'TPA': 'United States', 'SJC': 'United States', 'SFO': 'United States',
            'LAX': 'United States', 'JFK': 'United States', 'MAN': 'United Kingdom',
            'DUB': 'Ireland', 'MUC': 'Germany', 'MAD': 'Spain', 'FCO': 'Italy',
            'DXB': 'UAE', 'SIN': 'Singapore', 'HKG': 'China', 'ICN': 'South Korea',
            'NRT': 'Japan', 'TPE': 'Taiwan', 'BKK': 'Thailand', 'DOH': 'Qatar',
            'IST': 'Turkey', 'ATH': 'Greece', 'DEL': 'India', 'BOM': 'India',
        }
        _def_origin_city = _city_map.get(_def_origin, '')
        _def_dest_city = _city_map.get(_def_dest, '')
        _def_origin_country = _country_map.get(_def_origin, '')
        _def_dest_country = _country_map.get(_def_dest, '')

        # Route type options matching the actual RouteType enum
        _rt_options = ['HUB_LONGHAUL', 'LCC_P2P', 'CHARTER_LEISURE', 'SIXTH_FREEDOM']
        _rt_index = _rt_options.index(_def_rt_key) if _def_rt_key in _rt_options else 0

        col1, col2 = st.columns(2)
        with col1:
            mr_origin = st.text_input("Origin airport (IATA)", value=_def_origin or "TPA", key='mr_origin')
            mr_dest = st.text_input("Destination airport (IATA)", value=_def_dest or "CDG", key='mr_dest')
            mr_airline = st.text_input("Airline", value=_def_airline or "Air France", key='mr_airline')
            mr_origin_city = st.text_input("Origin city", value=_def_origin_city or "Tampa", key='mr_origin_city')
            mr_dest_city = st.text_input("Destination city", value=_def_dest_city or "Paris", key='mr_dest_city')
        with col2:
            mr_demand = st.selectbox("Demand profile", ['BUSINESS', 'LEISURE', 'VFR_DIASPORA', 'MIXED'],
                                     index=3, key='mr_demand')
            mr_route_type = st.selectbox("Route type", _rt_options,
                                         index=_rt_index, key='mr_route_type')
            mr_buyer = st.selectbox("Buyer type", ['AIRPORT', 'AIRLINE', 'FUND'],
                                    index=0, key='mr_buyer')
            mr_origin_country = st.text_input("Origin country", value=_def_origin_country or "USA", key='mr_origin_country')
            mr_dest_country = st.text_input("Destination country", value=_def_dest_country or "France", key='mr_dest_country')

        col3, col4 = st.columns(2)
        with col3:
            mr_hub = st.text_input("Hub airport (if applicable)", value=_def_dest if _def_rt_key == 'HUB_LONGHAUL' else "", key='mr_hub')
        with col4:
            mr_origin_region = st.text_input("Origin region (e.g. Silicon Valley)", value="", key='mr_origin_region')

        if st.button("Generate Research Brief", type="primary", key='gen_research'):
            try:
                demand_map = {'BUSINESS': DemandProfile.BUSINESS, 'LEISURE': DemandProfile.LEISURE,
                              'VFR_DIASPORA': DemandProfile.VFR_DIASPORA, 'MIXED': DemandProfile.MIXED}
                route_map = {'HUB_LONGHAUL': MRRouteType.HUB_LONGHAUL, 'LCC_P2P': MRRouteType.LCC_P2P,
                             'CHARTER_LEISURE': MRRouteType.CHARTER_LEISURE,
                             'SIXTH_FREEDOM': MRRouteType.SIXTH_FREEDOM}
                buyer_map = {'AIRPORT': BuyerType.AIRPORT, 'AIRLINE': BuyerType.AIRLINE,
                             'FUND': BuyerType.FUND}

                research_config = RouteResearchConfig(
                    origin=mr_origin,
                    destination=mr_dest,
                    airline=mr_airline,
                    origin_city=mr_origin_city,
                    destination_city=mr_dest_city,
                    origin_country=mr_origin_country,
                    destination_country=mr_dest_country,
                    demand_profile=demand_map[mr_demand],
                    route_type=route_map[mr_route_type],
                    buyer_type=buyer_map[mr_buyer],
                    hub_airport=mr_hub,
                    origin_region=mr_origin_region,
                )

                blocks = generate_queries(research_config)
                plan_text = summarise_research_plan(research_config, blocks)
                st.session_state['mr_blocks'] = blocks
                st.session_state['mr_plan'] = plan_text
                st.session_state['mr_config'] = research_config

                # Also create executor for Phase 2 handoff
                demand_lower = mr_demand.lower()
                route_lower = mr_route_type.lower().replace('_', ' ')
                executor_route_type = "hub_longhaul" if "long" in route_lower else route_lower.replace(' ', '_')
                executor_buyer = mr_buyer.lower()

                _executor = ResearchExecutor(
                    origin=mr_origin,
                    destination=mr_dest,
                    origin_city=mr_origin_city or mr_origin,
                    destination_city=mr_dest_city or mr_dest,
                    origin_country=mr_origin_country or "",
                    destination_country=mr_dest_country or "",
                    airline=mr_airline,
                    demand_profile=demand_lower,
                    route_type=executor_route_type,
                    buyer_type=executor_buyer,
                    hub_airport=mr_hub,
                    origin_region=mr_origin_region,
                )
                st.session_state['mr_executor'] = _executor

                # Generate execution plan
                exec_plan = generate_execution_plan(_executor)
                st.session_state['mr_exec_plan'] = exec_plan

                # Get execution sequence for block init
                exec_seq = get_execution_sequence(demand_lower, executor_route_type, executor_buyer)
                st.session_state['mr_exec_seq'] = exec_seq

            except Exception as e:
                st.error(f"Research brief generation failed: {e}")
                with st.expander("Traceback"):
                    st.code(traceback.format_exc())

        if 'mr_blocks' in st.session_state and st.session_state['mr_blocks']:
            blocks = st.session_state['mr_blocks']
            plan_text = st.session_state.get('mr_plan', '')

            essential = sum(1 for b in blocks if b.relevance == Relevance.ESSENTIAL)
            include = sum(1 for b in blocks if b.relevance == Relevance.INCLUDE)
            optional = sum(1 for b in blocks if b.relevance == Relevance.OPTIONAL)
            total_queries = sum(len(b.queries) for b in blocks)

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Research Blocks", len(blocks))
            col2.metric("Essential", essential)
            col3.metric("Include", include)
            col4.metric("Total Queries", total_queries)

            for block in blocks:
                rel_icon = {'ESSENTIAL': 'ESSENTIAL', 'INCLUDE': 'INCLUDE', 'OPTIONAL': 'OPTIONAL'}.get(
                    block.relevance.value, '')
                with st.expander(f"{block.name} [{rel_icon}]",
                                 expanded=(block.relevance == Relevance.ESSENTIAL)):
                    if block.rationale:
                        st.caption(block.rationale)
                    for q in block.queries:
                        st.markdown(f"- `{q.query}`")
                        if q.preferred_sources:
                            st.caption(f"  Sources: {', '.join(q.preferred_sources)}")

            st.markdown("---")
            dl_c1, dl_c2 = st.columns(2)
            with dl_c1:
                st.download_button(
                    "Download Research Plan (Markdown)",
                    data=plan_text,
                    file_name=f"Research_Brief_{mr_origin}_{mr_dest}_{datetime.now():%Y%m%d}.md",
                    mime="text/markdown",
                    key="dl_brief_md",
                )
            with dl_c2:
                if 'mr_exec_plan' in st.session_state:
                    st.download_button(
                        "Download Execution Plan",
                        data=st.session_state['mr_exec_plan'],
                        file_name=f"Execution_Plan_{mr_origin}_{mr_dest}_{datetime.now():%Y%m%d}.txt",
                        mime="text/plain",
                        key="dl_exec_plan",
                    )

            if 'mr_executor' in st.session_state:
                st.success("Executor ready -- switch to **Research Executor** phase to track findings.")

    # ================================================================
    # PHASE 2: Research Executor (new in v9)
    # ================================================================
    elif mr_phase == "Research Executor":
        st.markdown("#### Research Executor -- Findings Tracker")

        if 'mr_executor' not in st.session_state:
            st.warning("Generate a Research Brief first (Phase 1) to initialise the executor.")
            st.caption("The executor needs the route configuration from the brief generator.")

            # Quick-start option
            st.markdown("---")
            st.markdown("**Or quick-start an executor directly:**")
            qs_col1, qs_col2 = st.columns(2)
            with qs_col1:
                qs_origin = st.text_input("Origin (IATA)", value="LHR", key='qs_origin')
                qs_dest = st.text_input("Destination (IATA)", value="SJC", key='qs_dest')
                qs_origin_city = st.text_input("Origin city", value="London", key='qs_ocity')
                qs_dest_city = st.text_input("Dest city", value="San Jose", key='qs_dcity')
                qs_airline = st.text_input("Airline", value="British Airways", key='qs_airline')
            with qs_col2:
                qs_ocountry = st.text_input("Origin country", value="United Kingdom", key='qs_oc')
                qs_dcountry = st.text_input("Dest country", value="United States", key='qs_dc')
                qs_demand = st.selectbox("Demand", ['mixed', 'business', 'leisure', 'vfr_diaspora'],
                                         key='qs_demand')
                qs_rtype = st.selectbox("Route type", ['hub_longhaul', 'short_haul', 'medium_haul'],
                                        key='qs_rtype')
                qs_buyer = st.selectbox("Buyer", ['airport', 'airline', 'fund'], key='qs_buyer')

            if st.button("Initialise Executor", key="qs_init"):
                _executor = ResearchExecutor(
                    origin=qs_origin, destination=qs_dest,
                    origin_city=qs_origin_city, destination_city=qs_dest_city,
                    origin_country=qs_ocountry, destination_country=qs_dcountry,
                    airline=qs_airline, demand_profile=qs_demand,
                    route_type=qs_rtype, buyer_type=qs_buyer,
                )
                st.session_state['mr_executor'] = _executor
                exec_plan = generate_execution_plan(_executor)
                st.session_state['mr_exec_plan'] = exec_plan
                exec_seq = get_execution_sequence(qs_demand, qs_rtype, qs_buyer)
                st.session_state['mr_exec_seq'] = exec_seq
                st.rerun()

        else:
            _executor = st.session_state['mr_executor']
            exec_seq = st.session_state.get('mr_exec_seq', [])

            # ---- Executor status header ----
            st.markdown(f"**Route:** {_executor.origin} -- {_executor.destination} "
                        f"({_executor.origin_city} to {_executor.destination_city})  \n"
                        f"**Airline:** {_executor.airline} | "
                        f"**Profile:** {_executor.demand_profile} | "
                        f"**Buyer:** {_executor.buyer_type}")

            # Count initialised blocks and findings
            n_blocks_init = len(_executor.blocks)
            n_findings = sum(len(bf.findings) for bf in _executor.blocks.values())
            n_gaps = sum(len(bf.data_gaps) for bf in _executor.blocks.values())

            sc1, sc2, sc3, sc4 = st.columns(4)
            sc1.metric("Blocks Initialised", n_blocks_init)
            sc2.metric("Findings Recorded", n_findings)
            sc3.metric("Data Gaps", n_gaps)
            sc4.metric("Exec Sequence", len(exec_seq))

            st.markdown("---")

            # ---- Execution sequence display ----
            with st.expander("Execution Sequence & Prompts", expanded=False):
                if exec_seq:
                    for i, (block_id, relevance) in enumerate(exec_seq, 1):
                        status = "[done]" if block_id in _executor.blocks else "[    ]"
                        st.markdown(f"{status} **Step {i}:** `{block_id}` [{relevance}]")

                if st.button("Initialise All Blocks", key="init_all_blocks"):
                    block_meta = {
                        "economic_context": ("Economic Context", "S4"),
                        "corporate_links": ("Corporate / Technology Links", "S5A"),
                        "tourism": ("Tourism / Visitor Data", "S5B"),
                        "trade": ("Trade and Investment", "S5C"),
                        "education": ("Education / Student Links", "S5D"),
                        "diaspora": ("Diaspora / VFR Population", "S5E"),
                        "passenger_profile": ("Passenger Profile", "S5F"),
                        "non_cannibalization": ("Non-Cannibalization Evidence", "S5G"),
                        "case_study": ("Case Study / Comparable Route", "S5H"),
                        "airport_overview": ("Airport Overview", "S8"),
                    }
                    for block_id, relevance in exec_seq:
                        if block_id not in _executor.blocks:
                            name, ref = block_meta.get(block_id, (block_id, ""))
                            _executor.init_block(block_id, name, ref, relevance)
                    st.success(f"Initialised {len(_executor.blocks)} blocks")
                    st.rerun()

            st.markdown("---")

            # ---- AUTO-RESEARCH: Generate prompt & import results ----
            st.markdown("#### Auto-Research")
            st.caption("Generate a research prompt for Claude (with web search), "
                       "then paste back the JSON results to auto-populate all blocks.")

            if _executor.blocks:
                # Build the research prompt
                _blocks_data = st.session_state.get('mr_blocks', [])
                _mr_config = st.session_state.get('mr_config')

                if st.button("Generate Research Prompt", key="gen_auto_prompt"):
                    prompt_lines = [
                        f"You are researching the {_executor.origin_city} ({_executor.origin}) to "
                        f"{_executor.destination_city} ({_executor.destination}) air route for "
                        f"{_executor.airline}.",
                        f"",
                        f"Buyer type: {_executor.buyer_type} | Demand profile: {_executor.demand_profile} | "
                        f"Route type: {_executor.route_type}",
                        f"",
                        f"For each research block below, use web search to find 3-5 factual findings "
                        f"with full citations. Prefer authoritative sources: government statistics, "
                        f"IATA/ACI/ICAO publications, airline press releases, airport authority data, "
                        f"Reuters/Bloomberg/FT, national tourism boards.",
                        f"",
                        f"Return your findings as a JSON object with this structure:",
                        f'```json',
                        f'{{',
                        f'  "block_id": [',
                        f'    {{',
                        f'      "claim": "factual statement",',
                        f'      "value": "numeric value if applicable",',
                        f'      "unit": "USD/visitors/%/etc",',
                        f'      "year": "data year",',
                        f'      "source_name": "organisation name",',
                        f'      "source_title": "article or report title",',
                        f'      "source_date": "publication date",',
                        f'      "source_url": "full URL",',
                        f'      "source_type": "government|industry|news|academic|airline|international",',
                        f'      "confidence": "high|medium|low",',
                        f'      "relevance": "why this matters for the route case"',
                        f'    }}',
                        f'  ]',
                        f'}}',
                        f'```',
                        f"",
                        f"RESEARCH BLOCKS:",
                        f"",
                    ]

                    for block_id, bf in _executor.blocks.items():
                        prompt_lines.append(f"## {bf.block_name} (block_id: \"{block_id}\") [{bf.relevance}]")
                        # Find matching queries from the research brief
                        if _blocks_data:
                            for rb in _blocks_data:
                                if rb.block_id == block_id:
                                    for q in rb.queries:
                                        prompt_lines.append(f"  - {q.query}")
                                    break
                        else:
                            # Generate basic queries from block metadata
                            prompt_lines.append(f"  - Search for: {bf.block_name} {_executor.origin_city} {_executor.destination_city}")
                        prompt_lines.append("")

                    prompt_lines.append("Return ONLY the JSON object, no other text. "
                                        "Include at least 3 findings per ESSENTIAL block and "
                                        "2 per INCLUDE block. Also generate a 2-sentence summary for each block "
                                        "in a separate key called \"_summaries\" with block_id -> summary text.")

                    full_prompt = "\n".join(prompt_lines)
                    st.session_state['auto_research_prompt'] = full_prompt

                if 'auto_research_prompt' in st.session_state:
                    st.info("**Step 1:** Copy or download the prompt below.  \n"
                            "**Step 2:** Open a new Claude chat (with web search ON) and paste the prompt.  \n"
                            "**Step 3:** Wait for Claude to search the web and return a JSON object.  \n"
                            "**Step 4:** Copy Claude's JSON response and paste it into the box below.")
                    with st.expander("STEP 1: Research Prompt (copy this to a new Claude chat)", expanded=False):
                        st.code(st.session_state['auto_research_prompt'], language="text")
                        st.download_button(
                            "Download prompt as .txt",
                            data=st.session_state['auto_research_prompt'],
                            file_name=f"research_prompt_{_executor.origin}_{_executor.destination}.txt",
                            mime="text/plain",
                            key="dl_auto_prompt",
                        )

                st.markdown("---")
                st.markdown("**STEP 4: Paste Claude's JSON response here**")
                st.caption("This should be the JSON object that Claude returns after researching — "
                           "NOT the prompt above. It will look like: "
                           '`{"economic_context": [{"claim": "...", "source_name": "..."}], ...}`')
                auto_json = st.text_area(
                    "JSON response from Claude",
                    height=300,
                    key="auto_json_input",
                    placeholder='{\n  "economic_context": [\n    {"claim": "Tampa Bay GDP...", "value": "$180B", ...}\n  ],\n  ...\n}',
                )

                if st.button("Import Findings", type="primary", key="btn_import_json"):
                    if auto_json.strip():
                        try:
                            # Clean JSON: strip markdown code fences
                            clean = auto_json.strip()
                            if clean.startswith('```'):
                                clean = clean.split('\n', 1)[1] if '\n' in clean else clean[3:]
                            if clean.endswith('```'):
                                clean = clean[:-3]
                            clean = clean.strip()

                            # Attempt to parse JSON, with repair for truncated responses
                            data = None
                            try:
                                data = json.loads(clean)
                            except json.JSONDecodeError as je:
                                # Try to repair truncated JSON from Claude hitting output limit
                                st.warning(f"JSON appears truncated (Claude may have hit output limit). Attempting repair...")
                                repaired = _repair_truncated_json(clean)
                                if repaired:
                                    try:
                                        data = json.loads(repaired)
                                        st.info("JSON repaired successfully — some blocks at the end may be incomplete.")
                                    except json.JSONDecodeError:
                                        st.error(f"Could not repair JSON. Error: {je}\n\n"
                                                 f"**Tip:** Ask Claude to output the research in 2 parts. "
                                                 f"First paste part 1, import, then paste part 2 and import again.")
                                        data = None
                                else:
                                    st.error(f"Invalid JSON: {je}\n\n"
                                             f"First 200 chars: `{clean[:200]}`")

                            if data is None:
                                pass  # Error already shown
                            else:
                                imported = 0
                                summaries_imported = 0
                                skipped_blocks = []

                                # Import summaries if present (could be top-level or inside blocks)
                                summaries = data.pop('_summaries', {})

                                # Handle nested structure: Claude may wrap findings inside
                                # a "blocks" key or similar container
                                if 'blocks' in data and isinstance(data['blocks'], dict):
                                    # Findings are nested: {"blocks": {"economic_context": [...]}}
                                    findings_data = data['blocks']
                                    # Check for summaries inside blocks too
                                    if '_summaries' in findings_data:
                                        summaries.update(findings_data.pop('_summaries', {}))
                                else:
                                    # Findings are top-level: {"economic_context": [...]}
                                    findings_data = data

                                # Also check for summaries nested elsewhere
                                if not summaries:
                                    for k in ('_summaries', 'summaries', 'block_summaries'):
                                        if k in findings_data and isinstance(findings_data[k], dict):
                                            summaries = findings_data.pop(k)
                                            break
                                for bid, summary_text in summaries.items():
                                    if bid in _executor.blocks:
                                        _executor.set_block_summary(bid, summary_text, summary_text)
                                        summaries_imported += 1

                                # Show what we received vs what we expect
                                json_keys = [k for k in findings_data.keys()]
                                executor_keys = list(_executor.blocks.keys())

                                # Import findings
                                for block_id, findings_list in findings_data.items():
                                    if block_id not in _executor.blocks:
                                        skipped_blocks.append(block_id)
                                        continue
                                    if not isinstance(findings_list, list):
                                        skipped_blocks.append(f"{block_id} (not a list)")
                                        continue
                                    for f_data in findings_list:
                                        try:
                                            finding = make_finding(
                                                claim=f_data.get('claim', ''),
                                                value=str(f_data.get('value', '')),
                                                unit=f_data.get('unit', ''),
                                                year=f_data.get('year', ''),
                                                source_name=f_data.get('source_name', ''),
                                                source_title=f_data.get('source_title', ''),
                                                source_date=f_data.get('source_date', ''),
                                                source_url=f_data.get('source_url', ''),
                                                source_type=f_data.get('source_type', 'other'),
                                                confidence=f_data.get('confidence', 'medium'),
                                                relevance=f_data.get('relevance', ''),
                                            )
                                            _executor.add_finding(block_id, finding)
                                            imported += 1
                                        except Exception as fe:
                                            st.warning(f"Skipped finding in {block_id}: {fe}")

                                if imported > 0:
                                    st.success(f"Imported **{imported}** findings across "
                                               f"**{len([k for k in findings_data if k in _executor.blocks])}** blocks, "
                                               f"**{summaries_imported}** summaries")
                                else:
                                    st.warning(f"No findings imported. "
                                               f"JSON keys: {json_keys}  \n"
                                               f"Executor blocks: {executor_keys}  \n"
                                               f"Make sure the block_id values in the JSON match exactly.")

                                if skipped_blocks:
                                    st.info(f"Skipped blocks not in executor: {skipped_blocks}")

                        except Exception as e:
                            st.error(f"Import error: {e}")
                    else:
                        st.warning("Paste JSON findings first.")

            st.markdown("---")

            # ---- Add findings form ----
            st.markdown("#### Add Finding")
            st.caption("Record a research finding with full citation. Select the target block, "
                       "enter the claim and source details.")

            if not _executor.blocks:
                st.info("Initialise blocks first using the button above.")
            else:
                block_options = list(_executor.blocks.keys())
                af_block = st.selectbox("Target block", block_options, key="af_block")

                af_col1, af_col2 = st.columns(2)
                with af_col1:
                    af_claim = st.text_area("Factual claim", key="af_claim",
                                            placeholder="e.g., San Jose metro area median household income is $143,000")
                    af_value = st.text_input("Value (if numeric)", key="af_value",
                                             placeholder="e.g., $143,000")
                    af_unit = st.text_input("Unit", key="af_unit", placeholder="e.g., USD, visitors, %")
                    af_year = st.text_input("Data year", key="af_year", placeholder="e.g., 2023")
                    af_relevance = st.text_input("Relevance to route case", key="af_relevance",
                                                  placeholder="e.g., Premium fare demand")

                with af_col2:
                    st.markdown("**Citation details:**")
                    af_source = st.text_input("Source name", key="af_source",
                                              placeholder="e.g., US Census Bureau")
                    af_title = st.text_input("Article/page title", key="af_title",
                                             placeholder="e.g., ACS 2023")
                    af_date = st.text_input("Publication date", key="af_date",
                                            placeholder="e.g., 2024 or 2024-03-15")
                    af_url = st.text_input("URL", key="af_url",
                                           placeholder="https://...")
                    af_stype = st.selectbox("Source type",
                                            ['government', 'industry', 'news', 'academic',
                                             'airline', 'international', 'other'],
                                            key="af_stype")
                    af_conf = st.selectbox("Confidence", ['high', 'medium', 'low'], key="af_conf")

                if st.button("Record Finding", key="btn_add_finding"):
                    if af_claim.strip():
                        finding = make_finding(
                            claim=af_claim.strip(),
                            value=af_value.strip(),
                            unit=af_unit.strip(),
                            year=af_year.strip(),
                            source_name=af_source.strip(),
                            source_title=af_title.strip(),
                            source_date=af_date.strip(),
                            source_url=af_url.strip(),
                            source_type=af_stype,
                            confidence=af_conf,
                            relevance=af_relevance.strip(),
                        )
                        _executor.add_finding(af_block, finding)
                        st.success(f"Finding added to **{af_block}** "
                                   f"(now {len(_executor.blocks[af_block].findings)} findings)")
                    else:
                        st.warning("Enter a claim before recording.")

            st.markdown("---")

            # ---- Add data gap ----
            st.markdown("#### Record Data Gap")
            if _executor.blocks:
                gap_col1, gap_col2 = st.columns([2, 1])
                with gap_col1:
                    gap_block = st.selectbox("Block", list(_executor.blocks.keys()), key="gap_block")
                    gap_desc = st.text_input("Gap description", key="gap_desc",
                                             placeholder="e.g., US-UK bilateral air service agreement terms not found")
                with gap_col2:
                    st.markdown("")
                    st.markdown("")
                    if st.button("Record Gap", key="btn_gap"):
                        if gap_desc.strip():
                            _executor.add_data_gap(gap_block, gap_desc.strip())
                            st.success(f"Gap recorded in **{gap_block}**")

            st.markdown("---")

            # ---- Set block summary ----
            st.markdown("#### Set Block Summary")
            if _executor.blocks:
                sum_block = st.selectbox("Block", list(_executor.blocks.keys()), key="sum_block")
                sum_text = st.text_area("Summary (2-3 sentences)", key="sum_text",
                                        placeholder="Brief summary of key findings for this block...")
                sum_pres = st.text_area("Presentation-ready text", key="sum_pres",
                                        placeholder="Polished text suitable for a presentation slide...")
                if st.button("Save Summary", key="btn_sum"):
                    _executor.set_block_summary(sum_block, sum_text.strip(), sum_pres.strip())
                    st.success(f"Summary saved for **{sum_block}**")

            st.markdown("---")

            # ---- Current findings review ----
            st.markdown("#### Findings Review")
            for block_id, bf in _executor.blocks.items():
                finding_count = len(bf.findings)
                gap_count = len(bf.data_gaps)
                status_label = "COMPLETE" if finding_count > 0 and gap_count == 0 else (
                    "PARTIAL" if finding_count > 0 else "EMPTY")

                with st.expander(
                    f"[{status_label}] {bf.block_name} ({bf.relevance}) -- "
                    f"{finding_count} findings, {gap_count} gaps",
                    expanded=False
                ):
                    if bf.summary:
                        st.markdown(f"**Summary:** {bf.summary}")
                    if bf.presentation_text:
                        st.markdown(f"**Presentation text:** _{bf.presentation_text}_")

                    if bf.findings:
                        for i, f in enumerate(bf.findings, 1):
                            val_str = f" -- **{f.value}** {f.unit}" if f.value else ""
                            st.markdown(f"**{i}.** {f.claim}{val_str}")
                            for c in f.citations:
                                parts = [f"*{c.source_name}*"]
                                if c.title:
                                    parts.append(f'"{c.title}"')
                                if c.date:
                                    parts.append(f"({c.date})")
                                if c.url:
                                    parts.append(f"[link]({c.url})")
                                if c.confidence != "high":
                                    parts.append(f"[{c.confidence}]")
                                st.caption(f"  Source: {', '.join(parts)}")
                            if f.is_single_source:
                                st.caption("  * Single source -- verify independently")
                            if f.relevance_to_case:
                                st.caption(f"  -> {f.relevance_to_case}")
                    else:
                        st.caption("No findings recorded yet.")

                    if bf.data_gaps:
                        st.markdown("**Data gaps:**")
                        for gap in bf.data_gaps:
                            st.markdown(f"- DATA REQUIRED: {gap}")

                    # Block-level execution prompt
                    with st.expander("Execution prompt for this block"):
                        prompt = _executor.get_execution_prompt(block_id)
                        st.code(prompt, language="text")

            st.markdown("---")

            # ---- Finalise & generate outputs ----
            st.markdown("#### Finalise & Generate Outputs")

            fin_col1, fin_col2 = st.columns(2)
            with fin_col1:
                fin_summary = st.text_area(
                    "Executive summary",
                    key="fin_summary",
                    placeholder="Brief executive summary of overall research findings...",
                )
            with fin_col2:
                st.markdown("**Key statistics** (one per line, format: `Label = Value`)")
                fin_stats_raw = st.text_area(
                    "Key stats",
                    key="fin_stats",
                    placeholder="Median household income (SJC) = $143,000\nUK GDP growth = 1.1%",
                )

            if st.button("Finalise Research", type="primary", key="btn_finalise"):
                # Parse key stats
                key_stats = {}
                if fin_stats_raw.strip():
                    for line in fin_stats_raw.strip().split('\n'):
                        if '=' in line:
                            k, v = line.split('=', 1)
                            key_stats[k.strip()] = v.strip()

                mr_output = _executor.finalise(
                    executive_summary=fin_summary.strip(),
                    key_statistics=key_stats,
                )
                st.session_state['mr_output'] = mr_output
                st.success(f"Research finalised -- Quality score: **{mr_output.data_quality_score:.0f}/100** | "
                           f"Citations: {mr_output.total_citations} | Gaps: {mr_output.total_data_gaps}")

            # ---- Download outputs ----
            if 'mr_output' in st.session_state and st.session_state['mr_output'] is not None:
                mr_output = st.session_state['mr_output']

                st.markdown("---")
                st.markdown("#### Download Research Outputs")

                qc1, qc2, qc3, qc4 = st.columns(4)
                qc1.metric("Quality Score", f"{mr_output.data_quality_score:.0f}/100")
                qc2.metric("Citations", mr_output.total_citations)
                qc3.metric("Data Gaps", mr_output.total_data_gaps)
                qc4.metric("Blocks", len(mr_output.blocks))

                dl1, dl2, dl3 = st.columns(3)

                with dl1:
                    md_report = generate_markdown_report(mr_output)
                    st.download_button(
                        "Markdown Report",
                        data=md_report,
                        file_name=f"Research_{mr_output.route}_{mr_output.execution_date}.md",
                        mime="text/markdown",
                        key="dl_mr_md",
                    )

                with dl2:
                    json_report = generate_json_output(mr_output)
                    st.download_button(
                        "JSON (pipeline)",
                        data=json_report,
                        file_name=f"Research_{mr_output.route}_{mr_output.execution_date}.json",
                        mime="application/json",
                        key="dl_mr_json",
                    )

                with dl3:
                    try:
                        excel_path = os.path.join(
                            tempfile.gettempdir(),
                            f"Research_{mr_output.route}_{mr_output.execution_date}.xlsx"
                        )
                        generate_excel_output(mr_output, excel_path)
                        with open(excel_path, 'rb') as ef:
                            excel_bytes = ef.read()
                        st.download_button(
                            "Excel Workbook",
                            data=excel_bytes,
                            file_name=f"Research_{mr_output.route}_{mr_output.execution_date}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_mr_xlsx",
                        )
                    except Exception as e:
                        st.warning(f"Excel generation failed: {e}")

                # Show markdown preview
                with st.expander("Markdown Report Preview", expanded=False):
                    st.markdown(md_report)


# ============================================================================
# TAB 11: COMPARISON
# ============================================================================

with tab11:
    st.markdown("### Analyst Comparison")

    results = st.session_state.results
    if results is None:
        st.info("Run the pipeline first, then compare against analyst results here.")
    else:
        st.markdown(f"**Pipeline:** {results['grand_total']:,} pax, {results['load_factor']:.1%} LF")
        st.markdown("---")

        comp_mode = st.radio(
            "Comparison Mode",
            ["Manual Entry", "Workbook Upload"],
            horizontal=True,
            key="comp_mode",
            help="Manual: enter headline numbers. Workbook: upload analyst's forecast file "
                 "for city-by-city comparison and calibration learning extraction."
        )

        # ========== MODE 1: MANUAL ENTRY (existing) ==========
        if comp_mode == "Manual Entry":
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Analyst Results:**")
                at = st.number_input("Total Passengers", min_value=0, value=0, step=1000, key="at")
                ap = st.number_input("P2P Passengers", min_value=0, value=0, step=1000, key="ap")
                ah = st.number_input("Cnx Home", min_value=0, value=0, step=1000, key="ah")
                ad = st.number_input("Cnx Dest", min_value=0, value=0, step=100, key="ad")
                al = st.number_input("Load Factor (%)", min_value=0.0, max_value=100.0,
                                     value=0.0, step=0.1, key="al")

            with c2:
                st.markdown("**Variance:**")
                if at > 0:
                    vt = (results['grand_total'] - at) / at
                    within_5 = abs(vt) <= 0.05

                    if within_5:
                        st.markdown('<span class="verdict-pass">WITHIN 5% - TOOL TRUSTED</span>',
                                    unsafe_allow_html=True)
                    elif abs(vt) <= 0.10:
                        st.markdown('<span class="verdict-marginal">WITHIN 10% - REVIEW CALIBRATION</span>',
                                    unsafe_allow_html=True)
                    else:
                        st.markdown('<span class="verdict-fail">EXCEEDS 10% - INVESTIGATE</span>',
                                    unsafe_allow_html=True)

                    st.markdown("---")

                    comparisons = [
                        ("Grand Total", results['grand_total'], at, vt),
                        ("P2P", results['p2p_total'], ap,
                         (results['p2p_total'] - ap) / ap if ap > 0 else None),
                        ("Cnx Home", results['home_total'], ah,
                         (results['home_total'] - ah) / ah if ah > 0 else None),
                        ("Cnx Dest", results['dest_total'], ad,
                         (results['dest_total'] - ad) / ad if ad > 0 else None),
                    ]

                    for label, pv, av, var in comparisons:
                        if var is not None:
                            color = ("#155724" if abs(var) <= 0.05
                                     else ("#856404" if abs(var) <= 0.10 else "#721c24"))
                            diff = pv - av
                            st.markdown(
                                f"""**{label}** &nbsp; Pipeline: {pv:,} &nbsp;|&nbsp; Analyst: {av:,}
                                &nbsp;|&nbsp; <span style="color:{color};font-weight:700">{var:+.1%} ({diff:+,})</span>""",
                                unsafe_allow_html=True)

                    if al > 0:
                        lf_diff = results['load_factor'] * 100 - al
                        st.markdown(
                            f"""**Load Factor** &nbsp; Pipeline: {results['load_factor']:.1%}
                            &nbsp;|&nbsp; Analyst: {al:.1f}% &nbsp;|&nbsp; {lf_diff:+.1f} pp""",
                            unsafe_allow_html=True)

                    st.markdown("---")
                    if within_5:
                        st.success("Pipeline and analyst results are well-aligned. Calibration is sound.")
                    else:
                        st.warning(f"Total variance: {vt:+.1%}. Review P2P capture rates, "
                                   f"connecting city QSI calibration, and base demand inputs.")
                else:
                    st.caption("Enter analyst total passengers to see comparison.")

        # ========== MODE 2: WORKBOOK UPLOAD (learning extraction) ==========
        else:
            if not HAS_LEARNING:
                st.warning("learning_comparison module not available. "
                           "Place learning_comparison.py alongside the portal.")
            else:
                st.markdown("Upload the analyst's completed forecast workbook. The system will parse "
                            "city-by-city connecting captures, compare QSI values, and extract "
                            "calibration learnings the PCE can use for future routes.")

                analyst_upload = st.file_uploader(
                    "Analyst Forecast Workbook",
                    type=["xlsx", "xlsm", "xls"],
                    key="analyst_benchmark_upload",
                    help="Standard Avia forecast workbook with connecting city sheets"
                )

                if analyst_upload:
                    if st.button("Run Learning Comparison", key="btn_learning_compare"):
                        with st.spinner("Parsing analyst workbook and extracting learnings..."):
                            try:
                                tmp = tempfile.mktemp(
                                    suffix=os.path.splitext(analyst_upload.name)[1])
                                with open(tmp, 'wb') as f:
                                    f.write(analyst_upload.getbuffer())

                                cfg = st.session_state.get('last_config')
                                report = run_learning_comparison(results, tmp, cfg)
                                st.session_state.learning_report = report

                                st.success(f"Comparison complete: {len(report.city_comparisons)} cities, "
                                           f"{len(report.learnings)} learnings extracted")
                            except Exception as e:
                                st.error(f"Failed to parse workbook: {str(e)}")
                                with st.expander("Traceback"):
                                    st.code(traceback.format_exc())

                # Display learning report
                report = st.session_state.get('learning_report')
                if report:
                    # Headline comparison
                    st.markdown("#### Headline Comparison")
                    hl = report.headline
                    if hl:
                        hl_rows = [
                            {"Metric": "Grand Total",
                             "Pipeline": f"{hl.pipe_total:,.0f}",
                             "Analyst": f"{hl.analyst_total:,.0f}",
                             "Variance": f"{hl.total_variance:+.1%}"},
                            {"Metric": "P2P",
                             "Pipeline": f"{hl.pipe_p2p:,.0f}",
                             "Analyst": f"{hl.analyst_p2p:,.0f}",
                             "Variance": f"{hl.p2p_variance:+.1%}" if hl.p2p_variance is not None else "-"},
                            {"Metric": "Cnx Home",
                             "Pipeline": f"{hl.pipe_home:,.0f}",
                             "Analyst": f"{hl.analyst_home:,.0f}",
                             "Variance": f"{hl.home_variance:+.1%}" if hl.home_variance is not None else "-"},
                            {"Metric": "Cnx Dest",
                             "Pipeline": f"{hl.pipe_dest:,.0f}",
                             "Analyst": f"{hl.analyst_dest:,.0f}",
                             "Variance": f"{hl.dest_variance:+.1%}" if hl.dest_variance is not None else "-"},
                        ]
                        st.dataframe(hl_rows, use_container_width=True, hide_index=True)

                        # Verdict
                        if hl.total_variance is not None:
                            var_abs = abs(hl.total_variance)
                            if var_abs <= 0.05:
                                st.success("Within 5%: pipeline and analyst well-aligned")
                            elif var_abs <= 0.10:
                                st.warning("Within 10%: review calibration parameters")
                            else:
                                st.error(f"Variance {hl.total_variance:+.1%}: significant divergence")

                    # City-by-city comparison
                    st.markdown("---")
                    st.markdown("#### City-by-City QSI Comparison")
                    if report.city_comparisons:
                        city_rows = []
                        for cc in sorted(report.city_comparisons,
                                         key=lambda x: abs(x.forecast_variance or 0),
                                         reverse=True)[:30]:
                            city_rows.append({
                                'City': cc.city_code,
                                'Side': cc.side,
                                'Pipeline QSI': f"{cc.pipe_qsi_capture:.4f}" if cc.pipe_qsi_capture else "-",
                                'Analyst QSI': f"{cc.analyst_qsi_capture:.4f}" if cc.analyst_qsi_capture else "-",
                                'QSI Ratio': f"{cc.qsi_ratio:.2f}" if cc.qsi_ratio else "-",
                                'Pipe Pax': f"{cc.pipe_forecast:,.0f}" if cc.pipe_forecast else "-",
                                'Analyst Pax': f"{cc.analyst_forecast:,.0f}" if cc.analyst_forecast else "-",
                                'Pax Var': f"{cc.forecast_variance:+.1%}" if cc.forecast_variance is not None else "-",
                                'Severity': cc.severity if cc.severity else "-",
                            })
                        st.dataframe(city_rows, use_container_width=True, hide_index=True)
                        st.caption(f"Showing top 30 by variance magnitude "
                                   f"(of {len(report.city_comparisons)} total)")

                    # Statistics
                    with st.expander("QSI Ratio Statistics", expanded=False):
                        sc1, sc2, sc3, sc4 = st.columns(4)
                        with sc1:
                            st.metric("Median QSI Ratio (Home)",
                                      f"{report.median_qsi_ratio_home:.3f}")
                        with sc2:
                            st.metric("Median QSI Ratio (Dest)",
                                      f"{report.median_qsi_ratio_dest:.3f}")
                        with sc3:
                            st.metric("Implied QSI Adj",
                                      f"{report.implied_qsi_adjustment:.3f}")
                        with sc4:
                            st.metric("Over / Under / Match",
                                      f"{report.cities_overestimated} / "
                                      f"{report.cities_underestimated} / "
                                      f"{report.cities_matched}")

                    # Extracted learnings
                    st.markdown("---")
                    st.markdown("#### Calibration Learnings")
                    if report.learnings:
                        for i, learning in enumerate(report.learnings, 1):
                            with st.expander(f"Learning {i}: {learning.learning_type} "
                                             f"-- {learning.parameter} ({learning.direction})",
                                             expanded=(i <= 3)):
                                st.markdown(f"**{learning.parameter}**: pipeline={learning.pipe_value:.4f}, "
                                            f"analyst={learning.analyst_value:.4f} "
                                            f"(magnitude: {learning.magnitude:.2f})")
                                if learning.city_code:
                                    st.markdown(f"City: {learning.city_code} ({learning.side})")
                                if learning.recommendation:
                                    st.caption(f"Recommendation: {learning.recommendation}")
                    else:
                        st.caption("No specific learnings extracted (pipeline may already "
                                   "be well-calibrated for this route).")

                    # Summary
                    if report.summary_lines:
                        with st.expander("Comparison Summary", expanded=False):
                            for line in report.summary_lines:
                                st.markdown(f"- {line}")


# ============================================================================
# TAB 12: CROSS-ROUTE VALIDATION
# ============================================================================

with tab12:
    st.markdown("###  Cross-Route Validation Suite")
    st.markdown("Run the built-in validation cases against 6 historical routes. "
                 "Tests forecast math, capacity logic, connecting city aggregation, "
                 "and cross-route pattern consistency.")

    if st.button(" Run Full Validation Suite", key="btn_validate"):
        with st.spinner("Running validation across 6 routes..."):
            try:
                # run_all_validations returns (results_list, cases_list) or just list
                raw = run_all_validations()
                if isinstance(raw, tuple):
                    all_results, val_cases = raw
                else:
                    all_results = raw
                    val_cases = [case_ba_lhr_sjc(), case_ke_icn_sjc_7x(), case_ke_icn_sjc_5x(),
                                 case_sq_sin_sjc(), case_cx_hkg_sjc(), case_fi_kef_sjc()]

                st.session_state.validation_results = all_results

                # Generate Excel workbook
                try:
                    tmpf = tempfile.mktemp(suffix='.xlsx')
                    write_validation_workbook(all_results, val_cases, tmpf)
                    with open(tmpf, 'rb') as f:
                        st.session_state.validation_bytes = f.read()
                except Exception:
                    st.session_state.validation_bytes = None

                passed = sum(1 for r in all_results if r.status == 'PASS')
                total = len(all_results)
                if passed == total:
                    st.success(f"All {total} tests passed  ")
                else:
                    st.warning(f"{passed}/{total} tests passed, {total-passed} issues found")
            except Exception as e:
                st.error(f"Validation failed: {e}")
                with st.expander("Traceback"):
                    st.code(traceback.format_exc())

    vresults = st.session_state.validation_results
    if vresults is not None:
        # Summary table
        st.markdown("##### Test Results")
        vrows = []
        for vr in vresults:
            icon = '' if vr.status == 'PASS' else ('' if vr.status == 'WARN' else '')
            vrows.append({
                'Test': vr.test_name,
                'Status': f"{icon} {vr.status}",
                'Detail': vr.detail[:120] if vr.detail else '',
            })
        st.dataframe(pd.DataFrame(vrows), use_container_width=True, hide_index=True)

        # Detailed results in expanders
        for vr in vresults:
            if vr.status != 'PASS' or vr.detail:
                with st.expander(f"{vr.test_name}  {vr.status}"):
                    if vr.detail:
                        st.text(vr.detail)
                    if hasattr(vr, 'items') and vr.items:
                        for item in vr.items:
                            st.markdown(f"- {item}")

        # Download
        if st.session_state.validation_bytes:
            st.download_button(" Download Validation Report (Excel)",
                                data=st.session_state.validation_bytes,
                                file_name="CrossRoute_Validation_Report.xlsx",
                                type="primary",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_validation")


# ============================================================================
# TAB 13: TIME GRID (New Route path only) -- LIVE PIPELINE RUNS
# ============================================================================

if tab13 is not None:
    with tab13:
        st.markdown("### \u23f0 Departure Time Grid Search")
        st.caption("Run the full QSI pipeline at multiple departure times to identify "
                    "the optimal schedule. Each scenario shifts the proposed service timing, "
                    "rebuilds hub connections, and re-runs the complete engine.")

        if st.session_state.data_source != 'newroute':
            st.info("Time grid analysis is available in New Route Assessment mode. "
                     "Switch to 'New Route Assessment' in the sidebar.")
        elif not st.session_state.results:
            st.info("\u261d Run the pipeline once from the Data Upload tab to establish "
                     "the base case before running time grid analysis.")
        else:
            st.markdown("#### \U0001f4cb Grid Configuration")

            # Grid definition controls
            col_g1, col_g2, col_g3 = st.columns(3)
            with col_g1:
                grid_start = st.time_input("Grid start (dep from origin)",
                                           value=dtime(14, 0), key='grid_start')
                grid_end = st.time_input("Grid end (dep from origin)",
                                         value=dtime(23, 0), key='grid_end')
            with col_g2:
                grid_step = st.selectbox("Step interval", [30, 60, 90, 120], index=1,
                                         format_func=lambda x: f"{x} minutes",
                                         key='grid_step')
                n_scenarios = st.number_input("Max scenarios", min_value=2, max_value=12,
                                              value=6, key='grid_max')
            with col_g3:
                ft_out = st.number_input("Block time outbound (min)",
                                         min_value=60, max_value=1200, value=625,
                                         step=5, key='ft_out',
                                         help="Flying time origin to destination in minutes")
                ft_ret = st.number_input("Block time return (min)",
                                         min_value=60, max_value=1200, value=660,
                                         step=5, key='ft_ret',
                                         help="Flying time destination to origin in minutes")

            # Calculate preview of scenarios
            import datetime as dt_module
            scenarios_preview = []
            current = dt_module.datetime.combine(dt_module.date.today(), grid_start)
            end_dt = dt_module.datetime.combine(dt_module.date.today(), grid_end)
            while current <= end_dt and len(scenarios_preview) < n_scenarios:
                dep_hhmm = current.strftime("%H%M")
                dep_mins = int(dep_hhmm[:2]) * 60 + int(dep_hhmm[2:])
                arr_mins = (dep_mins + ft_out) % 1440
                arr_h, arr_m = divmod(arr_mins, 60)
                scenarios_preview.append({
                    'dep': current.strftime("%H:%M"),
                    'dep_hhmm': dep_hhmm,
                    'arr': f"{arr_h:02d}:{arr_m:02d}",
                })
                current += dt_module.timedelta(minutes=grid_step)

            if scenarios_preview:
                times_str = ', '.join(f"{s['dep']}\u2192{s['arr']}" for s in scenarios_preview)
                st.markdown(f"**{len(scenarios_preview)} scenarios:** {times_str}")
                est_secs = len(scenarios_preview) * 45
                st.caption(f"Estimated runtime: {est_secs // 60}m {est_secs % 60}s "
                           f"({len(scenarios_preview)} \u00d7 ~45s per scenario)")

            st.markdown("---")

            # Dest QSI file for bidirectional grid (optional)
            st.markdown("**Destination QSI Template** (optional, for bidirectional analysis)")
            dest_qsi_grid_upload = st.file_uploader(
                "Dest QSI for grid", type=['xlsx', 'xlsm', 'xls'],
                key='dest_qsi_grid', label_visibility="collapsed",
                help="If provided, each scenario also shifts connections at the destination hub. "
                     "Use the dest-perspective QSI template file.")

            st.markdown("---")

            # RUN GRID BUTTON
            can_run = (len(scenarios_preview) >= 2 and
                       st.session_state.results is not None)

            if st.button("\U0001f50d Run Time Grid Search", type="primary",
                        disabled=not can_run, use_container_width=True):

                grid_progress = st.progress(0)
                grid_status = st.empty()
                grid_log_area = st.empty()
                grid_log_lines = []

                def grid_callback(idx, total, dep_time, status):
                    pct = (idx + (1 if status in ('done', 'error') else 0)) / total
                    grid_progress.progress(min(pct, 1.0))
                    icon = "\u2705" if status == 'done' else ("\u274c" if status == 'error' else "\u23f3")
                    msg = f"{icon} Scenario {idx+1}/{total}: {dep_time} dep -- {status}"
                    grid_log_lines.append(msg)
                    grid_status.text(msg)

                try:
                    grid_status.text("Preparing time grid search...")

                    # We need to recreate the provider from the uploaded files
                    # (session state has the file paths from the original run)
                    # But the simpler approach: we check if the config was stored

                    # Recreate SingleExtractOAGProvider from uploaded template
                    qsi_template_path_grid = save_uploaded(qsi_template_upload)

                    mct_file_map_grid = {}
                    if mct_uploads:
                        for mf in mct_uploads:
                            mct_path = save_uploaded(mf)
                            name_upper = mf.name.upper()
                            for apt_code in KNOWN_AIRPORTS:
                                if apt_code in name_upper:
                                    mct_file_map_grid[apt_code] = mct_path
                                    break
                            else:
                                mct_file_map_grid[mf.name] = mct_path

                    city_lookup_path_grid = None
                    if city_lookup_upload:
                        city_lookup_path_grid = save_uploaded(city_lookup_upload)

                    base_provider = SingleExtractOAGProvider(
                        qsi_file=qsi_template_path_grid,
                        origin_airport=origin,
                        dest_airport=destination,
                        proposed_carrier=carrier_code or 'XX',
                        use_city_codes=True,
                        city_lookup_file=city_lookup_path_grid,
                        mct_files=mct_file_map_grid,
                    )

                    # Dest provider (optional)
                    dest_base_provider = None
                    dest_qsi_grid_path = None
                    if dest_qsi_grid_upload:
                        dest_qsi_grid_path = save_uploaded(dest_qsi_grid_upload)
                        dest_base_provider = SingleExtractOAGProvider(
                            qsi_file=dest_qsi_grid_path,
                            origin_airport=destination,
                            dest_airport=origin,
                            proposed_carrier=carrier_code or 'XX',
                            use_city_codes=True,
                            city_lookup_file=city_lookup_path_grid,
                            mct_files=mct_file_map_grid,
                        )

                    # Build RouteConfig for grid
                    grid_cfg = RouteConfig()
                    grid_cfg.airline_name = carrier_name or carrier_code
                    grid_cfg.airline_code = carrier_code
                    grid_cfg.home_airport_code = origin
                    grid_cfg.dest_airport_code = destination
                    o_info = lookup_airport(origin)
                    d_info = lookup_airport(destination)
                    grid_cfg.home_city_code = o_info[0] if o_info else origin
                    grid_cfg.dest_city_code = d_info[0] if d_info else destination
                    grid_cfg.frequency = frequency
                    grid_cfg.aircraft_type = aircraft_type
                    grid_cfg.seats = seats
                    grid_cfg.qsi_ceiling = qsi_ceiling
                    grid_cfg.qsi_adjustment = 1.0
                    grid_cfg.online_coeff = online_coeff
                    grid_cfg.alliance_coeff = alliance_coeff
                    grid_cfg.interline_coeff = interline_coeff
                    grid_cfg.et_decay_factor = 0.8
                    grid_cfg.et_decay_interval = 0.1
                    grid_cfg.target_total = 0
                    grid_cfg.target_p2p = 0
                    grid_cfg.target_cnx_home = 0
                    grid_cfg.target_cnx_dest = 0
                    grid_cfg.target_load_factor = 0.0

                    # Build demand provider (same as main run)
                    p2p_config_grid = {'segments': []}
                    for seg in st.session_state.p2p_segments:
                        seg_dict = {
                            'name': seg['name'], 'base_demand': seg['base_demand'],
                            'growth_rate': seg['growth'], 'seasonality': 1.0,
                            'stimulation': seg['stimulation'], 'capture_rate': seg['capture_rate'],
                        }
                        if seg['subsegments']:
                            seg_dict['subsegments'] = [{
                                'name': s['name'], 'base_demand': s['base_demand'],
                                'growth_rate': s['growth'], 'seasonality': 1.0,
                                'stimulation': s['stimulation'], 'capture_rate': s['capture_rate'],
                            } for s in seg['subsegments']]
                        p2p_config_grid['segments'].append(seg_dict)

                    # Re-save MIDT files
                    home_cnx_paths_g = save_multiple_uploads(home_cnx_uploads) if home_cnx_uploads else []
                    dest_cnx_paths_g = save_multiple_uploads(dest_cnx_uploads) if dest_cnx_uploads else []
                    p2p_paths_g = save_multiple_uploads(p2p_uploads) if p2p_uploads else []

                    grid_cfg.demand_provider = MIDTDemandProvider(
                        home_cnx_files=home_cnx_paths_g,
                        dest_cnx_files=dest_cnx_paths_g,
                        p2p_files=p2p_paths_g,
                        p2p_config=p2p_config_grid,
                        home_growth=home_growth,
                        dest_growth=dest_growth,
                        catchment_share_home=catchment_share_home,
                        catchment_share_dest=catchment_share_dest,
                        city_lookup_file=city_lookup_path_grid,
                        min_demand_threshold=min_demand_threshold,
                    )

                    grid_cfg.schedule_provider = base_provider

                    # Create TimeGridRunner
                    runner = TimeGridRunner(
                        base_config=grid_cfg,
                        base_provider=base_provider,
                        origin=origin,
                        destination=destination,
                        carrier=carrier_code or 'XX',
                        flying_time_outbound=ft_out,
                        flying_time_return=ft_ret,
                        dest_base_provider=dest_base_provider,
                        callback=grid_callback,
                    )

                    # Generate scenarios
                    start_str = grid_start.strftime("%H:%M")
                    end_str = grid_end.strftime("%H:%M")
                    grid_scenarios = runner.generate_grid(
                        start=start_str,
                        end=end_str,
                        step_minutes=grid_step,
                        max_scenarios=n_scenarios,
                    )

                    grid_status.text(f"Running {len(grid_scenarios)} scenarios...")

                    # Capture stdout
                    old_stdout = sys.stdout
                    sys.stdout = io.StringIO()
                    try:
                        grid_result = runner.run(grid_scenarios)
                    finally:
                        sys.stdout = old_stdout

                    grid_progress.progress(1.0)
                    grid_status.text(f"Grid search complete: {len(grid_result.scenarios)} scenarios "
                                     f"in {grid_result.total_elapsed:.0f}s")

                    # Write Excel output
                    import tempfile as tf_mod
                    grid_xlsx_path = tf_mod.mktemp(suffix='.xlsx')
                    write_grid_output(grid_result, grid_xlsx_path)
                    with open(grid_xlsx_path, 'rb') as gf:
                        st.session_state.grid_output_bytes = gf.read()

                    st.session_state.grid_search_result = grid_result
                    st.session_state.grid_results = [
                        {
                            'dep_time': s.dep_time_origin,
                            'arr_hub': s.arr_time_hub,
                            'dep_hub': s.dep_time_hub,
                            'arr_origin': s.arr_time_origin,
                            'grand_total': s.grand_total,
                            'p2p': round(s.p2p_total),
                            'cnx_home': round(s.cnx_home_total),
                            'cnx_dest': round(s.cnx_dest_total),
                            'load_factor': s.load_factor,
                            'n_cities_home': s.n_cities_home,
                            'n_cities_dest': s.n_cities_dest,
                            'itineraries_qsi1': s.itineraries_qsi1,
                            'itineraries_qsi2': s.itineraries_qsi2,
                            'elapsed': s.elapsed_seconds,
                            'error': s.error,
                        }
                        for s in grid_result.scenarios
                    ]

                    if grid_result.best_scenario:
                        best = grid_result.best_scenario
                        st.success(
                            f"\u2705 **Optimal: {best.dep_time_origin} departure** "
                            f"\u2192 {best.grand_total:,} pax, {best.load_factor:.1%} LF"
                        )

                except Exception as e:
                    grid_progress.progress(1.0)
                    grid_status.text(f"Grid search failed: {str(e)}")
                    st.error(f"Time grid search failed: {str(e)}")
                    with st.expander("Full traceback"):
                        st.code(traceback.format_exc())

                # Show log
                if grid_log_lines:
                    with st.expander("Grid Search Log", expanded=False):
                        st.code('\n'.join(grid_log_lines), language='text')

            # ============================================================
            # DISPLAY GRID RESULTS
            # ============================================================
            grid_result_obj = st.session_state.grid_search_result
            grid_data = st.session_state.grid_results

            if grid_result_obj and grid_data:
                st.markdown("---")
                st.markdown("#### \U0001f3c6 Grid Search Results")

                # Summary metrics
                best = grid_result_obj.best_scenario
                if best and len(grid_result_obj.ranked) >= 2:
                    worst = grid_result_obj.ranked[-1]
                    delta = best.grand_total - worst.grand_total
                    pct_range = (delta / worst.grand_total * 100) if worst.grand_total else 0

                    gm1, gm2, gm3, gm4 = st.columns(4)
                    with gm1:
                        st.markdown(metric_card("Best Time",
                                    best.dep_time_origin, f"dep {origin}"),
                                    unsafe_allow_html=True)
                    with gm2:
                        st.markdown(metric_card("Best Total",
                                    f"{best.grand_total:,}", f"{best.load_factor:.1%} LF"),
                                    unsafe_allow_html=True)
                    with gm3:
                        st.markdown(metric_card("Worst Total",
                                    f"{worst.grand_total:,}", worst.dep_time_origin),
                                    unsafe_allow_html=True)
                    with gm4:
                        st.markdown(metric_card("Range",
                                    f"{delta:+,}", f"{pct_range:+.1f}%"),
                                    unsafe_allow_html=True)

                # Ranked table
                st.markdown("##### Scenario Ranking")
                rows = []
                for rank, gr in enumerate(sorted(grid_data,
                                                  key=lambda x: -(x.get('grand_total', 0)
                                                                    if not x.get('error') else 0)),
                                           1):
                    if gr.get('error'):
                        rows.append({
                            'Rank': rank,
                            f'Dep {origin}': gr.get('dep_time', ''),
                            f'Arr {destination}': gr.get('arr_hub', ''),
                            f'Dep {destination}': gr.get('dep_hub', ''),
                            f'Arr {origin}': gr.get('arr_origin', ''),
                            'Grand Total': 'ERROR',
                            'LF': '',
                            'P2P': '',
                            'Cnx Home': '',
                            'Cnx Dest': '',
                            'Time (s)': f"{gr.get('elapsed', 0):.0f}",
                        })
                    else:
                        best_total = best.grand_total if best else 0
                        delta_v = gr.get('grand_total', 0) - best_total
                        rows.append({
                            'Rank': rank,
                            f'Dep {origin}': gr.get('dep_time', ''),
                            f'Arr {destination}': gr.get('arr_hub', ''),
                            f'Dep {destination}': gr.get('dep_hub', ''),
                            f'Arr {origin}': gr.get('arr_origin', ''),
                            'Grand Total': f"{gr.get('grand_total', 0):,}",
                            'LF': f"{gr.get('load_factor', 0):.1%}",
                            'P2P': f"{gr.get('p2p', 0):,}",
                            'Cnx Home': f"{gr.get('cnx_home', 0):,}",
                            'Cnx Dest': f"{gr.get('cnx_dest', 0):,}",
                            'Time (s)': f"{gr.get('elapsed', 0):.0f}",
                        })
                st.dataframe(rows, use_container_width=True, hide_index=True)

                # BAR CHART: Grand total by departure time
                st.markdown("##### Total Passengers by Departure Time")
                chart_data = [
                    {'dep_time': gr['dep_time'], 'Grand Total': gr['grand_total']}
                    for gr in grid_data if not gr.get('error')
                ]
                if chart_data:
                    import pandas as pd
                    df_chart = pd.DataFrame(chart_data)
                    st.bar_chart(df_chart.set_index('dep_time'), y='Grand Total',
                                 color='#e8a83e', use_container_width=True)

                # STACKED BAR: P2P / Cnx Home / Cnx Dest breakdown
                st.markdown("##### Traffic Breakdown by Departure Time")
                breakdown_data = [
                    {
                        'dep_time': gr['dep_time'],
                        'P2P': gr.get('p2p', 0),
                        'Cnx Home': gr.get('cnx_home', 0),
                        'Cnx Dest': gr.get('cnx_dest', 0),
                    }
                    for gr in grid_data if not gr.get('error')
                ]
                if breakdown_data:
                    df_bd = pd.DataFrame(breakdown_data)
                    st.bar_chart(df_bd.set_index('dep_time'),
                                 color=['#0d3b66', '#e8a83e', '#8ba4c4'],
                                 use_container_width=True)

                # Itinerary counts
                st.markdown("##### Connection Window Impact")
                st.caption("Number of valid itineraries at each departure time "
                           "-- shows how the connection window shifts.")
                itin_data = [
                    {
                        'dep_time': gr['dep_time'],
                        'QSI1 Itineraries': gr.get('itineraries_qsi1', 0),
                        'QSI2 Itineraries': gr.get('itineraries_qsi2', 0),
                    }
                    for gr in grid_data if not gr.get('error')
                ]
                if itin_data:
                    df_itin = pd.DataFrame(itin_data)
                    st.bar_chart(df_itin.set_index('dep_time'),
                                 color=['#0a1628', '#336699'],
                                 use_container_width=True)

                # City sensitivity analysis
                if grid_result_obj.city_sensitivity:
                    st.markdown("##### Top City Sensitivity")
                    st.caption("Cities most affected by departure time changes "
                               "-- ordered by absolute passenger range.")
                    sens_rows = []
                    for cs in grid_result_obj.city_sensitivity[:20]:
                        row_s = {
                            'City': cs.get('city', ''),
                            'Best Time': cs.get('best_time', ''),
                            'Max Pax': f"{cs.get('max_pax', 0):,.0f}",
                            'Min Pax': f"{cs.get('min_pax', 0):,.0f}",
                            'Range': f"{cs.get('range', 0):,.0f}",
                            'Range %': f"{cs.get('range_pct', 0):.1f}%",
                        }
                        sens_rows.append(row_s)
                    st.dataframe(sens_rows, use_container_width=True, hide_index=True)

                # Download buttons
                st.markdown("---")
                dl1, dl2, _ = st.columns([1, 1, 2])
                with dl1:
                    if st.session_state.grid_output_bytes:
                        fn = (f"TimeGrid_{origin}_{destination}_{carrier_code}_"
                              f"{datetime.now():%Y%m%d}.xlsx")
                        st.download_button(
                            "\U0001f4e5 Download Grid Workbook",
                            data=st.session_state.grid_output_bytes,
                            file_name=fn, type="primary",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with dl2:
                    if st.session_state.output_bytes:
                        fn2 = (f"QSI_Forecast_{origin}_{destination}_{carrier_code}_"
                               f"{datetime.now():%Y%m%d}.xlsx")
                        st.download_button(
                            "\U0001f4e5 Download Base Case Workbook",
                            data=st.session_state.output_bytes,
                            file_name=fn2,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================================
# TAB 14: PREDICTIVE CALIBRATION ENGINE
# ============================================================================

with tab14:
    st.markdown('<div class="section-label">Predictive Calibration Engine</div>',
                unsafe_allow_html=True)
    st.markdown("""
    Auto-suggests calibration parameters (stimulation, capture, QSI adjustment, growth rate)
    for a proposed route using the 22-case calibration library and 8 discovered rules.
    Run this **before** the full pipeline to get starting parameter values.
    """)

    # --- Market characterisation inputs ---
    st.markdown("#### Route & Market Characterisation")

    pc_col1, pc_col2, pc_col3 = st.columns(3)

    with pc_col1:
        pc_origin = st.text_input("Origin Airport", value=origin if origin else "",
                                   max_chars=3, key="pc_origin",
                                   help="IATA code for origin/hub airport").upper().strip()
        pc_destination = st.text_input("Destination Airport", value=destination if destination else "",
                                        max_chars=3, key="pc_dest",
                                        help="IATA code for destination airport").upper().strip()
        pc_carrier = st.text_input("Carrier Code", value=carrier_code if carrier_code else "",
                                    max_chars=2, key="pc_carrier").upper().strip()
        pc_carrier_name = st.text_input("Carrier Name", value=carrier_name if carrier_name else "",
                                         key="pc_carrier_name")

    with pc_col2:
        pc_hub_status = st.selectbox("Hub Status",
                                      ["Major Hub", "Secondary Hub", "Non-Hub"],
                                      index=0, key="pc_hub_status",
                                      help="Major Hub = top-tier connecting hub (LHR, CDG, ICN). "
                                           "Secondary = regional hub (DUB, KEF). Non-Hub = P2P only.")
        pc_alliance = st.selectbox("Alliance",
                                    ["OneWorld", "Star Alliance", "SkyTeam", "None"],
                                    index=0, key="pc_alliance")
        pc_carrier_type = st.selectbox("Carrier Type",
                                        ["Full Service", "LCC", "Ultra-LCC", "Hybrid", "Charter"],
                                        index=0, key="pc_carrier_type")
        pc_frequency = st.number_input("Frequency (per week)", min_value=1, max_value=21,
                                        value=5, key="pc_freq")

    with pc_col3:
        pc_seats = st.number_input("Seats per flight", min_value=50, max_value=600,
                                    value=280, key="pc_seats")
        pc_aircraft = st.text_input("Aircraft Type", value="", key="pc_aircraft",
                                     placeholder="B787-9")
        pc_business_split = st.slider("Business / Leisure Split (P2P)",
                                       min_value=0.10, max_value=0.90, value=0.55,
                                       step=0.05, key="pc_biz_split",
                                       help="Fraction of P2P demand that is business travel")

    # --- Key market booleans ---
    st.markdown("#### Market Maturity Inputs")
    st.markdown("""
    *These booleans drive the maturity classification and are the most important inputs.
    They require market knowledge  does the hub currently have any direct service to the
    destination metro area? Is there a nearby airport with direct service?*
    """)

    mb_col1, mb_col2 = st.columns(2)
    with mb_col1:
        pc_new_route = st.checkbox("New route (not currently operated)", value=True,
                                    key="pc_new_route",
                                    help="Tick if this carrier does not currently fly this pair")
        pc_hub_to_metro = st.checkbox("Hub has direct service to destination metro",
                                       value=True, key="pc_hub_metro",
                                       help="E.g., LH FRA-SFO exists when assessing LH FRA-SJC. "
                                            "Untick for Rule 1 (zero hub-metro)  highest stimulation.")
    with mb_col2:
        pc_nearby_direct = st.checkbox("Nearby airport has direct service on this pair",
                                        value=False, key="pc_nearby",
                                        help="E.g., SFO has EU nonstops when assessing SJC routes")
        pc_catchment_overlap = st.checkbox("Nearby airport shares same catchment",
                                            value=True, key="pc_catchment",
                                            help="If nearby airport catches the same passenger base "
                                                 "(overlap = lower stimulation). Untick for distinct catchment "
                                                 "(SJC vs SFO = distinct).")

    # --- Run button ---
    st.markdown("---")
    run_calib = st.button(" Generate Calibration Suggestion", type="primary",
                           use_container_width=True, key="run_calib_btn")

    if run_calib:
        if not pc_origin or not pc_destination:
            st.error("Please enter origin and destination airports.")
        else:
            with st.spinner("Running predictive calibration engine..."):
                try:
                    profile = RouteProfile(
                        origin=pc_origin,
                        destination=pc_destination,
                        carrier=pc_carrier,
                        carrier_name=pc_carrier_name,
                        alliance=pc_alliance,
                        carrier_type=pc_carrier_type,
                        hub_airport=pc_origin,
                        hub_status=pc_hub_status,
                        frequency=pc_frequency,
                        aircraft=pc_aircraft,
                        seats=pc_seats,
                        new_route=pc_new_route,
                        existing_service_hub_to_metro=pc_hub_to_metro,
                        nearby_airport_direct=pc_nearby_direct,
                        catchment_overlap=pc_catchment_overlap,
                        business_leisure_split=pc_business_split,
                    )
                    engine = PredictiveCalibrationEngine()
                    suggestion = engine.predict(profile)
                    st.session_state.calib_suggestion = suggestion

                    # Generate Excel
                    import tempfile as _tf
                    tmp = _tf.NamedTemporaryFile(suffix='.xlsx', delete=False)
                    write_suggestion_excel(suggestion, tmp.name)
                    with open(tmp.name, 'rb') as f:
                        st.session_state.calib_excel_bytes = f.read()
                    os.unlink(tmp.name)

                    st.success(f"Calibration suggestion generated for {suggestion.route_label}")
                except Exception as ex:
                    st.error(f"Calibration engine error: {ex}")
                    st.code(traceback.format_exc())

    # --- Display results ---
    suggestion = st.session_state.calib_suggestion
    if suggestion:
        # --- Confidence & maturity banner ---
        conf_color = {"HIGH": "pass", "MEDIUM": "marginal", "LOW": "fail"}.get(
            suggestion.confidence, "marginal")
        st.markdown(f"""
        <div style="display:flex;gap:1rem;margin:1rem 0;">
            <div class="verdict-{conf_color}" style="flex:1;text-align:center;">
                Confidence: {suggestion.confidence}</div>
            <div class="verdict-pass" style="flex:1;text-align:center;">
                Market Maturity: {suggestion.market_maturity}</div>
        </div>
        """, unsafe_allow_html=True)

        # --- Key metrics ---
        kc1, kc2, kc3, kc4, kc5 = st.columns(5)
        with kc1:
            st.markdown(metric_card("Blended Stimulation",
                                     f"{suggestion.blended_stimulation:.3f}"), unsafe_allow_html=True)
        with kc2:
            st.markdown(metric_card("Blended Capture",
                                     f"{suggestion.blended_capture:.1%}"), unsafe_allow_html=True)
        with kc3:
            st.markdown(metric_card("QSI Adjustment",
                                     f"{suggestion.qsi_adjustment:.2f}",
                                     "1.0 = new route"), unsafe_allow_html=True)
        with kc4:
            st.markdown(metric_card("Growth Rate",
                                     f"{suggestion.growth_rate:.1%}"), unsafe_allow_html=True)
        with kc5:
            cnx_lo, cnx_hi = suggestion.cnx_share_range
            st.markdown(metric_card("Cnx Share Range",
                                     f"{cnx_lo:.0%}{cnx_hi:.0%}"), unsafe_allow_html=True)

        # --- Segment detail table ---
        st.markdown("#### Segment-Level Predictions")
        seg_rows = []
        for seg_name, seg in suggestion.segments.items():
            seg_rows.append({
                'Segment': seg_name.replace('_', ' ').title(),
                'Stimulation': f"{seg.stimulation:.2f}",
                'Stim Range': f"{seg.stim_range[0]:.2f}{seg.stim_range[1]:.2f}",
                'Capture': f"{seg.capture:.1%}",
                'Cap Range': f"{seg.capture_range[0]:.1%}{seg.capture_range[1]:.1%}",
                'Confidence': seg.confidence,
            })
        if seg_rows:
            seg_df = pd.DataFrame(seg_rows)
            st.dataframe(seg_df, use_container_width=True, hide_index=True)

        # --- Reasoning panels ---
        with st.expander(" QSI Adjustment Reasoning"):
            st.write(suggestion.qsi_adj_reasoning or "Standard new-route QSI adjustment = 1.0")

        with st.expander(" Growth Rate Reasoning"):
            st.write(suggestion.growth_reasoning or "Default growth rate applied")

        with st.expander(" Connecting Share Reasoning"):
            st.write(suggestion.cnx_share_reasoning or "Range based on hub status and comparable cases")

        # --- Rules applied ---
        if suggestion.rules_applied:
            with st.expander(" Rules Applied", expanded=True):
                for rule in suggestion.rules_applied:
                    st.markdown(f"- {rule}")

        # --- Comparable cases ---
        if suggestion.comparable_cases:
            with st.expander(" Comparable Cases from Library"):
                comp_rows = []
                for comp in suggestion.comparable_cases:
                    comp_rows.append({
                        'Route': comp.get('route_id', ''),
                        'Relevance Score': comp.get('score', 0),
                        'Match Factors': comp.get('relevance', ''),
                    })
                st.dataframe(comp_rows, use_container_width=True, hide_index=True)

        # --- Warnings ---
        if suggestion.warnings:
            with st.expander(" Warnings", expanded=True):
                for w in suggestion.warnings:
                    st.warning(w)

        # --- Segment-level reasoning ---
        with st.expander(" Detailed Segment Reasoning"):
            for seg_name, seg in suggestion.segments.items():
                st.markdown(f"**{seg_name.replace('_', ' ').title()}**")
                st.markdown(f"- Stimulation: {seg.stim_reasoning}")
                st.markdown(f"- Capture: {seg.capture_reasoning}")
                st.markdown("")

        # --- Download ---
        st.markdown("---")
        dl_c1, dl_c2, _ = st.columns([1, 1, 2])
        with dl_c1:
            if st.session_state.calib_excel_bytes:
                fn = (f"Calibration_{pc_origin}_{pc_destination}_{pc_carrier}_"
                      f"{datetime.now():%Y%m%d}.xlsx")
                st.download_button(
                    " Download Calibration Workbook",
                    data=st.session_state.calib_excel_bytes,
                    file_name=fn, type="primary",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_calib_xlsx")
        with dl_c2:
            # JSON export
            json_data = json.dumps(suggestion.to_dict(), indent=2, default=str)
            st.download_button(
                " Download as JSON",
                data=json_data,
                file_name=f"Calibration_{pc_origin}_{pc_destination}_{pc_carrier}.json",
                mime="application/json",
                key="dl_calib_json")

    else:
        st.info("Configure the route characterisation above and click **Generate Calibration Suggestion** "
                "to get auto-suggested parameters before running the full pipeline.")



# ============================================================================
# TAB 15: FARE ALLOCATION
# ============================================================================

with tab15:
    st.markdown("### Fare Allocation")
    st.markdown("Upload Sabre fare extract files to compute per-market revenue using actual booking fares "
                "instead of single-average estimates.")

    results_fa = st.session_state.results

    # --- File Upload ---
    st.markdown("#### Upload Fare Files")
    st.markdown("Two files cover both directions. Each should contain an "
                "**>>OUTPUTS<<** sheet with DIRECT and INDIRECT sections.")

    fc1, fc2 = st.columns(2)
    with fc1:
        home_fare_file = st.file_uploader(
            f"Home direction ({origin or 'ORG'} -> {destination or 'DST'} -> XXX)",
            type=['xlsx', 'xls'], key='fare_home_upload',
            help="Sabre fare data for traffic originating at home hub")
    with fc2:
        dest_fare_file = st.file_uploader(
            f"Dest direction ({destination or 'DST'} -> {origin or 'ORG'} -> XXX)",
            type=['xlsx', 'xls'], key='fare_dest_upload',
            help="Sabre fare data for traffic originating at destination")

    fwc1, fwc2 = st.columns(2)
    with fwc1:
        fa_fw = st.number_input("Fare Weight (P2P)", min_value=0.50, max_value=1.00,
                                value=0.85, step=0.05, format="%.2f", key="fa_fw_p2p",
                                help="Standard Avia discount from raw Sabre fares")
    with fwc2:
        fa_fw_cnx = st.number_input("Fare Weight (Connecting)", min_value=0.50, max_value=1.00,
                                    value=0.85, step=0.05, format="%.2f", key="fa_fw_cnx",
                                    help="Set to 1.0 if connecting fares already net")

    if home_fare_file or dest_fare_file:
        if st.button("Parse & Compute Revenue", key="fare_parse_btn", type="primary"):
            with st.spinner("Parsing Sabre fare files..."):
                try:
                    fa_alloc = FareAllocator(fare_weight=fa_fw)
                    if home_fare_file:
                        fa_alloc.add_home_file(save_uploaded(home_fare_file))
                    if dest_fare_file:
                        fa_alloc.add_dest_file(save_uploaded(dest_fare_file))
                    fa_alloc.build()
                    st.session_state.fare_allocator = fa_alloc

                    if results_fa is not None:
                        fa_mr = fa_alloc.compute_market_revenue(
                            results_fa, fare_weight=fa_fw, fare_weight_cnx=fa_fw_cnx)
                        st.session_state.fare_market_revenue = fa_mr

                        route_lbl = f"{origin or ''}-{destination or ''}"
                        carrier_lbl = carrier_code if carrier_code else '??'
                        fa_writer = FareAllocationWorkbook(fa_alloc, fa_mr,
                                                           route_label=route_lbl, carrier=carrier_lbl)
                        fa_writer.write_all()
                        fa_buf = io.BytesIO()
                        fa_writer.save(fa_buf)
                        st.session_state.fare_excel_bytes = fa_buf.getvalue()

                    st.success(f"Parsed: {len(fa_alloc.cnx_home_fares)} cnx home, "
                               f"{len(fa_alloc.cnx_dest_fares)} cnx dest, "
                               f"{len(fa_alloc.p2p_fares)} P2P markets")
                except Exception as e:
                    st.error(f"Parse error: {e}")
                    st.code(traceback.format_exc())

    # --- Display ---
    fa_alloc_display = st.session_state.fare_allocator
    fa_mr_display = st.session_state.fare_market_revenue

    if fa_alloc_display is not None:
        st.markdown("---")
        st.markdown("#### Parsing Summary")
        fps1, fps2, fps3 = st.columns(3)
        with fps1:
            st.markdown(metric_card("Cnx @ Home", str(len(fa_alloc_display.cnx_home_fares)),
                        f"Beyond {origin or 'ORG'}"), unsafe_allow_html=True)
        with fps2:
            st.markdown(metric_card("Cnx @ Dest", str(len(fa_alloc_display.cnx_dest_fares)),
                        f"Beyond {destination or 'DST'}"), unsafe_allow_html=True)
        with fps3:
            st.markdown(metric_card("P2P Reference", str(len(fa_alloc_display.p2p_fares)),
                        "Direct traffic"), unsafe_allow_html=True)

        st.markdown("#### Weighted-Average Fares (Fallback)")
        st.markdown("Pax-weighted averages used when a market has no specific Sabre data.")
        ff_rows = []
        for lbl, fares in [("Cnx Home", fa_alloc_display.fallback_cnx_home),
                            ("Cnx Dest", fa_alloc_display.fallback_cnx_dest),
                            ("P2P", fa_alloc_display.fallback_p2p)]:
            if fares:
                ff_rows.append({'Segment': lbl,
                    'Y': f"${fares.get('Y', 0):,.0f}", 'PY': f"${fares.get('PY', 0):,.0f}",
                    'J': f"${fares.get('J', 0):,.0f}",
                    'F': f"${fares.get('F', 0):,.0f}" if fares.get('F') else '-'})
        if ff_rows:
            st.dataframe(ff_rows, use_container_width=True, hide_index=True)

        cs = fa_alloc_display.fallback_cabin_split
        if cs:
            st.markdown(f"**Cabin split:** Y={cs.get('Y',0):.1%}, PY={cs.get('PY',0):.1%}, "
                        f"J={cs.get('J',0):.1%}, F={cs.get('F',0):.1%}")

    if fa_mr_display is not None:
        st.markdown("---")
        st.markdown("#### Per-Market Revenue Results")
        fa_s = fa_mr_display['summary']

        fms1, fms2, fms3, fms4 = st.columns(4)
        with fms1:
            st.markdown(metric_card("Total Pax", f"{fa_s['grand_total_pax']:,.0f}"),
                        unsafe_allow_html=True)
        with fms2:
            st.markdown(metric_card("Total Revenue", f"${fa_s['grand_total_revenue']/1e6:.1f}M"),
                        unsafe_allow_html=True)
        with fms3:
            st.markdown(metric_card("Avg OW Fare", f"${fa_s['overall_avg_fare']:,.0f}"),
                        unsafe_allow_html=True)
        with fms4:
            n_spec = fa_s['markets_with_specific_fares']
            n_fb = fa_s['markets_using_fallback']
            st.markdown(metric_card("Specific Fares", f"{n_spec}/{n_spec+n_fb}",
                        "markets with data"), unsafe_allow_html=True)

        st.markdown("#### Revenue by Segment")
        seg_rows = []
        for sk, sl in [('p2p', 'Point-to-Point'),
                        ('cnx_home', f'Connecting @ {origin or "Home"}'),
                        ('cnx_dest', f'Connecting @ {destination or "Dest"}')]:
            seg = fa_mr_display.get(sk, {})
            seg_rows.append({
                'Segment': sl, 'Passengers': f"{seg.get('total_pax',0):,.0f}",
                'Revenue': f"${seg.get('total_revenue',0):,.0f}",
                'Avg Fare': f"${seg.get('avg_fare',0):,.0f}",
                'Share': f"{seg.get('total_revenue',0)/fa_s['grand_total_revenue']*100:.1f}%"
                         if fa_s['grand_total_revenue'] > 0 else '-'})
        seg_rows.append({
            'Segment': 'TOTAL', 'Passengers': f"{fa_s['grand_total_pax']:,.0f}",
            'Revenue': f"${fa_s['grand_total_revenue']:,.0f}",
            'Avg Fare': f"${fa_s['overall_avg_fare']:,.0f}", 'Share': '100.0%'})
        st.dataframe(seg_rows, use_container_width=True, hide_index=True)

        # Per-market detail
        for sk, sl in [('cnx_home', f'Connecting @ {origin or "Home"}'),
                        ('cnx_dest', f'Connecting @ {destination or "Dest"}')]:
            seg = fa_mr_display.get(sk, {})
            mkts = seg.get('markets', [])
            if not mkts:
                continue
            st.markdown("---")
            st.markdown(f"#### {sl} Detail")
            st.caption(f"{seg.get('markets_with_fares',0)} specific, "
                       f"{seg.get('markets_using_fallback',0)} fallback")
            mkt_rows = []
            for m in mkts[:30]:
                bc = m.get('by_cabin', {})
                mkt_rows.append({
                    'City': m['city'], 'Name': m.get('name',''),
                    'Pax': f"{m['pax']:,.0f}", 'Revenue': f"${m['revenue']:,.0f}",
                    'Avg Fare': f"${m['avg_fare']:,.0f}",
                    'Y': f"${bc['Y']['fare_ow']:,.0f}" if 'Y' in bc else '-',
                    'J': f"${bc['J']['fare_ow']:,.0f}" if 'J' in bc else '-',
                    'Source': m.get('fare_source','')})
            st.dataframe(mkt_rows, use_container_width=True, hide_index=True)
            if len(mkts) > 30:
                st.caption(f"Top 30 of {len(mkts)}. Download Excel for full list.")

        st.markdown("---")
        if st.session_state.fare_excel_bytes:
            st.download_button("Download Fare Allocation Workbook",
                data=st.session_state.fare_excel_bytes,
                file_name=f"Fare_Allocation_{origin or 'X'}_{destination or 'X'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="fare_dl_btn")

        if fa_alloc_display is not None:
            with st.expander("Audit Trail"):
                for line in fa_alloc_display.audit:
                    st.text(line)

    elif fa_alloc_display is None:
        st.info("Upload Sabre fare files above to begin. Run the pipeline first to combine with forecast data.")


# ============================================================================
# TAB 16: PRESENTATION GENERATOR
# ============================================================================

with tab16:
    st.markdown("###  City Pair Presentation Generator")
    st.markdown("Generate a branded Avia Solutions PPTX presentation from your pipeline results and research.")

    # Auto-populate from route config and research
    _cfg = st.session_state.get('last_config')
    _exec = st.session_state.get('mr_executor')
    _result = st.session_state.get('results') or st.session_state.get('last_result')

    # City/country lookup for auto-populate
    _city_map = {
        'LHR': 'London', 'CDG': 'Paris', 'AMS': 'Amsterdam', 'FRA': 'Frankfurt',
        'TPA': 'Tampa', 'SJC': 'San Jose', 'SFO': 'San Francisco', 'LAX': 'Los Angeles',
        'JFK': 'New York', 'ORD': 'Chicago', 'DXB': 'Dubai', 'SIN': 'Singapore',
        'HKG': 'Hong Kong', 'ICN': 'Seoul', 'NRT': 'Tokyo', 'TPE': 'Taipei',
        'MAN': 'Manchester', 'DUB': 'Dublin', 'MUC': 'Munich', 'MAD': 'Madrid',
    }
    _country_map = {
        'LHR': 'United Kingdom', 'CDG': 'France', 'AMS': 'Netherlands', 'FRA': 'Germany',
        'TPA': 'United States', 'SJC': 'United States', 'SFO': 'United States',
        'LAX': 'United States', 'JFK': 'United States', 'MAN': 'United Kingdom',
        'DUB': 'Ireland', 'DXB': 'UAE', 'SIN': 'Singapore', 'HKG': 'China',
        'ICN': 'South Korea', 'NRT': 'Japan', 'TPE': 'Taiwan',
    }

    _o = getattr(_cfg, 'home_airport_code', '') if _cfg else ''
    _d = getattr(_cfg, 'dest_airport_code', '') if _cfg else ''

    # Pre-set session state keys BEFORE widgets render (only if not already set by user)
    # This is the Streamlit pattern for auto-populating keyed widgets
    _pptx_defaults = {
        'pptx_airline': getattr(_cfg, 'airline_name', '') if _cfg else '',
        'pptx_airline_code': getattr(_cfg, 'airline_code', '') if _cfg else '',
        'pptx_origin': _o,
        'pptx_dest': _d,
        'pptx_origin_city': _city_map.get(_o, ''),
        'pptx_dest_city': _city_map.get(_d, ''),
        'pptx_origin_country': _country_map.get(_o, ''),
        'pptx_dest_country': _country_map.get(_d, ''),
        'pptx_freq': getattr(_cfg, 'frequency', 7) if _cfg else 7,
        'pptx_aircraft': getattr(_cfg, 'aircraft_type', '') if _cfg else '',
        'pptx_seats': getattr(_cfg, 'seats', 275) if _cfg else 275,
    }
    for k, v in _pptx_defaults.items():
        if k not in st.session_state and v:
            st.session_state[k] = v

    # Also pre-set Why This Route cards from research
    if _exec and _exec.blocks:
        _why_auto = [
            ("why_text_0", "tourism", ""),       # Market Size <- tourism
            ("why_text_1", "corporate_links", ""), # Strategic Fit <- corporate
            ("why_text_2", "non_cannibalization", ""), # Competitive Gap <- non-cannibalisation
            ("why_text_3", "economic_context", ""),    # Economic Drivers <- economic
        ]
        for key, block_id, _ in _why_auto:
            if key not in st.session_state:
                blk = _exec.blocks.get(block_id)
                if blk and blk.summary:
                    st.session_state[key] = blk.summary[:200]
                elif blk and blk.findings:
                    st.session_state[key] = blk.findings[0].claim[:200]

    # Show auto-populate status
    auto_sources = []
    if _cfg:
        auto_sources.append("Route Config")
    if _exec and any(len(bf.findings) > 0 for bf in _exec.blocks.values()):
        auto_sources.append("Research Findings")
    if _result:
        auto_sources.append("Forecast Results")
    if auto_sources:
        st.success(f"Auto-populated from: **{', '.join(auto_sources)}**")
    else:
        st.info("Run a forecast and/or research first to auto-populate. "
                "Or enter details manually below.")

    pptx_col1, pptx_col2 = st.columns(2)

    with pptx_col1:
        st.markdown("**Route & Client Details**")
        pptx_airline = st.text_input("Airline Name",
            value=getattr(_cfg, 'airline_name', '') if _cfg else '', key="pptx_airline")
        pptx_airline_code = st.text_input("Airline Code",
            value=getattr(_cfg, 'airline_code', '') if _cfg else '', key="pptx_airline_code")
        pptx_origin = st.text_input("Origin Airport (IATA)", value=_o, key="pptx_origin")
        pptx_dest = st.text_input("Destination Airport (IATA)", value=_d, key="pptx_dest")
        pptx_origin_city = st.text_input("Origin City",
            value=_city_map.get(_o, ''), key="pptx_origin_city")
        pptx_dest_city = st.text_input("Destination City",
            value=_city_map.get(_d, ''), key="pptx_dest_city")
        pptx_origin_country = st.text_input("Origin Country",
            value=_country_map.get(_o, ''), key="pptx_origin_country")
        pptx_dest_country = st.text_input("Destination Country",
            value=_country_map.get(_d, ''), key="pptx_dest_country")
        pptx_client = st.text_input("Client Name (for branding)", value="", key="pptx_client")

    with pptx_col2:
        st.markdown("**Presentation Settings**")
        pptx_buyer = st.selectbox("Buyer Type", [
            "Airport pitch", "Airline forecast", "Fund due diligence"
        ], key="pptx_buyer")
        pptx_demand = st.selectbox("Primary Demand Driver", [
            "Business", "Leisure", "Mixed", "VFR-diaspora"
        ], key="pptx_demand")
        pptx_headline = st.text_area("Cover Headline (leave blank for auto)",
            value="", height=68, key="pptx_headline")
        _def_freq = getattr(_cfg, 'frequency', 7) if _cfg else 7
        pptx_frequency = st.number_input("Frequency (flights/week)",
            value=_def_freq, min_value=1, max_value=28, key="pptx_freq")
        _def_ac = getattr(_cfg, 'aircraft_type', '') if _cfg else ''
        pptx_aircraft = st.text_input("Aircraft Type", value=_def_ac, key="pptx_aircraft")
        _def_seats = getattr(_cfg, 'seats', 275) if _cfg else 275
        pptx_seats = st.number_input("Seats per flight",
            value=_def_seats, min_value=50, max_value=600, key="pptx_seats")

        st.markdown("**Schedule**")
        pptx_sc1, pptx_sc2 = st.columns(2)
        with pptx_sc1:
            pptx_out_dep = st.text_input("Outbound Dep.", value="", key="pptx_out_dep")
            pptx_ret_dep = st.text_input("Return Dep.", value="", key="pptx_ret_dep")
        with pptx_sc2:
            pptx_out_arr = st.text_input("Outbound Arr.", value="", key="pptx_out_arr")
            pptx_ret_arr = st.text_input("Return Arr.", value="", key="pptx_ret_arr")

    st.markdown("---")

    # Image upload section
    st.markdown("**Background Images** (optional — presentation uses solid navy backgrounds by default)")
    pptx_img_col1, pptx_img_col2, pptx_img_col3 = st.columns(3)
    with pptx_img_col1:
        cover_img_file = st.file_uploader("Cover slide image", type=["jpg", "jpeg", "png"], key="pptx_cover_img")
    with pptx_img_col2:
        closing_img_file = st.file_uploader("Closing slide image", type=["jpg", "jpeg", "png"], key="pptx_closing_img")
    with pptx_img_col3:
        contents_img_file = st.file_uploader("Contents page image", type=["jpg", "jpeg", "png"], key="pptx_contents_img")

    # Image suggestions
    _img_suggestions = {
        "Business": "Tech campuses, city skylines, corporate HQ buildings",
        "Leisure": "Landscapes, beaches, tourist landmarks, national parks",
        "Mixed": "City panoramas with natural scenery, modern architecture",
        "VFR-diaspora": "Cultural landmarks, heritage sites, community spaces",
    }
    st.info(f"**Image suggestions for {pptx_demand} route:** {_img_suggestions.get(pptx_demand, '')}")

    st.markdown("---")

    # Why points editor — auto-populate from research if available
    st.markdown("**'Why This Route?' Key Points** (4 cards)")
    _why_defaults = [
        {"title": "Market Size", "text": ""},
        {"title": "Strategic Fit", "text": ""},
        {"title": "Competitive Gap", "text": ""},
        {"title": "Economic Drivers", "text": ""},
    ]

    # Try to auto-fill from research findings
    if _exec and _exec.blocks:
        econ = _exec.blocks.get("economic_context")
        if econ and econ.findings:
            _why_defaults[3]["text"] = econ.findings[0].claim if econ.findings else ""
        corp = _exec.blocks.get("corporate_links")
        if corp and corp.findings:
            _why_defaults[1]["text"] = corp.findings[0].claim if corp.findings else ""
        tourism = _exec.blocks.get("tourism")
        if tourism and tourism.findings:
            _why_defaults[0]["text"] = tourism.findings[0].claim if tourism.findings else ""
        noncan = _exec.blocks.get("non_cannibalization")
        if noncan and noncan.findings:
            _why_defaults[2]["text"] = noncan.findings[0].claim if noncan.findings else ""

    why_points_data = []
    for i in range(4):
        wcol1, wcol2 = st.columns([1, 3])
        with wcol1:
            wt = st.text_input(f"Title {i+1}", value=_why_defaults[i]["title"], key=f"why_title_{i}")
        with wcol2:
            wtx = st.text_input(f"Text {i+1}", value=_why_defaults[i]["text"], key=f"why_text_{i}")
        why_points_data.append({"title": wt, "text": wtx if wtx else "DATA REQUIRED"})

    st.markdown("---")

    # Generate button
    pptx_ready = bool(pptx_airline and pptx_origin and pptx_dest)

    if not pptx_ready:
        st.warning("Enter at least airline name, origin and destination to generate.")

    if st.button("Generate Presentation", disabled=not pptx_ready, type="primary", key="pptx_generate"):
        with st.spinner("Generating branded PPTX presentation..."):
            try:
                from city_pair_pptx_generator import generate_presentation, extract_research_for_pptx, HAS_PPTX as _has_pptx

                if not _has_pptx:
                    st.error("python-pptx is required. Install with: `pip install python-pptx`")
                else:
                    # Build config
                    pptx_config = {
                        "airline_name": pptx_airline,
                        "airline_code": pptx_airline_code,
                        "origin": pptx_origin.upper(),
                        "destination": pptx_dest.upper(),
                        "origin_city": pptx_origin_city,
                        "dest_city": pptx_dest_city,
                        "origin_country": pptx_origin_country,
                        "dest_country": pptx_dest_country,
                        "frequency": pptx_frequency,
                        "aircraft_type": pptx_aircraft,
                        "seats": pptx_seats,
                        "demand_driver": pptx_demand,
                        "buyer_type": pptx_buyer.lower().replace(" ", "_"),
                        "client_name": pptx_client,
                        "date": datetime.now().strftime("%B %Y"),
                        "headline": pptx_headline.strip() if pptx_headline.strip() else "",
                        "why_points": why_points_data,
                    }

                    # Back-fill any empty "Why This Route" cards from research
                    if _exec and _exec.blocks:
                        _why_fill_map = [
                            (0, "tourism", "Market Size"),
                            (1, "corporate_links", "Strategic Fit"),
                            (2, "non_cannibalization", "Competitive Gap"),
                            (3, "economic_context", "Economic Drivers"),
                        ]
                        for idx, block_id, title in _why_fill_map:
                            if idx < len(pptx_config["why_points"]):
                                wp = pptx_config["why_points"][idx]
                                if wp.get("text", "").strip() in ("", "DATA REQUIRED"):
                                    blk = _exec.blocks.get(block_id)
                                    if blk and blk.summary:
                                        wp["text"] = blk.summary[:250]
                                    elif blk and blk.findings:
                                        wp["text"] = blk.findings[0].claim[:250]

                    # Pull forecast data from session state
                    result = st.session_state.get("results") or st.session_state.get("last_result")
                    if result and isinstance(result, dict):
                        fc = {
                            "base_year": "2024",
                            "forecast_year": "2026",
                            "grand_total": result.get('grand_total', 0),
                            "p2p_total": result.get('p2p_total', 0),
                            "cnx_home_total": result.get('home_total', 0),
                            "cnx_dest_total": result.get('dest_total', 0),
                            "load_factor": result.get('load_factor', 0),
                        }
                        pptx_config["forecast"] = fc
                    elif result and hasattr(result, 'grand_total'):
                        fc = {
                            "base_year": "2024",
                            "forecast_year": "2026",
                            "grand_total": result.grand_total,
                            "p2p_total": getattr(result, 'p2p_total', 0),
                            "cnx_home_total": getattr(result, 'home_total', 0),
                            "cnx_dest_total": getattr(result, 'dest_total', 0),
                            "load_factor": getattr(result, 'load_factor', 0),
                        }
                        pptx_config["forecast"] = fc

                    # Extract research findings from executor
                    research_blocks = {}
                    if _exec and _exec.blocks:
                        research_blocks = extract_research_for_pptx(_exec)

                    # Extract assumptions from config
                    if _cfg:
                        pptx_config["assumptions"] = {
                            "qsi_adjustment": getattr(_cfg, 'qsi_adjustment', 1.0),
                            "qsi_ceiling": getattr(_cfg, 'qsi_ceiling', 1.0),
                        }

                    # Extract connecting cities from results
                    if result and isinstance(result, dict):
                        dest_results = result.get('dest_results', [])
                        if dest_results:
                            cnx_list = []
                            if isinstance(dest_results, dict):
                                for city, pax in sorted(dest_results.items(),
                                                         key=lambda x: x[1], reverse=True)[:20]:
                                    cnx_list.append({"city": city, "pax": pax})
                            elif isinstance(dest_results, list):
                                # List of tuples/dicts: [(city, pax), ...] or [{"city":..., "pax":...}]
                                sorted_results = []
                                for item in dest_results:
                                    if isinstance(item, dict):
                                        sorted_results.append((item.get('city', item.get('label', '')),
                                                                item.get('pax', item.get('forecast', 0))))
                                    elif isinstance(item, (list, tuple)) and len(item) >= 2:
                                        sorted_results.append((str(item[0]), float(item[1])))
                                sorted_results.sort(key=lambda x: x[1], reverse=True)
                                for city, pax in sorted_results[:20]:
                                    cnx_list.append({"city": city, "pax": pax})
                            if cnx_list:
                                pptx_config["connecting_cities"] = cnx_list

                    # Generate
                    output_filename = f"{pptx_origin.upper()}_{pptx_dest.upper()}_{pptx_airline.replace(' ', '_')}.pptx"
                    output_path = os.path.join(tempfile.mkdtemp(), output_filename)

                    generate_presentation(pptx_config, research_blocks, output_path)

                    st.success(f"Presentation generated: **{output_filename}**")

                    with open(output_path, "rb") as pf:
                        pptx_bytes = pf.read()

                    st.download_button(
                        label=f"Download {output_filename}",
                        data=pptx_bytes,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    )

                    with st.expander("View generated config JSON"):
                        st.json(pptx_config)

            except ImportError as ie:
                st.error(f"Missing dependency: {ie}. Install with: `pip install python-pptx`")
            except Exception as e:
                st.error(f"Presentation generation failed: {e}")
                import traceback
                st.code(traceback.format_exc())


# ============================================================================
# FOOTER
# ============================================================================

st.markdown("---")
st.markdown("""
<div style="text-align:center;color:#8ba4c4;font-size:0.75rem;padding:1rem;">
Avia Solutions -- QSI Route Forecast Portal v11.0<br>
32 modules | 16 tabs | Pre-computed + New Route Assessment paths<br>
Presentation Generator + Fare Allocation + Market Research Executor<br>
Predictive Calibration Engine + Q&A Checklist + Spill Analysis<br>
Seasonality + Revenue + Assumptions Log + Business Case + Output Workbook<br>
Cross-Route Validation + Departure Time Grid<br>
Validated: BA LHR-SJC, KLM AMS-TPA, KE ICN-SJC, SQ SIN-SJC, CX HKG-SJC, FI KEF-SJC<br>
<br>
<em>This tool is provided for internal use by Avia Solutions Limited and its authorised clients.
Forecasts are indicative and subject to professional review. All data remains confidential.</em>
</div>
""", unsafe_allow_html=True)
