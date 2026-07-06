#!/usr/bin/env python3
"""
Avia Solutions  SingleExtractOAGProvider (Chat 24 continuation)
================================================================
ScheduleProvider that reads OAG leg data directly from a QSI template file
(e.g., QSILHR_v1_OS_JZ_17Feb15.xlsx) and builds connections across all
auto-detected hubs.

This is the pragmatic middle ground between:
  - ExcelScheduleProvider (reads pre-computed QSI sheets  no rebuild capability)
  - RawOAGProvider (needs raw OAG exports per hub  requires multiple files)

The QSI template already contains leg data for ALL hubs in its Leg 1.1/2.1/1.2/2.2
sheets  the analyst pasted these in from OAG Analyser. This provider reads those
leg sheets, auto-detects the hub airports, runs the Connection Builder at each hub,
and produces Itinerary objects for the QSI engine.

Key insight from Chat 24:
  QSILHR Leg 1.1 = flights FROM Bay Area (SJC/SFO/OAK) TO 23 hub airports (346 flights)
  QSILHR Leg 2.1 = flights FROM those hubs TO beyond destinations (5,732 flights)
  QSILHR Leg 1.2 = return: beyond destinations TO hubs (5,698 flights)
  QSILHR Leg 2.2 = return: hubs TO Bay Area (318 flights)

The naming convention in QSILHR is:
  Direction 1: Beyond  Hub  Destination (Bay Area)
  Direction 2: Destination (Bay Area)  Hub  Beyond

But our pipeline convention is:
  QSI 1: Beyond home hub  Home  Dest (i.e., Beyond  LHR  SJC)
  QSI 2: Origin  Dest  Beyond dest (i.e., SJC  SJC/catchment  Beyond)

For QSILHR (LHR perspective):
  QSI 1 uses Leg 2.1 (hubbeyond) as leg1 and Leg 1.1 (BAhub) as leg2
  Actually  let's think about this carefully:

  QSILHR QSI 1 sheet shows connections like: BRU-SN-LHR-BA-SJC
  This means: BRU  LHR (Leg 1), LHR  SJC (Leg 2)
  So Leg 1.1 = flights from beyond to hub = arrivals at hub
  But Leg 1.1 contains SJCLHR, SFOORD etc. (origin  hub)

  Wait  the QSILHR file is from the HOME AIRPORT perspective (LHR).
  Direction 1 in the file = connections through the home hub.
  
  Leg 1.1 = Bay Area  Hub airports (all catchmenthub flights)
  Leg 2.1 = Hub airports  Beyond destinations
  
  For QSI 1 (beyond home hub): we need Beyond  Hub  Dest(Bay Area)
    - Leg1 of connection = Beyond  Hub = Leg 2.1 reversed? No...
    
  Actually, let's look at what QSI 1 in QSILHR contains:
    Route: BRU-SN-LHR-BA-SJC means BRULHRSJC
    - leg1: BRU  LHR (this is a flight arriving at LHR from BRU)
    - leg2: LHR  SJC (this is a flight departing LHR to SJC)
    
  But Leg 1.1 has SJCLHR, SFOORD (Bay Area  hubs), NOT BRULHR
  And Leg 2.1 has LHRBRU, FRACDG (hubs  beyond), NOT LHRSJC
  
  So for QSI 1 connections (Beyond  Hub  BayArea):
    leg1 (arrivals at hub from beyond) = these flights are in Leg 2.1 REVERSED 
    Actually no  Leg 2.1 has flights DEPARTING from hubs. The REVERSE direction
    (flights ARRIVING at hubs from beyond) would be Leg 1.2.
    
  Let me reconsider the direction numbering:
  
  File direction 1 (Leg 1.1, 2.1):
    Leg 1.1: BayArea  Hub (SJCLHR, SFOORD, etc.)
    Leg 2.1: Hub  Beyond (LHRBRU, FRACDG, etc.)
    Connection: BayArea  Hub  Beyond (outbound from Bay Area perspective)
    
  File direction 2 (Leg 1.2, 2.2):
    Leg 1.2: Beyond  Hub (BRULHR, CDGFRA, etc.)  
    Leg 2.2: Hub  BayArea (LHRSJC, ORDSFO, etc.)
    Connection: Beyond  Hub  BayArea (inbound to Bay Area perspective)

  Now mapping to pipeline QSI directions:
    QSI 1 (beyond home hub = beyond LHR):
      Connections: BeyondCity  LHR  SJC
      leg1 = beyond arriving at hub = Leg 1.2 data
      leg2 = hub departing to BayArea = Leg 2.2 data
      
    QSI 2 (beyond dest = beyond SJC, but from LHR perspective this is moot):
      Actually for QSILHR, QSI 2 shows connections from SJC side.
      QSI 2 in the file = connections via Bay Area hubs to beyond.
      Connections: Beyond  SJC/BayArea hubs  LHR
      But that doesn't make sense for the LHR perspective...
      
  Let me look at what QSI 2 in QSILHR actually contains:
    It has routes like: SEA-AS-SJC-BA-LHR
    This means SEA  SJC  LHR
    - leg1: SEA  SJC
    - leg2: SJC  LHR (BA service)
    
  So:
    QSI 2 connections go: Beyond  SJC/BayArea  LHR
    leg1 = beyond arriving at SJC-area = Leg 1.1 data REVERSED? No...
    
    Leg 1.1 = BayAreaHub (SJCLHR etc.)
    But QSI 2 needs: SEA  SJC (arrival at SJC from beyond)
    
    These flights aren't in Leg 1.1! Leg 1.1 has Bay Area DEPARTURES.
    QSI 2 needs Bay Area ARRIVALS from beyond cities.
    
    Hmm  this means the QSILHR file may have data structured differently
    than I assumed. Let me think again...

  CORRECT INTERPRETATION (from the Chat 24 discovery):
    
    The QSILHR file is built from the LHR HOME AIRPORT perspective.
    The QSI model evaluates competitive connections for markets reachable
    via LHR.
    
    QSI 1 in QSILHR = outbound connections: BayArea  Hub  Beyond
      These are scored to determine: given all ways to get from Bay Area
      to BRU, what share does BA-via-LHR capture?
      Leg 1.1 = BayArea  Hub (SJCLHR, SFOORD, etc.) 
      Leg 2.1 = Hub  Beyond (LHRBRU, FRACDG, etc.)
      Connection: SJCLHRBRU (BA), SFOORDBRU (UA), etc.
      
    QSI 2 in QSILHR = inbound connections: Beyond  Hub  BayArea
      Same markets, reverse direction. What share does BA capture for
      BRUBayArea traffic?
      Leg 1.2 = Beyond  Hub (BRULHR, CDGFRA, etc.)
      Leg 2.2 = Hub  BayArea (LHRSJC, ORDSFO, etc.)
      Connection: BRULHRSJC (BA), BRUFRASFO (LH), etc.

    The pipeline's QSI 1/QSI 2 directions are the SAME concept:
      QSI 1 = beyond home hub perspective
      QSI 2 = beyond dest hub perspective
      
    For QSILHR:
      home = LHR, dest = SJC
      QSI 1 = beyond LHR = cities in Europe/ME/Asia
      QSI 2 = beyond SJC = cities in Americas
      
    But in QSILHR the file actually evaluates BOTH directions of the
    SAME set of beyond cities (outbound and inbound for bidirectional avg).
    
    So:
      Pipeline QSI 1 (beyond home hub) maps to QSILHR directions 1+2 combined.
      The file's Direction 1 = outbound half, Direction 2 = inbound half.
      Both get bidirectionally averaged in QSI Calc.

  SIMPLIFICATION FOR SingleExtractOAGProvider:
    We DON'T need to map file directions to pipeline QSI directions.
    Instead, we:
    1. Read ALL 4 leg sheets
    2. Build connections for Direction 1: Leg 1.1 + Leg 2.1 (outbound)
    3. Build connections for Direction 2: Leg 1.2 + Leg 2.2 (inbound)
    4. Return Direction 1 connections as 'qsi1'
    5. Return Direction 2 connections as 'qsi2'
    
    The pipeline's QSI engine handles bidirectional averaging.
    The city codes in QSI 1 = beyond cities (from Leg 2.1 destinations)
    The city codes in QSI 2 = beyond cities (from Leg 1.2 origins)
    These should be the same set of cities  just different directions.

Dependencies:
    - providers.py (Itinerary, ScheduleProvider)
    - connection_builder.py (build_connections, alliances, MCT)
    - oag_parser.py (read_mct_xls, build_mct_lookup, read_city_lookup)
"""

from config import REFERENCE_CASE_DIR
import os
import sys
from collections import defaultdict, Counter
from typing import List, Dict, Optional, Tuple, Set, Any


from providers import Itinerary, ScheduleProvider
from connection_builder import (
    build_connections, create_proposed_service,
    load_alliance_data, load_lcc_list,
    parse_days_string, get_dom_int,
    ONEWORLD, STAR_ALLIANCE, SKYTEAM,
)
from oag_parser import (
    read_mct_xls, build_mct_lookup, read_city_lookup,
    _parse_hhmm_to_minutes, _parse_time_to_minutes,
)


# ============================================================================
# QSI TEMPLATE LEG SHEET READER
# ============================================================================

# Column indices in the QSI template leg sheets (standard OAG Analyser format)
COL_CARRIER       = 0   # Carrier1 (2-letter IATA)
COL_CARRIER_NAME  = 1   # Carrier1Name
COL_FLIGHT_NO     = 2   # FlightNo1
COL_DEP_AIRPORT   = 10  # DepAirport
COL_DEP_TERMINAL  = 12  # DepTerminal
COL_DEP_CITY      = 13  # DepCity
COL_DEP_COUNTRY   = 17  # DepIATACtry
COL_ARR_AIRPORT   = 23  # ArrAirport
COL_ARR_TERMINAL  = 25  # ArrTerminal
COL_ARR_CITY      = 26  # ArrCity
COL_ARR_COUNTRY   = 30  # ArrIATACtry
COL_DEP_TIME      = 36  # LocalDepTime
COL_ARR_TIME      = 37  # LocalArrTime
COL_ARR_DAY       = 38  # LocalArrday
COL_DAYS_OF_OP    = 39  # LocaldaysOfOp
COL_ELAPSED_TIME  = 55  # ElapsedTime
COL_FLYING_TIME   = 56  # FlyingTime
COL_FREQUENCY     = 60  # Frequency (may be None  derive from DaysOfOp)
COL_SEATS         = 42  # Seats

HEADER_ROW = 3  # 0-indexed row where column headers appear
DATA_START = 4  # 0-indexed first data row


def _safe_str(val, default=''):
    """Safely convert cell value to stripped string."""
    if val is None:
        return default
    return str(val).strip()


def _safe_int(val, default=0):
    """Safely convert cell value to int."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _parse_elapsed(val):
    """Parse elapsed time from OAG format (e.g., '001025' or '0409'  minutes)."""
    if val is None:
        return 0
    s = str(val).strip().replace(':', '')
    if not s:
        return 0
    # Try HHMM or HHHMM format
    try:
        if len(s) <= 4:
            h = int(s[:-2]) if len(s) > 2 else 0
            m = int(s[-2:])
        else:
            h = int(s[:-2])
            m = int(s[-2:])
        return h * 60 + m
    except (ValueError, IndexError):
        return 0


def _count_days(days_str):
    """Count operating days from OAG days string like '1234567' or '  3 5 7'."""
    if not days_str:
        return 7  # assume daily if not specified
    s = str(days_str).strip()
    return sum(1 for c in s if c.isdigit())


def load_legs_from_qsi_template(filepath: str, sheet_name: str) -> List[Dict]:
    """
    Load flight leg records from a QSI template leg sheet.
    
    Returns list of dicts compatible with connection_builder.build_connections().
    
    Args:
        filepath: Path to QSI template Excel file
        sheet_name: Sheet name (e.g., 'Leg 1.1', 'Leg 2.1', etc.)
    
    Returns:
        List of leg dicts with keys matching Connection Builder expectations.
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    legs = []
    for idx, row in enumerate(rows[DATA_START:], start=1):
        if not row or not row[COL_CARRIER]:
            continue
        
        carrier = _safe_str(row[COL_CARRIER])
        # Skip non-carrier rows (headers, labels, blanks)
        if len(carrier) != 2 or not carrier.isalpha() or not carrier.isupper():
            continue
        
        dep_airport = _safe_str(row[COL_DEP_AIRPORT])
        arr_airport = _safe_str(row[COL_ARR_AIRPORT])
        if not dep_airport or not arr_airport:
            continue
        
        dep_time = _safe_str(row[COL_DEP_TIME])
        arr_time = _safe_str(row[COL_ARR_TIME])
        days_of_op = _safe_str(row[COL_DAYS_OF_OP], '1234567')
        elapsed_str = _safe_str(row[COL_ELAPSED_TIME])
        flying_str = _safe_str(row[COL_FLYING_TIME])
        
        dep_time_mins = _parse_hhmm_to_minutes(dep_time)
        arr_time_mins = _parse_hhmm_to_minutes(arr_time)
        flying_mins = _parse_elapsed(flying_str) or _parse_elapsed(elapsed_str)
        
        dep_day_set = parse_days_string(days_of_op)
        arr_day_offset = _safe_str(row[COL_ARR_DAY]) if len(row) > COL_ARR_DAY else ''
        if arr_day_offset and arr_day_offset.strip():
            arr_day_set = parse_days_string(arr_day_offset)
        else:
            arr_day_set = dep_day_set.copy()
        
        # Frequency: use col 60 if populated, else count days of operation
        freq = _safe_int(row[COL_FREQUENCY]) if len(row) > COL_FREQUENCY else 0
        if freq == 0:
            freq = _count_days(days_of_op)
        
        dep_country = _safe_str(row[COL_DEP_COUNTRY])
        arr_country = _safe_str(row[COL_ARR_COUNTRY])
        
        leg = {
            'id': idx,
            'carrier': carrier,
            'carrier_name': _safe_str(row[COL_CARRIER_NAME]),
            'flight_no': _safe_str(row[COL_FLIGHT_NO]),
            'dep_airport': dep_airport,
            'dep_terminal': _safe_str(row[COL_DEP_TERMINAL]),
            'dep_city': _safe_str(row[COL_DEP_CITY]) or dep_airport,
            'dep_country': dep_country,
            'arr_airport': arr_airport,
            'arr_terminal': _safe_str(row[COL_ARR_TERMINAL]),
            'arr_city': _safe_str(row[COL_ARR_CITY]) or arr_airport,
            'arr_country': arr_country,
            'dep_time': dep_time,
            'arr_time': arr_time,
            'dep_time_mins': dep_time_mins,
            'arr_time_mins': arr_time_mins,
            'dep_days': days_of_op,
            'arr_days': arr_day_offset or days_of_op,
            'dep_day_set': dep_day_set,
            'arr_day_set': arr_day_set,
            'flying_time': flying_str,
            'flying_mins': flying_mins,
            'dom_int': get_dom_int(dep_country, arr_country),
            'seats': _safe_int(row[COL_SEATS]) if len(row) > COL_SEATS else 0,
            'is_proposed': False,
        }
        legs.append(leg)
    
    return legs


def load_city_lookup_from_qsi(filepath: str) -> Dict[str, str]:
    """
    Load airportcity code mapping from the Lookups sheet of a QSI template.
    
    Returns dict: airport_code  city_code
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if 'Lookups' not in wb.sheetnames:
        wb.close()
        return {}
    
    ws = wb['Lookups']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    mapping = {}
    for row in rows[4:]:  # Data starts at row 5 (0-indexed 4)
        if row and row[1] and row[2]:
            apt = _safe_str(row[1])
            city = _safe_str(row[2])
            if len(apt) == 3 and len(city) == 3:
                mapping[apt] = city
    
    return mapping


def connection_to_itinerary(cnx: Dict, city_code: str = None,
                            airport_code: str = None) -> Itinerary:
    """
    Convert a connection_builder output dict to an Itinerary object.
    
    Args:
        cnx: Connection dict from build_connections()
        city_code: Override city code (for city-level aggregation)
        airport_code: Override airport code
    """
    return Itinerary(
        city=city_code or cnx.get('dep_airport', ''),
        airport=airport_code or cnx.get('dep_airport', ''),
        route_label=cnx.get('route_label', ''),
        dep_airport=cnx.get('dep_airport', ''),
        cnx_airport=cnx.get('cnx_airport', ''),
        carrier_l1=cnx.get('leg1_carrier', ''),
        carrier_l2=cnx.get('leg2_carrier', ''),
        freq=cnx.get('frequency', 0),
        elapsed=int(cnx.get('elapsed_time', 0)),
        cnx_type=cnx.get('cnx_type', 'INTERLINING'),
        is_proposed=(cnx.get('leg1_is_proposed', False) or
                     cnx.get('leg2_is_proposed', False)),
    )


# ============================================================================
# SINGLE EXTRACT OAG PROVIDER
# ============================================================================

class SingleExtractOAGProvider(ScheduleProvider):
    """
    ScheduleProvider that reads OAG leg data from a single QSI template file
    and builds connections across all auto-detected hub airports.
    
    This provider:
    1. Reads Leg 1.1, 2.1, 1.2, 2.2 from the QSI template
    2. Auto-detects hub airports (common airports between leg1 arrivals and leg2 departures)
    3. Runs Connection Builder at each hub
    4. Applies airportcity code mapping from the Lookups sheet
    5. Returns Itinerary objects compatible with the QSI engine
    
    Direction mapping:
        'qsi1'  File Direction 1: Leg 1.1 (originhub) + Leg 2.1 (hubbeyond)
                  Connection: Origin  Hub  Beyond
                  City = beyond destination (from Leg 2.1 arr_airport)
                  
        'qsi2'  File Direction 2: Leg 1.2 (beyondhub) + Leg 2.2 (huborigin)
                  Connection: Beyond  Hub  Origin
                  City = beyond origin (from Leg 1.2 dep_airport)
    
    Args:
        qsi_file: Path to QSI template file with populated leg sheets
        origin_airport: Origin airport IATA code (e.g., 'LHR')
        dest_airport: Destination airport IATA code (e.g., 'SJC')
        proposed_carrier: Carrier code for the proposed service (e.g., 'BA')
        use_city_codes: If True, map airport codes to city codes via Lookups
        city_lookup_file: Optional external city lookup file
        mct_files: Dict of hub_code  MCT file path
        min_connect: Minimum connection time (minutes)
        max_connect: Maximum connection time (minutes)
        default_mct: Default MCT when no specific rule found (minutes)
    """

    def __init__(self,
                 qsi_file: str,
                 origin_airport: str,
                 dest_airport: str,
                 proposed_carrier: str = 'BA',
                 use_city_codes: bool = True,
                 city_lookup_file: str = None,
                 mct_files: Dict[str, str] = None,
                 min_connect: int = 20,
                 max_connect: int = 720,
                 default_mct: int = 90):

        self.qsi_file = qsi_file
        self.origin_airport = origin_airport.upper()
        self.dest_airport = dest_airport.upper()
        self.proposed_carrier = proposed_carrier.upper()
        self.use_city_codes = use_city_codes
        self.mct_files = mct_files or {}
        self.min_connect = min_connect
        self.max_connect = max_connect
        self.default_mct = default_mct

        self._cache: Dict[str, List[Itinerary]] = {}
        self._build_log: List[str] = []
        self._stats: Dict[str, Any] = {}

        # Load city mapping from Lookups sheet
        self._city_map: Dict[str, str] = {}
        if use_city_codes:
            self._city_map = load_city_lookup_from_qsi(qsi_file)
            if city_lookup_file and os.path.exists(city_lookup_file):
                try:
                    extra, _ = read_city_lookup(city_lookup_file)
                    self._city_map.update(extra)
                except Exception:
                    pass

        # Alliance data (use built-in defaults)
        self._alliances = [ONEWORLD, STAR_ALLIANCE, SKYTEAM]
        self._lcc_set = set()  # No LCC filtering for connection building

    def _log(self, msg: str):
        self._build_log.append(msg)

    def _apt_to_city(self, apt: str) -> str:
        """Map airport code to city code."""
        if self.use_city_codes and apt in self._city_map:
            return self._city_map[apt]
        return apt

    #  ScheduleProvider interface 

    def get_itineraries(self, direction: str) -> List[Itinerary]:
        if direction in self._cache:
            return self._cache[direction]

        if direction == 'qsi1':
            result = self._build_qsi1()
        elif direction == 'qsi2':
            result = self._build_qsi2()
        else:
            result = []

        self._cache[direction] = result
        return result

    def get_metadata(self) -> Dict[str, Any]:
        return {
            'provider_type': 'SingleExtractOAGProvider',
            'qsi_file': os.path.basename(self.qsi_file),
            'origin': self.origin_airport,
            'destination': self.dest_airport,
            'use_city_codes': self.use_city_codes,
            'city_mappings': len(self._city_map),
            'stats': self._stats,
            'build_log': self._build_log,
        }

    #  QSI Direction 1 

    def _build_qsi1(self) -> List[Itinerary]:
        """
        Build QSI 1 itineraries from Leg 1.1 and Leg 2.1.
        
        Leg 1.1: Origin/catchment  Hub airports
        Leg 2.1: Hub airports  Beyond destinations
        
        Connection: Origin  Hub  Beyond
        City label = beyond destination (Leg 2.1 arrival airport, mapped to city)
        
        For each auto-detected hub, runs the Connection Builder and collects
        all valid connecting itineraries.
        """
        self._log(f"\n{'='*60}")
        self._log(f"QSI 1: {self.origin_airport}  Hubs  Beyond")
        self._log(f"{'='*60}")

        # Load leg data
        leg1_all = load_legs_from_qsi_template(self.qsi_file, 'Leg 1.1')
        leg2_all = load_legs_from_qsi_template(self.qsi_file, 'Leg 2.1')
        self._log(f"  Leg 1.1: {len(leg1_all)} flights (origin  hubs)")
        self._log(f"  Leg 2.1: {len(leg2_all)} flights (hubs  beyond)")

        if not leg1_all or not leg2_all:
            self._log("   No leg data  skipping")
            return []

        # Auto-detect hubs: airports that appear as arrivals in Leg 1.1 
        # AND as departures in Leg 2.1
        leg1_arr = set(l['arr_airport'] for l in leg1_all)
        leg2_dep = set(l['dep_airport'] for l in leg2_all)
        hubs = sorted(leg1_arr & leg2_dep)
        self._log(f"  Auto-detected hubs: {len(hubs)}  {hubs}")

        # Build connections at each hub
        all_itineraries = []
        hub_stats = {}

        for hub in hubs:
            # Filter legs for this hub
            leg1 = [l for l in leg1_all if l['arr_airport'] == hub]
            leg2 = [l for l in leg2_all if l['dep_airport'] == hub]

            if not leg1 or not leg2:
                continue

            # Load MCT for this hub
            mct = self._load_mct(hub)

            # Build connections
            valid, failed = build_connections(
                leg1, leg2, self._alliances, mct, self._lcc_set,
                self.min_connect, self.max_connect, self.default_mct,
                hub_airport=hub)

            # Convert to itineraries
            # City = beyond destination (arr_airport of leg2 = connection dest)
            hub_itineraries = []
            for cnx in valid:
                # The "beyond" city is the arrival airport of the connection
                beyond_apt = cnx.get('airport_label', '') or cnx.get('arr_airport', '')
                beyond_city = self._apt_to_city(beyond_apt)
                it = connection_to_itinerary(cnx, city_code=beyond_city,
                                             airport_code=beyond_apt)
                hub_itineraries.append(it)

            cities = set(it.city for it in hub_itineraries)
            hub_stats[hub] = {
                'leg1': len(leg1),
                'leg2': len(leg2),
                'valid': len(valid),
                'failed': len(failed),
                'itineraries': len(hub_itineraries),
                'cities': len(cities),
            }
            self._log(f"  {hub}: {len(leg1)}+{len(leg2)} legs  "
                      f"{len(hub_itineraries)} itineraries, {len(cities)} cities")

            all_itineraries.extend(hub_itineraries)

        total_cities = set(it.city for it in all_itineraries)
        self._stats['qsi1'] = {
            'total_itineraries': len(all_itineraries),
            'total_cities': len(total_cities),
            'hubs': len(hub_stats),
            'hub_detail': hub_stats,
        }
        self._log(f"\n  TOTAL QSI 1: {len(all_itineraries)} itineraries, "
                  f"{len(total_cities)} cities across {len(hub_stats)} hubs")

        return all_itineraries

    #  QSI Direction 2 

    def _build_qsi2(self) -> List[Itinerary]:
        """
        Build QSI 2 itineraries from Leg 1.2 and Leg 2.2.
        
        Leg 1.2: Beyond destinations  Hub airports
        Leg 2.2: Hub airports  Origin/catchment
        
        Connection: Beyond  Hub  Origin
        City label = beyond origin (Leg 1.2 departure airport, mapped to city)
        
        This is the reverse direction for bidirectional QSI averaging.
        """
        self._log(f"\n{'='*60}")
        self._log(f"QSI 2: Beyond  Hubs  {self.origin_airport}")
        self._log(f"{'='*60}")

        # Load leg data
        leg1_all = load_legs_from_qsi_template(self.qsi_file, 'Leg 1.2')
        leg2_all = load_legs_from_qsi_template(self.qsi_file, 'Leg 2.2')
        self._log(f"  Leg 1.2: {len(leg1_all)} flights (beyond  hubs)")
        self._log(f"  Leg 2.2: {len(leg2_all)} flights (hubs  origin)")

        if not leg1_all or not leg2_all:
            self._log("   No leg data  skipping")
            return []

        # Auto-detect hubs: arrivals in Leg 1.2 AND departures in Leg 2.2
        leg1_arr = set(l['arr_airport'] for l in leg1_all)
        leg2_dep = set(l['dep_airport'] for l in leg2_all)
        hubs = sorted(leg1_arr & leg2_dep)
        self._log(f"  Auto-detected hubs: {len(hubs)}  {hubs}")

        # Build connections at each hub
        all_itineraries = []
        hub_stats = {}

        for hub in hubs:
            leg1 = [l for l in leg1_all if l['arr_airport'] == hub]
            leg2 = [l for l in leg2_all if l['dep_airport'] == hub]

            if not leg1 or not leg2:
                continue

            mct = self._load_mct(hub)

            valid, failed = build_connections(
                leg1, leg2, self._alliances, mct, self._lcc_set,
                self.min_connect, self.max_connect, self.default_mct,
                hub_airport=hub)

            # City = beyond origin (dep_airport of leg1 = connection origin)
            hub_itineraries = []
            for cnx in valid:
                beyond_apt = cnx.get('dep_airport', '')
                beyond_city = self._apt_to_city(beyond_apt)
                it = connection_to_itinerary(cnx, city_code=beyond_city,
                                             airport_code=beyond_apt)
                hub_itineraries.append(it)

            cities = set(it.city for it in hub_itineraries)
            hub_stats[hub] = {
                'leg1': len(leg1),
                'leg2': len(leg2),
                'valid': len(valid),
                'failed': len(failed),
                'itineraries': len(hub_itineraries),
                'cities': len(cities),
            }
            self._log(f"  {hub}: {len(leg1)}+{len(leg2)} legs  "
                      f"{len(hub_itineraries)} itineraries, {len(cities)} cities")

            all_itineraries.extend(hub_itineraries)

        total_cities = set(it.city for it in all_itineraries)
        self._stats['qsi2'] = {
            'total_itineraries': len(all_itineraries),
            'total_cities': len(total_cities),
            'hubs': len(hub_stats),
            'hub_detail': hub_stats,
        }
        self._log(f"\n  TOTAL QSI 2: {len(all_itineraries)} itineraries, "
                  f"{len(total_cities)} cities across {len(hub_stats)} hubs")

        return all_itineraries

    #  Helpers 

    def _load_mct(self, hub_code: str) -> Dict:
        """Load MCT rules for a hub airport."""
        mct_file = self.mct_files.get(hub_code)
        if mct_file and os.path.exists(mct_file):
            try:
                entries = read_mct_xls(mct_file)
                mct = build_mct_lookup(entries, hub_code)
                return mct
            except Exception:
                pass
        return {}

    def print_build_log(self):
        """Print the complete build log."""
        for line in self._build_log:
            print(line)

    def get_hub_summary(self) -> Dict[str, Dict]:
        """Get a summary of hub statistics for both directions."""
        return {
            'qsi1': self._stats.get('qsi1', {}),
            'qsi2': self._stats.get('qsi2', {}),
        }


# ============================================================================
# VALIDATION: Compare SingleExtractOAGProvider vs ExcelScheduleProvider
# ============================================================================

def compare_against_excel(single_prov: SingleExtractOAGProvider,
                          excel_prov: ScheduleProvider,
                          direction: str) -> Dict:
    """
    Compare itineraries from SingleExtractOAGProvider against
    the pre-computed ExcelScheduleProvider output.
    
    Returns comparison statistics.
    """
    single_its = single_prov.get_itineraries(direction)
    excel_its = excel_prov.get_itineraries(direction)

    def city_stats(its):
        d = defaultdict(lambda: {'count': 0, 'freq': 0, 'routes': set(), 'hubs': set()})
        for it in its:
            d[it.city]['count'] += 1
            d[it.city]['freq'] += it.freq
            d[it.city]['routes'].add(it.route_label)
            d[it.city]['hubs'].add(it.cnx_airport)
        return d

    sc = city_stats(single_its)
    ec = city_stats(excel_its)
    both = set(sc) & set(ec)
    
    # Per-city itinerary count comparison
    mismatches = []
    for city in sorted(both):
        s_count = sc[city]['count']
        e_count = ec[city]['count']
        if s_count != e_count:
            mismatches.append({
                'city': city,
                'single': s_count,
                'excel': e_count,
                'diff': s_count - e_count,
            })

    return {
        'direction': direction,
        'single_itineraries': len(single_its),
        'excel_itineraries': len(excel_its),
        'single_cities': len(sc),
        'excel_cities': len(ec),
        'cities_both': len(both),
        'cities_single_only': sorted(set(sc) - set(ec)),
        'cities_excel_only': sorted(set(ec) - set(sc)),
        'mismatches': mismatches[:20],  # Top 20
    }


# ============================================================================
# STANDALONE VALIDATION SCRIPT
# ============================================================================

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='SingleExtractOAGProvider  build and validate')
    parser.add_argument('--qsi-file', default=str(REFERENCE_CASE_DIR / 'QSILHR_v1_OS_JZ_17Feb15.xlsx'),
                        help='Path to QSI template file')
    parser.add_argument('--origin', default='LHR', help='Origin airport')
    parser.add_argument('--dest', default='SJC', help='Destination airport')
    parser.add_argument('--carrier', default='BA', help='Proposed carrier')
    parser.add_argument('--compare-excel', action='store_true',
                        help='Compare against ExcelScheduleProvider')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Print build log')
    args = parser.parse_args()

    # MCT files for BA LHR-SJC
    mct_files = {
        'LHR': str(REFERENCE_CASE_DIR / 'LHR_MCTs.xls'),
        'SJC': str(REFERENCE_CASE_DIR / 'Minimum_Cnx_Times_SJC.xls'),
    }

    print(f"Building SingleExtractOAGProvider from {os.path.basename(args.qsi_file)}")
    print(f"Route: {args.origin}-{args.dest} ({args.carrier})")
    print()

    prov = SingleExtractOAGProvider(
        qsi_file=args.qsi_file,
        origin_airport=args.origin,
        dest_airport=args.dest,
        proposed_carrier=args.carrier,
        use_city_codes=True,
        mct_files=mct_files,
    )

    # Build QSI 1
    q1 = prov.get_itineraries('qsi1')
    print(f"QSI 1: {len(q1)} itineraries, "
          f"{len(set(it.city for it in q1))} cities")

    # Build QSI 2
    q2 = prov.get_itineraries('qsi2')
    print(f"QSI 2: {len(q2)} itineraries, "
          f"{len(set(it.city for it in q2))} cities")

    if args.verbose:
        print("\n--- BUILD LOG ---")
        prov.print_build_log()

    # Optional: compare against pre-computed Excel
    if args.compare_excel:
        from providers import ExcelScheduleProvider

        excel = ExcelScheduleProvider(
            qsi1_file=args.qsi_file,
            qsi2_file=args.qsi_file,
        )

        for direction in ['qsi1', 'qsi2']:
            comp = compare_against_excel(prov, excel, direction)
            print(f"\n--- {direction.upper()} Comparison ---")
            print(f"  Single: {comp['single_itineraries']} its, {comp['single_cities']} cities")
            print(f"  Excel:  {comp['excel_itineraries']} its, {comp['excel_cities']} cities")
            print(f"  Both:   {comp['cities_both']} cities")
            if comp['cities_single_only']:
                print(f"  Single-only: {comp['cities_single_only'][:10]}")
            if comp['cities_excel_only']:
                print(f"  Excel-only:  {comp['cities_excel_only'][:10]}")
            if comp['mismatches']:
                print(f"  Itinerary count mismatches:")
                for m in comp['mismatches'][:10]:
                    print(f"    {m['city']}: single={m['single']} excel={m['excel']} "
                          f"diff={m['diff']:+d}")
