#!/usr/bin/env python3
"""
Avia Solutions  Predictive Calibration Engine
===============================================
Given a new route's characteristics, auto-suggests all calibration parameters
(stimulation, capture, QSI adjustment, connecting share expectations) using
the 22-case calibration library and 8 discovered rules.

This is the module that turns the calibration library from documentation
into automation. Instead of an analyst manually choosing stim=1.20, cap=35%
by reviewing precedents, this engine does it systematically.

ARCHITECTURE:
  RouteProfile  FeatureExtractor  RuleMatcher  ParameterPredictor  CalibrationSuggestion
  
  1. RouteProfile: structured input describing the proposed route
  2. FeatureExtractor: derives Boolean/categorical features (is_zero_metro, catchment_type, etc.)
  3. RuleMatcher: applies 8 calibration rules in priority order
  4. ParameterPredictor: selects parameter ranges and picks point estimates
  5. CalibrationSuggestion: output with suggested values, confidence, reasoning, comps

VALIDATED AGAINST:
  - CI TPE-SJC blind test: 0.6% error
  - LH FRA-SJC / MUC-SJC paired test: 0.0-0.4% error (with rules)
  - BR TPE-SJC Aug19B blind test: 0.0% error

CRITICAL DESIGN PRINCIPLE (Rule 4):
  This engine predicts P2P parameters (stim, capture, growth) and QSI adjustment.
  It does NOT predict blended connecting QSI  that must come from the pipeline's
  connection builder and QSI scorer. Connecting QSI depends on market-by-market
  competitive structures that cannot be benchmarked from hub size alone.

USAGE:
  from predictive_calibration_engine import PredictiveCalibrationEngine, RouteProfile
  
  profile = RouteProfile(
      origin='FRA', destination='SJC', carrier='LH', alliance='Star Alliance',
      carrier_type='Full Service', hub_airport='FRA', hub_status='Major Hub',
      frequency=5, aircraft='A340-300', seats=267,
      new_route=True, existing_service_same_pair=False,
      existing_service_hub_to_metro=True,  # LH FRA-SFO exists
      nearby_airport_direct=True,  # SFO has direct from FRA
      catchment_overlap=False,  # SJC  SFO catchment
      primary_demand='Mixed', business_leisure_split=0.60,
  )
  
  engine = PredictiveCalibrationEngine()
  suggestion = engine.predict(profile)
  suggestion.print_summary()

LAST UPDATED: February 2026
"""

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any
from enum import Enum
import json


# =============================================================================
# ENUMS AND CONSTANTS
# =============================================================================

class HubStatus(Enum):
    MAJOR_HUB = "Major Hub"
    SECONDARY_HUB = "Secondary Hub"
    NON_HUB = "Non-Hub"


class CarrierType(Enum):
    FULL_SERVICE = "Full Service"
    LCC = "LCC"
    ULTRA_LCC = "Ultra-LCC"
    HYBRID = "Hybrid"
    CHARTER = "Charter"


class MarketMaturity(Enum):
    NEW_UNSERVED = "New Unserved"          # No direct anywhere nearby
    ZERO_HUB_METRO = "Zero Hub-Metro"      # Hub has zero direct to dest metro
    VIRGIN_NONSTOP = "Virgin Nonstop"      # No direct on this pair, some indirect
    WELL_SERVED_INDIRECT = "Well Served"   # Good indirect options exist
    NEARBY_DIRECT_OVERLAP = "Nearby Overlap"   # Nearby airport direct, same catchment
    NEARBY_DIRECT_DISTINCT = "Nearby Distinct"  # Nearby airport direct, different catchment
    EXISTING_RE_ASSESSMENT = "Existing"    # Re-assessing current service


class Confidence(Enum):
    HIGH = "HIGH"        # 3+ library cases with similar profile
    MEDIUM = "MEDIUM"    # 1-2 comparable cases or strong rule match
    LOW = "LOW"          # No close precedent, extrapolating from rules


# Standard 8-segment structure used by Avia forecasts
SEGMENT_NAMES = [
    'origin_business', 'origin_leisure_primary', 'origin_leisure_secondary', 'origin_leisure_contested',
    'dest_business', 'dest_leisure_primary', 'dest_leisure_secondary', 'dest_leisure_contested',
]


# =============================================================================
# ROUTE PROFILE  INPUT
# =============================================================================

@dataclass
class RouteProfile:
    """Describes a proposed route for parameter prediction."""
    # Route identity
    origin: str                              # IATA airport code
    destination: str                         # IATA airport code
    origin_city: str = ""                    # IATA city code (if different)
    destination_city: str = ""               # IATA city code (if different)
    carrier: str = ""                        # IATA carrier code
    carrier_name: str = ""
    alliance: str = ""                       # OneWorld / Star Alliance / SkyTeam / None
    carrier_type: str = "Full Service"       # Full Service / LCC / etc.
    
    # Hub characteristics
    hub_airport: str = ""                    # Which end is the hub (if any)
    hub_status: str = "Major Hub"            # Major Hub / Secondary Hub / Non-Hub
    
    # Service parameters
    frequency: int = 5                       # flights per week
    aircraft: str = ""
    seats: int = 280
    departure_time_home: str = ""            # HH:MM local
    departure_time_dest: str = ""
    
    # Market characteristics  THE KEY INPUTS FOR PARAMETER SELECTION
    new_route: bool = True                   # Is this a new route proposal?
    existing_service_same_pair: bool = False  # Does this carrier already fly this pair?
    existing_service_hub_to_metro: bool = True  # Does hub have ANY direct to dest metro?
    nearby_airport_direct: bool = False       # Does a nearby airport have direct service?
    catchment_overlap: bool = True            # Do nearby airports share the same catchment?
    
    # Demand profile
    primary_demand: str = "Mixed"            # Business / Leisure / VFR / Mixed
    business_leisure_split: float = 0.55     # Business fraction of P2P
    seasonal_profile: str = "Year-Round"
    
    # P2P demand data (if known  from MIDT/Sabre)
    p2p_base_demand: Optional[float] = None
    p2p_growth_rate: Optional[float] = None
    
    # Connecting demand data (if known)
    cnx_home_pool: Optional[float] = None
    cnx_dest_pool: Optional[float] = None
    
    # Distance / ULR flag
    distance_nm: Optional[float] = None
    is_ulr: bool = False                     # Ultra-long-range (>8,000nm)
    
    # Data vintage
    data_vintage: str = ""                   # e.g., "2024"
    forecast_year: Optional[int] = None
    
    def __post_init__(self):
        if not self.origin_city:
            self.origin_city = self.origin
        if not self.destination_city:
            self.destination_city = self.destination
        if not self.hub_airport:
            self.hub_airport = self.origin


# =============================================================================
# CALIBRATION SUGGESTION  OUTPUT
# =============================================================================

@dataclass
class SegmentSuggestion:
    """Predicted parameters for one P2P demand segment."""
    segment: str
    stimulation: float
    capture: float
    stim_range: Tuple[float, float]
    capture_range: Tuple[float, float]
    stim_reasoning: str
    capture_reasoning: str
    confidence: str


@dataclass
class CalibrationSuggestion:
    """Complete set of predicted calibration parameters for a route."""
    # Route identity
    route_label: str = ""
    market_maturity: str = ""
    
    # P2P predictions  8 segments
    segments: Dict[str, SegmentSuggestion] = field(default_factory=dict)
    
    # Blended summaries
    blended_stimulation: float = 1.0
    blended_capture: float = 0.30
    
    # QSI adjustment
    qsi_adjustment: float = 1.0
    qsi_adj_reasoning: str = ""
    
    # Connecting expectations (ranges, not point estimates  per Rule 4)
    cnx_share_range: Tuple[float, float] = (0.30, 0.60)
    cnx_share_reasoning: str = ""
    
    # Growth
    growth_rate: float = 0.10
    growth_reasoning: str = ""
    
    # Suggested load factor range (learned from comparable cases)
    suggested_lf_range: Tuple[float, float] = (0.65, 0.82)
    suggested_lf_reasoning: str = ""
    
    # Overall confidence
    confidence: str = "MEDIUM"
    
    # Rules applied
    rules_applied: List[str] = field(default_factory=list)
    
    # Comparable cases
    comparable_cases: List[Dict] = field(default_factory=list)
    
    # Warnings
    warnings: List[str] = field(default_factory=list)
    
    def print_summary(self):
        """Print a human-readable summary of predictions."""
        print(f"\n{'='*70}")
        print(f"PREDICTIVE CALIBRATION: {self.route_label}")
        print(f"Market Maturity: {self.market_maturity}")
        print(f"Confidence: {self.confidence}")
        print(f"{'='*70}")
        
        print(f"\n--- P2P Segment Predictions ---")
        print(f"{'Segment':<30s} {'Stim':>6s} {'Capture':>8s} {'Conf':>8s}")
        print(f"{'-'*52}")
        for name, seg in self.segments.items():
            print(f"{name:<30s} {seg.stimulation:>6.2f} {seg.capture:>7.0%} {seg.confidence:>8s}")
        
        print(f"\nBlended Stim: {self.blended_stimulation:.3f}")
        print(f"Blended Capture: {self.blended_capture:.1%}")
        print(f"QSI Adjustment: {self.qsi_adjustment:.2f}  {self.qsi_adj_reasoning}")
        print(f"Growth Rate: {self.growth_rate:.1%}  {self.growth_reasoning}")
        print(f"Expected Cnx Share: {self.cnx_share_range[0]:.0%}-{self.cnx_share_range[1]:.0%}  {self.cnx_share_reasoning}")
        
        if self.rules_applied:
            print(f"\n--- Rules Applied ---")
            for r in self.rules_applied:
                print(f"   {r}")
        
        if self.comparable_cases:
            print(f"\n--- Comparable Cases ---")
            for c in self.comparable_cases:
                print(f"   {c['route_id']}: {c.get('relevance', '')}")
        
        if self.warnings:
            print(f"\n--- Warnings ---")
            for w in self.warnings:
                print(f"   {w}")
    
    def to_dict(self) -> Dict:
        """Convert to serialisable dictionary."""
        d = {
            'route_label': self.route_label,
            'market_maturity': self.market_maturity,
            'confidence': self.confidence,
            'blended_stimulation': self.blended_stimulation,
            'blended_capture': self.blended_capture,
            'qsi_adjustment': self.qsi_adjustment,
            'qsi_adj_reasoning': self.qsi_adj_reasoning,
            'growth_rate': self.growth_rate,
            'cnx_share_range': list(self.cnx_share_range),
            'rules_applied': self.rules_applied,
            'warnings': self.warnings,
            'segments': {},
        }
        for name, seg in self.segments.items():
            d['segments'][name] = {
                'stimulation': seg.stimulation,
                'capture': seg.capture,
                'stim_range': list(seg.stim_range),
                'capture_range': list(seg.capture_range),
                'confidence': seg.confidence,
            }
        return d


# =============================================================================
# PREDICTIVE CALIBRATION ENGINE
# =============================================================================

class PredictiveCalibrationEngine:
    """
    The main engine. Takes a RouteProfile, applies rules and library patterns,
    returns a CalibrationSuggestion with auto-suggested parameters.
    """
    
    def __init__(self, calibration_cases=None, calibration_patterns=None):
        """
        Initialise with calibration data. If not provided, imports from
        calibration_library_v8.
        """
        if calibration_cases is not None:
            self.cases = calibration_cases
            self.patterns = calibration_patterns or {}
        else:
            try:
                from calibration_library_v8 import CALIBRATION_CASES, CALIBRATION_PATTERNS
                self.cases = CALIBRATION_CASES
                self.patterns = CALIBRATION_PATTERNS
            except ImportError:
                self.cases = []
                self.patterns = {}
    
    # =========================================================================
    # MAIN PREDICTION FLOW
    # =========================================================================
    
    def predict(self, profile: RouteProfile) -> CalibrationSuggestion:
        """
        Main entry point. Returns complete parameter suggestions.
        """
        suggestion = CalibrationSuggestion(
            route_label=f"{profile.carrier} {profile.origin}-{profile.destination}",
        )
        
        # Step 1: Classify market maturity
        maturity = self._classify_maturity(profile)
        suggestion.market_maturity = maturity.value
        
        # Step 2: Find comparable cases
        comps = self._find_comparables(profile, maturity)
        suggestion.comparable_cases = comps
        
        # Step 3: Predict stimulation factors
        stim_params = self._predict_stimulation(profile, maturity, comps)
        
        # Step 4: Predict capture rates
        cap_params = self._predict_capture(profile, maturity, comps)
        
        # Step 5: Build segment suggestions
        segments = self._build_segments(profile, stim_params, cap_params, maturity, comps)
        suggestion.segments = segments
        
        # Step 6: Compute blended values
        suggestion.blended_stimulation = self._blend_stim(segments, profile)
        suggestion.blended_capture = self._blend_capture(segments, profile)
        
        # Step 7: QSI adjustment
        qsi_adj, qsi_reason = self._predict_qsi_adjustment(profile, maturity)
        suggestion.qsi_adjustment = qsi_adj
        suggestion.qsi_adj_reasoning = qsi_reason
        
        # Step 8: Growth rate
        growth, growth_reason = self._predict_growth(profile, comps)
        suggestion.growth_rate = growth
        suggestion.growth_reasoning = growth_reason
        
        # Step 9: Connecting share expectations
        cnx_range, cnx_reason = self._predict_cnx_share(profile, maturity, comps)
        suggestion.cnx_share_range = cnx_range
        suggestion.cnx_share_reasoning = cnx_reason
        
        # Step 9b: Suggested LF range from comparables and route type
        lf_range, lf_reason = self._predict_lf_range(profile, maturity, comps)
        suggestion.suggested_lf_range = lf_range
        suggestion.suggested_lf_reasoning = lf_reason
        
        # Step 10: Apply special rules and collect warnings
        self._apply_special_rules(profile, suggestion, maturity, comps)
        
        # Step 11: Set overall confidence
        suggestion.confidence = self._assess_confidence(profile, maturity, comps)
        
        return suggestion
    
    # =========================================================================
    # STEP 1: MARKET MATURITY CLASSIFICATION
    # =========================================================================
    
    def _classify_maturity(self, p: RouteProfile) -> MarketMaturity:
        """
        Classify the market maturity. This is the single most important
        determinant of parameter selection.
        
        Decision tree:
          Is this an existing service re-assessment?  EXISTING
          Is this a new route?
            Does hub have ANY direct to dest metro? (Rule 1 check)
              NO  ZERO_HUB_METRO (highest stim/capture)
              YES  Is there direct on this specific pair?
                YES with same catchment  NEARBY_DIRECT_OVERLAP
                YES with distinct catchment  NEARBY_DIRECT_DISTINCT (Rule 2)
                NO nearby direct at all  
                  Well-served indirectly?  WELL_SERVED_INDIRECT
                  No  NEW_UNSERVED or VIRGIN_NONSTOP
        """
        if not p.new_route:
            return MarketMaturity.EXISTING_RE_ASSESSMENT
        
        # Rule 1: Zero hub-metro direct service
        if not p.existing_service_hub_to_metro:
            return MarketMaturity.ZERO_HUB_METRO
        
        # Nearby airport has direct service?
        if p.nearby_airport_direct:
            # Rule 2: Catchment differentiation
            if p.catchment_overlap:
                return MarketMaturity.NEARBY_DIRECT_OVERLAP
            else:
                return MarketMaturity.NEARBY_DIRECT_DISTINCT
        
        # No direct to this specific pair, but hub serves metro
        # Distinguish between virgin nonstop and new unserved
        if p.existing_service_hub_to_metro and not p.nearby_airport_direct:
            # Hub flies to metro but no nearby airport direct = well-served indirectly
            return MarketMaturity.WELL_SERVED_INDIRECT
        
        # Truly new unserved market
        return MarketMaturity.NEW_UNSERVED
    
    # =========================================================================
    # STEP 2: FIND COMPARABLE CASES
    # =========================================================================
    
    def _find_comparables(self, p: RouteProfile, maturity: MarketMaturity) -> List[Dict]:
        """
        Find the most relevant cases from the calibration library.
        Scores each case on similarity to the proposed route.
        """
        scored = []
        
        for case in self.cases:
            score = 0
            relevance_notes = []
            
            # Hub status match (high weight)
            case_hub_status = case.get('hub_status', '')
            if case_hub_status == p.hub_status:
                score += 3
                relevance_notes.append("Same hub status")
            
            # Alliance match
            if case.get('alliance', '') == p.alliance:
                score += 1
                relevance_notes.append("Same alliance")
            
            # Carrier type match
            if case.get('carrier_type', '') == p.carrier_type:
                score += 1
                relevance_notes.append("Same carrier type")
            
            # Frequency band match
            case_freq = case.get('frequency')
            if case_freq and p.frequency:
                if abs(case_freq - p.frequency) <= 1:
                    score += 2
                    relevance_notes.append(f"Similar frequency ({case_freq}x)")
                elif abs(case_freq - p.frequency) <= 2:
                    score += 1
            
            # New route match
            case_new = case.get('new_route', not case.get('existing_service', False))
            if case_new == p.new_route:
                score += 2
                relevance_notes.append("Same route status (new/existing)")
            
            # Same destination city (very high relevance)
            if case.get('destination', '') == p.destination or case.get('destination_city', '') == p.destination_city:
                score += 4
                relevance_notes.append("Same destination")
            
            # Same origin region (if both European hubs, both Asian hubs, etc.)
            origin_region_case = self._get_region(case.get('origin', ''))
            origin_region_new = self._get_region(p.origin)
            if origin_region_case == origin_region_new and origin_region_case:
                score += 2
                relevance_notes.append(f"Same origin region ({origin_region_case})")
            
            # Has segment-level detail (more useful for comparison)
            if 'p2p_segments' in case:
                score += 1
            
            # ULR match
            if p.is_ulr and case.get('route_type', '') == 'ULR':
                score += 2
            
            # COVID-era penalty: forecasts from 2021-2024 used adapted methodology
            # due to disrupted base years and connections. Pre-2020 and 2025+
            # cases use the reliable original methodology and should be preferred.
            forecast_year = case.get('forecast_year', 0)
            if isinstance(forecast_year, int) and 2021 <= forecast_year <= 2024:
                score -= 3
                relevance_notes.append(f"COVID-era ({forecast_year})")
            elif isinstance(forecast_year, int) and forecast_year <= 2019:
                score += 1
                relevance_notes.append("Pre-COVID (reliable)")
            
            if score >= 3:  # minimum threshold
                scored.append({
                    'route_id': case.get('route_id', ''),
                    'score': score,
                    'relevance': '; '.join(relevance_notes),
                    'case': case,
                })
        
        # Sort by score descending, take top 5
        scored.sort(key=lambda x: -x['score'])
        return scored[:5]
    
    @staticmethod
    def _get_region(airport_code: str) -> str:
        """Map airport to region for comparability scoring."""
        european = {'LHR', 'CDG', 'AMS', 'FRA', 'MUC', 'DUB', 'KEF', 'MAD', 'FCO', 'ZRH', 'VIE', 'BRU', 'CPH', 'OSL', 'ARN', 'HEL'}
        asian = {'TPE', 'ICN', 'HKG', 'SIN', 'NRT', 'HND', 'CAN', 'PVG', 'PEK', 'BKK', 'KUL', 'DEL', 'BOM'}
        middle_east = {'DXB', 'DOH', 'AUH', 'IST', 'TLV'}
        north_america = {'SJC', 'SFO', 'LAX', 'JFK', 'ORD', 'ATL', 'DFW', 'IAH', 'SEA', 'BOS', 'TPA', 'MIA', 'YYZ', 'YVR'}
        
        if airport_code in european:
            return "Europe"
        elif airport_code in asian:
            return "Asia"
        elif airport_code in middle_east:
            return "Middle East"
        elif airport_code in north_america:
            return "North America"
        return ""
    
    # =========================================================================
    # STEP 3: PREDICT STIMULATION
    # =========================================================================
    
    def _predict_stimulation(self, p: RouteProfile, maturity: MarketMaturity,
                             comps: List[Dict]) -> Dict:
        """
        Predict stimulation factors by segment type.
        Returns dict with business/leisure_pri/leisure_sec/leisure_con values.
        """
        # Base ranges from CALIBRATION_PATTERNS by maturity class
        ranges = {
            MarketMaturity.ZERO_HUB_METRO: {
                'business': (1.45, 1.55), 'leisure_pri': (1.30, 1.40),
                'leisure_sec': (1.25, 1.35), 'leisure_con': (1.25, 1.35),
            },
            MarketMaturity.NEW_UNSERVED: {
                'business': (1.30, 1.40), 'leisure_pri': (1.25, 1.35),
                'leisure_sec': (1.20, 1.30), 'leisure_con': (1.15, 1.25),
            },
            MarketMaturity.VIRGIN_NONSTOP: {
                'business': (1.15, 1.30), 'leisure_pri': (1.10, 1.25),
                'leisure_sec': (1.05, 1.20), 'leisure_con': (1.00, 1.15),
            },
            MarketMaturity.WELL_SERVED_INDIRECT: {
                'business': (1.05, 1.15), 'leisure_pri': (1.05, 1.15),
                'leisure_sec': (1.00, 1.10), 'leisure_con': (1.00, 1.05),
            },
            MarketMaturity.NEARBY_DIRECT_DISTINCT: {
                # Rule 2: distinct catchment gets +0.05 above baseline
                'business': (1.10, 1.20), 'leisure_pri': (1.05, 1.15),
                'leisure_sec': (1.00, 1.10), 'leisure_con': (1.00, 1.10),
            },
            MarketMaturity.NEARBY_DIRECT_OVERLAP: {
                'business': (1.00, 1.05), 'leisure_pri': (1.00, 1.05),
                'leisure_sec': (1.00, 1.00), 'leisure_con': (1.00, 1.00),
            },
            MarketMaturity.EXISTING_RE_ASSESSMENT: {
                'business': (1.00, 1.15), 'leisure_pri': (1.00, 1.10),
                'leisure_sec': (1.00, 1.05), 'leisure_con': (1.00, 1.00),
            },
        }
        
        base = ranges.get(maturity, ranges[MarketMaturity.WELL_SERVED_INDIRECT])
        
        # Pick point estimates: use midpoint, adjusted by comparable cases
        result = {}
        for seg_type, (lo, hi) in base.items():
            mid = (lo + hi) / 2
            
            # If we have comps with segment data, use their values as anchor
            comp_values = self._extract_comp_stim(comps, seg_type)
            if comp_values:
                comp_avg = sum(comp_values) / len(comp_values)
                # More comps with higher relevance = more weight on comps
                comp_weight = min(0.75, 0.25 + 0.12 * len(comp_values))
                mid = (1.0 - comp_weight) * mid + comp_weight * comp_avg
            
            # Round to nearest 0.05
            mid = round(mid * 20) / 20
            # Allow slight range extension when comps pull outside
            effective_lo = lo - 0.05 if comp_values else lo
            effective_hi = hi + 0.05 if comp_values else hi
            mid = max(effective_lo, min(effective_hi, mid))
            
            result[seg_type] = {
                'value': mid,
                'range': (lo, hi),
                'comp_values': comp_values,
            }
        
        return result
    
    def _extract_comp_stim(self, comps: List[Dict], seg_type: str) -> List[float]:
        """Extract stimulation values for a segment type from comparable cases."""
        values = []
        for comp in comps:  # Check all comps
            case = comp['case']
            segs = case.get('p2p_segments', {})
            found = False
            for seg_name, seg_data in segs.items():
                if self._segment_type_match(seg_name, seg_type):
                    stim = seg_data.get('stimulation', seg_data.get('stim'))
                    if stim:
                        values.append(stim)
                        found = True
                        break
            # Fallback: blended stimulation if no segment match
            if not found and case.get('p2p_stimulation'):
                if seg_type == 'business':
                    values.append(case['p2p_stimulation'])
            if len(values) >= 5:
                break
        return values
    
    @staticmethod
    def _segment_type_match(seg_name: str, seg_type: str) -> bool:
        """Match segment name to type category."""
        seg_lower = seg_name.lower()
        if seg_type == 'business':
            return 'business' in seg_lower
        elif seg_type == 'leisure_pri':
            return ('primary' in seg_lower or 'leisure_pri' in seg_lower) and 'business' not in seg_lower
        elif seg_type == 'leisure_sec':
            return ('secondary' in seg_lower or 'leisure_sec' in seg_lower) and 'business' not in seg_lower
        elif seg_type == 'leisure_con':
            return ('contested' in seg_lower or 'leisure_con' in seg_lower) and 'business' not in seg_lower
        return False
    
    # =========================================================================
    # STEP 4: PREDICT CAPTURE RATES
    # =========================================================================
    
    def _predict_capture(self, p: RouteProfile, maturity: MarketMaturity,
                         comps: List[Dict]) -> Dict:
        """
        Predict capture rates by segment type.
        """
        ranges = {
            MarketMaturity.ZERO_HUB_METRO: {
                'business': (0.70, 0.80), 'leisure_pri': (0.55, 0.65),
                'leisure_sec': (0.55, 0.65), 'leisure_con': (0.35, 0.50),
            },
            MarketMaturity.NEW_UNSERVED: {
                'business': (0.35, 0.45), 'leisure_pri': (0.30, 0.40),
                'leisure_sec': (0.25, 0.35), 'leisure_con': (0.15, 0.25),
            },
            MarketMaturity.VIRGIN_NONSTOP: {
                'business': (0.25, 0.40), 'leisure_pri': (0.25, 0.35),
                'leisure_sec': (0.20, 0.30), 'leisure_con': (0.15, 0.25),
            },
            MarketMaturity.WELL_SERVED_INDIRECT: {
                'business': (0.20, 0.30), 'leisure_pri': (0.20, 0.30),
                'leisure_sec': (0.15, 0.25), 'leisure_con': (0.10, 0.20),
            },
            MarketMaturity.NEARBY_DIRECT_DISTINCT: {
                # Rule 2: +5-8pp above comparable hub benchmark
                'business': (0.35, 0.50), 'leisure_pri': (0.25, 0.40),
                'leisure_sec': (0.25, 0.35), 'leisure_con': (0.15, 0.25),
            },
            MarketMaturity.NEARBY_DIRECT_OVERLAP: {
                'business': (0.15, 0.25), 'leisure_pri': (0.15, 0.25),
                'leisure_sec': (0.10, 0.20), 'leisure_con': (0.05, 0.15),
            },
            MarketMaturity.EXISTING_RE_ASSESSMENT: {
                'business': (0.15, 0.30), 'leisure_pri': (0.15, 0.25),
                'leisure_sec': (0.10, 0.20), 'leisure_con': (0.10, 0.15),
            },
        }
        
        base = ranges.get(maturity, ranges[MarketMaturity.WELL_SERVED_INDIRECT])
        
        result = {}
        for seg_type, (lo, hi) in base.items():
            mid = (lo + hi) / 2
            
            # Adjust by comparable cases
            comp_values = self._extract_comp_capture(comps, seg_type)
            if comp_values:
                comp_avg = sum(comp_values) / len(comp_values)
                comp_weight = min(0.75, 0.25 + 0.12 * len(comp_values))
                mid = (1.0 - comp_weight) * mid + comp_weight * comp_avg
            
            # Rule 3: Secondary = Primary for strong hub carriers
            # (Applied later in _build_segments)
            
            # Round to nearest 5pp
            mid = round(mid * 20) / 20
            # Allow slight range extension when comps pull outside (max +5pp)
            effective_lo = lo - 0.05 if comp_values else lo
            effective_hi = hi + 0.05 if comp_values else hi
            mid = max(effective_lo, min(effective_hi, mid))
            
            result[seg_type] = {
                'value': mid,
                'range': (lo, hi),
                'comp_values': comp_values,
            }
        
        return result
    
    def _extract_comp_capture(self, comps: List[Dict], seg_type: str) -> List[float]:
        """Extract capture rates for a segment type from comparable cases."""
        values = []
        for comp in comps:  # Check all comps for segment data
            case = comp['case']
            segs = case.get('p2p_segments', {})
            for seg_name, seg_data in segs.items():
                if self._segment_type_match(seg_name, seg_type):
                    cap = seg_data.get('capture_rate', seg_data.get('capture'))
                    if cap:
                        values.append(cap)
                        break  # One value per case per segment type
            if len(values) >= 5:  # cap at 5 values
                break
        return values
    
    # =========================================================================
    # STEP 5: BUILD SEGMENT SUGGESTIONS
    # =========================================================================
    
    def _build_segments(self, p: RouteProfile, stim_params: Dict, cap_params: Dict,
                        maturity: MarketMaturity, comps: List[Dict]) -> Dict[str, SegmentSuggestion]:
        """
        Build the 8 P2P segment suggestions.
        Maps the 4 segment types (business, leisure_pri/sec/con) to origin and dest.
        """
        segments = {}
        
        # Determine origin/dest labels
        origin_label = p.origin_city or p.origin
        dest_label = p.destination_city or p.destination
        
        type_map = {
            f'{origin_label}_business': 'business',
            f'{origin_label}_leisure_primary': 'leisure_pri',
            f'{origin_label}_leisure_secondary': 'leisure_sec',
            f'{origin_label}_leisure_contested': 'leisure_con',
            f'{dest_label}_business': 'business',
            f'{dest_label}_leisure_primary': 'leisure_pri',
            f'{dest_label}_leisure_secondary': 'leisure_sec',
            f'{dest_label}_leisure_contested': 'leisure_con',
        }
        
        for seg_name, seg_type in type_map.items():
            stim_info = stim_params[seg_type]
            cap_info = cap_params[seg_type]
            
            stim_val = stim_info['value']
            cap_val = cap_info['value']
            
            # Rule 3: Secondary leisure = Primary for strong hub carriers
            if seg_type == 'leisure_sec' and self._is_strong_hub_carrier(p):
                pri_cap = cap_params['leisure_pri']['value']
                if cap_val < pri_cap:
                    cap_val = pri_cap
            
            # Rule 6: Flag capture asymmetry > 15pp between directions
            # (flagged in warnings, not auto-corrected)
            
            # Build reasoning
            stim_reason = f"Maturity={maturity.value}"
            if stim_info['comp_values']:
                stim_reason += f"; comps avg={sum(stim_info['comp_values'])/len(stim_info['comp_values']):.2f}"
            
            cap_reason = f"Maturity={maturity.value}"
            if cap_info['comp_values']:
                cap_reason += f"; comps avg={sum(cap_info['comp_values'])/len(cap_info['comp_values']):.2f}"
            
            # Confidence based on comp availability
            if len(stim_info.get('comp_values', [])) >= 2 and len(cap_info.get('comp_values', [])) >= 2:
                conf = Confidence.HIGH.value
            elif stim_info.get('comp_values') or cap_info.get('comp_values'):
                conf = Confidence.MEDIUM.value
            else:
                conf = Confidence.LOW.value
            
            segments[seg_name] = SegmentSuggestion(
                segment=seg_name,
                stimulation=stim_val,
                capture=cap_val,
                stim_range=stim_info['range'],
                capture_range=cap_info['range'],
                stim_reasoning=stim_reason,
                capture_reasoning=cap_reason,
                confidence=conf,
            )
        
        return segments
    
    def _is_strong_hub_carrier(self, p: RouteProfile) -> bool:
        """Rule 3 check: Is this a strong hub carrier where secondary = primary?"""
        strong_carriers = {'LH', 'BA', 'AF', 'KL', 'SQ', 'CX', 'KE', 'BR', 'CI', 'NH', 'JL'}
        return (p.carrier in strong_carriers and 
                p.hub_status in ('Major Hub', 'Secondary Hub'))
    
    # =========================================================================
    # STEP 6-7: BLENDED VALUES AND QSI ADJUSTMENT
    # =========================================================================
    
    def _blend_stim(self, segments: Dict[str, SegmentSuggestion], p: RouteProfile) -> float:
        """Compute demand-weighted blended stimulation."""
        if not segments:
            return 1.0
        # Weight by approximate demand share: business heavier if business-dominant
        biz_weight = p.business_leisure_split
        lei_weight = 1.0 - biz_weight
        
        stims = list(segments.values())
        biz_stims = [s.stimulation for s in stims if 'business' in s.segment]
        lei_stims = [s.stimulation for s in stims if 'business' not in s.segment]
        
        biz_avg = sum(biz_stims) / len(biz_stims) if biz_stims else 1.0
        lei_avg = sum(lei_stims) / len(lei_stims) if lei_stims else 1.0
        
        return round(biz_weight * biz_avg + lei_weight * lei_avg, 3)
    
    def _blend_capture(self, segments: Dict[str, SegmentSuggestion], p: RouteProfile) -> float:
        """Compute demand-weighted blended capture."""
        if not segments:
            return 0.25
        biz_weight = p.business_leisure_split
        lei_weight = 1.0 - biz_weight
        
        caps = list(segments.values())
        biz_caps = [s.capture for s in caps if 'business' in s.segment]
        lei_caps = [s.capture for s in caps if 'business' not in s.segment]
        
        biz_avg = sum(biz_caps) / len(biz_caps) if biz_caps else 0.25
        lei_avg = sum(lei_caps) / len(lei_caps) if lei_caps else 0.25
        
        return round(biz_weight * biz_avg + lei_weight * lei_avg, 3)
    
    def _predict_qsi_adjustment(self, p: RouteProfile, maturity: MarketMaturity) -> Tuple[float, str]:
        """
        Predict QSI adjustment factor.
        Pattern: ALL 21 new-route cases in library = 1.0
        Only existing re-assessment (BA LHR-SJC) needed heavy calibration.
        """
        if maturity == MarketMaturity.EXISTING_RE_ASSESSMENT:
            return 0.267, ("Existing service re-assessment. Raw QSI overestimates connecting ~5x. "
                          "Median factor from BA LHR-SJC = 0.267. REQUIRES expert calibration per city.")
        else:
            return 1.0, "New route proposal  all 21 library cases accepted raw QSI (factor = 1.0)"
    
    # =========================================================================
    # STEP 8: GROWTH RATE
    # =========================================================================
    
    def _predict_growth(self, p: RouteProfile, comps: List[Dict]) -> Tuple[float, str]:
        """Predict P2P demand growth rate to apply to base demand."""
        # Extract growth rates from comparable cases
        comp_growths = []
        for comp in comps[:3]:
            case = comp['case']
            g = case.get('p2p_growth_rate')
            if g:
                comp_growths.append(g)
        
        if comp_growths:
            avg = sum(comp_growths) / len(comp_growths)
            return round(avg, 3), f"Based on {len(comp_growths)} comparable cases (avg {avg:.1%})"
        
        # Default by market type
        if p.primary_demand == 'Business':
            return 0.10, "Default business-dominated market growth (10%)"
        elif p.primary_demand == 'Leisure':
            return 0.08, "Default leisure-dominated market growth (8%)"
        else:
            return 0.10, "Default mixed market growth (10%)"
    
    # =========================================================================
    # STEP 9: CONNECTING SHARE EXPECTATIONS
    # =========================================================================
    
    def _predict_cnx_share(self, p: RouteProfile, maturity: MarketMaturity,
                           comps: List[Dict]) -> Tuple[Tuple[float, float], str]:
        """
        Predict expected connecting traffic share range.
        Per Rule 4, this is a RANGE expectation, not a point estimate.
        The actual connecting QSI must come from the pipeline.
        """
        hub = p.hub_status
        
        if hub == 'Non-Hub' or p.hub_airport not in (p.origin, p.destination):
            return (0.0, 0.05), "Non-hub route  minimal connecting traffic expected"
        
        # Check comparable connecting shares
        comp_cnx = []
        for comp in comps[:3]:
            case = comp['case']
            mix = case.get('traffic_mix', {})
            if mix:
                home_pct = mix.get('cnx_home_pct', 0)
                dest_pct = mix.get('cnx_dest_pct', 0)
                comp_cnx.append(home_pct + dest_pct)
            else:
                # Try to compute from totals
                total = case.get('grand_total_forecast', 0)
                cnx_h = case.get('cnx_home_forecast', 0)
                cnx_d = case.get('cnx_dest_forecast', 0)
                if total and (cnx_h or cnx_d):
                    comp_cnx.append((cnx_h + cnx_d) / total)
        
        if hub == 'Major Hub':
            if p.is_ulr:
                base_range = (0.35, 0.50)
                reason = "Major hub ULR  connecting depressed by route length"
            else:
                base_range = (0.40, 0.65)
                reason = "Major hub  typical connecting share 40-65%"
        elif hub == 'Secondary Hub':
            base_range = (0.30, 0.65)
            reason = "Secondary hub  wide range depending on hub network strength"
        else:
            base_range = (0.0, 0.10)
            reason = "Non-hub or weak hub"
        
        # Adjust by departure time (Rule 7)
        if p.departure_time_home:
            hour = self._parse_hour(p.departure_time_home)
            if hour is not None:
                if 18 <= hour <= 21:
                    reason += "; evening departure optimal for connecting"
                elif 22 <= hour or hour <= 2:
                    reason += "; late-night departure  connecting 30-37% (between evening and midday)"
                    base_range = (max(base_range[0], 0.25), min(base_range[1], 0.45))
                elif 10 <= hour <= 15:
                    reason += "; midday departure  connecting share depressed"
                    base_range = (max(base_range[0] - 0.10, 0.10), base_range[1] - 0.10)
        
        if comp_cnx:
            comp_avg = sum(comp_cnx) / len(comp_cnx)
            reason += f"; comps avg={comp_avg:.0%}"
        
        return base_range, reason
    
    @staticmethod
    def _parse_hour(time_str: str) -> Optional[int]:
        """Parse HH:MM to hour integer."""
        try:
            parts = time_str.split(':')
            return int(parts[0])
        except (ValueError, IndexError):
            return None
    
    # =========================================================================
    # STEP 9b: SUGGESTED LOAD FACTOR RANGE
    # =========================================================================

    def _predict_lf_range(self, p: RouteProfile, maturity: MarketMaturity,
                          comps: List[Dict]) -> Tuple[Tuple[float, float], str]:
        """
        Predict an appropriate load factor range for this route.
        
        Airlines find LFs of 60-89% credible for new routes. Anything above
        ~85% on a new service looks artificially high; anything below 60%
        suggests the route is marginal. Mature re-assessments can be higher.
        
        Learn from comparable cases if available.
        """
        # Default ranges by maturity
        lf_defaults = {
            MarketMaturity.ZERO_HUB_METRO:        (0.70, 0.82),
            MarketMaturity.NEW_UNSERVED:           (0.65, 0.80),
            MarketMaturity.VIRGIN_NONSTOP:         (0.65, 0.80),
            MarketMaturity.WELL_SERVED_INDIRECT:   (0.65, 0.78),
            MarketMaturity.NEARBY_DIRECT_OVERLAP:  (0.60, 0.75),
            MarketMaturity.NEARBY_DIRECT_DISTINCT: (0.65, 0.78),
            MarketMaturity.EXISTING_RE_ASSESSMENT: (0.75, 0.88),
        }
        
        lf_low, lf_high = lf_defaults.get(maturity, (0.65, 0.82))
        reason_parts = [f"Default for {maturity.value}: {lf_low:.0%}-{lf_high:.0%}"]
        
        # LCC routes can sustain higher LFs (tighter revenue management)
        carrier_type = getattr(p, 'carrier_type', 'Full Service')
        if carrier_type and 'lcc' in str(carrier_type).lower():
            lf_low = min(lf_low + 0.05, 0.85)
            lf_high = min(lf_high + 0.05, 0.92)
            reason_parts.append(f"LCC adjustment: +5pp -> {lf_low:.0%}-{lf_high:.0%}")
        
        # Hub longhaul routes with large connecting pools tend toward higher LFs
        route_type = getattr(p, 'route_type', '')
        if route_type and 'hub' in str(route_type).lower() and 'longhaul' in str(route_type).lower():
            lf_low = min(lf_low + 0.03, 0.82)
            lf_high = min(lf_high + 0.03, 0.88)
            reason_parts.append(f"Hub longhaul: +3pp")
        
        # Learn from comparable cases
        comp_lfs = []
        for c in comps:
            case = c.get('case', {})
            # If calibration library stores actual LF, use it
            actual_lf = case.get('actual_load_factor') or case.get('load_factor')
            if actual_lf and 0.30 < actual_lf < 1.0:
                comp_lfs.append(actual_lf)
        
        if comp_lfs:
            comp_low = min(comp_lfs)
            comp_high = max(comp_lfs)
            comp_avg = sum(comp_lfs) / len(comp_lfs)
            # Blend: 60% default, 40% comparable evidence
            lf_low = 0.6 * lf_low + 0.4 * max(comp_low - 0.05, 0.50)
            lf_high = 0.6 * lf_high + 0.4 * min(comp_high + 0.05, 0.92)
            reason_parts.append(
                f"Comparable LFs: {', '.join(f'{x:.0%}' for x in comp_lfs)} "
                f"(avg {comp_avg:.0%})"
            )
        
        return (round(lf_low, 2), round(lf_high, 2)), "; ".join(reason_parts)

    # =========================================================================
    # STEP 10: SPECIAL RULES AND WARNINGS
    # =========================================================================
    
    def _apply_special_rules(self, p: RouteProfile, suggestion: CalibrationSuggestion,
                             maturity: MarketMaturity, comps: List[Dict]):
        """Apply the 8 calibration rules and generate warnings."""
        
        # Rule 1: Zero hub-metro direct service
        if maturity == MarketMaturity.ZERO_HUB_METRO:
            suggestion.rules_applied.append(
                "Rule 1 (Zero Hub-Metro): Hub has no direct service to any airport in "
                "destination metro area. Elevated stim (1.45-1.55) and capture (70-80% bus). "
                "Evidence: LH MUC-SJC."
            )
        
        # Rule 2: Catchment differentiation
        if maturity == MarketMaturity.NEARBY_DIRECT_DISTINCT:
            suggestion.rules_applied.append(
                "Rule 2 (Catchment Differentiation): Nearby airport has direct service but "
                "catchments are distinct. Stim +0.05 and capture +5-8pp above baseline. "
                "Evidence: LH FRA-SJC, CZ CAN-SJC."
            )
        
        # Rule 3: Secondary = Primary for strong hub carriers
        if self._is_strong_hub_carrier(p):
            suggestion.rules_applied.append(
                "Rule 3 (Secondary = Primary): Strong hub carrier  secondary leisure capture "
                "set equal to primary (no discount). Evidence: LH FRA-SJC, LH MUC-SJC."
            )
        
        # Rule 4: Connecting QSI must be computed
        suggestion.rules_applied.append(
            "Rule 4 (Connecting Must Be Computed): Connecting QSI share from pipeline, "
            "not benchmarked. Hub size is poor predictor  MUC (5.2%) beat FRA (3.6%)."
        )
        
        # Rule 5: Vintage-based parameter selection
        if p.data_vintage:
            suggestion.rules_applied.append(
                f"Rule 5 (Vintage): Data vintage {p.data_vintage}. Use most recent library "
                "cases for default parameter selection."
            )
        
        # Rule 6: Capture asymmetry check
        # Check if origin/dest capture differs by >15pp
        origin_biz = None
        dest_biz = None
        for name, seg in suggestion.segments.items():
            if 'business' in name:
                if p.origin_city.lower() in name.lower() or p.origin.lower() in name.lower():
                    origin_biz = seg.capture
                else:
                    dest_biz = seg.capture
        
        if origin_biz and dest_biz and abs(origin_biz - dest_biz) > 0.15:
            suggestion.warnings.append(
                f"Rule 6: Capture asymmetry detected ({origin_biz:.0%} vs {dest_biz:.0%}). "
                "Consider standardising to symmetric average."
            )
            suggestion.rules_applied.append("Rule 6 (Asymmetry Flag): Large capture difference between directions flagged.")
        
        # Rule 7: Departure time
        if p.departure_time_home:
            hour = self._parse_hour(p.departure_time_home)
            if hour is not None and (22 <= hour or hour <= 2):
                suggestion.rules_applied.append(
                    "Rule 7 (Late-Night Connecting): Departure after 22:00  connecting share "
                    "30-37%, between evening optimal and midday minimum."
                )
        
        # Rule 8: Hub-to-non-hub repurposing
        if p.hub_status == 'Non-Hub':
            suggestion.warnings.append(
                "Rule 8: Non-hub carrier  P2P forecast from hub carrier analysis can be reused, "
                "but set connecting = 0. Check LF viability without connecting traffic."
            )
            suggestion.rules_applied.append("Rule 8 (HubNon-Hub Repurposing): P2P component only, no connecting.")
        
        # Additional warnings
        if p.frequency and p.frequency >= 7 and p.seats and p.seats > 250:
            suggestion.warnings.append(
                "High capacity warning: Daily service with large aircraft. "
                "Library shows LF deteriorates significantly with excess frequency "
                "(EI DUB 4x7x: 70%47% LF)."
            )
        
        if p.is_ulr:
            suggestion.warnings.append(
                "ULR route: Only one library case (SQ SIN-SJC). Connecting share may be "
                "suppressed by route length despite major hub status."
            )
    
    # =========================================================================
    # STEP 11: CONFIDENCE ASSESSMENT
    # =========================================================================
    
    def _assess_confidence(self, p: RouteProfile, maturity: MarketMaturity,
                           comps: List[Dict]) -> str:
        """Assess overall prediction confidence."""
        # Count high-scoring comps
        good_comps = [c for c in comps if c['score'] >= 6]
        
        if len(good_comps) >= 3:
            return Confidence.HIGH.value
        elif len(good_comps) >= 1 or len(comps) >= 3:
            return Confidence.MEDIUM.value
        else:
            return Confidence.LOW.value


# =============================================================================
# CONVENIENCE: Quick prediction from minimal inputs
# =============================================================================

def quick_predict(origin: str, destination: str, carrier: str = "",
                  hub_status: str = "Major Hub", frequency: int = 5,
                  seats: int = 280, new_route: bool = True,
                  hub_to_metro_direct: bool = True,
                  nearby_direct: bool = False,
                  catchment_overlap: bool = True,
                  business_split: float = 0.55,
                  **kwargs) -> CalibrationSuggestion:
    """
    Quick prediction with minimal inputs.
    
    Example:
        result = quick_predict('FRA', 'SJC', 'LH', hub_to_metro_direct=True,
                               nearby_direct=True, catchment_overlap=False)
    """
    profile = RouteProfile(
        origin=origin, destination=destination, carrier=carrier,
        hub_status=hub_status, frequency=frequency, seats=seats,
        new_route=new_route,
        existing_service_hub_to_metro=hub_to_metro_direct,
        nearby_airport_direct=nearby_direct,
        catchment_overlap=catchment_overlap,
        business_leisure_split=business_split,
        **kwargs,
    )
    engine = PredictiveCalibrationEngine()
    return engine.predict(profile)


# =============================================================================
# VALIDATION: Test against known blind forecast results
# =============================================================================

def validate_against_blind_tests():
    """
    Test the engine's predictions against the three completed blind
    forecast tests from the calibration library development.
    """
    print("="*70)
    print("PREDICTIVE CALIBRATION ENGINE  VALIDATION")
    print("="*70)
    
    results = {}
    
    # -----------------------------------------------------------------------
    # TEST 1: CI TPE-SJC (Chat 43 blind test  target 0.6% error)
    # -----------------------------------------------------------------------
    print("\n--- TEST 1: CI TPE-SJC ---")
    ci = RouteProfile(
        origin='TPE', destination='SJC', carrier='CI', carrier_name='China Airlines',
        alliance='SkyTeam', carrier_type='Full Service',
        hub_airport='TPE', hub_status='Major Hub',
        frequency=4, aircraft='A350-900', seats=306,
        new_route=True, existing_service_hub_to_metro=True,
        nearby_airport_direct=True,   # CI TPE-SFO exists
        catchment_overlap=False,      # SJC distinct from SFO
        business_leisure_split=0.55,
        data_vintage='2019',
    )
    
    engine = PredictiveCalibrationEngine()
    ci_result = engine.predict(ci)
    ci_result.print_summary()
    
    # Actual analyst values: stim ~1.10 blended, capture ~30%, QSI adj = 1.0
    actual_stim = 1.10
    actual_capture = 0.30
    print(f"\n  Actual analyst: stim={actual_stim}, capture={actual_capture:.0%}")
    print(f"  Predicted:      stim={ci_result.blended_stimulation:.2f}, capture={ci_result.blended_capture:.0%}")
    stim_err = abs(ci_result.blended_stimulation - actual_stim) / actual_stim * 100
    cap_err = abs(ci_result.blended_capture - actual_capture) / actual_capture * 100
    print(f"  Stim error: {stim_err:.1f}%, Capture error: {cap_err:.1f}%")
    results['CI_TPE_SJC'] = {'stim_err': stim_err, 'cap_err': cap_err}
    
    # -----------------------------------------------------------------------
    # TEST 2: LH FRA-SJC (Chat 47 blind test  target 0.4% error with rules)
    # -----------------------------------------------------------------------
    print("\n--- TEST 2: LH FRA-SJC ---")
    fra = RouteProfile(
        origin='FRA', destination='SJC', carrier='LH', carrier_name='Lufthansa',
        alliance='Star Alliance', carrier_type='Full Service',
        hub_airport='FRA', hub_status='Major Hub',
        frequency=5, aircraft='A340-300', seats=267,
        new_route=True, existing_service_hub_to_metro=True,
        nearby_airport_direct=True,    # LH+UA FRA-SFO exists
        catchment_overlap=False,       # SJC distinct from SFO
        business_leisure_split=0.60,
        data_vintage='2019',
    )
    
    fra_result = engine.predict(fra)
    fra_result.print_summary()
    
    # Actual: stim ~1.12, capture ~42% blended
    actual_stim_fra = 1.12
    actual_capture_fra = 0.421
    print(f"\n  Actual analyst: stim={actual_stim_fra}, capture={actual_capture_fra:.1%}")
    print(f"  Predicted:      stim={fra_result.blended_stimulation:.3f}, capture={fra_result.blended_capture:.1%}")
    results['LH_FRA_SJC'] = {
        'stim_err': abs(fra_result.blended_stimulation - actual_stim_fra) / actual_stim_fra * 100,
        'cap_err': abs(fra_result.blended_capture - actual_capture_fra) / actual_capture_fra * 100,
    }
    
    # -----------------------------------------------------------------------
    # TEST 3: LH MUC-SJC (Chat 47  Rule 1 zero-metro test)
    # -----------------------------------------------------------------------
    print("\n--- TEST 3: LH MUC-SJC ---")
    muc = RouteProfile(
        origin='MUC', destination='SJC', carrier='LH', carrier_name='Lufthansa',
        alliance='Star Alliance', carrier_type='Full Service',
        hub_airport='MUC', hub_status='Secondary Hub',
        frequency=5, aircraft='A350-900', seats=293,
        new_route=True,
        existing_service_hub_to_metro=False,  # NO MUC-Bay Area service at all!
        nearby_airport_direct=False,
        catchment_overlap=False,
        business_leisure_split=0.55,
        data_vintage='2019',
    )
    
    muc_result = engine.predict(muc)
    muc_result.print_summary()
    
    # Actual: stim ~1.43 blended, capture ~69% blended
    actual_stim_muc = 1.43
    actual_capture_muc = 0.69
    print(f"\n  Actual analyst: stim={actual_stim_muc}, capture={actual_capture_muc:.0%}")
    print(f"  Predicted:      stim={muc_result.blended_stimulation:.3f}, capture={muc_result.blended_capture:.0%}")
    results['LH_MUC_SJC'] = {
        'stim_err': abs(muc_result.blended_stimulation - actual_stim_muc) / actual_stim_muc * 100,
        'cap_err': abs(muc_result.blended_capture - actual_capture_muc) / actual_capture_muc * 100,
    }
    
    # -----------------------------------------------------------------------
    # TEST 4: BR TPE-SJC Aug19B (Chat 49  target 0.0% error)
    # -----------------------------------------------------------------------
    print("\n--- TEST 4: BR TPE-SJC Aug19B ---")
    br = RouteProfile(
        origin='TPE', destination='SJC', carrier='BR', carrier_name='EVA Air',
        alliance='Star Alliance', carrier_type='Full Service',
        hub_airport='TPE', hub_status='Major Hub',
        frequency=5, aircraft='A350', seats=304,
        new_route=True, existing_service_hub_to_metro=True,
        nearby_airport_direct=True,   # BR TPE-SFO exists
        catchment_overlap=False,
        business_leisure_split=0.55,
        data_vintage='2019',
    )
    
    br_result = engine.predict(br)
    br_result.print_summary()
    
    # Actual: stim 1.10 blended, capture 40% blended
    actual_stim_br = 1.10
    actual_capture_br = 0.40
    print(f"\n  Actual analyst: stim={actual_stim_br}, capture={actual_capture_br:.0%}")
    print(f"  Predicted:      stim={br_result.blended_stimulation:.3f}, capture={br_result.blended_capture:.0%}")
    results['BR_TPE_SJC'] = {
        'stim_err': abs(br_result.blended_stimulation - actual_stim_br) / actual_stim_br * 100,
        'cap_err': abs(br_result.blended_capture - actual_capture_br) / actual_capture_br * 100,
    }
    
    # -----------------------------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------------------------
    print("\n" + "="*70)
    print("VALIDATION SUMMARY")
    print("="*70)
    print(f"{'Route':<20s} {'Stim Error':>12s} {'Capture Error':>14s}")
    print("-"*46)
    for route, errs in results.items():
        print(f"{route:<20s} {errs['stim_err']:>11.1f}% {errs['cap_err']:>13.1f}%")
    
    avg_stim = sum(r['stim_err'] for r in results.values()) / len(results)
    avg_cap = sum(r['cap_err'] for r in results.values()) / len(results)
    print(f"\n{'Average':<20s} {avg_stim:>11.1f}% {avg_cap:>13.1f}%")
    
    # Check relative ranking: MUC should beat FRA
    if 'LH_FRA_SJC' in results and 'LH_MUC_SJC' in results:
        fra_cap = fra_result.blended_capture
        muc_cap = muc_result.blended_capture
        if muc_cap > fra_cap:
            print("\n   Relative ranking CORRECT: MUC capture > FRA capture")
        else:
            print("\n   Relative ranking WRONG: MUC should beat FRA")
    
    return results


# =============================================================================
# INTEGRATION: Convert RouteConfig to RouteProfile
# =============================================================================

def from_route_config(cfg) -> RouteProfile:
    """
    Create a RouteProfile from an existing RouteConfig object.
    This bridges the pipeline's config system with the prediction engine.
    
    The analyst still needs to set the market characterisation booleans
    (existing_service_hub_to_metro, nearby_airport_direct, catchment_overlap)
    as these require market knowledge.
    """
    return RouteProfile(
        origin=getattr(cfg, 'home_airport_code', ''),
        destination=getattr(cfg, 'dest_airport_code', ''),
        origin_city=getattr(cfg, 'home_city_code', ''),
        destination_city=getattr(cfg, 'dest_city_code', ''),
        carrier=getattr(cfg, 'airline_code', ''),
        carrier_name=getattr(cfg, 'airline_name', ''),
        alliance=getattr(cfg, 'alliance', ''),
        carrier_type=getattr(cfg, 'carrier_type', 'Full Service'),
        hub_airport=getattr(cfg, 'home_airport_code', ''),
        hub_status=getattr(cfg, 'hub_status', 'Major Hub'),
        frequency=getattr(cfg, 'frequency', 5),
        aircraft=getattr(cfg, 'aircraft_type', ''),
        seats=getattr(cfg, 'seats', 280),
        new_route=not getattr(cfg, 'existing_service', False),
    )


# =============================================================================
# EXCEL OUTPUT: Generate calibration suggestion workbook
# =============================================================================

def write_suggestion_excel(suggestion: CalibrationSuggestion, filepath: str):
    """
    Write the calibration suggestion to a branded Excel workbook.
    Designed for analyst review and sign-off.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not available  skipping Excel output")
        return
    
    wb = Workbook()
    
    # -- Colours --
    AVIA_BLUE = PatternFill('solid', fgColor='1F4E79')
    LIGHT_BLUE = PatternFill('solid', fgColor='D6E4F0')
    WHITE = PatternFill('solid', fgColor='FFFFFF')
    GREEN = PatternFill('solid', fgColor='C6EFCE')
    YELLOW = PatternFill('solid', fgColor='FFEB9C')
    RED = PatternFill('solid', fgColor='FFC7CE')
    
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    BOLD = Font(bold=True, size=11)
    NORMAL = Font(size=11)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "Calibration Suggestion"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Predictive Calibration: {suggestion.route_label}"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = AVIA_BLUE
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    for label, value in [
        ("Market Maturity", suggestion.market_maturity),
        ("Confidence", suggestion.confidence),
        ("Blended Stimulation", f"{suggestion.blended_stimulation:.3f}"),
        ("Blended Capture Rate", f"{suggestion.blended_capture:.1%}"),
        ("QSI Adjustment", f"{suggestion.qsi_adjustment:.2f}"),
        ("Growth Rate", f"{suggestion.growth_rate:.1%}"),
        ("Expected Cnx Share", f"{suggestion.cnx_share_range[0]:.0%} - {suggestion.cnx_share_range[1]:.0%}"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=value).font = NORMAL
        row += 1
    
    # Segment table
    row += 1
    ws.cell(row=row, column=1, value="Segment").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = AVIA_BLUE
    ws.cell(row=row, column=2, value="Stimulation").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = AVIA_BLUE
    ws.cell(row=row, column=3, value="Capture").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = AVIA_BLUE
    ws.cell(row=row, column=4, value="Confidence").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = AVIA_BLUE
    
    row += 1
    for name, seg in suggestion.segments.items():
        ws.cell(row=row, column=1, value=name).font = NORMAL
        ws.cell(row=row, column=2, value=seg.stimulation).font = NORMAL
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=3, value=seg.capture).font = NORMAL
        ws.cell(row=row, column=3).number_format = '0.0%'
        
        conf_cell = ws.cell(row=row, column=4, value=seg.confidence)
        if seg.confidence == 'HIGH':
            conf_cell.fill = GREEN
        elif seg.confidence == 'MEDIUM':
            conf_cell.fill = YELLOW
        else:
            conf_cell.fill = RED
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    
    # Rules applied
    row += 1
    ws.cell(row=row, column=1, value="Rules Applied").font = BOLD
    ws.cell(row=row, column=1).fill = LIGHT_BLUE
    row += 1
    for rule in suggestion.rules_applied:
        ws.cell(row=row, column=1, value=rule).font = NORMAL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    
    # Warnings
    if suggestion.warnings:
        row += 1
        ws.cell(row=row, column=1, value="Warnings").font = BOLD
        ws.cell(row=row, column=1).fill = YELLOW
        row += 1
        for w in suggestion.warnings:
            ws.cell(row=row, column=1, value=w).font = NORMAL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
    
    # === Sheet 2: Analyst Override ===
    ws2 = wb.create_sheet("Analyst Override")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 40
    
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "Analyst Parameter Override Sheet"
    ws2['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws2['A1'].fill = AVIA_BLUE
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    headers = ["Segment", "Engine Suggestion", "Analyst Override", "Justification"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = AVIA_BLUE
    
    row = 4
    for name, seg in suggestion.segments.items():
        ws2.cell(row=row, column=1, value=f"{name}  Stim").font = NORMAL
        ws2.cell(row=row, column=2, value=seg.stimulation).font = NORMAL
        ws2.cell(row=row, column=2).number_format = '0.00'
        ws2.cell(row=row, column=3).fill = YELLOW  # Override cell
        ws2.cell(row=row, column=4).fill = YELLOW  # Justification
        for col in range(1, 5):
            ws2.cell(row=row, column=col).border = thin_border
        row += 1
        
        ws2.cell(row=row, column=1, value=f"{name}  Capture").font = NORMAL
        ws2.cell(row=row, column=2, value=seg.capture).font = NORMAL
        ws2.cell(row=row, column=2).number_format = '0.0%'
        ws2.cell(row=row, column=3).fill = YELLOW
        ws2.cell(row=row, column=4).fill = YELLOW
        for col in range(1, 5):
            ws2.cell(row=row, column=col).border = thin_border
        row += 1
    
    # Reviewer sign-off
    row += 2
    ws2.cell(row=row, column=1, value="Reviewed by:").font = BOLD
    ws2.cell(row=row, column=2).fill = YELLOW
    row += 1
    ws2.cell(row=row, column=1, value="Date:").font = BOLD
    ws2.cell(row=row, column=2).fill = YELLOW
    row += 1
    ws2.cell(row=row, column=1, value="Notes:").font = BOLD
    ws2.merge_cells(start_row=row, start_column=2, end_row=row+2, end_column=4)
    ws2.cell(row=row, column=2).fill = YELLOW
    
    wb.save(filepath)
    print(f"Saved: {filepath}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    validate_against_blind_tests()
