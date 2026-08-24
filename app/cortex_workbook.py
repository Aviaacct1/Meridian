#!/usr/bin/env python3
"""
Avia Cortex - detailed route workbook (built straight from the calibrated forecast).
====================================================================================
build_workbook(out_path, fc, meta) writes the client-ready Excel with the full tables that
sit behind the deck: the forecast breakdown, the connecting-feed detail each way (PTEW), the
catchment split, the turnaround P&L line items, and an assumptions / methodology log. It reads
the calibrated_forecast() dict directly, so every number matches the portal, and needs nothing
from the old pipeline. Author is set to Avia Solutions.

  fc   = the calibrated_forecast() output dict (ok, origin, dest, demand, capacity, catchment,
         economics{raw}, distance_nm, block_min, week, year)
  meta = dict(airline_name, analyst, date, plan_lf, capture_basis, econ_fare)
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import fare_bands   # R5: measured fares leave as bands; the grid lives there
import attribution  # the one source line (R3); the Departure curve sheet prints it

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def carried_split(dem):
    """(p2p, behind, beyond, total) each way, on the CARRIED allocation, so any
    surface printing the three legs beside the total shows rows that sum to it.
    Prefers the payload's own carried split; an older payload falls back to scaling
    the uncapped legs pro-rata to the total, the same allocation
    forecast_to_contract renders, so the surfaces can never disagree."""
    n = lambda v: float(v or 0)
    p2p = n(dem.get("captured"))
    behind, beyond = n(dem.get("feed_behind")), n(dem.get("feed_beyond"))
    tot = n(dem.get("total"))
    p2p_c, conn_c = n(dem.get("p2p_carried")), n(dem.get("connecting_carried"))
    if p2p_c > 0 or conn_c > 0:
        p2p = p2p_c if p2p_c > 0 else p2p
        fsum = behind + beyond
        if conn_c > 0 and fsum > 0:
            behind, beyond = behind * conn_c / fsum, beyond * conn_c / fsum
    elif tot > 0 and (p2p + behind + beyond) > tot:
        sc = tot / (p2p + behind + beyond)
        p2p, behind, beyond = p2p * sc, behind * sc, beyond * sc
    return p2p, behind, beyond, tot


def _curve_series(fc):
    """The departure-time curve, on the CARRIED allocation, shared by the Excel sheet, the
    dashboard's own drawCurve() JS and (24 August 2026) the standalone PNG (John Carter's
    curve-picture request) - one calculation, three renderings, so all three always show the
    same numbers for the same run rather than three copies of the same maths drifting apart.
    Mirrors the dashboard slider's own transform: anchor on the curve point nearest the chosen
    departure, scale the raw connecting scores by connecting_carried at that anchor, hold
    point-to-point constant across the day, cap the route total at annual capacity x the plan
    cap. Returns None if there is no optimiser curve to plot - never a fabricated one - else a
    dict with everything a renderer needs, each-way throughout (multiply by 2 for two-way)."""
    dem = fc["demand"]; cap = fc["capacity"]
    sch = fc.get("schedule") or {}
    op = sch.get("optimised") or {}
    curve = op.get("curve") or []
    n0 = lambda v: float(v or 0)
    p2p_c = n0(dem.get("p2p_carried")); conn_c = n0(dem.get("connecting_carried"))
    cap_ew = n0(cap.get("annual_capacity")) * n0(cap.get("plan_cap"))
    chosen = ((sch.get("outbound") or {}).get("dep") or "")
    parts = str(chosen).split(":")
    cmins = (int(parts[0]) * 60 + int(parts[1])) if len(parts) == 2 and parts[0].isdigit() else None
    if len(curve) < 4 or conn_c <= 0 or p2p_c <= 0 or cmins is None:
        return None
    anchor = min(curve, key=lambda q: abs(float(q.get("dep") or 0) - cmins))
    at = float(anchor.get("total") or 0)
    if at <= 0:
        return None
    scl = conn_c / at
    points = []
    for p in curve:
        conn_ew = float(p.get("total") or 0) * scl
        demand_ew = p2p_c + conn_ew   # the UNCAPPED total: what the market would carry with
                                      # no aircraft ceiling, before spill (John Carter, 24
                                      # August 2026: the capped total line alone hides how
                                      # much demand a poor departure time actually spills)
        tot_ew = min(demand_ew, cap_ew) if cap_ew > 0 else demand_ew
        points.append({"dep": float(p.get("dep") or 0), "hhmm": p.get("hhmm") or "",
                        "permitted": bool(p.get("permitted")),
                        "beyond_ew": float(p.get("beyond") or 0) * scl,
                        "behind_ew": float(p.get("behind") or 0) * scl,
                        "connecting_ew": conn_ew, "total_ew": tot_ew, "demand_ew": demand_ew})
    return {"points": points, "p2p_ew": p2p_c, "cap_ew": cap_ew, "chosen": chosen,
            "chosen_mins": cmins, "unrestricted_dep": op.get("unrestricted_dep"),
            "restricted": op.get("restricted") or [],
            "forecast_year": sch.get("forecast_year") or fc.get("year", "")}


def render_curve_png(fc, meta, out_path):
    """The departure-curve picture (John Carter, 24 August 2026: running a batch of EVA/CI/JX
    forecasts at several frequencies and wanting the curve as a picture for each one, not a
    workbook chart he has to screenshot). Built with matplotlib, not openpyxl's native chart -
    shaded restricted-hour bands and an annotated callout are not something openpyxl's chart
    object can do without hand-editing chart XML, so this uses the one charting library
    actually available rather than fighting the Excel chart model into a shape it does not
    support. Saved as "<out_path base>_curve.png" alongside the workbook; only when the same
    _curve_series() data used by the Excel sheet exists - never a fabricated picture, same
    discipline as the sheet it sits beside. Returns the PNG path, or None if there was no
    curve to draw."""
    cs = _curve_series(fc)
    if not cs:
        return None
    # NOT matplotlib.pyplot (24 August 2026, second live-run miss after a first fix that
    # forced the Agg backend and still didn't hold under a real batch). api_report_start
    # runs every download in its own background thread (threading.Thread, cortex_app.py),
    # and pyplot's state - the current-figure registry, plt.subplots(), matplotlib.use()
    # itself - is one set of global variables shared by the whole process, not one per
    # thread. Two downloads built close together, exactly what a batch run is, can have
    # two threads mutating that same global state inside the same few milliseconds: one
    # thread's matplotlib.use() or plt.subplots() can invalidate the figure another thread
    # is mid-render on, with no guarantee either side raises loud enough to be caught.
    # The Figure/FigureCanvasAgg pair below is matplotlib's own documented route round
    # this - each call gets its own Figure object, own canvas, nothing global, nothing to
    # race. No backend to select either, so today's earlier force=True fix is now moot.
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    import matplotlib.dates as mdates
    import datetime as _dt

    o = fc["origin"]; home = fc["catchment"]["home"]
    mult = 2  # the picture is always two-way, matching the headline figures a client reads
    pts = cs["points"]
    times = [_dt.datetime(2000, 1, 1) + _dt.timedelta(minutes=p["dep"]) for p in pts]
    conn = [p["connecting_ew"] * mult / 1000.0 for p in pts]
    tot = [p["total_ew"] * mult / 1000.0 for p in pts]
    dem = [p["demand_ew"] * mult / 1000.0 for p in pts]

    fig = Figure(figsize=(11, 6.2))
    FigureCanvasAgg(fig)
    ax = fig.add_subplot(111)
    NAVY, STEEL, AMBER = "#" + AVIA, "#4C8FBF", "#C9822E"
    # THREE LINES, not two (John Carter, 24 August 2026): the capped total alone hides how
    # much demand a poor departure time actually spills - P2P is not itself capped anywhere
    # in the engine, it is the ROUTE TOTAL that is capped at the aircraft ceiling, so a
    # departure with strong connecting demand can show an uncapped total well above what is
    # actually carried. Drawn first/thin/dashed so the two solid lines that were already
    # here read exactly as they did before wherever demand sits at or under the ceiling.
    ax.plot(times, dem, color=AMBER, linewidth=1.6, linestyle="--", zorder=1,
            label="Total demand (uncapped)")
    ax.plot(times, conn, color=STEEL, linewidth=2, zorder=2,
            label="Connecting demand at the departure time")
    ax.plot(times, tot, color=NAVY, linewidth=2.4, zorder=3, label="Route total carried")

    # Restricted-hour shading, wraparound-aware (e.g. "21:00-06:00" spans midnight).
    def _hhmm_to_dt(s):
        h, m = (s.split(":") + ["0"])[:2]
        return _dt.datetime(2000, 1, 1) + _dt.timedelta(hours=int(h), minutes=int(m))
    day0, day1 = _dt.datetime(2000, 1, 1), _dt.datetime(2000, 1, 2)
    for band in (cs["restricted"] or []):
        try:
            a, b = [s.strip() for s in band.split("-")]
            ta, tb = _hhmm_to_dt(a), _hhmm_to_dt(b)
        except Exception:
            continue
        if tb <= ta:
            ax.axvspan(ta, day1, color="#DCE3EE", zorder=0)
            ax.axvspan(day0, tb, color="#DCE3EE", zorder=0)
        else:
            ax.axvspan(ta, tb, color="#DCE3EE", zorder=0)

    # Headroom at the top so the callout box and legend never sit flush against the highest
    # line (24 August 2026: the first draft cramped both against the axis ceiling; extended
    # the same day to include the new uncapped demand line, which can now be the highest of
    # the three wherever a departure spills demand past the aircraft ceiling).
    y_top = max(tot + conn + dem) if (tot or conn or dem) else 1.0
    ax.set_ylim(0, y_top * 1.18)

    # Chosen-departure marker and callout - offset a short, fixed distance from the point
    # itself (not a fixed spot on the chart), so the leader line stays short and the box
    # sits near what it describes regardless of where in the day the departure falls.
    # Pushed left when the point is in the second half of the day so it clears the legend
    # (top-right) rather than sitting under it.
    chosen_total = None
    if cs["chosen_mins"] is not None:
        anchor = min(pts, key=lambda p: abs(p["dep"] - cs["chosen_mins"]))
        chosen_total = anchor["total_ew"] * mult
        cx = _dt.datetime(2000, 1, 1) + _dt.timedelta(minutes=cs["chosen_mins"])
        ax.axvline(cx, color="#555555", linestyle="--", linewidth=1)
        second_half = cs["chosen_mins"] >= 720
        dx = -170 if second_half else 20
        ax.annotate(f"Chosen Dep {cs['chosen']}\n{round(chosen_total):,} passengers 2-way",
                    xy=(cx, chosen_total / 1000.0), xytext=(dx, 28),
                    textcoords="offset points", fontsize=10,
                    ha=("right" if second_half else "left"), va="bottom",
                    bbox=dict(boxstyle="round,pad=0.4", fc="#EEF1F6", ec="#AAB4C4"),
                    arrowprops=dict(arrowstyle="-", color="#777777", lw=0.8, shrinkB=4))

    # REMOVED (John Carter, 24 August 2026): this caption made the same "the restriction"
    # claim as the title clause dropped alongside it, and for the same reason - a run's
    # restricted_hours input is not necessarily the airport's real curfew (John's own SJC
    # runs enter 21:00-06:00, wider than SJC's actual 23:30-06:00, purely to force the
    # optimiser onto the 20:59 departure, a workaround for the separate dep-time-pinning
    # gap where selecting 21:00 directly loses the curve entirely). Confusing rather than
    # informative while that workaround is the only way to see a curve for a chosen time.
    # THE REAL FIX, noted for later, not attempted now: let a caller fix a departure time
    # AND still get the curve, so the picture can show where that choice sits against the
    # optimum for the same frequency - the dashboard's own dep_time-blank-only rule
    # (cortex_app.py line ~1046, optimise_departure only runs when dep_mins is None) is
    # the thing to change, not this chart.

    yr = cs["forecast_year"]
    # NOT labelled as the restriction (John Carter, 24 August 2026): the shaded band is
    # whatever restricted_hours the run was actually given, which is a real input but not
    # necessarily the airport's own published curfew - John's own SJC runs enter 21:00-06:00
    # to force the optimiser onto 20:59, wider than SJC's real 23:30-06:00 curfew. Calling it
    # "restricted departures" reads as an operational fact when it may be a modelling choice,
    # so the band is shown unlabelled rather than asserted as something it may not be.
    ax.set_title(f"Year 1 ({yr}) forecast by outbound departure time", fontsize=13, loc="left")
    ax.set_xlabel(f"Departure time, {home or o.get('iata','')} local")
    ax.set_ylabel("Passengers a year, two-way (000s)")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=3))
    ax.set_xlim(day0, day1)
    ax.grid(True, color="#E4E8EE", linewidth=0.8)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    ax.legend(loc="upper right", frameon=False, fontsize=9)
    fig.text(0.01, 0.01, attribution.SOURCE_LINE, fontsize=7, color="#888888")
    fig.tight_layout(rect=[0, 0.02, 1, 1])

    png_path = os.path.splitext(out_path)[0] + "_curve.png"
    fig.savefig(png_path, dpi=150)
    # No plt.close(fig): this Figure was never registered with pyplot's global figure
    # manager (that IS the point, see the note above), so there is nothing there to leak
    # or clean up - it is just released when fig goes out of scope, like any other object.
    return png_path


AVIA = "1F3864"; MID = "2F6BF0"; LIGHT = "EAF0FE"; GREENF = "E6F6EF"
HDR = PatternFill("solid", fgColor=AVIA)
SEC = PatternFill("solid", fgColor="4472C4")
LFILL = PatternFill("solid", fgColor=LIGHT)
TOTF = PatternFill("solid", fgColor="D6E4F0")
WHITE = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE = Font(name="Arial", bold=True, color=AVIA, size=15)
BOLD = Font(name="Arial", bold=True, size=10)
NORM = Font(name="Arial", size=10)
NOTE = Font(name="Arial", size=9, italic=True, color="666666")
TOTF_FONT = Font(name="Arial", bold=True, color=AVIA, size=11)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LFT_NW = Alignment(horizontal="left", vertical="center", wrap_text=False)
RGT = Alignment(horizontal="right", vertical="center")
THIN = Border(*(Side("thin", color="C9D3E4"),) * 4)


def _c(ws, r, c, v=None, font=NORM, fill=None, fmt=None, align=CTR, border=THIN):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.alignment = align
    cell.border = border if border is not None else Border()  # a real (empty) border, never None, so merge_cells can read its sides
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    return cell


def _hdr(ws, r, headers, widths=None):
    for i, h in enumerate(headers, 1):
        _c(ws, r, i, h, font=WHITE, fill=HDR)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _sec(ws, r, text, span):
    for i in range(1, span + 1):
        _c(ws, r, i, text if i == 1 else None, font=WHITE, fill=SEC, align=LFT_NW)
    if span > 1:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)


def _title(ws, text, sub="", span=6):
    _c(ws, 1, 1, text, font=TITLE, align=LFT_NW, border=None)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    if sub:
        _c(ws, 2, 1, sub, font=NOTE, align=LFT_NW, border=None)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)


def build_workbook(out_path, fc, meta=None):
    meta = meta or {}
    o = fc["origin"]; d = fc["dest"]; dem = fc["demand"]; cap = fc["capacity"]
    ec = fc.get("economics") or {}; raw = ec.get("raw") or {}
    home = fc["catchment"]["home"]; nm = fc["catchment"]["names"]; sh = fc["catchment"]["observed_share"]
    airline = meta.get("airline_name") or fc.get("airline") or "New entrant"
    n0 = lambda v: float(v or 0)
    # SEASON labelling: volume figures (forecast, feed, carried, seats, pax, profit) are for the
    # operating season; the addressable market stays annual. weeks = the season's operating weeks.
    _season = fc.get("season") or {}
    _smode = _season.get("mode", "annual")
    weeks = float(_season.get("weeks") or 52)
    _sshare = float(_season.get("share") or 1.0)      # 1.0 = annual; the season's share of the year
    _seasonal = _smode in ("summer", "winter")
    _pnoun = _smode if _seasonal else "year"          # "summer" / "winter" / "year"
    _padj = _smode.capitalize() if _seasonal else "Annual"   # "Summer" / "Winter" / "Annual"
    wb = openpyxl.Workbook()
    # Moved up from the Forecast tab section (24 August 2026, Jol Kingham: "Connecting
    # feed 2-way" H22+H43 = 147.4, "Forecast 2-way"/Cover show 148 - is that ok?"). Cover's
    # own PTEW used to come from dem["pdew_total"], a figure computed independently in
    # cortex_app.py; Forecast's PTEW comes from this same freq/weeks basis but footed
    # through ptew(). Two different code paths computing "the same" figure is exactly the
    # class of problem this whole day has been closing out one tab at a time - so Cover now
    # uses this identical, already-footed calculation directly, not a separately-sourced
    # figure that can only ever be close, never guaranteed equal.
    freq = n0(cap.get("freq")) or 7.0
    ptew = lambda x: round(x / (freq * weeks), 1) if freq else 0

    # ---- 1. Cover -----------------------------------------------------------
    ws = wb.active; ws.title = "Cover"
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 40
    _title(ws, f'{o["city"]} to {d["city"]}  ({home}-{d["iata"]})', "Meridian route forecast - detailed workbook")
    r = 4
    # Each-way figures, computed once; every 2-way figure below is this doubled at this
    # one point, per the workbook's standing EW/2-way discipline (Jol Kingham, 24 Aug 2026).
    # Named _cover_split, not _cs, because _cs is reused below for the Forecast tab's
    # competition_split dict - two different things, kept apart deliberately.
    _cover_split = carried_split(dem)
    _p2p_ew, _beh_ew, _bey_ew = round(_cover_split[0]), round(_cover_split[1]), round(_cover_split[2])
    _tot_ew = round(n0(dem.get("total")))
    # Cover's PTEW: the same footed sum the Forecast tab prints on its own GRAND TOTAL row
    # (each leg's own PTEW, footed, then summed - not an independent round of the total),
    # so this figure can never again disagree with Forecast, by construction rather than
    # by coincidence.
    _cover_ptew = ptew(_p2p_ew) + ptew(_beh_ew) + ptew(_bey_ew)
    blocks = [
        ("ROUTE", [("Origin", f'{o["city"]} ({home})'), ("Destination", f'{d["city"]} ({d["iata"]})'),
                   ("Country", f'{o.get("country","")} to {d.get("country","")}'),
                   ("Sector", f'{fc.get("distance_nm",0):,} nm'), ("Block time", f'{fc.get("block_min",0)} min')]),
        ("SERVICE", [("Airline", airline), ("Carrier type", fc.get("carrier_type","")),
                     ("Aircraft", cap.get("aircraft","")), ("Seats", ec.get("seats","")),
                     ("Frequency", f'{cap.get("freq","")}x/week')]),
        # Carried allocation on the cover too (18 Aug 2026): the cover printed the
        # same uncapped legs beside the capped total as the Forecast sheet did.
        #
        # JOL KINGHAM, 24 August 2026, on the corrected CI/BR/JX files: (1) row 24's label
        # should say PTEW, not a plain-English paraphrase, so it sets the term up for the
        # rest of the workbook (John: spell it out in full on first use here - "Passenger
        # Trip Each Way (PTEW)" - the same house-style rule as any other sector-specific
        # abbreviation); (2) include the 2-way figure too. Each passenger-count row now
        # states its own basis in its own label and the 2-way figure is the each-way figure
        # doubled at this one point, never recomputed a second way - the same discipline as
        # every other EW/2-way pairing in this workbook.
        ("FORECAST (per year)", [
                     ("Point to point (each way)", _p2p_ew),
                     ("Point to point (2-way)", _p2p_ew * 2),
                     ("Connecting behind " + home + " (each way)", _beh_ew),
                     ("Connecting behind " + home + " (2-way)", _beh_ew * 2),
                     ("Connecting beyond " + d["iata"] + " (each way)", _bey_ew),
                     ("Connecting beyond " + d["iata"] + " (2-way)", _bey_ew * 2),
                     ("Total forecast (each way)", _tot_ew),
                     ("Total forecast (2-way)", _tot_ew * 2),
                     ("Planned load factor", n0(cap.get("load"))),
                     ("Passenger Trip Each Way (PTEW)", _cover_ptew)]),
        # JOL, same email: (3) row 28's year printed "2,025" - the generic row-renderer
        # below applies a thousands-separator format to every int/float, and a year is an
        # int with no digit-grouping meaning at all. Cast to a string here so it falls
        # through to the plain-text branch instead. (4) row 27, is the OAG schedule week
        # beginning or ending - sourced, not guessed: resolve_oag_week()'s own docstring
        # (this file's caller, cortex_app.py line ~470) documents the single-week store
        # label as "week commencing", so it is beginning, and the header now says so.
        ("BASIS", [("OAG schedule week (beginning)", fc.get("week","")),
                   ("Sabre Global Demand Data year", str(fc.get("year") or "")),
                   ("Analyst", meta.get("analyst","Avia Solutions")), ("Date", meta.get("date",""))]),
    ]
    for title, rows in blocks:
        _sec(ws, r, title, 2); r += 1
        for k, v in rows:
            _c(ws, r, 1, k, font=BOLD, align=LFT)
            if isinstance(v, float) and 0 < v < 1:
                _c(ws, r, 2, v, fmt="0.0%", align=RGT)
            elif isinstance(v, (int, float)):
                _c(ws, r, 2, v, fmt="#,##0", align=RGT)
            else:
                _c(ws, r, 2, v, align=RGT)
            r += 1
        r += 1

    # ---- 2. Traffic forecast (Avia gold-standard layout) -------------------
    # EW / 2-way pair (John, 22 August 2026): every tab that carries a passenger
    # COUNT is now built once, natively each way, and written to two tabs - "EW"
    # at the native value, "2-way" at the same value x2 - so the basis is stated
    # in the tab name itself and never has to be inferred from a footnote. Rates
    # (CAGR, capture share, PTEW - passengers per departure, which nets to the
    # same figure either way since passengers and departures scale together) are
    # written unchanged on both tabs; only the counted columns take the
    # multiplier, at this single point, never recomputed a second way.
    #
    # BASE / GROWTH / GROWN decomposed (18 August 2026): with a forecast year set,
    # this sheet printed the GROWN market in the base column with growth showing 0,
    # so a client table could not state its own growth. The payload now carries the
    # numbers behind the growth-basis prose; a steady-state run decomposes to
    # growth 0 and prints as before.
    _sch = fc.get("schedule") or {}
    _gy = n0(_sch.get("growth_years")); _gr = n0(_sch.get("growth_rate"))
    _cum = ((1.0 + _gr) ** _gy - 1.0) if _gy else 0.0
    _byr = _sch.get("base_year") or fc.get("year") or ""
    _fyr = _sch.get("forecast_year") or _byr
    _debase = lambda grown: (grown / (1.0 + _cum)) if _cum else grown
    cap_share = n0(dem.get("qsi_share")); stim = n0(dem.get("stimulation")) or 1.0
    natural = n0(dem.get("natural")) * _sshare
    # THE ROWS ARE THE CARRIED ALLOCATION (18 August 2026). This sheet printed
    # UNCAPPED legs under the CAPPED total, so its rows did not sum to its own grand
    # total (76.0 over a 55.7 total on the SJC-TPE airline tables, found the moment a
    # client table was built from it; a network planner adds the rows in the room).
    # The payload carries the engine's carried split; the rows print it, the capture
    # column becomes the EFFECTIVE rate after the capacity allocation, and the legs
    # sum to the total by construction. Older payloads without the carried split fall
    # back to scaling the uncapped legs pro-rata to the total, the same allocation
    # forecast_to_contract renders, so the two surfaces can never disagree.
    p2p, behind, beyond, tot = carried_split(dem)
    # effective P2P capture after the allocation, so the row multiplies through
    if natural * stim > 0:
        cap_share = p2p / (natural * stim)
    # freq and ptew() are now defined once, up in the Cover section, so Cover and this
    # tab always use the identical figure - see the note there (24 August 2026, Jol
    # Kingham's "148 v 147" question).
    _cs = fc.get("competition_split") or {}
    for _suf, _mult, _bw in (("EW", 1, "each way"), ("2-way", 2, "two way")):
        ws = wb.create_sheet(f"Forecast {_suf}")
        # Bumped 1dp -> 3dp (24 August 2026, Jol Kingham: "can demand columns and
        # forecast column show figures to at least 3 decimal places?"). The fmt="#,##0.000"
        # display format was changed earlier the same day but this rounding was missed -
        # the value itself was still baked to 1dp before the format string ever saw it, so
        # the tab was showing "160.900" rather than genuine added precision. Caught by
        # John's own check against a live regeneration, not by the sandbox test suite,
        # which only asserts against a tolerance and never noticed the trailing zeros.
        k = lambda x, _m=_mult: round(x * _m / 1000.0, 3)
        # CAGR, not the cumulative (20 August 2026, Mark Kiehl/SJC, reviewing the PPTX packs,
        # then applied here for the same reason John raised about the identical growth-rate
        # display in the deck: one basis everywhere this table appears). 18.3% over two years
        # reads as a big, alarming number; _gr is the per-annum rate the cumulative was built
        # from, roughly half of it and in single digits. The cumulative is stated in the note.
        _title(ws, "Traffic forecast", f"{_bw} per {_pnoun}; connecting rows show the captured feed")
        _hdr(ws, 4, ["Market", f"Base demand {_byr} (000s)", f"CAGR {_fyr} v {_byr}",
                     f"Grown demand {_fyr} (000s)", "Stimulation",
                     "Stimulated demand (000s)", "Capture rate", "Forecast (000s)", "PTEW"],
             [30, 15, 12, 15, 11, 15, 11, 12, 9])
        r = 5
        # JOL KINGHAM, 24 August 2026: the three main rows' PTEW column did not sum to
        # the GRAND TOTAL row's own PTEW (267 v 268). Both figures were true, and this
        # was never a basis bug like the others this week - p2p+behind+beyond equals tot
        # exactly (carried_split's own invariant), but ptew() rounds each to a whole
        # number independently, so round(p2p)+round(behind)+round(beyond) does not
        # generally equal round(p2p+behind+beyond); a 1-unit drift is ordinary rounding,
        # not an error. A table that a reader adds up by hand has to foot, so the GRAND
        # TOTAL row below now sums the three displayed PTEW figures instead of rounding
        # the true total separately - the row a reader checks against is the row that is
        # actually right when checked.
        _p2p_ptew = ptew(p2p)
        _legs_ptew = _p2p_ptew
        _c(ws, r, 1, "Total point to point", font=BOLD, align=LFT)
        _c(ws, r, 2, k(_debase(natural)), fmt="#,##0.000", align=RGT); _c(ws, r, 3, _gr, fmt="0.0%", align=RGT)
        _c(ws, r, 4, k(natural), fmt="#,##0.000", align=RGT); _c(ws, r, 5, round(stim, 2), fmt="0.00", align=RGT)
        _c(ws, r, 6, k(natural * stim), fmt="#,##0.000", align=RGT); _c(ws, r, 7, cap_share, fmt="0.0%", align=RGT)
        _c(ws, r, 8, k(p2p), fmt="#,##0.000", align=RGT); _c(ws, r, 9, _p2p_ptew, fmt="#,##0.0", align=RGT)
        r += 1
        for label, val, cbase, _leg in [(f"Total connecting behind {home}", behind, n0(dem.get("feed_behind_base")) * _sshare, "behind"),
                                        (f"Total connecting beyond {d['iata']}", beyond, n0(dem.get("feed_beyond_base")) * _sshare, "beyond")]:
            _leg_ptew = ptew(val)
            _legs_ptew += _leg_ptew
            _c(ws, r, 1, label, font=BOLD, align=LFT)
            _c(ws, r, 2, k(_debase(cbase)) if cbase else "-", fmt="#,##0.000", align=RGT)
            _c(ws, r, 3, _gr if cbase else "-", fmt="0.0%", align=RGT)
            _c(ws, r, 4, k(cbase) if cbase else "-", fmt="#,##0.000", align=RGT)
            _c(ws, r, 5, 1.00, fmt="0.00", align=RGT)
            _c(ws, r, 6, k(cbase) if cbase else "-", fmt="#,##0.000", align=RGT)
            _c(ws, r, 7, (val / cbase) if cbase else "-", fmt="0.0%", align=RGT)
            _c(ws, r, 8, k(val), fmt="#,##0.000", align=RGT); _c(ws, r, 9, _leg_ptew, fmt="#,##0.0", align=RGT)
            r += 1
            # THE COMPETITION SUB-ROWS (John's ruling, 18 August 2026, validated against
            # the 2025 analyst's split): direct / without direct competition beneath each
            # leg, scaled to the displayed carried leg so the two sum to the row above.
            # Absent block, no sub-rows, never zeros.
            _tt = ((_cs.get(_leg) or {}).get("totals") or {})
            _dsum = sum(((_tt.get(_k) or {}).get("forecast") or 0)
                        for _k in ("direct", "no_direct"))
            if _dsum > 0 and val > 0:
                _scl = val / _dsum
                # SUB-ROW PTEW FOOTING (24 August 2026, Jol Kingham, one level deeper than
                # the GRAND TOTAL footing fixed earlier today: "the PTEW column in the
                # excels don't add - CI says 268 but is 270 sum of the parts"). His own
                # annotated sum used these two sub-rows, not the leg-total row above them,
                # and got 267 against a printed 268 - the same independent-rounding drift,
                # one row down: the forecast (000s) column already sums exactly here
                # (scaled by _scl for that reason), but PTEW was still rounded per sub-row
                # independently. The "with direct competition" row keeps its own rounded
                # figure; "without" is now the remainder against the leg's own PTEW
                # (already proven to foot into the grand total), so direct + without always
                # equals the leg row above, and the whole table foots at every level, not
                # only the top.
                _fv_direct = float((_tt.get("direct") or {}).get("forecast") or 0) * _scl
                _fv_nodirect = float((_tt.get("no_direct") or {}).get("forecast") or 0) * _scl
                _ptew_direct = ptew(_fv_direct)
                _ptew_nodirect = _leg_ptew - _ptew_direct
                for _bk, _bl, _fv, _pt in (
                        ("direct", "   O&Ds with direct competition", _fv_direct, _ptew_direct),
                        ("no_direct", "   O&Ds without direct competition", _fv_nodirect, _ptew_nodirect)):
                    _t = _tt.get(_bk) or {}
                    _fb = float(_t.get("base") or 0)
                    _c(ws, r, 1, _bl, align=LFT)
                    _c(ws, r, 2, k(_debase(_fb)) if _fb else "-", fmt="#,##0.000", align=RGT)
                    _c(ws, r, 3, _gr if _fb else "-", fmt="0.0%", align=RGT)
                    _c(ws, r, 4, k(_fb) if _fb else "-", fmt="#,##0.000", align=RGT)
                    _c(ws, r, 5, 1.00, fmt="0.00", align=RGT)
                    _c(ws, r, 6, k(_fb) if _fb else "-", fmt="#,##0.000", align=RGT)
                    _c(ws, r, 7, (_fv / _fb) if _fb else "-", fmt="0.0%", align=RGT)
                    _c(ws, r, 8, k(_fv), fmt="#,##0.000", align=RGT)
                    _c(ws, r, 9, _pt, fmt="#,##0.0", align=RGT)
                    r += 1
        _c(ws, r, 1, "GRAND TOTAL", font=TOTF_FONT, fill=TOTF, align=LFT)
        _bb = n0(dem.get("feed_behind_base")) * _sshare; _yb = n0(dem.get("feed_beyond_base")) * _sshare
        _c(ws, r, 2, k(_debase(natural + _bb + _yb)), font=TOTF_FONT, fill=TOTF, fmt="#,##0.000", align=RGT)
        _c(ws, r, 3, None, fill=TOTF)
        _c(ws, r, 4, k(natural + _bb + _yb), font=TOTF_FONT, fill=TOTF, fmt="#,##0.000", align=RGT)
        _c(ws, r, 5, None, fill=TOTF)
        _c(ws, r, 6, k(natural * stim + _bb + _yb), font=TOTF_FONT, fill=TOTF, fmt="#,##0.000", align=RGT)
        _c(ws, r, 7, (tot / (natural * stim + _bb + _yb)) if (natural * stim + _bb + _yb) > 0 else None,
           font=TOTF_FONT, fill=TOTF, fmt="0.0%", align=RGT)
        _c(ws, r, 8, k(tot), font=TOTF_FONT, fill=TOTF, fmt="#,##0.000", align=RGT)
        _c(ws, r, 9, _legs_ptew, font=TOTF_FONT, fill=TOTF, fmt="#,##0.0", align=RGT)
        _c(ws, r + 2, 1, f"Figures on this tab are {_bw}. "
                         "Base demand is the addressable each-way O&D market from Sabre Global Demand Data in the origin catchment. "
                         "Rows show the carried allocation after the planned load factor cap and they sum to the total; "
                         "capture rates are the effective rates after that allocation, so each row multiplies through. "
                         "Where capacity binds, unconstrained demand exceeds the figures shown. "
                         + ("Growth is shown as a compound annual rate; the cumulative growth from %s to %s is %.1f%%. "
                            % (_byr, _fyr, _cum * 100))
                         + "PTEW = passengers per departure each way, the same figure on the EW and 2-way tabs."
                         + (f"  Figures are for the {_pnoun} service (the season's share of the annual O&D)."
                            if _seasonal else ""),
           font=NOTE, align=LFT, border=None)

    # ---- 3. Connecting feed detail (base demand, share, forecast, PDEW) -----
    # JOL'S REVIEW, 19 August 2026, two findings both fixed here.
    # (1) The Total row summed only the PRINTED top-15 rows, a subtotal wearing a
    #     total's label: on CI the beyond column showed 24,325 against a carried leg
    #     of 29,063 each way, and the missing 4,738 is the tail of small markets. An
    #     "All other" row now completes each leg to the SAME figure the Cover and the
    #     Forecast sheet print (carried_split), so the surfaces agree by construction.
    #     The market-demand column takes NO other-row figure: each row's demand is
    #     that market's own O&D size, a different quantity from the leg's connecting
    #     base, and a filler there would paper over a definition, so it stays "-".
    # (2) Year labels (John's rule: every column names its year): route_forecast
    #     line ~642 grows the city detail's base AND captured to the forecast year,
    #     so both figure columns are AT the service year, not the base year, and the
    #     headers now say so.
    _svc_yr = (fc.get("schedule") or {}).get("forecast_year") or ""
    _yrtag = f" {_svc_yr}" if _svc_yr else ""
    _csp, _csb, _csy, _cst = carried_split(dem)
    for _suf, _mult, _bw in (("EW", 1, "each way"), ("2-way", 2, "two way")):
        ws = wb.create_sheet(f"Connecting feed {_suf}")
        _title(ws, "Connecting feed detail",
               f"connecting markets {_bw} at{_yrtag or ' the service year'}: market O&D demand, captured share, forecast, PTEW")
        r = 4
        # DEMAND-COLUMN TOTAL, 20 August 2026 (John, checking the EVA pack against the deck's
        # completed forecast column): feed_beyond_base/feed_behind_base are the FULL uncapped
        # market before capture, the same quantity each city's own "base"/demand figure is
        # drawn from, additive with them. The 19 August note calling this "a different
        # quantity, leave it blank" was wrong; mirrored in deck/forecast_pack.py the same day.
        for label, key, _leg_ew, _mkt_ew in [
                (f"Connecting at {home} (behind the origin)", "behind_pdew", _csb, n0(dem.get("feed_behind_base"))),
                (f"Connecting at {d['iata']} (beyond the destination)", "beyond_pdew", _csy, n0(dem.get("feed_beyond_base")))]:
            # JOL KINGHAM, 24 August 2026: "Code" is ambiguous (airport or IATA city code);
            # "City" and "Country" likewise don't say name or code. Traced against
            # _feed_list()'s actual field semantics in cortex_app.py: "code" is keyed into
            # the airport table (an airport code, not a city code - it disambiguates
            # airports sharing a city, e.g. the London group), the city field is a name,
            # and country is the 2-letter ISO code (confirmed against pitch_html.py's own
            # sample payloads, e.g. "PH", "VN"). Headers now state exactly what each is.
            _sec(ws, r, label, 8); r += 1
            _hdr(ws, r, ["Nr", "Airport Code", "City Name", "Country Code", f"Market demand{_yrtag}", "Share",
                         f"{_padj} forecast{_yrtag}", "PTEW"],
                 [6, 12, 24, 14, 15, 10, 15, 9]); r += 1
            lst = dem.get(key) or []; sub_base = 0.0; sub_fc = 0.0; sub_ptew = 0.0
            for i, row in enumerate(lst, 1):
                base = n0(row.get("base")); shr = n0(row.get("share"))
                # SEVENTH INSTANCE OF THE FLAT-DAY PTEW BUG (24 August 2026, Jol Kingham: the
                # Connecting feed tabs' PTEW column did not sum to Cover row 24 nor the
                # Forecast tabs). This fallback (only used when a row's own "forecast" field
                # is missing) reconstructed it from pdew x weeks x 7.0 - a flat daily-service
                # assumption, the same wrong basis as the six other instances fixed this week -
                # instead of the route's actual freq x weeks departures.
                fcv = n0(row.get("forecast")) or (n0(row.get("pdew")) * freq * weeks); pdv = n0(row.get("pdew"))
                if fcv <= 0 and pdv <= 0:
                    continue
                sub_base += base; sub_fc += fcv; sub_ptew += round(pdv, 1)
                _c(ws, r, 1, i, align=CTR); _c(ws, r, 2, row.get("code"), align=CTR)
                _c(ws, r, 3, row.get("name"), align=LFT); _c(ws, r, 4, row.get("country") or "", align=LFT)
                _c(ws, r, 5, round(base * _mult) if base else "-", fmt="#,##0", align=RGT)
                _c(ws, r, 6, shr if base else "-", fmt="0.0%", align=RGT)
                _c(ws, r, 7, round(fcv * _mult), fmt="#,##0", align=RGT); _c(ws, r, 8, round(pdv, 1), fmt="#,##0.0", align=RGT)
                r += 1
            # The tail: every connecting market smaller than the listed ones. Only drawn
            # when the carried leg genuinely exceeds the listed sum; a seasonal or older
            # payload where the two bases differ keeps the honest subtotal instead. Demand
            # completes to _mkt_ew the same way; a run without feed_beyond_base/behind_base
            # (an older payload) keeps the honest "-" rather than a fabricated figure.
            _other = (_leg_ew - sub_fc) if (_leg_ew and _leg_ew > sub_fc + 0.5) else 0.0
            _other_dem = (_mkt_ew - sub_base) if (_mkt_ew and _mkt_ew > sub_base + 0.5) else 0.0
            # Total row's own PTEW (the authoritative figure, see below) computed here so
            # the All-other row can foot against it.
            _total_ptew = round(_leg_ew / (freq * weeks), 1) if freq else None
            if _other or _other_dem:
                # ALL-OTHER PTEW FOOTING (24 August 2026, Jol Kingham: his own annotated sum
                # of the SJC behind leg's PTEW column read 35.5 against a printed Total of
                # 35.6 - each of the ~15 listed cities is independently rounded to 1dp, and
                # that rounding accumulates over that many rows). All-other is already a
                # residual row for the demand and forecast columns (it completes the listed
                # cities to the leg's true total); PTEW now follows the same design - the
                # remainder against the leg's own Total PTEW, not an independent calculation
                # from _other - so the full column, cities plus All-other, always foots
                # exactly to the Total row beneath it.
                _other_ptew = (round(_total_ptew - sub_ptew, 1) if _total_ptew is not None else None)
                _c(ws, r, 1, "", align=CTR)
                _c(ws, r, 3, "All other connecting markets", align=LFT)
                _c(ws, r, 5, round(_other_dem * _mult) if _other_dem else "-", fmt="#,##0", align=RGT)
                _c(ws, r, 6, "-", align=RGT)
                _c(ws, r, 7, round(_other * _mult), fmt="#,##0", align=RGT)
                _c(ws, r, 8, _other_ptew if _other_ptew is not None else "-", fmt="#,##0.0", align=RGT)
                r += 1
            # Total row PTEW (24 August 2026, Jol Kingham): this cell was blank, so a
            # reader summing the column by hand was building their own total from
            # individually-rounded per-city figures, which drifts from the true leg total
            # by ordinary rounding error as the list gets longer. Printed directly instead,
            # from the same carried-leg figure and the same freq x weeks basis Cover and
            # the Forecast tabs use for this leg, so this cell agrees with both by
            # construction rather than by chance - and now, with the All-other row above
            # footing to it too, the whole column sums to this figure exactly.
            _c(ws, r, 1, "Total", font=BOLD, fill=TOTF, align=LFT)
            for cc in (2, 3, 4, 6):
                _c(ws, r, cc, None, fill=TOTF)
            _c(ws, r, 5, round((sub_base + _other_dem) * _mult) if (sub_base or _other_dem) else "-",
               font=BOLD, fill=TOTF, fmt="#,##0", align=RGT)
            _c(ws, r, 7, round((sub_fc + _other) * _mult), font=BOLD, fill=TOTF, fmt="#,##0", align=RGT)
            _c(ws, r, 8, _total_ptew if _total_ptew is not None else "-",
               font=BOLD, fill=TOTF, fmt="#,##0.0", align=RGT); r += 1
            _c(ws, r, 1, f"Figures on this tab are {_bw}; PTEW (passengers per departure) is a rate and reads the "
                         "same on the EW and 2-way tabs. Market demand, forecast and PTEW all complete to their "
                         "carried-leg and market totals via the All-other row, so every column foots exactly to the "
                         "Total row and agrees with the Forecast tab and Cover. Where no All-other row is drawn (the "
                         "listed markets already account for the full leg), the city rows sum to the Total within "
                         "ordinary rounding.",
               font=NOTE, align=LFT, border=None)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 2

    # ---- 3b. Schedule and capacity ----------------------------------------
    # THE TITLE AND NOTE STATE THE TIMES' BASIS (19 August 2026): the sheet always
    # read the payload's times, but its wording claimed "indicative" even when the
    # optimiser chose them, which read as a mismatch against the screen.
    _sch0 = fc.get("schedule") or {}
    _tbasis = ("indicative times" if _sch0.get("indicative")
               else "optimised departure" if not _sch0.get("basis")
               else str(_sch0.get("basis"))[:40])
    sched = fc.get("schedule") or {}
    seats = n0(ec.get("seats")); load = n0(cap.get("load"))
    ann_seats_dir = seats * freq * weeks
    # EW / 2-way pair (22 August 2026): the outbound and inbound ROWS are each one
    # direction by definition, so they print the same on both tabs, never doubled.
    # Only the Total row differs - the EW tab's total is one direction's own annual
    # figures, the 2-way tab's is both directions combined, the figure this sheet
    # always showed before this change.
    for _suf, _mult, _bw in (("EW", 1, "each way"), ("2-way", 2, "two way")):
        ws = wb.create_sheet(f"Schedule {_suf}")
        _title(ws, "Schedule and capacity", f'{cap.get("aircraft","")} at {int(freq)}x/week ({_tbasis})')
        _hdr(ws, 4, ["Sector", "Dep", "Arr", "Op days/week", "Aircraft", "Seats", f"{_padj} seats", f"{_padj} pax", "Seat factor"],
             [13, 9, 9, 13, 12, 9, 14, 13, 11])
        r = 5
        for leg in ("outbound", "inbound"):
            s = sched.get(leg) or {}
            sector = s.get("sector") or (f'{home}-{d["iata"]}' if leg == "outbound" else f'{d["iata"]}-{home}')
            _c(ws, r, 1, sector, font=BOLD, align=LFT); _c(ws, r, 2, s.get("dep") or "-", align=CTR)
            _c(ws, r, 3, s.get("arr") or "-", align=CTR); _c(ws, r, 4, int(freq), align=RGT)
            _c(ws, r, 5, cap.get("aircraft", ""), align=CTR); _c(ws, r, 6, round(seats), fmt="#,##0", align=RGT)
            _c(ws, r, 7, round(ann_seats_dir), fmt="#,##0", align=RGT); _c(ws, r, 8, round(tot), fmt="#,##0", align=RGT)
            _c(ws, r, 9, load, fmt="0.0%", align=RGT); r += 1
        _c(ws, r, 1, "Total", font=BOLD, fill=TOTF, align=LFT)
        for cc in (2, 3, 4, 5, 6):
            _c(ws, r, cc, None, fill=TOTF)
        _c(ws, r, 7, round(ann_seats_dir * _mult), font=BOLD, fill=TOTF, fmt="#,##0", align=RGT)
        _c(ws, r, 8, round(tot * _mult), font=BOLD, fill=TOTF, fmt="#,##0", align=RGT)
        _c(ws, r, 9, load, font=BOLD, fill=TOTF, fmt="0.0%", align=RGT)
        _c(ws, r + 2, 1,
           (("Departure and arrival are indicative local times derived from block time and "
             "timezone; not curfew- or slot-optimised. ") if _sch0.get("indicative") else
            ("Departure and arrival are the run's chosen schedule"
             + (f" ({_sch0.get('basis')})" if _sch0.get("basis") else "") + ". "))
           + "The outbound and inbound rows above are each one direction, so they read the same on "
             f"the EW and 2-way tabs; only the Total row differs, {'one direction' if _mult == 1 else 'both directions combined'}. "
           + f"{_padj} seats = seats x frequency x {int(weeks)}, each "
             f"direction; {_pnoun} pax is the forecast each way; seat factor is the planned load.",
           font=NOTE, align=LFT, border=None)

    # ---- 3c. Departure curve (John, 19 August 2026) ------------------------
    # The day curve with its raw data and a native Excel chart, so the analysis
    # travels in the workbook and can later become a presentation slide. The
    # transform MIRRORS the dashboard slider's, which is the spec: anchor on the
    # curve point nearest the chosen departure, scale the raw connecting scores by
    # connecting_carried at that anchor, hold point-to-point constant across the
    # day, cap the route total at annual capacity x the plan cap. Absent an
    # optimiser curve (no airline named), no sheet, never a fabricated one.
    _op0 = (_sch0.get("optimised") or {})
    _curve = _op0.get("curve") or []
    _p2pc2 = n0(dem.get("p2p_carried")); _connc2 = n0(dem.get("connecting_carried"))
    _capew2 = n0(cap.get("annual_capacity")) * n0(cap.get("plan_cap"))
    _chosen = ((_sch0.get("outbound") or {}).get("dep") or "")
    _cparts = str(_chosen).split(":")
    _cmins = (int(_cparts[0]) * 60 + int(_cparts[1])) if len(_cparts) == 2 and _cparts[0].isdigit() else None
    if len(_curve) >= 4 and _connc2 > 0 and _p2pc2 > 0 and _cmins is not None:
        _anchor = min(_curve, key=lambda q: abs(float(q.get("dep") or 0) - _cmins))
        _at = float(_anchor.get("total") or 0)
        if _at > 0:
            _scl2 = _connc2 / _at
            # THE CURVE PICTURE, embedded (24 August 2026, found live): render_curve_png()
            # was built and wired as a side effect of build_workbook(), called AFTER
            # wb.save() and saving only to a sibling file next to out_path on the SERVER's
            # own temp directory. Every api_report() download path (part=xlsx, the one John
            # actually clicks; the part=both zip) only ever returns the xlsx or the
            # deck+xlsx zip - the sibling PNG was never zipped, never attached, never on
            # any path back to the browser. Three real rendering bugs got found and fixed
            # today (the pinned-departure gap, the backend no-op, the pyplot thread race)
            # and every one of them was necessary, but none of them was sufficient,
            # because even a perfectly-rendered PNG was sitting on a machine John never
            # gets a file listing of. Rendered here, BEFORE wb.save(), and embedded
            # directly into the workbook that actually reaches him, so it travels with a
            # single xlsx download regardless of which button he clicks. The standalone
            # file next to out_path is still written too (render_curve_png's own contract,
            # what test_workbook_table.py already checks) - harmless, and useful if a
            # future caller does have real file-system access to the server.
            try:
                _png_path = render_curve_png(fc, meta, out_path)
            except Exception:                                        # noqa: BLE001
                import traceback
                print("render_curve_png: not rendered -")
                traceback.print_exc()
                _png_path = None
            # EW / 2-way pair (22 August 2026): this sheet was always built two-way
            # (every figure here already carried a x2). That existing content is now
            # the "2-way" tab, unchanged; a new "EW" tab sits beside it built from the
            # same native, undoubled figures. Do not read the old single tab as "the
            # EW basis" - it never was.
            for _suf, _mult, _bw in (("EW", 1, "each way"), ("2-way", 2, "two way")):
                ws = wb.create_sheet(f"Departure curve {_suf}")
                _title(ws, "Connecting traffic by departure time",
                       f"forecast {_sch0.get('forecast_year') or fc.get('year', '')}; {_bw} annual passengers; "
                       "derived from this run's own carried figures")
                _hdr(ws, 4, ["Departure (origin local)", "Permitted", "Connecting score (raw)",
                             f"Connecting, {_bw}", "of which beyond", "of which behind",
                             f"Route total carried, {_bw}"], [18, 10, 15, 15, 13, 13, 18])
                r = 5
                for p in _curve:
                    _fe = float(p.get("total") or 0) * _scl2
                    _tot2 = (min(_p2pc2 + _fe, _capew2) if _capew2 > 0 else (_p2pc2 + _fe)) * _mult
                    _c(ws, r, 1, p.get("hhmm") or "", align=CTR)
                    _c(ws, r, 2, "yes" if p.get("permitted") else "no", align=CTR)
                    _c(ws, r, 3, round(float(p.get("total") or 0)), fmt="#,##0", align=RGT)
                    _c(ws, r, 4, round(_fe * _mult), fmt="#,##0", align=RGT)
                    _c(ws, r, 5, round(float(p.get("beyond") or 0) * _scl2 * _mult), fmt="#,##0", align=RGT)
                    _c(ws, r, 6, round(float(p.get("behind") or 0) * _scl2 * _mult), fmt="#,##0", align=RGT)
                    _c(ws, r, 7, round(_tot2), fmt="#,##0", align=RGT)
                    r += 1
                _rw = ", ".join(_op0.get("restricted") or []) or "none"
                _c(ws, r + 1, 1,
                   f"Chosen departure {_chosen}"
                   + (f"; unrestricted optimum {_op0.get('unrestricted_dep')}" if _op0.get("unrestricted_dep") else "")
                   + f"; restricted hours {_rw}. Point to point is constant across the day "
                   f"({round(_p2pc2 * _mult):,} {_bw}); every movement in the total comes from the "
                   f"connecting side, scored on the outbound departure. The route total is capped "
                   f"at the aircraft ceiling ({round(_capew2 * _mult):,} {_bw} at the plan load factor). "
                   "Connecting is shown uncapped, as on the dashboard chart. "
                   + attribution.SOURCE_LINE,
                   font=NOTE, align=LFT, border=None)
                try:
                    from openpyxl.chart import LineChart, Reference
                    _ch = LineChart()
                    _ch.title = (f"{_bw.capitalize()} annual passengers by outbound departure "
                                 f"(forecast {_sch0.get('forecast_year') or fc.get('year', '')})")
                    _ch.y_axis.title = f"passengers / yr, {_bw}"
                    _ch.x_axis.title = "outbound departure, origin local"
                    _n = len(_curve)
                    _ch.add_data(Reference(ws, min_col=7, min_row=4, max_row=4 + _n), titles_from_data=True)
                    _ch.add_data(Reference(ws, min_col=4, min_row=4, max_row=4 + _n), titles_from_data=True)
                    _ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + _n))
                    _ch.height, _ch.width = 9, 24
                    ws.add_chart(_ch, "I4")
                except Exception as _ce:                         # noqa: BLE001
                    _c(ws, r + 3, 1, "chart not rendered: %s" % _ce, font=NOTE, align=LFT, border=None)
                # The picture, on the two-way tab only (matplotlib.mult=2 in render_curve_png,
                # so its own axis already reads "two way" - embedding it a second time on the
                # EW tab would print a two-way picture under an each-way header, which is
                # exactly the kind of basis mismatch this workbook is otherwise strict about).
                # Below the note, well clear of the native chart's own anchor at I4.
                if _suf == "2-way" and _png_path and os.path.exists(_png_path):
                    try:
                        from openpyxl.drawing.image import Image as _XLImage
                        _img = _XLImage(_png_path)
                        _img.width, _img.height = 760, 429   # matplotlib's 11x6.2in @150dpi scaled to fit a sheet
                        ws.add_image(_img, f"A{r + 5}")
                    except Exception as _ie:                       # noqa: BLE001
                        _c(ws, r + 5, 1, "picture not embedded: %s" % _ie, font=NOTE, align=LFT, border=None)

    # ---- 4. Catchment split ------------------------------------------------
    # EW / 2-way pair, built for structural consistency with the other tabs (John,
    # 22 August 2026): this table is a SHARE of catchment demand, not a passenger
    # count, so it has no each-way/two-way basis and prints identically on both
    # tabs. Said so on the tab itself, so an identical pair reads as by design.
    for _suf in ("EW", "2-way"):
        ws = wb.create_sheet(f"Catchment {_suf}")
        _title(ws, "Catchment airport split", "how the origin catchment's demand splits across competing airports today")
        _hdr(ws, 4, ["Airport", "Share of catchment"], [40, 18])
        r = 5
        for c, v in sorted(sh.items(), key=lambda kv: -kv[1]):
            lbl = nm.get(c) or c
            _c(ws, r, 1, lbl + ("  (this route's origin)" if c == home else ""),
               font=(BOLD if c == home else NORM), fill=(GREENF and PatternFill("solid", fgColor=GREENF)) if c == home else None, align=LFT)
            _c(ws, r, 2, round(n0(v), 4), fmt="0.0%", align=RGT); r += 1
        cb = meta.get("capture_basis", "modelled from drive time and competing service")
        _c(ws, r + 1, 1, f"Assumed capture with this route's own nonstop: {cap_share*100:.1f}%  ({cb}). "
                         "This tab is a share of catchment demand, not a passenger count, so it reads "
                         "identically on the EW and 2-way tabs.",
           font=NOTE, align=LFT, border=None)

    # ---- 5. Route economics (turnaround P&L) -------------------------------
    # NOT given an EW/2-way pair (John's 22 August EW/2-way project; checked against
    # aircraft_economics.py before deciding, not assumed). A rotation is an out-
    # and-back unit and every cost line already reflects that: pax_turn = 2 x
    # (econ_ow + bus_ow), and fuel, maintenance, landing and ground handling are
    # priced per ROTATION, not per passenger, so they have no coherent each-way
    # half - an aircraft does not burn half its fuel for "half a rotation."
    # Halving only the passenger-linked lines (revenue, catering, per-pax charges)
    # while leaving the block-hour and per-turn costs whole would set a halved
    # revenue against a full-rotation cost base and understate profit, a wrong P&L,
    # not a basis choice. This tab stays single and is now labelled explicitly.
    ws = wb.create_sheet("Economics")
    _title(ws, "Route economics", f'turnaround P&L on the {cap.get("aircraft","")}, one rotation out and back, two way')
    _hdr(ws, 4, ["Line item", "Per rotation ($)"], [40, 20]); r = 5
    def prow(k, v, bold=False, fill=None):
        nonlocal r
        _c(ws, r, 1, k, font=(BOLD if bold else NORM), fill=fill, align=LFT)
        _c(ws, r, 2, round(n0(v)), font=(BOLD if bold else NORM), fill=fill, fmt="#,##0", align=RGT); r += 1
    _sec(ws, r, "REVENUE", 2); r += 1
    prow("Passenger revenue (net)", raw.get("net_rev")); prow("Cargo", raw.get("cargo_rev"))
    prow("Charges recovery", raw.get("charges_recovery")); prow("Gross revenue", raw.get("gross_rev"), True, LFILL)
    _sec(ws, r, "OPERATING COST", 2); r += 1
    # The two plug lines are labelled as such. Airport and handling charges are a generic placeholder
    # rather than this airport pair, and on a short sector they are the largest cost in the P&L;
    # ownership rests on a lease rate that is not publicly available and which Avia does not publish.
    # A reader who cannot tell a measured line from a plug will quote the plug.
    for k, key in [("Fuel", "fuel"), ("Maintenance", "maintenance"), ("Crew", "crew"),
                   ("Landing (PLUG, generic)", "landing"),
                   ("Passenger charges (PLUG, generic)", "per_pax"),
                   ("Ground handling (PLUG, generic)", "handling"),
                   ("En-route navigation (PLUG, generic)", "nav"), ("Catering", "catering"),
                   ("Admin", "admin"), ("Sales", "sales")]:
        prow(k, -abs(n0(raw.get(key))))
    prow("Cash operating cost before ownership", -abs(n0(raw.get("total_cost")))
         + abs(n0(raw.get("ownership"))) + abs(n0(raw.get("insurance"))), True, LFILL)
    prow("Contribution towards ownership per rotation", n0(ec.get("contribution_before_ownership")),
         True, TOTF)
    prow("Ownership and insurance (PLUG, not published)",
         -abs(n0(raw.get("ownership"))) - abs(n0(raw.get("insurance"))))
    prow("Operating profit per rotation, both plugs as set", raw.get("profit"), True, TOTF)
    r += 1
    _sec(ws, r, "SUMMARY", 2); r += 1
    for k, v, fmt in [("Breaks even at ownership per block hour",
                       n0(ec.get("ownership_breakeven_per_bh")), "#,##0"),
                      ("Equivalent monthly lease at type utilisation",
                       n0(ec.get("ownership_breakeven_per_month")), "#,##0"),
                      ("Multiple of the model's ownership plug",
                       n0(ec.get("ownership_breakeven_multiple")), "0.00"),
                      ("Operating margin, both plugs as set", n0(raw.get("margin")), "0.0%"),
                      ("Breakeven load factor", n0(raw.get("breakeven_lf")), "0.0%"),
                      ("Passengers per rotation", n0(raw.get("pax_turn")), "#,##0"),
                      (f"{_padj} contribution towards ownership",
                       n0(ec.get("annual_contribution_before_ownership")), "#,##0"),
                      (f"{_padj} profit, both plugs as set", n0(ec.get("annual_profit")), "#,##0"),
                      ("Aircraft required", ec.get("aircraft_required") or 0, "0.00")]:
        _c(ws, r, 1, k, font=BOLD, align=LFT); _c(ws, r, 2, v, fmt=fmt, align=RGT); r += 1
    r += 1
    _c(ws, r, 1, "Two inputs above are plugs, not measurements. Airport and handling charges are a "
                 "generic placeholder, not this airport pair, and Avia does not hold a charges "
                 "database; published charges are in any case a ceiling, since most carriers "
                 "negotiate below them. Ownership rests on a lease rate that is not available in "
                 "public form and which Avia does not publish. Set both to your own figures. The "
                 "contribution line is unaffected by the ownership plug.", align=LFT); r += 2

    # ---- 5b. Competition: alliance seat share (18 Aug 2026) ----------------
    # One download must be enough to populate a client deck's competition slide
    # (the EVA review found this was the one figure no Meridian output carried).
    # Passed in via meta by the caller that has store access; absent, no sheet,
    # never an empty one.
    _alli = (meta or {}).get("alliance") or {}
    _ends = [e for e in (_alli.get("origin"), _alli.get("dest"))
             if isinstance(e, dict) and e.get("ok")]
    if _ends:
        # EW / 2-way pair, built for structural consistency (John, 22 August 2026):
        # seat share is a percentage, not a passenger count, so it prints
        # identically on both tabs; said so on the tab itself.
        for _suf in ("EW", "2-way"):
            ws = wb.create_sheet(f"Competition {_suf}")
            _title(ws, "Alliance seat share", "departing seats by alliance at each end, OAG snapshot week")
            r = 4
            for e in _ends:
                _c(ws, r, 1, "%s  -  week %s, %s weekly departing seats"
                   % (e.get("airport"), e.get("week"), f"{e.get('weekly_seats', 0):,}"),
                   font=BOLD); r += 1
                _hdr(ws, r, ["Alliance", "Share of seats"], [30, 20]); r += 1
                for name, share in (e.get("rows") or []):
                    _c(ws, r, 1, name)
                    _c(ws, r, 2, f"{share * 100:.1f}%")
                    r += 1
                r += 1
            _c(ws, r, 1, "Source: OAG schedules, Meridian analysis. Seat shares are a percentage, not a "
                         "passenger count, so this tab reads identically on the EW and 2-way tabs.")

    # ---- 6. Assumptions & methodology --------------------------------------
    ws = wb.create_sheet("Assumptions")
    _title(ws, "Assumptions and methodology", "every key parameter behind this forecast")
    _hdr(ws, 4, ["Parameter", "Value", "Basis"], [30, 20, 60]); r = 5
    A = [
        ("Addressable market", f'{round(n0(dem.get("natural"))):,} each way/yr', "Sabre Global Demand Data (point of origin) in the origin catchment"),
        ("Origin QSI capture", f'{cap_share*100:.1f}%', meta.get("capture_basis", "modelled from drive time and competing service")),
        ("Coverage gross-up", f'x{n0(dem.get("coverage_gross_up")) or 1:.2f}', "uplift from surveyed to full O&D coverage"),
        ("Stimulation", f'x{stim:.2f}', "new nonstop demand uplift by carrier type"),
        ("Connecting feed", f'behind {round(behind):,}, beyond {round(beyond):,}', "alliance-weighted, circuity-screened onward O&D"),
        ("Planned load factor", f'{n0(cap.get("load"))*100:.0f}%', f'cap {n0(meta.get("plan_lf") or cap.get("load"))*100:.0f}%'),
        # R5: the fare line is a BAND on this self-serve surface. The P&L inside the
        # workbook still runs on the exact assumption server-side; what is stated
        # here is the band it falls in, which is what the licence permits to travel.
        ("Economy fare (one way)",
         (lambda _b: f'${_b["label"]} (band)' if _b else "-")(fare_bands.band(ec.get("econ_fare"))),
         "Sabre Global Demand Data implied / distance floor; stated as a band"),
        ("Aircraft", cap.get("aircraft",""), "seats and burn from validated type economics"),
        ("Maintenance basis", raw.get("maint_basis",""), "sector-aware Airbus reserves"),
        ("Ownership basis", raw.get("own_basis",""), "blended owned/leased cost of capital by type and age"),
    ]
    for k, v, b in A:
        _c(ws, r, 1, k, font=BOLD, align=LFT); _c(ws, r, 2, v, align=RGT); _c(ws, r, 3, b, align=LFT); r += 1
    r += 1
    _sec(ws, r, "METHODOLOGY", 3); r += 1
    method = ("1. Catchment: the resident population within drive time of the origin, from GeoNames and "
              "least-cost road times.  2. Market: measured Sabre Global Demand Data O&D each way in that catchment.  3. Capture: "
              "the QSI + access share the new nonstop takes from competing airports and airlines; measured "
              "survey/mobility data overrides the model where held.  4. Stimulation: uplift for the new nonstop.  "
              "5. Connecting feed: onward O&D behind the origin and beyond the destination on the chosen "
              "airline, alliance-weighted and circuity-screened.  6. Capacity cap: demand bounded by the "
              "aircraft and frequency at the planned load factor.  7. Economics: turnaround and "
              f"{_pnoun} P&L on validated type costs.  Indicative central estimate, for directional guidance.")
    _c(ws, r, 1, method, font=NORM, align=LFT); ws.merge_cells(start_row=r, start_column=1, end_row=r + 6, end_column=3)
    ws.row_dimensions[r].height = 120

    cp = wb.properties
    cp.creator = "Avia Solutions"; cp.lastModifiedBy = "Avia Solutions"
    cp.title = f'{o["city"]} to {d["city"]} route workbook'
    wb.save(out_path)
    # render_curve_png() already ran earlier, in the Departure curve section above, so its
    # output could be embedded into the "Departure curve 2-way" sheet before this save -
    # a workbook once saved cannot have an image added after the fact without reopening it,
    # so this has to happen before wb.save(), not after it. See that section's own note
    # (24 August 2026) for why embedding, not a sibling file, is what actually reaches John.
    return out_path
